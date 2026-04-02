# ============================================================================
# FRS.Simulator - 関節力学試験ロボットシステム FRS-2015 可視化プログラム
# ============================================================================
# バージョン: 1.0
# 作成日: 2024
# 説明: 股関節/膝関節の3Dモデル可視化、アニメーション、フィッティング
# ============================================================================
VERSION = "2.2"  # ここでバージョンを変更してください
# ============================================================================

# ============================================================================
# 【目次 - Table of Contents】
# ============================================================================
# このファイルは単一ファイルとして実装されており、約7900行のコードで構成されています。
# 以下の目次を使用して、必要なセクションに素早くジャンプできます。
#
# ■ 主要セクション
#   1. インポートと初期設定                           (行 30-50)
#   2. グローバルヘルパー関数                         (行 52-70)
#   3. MainMenuGUIクラス                              (行 72-7950)
#       3.1 初期化・UI構築                            (行 72-500)
#       3.2 ファイル選択メソッド                      (行 775-950)
#       3.3 可視化メソッド（近位・遠位・全体）        (行 951-2100)
#       3.4 変換行列管理                              (行 2101-2370)
#       3.5 キャッシュ管理                            (行 2369-2815)
#       3.6 フィッティング処理【重要】                (行 3043-4300)
#       3.7 ORG可視化                                 (行 4302-4670)
#       3.8 C Fix機能                                 (行 4671-5280)
#       3.9 アニメーション処理【最重要・1500行】      (行 5279-6750)
#       3.10 ボタン状態管理                           (行 6753-6845)
#       3.11 状態保存・復元                           (行 6846-7045)
#       3.12 座標系・幾何処理                         (行 7045-7440)
#       3.13 並列処理ヘルパー                         (行 7441-7850)
#       3.14 その他ユーティリティ                     (行 7851-7950)
#   4. main()エントリーポイント                       (行 7952-7960)
#
# ■ 重要メソッド一覧（改修頻度が高い箇所）
#   - on_animate():               シミュレーションのメインループ（行 5279）
#   - on_fitting_execute():       RANSAC/ICPフィッティング（行 3043）
#   - on_visualize_all():         関節全体の可視化（行 1382）
#   - _compute_distance_heatmap(): ヒートマップ計算（行 7070）
#   - _load_transform_matrices():  Excelデータ読み込み（行 2121）
#   - _build_coordinate_system():  座標系構築（近位）（行 7249）
#
# ■ VSCode region/endregion マーカー
#   このファイルはVSCodeのregion/endregion機能に対応しています。
#   エディタの折りたたみ機能を活用して、セクションごとに表示/非表示を切り替えられます。
#
# ■ コード検索Tips
#   - メソッド定義: "def method_name"で検索
#   - 座標変換: "transform"または"matrix"で検索
#   - PyVistaプロッター: "pv.Plotter"で検索
#   - Open3D処理: "o3d."で検索
#   - ボタンコールバック: "on_"で検索（例: on_visualize_prox）
#
# ============================================================================

# region インポートと初期設定
import sys
import platform
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from tkinter import font as tkfont
import urllib.parse

# プラットフォーム判定
IS_WINDOWS = sys.platform == 'win32'
IS_MACOS = sys.platform == 'darwin'
IS_LINUX = sys.platform.startswith('linux')

# ドラッグ&ドロップ対応（tkinterdnd2）
try:
	from tkinterdnd2 import TkinterDnD, DND_FILES
	_HAS_DND = True
except ImportError:
	_HAS_DND = False
import json
from pathlib import Path
import openpyxl
import numpy as np
import pyvista as pv
import open3d as o3d # type: ignore
import traceback
import threading
import time
from multiprocessing import Pool, cpu_count
import os
import pickle
import hashlib
import io
import xml.etree.ElementTree as ET
import copy
from typing import Optional

# FEM接触解析ソルバー
try:
	from fem_contact_solver_2 import (
		FEMContactSolver, MaterialProperties, ContactParameters,
		FEMResults, apply_fem_results_to_mesh,
	)
	_HAS_FEM = True
except ImportError:
	_HAS_FEM = False

# レポート自動生成モジュール
try:
	from fem_report_generator import generate_report_from_simulator
	_HAS_REPORT = True
except ImportError:
	_HAS_REPORT = False

# Open3Dの警告を抑制
o3d.utility.set_verbosity_level(o3d.utility.VerbosityLevel.Error)
# endregion インポートと初期設定

# region グローバルヘルパー関数
def save_child_pp_file(save_path: str, points: np.ndarray, labels: list) -> None:
	"""子特徴点（PP）をXML形式で保存する
	<PickedPoints><point x="..." y="..." z="..." active="1" name="..."/></PickedPoints>
	モジュールレベルのヘルパー（クラス外）
	"""
	root = ET.Element('PickedPoints')
	for pt, label in zip(points, labels):
		point_elem = ET.SubElement(root, 'point')
		point_elem.set('x', f"{pt[0]:.8f}")
		point_elem.set('y', f"{pt[1]:.8f}")
		point_elem.set('z', f"{pt[2]:.8f}")
		point_elem.set('active', "1")
		point_elem.set('name', label)
	tree = ET.ElementTree(root)
	tree.write(save_path, encoding="utf-8", xml_declaration=True)
# endregion グローバルヘルパー関数


# region MainMenuGUIクラス
# tkinterdnd2がある場合はTkinterDnD.Tkを基底に、なければtk.Tkを使用
_BaseWindow = TkinterDnD.Tk if _HAS_DND else tk.Tk

class MainMenuGUI(_BaseWindow):
	"""
	FRS-2015 可視化プログラムのメインGUIクラス
	
	このクラスは以下の主要機能を提供します：
	- Simulator: 近位・遠位モデルの可視化とアニメーション
	- Fitting: RANSAC/ICPによる3Dモデル位置合わせ
	- ORG: RobotScanデータの可視化
	- C Fix: 座標系修正ツール
	
	【改修ポイント】
	- on_animate(): アニメーション処理（約1500行）
	- on_fitting_execute(): フィッティング処理
	- 各種可視化メソッド: on_visualize_prox(), on_visualize_dist(), on_visualize_all()
	"""
	
	# region 初期化とUI構築
	def __init__(self) -> None:
		super().__init__()
		# Window
		self.title(f"関節力学試験ロボットシステム FRS-2015：可視化プログラム Ver.{VERSION}")
		self.minsize(820, 680)

		# macOS固有のウィンドウ設定
		if IS_MACOS:
			try:
				# macOSでのネイティブメニューバー対応
				self.createcommand('tk::mac::ShowPreferences', lambda: None)
				# Retinaディスプレイでのスケーリング
				self.tk.call('tk', 'scaling', self.winfo_fpixels('1i') / 72.0)
			except Exception:
				pass

		# ウィンドウサイズを保存（後で復元用）
		self._initial_window_geometry = None

		# フォントを環境に合わせて統一
		self._setup_fonts()

		# Variables (Simulator)
		self.joint_var = tk.IntVar(value=0)  # 0: 未選択, 1: 股関節, 2: 膝関節
		self.prox_model_path = tk.StringVar(value="")
		self.prox_pp_abcd_path = tk.StringVar(value="")  # 近位ABCD特徴点
		self.prox_pp_olmn_path = tk.StringVar(value="")  # 近位OLMN特徴点
		self.dist_model_path = tk.StringVar(value="")
		self.dist_pp_abc_path = tk.StringVar(value="")   # 遠位ABC特徴点
		self.dist_pp_olmn_path = tk.StringVar(value="")  # 遠位OLMN特徴点
		self.transform_group_path = tk.StringVar(value="")
		
		# 可視化オプション
		self.show_ao_angle = tk.BooleanVar(value=True) # 骨盤A-O線と大腿骨A-O線の表示

		# Variables (Fitting)
		self.fitting_parent_model_path = tk.StringVar(value="")
		self.fitting_parent_region_path = tk.StringVar(value="")
		self.fitting_child_model_path = tk.StringVar(value="")
		self.fitting_child_region_path = tk.StringVar(value="")
		self.fitting_child_pp_path = tk.StringVar(value="")
		# ヒートマップ表示用モデル（任意）
		self.fitting_child_heatmap_model_path = tk.StringVar(value="")
		
		# Variables (ORG)
		self.org_model_path = tk.StringVar(value="")
		self.org_pp_path = tk.StringVar(value="")
		self.org_coord_x = tk.DoubleVar(value=0.0)
		self.org_coord_y = tk.DoubleVar(value=0.0)
		self.org_coord_z = tk.DoubleVar(value=0.0)

		# Variables (A-B)
		self.a_b_model_a_path = tk.StringVar(value="")
		self.a_b_model_b_path = tk.StringVar(value="")
		
		# Variables (HeatMap)
		self.heatmap_model_a_path = tk.StringVar(value="")
		self.heatmap_model_b_path = tk.StringVar(value="")
		self.heatmap_show_model_a = tk.BooleanVar(value=True)
		self.heatmap_show_model_b = tk.BooleanVar(value=True)
		
		# Variables (Randomizer)
		self.randomizer_input_path = tk.StringVar(value="")
		self.random_trans_range = tk.DoubleVar(value=50.0)
		self.random_rot_range = tk.DoubleVar(value=180.0)

		# Variables (Cartilage Separation / 軟骨分離)
		# 近位（骨盤）
		self.cs_prox_model1_whole_path = tk.StringVar(value="")   # 骨盤: 骨+軟骨 全体
		self.cs_prox_model1_region_path = tk.StringVar(value="")  # 骨盤: 骨+軟骨 任意領域
		self.cs_prox_model2_whole_path = tk.StringVar(value="")   # 骨盤: 骨のみ 全体
		self.cs_prox_model2_region_path = tk.StringVar(value="")  # 骨盤: 骨のみ 任意領域
		# 遠位（大腿骨）
		self.cs_dist_model1_whole_path = tk.StringVar(value="")   # 大腿骨: 骨+軟骨 全体
		self.cs_dist_model1_region_path = tk.StringVar(value="")  # 大腿骨: 骨+軟骨 任意領域
		self.cs_dist_model2_whole_path = tk.StringVar(value="")   # 大腿骨: 骨のみ 全体
		self.cs_dist_model2_region_path = tk.StringVar(value="")  # 大腿骨: 骨のみ 任意領域

		# 軟骨分離 専用RANSACパラメータ
		self.cs_ransac_distance_threshold = tk.DoubleVar(value=1.0)
		self.cs_ransac_max_iterations = tk.IntVar(value=1000)
		self.cs_ransac_confidence = tk.DoubleVar(value=0.99)

		# 軟骨分離 専用ICPパラメータ
		self.cs_icp_threshold = tk.DoubleVar(value=5.0)
		self.cs_icp_max_iterations = tk.IntVar(value=2000)
		self.cs_sample_points = tk.IntVar(value=10000)

		# 軟骨分離 距離閾値（セグメンテーション用）
		self.cs_distance_threshold = tk.DoubleVar(value=0.5)  # mm
		# 軟骨分離 スケール補正（3Dスキャナ vs CT のスケール差自動補正）
		self.cs_enable_scaling = tk.BooleanVar(value=False)

		# 軟骨分離 可視化オプション
		self.cs_bone_color = "#F5DEB3"       # 骨: 小麦色
		self.cs_cartilage_color = "#00CED1"  # 軟骨: ダークターコイズ
		self.cs_bone_opacity = tk.DoubleVar(value=0.4)
		self.cs_cartilage_opacity = tk.DoubleVar(value=1.0)

		# Simulator用: 軟骨モデルパス（任意・分離後に設定）
		self.prox_cartilage_model_path = tk.StringVar(value="")  # 近位軟骨モデル
		self.dist_cartilage_model_path = tk.StringVar(value="")  # 遠位軟骨モデル

		# Variables (FEM解析)
		# 入力メッシュパス（複合モデル: 骨+軟骨）
		self.fem_prox_bone_path = tk.StringVar(value="")       # 近位骨モデル
		self.fem_prox_cartilage_path = tk.StringVar(value="")  # 近位軟骨モデル
		self.fem_dist_bone_path = tk.StringVar(value="")       # 遠位骨モデル
		self.fem_dist_cartilage_path = tk.StringVar(value="")  # 遠位軟骨モデル
		# 材料パラメータ（軟骨）
		self.fem_cart_E = tk.DoubleVar(value=10.0)        # ヤング率 [MPa]
		self.fem_cart_nu = tk.DoubleVar(value=0.45)       # ポアソン比
		self.fem_cart_thickness = tk.DoubleVar(value=2.0)  # 軟骨厚さ [mm]
		# 材料パラメータ（骨 — 参考表示）
		self.fem_bone_E = tk.DoubleVar(value=17000.0)     # ヤング率 [MPa]
		self.fem_bone_nu = tk.DoubleVar(value=0.3)        # ポアソン比
		# 接触パラメータ
		self.fem_penalty_stiffness = tk.DoubleVar(value=500.0)  # ペナルティ剛性 [MPa/mm]
		self.fem_contact_tolerance = tk.DoubleVar(value=2.0)    # 接触判定閾値 [mm]
		# 解析オプション
		self.fem_boundary_mode = tk.StringVar(value="auto")     # 境界条件モード
		self.fem_max_nodes = tk.IntVar(value=30000)             # 最大節点数
		# 可視化オプション
		self.fem_scalar_name = tk.StringVar(value="contact_pressure")
		self.fem_cmap = tk.StringVar(value="jet")
		self.fem_show_dist_mesh = tk.BooleanVar(value=True)
		self.fem_show_bone = tk.BooleanVar(value=True)
		# Simulator連携
		self.show_fem_analysis = tk.BooleanVar(value=False)
		# ランタイム状態（非永続）
		self.fem_last_results = None
		self.fem_last_prox_cart_mesh = None
		self.fem_last_dist_cart_mesh = None
		self.fem_last_prox_bone_mesh = None
		self.fem_last_dist_bone_mesh = None
		self.fem_is_running = False

		# Fittingパラメータ
		# RANSACパラメータ
		self.ransac_distance_threshold = tk.DoubleVar(value=1.0)  # 距離閾値 (mm)
		self.ransac_max_iterations = tk.IntVar(value=1000)  # 最大反復回数
		self.ransac_confidence = tk.DoubleVar(value=0.99)  # 信頼度
		
		# ICPパラメータ
		self.icp_threshold = tk.DoubleVar(value=5.0)  # 閾値 (mm)
		self.icp_max_iterations = tk.IntVar(value=2000)  # 最大反復回数
		self.sample_points = tk.IntVar(value=10000)  # サンプル点数
		
		# 可視化オプション（色設定）
		self.prox_color = "#B0C4DE"  # デフォルト: 水色
		self.dist_color = "#FFB6C1"  # デフォルト: ピンク
		self.overlap_color = "#BC0051"  # デフォルト: 黄色
		
		# 関節範囲の半径（球体抽出用）
		self.prox_radius = tk.DoubleVar(value=30.0)  # 近位側の半径（mm）
		self.dist_radius = tk.DoubleVar(value=30.0)  # 遠位側の半径（mm）
		
		# O'原点位置オフセット（ローカル座標系に沿った移動量）
		self.prox_offset_x = tk.DoubleVar(value=0.0)  # 近位O' X方向オフセット（mm）
		self.prox_offset_y = tk.DoubleVar(value=0.0)  # 近位O' Y方向オフセット（mm）
		self.prox_offset_z = tk.DoubleVar(value=0.0)  # 近位O' Z方向オフセット（mm）
		self.dist_offset_x = tk.DoubleVar(value=0.0)  # 遠位O' X方向オフセット（mm）
		self.dist_offset_y = tk.DoubleVar(value=0.0)  # 遠位O' Y方向オフセット（mm）
		self.dist_offset_z = tk.DoubleVar(value=0.0)  # 遠位O' Z方向オフセット（mm）

		# 状態復元（可能なら）
		self._load_state()

		# スタイル設定（LabelFrameのタイトルを太字に）
		style = ttk.Style()
		style.configure("Bold.TLabelframe.Label", font=(self.ui_font_family, 9, "bold"))

		# Layout root container
		main_container = ttk.Frame(self, padding=12)
		main_container.grid(row=0, column=0, sticky="nsew")
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		main_container.columnconfigure(0, weight=1)
		main_container.rowconfigure(1, weight=1)

		# Title
		title_label = ttk.Label(
			main_container,
			text="関節力学試験ロボットシステム FRS-2015：可視化プログラム",
			font=(self.ui_font_family, 13, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", pady=(0, 12))

		# タブコントロールを作成
		self.notebook = ttk.Notebook(main_container)
		self.notebook.grid(row=1, column=0, sticky="nsew")

		# タブ1: Simulator
		self.simulator_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.simulator_tab, text="Simulator")

		# タブ2: Fitting
		self.fitting_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.fitting_tab, text="Fitting")

		# タブ3: 軟骨分離 (Cartilage Separation)
		self.cartilage_sep_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.cartilage_sep_tab, text="軟骨分離")

		# タブ4: FEM解析
		self.fem_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.fem_tab, text="FEM解析")

		# タブ5: Utilities (サブタブを含むコンテナ)
		self.utilities_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.utilities_tab, text="Utilities")
		
		# Utilities内のサブノートブック
		self.utilities_notebook = ttk.Notebook(self.utilities_tab)
		self.utilities_notebook.pack(fill="both", expand=True)

		# Utilities > ORG
		self.org_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.org_tab, text="ORG")

		# Utilities > A-B
		self.a_b_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.a_b_tab, text="A-B")

		# Utilities > HeatMap
		self.heatmap_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.heatmap_tab, text="HeatMap")

		# Utilities > C fix
		self.c_fix_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.c_fix_tab, text="C fix")

		# Utilities > STL2ASC
		self.stl2asc_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.stl2asc_tab, text="STL2ASC")

		# Utilities > Randomizer
		self.randomizer_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.randomizer_tab, text="Randomizer")

		# Utilities > PosChecker
		self.pos_checker_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.pos_checker_tab, text="PosChecker")

		# Utilities > Coordinate Creater
		self.coordinate_creater_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.coordinate_creater_tab, text="Coordinate Creater")

		# Utilities > pp merge
		self.pp_merge_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.pp_merge_tab, text="pp merge")

		# Simulatorタブのコンテンツを作成
		self._create_simulator_tab()

		# Fittingタブのコンテンツを作成
		self._create_fitting_tab()

		# 軟骨分離タブのコンテンツを作成
		self._create_cartilage_separation_tab()

		# FEM解析タブのコンテンツを作成
		self._create_fem_tab()

		# ORGタブのコンテンツを作成
		# self._create_org_tab() は存在しないためダミー実装を呼び出し
		self._create_org_tab()

		# A-Bタブのコンテンツを作成
		self._create_a_b_tab()

		# HeatMapタブのコンテンツを作成
		self._create_heatmap_tab()

		# C fixタブのコンテンツを作成
		self._create_c_fix_tab()

		# STL2ASCタブのコンテンツを作成
		self._create_stl2asc_tab()

		# Randomizerタブのコンテンツを作成
		self._create_randomizer_tab()

		# PosCheckerタブのコンテンツを作成
		self._create_pos_checker_tab()

		# Coordinate Createrタブのコンテンツを作成
		self._create_coordinate_creater_tab()

		# pp mergeタブのコンテンツを作成
		self._create_pp_merge_tab()

		# 初期状態のボタン活性制御
		self.update_button_states()

		# Variable trace for dynamic state updates
		self.prox_model_path.trace_add("write", lambda *_: self.update_button_states())
		self.prox_pp_abcd_path.trace_add("write", lambda *_: self.update_button_states())
		self.prox_pp_olmn_path.trace_add("write", lambda *_: self.update_button_states())
		self.dist_model_path.trace_add("write", lambda *_: self.update_button_states())
		self.dist_pp_abc_path.trace_add("write", lambda *_: self.update_button_states())
		self.dist_pp_olmn_path.trace_add("write", lambda *_: self.update_button_states())
		self.transform_group_path.trace_add("write", lambda *_: self.update_button_states())

		# 終了時に状態保存
		self.protocol("WM_DELETE_WINDOW", self._on_close)
		
		# ウィンドウが完全に表示された後にサイズを保存
		self.after(100, self._save_initial_geometry)
	# endregion 初期化とUI構築

	# region UI構築メソッド
	def _create_simulator_tab(self) -> None:
		"""Simulatorタブのコンテンツを作成"""
		container = self.simulator_tab
		container.columnconfigure(0, weight=1)

		# Section: 使用する関節
		joint_frame = ttk.LabelFrame(container, text="使用する関節", style="Bold.TLabelframe")
		joint_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 12))
		joint_frame.columnconfigure(0, weight=1)

		rb1 = ttk.Radiobutton(
			joint_frame, text="1. 股関節", value=1, variable=self.joint_var, command=self.update_button_states
		)
		rb2 = ttk.Radiobutton(
			joint_frame, text="2. 膝関節 (未実装)", value=2, variable=self.joint_var, command=self.update_button_states, state="disabled"
		)
		# 左詰め・上下配置
		rb1.grid(row=0, column=0, sticky="w", padx=12, pady=6)
		rb2.grid(row=1, column=0, sticky="w", padx=12, pady=6)

		# Section: 近位ファイル選択
		prox_file_frame = ttk.LabelFrame(container, text="近位ファイル選択（寛骨臼側）", style="Bold.TLabelframe")
		prox_file_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			prox_file_frame.columnconfigure(i, weight=1)

		# 近位モデル (OBJ/STL)
		self._add_file_row(
			parent=prox_file_frame,
			row=0,
			label_text="近位モデル (OBJ/STL)",
			textvariable=self.prox_model_path,
			command=self.choose_prox_model, 
		)

		# 近位特徴点ABCD (PP)
		self._add_file_row(
			parent=prox_file_frame,
			row=1,
			label_text="近位特徴点 ABCD (PP)",
			textvariable=self.prox_pp_abcd_path,
			command=self.choose_prox_pp_abcd,
		)

		# 近位特徴点OLMN (PP)
		self._add_file_row(
			parent=prox_file_frame,
			row=2,
			label_text="近位特徴点 OLMN (PP)",
			textvariable=self.prox_pp_olmn_path,
			command=self.choose_prox_pp_olmn,
		)

		# 近位軟骨モデル（任意）
		self._add_file_row(
			parent=prox_file_frame,
			row=3,
			label_text="近位軟骨モデル (STL/OBJ) ※任意",
			textvariable=self.prox_cartilage_model_path,
			command=self.choose_prox_cartilage_model,
		)

		# Section: 遠位ファイル選択
		dist_file_frame = ttk.LabelFrame(container, text="遠位ファイル選択（大腿骨側）", style="Bold.TLabelframe")
		dist_file_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 12))
		for i in range(3):
			dist_file_frame.columnconfigure(i, weight=1)

		# 遠位モデル (OBJ/STL)
		self._add_file_row(
			parent=dist_file_frame,
			row=0,
			label_text="遠位モデル (OBJ/STL)",
			textvariable=self.dist_model_path,
			command=self.choose_dist_model,
		)

		# 遠位特徴点ABC (PP)
		self._add_file_row(
			parent=dist_file_frame,
			row=1,
			label_text="遠位特徴点 ABC (PP)",
			textvariable=self.dist_pp_abc_path,
			command=self.choose_dist_pp_abc,
		)

		# 遠位特徴点OLMN (PP)
		self._add_file_row(
			parent=dist_file_frame,
			row=2,
			label_text="遠位特徴点 OLMN (PP)",
			textvariable=self.dist_pp_olmn_path,
			command=self.choose_dist_pp_olmn,
		)

		# 遠位軟骨モデル（任意）
		self._add_file_row(
			parent=dist_file_frame,
			row=3,
			label_text="遠位軟骨モデル (STL/OBJ) ※任意",
			textvariable=self.dist_cartilage_model_path,
			command=self.choose_dist_cartilage_model,
		)

		# 操作ボタン段（可視化）
		action_frame = ttk.Frame(container)
		action_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.visualize_prox_button = ttk.Button(action_frame, text="近位モデル/特徴点を可視化", command=self.on_visualize_prox)
		self.visualize_prox_button.grid(row=0, column=0, sticky="w")

		self.visualize_dist_button = ttk.Button(action_frame, text="遠位モデル/特徴点を可視化", command=self.on_visualize_dist)
		self.visualize_dist_button.grid(row=1, column=0, sticky="w", pady=(6, 0))

		self.visualize_all_button = ttk.Button(action_frame, text="関節全体モデル/特徴点を可視化", command=self.on_visualize_all)
		self.visualize_all_button.grid(row=2, column=0, sticky="w", pady=(6, 0))

		self.visualize_options_button = ttk.Button(action_frame, text="可視化オプション", command=self.on_visualize_options)
		self.visualize_options_button.grid(row=3, column=0, sticky="w", pady=(6, 0))

		# 同次変換行列群 入力 & 確認
		tf_frame = ttk.LabelFrame(container, text="変位・姿勢変化データファイルの選択", style="Bold.TLabelframe")
		tf_frame.grid(row=4, column=0, sticky="nsew", pady=(0, 12))
		for i in range(3):
			tf_frame.columnconfigure(i, weight=1)

		ttk.Label(tf_frame, text="変位・姿勢変化データファイルを入力").grid(row=0, column=0, sticky="w", padx=12, pady=6)
		tf_entry = ttk.Entry(tf_frame, textvariable=self.transform_group_path, state="readonly")
		tf_entry.grid(row=0, column=1, sticky="ew", padx=12, pady=6)
		self._register_dnd(tf_entry, self.transform_group_path)
		ttk.Button(tf_frame, text="参照...", command=self.choose_transform_group).grid(row=0, column=2, sticky="e", padx=12, pady=6)

		self.check_tf_button = ttk.Button(tf_frame, text="変位・姿勢変化データファイルの確認", command=self.on_check_transform)
		self.check_tf_button.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 6))

		# シミュレーション実行（ヒートマップ事前計算→アニメーション）
		self.animate_button = ttk.Button(tf_frame, text="シミュレーション実行", command=self.on_animate)
		self.animate_button.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 6))

	def _create_org_tab(self) -> None:
		"""ORGタブのUIを構築"""
		# スクロール可能なメインフレーム
		canvas = tk.Canvas(self.org_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.org_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)

		scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
		)

		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)

		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")

		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="ORG - RobotScan データ可視化",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# ファイル選択フレーム
		file_frame = ttk.LabelFrame(container, text="ファイル選択", style="Bold.TLabelframe")
		file_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
		for i in range(3):
			file_frame.columnconfigure(i, weight=[0, 1, 0][i])

		# RobotScan STL/OBJ
		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="RobotScan モデル (STL/OBJ)",
			textvariable=self.org_model_path,
			command=self.choose_org_model,
		)

		# RobotScan PP
		self._add_file_row(
			parent=file_frame,
			row=1,
			label_text="RobotScan 座標ポイント (PP)",
			textvariable=self.org_pp_path,
			command=self.choose_org_pp,
		)

		# ORG座標入力フレーム
		coord_frame = ttk.LabelFrame(container, text="ORG座標入力（C_o座標系）", style="Bold.TLabelframe")
		coord_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		coord_frame.columnconfigure(1, weight=1)
		coord_frame.columnconfigure(3, weight=1)
		coord_frame.columnconfigure(5, weight=1)

		ttk.Label(coord_frame, text="X (mm):").grid(row=0, column=0, sticky="e", padx=(12, 5), pady=8)
		self.org_x_entry = ttk.Entry(coord_frame, textvariable=self.org_coord_x, width=15)
		self.org_x_entry.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=8)

		ttk.Label(coord_frame, text="Y (mm):").grid(row=0, column=2, sticky="e", padx=(12, 5), pady=8)
		self.org_y_entry = ttk.Entry(coord_frame, textvariable=self.org_coord_y, width=15)
		self.org_y_entry.grid(row=0, column=3, sticky="w", padx=(0, 12), pady=8)

		ttk.Label(coord_frame, text="Z (mm):").grid(row=0, column=4, sticky="e", padx=(12, 5), pady=8)
		self.org_z_entry = ttk.Entry(coord_frame, textvariable=self.org_coord_z, width=15)
		self.org_z_entry.grid(row=0, column=5, sticky="w", padx=(0, 12), pady=8)

		# 可視化ボタン
		action_frame = ttk.Frame(container)
		action_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.org_visualize_button = ttk.Button(
			action_frame,
			text="モデル + 座標ポイントを可視化",
			command=self.on_visualize_org,
			state="disabled"
		)
		self.org_visualize_button.grid(row=0, column=0, sticky="ew")

		# Variable trace
		self.org_model_path.trace_add("write", lambda *_: self.update_org_button_states())
		self.org_pp_path.trace_add("write", lambda *_: self.update_org_button_states())

		# 初期状態を更新
		self.update_org_button_states()

	def _create_a_b_tab(self) -> None:
		"""A-BタブのUIを構築"""
		container = self.a_b_tab
		container.columnconfigure(0, weight=1)

		title_label = ttk.Label(
			container,
			text="A-B モデル表示",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		file_frame = ttk.LabelFrame(container, text="モデル選択", style="Bold.TLabelframe")
		file_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
		for i in range(3):
			file_frame.columnconfigure(i, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="モデルA (STL)",
			textvariable=self.a_b_model_a_path,
			command=self.choose_a_b_model_a,
		)

		self._add_file_row(
			parent=file_frame,
			row=1,
			label_text="モデルB (STL)",
			textvariable=self.a_b_model_b_path,
			command=self.choose_a_b_model_b,
		)

		action_frame = ttk.Frame(container)
		action_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.a_b_visualize_button = ttk.Button(
			action_frame,
			text="AとBを同時表示",
			command=self.on_visualize_a_b,
			state="disabled",
		)
		self.a_b_visualize_button.grid(row=0, column=0, sticky="ew")

		self.a_b_diff_button = ttk.Button(
			action_frame,
			text="A-Bを実行",
			command=self.on_execute_a_b_heatmap,
			state="disabled",
		)
		self.a_b_diff_button.grid(row=1, column=0, sticky="ew", pady=(8, 0))

		# Variable trace
		self.a_b_model_a_path.trace_add("write", lambda *_: self.update_a_b_button_states())
		self.a_b_model_b_path.trace_add("write", lambda *_: self.update_a_b_button_states())

		# 初期状態を更新
		self.update_a_b_button_states()

	def _create_heatmap_tab(self) -> None:
		"""HeatMapタブのUIを構築"""
		container = self.heatmap_tab
		container.columnconfigure(0, weight=1)

		title_label = ttk.Label(
			container,
			text="HeatMap (B→A 距離)",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		info_label = ttk.Label(
			container,
			text="モデルBの各点からモデルA表面までの距離をヒートマップ表示します。\n可視化画面で点をクリックすると距離(mm)が表示されます。",
			justify=tk.LEFT
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 8))

		file_frame = ttk.LabelFrame(container, text="モデル選択", style="Bold.TLabelframe")
		file_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		for i in range(3):
			file_frame.columnconfigure(i, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="モデルA (参照: 表面) (STL/OBJ)",
			textvariable=self.heatmap_model_a_path,
			command=self.choose_heatmap_model_a,
		)

		self._add_file_row(
			parent=file_frame,
			row=1,
			label_text="モデルB (距離計算対象) (STL/OBJ)",
			textvariable=self.heatmap_model_b_path,
			command=self.choose_heatmap_model_b,
		)

		# 表示モデル切り替え
		display_frame = ttk.LabelFrame(container, text="表示モデル切り替え", style="Bold.TLabelframe")
		display_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		display_frame.columnconfigure(0, weight=1)
		display_frame.columnconfigure(1, weight=1)

		self.heatmap_show_model_a_check = ttk.Checkbutton(
			display_frame,
			text="モデルAを表示",
			variable=self.heatmap_show_model_a,
		)
		self.heatmap_show_model_a_check.grid(row=0, column=0, sticky="w", padx=12, pady=8)

		self.heatmap_show_model_b_check = ttk.Checkbutton(
			display_frame,
			text="モデルBを表示",
			variable=self.heatmap_show_model_b,
		)
		self.heatmap_show_model_b_check.grid(row=0, column=1, sticky="w", padx=12, pady=8)

		action_frame = ttk.Frame(container)
		action_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.heatmap_visualize_button = ttk.Button(
			action_frame,
			text="AとBを同時表示",
			command=self.on_visualize_heatmap_models,
			state="disabled",
		)
		self.heatmap_visualize_button.grid(row=0, column=0, sticky="ew")

		self.heatmap_execute_button = ttk.Button(
			action_frame,
			text="B→A 距離ヒートマップ",
			command=self.on_execute_heatmap_b_to_a,
			state="disabled",
		)
		self.heatmap_execute_button.grid(row=1, column=0, sticky="ew", pady=(8, 0))

		# Variable trace
		self.heatmap_model_a_path.trace_add("write", lambda *_: self.update_heatmap_button_states())
		self.heatmap_model_b_path.trace_add("write", lambda *_: self.update_heatmap_button_states())

		# 初期状態を更新
		self.update_heatmap_button_states()

	def _create_c_fix_tab(self) -> None:
		"""C fixタブのコンテンツを作成"""
		container = self.c_fix_tab
		container.columnconfigure(0, weight=1)
		
		# タイトル
		title_label = ttk.Label(
			container,
			text="Coordinate Fix ツール",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))
		
		# 説明
		info_label = ttk.Label(
			container,
			text="Simulatorタブで選択されたモデルと特徴点を使用して座標系を確認・修正します。",
			justify=tk.LEFT,
			wraplength=600
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))
		
		# 修正量入力フレーム
		correction_frame = ttk.LabelFrame(container, text="修正量", style="Bold.TLabelframe")
		correction_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		correction_frame.columnconfigure(1, weight=1)
		correction_frame.columnconfigure(3, weight=1)
		correction_frame.columnconfigure(5, weight=1)
		
		# 修正量の変数
		self.c_fix_correction_x = tk.DoubleVar(value=0.0)
		self.c_fix_correction_y = tk.DoubleVar(value=0.0)
		self.c_fix_correction_z = tk.DoubleVar(value=0.0)
		
		ttk.Label(correction_frame, text="X (mm):").grid(row=0, column=0, sticky="e", padx=(12, 5), pady=8)
		self.c_fix_x_entry = ttk.Entry(correction_frame, textvariable=self.c_fix_correction_x, width=15)
		self.c_fix_x_entry.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=8)
		
		ttk.Label(correction_frame, text="Y (mm):").grid(row=0, column=2, sticky="e", padx=(12, 5), pady=8)
		self.c_fix_y_entry = ttk.Entry(correction_frame, textvariable=self.c_fix_correction_y, width=15)
		self.c_fix_y_entry.grid(row=0, column=3, sticky="w", padx=(0, 12), pady=8)
		
		ttk.Label(correction_frame, text="Z (mm):").grid(row=0, column=4, sticky="e", padx=(12, 5), pady=8)
		self.c_fix_z_entry = ttk.Entry(correction_frame, textvariable=self.c_fix_correction_z, width=15)
		self.c_fix_z_entry.grid(row=0, column=5, sticky="w", padx=(0, 12), pady=8)
		
		# 可視化ボタン
		viz_frame = ttk.LabelFrame(container, text="可視化", style="Bold.TLabelframe")
		viz_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		viz_frame.columnconfigure(0, weight=1)
		
		self.c_fix_visualize_button = ttk.Button(
			viz_frame,
			text="モデル + 特徴点を可視化",
			command=self.on_c_fix_visualize,
			state="disabled"
		)
		self.c_fix_visualize_button.grid(row=0, column=0, sticky="ew", padx=12, pady=12)
		
		# 変換ボタン
		transform_frame = ttk.LabelFrame(container, text="変換", style="Bold.TLabelframe")
		transform_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 12))
		transform_frame.columnconfigure(0, weight=1)
		transform_frame.columnconfigure(1, weight=1)
		
		self.c_fix_transform_prox_button = ttk.Button(
			transform_frame,
			text="近位ABCD修正後を保存",
			command=self.on_c_fix_transform_prox,
			state="disabled"
		)
		self.c_fix_transform_prox_button.grid(row=0, column=0, sticky="ew", padx=12, pady=12)
		
		self.c_fix_transform_dist_button = ttk.Button(
			transform_frame,
			text="遠位ABC修正後を保存",
			command=self.on_c_fix_transform_dist,
			state="disabled"
		)
		self.c_fix_transform_dist_button.grid(row=0, column=1, sticky="ew", padx=12, pady=12)
		
		# Variable trace for button state updates
		self.prox_model_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.prox_pp_abcd_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.prox_pp_olmn_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.dist_model_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.dist_pp_abc_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.dist_pp_olmn_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		
		# 初期状態を更新
		self.update_c_fix_button_states()

	def _create_fitting_tab(self) -> None:
		"""Fittingタブのコンテンツを作成"""
		# スクロール可能なキャンバスを作成
		canvas = tk.Canvas(self.fitting_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.fitting_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)
		
		scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
		)
		
		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)
		
		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")
		
		# マウスホイールでスクロール
		def _on_mousewheel(event):
			canvas.yview_scroll(int(-1*(event.delta/120)), "units")
		canvas.bind_all("<MouseWheel>", _on_mousewheel)
		
		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="3Dモデル フィッティングツール",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# 説明（コンパクト版）
		description_text = (
			"親任意領域に対して子任意領域をフィッティングします (RANSAC/ICP)\n"
			"その後、親モデルファイルに対して同じ動き方で子モデルファイルを移動させます。"
		)
		description = ttk.Label(
			container,
			text=description_text,
			justify=tk.LEFT
		)
		description.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 4))
		
		# パラメータ説明ボタン
		help_button = ttk.Button(container, text="パラメータの意味を表示", command=self.show_fitting_help)
		help_button.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 8))

		# Section: ファイル選択（親）
		parent_frame = ttk.LabelFrame(container, text="親ファイル（固定側）", style="Bold.TLabelframe")
		parent_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 8))
		for i in range(3):
			parent_frame.columnconfigure(i, weight=1)

		# 親モデルファイル
		self._add_file_row(
			parent=parent_frame,
			row=0,
			label_text="親モデルファイル (STL/OBJ)",
			textvariable=self.fitting_parent_model_path,
			command=self.choose_fitting_parent_model,
		)

		# 親任意領域ファイル
		self._add_file_row(
			parent=parent_frame,
			row=1,
			label_text="親任意領域ファイル (STL/OBJ)",
			textvariable=self.fitting_parent_region_path,
			command=self.choose_fitting_parent_region,
		)

		# Section: ファイル選択（子）
		child_frame = ttk.LabelFrame(container, text="子ファイル（移動側）", style="Bold.TLabelframe")
		child_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 8))
		for i in range(3):
			child_frame.columnconfigure(i, weight=1)

		# 子モデルファイル
		self._add_file_row(
			parent=child_frame,
			row=0,
			label_text="子モデルファイル (STL/OBJ)",
			textvariable=self.fitting_child_model_path,
			command=self.choose_fitting_child_model,
		)

		# 子任意領域ファイル
		self._add_file_row(
			parent=child_frame,
			row=1,
			label_text="子任意領域ファイル (STL/OBJ)",
			textvariable=self.fitting_child_region_path,
			command=self.choose_fitting_child_region,
		)

		# 子特徴点ファイル
		self._add_file_row(
			parent=child_frame,
			row=2,
			label_text="子特徴点ファイル (PP)",
			textvariable=self.fitting_child_pp_path,
			command=self.choose_fitting_child_pp,
		)

		# 子ヒートマップ表示用モデル（任意）
		self._add_file_row(
			parent=child_frame,
			row=3,
			label_text="子ヒートマップ表示用モデル (STL/OBJ)",
			textvariable=self.fitting_child_heatmap_model_path,
			command=self.choose_fitting_child_heatmap_model,
		)


		# 操作ボタン段（可視化）
		viz_frame = ttk.LabelFrame(container, text="モデル確認", style="Bold.TLabelframe")
		viz_frame.grid(row=5, column=0, sticky="nsew", padx=12, pady=(0, 8))
		viz_frame.columnconfigure(0, weight=1)
		viz_frame.columnconfigure(1, weight=1)

		self.fitting_viz_parent_button = ttk.Button(
			viz_frame, 
			text="親を可視化（モデル + 任意領域）",
			command=self.on_visualize_fitting_parent,
			width=30
		)
		self.fitting_viz_parent_button.grid(row=0, column=0, sticky="ew", padx=8, pady=6)

		self.fitting_viz_child_button = ttk.Button(
			viz_frame, 
			text="子を可視化（モデル + 任意領域）",
			command=self.on_visualize_fitting_child,
			width=30
		)
		self.fitting_viz_child_button.grid(row=0, column=1, sticky="ew", padx=8, pady=6)

		# パラメータ設定フレーム
		param_frame = ttk.LabelFrame(container, text="フィッティングパラメータ", style="Bold.TLabelframe")
		param_frame.grid(row=6, column=0, sticky="nsew", padx=12, pady=(0, 8))
		param_frame.columnconfigure(1, weight=1)
		param_frame.columnconfigure(3, weight=1)
		param_frame.columnconfigure(5, weight=1)
		
		# RANSACパラメータ
		ransac_label = ttk.Label(param_frame, text="RANSACパラメータ", font=(self.ui_font_family, 9, "bold"))
		ransac_label.grid(row=0, column=0, columnspan=6, sticky="w", padx=8, pady=(6, 3))
		
		ttk.Label(param_frame, text="距離閾値 (mm):").grid(row=1, column=0, sticky="e", padx=(8, 5), pady=4)
		self.ransac_distance_entry = ttk.Entry(param_frame, textvariable=self.ransac_distance_threshold, width=10)
		self.ransac_distance_entry.grid(row=1, column=1, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="最大反復回数:").grid(row=1, column=2, sticky="e", padx=(8, 5), pady=4)
		self.ransac_max_iterations_entry = ttk.Entry(param_frame, textvariable=self.ransac_max_iterations, width=10)
		self.ransac_max_iterations_entry.grid(row=1, column=3, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="信頼度:").grid(row=1, column=4, sticky="e", padx=(8, 5), pady=4)
		self.ransac_confidence_entry = ttk.Entry(param_frame, textvariable=self.ransac_confidence, width=10)
		self.ransac_confidence_entry.grid(row=1, column=5, sticky="w", padx=(0, 8), pady=4)
		
		# ICPパラメータ
		icp_label = ttk.Label(param_frame, text="ICPパラメータ", font=(self.ui_font_family, 9, "bold"))
		icp_label.grid(row=2, column=0, columnspan=6, sticky="w", padx=8, pady=(8, 3))
		
		ttk.Label(param_frame, text="閾値 (mm):").grid(row=3, column=0, sticky="e", padx=(8, 5), pady=4)
		self.icp_threshold_entry = ttk.Entry(param_frame, textvariable=self.icp_threshold, width=10)
		self.icp_threshold_entry.grid(row=3, column=1, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="最大反復回数:").grid(row=3, column=2, sticky="e", padx=(8, 5), pady=4)
		self.icp_max_iterations_entry = ttk.Entry(param_frame, textvariable=self.icp_max_iterations, width=10)
		self.icp_max_iterations_entry.grid(row=3, column=3, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="サンプル点数:").grid(row=3, column=4, sticky="e", padx=(8, 5), pady=4)
		self.sample_points_entry = ttk.Entry(param_frame, textvariable=self.sample_points, width=10)
		self.sample_points_entry.grid(row=3, column=5, sticky="w", padx=(0, 8), pady=4)
		
		# デフォルトに戻すボタン
		btn_reset_params = ttk.Button(param_frame, text="デフォルトに戻す", command=self.reset_fitting_parameters)
		btn_reset_params.grid(row=4, column=0, columnspan=6, pady=(6, 6))

		# 操作ボタン段（実行）
		action_frame = ttk.Frame(container)
		action_frame.grid(row=7, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.fitting_execute_button = ttk.Button(
			action_frame, 
			text="フィッティング実行",
			command=self.on_fitting_execute,
			state="disabled"
		)
		self.fitting_execute_button.grid(row=0, column=0, sticky="ew", pady=4)

		# Variable trace for dynamic state updates
		self.fitting_parent_model_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_parent_region_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_child_model_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_child_region_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_child_heatmap_model_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		
		# 初期状態を更新
		self.update_fitting_button_states()

	def _create_cartilage_separation_tab(self) -> None:
		"""軟骨分離タブ: 近位（骨盤）・遠位（大腿骨）それぞれのサブタブを持つ"""
		# スクロール可能なキャンバス
		canvas = tk.Canvas(self.cartilage_sep_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.cartilage_sep_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)
		scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")
		def _on_mousewheel_cs(event):
			canvas.yview_scroll(int(-1*(event.delta/120)), "units")
		canvas.bind_all("<MouseWheel>", _on_mousewheel_cs)
		
		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		ttk.Label(container, text="軟骨分離ツール (Cartilage Separation)",
			font=(self.ui_font_family, 12, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))

		ttk.Label(container, justify=tk.LEFT, wraplength=700, text=(
			"骨+軟骨モデル（体積メッシュ）と骨のみモデル（体積メッシュ）を位置合わせし、\n"
			"ブーリアン差分（モデル1 − モデル2）で軟骨領域を分離します。\n"
			"近位（骨盤）・遠位（大腿骨）それぞれで実行してください。\n"
			"分離後のモデルは Simulator タブの軟骨モデル入力欄から読み込めます。")
		).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 8))

		# === 近位 / 遠位 サブノートブック ===
		cs_notebook = ttk.Notebook(container)
		cs_notebook.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 8))

		# --- 近位（骨盤）サブタブ ---
		prox_tab = ttk.Frame(cs_notebook, padding=8)
		cs_notebook.add(prox_tab, text="近位（骨盤）")
		self._build_cs_side_ui(prox_tab, side="prox")

		# --- 遠位（大腿骨）サブタブ ---
		dist_tab = ttk.Frame(cs_notebook, padding=8)
		cs_notebook.add(dist_tab, text="遠位（大腿骨）")
		self._build_cs_side_ui(dist_tab, side="dist")

		# === 共有: RANSAC/ICPパラメータ ===
		param_frame = ttk.LabelFrame(container, text="フィッティングパラメータ（共有）", style="Bold.TLabelframe")
		param_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 8))
		param_frame.columnconfigure(1, weight=1)
		param_frame.columnconfigure(3, weight=1)
		param_frame.columnconfigure(5, weight=1)

		ttk.Label(param_frame, text="RANSACパラメータ", font=(self.ui_font_family, 9, "bold")).grid(row=0, column=0, columnspan=6, sticky="w", padx=8, pady=(6, 3))
		ttk.Label(param_frame, text="距離閾値 (mm):").grid(row=1, column=0, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_ransac_distance_threshold, width=10).grid(row=1, column=1, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="最大反復回数:").grid(row=1, column=2, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_ransac_max_iterations, width=10).grid(row=1, column=3, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="信頼度:").grid(row=1, column=4, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_ransac_confidence, width=10).grid(row=1, column=5, sticky="w", padx=(0, 8), pady=4)

		ttk.Label(param_frame, text="ICPパラメータ", font=(self.ui_font_family, 9, "bold")).grid(row=2, column=0, columnspan=6, sticky="w", padx=8, pady=(8, 3))
		ttk.Label(param_frame, text="閾値 (mm):").grid(row=3, column=0, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_icp_threshold, width=10).grid(row=3, column=1, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="最大反復回数:").grid(row=3, column=2, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_icp_max_iterations, width=10).grid(row=3, column=3, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="サンプル点数:").grid(row=3, column=4, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_sample_points, width=10).grid(row=3, column=5, sticky="w", padx=(0, 8), pady=4)
		btn_row = ttk.Frame(param_frame)
		btn_row.grid(row=4, column=0, columnspan=6, pady=(6, 6))
		ttk.Button(btn_row, text="デフォルトに戻す", command=self.reset_cs_fitting_parameters).pack(side="left", padx=4)
		ttk.Button(btn_row, text="現在の値をデフォルトとして保存", command=self._save_cs_fitting_defaults).pack(side="left", padx=4)

		# === 共有: スケール補正 ===
		scale_frame = ttk.LabelFrame(container, text="スケール補正", style="Bold.TLabelframe")
		scale_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 8))

		ttk.Checkbutton(scale_frame, text="スケール補正を有効化",
			variable=self.cs_enable_scaling).grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		ttk.Label(scale_frame, text="3Dスキャナ（骨+軟骨）と CT（骨のみ）のように撮影装置が異なる場合、\n"
			"モデル間の寸法差（スケール差）を ICP で自動補正します。\n"
			"同一装置で撮影した場合はOFFのままで構いません。",
			foreground="gray", justify="left").grid(row=1, column=0, sticky="w", padx=24, pady=(0, 8))

		# === 共有: 軟骨抽出パラメータ ===
		extract_frame = ttk.LabelFrame(container, text="軟骨抽出パラメータ", style="Bold.TLabelframe")
		extract_frame.grid(row=5, column=0, sticky="nsew", padx=12, pady=(0, 8))

		ttk.Label(extract_frame, text="距離閾値 [mm]:").grid(row=0, column=0, sticky="e", padx=(8, 5), pady=(6, 2))
		ttk.Entry(extract_frame, textvariable=self.cs_distance_threshold, width=8).grid(row=0, column=1, sticky="w", padx=(0, 8), pady=(6, 2))
		ttk.Label(extract_frame, text="骨表面からこの距離以上離れた部分を軟骨として抽出します。\n"
			"値を小さくすると薄い軟骨も拾えます（推奨: 0.1〜0.5mm）。",
			foreground="gray", justify="left").grid(row=1, column=0, columnspan=2, sticky="w", padx=24, pady=(0, 8))

		# === 共有: 可視化オプション ===
		vis_frame = ttk.LabelFrame(container, text="可視化オプション（共有）", style="Bold.TLabelframe")
		vis_frame.grid(row=6, column=0, sticky="nsew", padx=12, pady=(0, 8))
		vis_frame.columnconfigure(1, weight=1); vis_frame.columnconfigure(3, weight=1)
		ttk.Label(vis_frame, text="骨の色:").grid(row=0, column=0, sticky="e", padx=(8, 5), pady=6)
		ttk.Button(vis_frame, text="■ 変更", command=self._choose_cs_bone_color).grid(row=0, column=1, sticky="w", padx=(0, 8), pady=6)
		ttk.Label(vis_frame, text="骨の透明度:").grid(row=0, column=2, sticky="e", padx=(8, 5), pady=6)
		ttk.Entry(vis_frame, textvariable=self.cs_bone_opacity, width=8).grid(row=0, column=3, sticky="w", padx=(0, 8), pady=6)
		ttk.Label(vis_frame, text="軟骨の色:").grid(row=1, column=0, sticky="e", padx=(8, 5), pady=6)
		ttk.Button(vis_frame, text="■ 変更", command=self._choose_cs_cartilage_color).grid(row=1, column=1, sticky="w", padx=(0, 8), pady=6)
		ttk.Label(vis_frame, text="軟骨の透明度:").grid(row=1, column=2, sticky="e", padx=(8, 5), pady=6)
		ttk.Entry(vis_frame, textvariable=self.cs_cartilage_opacity, width=8).grid(row=1, column=3, sticky="w", padx=(0, 8), pady=6)

	def _build_cs_side_ui(self, parent, side: str) -> None:
		"""近位/遠位共通のファイル入力・ボタンUIを構築
		Args:
			side: "prox" or "dist"
		"""
		parent.columnconfigure(0, weight=1)
		side_label = "近位（骨盤）" if side == "prox" else "遠位（大腿骨）"

		# 変数を取得
		if side == "prox":
			v_m1w = self.cs_prox_model1_whole_path
			v_m1r = self.cs_prox_model1_region_path
			v_m2w = self.cs_prox_model2_whole_path
			v_m2r = self.cs_prox_model2_region_path
		else:
			v_m1w = self.cs_dist_model1_whole_path
			v_m1r = self.cs_dist_model1_region_path
			v_m2w = self.cs_dist_model2_whole_path
			v_m2r = self.cs_dist_model2_region_path

		# モデル1（骨+軟骨）
		m1f = ttk.LabelFrame(parent, text=f"{side_label}: モデル1（骨＋軟骨）", style="Bold.TLabelframe")
		m1f.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3): m1f.columnconfigure(i, weight=1)
		self._add_file_row(parent=m1f, row=0, label_text="全体モデル (STL/OBJ)", textvariable=v_m1w,
			command=lambda s=side: self._choose_cs_file(s, "m1w"))
		self._add_file_row(parent=m1f, row=1, label_text="任意領域 (STL/OBJ) ※共通骨領域", textvariable=v_m1r,
			command=lambda s=side: self._choose_cs_file(s, "m1r"))

		# モデル2（骨のみ）
		m2f = ttk.LabelFrame(parent, text=f"{side_label}: モデル2（骨のみ）", style="Bold.TLabelframe")
		m2f.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3): m2f.columnconfigure(i, weight=1)
		self._add_file_row(parent=m2f, row=0, label_text="全体モデル (STL/OBJ)", textvariable=v_m2w,
			command=lambda s=side: self._choose_cs_file(s, "m2w"))
		self._add_file_row(parent=m2f, row=1, label_text="任意領域 (STL/OBJ) ※共通骨領域", textvariable=v_m2r,
			command=lambda s=side: self._choose_cs_file(s, "m2r"))

		# 確認 + 実行ボタン
		action_frame = ttk.Frame(parent)
		action_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 8))
		action_frame.columnconfigure(0, weight=1)
		action_frame.columnconfigure(1, weight=1)

		viz_btn = ttk.Button(action_frame, text="モデルを確認",
			command=lambda s=side: self._on_cs_visualize_side(s))
		viz_btn.grid(row=0, column=0, sticky="ew", padx=(0, 4), pady=4)

		exec_btn = ttk.Button(action_frame, text=f"{side_label} 軟骨分離を実行",
			command=lambda s=side: self._on_cs_execute_side(s))
		exec_btn.grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=4)

		# ボタン参照を保持
		if side == "prox":
			self.cs_prox_viz_btn = viz_btn
			self.cs_prox_exec_btn = exec_btn
		else:
			self.cs_dist_viz_btn = viz_btn
			self.cs_dist_exec_btn = exec_btn

		# 変数のtrace
		for v in [v_m1w, v_m1r, v_m2w, v_m2r]:
			v.trace_add("write", lambda *_, s=side: self._update_cs_side_buttons(s))

	# region FEM解析タブ
	def _create_fem_tab(self) -> None:
		"""FEM解析タブのコンテンツを作成"""
		container = self.fem_tab
		container.columnconfigure(0, weight=1)

		if not _HAS_FEM:
			ttk.Label(container, text="FEM解析モジュール (fem_contact_solver_2.py) が見つかりません。\n同じディレクトリに配置してください。",
				foreground="red").grid(row=0, column=0, pady=20)
			return

		row_idx = 0

		# タイトル
		ttk.Label(container, text="FEM 接触解析（複合モデル: 骨+軟骨）",
			font=(self.ui_font_family, 11, "bold")).grid(row=row_idx, column=0, sticky="w", pady=(0, 4))
		row_idx += 1
		ttk.Label(container, text="骨と軟骨を分離したモデルに対し、軟骨面同士の接触応力を有限要素法で解析します。",
			foreground="gray").grid(row=row_idx, column=0, sticky="w", pady=(0, 8))
		row_idx += 1

		# === 入力メッシュ ===
		mesh_frame = ttk.LabelFrame(container, text="入力メッシュ（複合モデル）", style="Bold.TLabelframe")
		mesh_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		mesh_frame.columnconfigure(1, weight=1)
		row_idx += 1

		# 近位（骨盤）
		ttk.Label(mesh_frame, text="【近位（骨盤）】", font=(self.ui_font_family, 9, "bold")).grid(
			row=0, column=0, columnspan=3, sticky="w", padx=8, pady=(6, 2))
		ttk.Label(mesh_frame, text="骨モデル:").grid(row=1, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_prox_bone_entry = ttk.Entry(mesh_frame, textvariable=self.fem_prox_bone_path, width=60)
		fem_prox_bone_entry.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_prox_bone_entry, self.fem_prox_bone_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("prox_bone")).grid(row=1, column=2, padx=(4, 8), pady=2)
		ttk.Label(mesh_frame, text="軟骨モデル:").grid(row=2, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_prox_cart_entry = ttk.Entry(mesh_frame, textvariable=self.fem_prox_cartilage_path, width=60)
		fem_prox_cart_entry.grid(row=2, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_prox_cart_entry, self.fem_prox_cartilage_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("prox_cartilage")).grid(row=2, column=2, padx=(4, 8), pady=2)

		# 遠位（大腿骨）
		ttk.Label(mesh_frame, text="【遠位（大腿骨）】", font=(self.ui_font_family, 9, "bold")).grid(
			row=3, column=0, columnspan=3, sticky="w", padx=8, pady=(8, 2))
		ttk.Label(mesh_frame, text="骨モデル:").grid(row=4, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_dist_bone_entry = ttk.Entry(mesh_frame, textvariable=self.fem_dist_bone_path, width=60)
		fem_dist_bone_entry.grid(row=4, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_dist_bone_entry, self.fem_dist_bone_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("dist_bone")).grid(row=4, column=2, padx=(4, 8), pady=2)
		ttk.Label(mesh_frame, text="軟骨モデル:").grid(row=5, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_dist_cart_entry = ttk.Entry(mesh_frame, textvariable=self.fem_dist_cartilage_path, width=60)
		fem_dist_cart_entry.grid(row=5, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_dist_cart_entry, self.fem_dist_cartilage_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("dist_cartilage")).grid(row=5, column=2, padx=(4, 8), pady=(2, 8))

		# === 材料パラメータ ===
		mat_frame = ttk.LabelFrame(container, text="材料パラメータ", style="Bold.TLabelframe")
		mat_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		# 軟骨
		ttk.Label(mat_frame, text="《軟骨》", font=(self.ui_font_family, 9, "bold")).grid(
			row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		ttk.Label(mat_frame, text="ヤング率 E [MPa]:").grid(row=0, column=1, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_cart_E, width=10).grid(row=0, column=2, padx=4, pady=2)
		ttk.Label(mat_frame, text="ポアソン比 ν:").grid(row=0, column=3, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_cart_nu, width=10).grid(row=0, column=4, padx=4, pady=2)
		ttk.Label(mat_frame, text="厚さ [mm]:").grid(row=0, column=5, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_cart_thickness, width=10).grid(row=0, column=6, padx=(4, 8), pady=2)

		# 骨（参考表示）
		ttk.Label(mat_frame, text="《骨》", font=(self.ui_font_family, 9, "bold")).grid(
			row=1, column=0, sticky="w", padx=8, pady=(4, 2))
		ttk.Label(mat_frame, text="ヤング率 E [MPa]:").grid(row=1, column=1, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_bone_E, width=10).grid(row=1, column=2, padx=4, pady=2)
		ttk.Label(mat_frame, text="ポアソン比 ν:").grid(row=1, column=3, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_bone_nu, width=10).grid(row=1, column=4, padx=4, pady=2)
		ttk.Label(mat_frame, text="（剛体BC扱い）", foreground="gray").grid(
			row=1, column=5, columnspan=2, sticky="w", padx=4, pady=(4, 8))

		# === 接触パラメータ ===
		contact_frame = ttk.LabelFrame(container, text="接触パラメータ", style="Bold.TLabelframe")
		contact_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		ttk.Label(contact_frame, text="ペナルティ剛性 [MPa/mm]:").grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		ttk.Entry(contact_frame, textvariable=self.fem_penalty_stiffness, width=10).grid(row=0, column=1, padx=4, pady=(6, 2))
		ttk.Label(contact_frame, text="接触判定閾値 [mm]:").grid(row=0, column=2, sticky="w", padx=(16, 4), pady=(6, 2))
		ttk.Entry(contact_frame, textvariable=self.fem_contact_tolerance, width=10).grid(row=0, column=3, padx=(4, 8), pady=(6, 8))

		# === 解析オプション ===
		opt_frame = ttk.LabelFrame(container, text="解析オプション", style="Bold.TLabelframe")
		opt_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		ttk.Label(opt_frame, text="境界条件:").grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		bc_combo = ttk.Combobox(opt_frame, textvariable=self.fem_boundary_mode, values=["auto", "rim"],
			state="readonly", width=8)
		bc_combo.grid(row=0, column=1, padx=4, pady=(6, 2))
		ttk.Label(opt_frame, text="最大節点数:").grid(row=0, column=2, sticky="w", padx=(16, 4), pady=(6, 2))
		ttk.Entry(opt_frame, textvariable=self.fem_max_nodes, width=10).grid(row=0, column=3, padx=4, pady=(6, 2))
		ttk.Button(opt_frame, text="デフォルトに戻す", command=self._reset_fem_parameters).grid(
			row=0, column=4, padx=(16, 8), pady=(6, 8))

		# === 実行ボタン ===
		btn_frame = ttk.Frame(container)
		btn_frame.grid(row=row_idx, column=0, sticky="w", pady=(0, 8))
		row_idx += 1

		self.fem_execute_btn = ttk.Button(btn_frame, text="解析を実行", command=self._on_fem_execute)
		self.fem_execute_btn.pack(side="left", padx=(0, 8))
		self.fem_visualize_btn = ttk.Button(btn_frame, text="結果を可視化", command=self._on_fem_visualize, state="disabled")
		self.fem_visualize_btn.pack(side="left", padx=(0, 8))
		self.fem_report_btn = ttk.Button(btn_frame, text="レポート生成", command=self._on_fem_generate_report, state="disabled")
		self.fem_report_btn.pack(side="left", padx=(0, 8))

		# === 結果サマリー ===
		result_frame = ttk.LabelFrame(container, text="結果サマリー", style="Bold.TLabelframe")
		result_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		result_frame.columnconfigure(0, weight=1)
		row_idx += 1

		self.fem_result_text = tk.Text(result_frame, height=10, state="disabled", wrap="word",
			font=(self.ui_font_family, 9))
		fem_result_scroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.fem_result_text.yview)
		self.fem_result_text.configure(yscrollcommand=fem_result_scroll.set)
		self.fem_result_text.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
		fem_result_scroll.grid(row=0, column=1, sticky="ns", padx=(0, 8), pady=8)

		# === 可視化オプション ===
		vis_frame = ttk.LabelFrame(container, text="可視化オプション", style="Bold.TLabelframe")
		vis_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		ttk.Label(vis_frame, text="表示スカラー:").grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		scalar_combo = ttk.Combobox(vis_frame, textvariable=self.fem_scalar_name,
			values=["contact_pressure", "von_mises_stress", "max_principal_stress",
				"displacement_magnitude", "penetration_depth"],
			state="readonly", width=24)
		scalar_combo.grid(row=0, column=1, padx=4, pady=(6, 2))

		ttk.Label(vis_frame, text="カラーマップ:").grid(row=0, column=2, sticky="w", padx=(16, 4), pady=(6, 2))
		cmap_combo = ttk.Combobox(vis_frame, textvariable=self.fem_cmap,
			values=["jet", "plasma", "coolwarm", "viridis", "hot"],
			state="readonly", width=12)
		cmap_combo.grid(row=0, column=3, padx=4, pady=(6, 2))

		ttk.Checkbutton(vis_frame, text="相手軟骨を表示", variable=self.fem_show_dist_mesh).grid(
			row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(2, 8))
		ttk.Checkbutton(vis_frame, text="骨を半透明表示", variable=self.fem_show_bone).grid(
			row=1, column=2, columnspan=2, sticky="w", padx=8, pady=(2, 8))

		# ボタン状態の初期設定
		self._update_fem_buttons()

		# パス変更時にボタン状態を更新
		for v in [self.fem_prox_cartilage_path, self.fem_dist_cartilage_path]:
			v.trace_add("write", lambda *_: self._update_fem_buttons())
	# --- FEM解析 コールバックメソッド群 ---

	def _choose_fem_file(self, which: str) -> None:
		"""FEM解析タブ用ファイル選択ダイアログ"""
		titles = {
			"prox_bone": "近位 骨モデル",
			"prox_cartilage": "近位 軟骨モデル",
			"dist_bone": "遠位 骨モデル",
			"dist_cartilage": "遠位 軟骨モデル",
		}
		var_map = {
			"prox_bone": self.fem_prox_bone_path,
			"prox_cartilage": self.fem_prox_cartilage_path,
			"dist_bone": self.fem_dist_bone_path,
			"dist_cartilage": self.fem_dist_cartilage_path,
		}
		path = filedialog.askopenfilename(
			title=f"FEM解析: {titles.get(which, which)}を選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			var_map[which].set(path)

	def _update_fem_buttons(self) -> None:
		"""FEM解析タブのボタン有効/無効制御"""
		has_cartilage = (bool(self.fem_prox_cartilage_path.get().strip()) and
						 bool(self.fem_dist_cartilage_path.get().strip()))
		try:
			self.fem_execute_btn.config(state="normal" if (has_cartilage and not self.fem_is_running) else "disabled")
		except Exception:
			pass
		try:
			self.fem_visualize_btn.config(state="normal" if self.fem_last_results is not None else "disabled")
		except Exception:
			pass
		try:
			self.fem_report_btn.config(state="normal" if self.fem_last_results is not None else "disabled")
		except Exception:
			pass

	def _reset_fem_parameters(self) -> None:
		"""FEM解析パラメータをデフォルト値に戻す"""
		self.fem_cart_E.set(10.0)
		self.fem_cart_nu.set(0.45)
		self.fem_cart_thickness.set(2.0)
		self.fem_bone_E.set(17000.0)
		self.fem_bone_nu.set(0.3)
		self.fem_penalty_stiffness.set(500.0)
		self.fem_contact_tolerance.set(2.0)
		self.fem_boundary_mode.set("auto")
		self.fem_max_nodes.set(30000)

	def _on_fem_execute(self) -> None:
		"""FEM接触解析を実行（バックグラウンドスレッド）"""
		if not _HAS_FEM:
			messagebox.showerror("エラー", "FEMモジュールが利用できません。")
			return

		prox_cart_path = self.fem_prox_cartilage_path.get().strip()
		dist_cart_path = self.fem_dist_cartilage_path.get().strip()

		if not prox_cart_path or not dist_cart_path:
			messagebox.showerror("エラー", "近位・遠位の軟骨モデルを選択してください。")
			return

		for p, name in [(prox_cart_path, "近位軟骨"), (dist_cart_path, "遠位軟骨")]:
			if not Path(p).exists():
				messagebox.showerror("エラー", f"{name}ファイルが見つかりません:\n{p}")
				return

		# ボタン無効化
		self.fem_is_running = True
		self._update_fem_buttons()

		# 結果テキストをクリア
		self.fem_result_text.config(state="normal")
		self.fem_result_text.delete("1.0", "end")
		self.fem_result_text.insert("1.0", "解析を実行中...")
		self.fem_result_text.config(state="disabled")

		# 骨モデルパス（可視化用、解析には不要）
		prox_bone_path = self.fem_prox_bone_path.get().strip()
		dist_bone_path = self.fem_dist_bone_path.get().strip()

		def run_analysis():
			try:
				prox_cart = pv.read(prox_cart_path)
				dist_cart = pv.read(dist_cart_path)

				# 骨メッシュ（可視化用に読み込み）
				prox_bone = pv.read(prox_bone_path) if prox_bone_path and Path(prox_bone_path).exists() else None
				dist_bone = pv.read(dist_bone_path) if dist_bone_path and Path(dist_bone_path).exists() else None

				material = MaterialProperties(
					E=self.fem_cart_E.get(),
					nu=self.fem_cart_nu.get(),
					thickness=self.fem_cart_thickness.get(),
				)
				contact = ContactParameters(
					penalty_stiffness=self.fem_penalty_stiffness.get(),
					contact_tolerance=self.fem_contact_tolerance.get(),
				)
				solver = FEMContactSolver(material=material, contact=contact, verbose=True)
				results = solver.analyze(
					prox_cart, dist_cart,
					boundary_mode=self.fem_boundary_mode.get(),
					max_nodes=self.fem_max_nodes.get(),
				)
				self._fem_thread_result = results
				self._fem_thread_prox_cart = prox_cart
				self._fem_thread_dist_cart = dist_cart
				self._fem_thread_prox_bone = prox_bone
				self._fem_thread_dist_bone = dist_bone
				self._fem_thread_error = None
			except Exception as e:
				self._fem_thread_result = None
				self._fem_thread_error = e
				import traceback
				traceback.print_exc()

		self._fem_thread = threading.Thread(target=run_analysis, daemon=True)
		self._fem_thread.start()
		self._poll_fem_thread()

	def _poll_fem_thread(self) -> None:
		"""FEM解析スレッドの完了チェック"""
		if self._fem_thread.is_alive():
			self.after(200, self._poll_fem_thread)
			return

		self.fem_is_running = False

		if hasattr(self, '_fem_thread_error') and self._fem_thread_error is not None:
			self.fem_result_text.config(state="normal")
			self.fem_result_text.delete("1.0", "end")
			self.fem_result_text.insert("1.0", f"解析エラー:\n{self._fem_thread_error}")
			self.fem_result_text.config(state="disabled")
			messagebox.showerror("FEM解析エラー", str(self._fem_thread_error))
		elif hasattr(self, '_fem_thread_result') and self._fem_thread_result is not None:
			self.fem_last_results = self._fem_thread_result
			self.fem_last_prox_cart_mesh = self._fem_thread_prox_cart
			self.fem_last_dist_cart_mesh = self._fem_thread_dist_cart
			self.fem_last_prox_bone_mesh = getattr(self, '_fem_thread_prox_bone', None)
			self.fem_last_dist_bone_mesh = getattr(self, '_fem_thread_dist_bone', None)

			# サマリー表示
			self.fem_result_text.config(state="normal")
			self.fem_result_text.delete("1.0", "end")
			self.fem_result_text.insert("1.0", self.fem_last_results.summary())
			self.fem_result_text.config(state="disabled")
			messagebox.showinfo("完了", "FEM接触解析が完了しました。")

		self._update_fem_buttons()

	def _on_fem_visualize(self) -> None:
		"""FEM解析結果をPyVistaで可視化"""
		if self.fem_last_results is None or self.fem_last_prox_cart_mesh is None:
			messagebox.showwarning("警告", "先に解析を実行してください。")
			return

		results = self.fem_last_results
		scalar_name = self.fem_scalar_name.get()
		cmap = self.fem_cmap.get()

		# 近位軟骨メッシュにFEM結果を適用
		mesh_vis = self.fem_last_prox_cart_mesh.copy()
		if results.n_nodes == mesh_vis.n_points:
			apply_fem_results_to_mesh(mesh_vis, results)
		else:
			messagebox.showwarning("警告",
				f"FEM結果の節点数({results.n_nodes})がメッシュ({mesh_vis.n_points})と不一致です。\n"
				"デシメーションが適用された可能性があります。")
			return

		# スカラーラベル定義
		scalar_labels = {
			'contact_pressure': '接触圧 [MPa]',
			'von_mises_stress': 'von Mises応力 [MPa]',
			'max_principal_stress': '最大主応力 [MPa]',
			'displacement_magnitude': '変位量 [mm]',
			'penetration_depth': '侵入量 [mm]',
		}
		label = scalar_labels.get(scalar_name, scalar_name)

		# プロッター構築
		plotter = pv.Plotter(window_size=(1200, 800))
		plotter.set_background('white')

		# 近位軟骨（FEM結果）
		plotter.add_mesh(
			mesh_vis, scalars=scalar_name, cmap=cmap, show_edges=False,
			scalar_bar_args={'title': label, 'color': 'black'},
			opacity=1.0,
		)

		# 遠位軟骨（半透明）
		if self.fem_show_dist_mesh.get() and self.fem_last_dist_cart_mesh is not None:
			plotter.add_mesh(
				self.fem_last_dist_cart_mesh, color='#FFB6C1', opacity=0.3, show_edges=False,
			)

		# 骨（半透明）
		if self.fem_show_bone.get():
			if self.fem_last_prox_bone_mesh is not None:
				plotter.add_mesh(
					self.fem_last_prox_bone_mesh, color='#F5DEB3', opacity=0.2, show_edges=False,
				)
			if self.fem_last_dist_bone_mesh is not None:
				plotter.add_mesh(
					self.fem_last_dist_bone_mesh, color='#DEB887', opacity=0.2, show_edges=False,
				)

		plotter.add_axes()

		# サマリーテキスト
		text = (
			f"接触面積: {results.contact_area:.1f} mm²\n"
			f"最大接触圧: {results.peak_contact_pressure:.3f} MPa\n"
			f"接触節点: {results.n_contact_nodes:,}\n"
			f"解析時間: {results.solve_time_sec:.1f} 秒"
		)
		plotter.add_text(text, position='upper_right', font_size=10, color='black')

		plotter.show()

	def _on_fem_generate_report(self) -> None:
		"""FEM解析レポートをWord形式で自動生成"""
		if not _HAS_REPORT:
			messagebox.showerror("エラー",
				"レポート生成モジュールが利用できません。\n"
				"fem_report_generator.py が同じフォルダにあることを確認してください。\n"
				"また python-docx がインストールされていることを確認してください:\n"
				"  pip install python-docx matplotlib")
			return

		if self.fem_last_results is None:
			messagebox.showwarning("警告", "先にFEM解析を実行してください。")
			return

		# 保存先を選択
		from tkinter import filedialog
		import datetime
		default_name = f"FEM_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
		filepath = filedialog.asksaveasfilename(
			title="レポート保存先を選択",
			defaultextension=".docx",
			filetypes=[("Word Document", "*.docx"), ("All Files", "*.*")],
			initialfile=default_name,
		)
		if not filepath:
			return

		# プログレス表示
		self.fem_report_btn.config(state="disabled", text="生成中...")
		self.root.update_idletasks()

		try:
			generate_report_from_simulator(self, output_path=filepath)
			messagebox.showinfo("完了", f"レポートを保存しました:\n{filepath}")
		except Exception as e:
			messagebox.showerror("エラー", f"レポート生成中にエラーが発生しました:\n{e}")
		finally:
			self.fem_report_btn.config(state="normal", text="レポート生成")

	def _run_fem_on_current_frame(self, prox_cartilage_mesh, dist_cartilage_mesh,
								   prox_bone_mesh=None, dist_bone_mesh=None) -> None:
		"""Simulator連携: 現在フレームの姿勢でFEM解析を実行

		on_animate() 内のFEMボタンから呼ばれる。
		メッシュは既にアニメーション姿勢に変換済み。
		"""
		if not _HAS_FEM:
			messagebox.showerror("エラー", "FEMモジュールが利用できません。")
			return

		if prox_cartilage_mesh is None or dist_cartilage_mesh is None:
			messagebox.showwarning("警告", "軟骨モデルが読み込まれていません。\nSimulatorタブで軟骨モデルを指定してください。")
			return

		try:
			material = MaterialProperties(
				E=self.fem_cart_E.get(),
				nu=self.fem_cart_nu.get(),
				thickness=self.fem_cart_thickness.get(),
			)
			contact = ContactParameters(
				penalty_stiffness=self.fem_penalty_stiffness.get(),
				contact_tolerance=self.fem_contact_tolerance.get(),
			)
			solver = FEMContactSolver(material=material, contact=contact, verbose=True)

			# 現在のメッシュ位置のコピーで解析
			prox_cart_copy = prox_cartilage_mesh.copy()
			dist_cart_copy = dist_cartilage_mesh.copy()

			results = solver.analyze(
				prox_cart_copy, dist_cart_copy,
				boundary_mode=self.fem_boundary_mode.get(),
				max_nodes=self.fem_max_nodes.get(),
			)

			# 結果を可視化
			scalar_name = self.fem_scalar_name.get()
			cmap = self.fem_cmap.get()

			mesh_vis = prox_cart_copy.copy()
			if results.n_nodes == mesh_vis.n_points:
				apply_fem_results_to_mesh(mesh_vis, results)
			else:
				# デシメーション時は解析メッシュを使用
				mesh_vis = pv.PolyData(solver._nodes, prox_cart_copy.faces)
				apply_fem_results_to_mesh(mesh_vis, results)

			scalar_labels = {
				'contact_pressure': '接触圧 [MPa]',
				'von_mises_stress': 'von Mises応力 [MPa]',
				'max_principal_stress': '最大主応力 [MPa]',
				'displacement_magnitude': '変位量 [mm]',
				'penetration_depth': '侵入量 [mm]',
			}
			label = scalar_labels.get(scalar_name, scalar_name)

			plotter = pv.Plotter(window_size=(1200, 800))
			plotter.set_background('white')

			plotter.add_mesh(
				mesh_vis, scalars=scalar_name, cmap=cmap, show_edges=False,
				scalar_bar_args={'title': label, 'color': 'black'},
				opacity=1.0,
			)

			# 遠位軟骨
			plotter.add_mesh(dist_cart_copy, color='#FFB6C1', opacity=0.3, show_edges=False)

			# 骨（半透明）
			if prox_bone_mesh is not None:
				plotter.add_mesh(prox_bone_mesh.copy(), color='#F5DEB3', opacity=0.2, show_edges=False)
			if dist_bone_mesh is not None:
				plotter.add_mesh(dist_bone_mesh.copy(), color='#DEB887', opacity=0.2, show_edges=False)

			plotter.add_axes()
			text = (
				f"接触面積: {results.contact_area:.1f} mm²\n"
				f"最大接触圧: {results.peak_contact_pressure:.3f} MPa\n"
				f"接触節点: {results.n_contact_nodes:,}\n"
				f"解析時間: {results.solve_time_sec:.1f} 秒"
			)
			plotter.add_text(text, position='upper_right', font_size=10, color='black')
			plotter.show()

		except Exception as e:
			messagebox.showerror("FEM解析エラー", f"解析中にエラーが発生しました:\n{e}")
			import traceback
			traceback.print_exc()

	# endregion FEM解析タブ

	def _create_stl2asc_tab(self) -> None:
		"""STL2ASCタブのコンテンツを作成"""
		container = self.stl2asc_tab
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="STL to ASC Converter",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# 説明
		info_label = ttk.Label(
			container,
			text="STLファイルから点群データ(ASC形式)を抽出・変換します。",
			justify=tk.LEFT,
			wraplength=600
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))

		# ファイル選択変数
		self.stl2asc_input_path = tk.StringVar(value="")

		# ファイル選択UI
		file_frame = ttk.LabelFrame(container, text="STLファイル選択", style="Bold.TLabelframe")
		file_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		file_frame.columnconfigure(1, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="STLファイル (.stl)",
			textvariable=self.stl2asc_input_path,
			command=self.choose_stl_file
		)

		# 変換ボタン
		action_frame = ttk.LabelFrame(container, text="操作", style="Bold.TLabelframe")
		action_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		
		self.stl2asc_convert_button = ttk.Button(
			action_frame,
			text="ASCとして保存",
			command=self.on_stl2asc_convert
		)
		self.stl2asc_convert_button.pack(fill="x", padx=12, pady=12)

	def choose_stl_file(self) -> None:
		"""STLファイルを選択"""
		path = filedialog.askopenfilename(
			title="STLファイルを選択",
			filetypes=[("STL files", "*.stl"), ("All files", "*.*")]
		)
		if path:
			self.stl2asc_input_path.set(path)

	def on_stl2asc_convert(self) -> None:
		"""STLをASCに変換して保存"""
		stl_path = self.stl2asc_input_path.get().strip()
		if not stl_path:
			messagebox.showwarning("警告", "STLファイルを選択してください。")
			return
		
		if not os.path.exists(stl_path):
			messagebox.showerror("エラー", f"ファイルが見つかりません:\n{stl_path}")
			return
			
		# 保存先を選択
		asc_path = filedialog.asksaveasfilename(
			title="ASCファイルとして保存",
			defaultextension=".asc",
			filetypes=[("ASC files", "*.asc"), ("All files", "*.*")],
			initialfile=Path(stl_path).stem + ".asc"
		)
		
		if not asc_path:
			return

		try:
			# PyVistaを使って読み込み
			mesh = pv.read(stl_path)
			points = mesh.points
			
			if points is None or len(points) == 0:
				messagebox.showerror("エラー", "点群データを読み込めませんでした。")
				return

			# ASCとして保存 (X Y Z format)
			np.savetxt(asc_path, points, fmt="%.8f", delimiter=" ")
			
			messagebox.showinfo("完了", f"保存しました:\n{asc_path}\n点数: {len(points)}")
			
		except Exception as e:
			messagebox.showerror("エラー", f"変換に失敗しました:\n{e}\n{traceback.format_exc()}")

	def _create_randomizer_tab(self) -> None:
		"""Randomizerタブのコンテンツを作成"""
		container = self.randomizer_tab
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="Random Transform (ワールド座標ランダム化)",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# 説明
		info_label = ttk.Label(
			container,
			text="選択したSTL/OBJモデルに対し、指定範囲内でランダムな平行移動・回転を適用して保存します。",
			justify=tk.LEFT,
			wraplength=600
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))

		# ファイル選択変数
		# (Note: self.randomizer_input_path, self.random_trans_range, self.random_rot_range are initialized in __init__)

		# ファイル選択
		file_frame = ttk.LabelFrame(container, text="対象モデル選択", style="Bold.TLabelframe")
		file_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		file_frame.columnconfigure(1, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="モデルファイル (STL/OBJ)",
			textvariable=self.randomizer_input_path,
			command=self.choose_randomizer_input
		)

		# パラメータ設定
		param_frame = ttk.LabelFrame(container, text="ランダム化パラメータ", style="Bold.TLabelframe")
		param_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		param_frame.columnconfigure(1, weight=1)

		# 平行移動範囲
		ttk.Label(param_frame, text="平行移動範囲 (±mm):").grid(row=0, column=0, sticky="e", padx=(12, 6), pady=6)
		ttk.Entry(param_frame, textvariable=self.random_trans_range).grid(row=0, column=1, sticky="w", padx=(0, 12), pady=6)
		
		# 回転範囲
		ttk.Label(param_frame, text="回転範囲 (±deg):").grid(row=1, column=0, sticky="e", padx=(12, 6), pady=6)
		ttk.Entry(param_frame, textvariable=self.random_rot_range).grid(row=1, column=1, sticky="w", padx=(0, 12), pady=6)

		# 実行ボタン
		action_frame = ttk.Frame(container)
		action_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 12))

		self.randomizer_execute_button = ttk.Button(
			action_frame,
			text="ランダム変換を実行して保存...",
			command=self.on_randomizer_execute
		)
		self.randomizer_execute_button.pack(fill="x", pady=12)
		
	def _create_pos_checker_tab(self) -> None:
		"""PosCheckerタブのコンテンツを作成"""
		container = self.pos_checker_tab
		container.columnconfigure(0, weight=1)

		# Variables
		self.pos_check_model_path = tk.StringVar(value="")
		self.pos_check_pp_pv_path = tk.StringVar(value="")
		self.pos_check_pp_fm_path = tk.StringVar(value="")

		# Model Input
		model_frame = ttk.LabelFrame(container, text="1. モデル選択 (STL/OBJ)", style="Bold.TLabelframe")
		model_frame.pack(fill="x", padx=10, pady=5)
		model_frame.columnconfigure(1, weight=1)

		ttk.Label(model_frame, text="モデルファイル:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
		pc_model_entry = ttk.Entry(model_frame, textvariable=self.pos_check_model_path)
		pc_model_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_model_entry, self.pos_check_model_path)
		ttk.Button(model_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_model_path, "model")).grid(row=0, column=2, padx=5, pady=5)

		# PP Inputs
		pp_frame = ttk.LabelFrame(container, text="2. PPファイル選択", style="Bold.TLabelframe")
		pp_frame.pack(fill="x", padx=10, pady=5)
		pp_frame.columnconfigure(1, weight=1)

		# PV (Proximal/Pelvis)
		ttk.Label(pp_frame, text="PP (PV/Proximal):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
		pc_pv_entry = ttk.Entry(pp_frame, textvariable=self.pos_check_pp_pv_path)
		pc_pv_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_pv_entry, self.pos_check_pp_pv_path)
		ttk.Button(pp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_pp_pv_path, "pp")).grid(row=0, column=2, padx=5, pady=5)
		ttk.Label(pp_frame, text="※ Simulator基本機能(A,B,C,D,O)で構築").grid(row=1, column=1, sticky="w", padx=5)

		# FM (Distal/Femur)
		ttk.Label(pp_frame, text="PP (FM/Distal):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
		pc_fm_entry = ttk.Entry(pp_frame, textvariable=self.pos_check_pp_fm_path)
		pc_fm_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_fm_entry, self.pos_check_pp_fm_path)
		ttk.Button(pp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_pp_fm_path, "pp")).grid(row=2, column=2, padx=5, pady=5)
		ttk.Label(pp_frame, text="※ Simulator遠位機能(A,B,O)で構築").grid(row=3, column=1, sticky="w", padx=5)

		# Execute 1
		btn_frame = ttk.Frame(container)
		btn_frame.pack(fill="x", padx=10, pady=10)
		ttk.Button(btn_frame, text="機能1: 可視化 & 変換行列(Cp->Cf)算出", command=self.on_pos_checker_execute).pack(fill="x", ipady=5)
		
		# Separator
		ttk.Separator(container, orient="horizontal").pack(fill="x", padx=10, pady=10)
		
		# Function 2: Matrix Comparison
		# Variables for Function 2
		self.pos_check_mat1_path = tk.StringVar(value="")
		self.pos_check_mat2_path = tk.StringVar(value="")

		comp_frame = ttk.LabelFrame(container, text="機能2: 変換行列の比較 (Input 1 ➙ Input 2)", style="Bold.TLabelframe")
		comp_frame.pack(fill="x", padx=10, pady=5)
		comp_frame.columnconfigure(1, weight=1)

		# Input 1
		ttk.Label(comp_frame, text="Input 1 (行列ファイル):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
		pc_mat1_entry = ttk.Entry(comp_frame, textvariable=self.pos_check_mat1_path)
		pc_mat1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_mat1_entry, self.pos_check_mat1_path)
		ttk.Button(comp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_mat1_path, "txt")).grid(row=0, column=2, padx=5, pady=5)

		# Input 2
		ttk.Label(comp_frame, text="Input 2 (行列ファイル):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
		pc_mat2_entry = ttk.Entry(comp_frame, textvariable=self.pos_check_mat2_path)
		pc_mat2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_mat2_entry, self.pos_check_mat2_path)
		ttk.Button(comp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_mat2_path, "txt")).grid(row=1, column=2, padx=5, pady=5)

		# Execute 2
		btn_frame2 = ttk.Frame(container)
		btn_frame2.pack(fill="x", padx=10, pady=10)
		ttk.Button(btn_frame2, text="機能2: 比較行列算出 & 保存", command=self.on_pos_checker_compare).pack(fill="x", ipady=5)

	def _choose_pos_check_file(self, target_var: tk.StringVar, ftype: str) -> None:
		if ftype == "model":
			filetypes = [("3D Model", "*.stl *.obj"), ("All files", "*.*")]
			title = "モデルファイルを選択"
		elif ftype == "pp":
			filetypes = [("PickedPoints", "*.pp"), ("All files", "*.*")]
			title = "PPファイルを選択"
		else: # txt
			filetypes = [("Text files", "*.txt"), ("All files", "*.*")]
			title = "変換行列ファイルを選択"
			
		path = filedialog.askopenfilename(title=title, filetypes=filetypes)
		if path:
			target_var.set(path)

	def on_pos_checker_execute(self) -> None:
		"""PosChecker実行: 可視化と行列計算"""
		model_path = self.pos_check_model_path.get().strip()
		pp_pv_path = self.pos_check_pp_pv_path.get().strip()
		pp_fm_path = self.pos_check_pp_fm_path.get().strip()

		if not all([model_path, pp_pv_path, pp_fm_path]):
			messagebox.showwarning("入力不足", "全てのファイルを選択してください。")
			return

		try:
			import pyvista as pv
			
			# 1. Load Data
			mesh = pv.read(model_path)
			pts_pv, lbl_pv = self._parse_pp_file(pp_pv_path)
			pts_fm, lbl_fm = self._parse_pp_file(pp_fm_path)
			
			# 2. Build Coordinate Systems
			# PV (Proximal): A, B, C, D, O
			try:
				res_pv = self._build_coordinate_system(pts_pv, lbl_pv)
				origin_pv, x_pv, y_pv, z_pv = res_pv
			except Exception as e:
				messagebox.showerror("座標系構築エラー (PV)", f"PV(Proximal)側の座標系構築に失敗しました:\n{e}")
				return

			# FM (Distal): A, B, O
			try:
				res_fm = self._build_coordinate_system_dist(pts_fm, lbl_fm)
				origin_fm, x_fm, y_fm, z_fm = res_fm
			except Exception as e:
				messagebox.showerror("座標系構築エラー (FM)", f"FM(Distal)側の座標系構築に失敗しました:\n{e}")
				return

			# 3. Calculate Transform Cp -> Cf
			# T_pv_world: PV to World
			R_pv = np.column_stack([x_pv, y_pv, z_pv])
			T_pv_world = np.eye(4)
			T_pv_world[:3, :3] = R_pv
			T_pv_world[:3, 3] = origin_pv

			# T_fm_world: FM to World
			R_fm = np.column_stack([x_fm, y_fm, z_fm])
			T_fm_world = np.eye(4)
			T_fm_world[:3, :3] = R_fm
			T_fm_world[:3, 3] = origin_fm
			
			# T_pv_fm: FM relative to PV (Coordinate of FM origin observed from PV, etc?)
			# "CpからCfまでの同時変換行列" implies taking a point in Cf and expressing it in Cp?
			# Or expressing the frame Cf in Cp.
			# T_relative = T_pv_world^(-1) * T_fm_world
			T_world_pv = np.linalg.inv(T_pv_world)
			T_rel = T_world_pv @ T_fm_world
			
			# 4. Display Result
			self._show_pos_checker_result(T_rel, mesh, pts_pv, lbl_pv, pts_fm, lbl_fm, 
										  (origin_pv, x_pv, y_pv, z_pv), 
										  (origin_fm, x_fm, y_fm, z_fm))

		except Exception as e:
			messagebox.showerror("エラー", f"実行中にエラーが発生しました:\n{e}")
			traceback.print_exc()

	def _show_pos_checker_result(self, matrix, mesh, pts_pv, lbl_pv, pts_fm, lbl_fm, cs_pv, cs_fm):
		# Result Window
		res_win = tk.Toplevel(self)
		res_win.title("PosChecker 結果")
		res_win.geometry("600x500")
		
		# Text Area
		text_frame = ttk.Frame(res_win)
		text_frame.pack(fill="both", expand=True, padx=10, pady=10)
		
		result_text = "【同時変換行列 (Cp -> Cf)】\n"
		result_text += "※ Cp(PV)座標系から見たCf(FM)座標系の位置姿勢\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{matrix[0][0]:12.6f} {matrix[0][1]:12.6f} {matrix[0][2]:12.6f} {matrix[0][3]:12.6f}\n"
		result_text += f"{matrix[1][0]:12.6f} {matrix[1][1]:12.6f} {matrix[1][2]:12.6f} {matrix[1][3]:12.6f}\n"
		result_text += f"{matrix[2][0]:12.6f} {matrix[2][1]:12.6f} {matrix[2][2]:12.6f} {matrix[2][3]:12.6f}\n"
		result_text += f"{matrix[3][0]:12.6f} {matrix[3][1]:12.6f} {matrix[3][2]:12.6f} {matrix[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"
		
		# Translation / Rotation analysis
		trans = matrix[:3, 3]
		rot_mat = matrix[:3, :3]
		# Simple Euler (XYZ)
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0

		result_text += "【解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"

		st = tk.Text(text_frame, wrap="none", font=(self.mono_font_family, 10))
		st.pack(side="left", fill="both", expand=True)
		sb = ttk.Scrollbar(text_frame, orient="vertical", command=st.yview)
		sb.pack(side="right", fill="y")
		st.config(yscrollcommand=sb.set)
		st.insert("1.0", result_text)
		st.config(state="disabled") # Read-only
		
		# Buttons
		btn_frame = ttk.Frame(res_win)
		btn_frame.pack(fill="x", padx=10, pady=10)
		
		def save_txt():
			path = filedialog.asksaveasfilename(
				title="結果を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if path:
				try:
					with open(path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"保存しました:\n{path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(btn_frame, text="テキストとして保存...", command=save_txt).pack(side="left", padx=5)
		ttk.Button(btn_frame, text="閉じる", command=res_win.destroy).pack(side="right", padx=5)

		# Visualize (PyVista)
		# numpy array の真偽値判定エラーを回避するため、Noneチェックを明示的に行う
		if mesh is not None and pts_pv is not None and pts_fm is not None: 
			# Create Plotter
			plotter = pv.Plotter(title="PosChecker Visualization")
			plotter.set_background("white")
			
			# Add Mesh
			plotter.add_mesh(mesh, color="lightgray", opacity=0.5, show_edges=False)
			
			# Add PV Points (Blue-ish)
			plotter.add_points(pts_pv, color="blue", point_size=10, render_points_as_spheres=True)
			plotter.add_point_labels(pts_pv, lbl_pv, font_size=12, text_color="blue", show_points=False)
			
			# Add FM Points (Red-ish)
			plotter.add_points(pts_fm, color="red", point_size=10, render_points_as_spheres=True)
			plotter.add_point_labels(pts_fm, lbl_fm, font_size=12, text_color="red", show_points=False)
			
			# Visualize Coordinate Systems
			origin_pv, x_pv, y_pv, z_pv = cs_pv
			origin_fm, x_fm, y_fm, z_fm = cs_fm
			
			axis_len = 50.0
			
			# PV System (Solid lines)
			plotter.add_arrows(origin_pv, x_pv, mag=axis_len, color='red', label='Cp X')
			plotter.add_arrows(origin_pv, y_pv, mag=axis_len, color='green', label='Cp Y')
			plotter.add_arrows(origin_pv, z_pv, mag=axis_len, color='blue', label='Cp Z')
			plotter.add_mesh(pv.Sphere(radius=2, center=origin_pv), color="black")
			plotter.add_point_labels(np.array([origin_pv]), ["Cp"], point_size=0, text_color="black", font_size=16, show_points=False)

			# FM System (Dotted lines? PyVista arrows usually solid. Use lighter colors or tubes)
			# Use slightly different colors (Magenta, Yellow, Cyan) to distinguish
			plotter.add_arrows(origin_fm, x_fm, mag=axis_len, color='magenta', label='Cf X')
			plotter.add_arrows(origin_fm, y_fm, mag=axis_len, color='orange', label='Cf Y') # Changed yellow to orange for visibility
			plotter.add_arrows(origin_fm, z_fm, mag=axis_len, color='cyan', label='Cf Z')
			plotter.add_mesh(pv.Sphere(radius=2, center=origin_fm), color="gray")
			plotter.add_point_labels(np.array([origin_fm]), ["Cf"], point_size=0, text_color="gray", font_size=16, show_points=False)

			plotter.add_legend()
			plotter.show()

	def on_pos_checker_compare(self) -> None:
		"""PosChecker 機能2: 変換行列の比較 (Input 1 -> Input 2)"""
		mat1_path = self.pos_check_mat1_path.get().strip()
		mat2_path = self.pos_check_mat2_path.get().strip()

		if not all([mat1_path, mat2_path]):
			messagebox.showwarning("入力不足", "Input 1 と Input 2 の両方の行列ファイルを選択してください。")
			return

		try:
			# 1. Parse Matrices
			T1 = self._parse_matrix_file(mat1_path)
			T2 = self._parse_matrix_file(mat2_path)

			if T1 is None:
				messagebox.showerror("形式エラー", f"Input 1 のファイルを解析できませんでした:\n{mat1_path}")
				return
			if T2 is None:
				messagebox.showerror("形式エラー", f"Input 2 のファイルを解析できませんでした:\n{mat2_path}")
				return

			# 2. Calculate Relative Transform Input 1 -> Input 2
			# Assuming T1 and T2 are poses relative to a common base (Cp)
			# P_Cp = T1 * P_Input1
			# P_Cp = T2 * P_Input2
			# T1 * P_Input1 = T2 * P_Input2
			# P_Input1 = T1^-1 * T2 * P_Input2
			# So Transform (1 -> 2) is T1^-1 * T2. This matrix maps points in Frame 2 to Frame 1.
			# This represents the pose of Frame 2 relative to Frame 1.
			T_12 = np.linalg.inv(T1) @ T2
			
			# 3. Show Result
			# We reuse _show_pos_checker_result but pass None for visualization args
			# And we need to adjust the title in the result text, but _show_pos_checker_result has hardcoded title.
			# So we might want to make _show_pos_checker_result more flexible or create a new method.
			# I'll create a dedicated display method to match the request exactly.
			
			self._show_pos_checker_compare_result(T_12, mat1_path, mat2_path)

		except Exception as e:
			messagebox.showerror("エラー", f"実行中にエラーが発生しました:\n{e}")
			traceback.print_exc()

	def _parse_matrix_file(self, path: str) -> Optional[np.ndarray]:
		"""テキストファイルから4x4行列を読み取る (PosChecker出力形式対応)"""
		try:
			matrix_lines = []
			with open(path, "r", encoding="utf-8") as f:
				lines = f.readlines()
			
			# シンプルなヒューリスティック: 4つの数値が並んでいる行を4行連続で探す
			for i in range(len(lines) - 3):
				candidates = lines[i:i+4]
				temp_matrix = []
				failed = False
				for line in candidates:
					parts = line.replace(",", " ").split() # カンマかスペース区切り
					# filter non-numeric parts if necessary, but assuming clean lines inside dashes
					# Try parsing 4 floats
					floats = []
					for p in parts:
						try:
							floats.append(float(p))
						except ValueError:
							pass
					
					if len(floats) == 4:
						temp_matrix.append(floats)
					else:
						failed = True
						break
				
				if not failed and len(temp_matrix) == 4:
					return np.array(temp_matrix)
			
			return None
		except Exception:
			return None

	def _show_pos_checker_compare_result(self, matrix, path1, path2):
		res_win = tk.Toplevel(self)
		res_win.title("PosChecker 比較結果")
		res_win.geometry("600x500")

		text_frame = ttk.Frame(res_win)
		text_frame.pack(fill="both", expand=True, padx=10, pady=10)

		name1 = os.path.basename(path1)
		name2 = os.path.basename(path2)

		result_text = "【変換行列比較 (Input 1 ➙ Input 2)】\n"
		result_text += f"Input 1: {path1}\n"
		result_text += f"Input 2: {path2}\n"
		result_text += "※ Input 1 のCf座標系から見た Input 2 のCf座標系の位置姿勢 (T_1_inv * T_2)\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{matrix[0][0]:12.6f} {matrix[0][1]:12.6f} {matrix[0][2]:12.6f} {matrix[0][3]:12.6f}\n"
		result_text += f"{matrix[1][0]:12.6f} {matrix[1][1]:12.6f} {matrix[1][2]:12.6f} {matrix[1][3]:12.6f}\n"
		result_text += f"{matrix[2][0]:12.6f} {matrix[2][1]:12.6f} {matrix[2][2]:12.6f} {matrix[2][3]:12.6f}\n"
		result_text += f"{matrix[3][0]:12.6f} {matrix[3][1]:12.6f} {matrix[3][2]:12.6f} {matrix[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"

		# Analysis
		trans = matrix[:3, 3]
		rot_mat = matrix[:3, :3]
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0

		dist_mm = np.linalg.norm(trans)

		result_text += "【相対変位解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Straight Dist(mm): {dist_mm:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"

		st = tk.Text(text_frame, wrap="none", font=(self.mono_font_family, 10))
		st.pack(side="left", fill="both", expand=True)
		sb = ttk.Scrollbar(text_frame, orient="vertical", command=st.yview)
		sb.pack(side="right", fill="y")
		st.config(yscrollcommand=sb.set)
		st.insert("1.0", result_text)
		st.config(state="disabled")

		btn_frame = ttk.Frame(res_win)
		btn_frame.pack(fill="x", padx=10, pady=10)

		def save_txt():
			path = filedialog.asksaveasfilename(
				title="結果を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if path:
				try:
					with open(path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"保存しました:\n{path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(btn_frame, text="テキストとして保存...", command=save_txt).pack(side="left", padx=5)
		ttk.Button(btn_frame, text="閉じる", command=res_win.destroy).pack(side="right", padx=5)


	def _create_coordinate_creater_tab(self) -> None:
		"""Coordinate Createrタブのコンテンツを作成"""
		container = self.coordinate_creater_tab
		container.columnconfigure(0, weight=1)
		
		# Variables
		self.cc_fe = tk.DoubleVar(value=0.0)
		self.cc_vv = tk.DoubleVar(value=0.0)
		self.cc_ie = tk.DoubleVar(value=0.0)
		self.cc_ml = tk.DoubleVar(value=0.0)
		self.cc_ap = tk.DoubleVar(value=0.0)
		self.cc_pd = tk.DoubleVar(value=0.0)
		
		# Container Frame
		input_frame = ttk.LabelFrame(container, text="パラメータ入力", style="Bold.TLabelframe")
		input_frame.pack(fill="x", padx=10, pady=10)
		input_frame.columnconfigure(1, weight=1)
		input_frame.columnconfigure(4, weight=1)
		
		# Rotations (FE, VV, IE)
		ttk.Label(input_frame, text="Rotation Parameters (deg)", font=(self.ui_font_family, 10, "bold")).grid(row=0, column=0, columnspan=2, pady=(5, 5), sticky="w", padx=5)
		
		ttk.Label(input_frame, text="FE (Flexion/Extension):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_fe).grid(row=1, column=1, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="VV (Varus/Valgus):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_vv).grid(row=2, column=1, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="IE (Internal/External):").grid(row=3, column=0, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_ie).grid(row=3, column=1, padx=5, pady=5, sticky="ew")
		
		ttk.Separator(input_frame, orient="vertical").grid(row=1, column=2, rowspan=3, sticky="ns", padx=10, pady=5)
		
		# Translations (ML, AP, PD)
		ttk.Label(input_frame, text="Translation Parameters (mm)", font=(self.ui_font_family, 10, "bold")).grid(row=0, column=3, columnspan=2, pady=(5, 5), sticky="w", padx=5)
		
		ttk.Label(input_frame, text="ML (Medial/Lateral):").grid(row=1, column=3, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_ml).grid(row=1, column=4, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="AP (Anterior/Posterior):").grid(row=2, column=3, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_ap).grid(row=2, column=4, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="PD (Proximal/Distal):").grid(row=3, column=3, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_pd).grid(row=3, column=4, padx=5, pady=5, sticky="ew")

		# Execute
		btn_frame = ttk.Frame(container)
		btn_frame.pack(fill="x", padx=10, pady=15)
		ttk.Button(btn_frame, text="変換行列を生成 & 表示", command=self.on_coordinate_creater_execute).pack(fill="x", ipady=5)

	def on_coordinate_creater_execute(self) -> None:
		"""Coordinate Creater実行"""
		try:
			fe = self.cc_fe.get()
			vv = self.cc_vv.get()
			ie = self.cc_ie.get()
			ml = self.cc_ml.get()
			ap = self.cc_ap.get()
			pd = self.cc_pd.get()
			
			# Build Matrix (Order should match _build_transform_matrix: rz(FE), rx(VV), ry(IE), ml, ap, pd)
			matrix = self._build_transform_matrix(fe, vv, ie, ml, ap, pd)
			
			self._show_coordinate_creater_result(matrix, fe, vv, ie, ml, ap, pd)
			
		except Exception as e:
			messagebox.showerror("エラー", f"値の取得または計算に失敗しました:\n{e}")

	def _show_coordinate_creater_result(self, matrix, fe, vv, ie, ml, ap, pd):
		res_win = tk.Toplevel(self)
		res_win.title("Coordinate Creater 結果")
		res_win.geometry("600x500")

		text_frame = ttk.Frame(res_win)
		text_frame.pack(fill="both", expand=True, padx=10, pady=10)

		# PosChecker形式に合わせる
		result_text = "【変換行列生成結果】\n"
		result_text += f"Input: FE={fe}, VV={vv}, IE={ie}, ML={ml}, AP={ap}, PD={pd}\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{matrix[0][0]:12.6f} {matrix[0][1]:12.6f} {matrix[0][2]:12.6f} {matrix[0][3]:12.6f}\n"
		result_text += f"{matrix[1][0]:12.6f} {matrix[1][1]:12.6f} {matrix[1][2]:12.6f} {matrix[1][3]:12.6f}\n"
		result_text += f"{matrix[2][0]:12.6f} {matrix[2][1]:12.6f} {matrix[2][2]:12.6f} {matrix[2][3]:12.6f}\n"
		result_text += f"{matrix[3][0]:12.6f} {matrix[3][1]:12.6f} {matrix[3][2]:12.6f} {matrix[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"
		
		# Analysis
		trans = matrix[:3, 3]
		rot_mat = matrix[:3, :3]
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0

		result_text += "【解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"

		st = tk.Text(text_frame, wrap="none", font=(self.mono_font_family, 10))
		st.pack(side="left", fill="both", expand=True)
		sb = ttk.Scrollbar(text_frame, orient="vertical", command=st.yview)
		sb.pack(side="right", fill="y")
		st.config(yscrollcommand=sb.set)
		st.insert("1.0", result_text)
		st.config(state="disabled")

		btn_frame = ttk.Frame(res_win)
		btn_frame.pack(fill="x", padx=10, pady=10)

		def save_txt():
			path = filedialog.asksaveasfilename(
				title="結果を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if path:
				try:
					with open(path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"保存しました:\n{path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(btn_frame, text="テキストとして保存...", command=save_txt).pack(side="left", padx=5)
		ttk.Button(btn_frame, text="閉じる", command=res_win.destroy).pack(side="right", padx=5)


	def _create_pp_merge_tab(self) -> None:
		"""pp mergeタブのコンテンツを作成"""
		container = self.pp_merge_tab
		container.columnconfigure(0, weight=1)

		# PPファイル入力エリア (最大6つ)
		list_frame = ttk.LabelFrame(container, text="統合するPPファイル (最大6つ)", style="Bold.TLabelframe")
		list_frame.pack(fill="both", expand=True, padx=5, pady=5)
		list_frame.columnconfigure(1, weight=1)

		self.pp_merge_vars = []
		
		# 6つの入力行を作成
		for i in range(6):
			var = tk.StringVar(value="")
			self.pp_merge_vars.append(var)
			
			row = i
			ttk.Label(list_frame, text=f"File {i+1}:").grid(row=row, column=0, sticky="w", padx=5, pady=5)
			
			entry = ttk.Entry(list_frame, textvariable=var)
			entry.grid(row=row, column=1, sticky="ew", padx=5, pady=5)
			
			# 参照ボタン
			ttk.Button(list_frame, text="参照...", width=8, 
					   command=lambda v=var: self._choose_pp_merge_file(v)).grid(row=row, column=2, sticky="e", padx=2, pady=5)
			
			# 削除（クリア）ボタン
			ttk.Button(list_frame, text="クリア", width=6,
					   command=lambda v=var: v.set("")).grid(row=row, column=3, sticky="e", padx=2, pady=5)

		# 実行ボタン
		ttk.Button(container, text="PPファイルを統合して保存...", command=self.on_pp_merge_execute).pack(fill="x", padx=5, pady=10)

	def _choose_pp_merge_file(self, target_var: tk.StringVar) -> None:
		path = filedialog.askopenfilename(
			title="PPファイルを選択",
			filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")]
		)
		if path:
			target_var.set(path)
			
	def on_pp_merge_execute(self) -> None:
		"""複数のPPファイルを統合して保存"""
		# 空でないパスを収集
		paths = [v.get().strip() for v in self.pp_merge_vars if v.get().strip()]
		
		if len(paths) < 2:
			messagebox.showwarning("入力不足", "少なくとも2つのPPファイルを選択してください。")
			return
			
		try:
			all_pts = []
			all_lbls = []
			
			print(f"[統合] {len(paths)}個のファイルを処理します...")
			
			for i, path in enumerate(paths):
				print(f"[統合] 読み込み中 ({i+1}/{len(paths)}): {path}")
				pts, lbls = self._parse_pp_file(path)
				all_pts.append(pts)
				all_lbls.extend(lbls)
				print(f" -> {len(pts)}点")
			
			# 統合
			if not all_pts:
				messagebox.showwarning("データなし", "有効な点データが見つかりませんでした。")
				return

			merged_pts = np.vstack(all_pts)
			merged_lbls = all_lbls
			print(f"[統合] 合計: {len(merged_pts)}点")
			
			# 保存先選択
			default_name = Path(paths[0]).stem + "_merged.pp"
			save_path = filedialog.asksaveasfilename(
				title="統合されたPPファイルを保存",
				defaultextension=".pp",
				filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")],
				initialfile=default_name
			)
			
			if not save_path:
				return
				
			# 保存
			save_child_pp_file(save_path, merged_pts, merged_lbls)
			messagebox.showinfo("完了", f"{len(paths)}個のPPファイルを統合しました:\n合計点数: {len(merged_pts)}\n\n保存先:\n{save_path}")
			
		except Exception as e:
			messagebox.showerror("エラー", f"PPファイルの統合に失敗しました:\n{e}")
			traceback.print_exc()

	def choose_randomizer_input(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルファイルを選択",
			filetypes=[("3D Model", "*.stl *.obj"), ("All files", "*.*")]
		)
		if path:
			self.randomizer_input_path.set(path)

	def on_randomizer_execute(self) -> None:
		"""ランダム変換を実行して保存"""
		input_path = self.randomizer_input_path.get().strip()
		if not input_path:
			messagebox.showwarning("警告", "モデルファイルを選択してください。")
			return
		
		if not os.path.exists(input_path):
			messagebox.showerror("エラー", f"ファイルが見つかりません:\n{input_path}")
			return

		# 保存先を選択
		save_path = filedialog.asksaveasfilename(
			title="変換後のモデルを保存",
			defaultextension=".stl",
			filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj"), ("All files", "*.*")],
			initialfile=Path(input_path).stem + "_random" + Path(input_path).suffix
		)
		
		if not save_path:
			return

		try:
			# パラメータ取得
			try:
				trans_range = self.random_trans_range.get()
				rot_range = self.random_rot_range.get()
			except tk.TclError:
				messagebox.showerror("エラー", "数値パラメータが不正です。")
				return

			# モデル読み込み
			mesh = pv.read(input_path)
			
			# ランダム変換生成
			# Translation
			t = np.random.uniform(-trans_range, trans_range, 3)
			
			# Rotation (Euler angles -> Rotation Matrix)
			# deg -> rad
			rx = np.deg2rad(np.random.uniform(-rot_range, rot_range))
			ry = np.deg2rad(np.random.uniform(-rot_range, rot_range))
			rz = np.deg2rad(np.random.uniform(-rot_range, rot_range))
			
			Rx = np.array([
				[1, 0, 0, 0],
				[0, np.cos(rx), -np.sin(rx), 0],
				[0, np.sin(rx), np.cos(rx), 0],
				[0, 0, 0, 1]
			])
			Ry = np.array([
				[np.cos(ry), 0, np.sin(ry), 0],
				[0, 1, 0, 0],
				[-np.sin(ry), 0, np.cos(ry), 0],
				[0, 0, 0, 1]
			])
			Rz = np.array([
				[np.cos(rz), -np.sin(rz), 0, 0],
				[np.sin(rz), np.cos(rz), 0, 0],
				[0, 0, 1, 0],
				[0, 0, 0, 1]
			])
			
			# Combined Rotation
			R = Rz @ Ry @ Rx
			
			# Translation Matrix
			T = np.eye(4)
			T[:3, 3] = t
			
			# Final Transform
			M = T @ R
			
			# Apply transform
			mesh.transform(M)
			
			# Save
			mesh.save(save_path)
			
			msg = (
				f"完了しました。\n\n"
				f"Translation (mm):\n  X={t[0]:.2f}, Y={t[1]:.2f}, Z={t[2]:.2f}\n"
				f"Rotation (deg):\n  X={np.rad2deg(rx):.1f}, Y={np.rad2deg(ry):.1f}, Z={np.rad2deg(rz):.1f}\n\n"
				f"Saved to:\n{save_path}"
			)
			messagebox.showinfo("成功", msg)
			
		except Exception as e:
			messagebox.showerror("エラー", f"処理に失敗しました:\n{e}\n{traceback.format_exc()}")

	# ----- UI Helpers -----
	def _setup_fonts(self) -> None:
		"""環境に合わせたUI/等幅フォントを選択し、Tkのデフォルトに適用する"""
		try:
			families = set(tkfont.families())
		except Exception:
			families = set()

		def pick(candidates, fallback=None):
			for fam in candidates:
				if fam in families:
					return fam
			return fallback or tkfont.nametofont("TkDefaultFont").cget("family")

		# 日本語表示を考慮した候補（プラットフォーム別に優先順位を設定）
		if IS_MACOS:
			ui_candidates = [
				"Hiragino Sans", "Hiragino Kaku Gothic Pro", "Hiragino Kaku Gothic ProN",
				".AppleSystemUIFont", "Helvetica Neue", "Lucida Grande", "Arial",
			]
			mono_candidates = [
				"Menlo", "Monaco", "SF Mono", "Osaka-Mono", "Courier New",
			]
		elif IS_WINDOWS:
			ui_candidates = [
				"Yu Gothic UI", "Meiryo UI", "Segoe UI", "メイリオ",
				"Yu Gothic", "MS UI Gothic", "MS Gothic", "Arial",
			]
			mono_candidates = [
				"Consolas", "MS Gothic", "MeiryoKe_Console", "Courier New",
			]
		else:  # Linux
			ui_candidates = [
				"Noto Sans CJK JP", "Noto Sans JP", "Takao Gothic",
				"IPAGothic", "VL Gothic", "DejaVu Sans", "Arial",
			]
			mono_candidates = [
				"Noto Sans Mono CJK JP", "DejaVu Sans Mono", "Monospace", "Courier New",
			]

		self.ui_font_family = pick(ui_candidates)
		self.mono_font_family = pick(mono_candidates, fallback=self.ui_font_family)

		# Tkのデフォルトフォント更新
		try:
			default_font = tkfont.nametofont("TkDefaultFont")
			default_font.configure(family=self.ui_font_family, size=11)
			for name in ("TkTextFont", "TkHeadingFont", "TkMenuFont"):
				try:
					f = tkfont.nametofont(name)
					f.configure(family=self.ui_font_family, size=11)
				except Exception:
					pass
			try:
				fixed = tkfont.nametofont("TkFixedFont")
				fixed.configure(family=self.mono_font_family, size=11)
			except Exception:
				pass
		except Exception:
			pass

	# ----- ドラッグ&ドロップ ヘルパー -----
	@staticmethod
	def _parse_dnd_path(data: str) -> str:
		"""ドロップされたデータからファイルパスを抽出する。
		macOS/Windows/Linuxのいずれの形式にも対応。
		"""
		raw = data.strip()
		# 複数ファイルの場合は最初の1つだけ使用
		# macOSでは中括弧で囲まれることがある: {/path/to/file}
		if raw.startswith("{") and "}" in raw:
			raw = raw[1:raw.index("}")]
		# 改行区切りの場合
		if "\n" in raw:
			raw = raw.split("\n")[0].strip()
		# スペースを含むパスが複数トークンに分割される場合への対処
		# file:// URI形式の場合
		if raw.startswith("file://"):
			raw = urllib.parse.unquote(raw[7:])
			# macOSでは file:///path の形式
			if raw.startswith("//"):
				raw = raw[2:]
		return raw

	def _register_dnd(self, widget, textvariable: tk.StringVar) -> None:
		"""ウィジェットにドラッグ&ドロップを登録する"""
		if not _HAS_DND:
			return
		try:
			widget.drop_target_register(DND_FILES)
			def on_drop(event):
				path = self._parse_dnd_path(event.data)
				if path and os.path.exists(path):
					textvariable.set(path)
			widget.dnd_bind("<<Drop>>", on_drop)
		except Exception:
			pass  # DnD非対応環境ではスキップ

	def _add_file_row(self, parent: ttk.Frame, row: int, label_text: str, textvariable: tk.StringVar, command) -> None:
		"""Add a file selection row with label, readonly entry (showing filename only), and browse button.
		ドラッグ&ドロップにも対応（tkinterdnd2が利用可能な場合）。
		"""
		ttk.Label(parent, text=label_text).grid(row=row, column=0, sticky="w", padx=12, pady=6)

		# Create a frame to hold the entry with custom display
		entry_frame = ttk.Frame(parent)
		entry_frame.grid(row=row, column=1, sticky="ew", padx=12, pady=6)

		# Create a StringVar that displays only the filename
		display_var = tk.StringVar(value="")

		# Update display when the actual variable changes
		def update_display(*args):
			full_path = textvariable.get()
			if full_path:
				# Show only filename
				display_var.set(Path(full_path).name)
			else:
				display_var.set("")

		textvariable.trace_add("write", update_display)
		update_display()  # Initialize

		entry = ttk.Entry(entry_frame, textvariable=display_var, state="readonly")
		entry.pack(fill="x", expand=True)

		# ドラッグ&ドロップ対応
		self._register_dnd(entry, textvariable)
		self._register_dnd(entry_frame, textvariable)

		# Add tooltip showing full path on hover
		def show_tooltip(event):
			full_path = textvariable.get()
			if full_path:
				# Create tooltip
				tooltip = tk.Toplevel()
				tooltip.wm_overrideredirect(True)
				tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
				label = ttk.Label(tooltip, text=full_path, background="lightyellow", relief="solid", borderwidth=1, padding=4)
				label.pack()
				entry.tooltip = tooltip

		def hide_tooltip(event):
			if hasattr(entry, 'tooltip'):
				entry.tooltip.destroy()
				del entry.tooltip

		entry.bind("<Enter>", show_tooltip)
		entry.bind("<Leave>", hide_tooltip)

		ttk.Button(parent, text="参照...", command=command).grid(row=row, column=2, sticky="e", padx=12, pady=6)

	# ----- File pickers -----
	def choose_prox_model(self) -> None:
		path = filedialog.askopenfilename(
			title="近位モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.prox_model_path.set(path)

	def choose_prox_pp_abcd(self) -> None:
		path = filedialog.askopenfilename(
			title="近位特徴点 ABCD (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.prox_pp_abcd_path.set(path)

	def choose_prox_pp_olmn(self) -> None:
		path = filedialog.askopenfilename(
			title="近位特徴点 OLMN (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.prox_pp_olmn_path.set(path)

	def choose_dist_model(self) -> None:
		path = filedialog.askopenfilename(
			title="遠位モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.dist_model_path.set(path)

	def choose_dist_pp_abc(self) -> None:
		path = filedialog.askopenfilename(
			title="遠位特徴点 ABC (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.dist_pp_abc_path.set(path)

	def choose_dist_pp_olmn(self) -> None:
		path = filedialog.askopenfilename(
			title="遠位特徴点 OLMN (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.dist_pp_olmn_path.set(path)

	def choose_prox_cartilage_model(self) -> None:
		"""近位軟骨モデルを選択"""
		path = filedialog.askopenfilename(
			title="近位軟骨モデルを選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			self.prox_cartilage_model_path.set(path)

	def choose_dist_cartilage_model(self) -> None:
		"""遠位軟骨モデルを選択"""
		path = filedialog.askopenfilename(
			title="遠位軟骨モデルを選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			self.dist_cartilage_model_path.set(path)

	# ----- File pickers (Fitting) -----
	def choose_fitting_parent_model(self) -> None:
		path = filedialog.askopenfilename(
			title="親モデルファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_parent_model_path.set(path)

	def choose_fitting_parent_region(self) -> None:
		path = filedialog.askopenfilename(
			title="親任意領域ファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_parent_region_path.set(path)

	def choose_fitting_child_model(self) -> None:
		path = filedialog.askopenfilename(
			title="子モデルファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_model_path.set(path)

	def choose_fitting_child_region(self) -> None:
		path = filedialog.askopenfilename(
			title="子任意領域ファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_region_path.set(path)

	def choose_fitting_child_pp(self) -> None:
		path = filedialog.askopenfilename(
			title="子特徴点 (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_pp_path.set(path)

	def choose_fitting_child_heatmap_model(self) -> None:
		path = filedialog.askopenfilename(
			title="子ヒートマップ表示用モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_heatmap_model_path.set(path)

	def choose_transform_group(self) -> None:
		path = filedialog.askopenfilename(
			title="同次変換行列群ファイルを選択",
			filetypes=[
				("すべてのファイル", "*.*"),
				("Excelファイル", "*.xlsx"),
				("KKRファイル", "*.KKR *.kkr"),
			],
		)
		if path:
			# 拡張子を確認
			ext = Path(path).suffix.lower()
			if ext not in ['.xlsx', '.kkr']:
				messagebox.showwarning("ファイル形式エラー", "選択されたファイル形式はサポートされていません。\n.xlsx または .kkr ファイルを選択してください。")
				return
			
			self.transform_group_path.set(path)

	# ----- File pickers (ORG) -----
	def choose_org_model(self) -> None:
		path = filedialog.askopenfilename(
			title="RobotScan モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.org_model_path.set(path)

	def choose_org_pp(self) -> None:
		path = filedialog.askopenfilename(
			title="RobotScan 座標ポイント (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.org_pp_path.set(path)

	# ----- File pickers (A-B) -----
	def choose_a_b_model_a(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルA (STL) を選択",
			filetypes=[
				("STL", "*.stl"),
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.a_b_model_a_path.set(path)

	def choose_a_b_model_b(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルB (STL) を選択",
			filetypes=[
				("STL", "*.stl"),
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.a_b_model_b_path.set(path)

	# ----- File pickers (HeatMap) -----
	def choose_heatmap_model_a(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルA (参照) を選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.heatmap_model_a_path.set(path)

	def choose_heatmap_model_b(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルB (距離計算対象) を選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.heatmap_model_b_path.set(path)

	# ----- Cartilage Separation: Generic File Picker -----
	def _choose_cs_file(self, side: str, which: str) -> None:
		"""軟骨分離用ファイル選択（side='prox'|'dist', which='m1w'|'m1r'|'m2w'|'m2r'）"""
		titles = {
			"m1w": "骨＋軟骨 全体モデル", "m1r": "骨＋軟骨 任意領域",
			"m2w": "骨のみ 全体モデル", "m2r": "骨のみ 任意領域",
		}
		var_map = {
			("prox", "m1w"): self.cs_prox_model1_whole_path, ("prox", "m1r"): self.cs_prox_model1_region_path,
			("prox", "m2w"): self.cs_prox_model2_whole_path, ("prox", "m2r"): self.cs_prox_model2_region_path,
			("dist", "m1w"): self.cs_dist_model1_whole_path, ("dist", "m1r"): self.cs_dist_model1_region_path,
			("dist", "m2w"): self.cs_dist_model2_whole_path, ("dist", "m2r"): self.cs_dist_model2_region_path,
		}
		side_label = "近位" if side == "prox" else "遠位"
		path = filedialog.askopenfilename(
			title=f"{side_label}: {titles[which]}を選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			var_map[(side, which)].set(path)

	def _update_cs_side_buttons(self, side: str) -> None:
		"""近位/遠位ボタンの活性制御"""
		if side == "prox":
			m1w = bool(self.cs_prox_model1_whole_path.get().strip())
			m1r = bool(self.cs_prox_model1_region_path.get().strip())
			m2w = bool(self.cs_prox_model2_whole_path.get().strip())
			m2r = bool(self.cs_prox_model2_region_path.get().strip())
			try: self.cs_prox_viz_btn.config(state="normal" if m1w else "disabled")
			except: pass
			try: self.cs_prox_exec_btn.config(state="normal" if (m1w and m1r and m2w and m2r) else "disabled")
			except: pass
		else:
			m1w = bool(self.cs_dist_model1_whole_path.get().strip())
			m1r = bool(self.cs_dist_model1_region_path.get().strip())
			m2w = bool(self.cs_dist_model2_whole_path.get().strip())
			m2r = bool(self.cs_dist_model2_region_path.get().strip())
			try: self.cs_dist_viz_btn.config(state="normal" if m1w else "disabled")
			except: pass
			try: self.cs_dist_exec_btn.config(state="normal" if (m1w and m1r and m2w and m2r) else "disabled")
			except: pass

	def _on_cs_visualize_side(self, side: str) -> None:
		"""近位/遠位モデルの確認可視化"""
		if side == "prox":
			w_path = self.cs_prox_model1_whole_path.get().strip()
			r_path = self.cs_prox_model1_region_path.get().strip()
			w2_path = self.cs_prox_model2_whole_path.get().strip()
		else:
			w_path = self.cs_dist_model1_whole_path.get().strip()
			r_path = self.cs_dist_model1_region_path.get().strip()
			w2_path = self.cs_dist_model2_whole_path.get().strip()
		side_label = "近位（骨盤）" if side == "prox" else "遠位（大腿骨）"
		if not w_path:
			messagebox.showwarning("入力不足", "モデル1の全体モデルを選択してください。")
			return
		try:
			plotter = pv.Plotter(title=f"{side_label} モデル確認")
			plotter.set_background('white')
			plotter.add_mesh(pv.read(w_path), color='lightblue', opacity=0.5, show_edges=False, label='骨+軟骨 全体')
			if r_path:
				plotter.add_mesh(pv.read(r_path), color='red', opacity=0.8, show_edges=False, label='任意領域')
			if w2_path:
				plotter.add_mesh(pv.read(w2_path), color='wheat', opacity=0.3, show_edges=False, label='骨のみ 全体')
			plotter.add_axes(); plotter.add_legend(); plotter.show()
			self._restore_window_geometry()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{str(e)}")

	def _on_cs_execute_side(self, side: str) -> None:
		"""近位/遠位の軟骨分離を実行（ブーリアン差分）"""
		if side == "prox":
			m1w_path = self.cs_prox_model1_whole_path.get().strip()
			m1r_path = self.cs_prox_model1_region_path.get().strip()
			m2w_path = self.cs_prox_model2_whole_path.get().strip()
			m2r_path = self.cs_prox_model2_region_path.get().strip()
		else:
			m1w_path = self.cs_dist_model1_whole_path.get().strip()
			m1r_path = self.cs_dist_model1_region_path.get().strip()
			m2w_path = self.cs_dist_model2_whole_path.get().strip()
			m2r_path = self.cs_dist_model2_region_path.get().strip()

		side_label = "近位（骨盤）" if side == "prox" else "遠位（大腿骨）"
		# 既存のon_cs_executeを呼び出す（パスを一時的にprox変数にセット）
		# on_cs_executeはcs_prox_*を参照するので、dist実行時はswap
		if side == "dist":
			# 一時的にprox変数にdist値をセット
			save = [self.cs_prox_model1_whole_path.get(), self.cs_prox_model1_region_path.get(),
					self.cs_prox_model2_whole_path.get(), self.cs_prox_model2_region_path.get()]
			self.cs_prox_model1_whole_path.set(m1w_path)
			self.cs_prox_model1_region_path.set(m1r_path)
			self.cs_prox_model2_whole_path.set(m2w_path)
			self.cs_prox_model2_region_path.set(m2r_path)

		self.on_cs_execute()

		if side == "dist":
			# 元に戻す
			self.cs_prox_model1_whole_path.set(save[0])
			self.cs_prox_model1_region_path.set(save[1])
			self.cs_prox_model2_whole_path.set(save[2])
			self.cs_prox_model2_region_path.set(save[3])

	def _choose_cs_bone_color(self) -> None:
		"""骨の色を選択"""
		color = colorchooser.askcolor(initialcolor=self.cs_bone_color, title="骨の表示色を選択")
		if color[1]:
			self.cs_bone_color = color[1]

	def _choose_cs_cartilage_color(self) -> None:
		"""軟骨の色を選択"""
		color = colorchooser.askcolor(initialcolor=self.cs_cartilage_color, title="軟骨の表示色を選択")
		if color[1]:
			self.cs_cartilage_color = color[1]

	# --- フィッティングパラメータ デフォルト保存/復元 ---

	_CS_FITTING_DEFAULTS_BUILTIN = {
		"ransac_distance_threshold": 1.0,
		"ransac_max_iterations": 1000,
		"ransac_confidence": 0.99,
		"icp_threshold": 5.0,
		"icp_max_iterations": 2000,
		"sample_points": 10000,
		"distance_threshold": 0.5,
	}

	def _cs_fitting_defaults_path(self) -> Path:
		"""ユーザーデフォルト保存先"""
		return Path(__file__).parent / ".frs_fitting_defaults.json"

	def _load_cs_fitting_defaults(self) -> dict:
		"""ユーザーデフォルトを読み込み。なければビルトインを返す"""
		p = self._cs_fitting_defaults_path()
		if p.exists():
			try:
				with p.open("r", encoding="utf-8") as f:
					data = json.load(f)
				# ビルトインのキーが全て揃っているか補完
				merged = dict(self._CS_FITTING_DEFAULTS_BUILTIN)
				merged.update(data)
				return merged
			except Exception:
				pass
		return dict(self._CS_FITTING_DEFAULTS_BUILTIN)

	def _save_cs_fitting_defaults(self) -> None:
		"""現在のフィッティングパラメータをユーザーデフォルトとして保存"""
		data = {
			"ransac_distance_threshold": self.cs_ransac_distance_threshold.get(),
			"ransac_max_iterations": self.cs_ransac_max_iterations.get(),
			"ransac_confidence": self.cs_ransac_confidence.get(),
			"icp_threshold": self.cs_icp_threshold.get(),
			"icp_max_iterations": self.cs_icp_max_iterations.get(),
			"sample_points": self.cs_sample_points.get(),
			"distance_threshold": self.cs_distance_threshold.get(),
		}
		try:
			p = self._cs_fitting_defaults_path()
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
			messagebox.showinfo("保存完了",
				"現在のフィッティングパラメータをデフォルトとして保存しました。\n\n"
				f"RANSAC 距離閾値: {data['ransac_distance_threshold']}\n"
				f"RANSAC 最大反復: {data['ransac_max_iterations']}\n"
				f"RANSAC 信頼度: {data['ransac_confidence']}\n"
				f"ICP 閾値: {data['icp_threshold']}\n"
				f"ICP 最大反復: {data['icp_max_iterations']}\n"
				f"サンプル点数: {data['sample_points']}\n"
				f"距離閾値: {data['distance_threshold']}")
		except Exception as e:
			messagebox.showerror("保存エラー", f"デフォルト保存に失敗しました:\n{e}")

	def reset_cs_fitting_parameters(self) -> None:
		"""フィッティングパラメータをユーザーデフォルト（なければビルトイン）に戻す"""
		d = self._load_cs_fitting_defaults()
		self.cs_ransac_distance_threshold.set(d["ransac_distance_threshold"])
		self.cs_ransac_max_iterations.set(int(d["ransac_max_iterations"]))
		self.cs_ransac_confidence.set(d["ransac_confidence"])
		self.cs_icp_threshold.set(d["icp_threshold"])
		self.cs_icp_max_iterations.set(int(d["icp_max_iterations"]))
		self.cs_sample_points.set(int(d["sample_points"]))
		self.cs_distance_threshold.set(d["distance_threshold"])

	def on_cs_execute(self) -> None:
		"""
		【軟骨分離】メイン実行メソッド
		
		■ 処理フロー:
		  1. 任意領域同士を RANSAC + ICP でフィッティング → 変換行列T取得
		  2. T をモデル2（骨のみ）全体に適用 → モデル1と同じ座標系に
		  3. ブーリアン差分: モデル1（骨+軟骨） − モデル2（骨のみ） = 軟骨（＋関節唇）
		  4. 結果を可視化＋STL保存オプション
		
		■ 前提条件:
		  すべてのモデルは体積メッシュ（内部にもメッシュがある）であること。
		  FEMに使用するため、表面だけでなく内部も埋まっている必要がある。
		"""
		# 入力チェック
		if not all([
			self.cs_prox_model1_whole_path.get().strip(),
			self.cs_prox_model1_region_path.get().strip(),
			self.cs_prox_model2_whole_path.get().strip(),
			self.cs_prox_model2_region_path.get().strip(),
		]):
			messagebox.showerror("エラー", "4つのモデルファイルをすべて選択してください。")
			return
		
		try:
			import copy
			import trimesh  # type: ignore
			
			def load_mesh_o3d(file_path):
				"""メッシュをOpen3Dで読み込む"""
				try:
					print(f"ファイル読み込み中: {Path(file_path).name}")
					tm = trimesh.load(file_path, force='mesh')
					vertices = np.array(tm.vertices)
					triangles = np.array(tm.faces)
					mesh = o3d.geometry.TriangleMesh()
					mesh.vertices = o3d.utility.Vector3dVector(vertices)
					mesh.triangles = o3d.utility.Vector3iVector(triangles)
					mesh.compute_vertex_normals()
					print(f"  ✓ 成功: 頂点数={len(vertices)}, 三角形数={len(triangles)}")
					return mesh
				except Exception as e:
					print(f"  ✗ エラー: {e}")
					messagebox.showerror("エラー", f"ファイル読み込み失敗:\n{Path(file_path).name}\n\n{str(e)}")
					return None
			
			print("\n" + "="*60)
			print("=== 軟骨分離処理開始 ===")
			print("="*60)
			print(f"モデル1全体: {Path(self.cs_prox_model1_whole_path.get()).name}")
			print(f"モデル1任意領域: {Path(self.cs_prox_model1_region_path.get()).name}")
			print(f"モデル2全体: {Path(self.cs_prox_model2_whole_path.get()).name}")
			print(f"モデル2任意領域: {Path(self.cs_prox_model2_region_path.get()).name}")
			
			# --- Phase 1: 任意領域の RANSAC + ICP フィッティング ---
			print("\n--- Phase 1: 任意領域のフィッティング ---")
			
			region1_mesh = load_mesh_o3d(self.cs_prox_model1_region_path.get())
			region2_mesh = load_mesh_o3d(self.cs_prox_model2_region_path.get())
			
			if not all([region1_mesh, region2_mesh]):
				return
			
			# 色設定して初期状態を表示
			region1_mesh.paint_uniform_color([1.0, 0.0, 0.0])  # 赤 = モデル1任意領域
			region2_mesh.paint_uniform_color([0.0, 0.0, 1.0])  # 青 = モデル2任意領域
			
			print("\n[初期状態] 任意領域を表示中...")
			o3d.visualization.draw_geometries(
				[copy.deepcopy(region1_mesh), copy.deepcopy(region2_mesh)],
				window_name="初期状態: モデル1任意領域(赤) + モデル2任意領域(青)")
			self._restore_window_geometry()
			
			# 主軸による事前アライメント
			print("\n=== 主軸による事前アライメント ===")
			
			def get_principal_axis(mesh):
				pcd = mesh.sample_points_uniformly(number_of_points=20000)
				pts = np.asarray(pcd.points)
				center = np.mean(pts, axis=0)
				cov = np.cov((pts - center).T)
				eigvals, eigvecs = np.linalg.eigh(cov)
				return center, eigvecs[:, np.argmax(eigvals)]
			
			def align_axis(mesh, src_axis, tgt_axis):
				v = np.cross(src_axis, tgt_axis)
				c = np.dot(src_axis, tgt_axis)
				if np.linalg.norm(v) < 1e-8:
					R = np.eye(3)
				else:
					vx = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
					R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
				mesh_rot = copy.deepcopy(mesh)
				mesh_rot.rotate(R, center=(0, 0, 0))
				return mesh_rot, R
			
			c_r1, axis_r1 = get_principal_axis(region1_mesh)
			c_r2, axis_r2 = get_principal_axis(region2_mesh)
			
			if np.dot(axis_r1, axis_r2) < 0:
				axis_r2 = -axis_r2
				print("主軸の向きを反転しました")
			
			region2_aligned, R_align = align_axis(region2_mesh, axis_r2, axis_r1)
			c_r2_aligned, _ = get_principal_axis(region2_aligned)
			translation_vector = c_r1 - c_r2_aligned
			region2_aligned.translate(translation_vector, relative=True)
			
			# 主軸アライメントの変換行列
			axis_transform = np.eye(4)
			axis_transform[:3, :3] = R_align
			axis_transform[:3, 3] = c_r1 - R_align @ c_r2
			
			print(f"主軸アライメント変換行列:\n{axis_transform}")
			
			# 主軸アライメント後を表示
			print("\n[主軸アライメント後] 表示中...")
			r1_disp = copy.deepcopy(region1_mesh)
			r2_disp = copy.deepcopy(region2_aligned)
			r1_disp.paint_uniform_color([1.0, 0.0, 0.0])
			r2_disp.paint_uniform_color([0.0, 1.0, 0.0])
			o3d.visualization.draw_geometries(
				[r1_disp, r2_disp],
				window_name="主軸・重心一致後: モデル1(赤) + モデル2(緑)")
			self._restore_window_geometry()
			
			# RANSAC
			print("\n=== RANSAC位置合わせ ===")
			parent_pcd = region1_mesh.sample_points_uniformly(number_of_points=self.cs_sample_points.get())
			child_pcd = region2_aligned.sample_points_uniformly(number_of_points=self.cs_sample_points.get())
			
			# voxel_size自動推定
			try:
				num_points = len(parent_pcd.points)
				sample_n = min(500, max(10, num_points))
				sample_indices = np.random.choice(num_points, sample_n, replace=False)
				kdt = o3d.geometry.KDTreeFlann(parent_pcd)
				nn_dists = []
				for idx in sample_indices:
					[_, idxs, dists] = kdt.search_knn_vector_3d(parent_pcd.points[idx], 2)
					if len(dists) >= 2:
						nn_dists.append(float(np.sqrt(dists[1])))
				voxel_size = max(float(np.mean(nn_dists)) * 1.5, 1e-6) if nn_dists else max(self.cs_ransac_distance_threshold.get(), 1e-6)
			except Exception:
				voxel_size = max(self.cs_ransac_distance_threshold.get(), 1e-6)
			
			print(f"voxel_size={voxel_size:.6f}")
			
			parent_down = parent_pcd.voxel_down_sample(voxel_size)
			child_down = child_pcd.voxel_down_sample(voxel_size)
			
			radius_normal = voxel_size * 2
			parent_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			child_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			
			radius_feature = voxel_size * 5
			parent_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				parent_down, o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			child_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				child_down, o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			
			dist_threshold_ransac = self.cs_ransac_distance_threshold.get()
			result_ransac = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
				child_down, parent_down, child_fpfh, parent_fpfh,
				mutual_filter=True,
				max_correspondence_distance=dist_threshold_ransac,
				estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPoint(self.cs_enable_scaling.get()),
				ransac_n=3,
				checkers=[
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(0.9),
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(dist_threshold_ransac)
				],
				criteria=o3d.pipelines.registration.RANSACConvergenceCriteria(
					self.cs_ransac_max_iterations.get(), self.cs_ransac_confidence.get())
			)
			
			ransac_transform = result_ransac.transformation
			print(f"RANSAC fitness={result_ransac.fitness:.6f}, RMSE={result_ransac.inlier_rmse:.6f}")
			
			# RANSAC後を表示
			region2_ransac = copy.deepcopy(region2_aligned)
			region2_ransac.transform(ransac_transform)
			
			print("\n[RANSAC後] 表示中...")
			r1_disp = copy.deepcopy(region1_mesh)
			r2_disp = copy.deepcopy(region2_ransac)
			r1_disp.paint_uniform_color([1.0, 0.0, 0.0])
			r2_disp.paint_uniform_color([1.0, 1.0, 0.0])
			o3d.visualization.draw_geometries(
				[r1_disp, r2_disp],
				window_name="RANSAC後: モデル1(赤) + モデル2(黄)")
			self._restore_window_geometry()
			
			# ICP
			print("\n=== ICP精密位置合わせ ===")
			result_icp = o3d.pipelines.registration.registration_icp(
				child_down, parent_down,
				self.cs_icp_threshold.get(),
				ransac_transform,
				o3d.pipelines.registration.TransformationEstimationPointToPoint(self.cs_enable_scaling.get()),
				o3d.pipelines.registration.ICPConvergenceCriteria(max_iteration=self.cs_icp_max_iterations.get())
			)
			
			icp_transform = result_icp.transformation
			print(f"ICP fitness={result_icp.fitness:.6f}, RMSE={result_icp.inlier_rmse:.6f}")

			# スケール補正が有効な場合、スケール係数をログ出力
			if self.cs_enable_scaling.get():
				scale_factor = np.cbrt(np.linalg.det(np.array(icp_transform[:3, :3])))
				print(f"  スケール補正係数: {scale_factor:.6f} ({(scale_factor-1)*100:.2f}%)")
			
			# 最終変換行列 = ICP @ axis_transform （元の座標系からの合成変換）
			final_transform = icp_transform @ axis_transform
			print(f"\n最終変換行列:\n{final_transform}")
			
			# ICP後を表示
			region2_final = copy.deepcopy(region2_mesh)
			region2_final.transform(final_transform)
			
			print("\n[ICP後] 表示中...")
			r1_disp = copy.deepcopy(region1_mesh)
			r2_disp = copy.deepcopy(region2_final)
			r1_disp.paint_uniform_color([1.0, 0.0, 0.0])
			r2_disp.paint_uniform_color([0.0, 1.0, 0.0])
			o3d.visualization.draw_geometries(
				[r1_disp, r2_disp],
				window_name="ICP後（最終）: モデル1(赤) + モデル2(緑)")
			self._restore_window_geometry()

			# --- フィッティング確認 → キャンセル可能 ---
			scale_info = ""
			if self.cs_enable_scaling.get():
				sf = np.cbrt(np.linalg.det(np.array(icp_transform[:3, :3])))
				scale_info = f"\n  スケール補正: {sf:.6f} ({(sf-1)*100:.2f}%)\n"
			proceed = messagebox.askyesno("フィッティング確認",
				f"位置合わせの結果を確認してください。\n\n"
				f"  ICP fitness: {result_icp.fitness:.6f}\n"
				f"  ICP RMSE:    {result_icp.inlier_rmse:.6f}\n"
				f"{scale_info}\n"
				f"このまま軟骨分離を実行しますか？\n"
				f"（フィッティングが不十分な場合は「いいえ」でキャンセル）")
			if not proceed:
				print("ユーザーがキャンセルしました。")
				messagebox.showinfo("キャンセル", "軟骨分離をキャンセルしました。\nパラメータを調整して再実行してください。")
				return

			# --- Phase 2: モデル2全体を変換 ---
			print("\n--- Phase 2: モデル2全体に変換行列を適用 ---")
			
			model1_whole = pv.read(self.cs_prox_model1_whole_path.get())
			model2_whole = pv.read(self.cs_prox_model2_whole_path.get())
			
			# 変換を適用
			model2_transformed = model2_whole.copy()
			model2_transformed.transform(final_transform, inplace=True)
			
			print(f"モデル1: {model1_whole.n_points}点, {model1_whole.n_cells}面")
			print(f"モデル2（変換後）: {model2_transformed.n_points}点, {model2_transformed.n_cells}面")
			
			# --- Phase 3: 距離ベース軟骨抽出 ---
			print("\n--- Phase 3: 軟骨領域の抽出（距離ベース）---")
			print("  モデル1(骨+軟骨)の表面から、モデル2(骨)の外側にある領域を軟骨として抽出します。")

			cartilage_mesh = None

			try:
				# 両モデルの表面メッシュを取得
				m1_surface = model1_whole.extract_surface()
				m2_surface = model2_transformed.extract_surface()

				print(f"  モデル1 表面: {m1_surface.n_points}点, {m1_surface.n_cells}面")
				print(f"  モデル2 表面: {m2_surface.n_points}点, {m2_surface.n_cells}面")

				# --- VTK implicit distance による符号付き距離計算 ---
				# vtkImplicitPolyDataDistance: 幾何学的に正確な符号付き距離
				# 正 = Model2の外側（軟骨）、負 = Model2の内側（骨）
				print("  VTK implicit distance で符号付き距離を計算中...")
				m1_with_dist = m1_surface.compute_implicit_distance(m2_surface, inplace=False)
				signed_distances = m1_with_dist['implicit_distance']

				# 統計出力
				n_outside = np.sum(signed_distances > 0)
				n_inside = np.sum(signed_distances <= 0)
				print(f"  外側（軟骨候補）: {n_outside}点 ({100*n_outside/max(1,m1_surface.n_points):.1f}%)")
				print(f"  内側（骨）:       {n_inside}点 ({100*n_inside/max(1,m1_surface.n_points):.1f}%)")
				if n_outside > 0:
					outside_dists = signed_distances[signed_distances > 0]
					print(f"  外側距離: 平均 {outside_dists.mean():.3f}mm, 最大 {outside_dists.max():.3f}mm")
				if n_inside > 0:
					inside_dists = signed_distances[signed_distances <= 0]
					print(f"  内側距離: 平均 {inside_dists.mean():.3f}mm, 最小 {inside_dists.min():.3f}mm")

				# 閾値適用: signed_distance > threshold の頂点を軟骨として抽出
				threshold = self.cs_distance_threshold.get()
				mask = signed_distances > threshold
				n_selected = np.sum(mask)
				print(f"\n  距離 > {threshold:.2f}mm の頂点: {n_selected}/{m1_surface.n_points} ({100*n_selected/max(1,m1_surface.n_points):.1f}%)")

				if n_selected > 0:
					# 頂点マスクからセル（面）を抽出
					cartilage_mesh = m1_surface.extract_points(mask, adjacent_cells=True)
					cartilage_mesh = cartilage_mesh.extract_surface()

					if cartilage_mesh.n_cells > 0:
						# クリーンアップ: 小さい断片を除去
						try:
							connected = cartilage_mesh.connectivity(largest=False)
							if 'RegionId' in connected.array_names:
								region_ids = connected['RegionId']
								unique_ids, counts = np.unique(region_ids, return_counts=True)
								min_cells = max(10, cartilage_mesh.n_cells * 0.01)
								large_regions = unique_ids[counts >= min_cells]
								if len(large_regions) > 0 and len(large_regions) < len(unique_ids):
									keep_mask = np.isin(region_ids, large_regions)
									cartilage_mesh = connected.extract_points(keep_mask, adjacent_cells=True).extract_surface()
									removed = len(unique_ids) - len(large_regions)
									print(f"  小断片 {removed}個を除去")
						except Exception:
							pass

						print(f"  ✓ 軟骨抽出 成功: {cartilage_mesh.n_points}点, {cartilage_mesh.n_cells}面")
					else:
						cartilage_mesh = None
						print("  ✗ セル抽出結果が空")
				else:
					print("  ✗ 閾値を超える頂点なし（閾値を小さくしてみてください）")

			except Exception as e_dist:
				print(f"  ✗ 距離ベース抽出 失敗: {e_dist}")
				import traceback; traceback.print_exc()

			if cartilage_mesh is None or cartilage_mesh.n_cells == 0:
				messagebox.showerror("エラー",
					"軟骨領域の抽出に失敗しました。\n\n"
					"考えられる対処:\n"
					"・「距離閾値」を小さくする（現在: {:.2f}mm → 0.1mm等）\n"
					"・位置合わせを改善（RANSAC/ICPパラメータを調整）\n"
					"・「スケール補正」を有効にする".format(self.cs_distance_threshold.get()))
				return

			print(f"  ✓ 軟骨＋関節唇: {cartilage_mesh.n_points}点, {cartilage_mesh.n_cells}面")
			
			# --- Phase 4: 結果の可視化 ---
			print("\n--- Phase 4: 結果可視化 ---")
			
			plotter = pv.Plotter()
			plotter.set_background('white')
			
			cart_opacity = self.cs_cartilage_opacity.get()
			bone_opacity = self.cs_bone_opacity.get()
			
			# 軟骨（差分結果）
			plotter.add_mesh(cartilage_mesh, color=self.cs_cartilage_color, opacity=cart_opacity,
				show_edges=False, label=f'軟骨+関節唇 ({cartilage_mesh.n_cells}面)')
			
			# モデル2（座標一致済み・骨のみ）半透明表示
			plotter.add_mesh(model2_transformed, color=self.cs_bone_color, opacity=bone_opacity,
				show_edges=False, label=f'骨のみモデル ({model2_transformed.n_cells}面)')
			
			# モデル1全体を薄く重ねて確認用
			plotter.add_mesh(model1_whole, color='gray', opacity=0.08,
				show_edges=False, label='モデル1全体（参照）')
			
			plotter.add_axes()
			plotter.add_legend()
			plotter.add_text("ブーリアン差分: モデル1 − モデル2", position='upper_right', font_size=10)
			plotter.show()
			# PyVistaリソースを明示解放（ウィンドウ閉じた後のフリーズを軽減）
			plotter.close()
			del plotter
			import gc; gc.collect()
			self._restore_window_geometry()
			self.update_idletasks()  # tkinterイベントを処理してUI応答性を回復

			# --- Phase 5: 保存オプション ---
			print("\n--- Phase 5: 保存 ---")
			
			want_save = messagebox.askyesno("保存確認",
				f"分離結果を保存しますか？\n\n"
				f"・軟骨（＋関節唇）: {cartilage_mesh.n_cells}面\n"
				f"・モデル2（座標系一致済み）: {model2_transformed.n_cells}面")
			
			if want_save:
				# 軟骨メッシュを保存
				original = Path(self.cs_prox_model1_whole_path.get())
				default_name = original.stem + "_cartilage.stl"
				save_path = filedialog.asksaveasfilename(
					title="軟骨モデルを保存",
					defaultextension=".stl",
					initialfile=default_name,
					filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj"), ("All files", "*.*")]
				)
				if save_path:
					cartilage_mesh.save(save_path)
					print(f"軟骨モデルを保存: {Path(save_path).name}")
				
				# 座標系一致済みのモデル2を保存
				want_m2 = messagebox.askyesno("座標系一致済みモデル保存",
					"モデル2（骨のみ）の座標系一致済みモデルも保存しますか？\n"
					"（FEMで骨と軟骨に異なる物性値を設定する際に使用）")
				if want_m2:
					original2 = Path(self.cs_prox_model2_whole_path.get())
					default_name2 = original2.stem + "_aligned.stl"
					save_path2 = filedialog.asksaveasfilename(
						title="座標系一致済みモデル2を保存",
						defaultextension=".stl",
						initialfile=default_name2,
						filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj"), ("All files", "*.*")]
					)
					if save_path2:
						model2_transformed.save(save_path2)
						print(f"座標系一致済みモデル2を保存: {Path(save_path2).name}")
			
			print("\n=== 軟骨分離処理完了 ===")
		
		except Exception as e:
			messagebox.showerror("エラー", f"軟骨分離中にエラーが発生しました:\n{str(e)}")
			print(f"エラー詳細: {e}")
			traceback.print_exc()

	# ----- Actions (Dummy) -----
	def on_visualize_prox(self) -> None:
		# 近位モデルと近位PPを表示（PyVista）
		model_path = self.prox_model_path.get().strip()
		pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		if not model_path or not pp_abcd_path or not pp_olmn_path:
			messagebox.showwarning("入力不足", "近位モデルと近位特徴点(ABCD, OLMN)を選択してください。")
			return

		# ライブラリの確認（遅延インポート）
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return

		# PP読み込み（PickedPoints XML対応）- 2つのファイルを統合
		try:
			points_abcd, labels_abcd = self._parse_pp_file(pp_abcd_path)
			points_olmn, labels_olmn = self._parse_pp_file(pp_olmn_path)
			points = np.vstack([points_abcd, points_olmn])
			labels = labels_abcd + labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 座標系構築（A, B, C, D, O点から）
		try:
			origin, x_axis, y_axis, z_axis = self._build_coordinate_system(points, labels)
			# 自動オフセット設定（O'点がある場合）
			self._try_auto_set_offset(points, labels, origin, x_axis, y_axis, z_axis,
			                          self.prox_offset_x, self.prox_offset_y, self.prox_offset_z)
		except Exception as e:
			messagebox.showwarning("座標系構築", f"座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			origin, x_axis, y_axis, z_axis = None, None, None, None

		# メッシュ読み込み（PyVista）
		try:
			mesh_full = pv.read(model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"近位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# 球体で寛骨臼領域を抽出（座標系が構築できた場合）
		mesh_extracted = None
		origin_prime = None
		if origin is not None:
			try:
				# O'原点位置を計算（ローカル座標系に沿ったオフセット）
				origin_prime = origin + (
					self.prox_offset_x.get() * x_axis +
					self.prox_offset_y.get() * y_axis +
					self.prox_offset_z.get() * z_axis
				)
				print(f"[可視化] 近位O: {origin}, 近位O': {origin_prime}, オフセット: [{self.prox_offset_x.get()}, {self.prox_offset_y.get()}, {self.prox_offset_z.get()}]")
				
				radius = self.prox_radius.get()
				mesh_extracted = self._extract_region_by_sphere(mesh_full, origin_prime, radius)
				if mesh_extracted.n_points == 0:
					messagebox.showwarning("抽出失敗", f"半径 {radius}mm の球体内にメッシュが見つかりませんでした。")
					mesh_extracted = None
			except Exception as e:
				messagebox.showwarning("球体抽出", f"球体抽出に失敗しました:\n{e}\n\n元のメッシュのみ表示します。")
				mesh_extracted = None

		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		# ウィンドウサイズをディスプレイの90%に設定
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)

		# プロッタ作成（格子/目盛りなし、モデル+点のみ）
		plotter = pv.Plotter(title="近位モデル/特徴点の可視化", window_size=(window_width, window_height))
		plotter.set_background("white")
		
		# 元のメッシュ全体（半透明・グレー）
		mesh_full_actor = plotter.add_mesh(mesh_full, color="lightgray", smooth_shading=True, show_edges=False, opacity=0.3)
		
		# 抽出された領域（設定したカラー・不透明）
		mesh_extracted_actor = None
		if mesh_extracted is not None and mesh_extracted.n_points > 0:
			mesh_extracted_actor = plotter.add_mesh(mesh_extracted, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=1.0)
		
		# 点群（設定したカラーを使用）
		points_actor = plotter.add_points(points, color=self.prox_color, point_size=12, render_points_as_spheres=True)
		
		# 点のラベルを表示
		label_actors = []
		for i, (pt, label) in enumerate(zip(points, labels)):
			label_actor = plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.prox_color, bold=True, shadow=True, show_points=False)
			label_actors.append(label_actor)
		
		# 座標系の表示（構築できた場合）
		axis_actors = []
		axis_prime_actors = []
		if origin is not None and x_axis is not None and y_axis is not None and z_axis is not None:
			axis_length = 50.0  # 軸の長さ
			# O座標系（実線）
			# X軸（赤）
			x_line = pv.Line(origin, origin + x_axis * axis_length)
			x_actor = plotter.add_mesh(x_line, color="red", line_width=3)
			axis_actors.append(x_actor)
			# Y軸（緑）
			y_line = pv.Line(origin, origin + y_axis * axis_length)
			y_actor = plotter.add_mesh(y_line, color="green", line_width=3)
			axis_actors.append(y_actor)
			# Z軸（青）
			z_line = pv.Line(origin, origin + z_axis * axis_length)
			z_actor = plotter.add_mesh(z_line, color="blue", line_width=3)
			axis_actors.append(z_actor)
			
			# O'座標系（破線風・細め）
			if origin_prime is not None:
				axis_prime_length = 40.0  # O'の軸は少し短く
				dash_length = 3.0  # 破線の長さ
				gap_length = 2.0   # 破線の間隔
				
				# X'軸（赤・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = origin_prime + x_axis * current_pos
					end = origin_prime + x_axis * min(current_pos + dash_length, axis_prime_length)
					x_prime_segment = pv.Line(start, end)
					x_prime_actor = plotter.add_mesh(x_prime_segment, color="darkred", line_width=2)
					axis_prime_actors.append(x_prime_actor)
					current_pos += dash_length + gap_length
				
				# Y'軸（緑・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = origin_prime + y_axis * current_pos
					end = origin_prime + y_axis * min(current_pos + dash_length, axis_prime_length)
					y_prime_segment = pv.Line(start, end)
					y_prime_actor = plotter.add_mesh(y_prime_segment, color="darkgreen", line_width=2)
					axis_prime_actors.append(y_prime_actor)
					current_pos += dash_length + gap_length
				
				# Z'軸（青・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = origin_prime + z_axis * current_pos
					end = origin_prime + z_axis * min(current_pos + dash_length, axis_prime_length)
					z_prime_segment = pv.Line(start, end)
					z_prime_actor = plotter.add_mesh(z_prime_segment, color="darkblue", line_width=2)
					axis_prime_actors.append(z_prime_actor)
					current_pos += dash_length + gap_length
				
				# O'の原点を球で表示
				o_prime_sphere = pv.Sphere(radius=2.0, center=origin_prime)
				o_prime_actor = plotter.add_mesh(o_prime_sphere, color="orange", opacity=0.8)
				axis_prime_actors.append(o_prime_actor)
		
		# 左上にチェックボックスUI
		def toggle_full_model(state):
			"""元のメッシュ全体の表示/非表示"""
			mesh_full_actor.SetVisibility(state)
		
		def toggle_extracted_model(state):
			"""抽出された領域の表示/非表示"""
			if mesh_extracted_actor is not None:
				mesh_extracted_actor.SetVisibility(state)
		
		def toggle_points(state):
			points_actor.SetVisibility(state)
			for label_actor in label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_axes(state):
			"""O座標系の表示/非表示を切り替え"""
			for actor in axis_actors:
				actor.SetVisibility(state)
		
		def toggle_axes_prime(state):
			"""O'座標系の表示/非表示を切り替え"""
			for actor in axis_prime_actors:
				actor.SetVisibility(state)
		
		def update_full_opacity(value):
			"""元のメッシュの透明度を変更"""
			mesh_full_actor.GetProperty().SetOpacity(value)
		
		def update_extracted_opacity(value):
			"""抽出された領域の透明度を変更"""
			if mesh_extracted_actor is not None:
				mesh_extracted_actor.GetProperty().SetOpacity(value)
		
		# チェックボックスとラベル（設定したカラーを使用）
		plotter.add_text("Full", position=(10, 10), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_full_model, value=True, position=(100, 12), size=20, border_size=1, color_on="lightgray", color_off="white")
		
		if mesh_extracted_actor is not None:
			plotter.add_text("Ext", position=(10, 40), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_extracted_model, value=True, position=(100, 42), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			checkbox_offset = 70
		else:
			checkbox_offset = 40
		
		plotter.add_text("PP", position=(10, checkbox_offset), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_points, value=True, position=(100, checkbox_offset+2), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		# 座標系の表示チェックボックス（構築できた場合のみ）
		if axis_actors:
			plotter.add_text("O", position=(10, checkbox_offset+30), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_axes, value=True, position=(100, checkbox_offset+32), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			checkbox_offset += 30
		
		# O'座標系の表示チェックボックス（O'が存在する場合のみ）
		if axis_prime_actors:
			plotter.add_text("O'", position=(10, checkbox_offset+30), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_axes_prime, value=True, position=(100, checkbox_offset+32), size=20, border_size=1, color_on="orange", color_off="white")
		
		# 透明度スライダー（元のメッシュ用）
		plotter.add_slider_widget(
			update_full_opacity,
			rng=[0.0, 1.0],
			value=0.3,
			title="Full ",
			pointa=(0.05, 0.02),
			pointb=(0.15, 0.02),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color="lightgray",
		)
		
		# 透明度スライダー（抽出領域用）
		if mesh_extracted_actor is not None:
			plotter.add_slider_widget(
				update_extracted_opacity,
				rng=[0.0, 1.0],
				value=1.0,
				title=" ",
				pointa=(0.05, 0.06),
				pointb=(0.15, 0.06),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.prox_color,
			)

		# 不要な装飾は追加しない（grid/axes/bounds未追加）
		# 自動でカメラをフィット
		plotter.show()  # ユーザーはマウスで視点操作可能

	def on_visualize_dist(self) -> None:
		# 遠位モデルと遠位PPを表示（PyVista）
		dist_model_path = self.dist_model_path.get().strip()
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		if not dist_model_path or not dist_pp_abc_path or not dist_pp_olmn_path:
			messagebox.showwarning("入力不足", "遠位モデルと遠位特徴点(ABC, OLMN)を選択してください。")
			return

		# ライブラリの確認（遅延インポート）
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return

		# PP読み込み（PickedPoints XML対応）- 2つのファイルを統合
		try:
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 座標系構築（A, B, O点から - 遠位用）
		try:
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = self._build_coordinate_system_dist(dist_points, dist_labels)
			# 自動オフセット設定（遠位O'点がある場合）
			self._try_auto_set_offset(dist_points, dist_labels, dist_origin, dist_x_axis, dist_y_axis, dist_z_axis,
			                          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z)
		except Exception as e:
			messagebox.showwarning("座標系構築", f"座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = None, None, None, None

		# メッシュ読み込み（PyVista）
		try:
			dist_mesh_full = pv.read(dist_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"遠位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# 球体で大腿骨頭領域を抽出（座標系が構築できた場合）
		dist_mesh_extracted = None
		dist_origin_prime = None
		if dist_origin is not None:
			try:
				# O'原点位置を計算（ローカル座標系に沿ったオフセット）
				# 遠位座標系は dist_origin, dist_x_axis, ... で定義されている
				dist_origin_prime = dist_origin + (
					self.dist_offset_x.get() * dist_x_axis +
					self.dist_offset_y.get() * dist_y_axis +
					self.dist_offset_z.get() * dist_z_axis
				)
				print(f"[可視化] 遠位O: {dist_origin}, 遠位O': {dist_origin_prime}, オフセット: [{self.dist_offset_x.get()}, {self.dist_offset_y.get()}, {self.dist_offset_z.get()}]")
				
				radius = self.dist_radius.get()
				dist_mesh_extracted = self._extract_region_by_sphere(dist_mesh_full, dist_origin_prime, radius)
				if dist_mesh_extracted.n_points == 0:
					messagebox.showwarning("抽出失敗", f"半径 {radius}mm の球体内にメッシュが見つかりませんでした。")
					dist_mesh_extracted = None
			except Exception as e:
				messagebox.showwarning("球体抽出", f"球体抽出に失敗しました:\n{e}\n\n元のメッシュのみ表示します。")
				dist_mesh_extracted = None

		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		# ウィンドウサイズをディスプレイの90%に設定
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)

		# プロッタ作成（格子/目盛りなし、モデル+点のみ）
		dist_plotter = pv.Plotter(title="遠位モデル/特徴点の可視化", window_size=(window_width, window_height))
		dist_plotter.set_background("white")
		
		# 元のメッシュ全体（半透明・グレー）
		dist_mesh_full_actor = dist_plotter.add_mesh(dist_mesh_full, color="lightgray", smooth_shading=True, show_edges=False, opacity=0.3)
		
		# 抽出された領域（設定したカラー・不透明）
		dist_mesh_extracted_actor = None
		if dist_mesh_extracted is not None and dist_mesh_extracted.n_points > 0:
			dist_mesh_extracted_actor = dist_plotter.add_mesh(dist_mesh_extracted, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=1.0)
		
		# 点群（設定したカラーを使用）
		dist_points_actor = dist_plotter.add_points(dist_points, color=self.dist_color, point_size=12, render_points_as_spheres=True)
		
		# 点のラベルを表示
		dist_label_actors = []
		for i, (pt, label) in enumerate(zip(dist_points, dist_labels)):
			dist_label_actor = dist_plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.dist_color, bold=True, shadow=True, show_points=False)
			dist_label_actors.append(dist_label_actor)
		
		# 座標系の表示（構築できた場合）
		dist_axis_actors = []
		dist_axis_prime_actors = []  # O'座標系用
		if dist_origin is not None and dist_x_axis is not None and dist_y_axis is not None and dist_z_axis is not None:
			axis_length = 50.0  # 軸の長さ
			# O座標系（実線）
			# X軸（赤）
			dist_x_line = pv.Line(dist_origin, dist_origin + dist_x_axis * axis_length)
			dist_x_actor = dist_plotter.add_mesh(dist_x_line, color="red", line_width=3)
			dist_axis_actors.append(dist_x_actor)
			# Y軸（緑）
			dist_y_line = pv.Line(dist_origin, dist_origin + dist_y_axis * axis_length)
			dist_y_actor = dist_plotter.add_mesh(dist_y_line, color="green", line_width=3)
			dist_axis_actors.append(dist_y_actor)
			# Z軸（青）
			dist_z_line = pv.Line(dist_origin, dist_origin + dist_z_axis * axis_length)
			dist_z_actor = dist_plotter.add_mesh(dist_z_line, color="blue", line_width=3)
			dist_axis_actors.append(dist_z_actor)

			# O'座標系（破線風・細め）
			if dist_origin_prime is not None:
				axis_prime_length = 40.0  # O'の軸は少し短く
				dash_length = 3.0  # 破線の長さ
				gap_length = 2.0   # 破線の間隔
				
				# X'軸（赤・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = dist_origin_prime + dist_x_axis * current_pos
					end = dist_origin_prime + dist_x_axis * min(current_pos + dash_length, axis_prime_length)
					x_prime_segment = pv.Line(start, end)
					x_prime_actor = dist_plotter.add_mesh(x_prime_segment, color="darkred", line_width=2)
					dist_axis_prime_actors.append(x_prime_actor)
					current_pos += dash_length + gap_length
				
				# Y'軸（緑・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = dist_origin_prime + dist_y_axis * current_pos
					end = dist_origin_prime + dist_y_axis * min(current_pos + dash_length, axis_prime_length)
					y_prime_segment = pv.Line(start, end)
					y_prime_actor = dist_plotter.add_mesh(y_prime_segment, color="darkgreen", line_width=2)
					dist_axis_prime_actors.append(y_prime_actor)
					current_pos += dash_length + gap_length
				
				# Z'軸（青・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = dist_origin_prime + dist_z_axis * current_pos
					end = dist_origin_prime + dist_z_axis * min(current_pos + dash_length, axis_prime_length)
					z_prime_segment = pv.Line(start, end)
					z_prime_actor = dist_plotter.add_mesh(z_prime_segment, color="darkblue", line_width=2)
					dist_axis_prime_actors.append(z_prime_actor)
					current_pos += dash_length + gap_length
				
				# O'の原点を球で表示
				o_prime_sphere = pv.Sphere(radius=2.0, center=dist_origin_prime)
				o_prime_actor = dist_plotter.add_mesh(o_prime_sphere, color="orange", opacity=0.8)
				dist_axis_prime_actors.append(o_prime_actor)
		
		# 左上にチェックボックスUI
		def toggle_dist_full_model(state):
			"""元のメッシュ全体の表示/非表示"""
			dist_mesh_full_actor.SetVisibility(state)
		
		def toggle_dist_extracted_model(state):
			"""抽出された領域の表示/非表示"""
			if dist_mesh_extracted_actor is not None:
				dist_mesh_extracted_actor.SetVisibility(state)
		
		def toggle_dist_points(state):
			dist_points_actor.SetVisibility(state)
			for dist_label_actor in dist_label_actors:
				dist_label_actor.SetVisibility(state)
		
		def toggle_dist_axes(state):
			"""座標系の表示/非表示を切り替え"""
			for actor in dist_axis_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_axes_prime(state):
			"""O'座標系の表示/非表示を切り替え"""
			for actor in dist_axis_prime_actors:
				actor.SetVisibility(state)
		
		def update_dist_full_opacity(value):
			"""元のメッシュの透明度を変更"""
			dist_mesh_full_actor.GetProperty().SetOpacity(value)
		
		def update_dist_extracted_opacity(value):
			"""抽出された領域の透明度を変更"""
			if dist_mesh_extracted_actor is not None:
				dist_mesh_extracted_actor.GetProperty().SetOpacity(value)
		
		# チェックボックスとラベル（設定したカラーを使用）
		dist_plotter.add_text("Full", position=(10, 10), font_size=10, color="black")
		dist_plotter.add_checkbox_button_widget(toggle_dist_full_model, value=True, position=(100, 12), size=20, border_size=1, color_on="lightgray", color_off="white")
		
		if dist_mesh_extracted_actor is not None:
			dist_plotter.add_text("Ext", position=(10, 40), font_size=10, color="black")
			dist_plotter.add_checkbox_button_widget(toggle_dist_extracted_model, value=True, position=(100, 42), size=20, border_size=1, color_on=self.dist_color, color_off="white")
			dist_checkbox_offset = 70
		else:
			dist_checkbox_offset = 40
		
		dist_plotter.add_text("PP", position=(10, dist_checkbox_offset), font_size=10, color="black")
		dist_plotter.add_checkbox_button_widget(toggle_dist_points, value=True, position=(100, dist_checkbox_offset+2), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		# 座標系の表示チェックボックス（構築できた場合のみ）
		if dist_axis_actors:
			dist_plotter.add_text("O", position=(10, dist_checkbox_offset+30), font_size=10, color="black")
			dist_plotter.add_checkbox_button_widget(toggle_dist_axes, value=True, position=(100, dist_checkbox_offset+32), size=20, border_size=1, color_on=self.dist_color, color_off="white")
			dist_checkbox_offset += 30
			
		# O'座標系の表示チェックボックス（O'が存在する場合のみ）
		if dist_axis_prime_actors:
			dist_plotter.add_text("O'", position=(10, dist_checkbox_offset+30), font_size=10, color="black")
			dist_plotter.add_checkbox_button_widget(toggle_dist_axes_prime, value=True, position=(100, dist_checkbox_offset+32), size=20, border_size=1, color_on="orange", color_off="white")
		
		# 透明度スライダー（元のメッシュ用）
		dist_plotter.add_slider_widget(
			update_dist_full_opacity,
			rng=[0.0, 1.0],
			value=0.3,
			title="Full ",
			pointa=(0.05, 0.02),
			pointb=(0.15, 0.02),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color="lightgray",
		)
		
		# 透明度スライダー（抽出領域用）
		if dist_mesh_extracted_actor is not None:
			dist_plotter.add_slider_widget(
				update_dist_extracted_opacity,
				rng=[0.0, 1.0],
				value=1.0,
				title="",
				pointa=(0.05, 0.06),
				pointb=(0.15, 0.06),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.dist_color,
			)

		# 不要な装飾は追加しない（grid/axes/bounds未追加）
		# 自動でカメラをフィット
		dist_plotter.show()  # ユーザーはマウスで視点操作可能

	def on_visualize_all(self) -> None:
		# 近位・遠位モデルと特徴点を同時表示（PyVista）
		prox_model_path = self.prox_model_path.get().strip()
		prox_pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		prox_pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		dist_model_path = self.dist_model_path.get().strip()
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		
		if not all([prox_model_path, prox_pp_abcd_path, prox_pp_olmn_path, dist_model_path, dist_pp_abc_path, dist_pp_olmn_path]):
			messagebox.showwarning("入力不足", "近位・遠位のモデルと特徴点(ABCD, OLMN, ABC, OLMN)をすべて選択してください。")
			return

		# ライブラリの確認（遅延インポート）
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return

		# 近位PP読み込み - 2つのファイルを統合
		try:
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(prox_pp_abcd_path)
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(prox_pp_olmn_path)
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
		except Exception as e:
			messagebox.showerror("近位PP読み込み失敗", f"近位PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 遠位PP読み込み - 2つのファイルを統合
		try:
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
		except Exception as e:
			messagebox.showerror("遠位PP読み込み失敗", f"遠位PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 近位座標系構築
		try:
			prox_origin, prox_x_axis, prox_y_axis, prox_z_axis = self._build_coordinate_system(prox_points, prox_labels)
			# 自動オフセット設定
			self._try_auto_set_offset(prox_points, prox_labels, prox_origin, prox_x_axis, prox_y_axis, prox_z_axis,
			                          self.prox_offset_x, self.prox_offset_y, self.prox_offset_z)
		except Exception as e:
			messagebox.showwarning("近位座標系構築", f"近位座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			prox_origin, prox_x_axis, prox_y_axis, prox_z_axis = None, None, None, None

		# 遠位座標系構築
		try:
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = self._build_coordinate_system_dist(dist_points, dist_labels)
			# 自動オフセット設定
			self._try_auto_set_offset(dist_points, dist_labels, dist_origin, dist_x_axis, dist_y_axis, dist_z_axis,
			                          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z)
		except Exception as e:
			messagebox.showwarning("遠位座標系構築", f"遠位座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = None, None, None, None

		# 近位メッシュ読み込み
		try:
			prox_mesh = pv.read(prox_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"近位モデルの読み込みに失敗しました:\n{e}")
			return

		# 遠位メッシュ読み込み
		try:
			dist_mesh = pv.read(dist_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"遠位モデルの読み込みに失敗しました:\n{e}")
			return

		# 遠位メッシュの元のコピーを保存（変換前）
		dist_mesh_original = dist_mesh.copy()
		
		# 遠位座標系の元の値を保存（変換前）
		dist_origin_original = dist_origin.copy() if dist_origin is not None else None
		dist_x_axis_original = dist_x_axis.copy() if dist_x_axis is not None else None
		dist_y_axis_original = dist_y_axis.copy() if dist_y_axis is not None else None
		dist_z_axis_original = dist_z_axis.copy() if dist_z_axis is not None else None

		# ========== ステップ1: 近位座標系をワールド座標系に揃える ==========
		prox_to_world = None
		if prox_origin is not None and prox_x_axis is not None and prox_y_axis is not None and prox_z_axis is not None:
			# 近位座標系 → ワールド座標系への変換行列を構築
			# 近位座標系の軸ベクトル [x, y, z] を列に持つ行列の逆行列
			prox_rotation_matrix = np.column_stack([prox_x_axis, prox_y_axis, prox_z_axis])
			prox_to_world_rotation = prox_rotation_matrix.T  # 逆行列（直交行列なので転置）
			
			# 4x4同次変換行列
			prox_to_world = np.eye(4)
			prox_to_world[:3, :3] = prox_to_world_rotation
			prox_to_world[:3, 3] = -prox_to_world_rotation @ prox_origin
			
			# 近位メッシュを変換
			prox_mesh_homo = np.hstack([prox_mesh.points, np.ones((prox_mesh.points.shape[0], 1))])
			prox_mesh_transformed = (prox_to_world @ prox_mesh_homo.T).T
			prox_mesh.points = prox_mesh_transformed[:, :3]
			
			# 近位特徴点を変換
			prox_points_homo = np.hstack([prox_points, np.ones((prox_points.shape[0], 1))])
			prox_points_transformed = (prox_to_world @ prox_points_homo.T).T
			prox_points = prox_points_transformed[:, :3]
			
			# 近位座標系を更新（ワールド座標系と一致）
			prox_origin = np.array([0.0, 0.0, 0.0])
			prox_x_axis = np.array([1.0, 0.0, 0.0])
			prox_y_axis = np.array([0.0, 1.0, 0.0])
			prox_z_axis = np.array([0.0, 0.0, 1.0])
		
		# ========== ステップ2: 遠位を近位と同じ変換でワールド座標系に移動 ==========
		# 遠位メッシュと遠位座標系をワールド座標系に変換（近位と同じ変換）
		# この段階では遠位メッシュと遠位座標系の関係は保たれる
		dist_origin_in_world = None
		dist_x_axis_in_world = None
		dist_y_axis_in_world = None
		dist_z_axis_in_world = None
		
		if dist_origin_original is not None and dist_x_axis_original is not None and dist_y_axis_original is not None and dist_z_axis_original is not None:
			if prox_to_world is not None:
				# 遠位メッシュをワールド座標系に変換
				dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
				dist_mesh_transformed = (prox_to_world @ dist_mesh_homo.T).T
				dist_mesh.points = dist_mesh_transformed[:, :3]
				
				# 遠位特徴点をワールド座標系に変換
				dist_points_homo = np.hstack([dist_points, np.ones((dist_points.shape[0], 1))])
				dist_points_transformed = (prox_to_world @ dist_points_homo.T).T
				dist_points = dist_points_transformed[:, :3]
				
				# 遠位座標系をワールド座標系に変換
				dist_origin_homo = np.append(dist_origin_original, 1)
				dist_origin_in_world = (prox_to_world @ dist_origin_homo)[:3]
				
				prox_to_world_rotation = prox_to_world[:3, :3]
				dist_x_axis_in_world = prox_to_world_rotation @ dist_x_axis_original
				dist_y_axis_in_world = prox_to_world_rotation @ dist_y_axis_original
				dist_z_axis_in_world = prox_to_world_rotation @ dist_z_axis_original
			else:
				# 近位変換がない場合はそのまま
				dist_origin_in_world = dist_origin_original
				dist_x_axis_in_world = dist_x_axis_original
				dist_y_axis_in_world = dist_y_axis_original
				dist_z_axis_in_world = dist_z_axis_original
			
			# この時点で、遠位メッシュと遠位座標系はワールド座標系に配置されているが
			# 遠位メッシュと遠位座標系の相対関係は元のまま保たれている
			
			# ========== ステップ3: 遠位座標系を近位座標系（=ワールド座標系）に揃える ==========
			# 遠位座標系 → ワールド座標系への変換行列を構築
			dist_rotation_matrix = np.column_stack([dist_x_axis_in_world, dist_y_axis_in_world, dist_z_axis_in_world])
			dist_to_world_rotation = dist_rotation_matrix.T
			
			dist_to_world = np.eye(4)
			dist_to_world[:3, :3] = dist_to_world_rotation
			dist_to_world[:3, 3] = -dist_to_world_rotation @ dist_origin_in_world
			
			# 遠位メッシュを遠位座標系基準に変換（原点中心、軸を単位ベクトルに）
			dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
			dist_mesh_transformed = (dist_to_world @ dist_mesh_homo.T).T
			dist_mesh.points = dist_mesh_transformed[:, :3]
			
			# 遠位特徴点を遠位座標系基準に変換
			dist_points_homo = np.hstack([dist_points, np.ones((dist_points.shape[0], 1))])
			dist_points_transformed = (dist_to_world @ dist_points_homo.T).T
			dist_points = dist_points_transformed[:, :3]
			
			# 遠位座標系を更新（近位座標系=ワールド座標系と一致）
			dist_origin = np.array([0.0, 0.0, 0.0])
			dist_x_axis = np.array([1.0, 0.0, 0.0])
			dist_y_axis = np.array([0.0, 1.0, 0.0])
			dist_z_axis = np.array([0.0, 0.0, 1.0])

		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)

		# プロッタ作成
		all_plotter = pv.Plotter(title="股関節全体モデル/特徴点の可視化", window_size=(window_width, window_height))
		all_plotter.set_background("white")
		
		# 近位モデル（設定したカラーを使用）
		prox_mesh_actor = all_plotter.add_mesh(prox_mesh, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=1.0)
		# 遠位モデル（設定したカラーを使用）
		dist_mesh_actor = all_plotter.add_mesh(dist_mesh, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=1.0)
		
		# 軟骨モデルの読み込みと表示（任意）
		prox_cart_actor = None
		dist_cart_actor = None
		prox_cart_path = self.prox_cartilage_model_path.get().strip()
		dist_cart_path = self.dist_cartilage_model_path.get().strip()
		
		if prox_cart_path:
			try:
				prox_cart_mesh = pv.read(prox_cart_path)
				# 骨と同じ変換を適用
				if prox_to_world is not None:
					prox_cart_homo = np.hstack([prox_cart_mesh.points, np.ones((prox_cart_mesh.points.shape[0], 1))])
					prox_cart_mesh.points = (prox_to_world @ prox_cart_homo.T).T[:, :3]
				prox_cart_actor = all_plotter.add_mesh(prox_cart_mesh, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='近位軟骨')
				print(f"近位軟骨モデル読み込み成功: {Path(prox_cart_path).name}")
			except Exception as e:
				print(f"近位軟骨モデル読み込み失敗: {e}")
		
		if dist_cart_path:
			try:
				dist_cart_mesh_viz = pv.read(dist_cart_path)
				# 骨と同じ変換を適用（prox_to_world → dist_to_world）
				if prox_to_world is not None:
					dist_cart_homo = np.hstack([dist_cart_mesh_viz.points, np.ones((dist_cart_mesh_viz.points.shape[0], 1))])
					dist_cart_mesh_viz.points = (prox_to_world @ dist_cart_homo.T).T[:, :3]
				if dist_origin_in_world is not None:
					dist_cart_homo2 = np.hstack([dist_cart_mesh_viz.points, np.ones((dist_cart_mesh_viz.points.shape[0], 1))])
					dist_cart_mesh_viz.points = (dist_to_world @ dist_cart_homo2.T).T[:, :3]
				dist_cart_actor = all_plotter.add_mesh(dist_cart_mesh_viz, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='遠位軟骨')
				print(f"遠位軟骨モデル読み込み成功: {Path(dist_cart_path).name}")
			except Exception as e:
				print(f"遠位軟骨モデル読み込み失敗: {e}")
		
		# 近位特徴点の表示
		prox_points_actor = all_plotter.add_points(prox_points, color=self.prox_color, point_size=12, render_points_as_spheres=True)
		
		# 近位点のラベルを表示
		prox_label_actors = []
		for i, (pt, label) in enumerate(zip(prox_points, prox_labels)):
			prox_label_actor = all_plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.prox_color, bold=True, shadow=True, show_points=False)
			prox_label_actors.append(prox_label_actor)
		
		# 遠位特徴点の表示
		dist_points_actor = all_plotter.add_points(dist_points, color=self.dist_color, point_size=12, render_points_as_spheres=True)
		
		# 遠位点のラベルを表示
		dist_label_actors = []
		for i, (pt, label) in enumerate(zip(dist_points, dist_labels)):
			dist_label_actor = all_plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.dist_color, bold=True, shadow=True, show_points=False)
			dist_label_actors.append(dist_label_actor)
		
		# 近位座標系の表示
		prox_axis_actors = []
		if prox_origin is not None and prox_x_axis is not None and prox_y_axis is not None and prox_z_axis is not None:
			axis_length = 50.0
			# X軸（赤）
			x_line = pv.Line(prox_origin, prox_origin + prox_x_axis * axis_length)
			x_actor = all_plotter.add_mesh(x_line, color="red", line_width=3)
			prox_axis_actors.append(x_actor)
			# Y軸（緑）
			y_line = pv.Line(prox_origin, prox_origin + prox_y_axis * axis_length)
			y_actor = all_plotter.add_mesh(y_line, color="green", line_width=3)
			prox_axis_actors.append(y_actor)
			# Z軸（青）
			z_line = pv.Line(prox_origin, prox_origin + prox_z_axis * axis_length)
			z_actor = all_plotter.add_mesh(z_line, color="blue", line_width=3)
			prox_axis_actors.append(z_actor)

		# 【追加】近位A-O線とO-遠位A線の角度表示
		# 動的に更新するために変数を初期化
		dist_a_initial = None
		line_dist_actor = None
		line_dist_source = None # PyVistaのラインソース
		angle_text_actor = None
		prox_a_static = None # 固定
		prox_o_static = np.array([0., 0., 0.])
		
		# オプションが有効な場合のみ実行
		if self.show_ao_angle.get():
			try:
				# 点Aの検索（大文字小文字区別なし）
				prox_a_idx = -1
				for i, lbl in enumerate(prox_labels):
					if lbl.strip().upper() == 'A':
						prox_a_idx = i
						break
				
				dist_a_idx = -1
				for i, lbl in enumerate(dist_labels):
					if lbl.strip().upper() == 'A':
						dist_a_idx = i
						break
				
				if prox_a_idx != -1 and dist_a_idx != -1:
					pt_prox_a = prox_points[prox_a_idx]
					pt_dist_a = dist_points[dist_a_idx]
					origin_pt = prox_o_static
					
					# 状態保存（スライダー更新用）
					prox_a_static = pt_prox_a
					dist_a_initial = pt_dist_a.copy() # 初期位置（ワールド変換後、変形前）
					
					# ベクトル計算: A_prox -> O
					vec_prox_ao = origin_pt - pt_prox_a
					norm_prox = np.linalg.norm(vec_prox_ao)
					
					# ベクトル計算: O -> A_dist
					vec_dist_oa = pt_dist_a - origin_pt
					norm_dist = np.linalg.norm(vec_dist_oa)
					
					if norm_prox > 1e-6 and norm_dist > 1e-6:
						# 角度計算 (cos theta)
						cos_theta = np.dot(vec_prox_ao, vec_dist_oa) / (norm_prox * norm_dist)
						cos_theta = np.clip(cos_theta, -1.0, 1.0)
						angle_deg = np.degrees(np.arccos(cos_theta))
						
						# 線分の可視化
						# Prox A -> O (マゼンタ)
						line_prox = pv.Line(pt_prox_a, origin_pt)
						all_plotter.add_mesh(line_prox, color="magenta", line_width=4, label="Prox A-O")
						
						# O -> Dist A (シアン)
						# ここでLineオブジェクトを保持しておき、後でpointsを更新する
						line_dist_source = pv.Line(origin_pt, pt_dist_a)
						line_dist_actor = all_plotter.add_mesh(line_dist_source, color="cyan", line_width=4, label="O-Dist A")
						
						# テキスト表示
						info_text = (
							f"Angle: {angle_deg:.1f} deg\n"
							f"(Prox A-O vs O-Dist A)"
						)
						
						angle_text_actor = all_plotter.add_text(
							info_text,
							position=(10, window_height - 150),
							font_size=12,
							color="black",
							shadow=True
						)
						print(f"[Angle Analysis] {info_text.replace(chr(10), ', ')}")

			except Exception as e:
				print(f"角度計算エラー: {e}")

		
		# 遠位座標系の表示

		
		# 遠位座標系の表示
		dist_axis_actors = []
		# ※ここでは追加しない（後述のスライダー用コードで追加し、重複を避ける）

		# 同次変換行列用のスライダー（遠位モデルの変換）
		# 遠位メッシュの変換後のコピーを保持（スライダーの基準）
		dist_mesh_for_slider = dist_mesh.copy()
		
		# 遠位座標系の初期値を保存（変換後の値）
		dist_origin_initial = dist_origin.copy() if dist_origin is not None else None
		dist_x_axis_initial = dist_x_axis.copy() if dist_x_axis is not None else None
		dist_y_axis_initial = dist_y_axis.copy() if dist_y_axis is not None else None
		dist_z_axis_initial = dist_z_axis.copy() if dist_z_axis is not None else None
		
		# 遠位座標系の軸ラインを保持（動的更新用）
		dist_x_line = None
		dist_y_line = None
		dist_z_line = None
		if dist_origin is not None and dist_x_axis is not None and dist_y_axis is not None and dist_z_axis is not None:
			axis_length = 50.0
			dist_x_line = pv.Line(dist_origin, dist_origin + dist_x_axis * axis_length)
			dist_y_line = pv.Line(dist_origin, dist_origin + dist_y_axis * axis_length)
			dist_z_line = pv.Line(dist_origin, dist_origin + dist_z_axis * axis_length)
			
			# 遠位座標系の軸をプロッターに追加（スライダーで動的に更新される）
			dx_actor = all_plotter.add_mesh(dist_x_line, color='red', line_width=3, label='Dist X')
			dy_actor = all_plotter.add_mesh(dist_y_line, color='green', line_width=3, label='Dist Y')
			dz_actor = all_plotter.add_mesh(dist_z_line, color='blue', line_width=3, label='Dist Z')
			
			# トグル制御用にアクターをリストに追加
			dist_axis_actors.append(dx_actor)
			dist_axis_actors.append(dy_actor)
			dist_axis_actors.append(dz_actor)
		
		# チェックボックスUI
		def toggle_prox_model(state):
			prox_mesh_actor.SetVisibility(state)
		
		def toggle_prox_points(state):
			prox_points_actor.SetVisibility(state)
			for label_actor in prox_label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_dist_model(state):
			dist_mesh_actor.SetVisibility(state)
		
		def toggle_dist_points(state):
			dist_points_actor.SetVisibility(state)
			for label_actor in dist_label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_prox_axes(state):
			for actor in prox_axis_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_axes(state):
			for actor in dist_axis_actors:
				actor.SetVisibility(state)
		
		def update_prox_opacity(value):
			prox_mesh_actor.GetProperty().SetOpacity(value)
		
		def update_dist_opacity(value):
			dist_mesh_actor.GetProperty().SetOpacity(value)
		
		# チェックボックスとラベル - 近位（設定したカラーを使用）
		all_plotter.add_text("Prox Model", position=(10, 10), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_prox_model, value=True, position=(130, 12), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		all_plotter.add_text("Prox PP", position=(10, 40), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_prox_points, value=True, position=(130, 42), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		prox_y_offset = 70
		if prox_axis_actors:
			all_plotter.add_text("Prox Axes", position=(10, prox_y_offset), font_size=10, color="black")
			all_plotter.add_checkbox_button_widget(toggle_prox_axes, value=True, position=(130, prox_y_offset+2), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			prox_y_offset += 30
		
		# チェックボックスとラベル - 遠位（設定したカラーを使用）
		all_plotter.add_text("Dist Model", position=(10, prox_y_offset), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_dist_model, value=True, position=(130, prox_y_offset+2), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		all_plotter.add_text("Dist PP", position=(10, prox_y_offset+30), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_dist_points, value=True, position=(130, prox_y_offset+32), size=20, border_size=1, color_on=self.dist_color, color_off="white")

		if dist_axis_actors:
			all_plotter.add_text("Dist Axes", position=(10, prox_y_offset+60), font_size=10, color="black")
			all_plotter.add_checkbox_button_widget(toggle_dist_axes, value=True, position=(130, prox_y_offset+62), size=20, border_size=1, color_on=self.dist_color, color_off="white")

		# 透明度スライダー - 近位（設定したカラーを使用）
		all_plotter.add_slider_widget(
			update_prox_opacity,
			rng=[0.0, 1.0],
			value=1.0,
			title=" ",
			pointa=(0.08, 0.025),
			pointb=(0.18, 0.025),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color=self.prox_color,
		)
		
		# 透明度スライダー - 遠位（設定したカラーを使用）
		all_plotter.add_slider_widget(
			update_dist_opacity,
			rng=[0.0, 1.0],
			value=1.0,
			title=" ",
			pointa=(0.08, 0.085),
			pointb=(0.18, 0.085),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color=self.dist_color,
		)
		
		
		# 変換パラメータの初期値
		transform_params = {
			'FE': 0.0,  # Flexion/Extension (Z軸回転)
			'VV': 0.0,  # Varus/Valgus (X軸回転)
			'IE': 0.0,  # Internal/External rotation (Y軸回転)
			'ML': 0.0,  # Medial/Lateral translation
			'AP': 0.0,  # Anterior/Posterior translation
			'PD': 0.0,  # Proximal/Distal translation
		}
		
		def apply_transform():
			"""現在のパラメータで遠位モデルと座標系を変換"""
			# 同次変換行列を構築（ワールド座標系基準）
			matrix = self._build_transform_matrix(
				rz=transform_params['FE'],
				rx=transform_params['VV'],
				ry=transform_params['IE'],
				ml=transform_params['ML'],
				ap=transform_params['AP'],
				pd=transform_params['PD']
			)
			
			# 遠位メッシュを変換
			# dist_mesh_for_sliderは遠位座標系の初期位置からの相対位置を保持
			# 1. 遠位座標系の原点を中心に配置
			if dist_origin_initial is not None:
				centered_points = dist_mesh_for_slider.points.copy() - dist_origin_initial
			else:
				centered_points = dist_mesh_for_slider.points.copy()
			
			# 2. 同次座標に変換
			ones = np.ones((centered_points.shape[0], 1))
			points_homogeneous = np.hstack([centered_points, ones])
			
			# 3. 同次変換行列を適用
			transformed_homogeneous = (matrix @ points_homogeneous.T).T
			transformed_points = transformed_homogeneous[:, :3]
			
			# 4. 原点に戻す
			if dist_origin_initial is not None:
				transformed_points = transformed_points + dist_origin_initial
			
			# 5. メッシュの頂点を更新
			dist_mesh.points = transformed_points
			
			# 【追加】角度線の動的更新
			if (dist_a_initial is not None and line_dist_source is not None and 
				angle_text_actor is not None and prox_a_static is not None):
				
				# Dist A点にも同じ変換を適用
				# dist_a_initialはワールド同次変換適用後の初期位置。
				# ここではapply_transformと同じロジックで変換する
				
				# 1. 中心化（dist_origin_initialは[0,0,0]なのでそのまま）
				pt_centered = dist_a_initial
				
				# 2. 同次座標 & 変換
				pt_homo = np.append(pt_centered, 1.0)
				pt_transformed = (matrix @ pt_homo)[:3]
				
				# 3. 原点戻し（+0 なのでそのまま）
				new_dist_a = pt_transformed
				
				# ライン更新: O(0,0,0) -> New Dist A
				line_dist_source.points = np.array([prox_o_static, new_dist_a])
				
				# 角度再計算
				vec_prox_ao = prox_o_static - prox_a_static
				vec_dist_oa = new_dist_a - prox_o_static
				
				n_p = np.linalg.norm(vec_prox_ao)
				n_d = np.linalg.norm(vec_dist_oa)
				
				if n_p > 1e-6 and n_d > 1e-6:
					cos_t = np.dot(vec_prox_ao, vec_dist_oa) / (n_p * n_d)
					cos_t = np.clip(cos_t, -1.0, 1.0)
					deg = np.degrees(np.arccos(cos_t))
					
					# テキスト更新
					new_text = (
						f"Angle: {deg:.1f} deg\n"
						f"(Prox A-O vs O-Dist A)"
					)
					angle_text_actor.SetInput(new_text)

			# 遠位座標系の軸を更新
			if (dist_origin_initial is not None and dist_x_axis_initial is not None and 
				dist_y_axis_initial is not None and dist_z_axis_initial is not None and
				dist_x_line is not None and dist_y_line is not None and dist_z_line is not None):
				
				# 遠位座標系の原点を変換
				origin_centered = np.array([0, 0, 0, 1])
				transformed_origin = (matrix @ origin_centered)[:3]
				transformed_origin = transformed_origin + dist_origin_initial
				
				# 遠位座標系の軸を変換
				rotation_matrix = matrix[:3, :3]
				transformed_x_axis = rotation_matrix @ dist_x_axis_initial
				transformed_y_axis = rotation_matrix @ dist_y_axis_initial
				transformed_z_axis = rotation_matrix @ dist_z_axis_initial
				
				axis_length = 50.0
				# X軸の更新
				dist_x_line.points = np.array([transformed_origin, transformed_origin + transformed_x_axis * axis_length])
				# Y軸の更新
				dist_y_line.points = np.array([transformed_origin, transformed_origin + transformed_y_axis * axis_length])
				# Z軸の更新
				dist_z_line.points = np.array([transformed_origin, transformed_origin + transformed_z_axis * axis_length])
		
		# スライダー関数（角度用）
		def update_FE(value):
			transform_params['FE'] = value
			apply_transform()
		
		def update_VV(value):
			transform_params['VV'] = value
			apply_transform()
		
		def update_IE(value):
			transform_params['IE'] = value
			apply_transform()
		
		# スライダー関数（変位用）
		def update_ML(value):
			transform_params['ML'] = value
			apply_transform()
		
		def update_AP(value):
			transform_params['AP'] = value
			apply_transform()
		
		def update_PD(value):
			transform_params['PD'] = value
			apply_transform()
		
		# スライダーを追加（右側に配置）
		slider_x_start = 0.75
		slider_x_end = 0.95
		
		# スライダーウィジェットの参照を保持
		slider_widgets = {}
		
		# 角度スライダー（度）
		slider_widgets['FE'] = all_plotter.add_slider_widget(
			update_FE,
			rng=[-180.0, 180.0],
			value=0.0,
			title="FE",
			pointa=(slider_x_start, 0.85),
			pointb=(slider_x_end, 0.85),
			style='modern',
			tube_width=0.005,
			slider_width=0.015,
			color="blue",
		)
		
		slider_widgets['VV'] = all_plotter.add_slider_widget(
			update_VV,
			rng=[-180.0, 180.0],
			value=0.0,
			title="VV",
			pointa=(slider_x_start, 0.75),
			pointb=(slider_x_end, 0.75),
			style='modern',
			tube_width=0.005,
			slider_width=0.015,
			color="red",
		)
		
		slider_widgets['IE'] = all_plotter.add_slider_widget(
			update_IE,
			rng=[-180.0, 180.0],
			value=0.0,
			title="IE",
			pointa=(slider_x_start, 0.65),
			pointb=(slider_x_end, 0.65),
			style='modern',
			tube_width=0.005,
			slider_width=0.015,
			color="green",
		)
		
		# 変位スライダー（mm）
		slider_widgets['ML'] = all_plotter.add_slider_widget(
			update_ML,
			rng=[-50.0, 50.0],
			value=0.0,
			title="ML",
			pointa=(slider_x_start, 0.50),
			pointb=(slider_x_end, 0.50),
			style='modern',
			tube_width=0.005,
			slider_width=0.015,
			color="blue",  # Z軸と同じ色
		)
		
		slider_widgets['AP'] = all_plotter.add_slider_widget(
			update_AP,
			rng=[-50.0, 50.0],
			value=0.0,
			title="AP",
			pointa=(slider_x_start, 0.40),
			pointb=(slider_x_end, 0.40),
			style='modern',
			tube_width=0.005,
			slider_width=0.015,
			color="red",  # X軸と同じ色
		)
		
		slider_widgets['PD'] = all_plotter.add_slider_widget(
			update_PD,
			rng=[-50.0, 50.0],
			value=0.0,
			title="PD",
			pointa=(slider_x_start, 0.30),
			pointb=(slider_x_end, 0.30),
			style='modern',
			tube_width=0.005,
			slider_width=0.015,
			color="green",  # Y軸と同じ色
		)
		
		# Resetボタンの追加
		def reset_all():
			"""全てのパラメータを0にリセット"""
			# パラメータを0に設定
			transform_params['FE'] = 0.0
			transform_params['VV'] = 0.0
			transform_params['IE'] = 0.0
			transform_params['ML'] = 0.0
			transform_params['AP'] = 0.0
			transform_params['PD'] = 0.0
			
			# スライダーの値を0に更新
			for key, widget in slider_widgets.items():
				# PyVistaのスライダーウィジェットの値を更新
				widget.GetRepresentation().SetValue(0.0)
			
			# 変換を適用
			apply_transform()
		
		# Resetボタンを追加
		all_plotter.add_text(
			"Reset",
			position=(slider_x_start * all_plotter.window_size[0], 0.15 * all_plotter.window_size[1]),
			font_size=12,
			color="white",
			name="reset_button"
		)
		
		# ボタンクリックのためのキーボードショートカット（'r'キー）
		all_plotter.add_key_event('r', reset_all)
		all_plotter.add_key_event('R', reset_all)

		all_plotter.show()

	def on_visualize_options(self) -> None:
		"""可視化オプションウィンドウを開く"""
		# 既に開いている場合はフォーカス
		if hasattr(self, '_options_window') and self._options_window is not None and self._options_window.winfo_exists():
			self._options_window.lift()
			self._options_window.focus_force()
			return

		from tkinter import colorchooser

		# 新しいウィンドウを作成
		self._options_window = tk.Toplevel(self)
		options_window = self._options_window
		options_window.title("可視化オプション")
		options_window.geometry("650x750")
		options_window.resizable(True, True)
		options_window.transient(self)  # メインウィンドウの前面に固定

		# タイトル
		title_label = ttk.Label(options_window, text="可視化オプション", font=(self.ui_font_family, 12, "bold"))
		title_label.pack(pady=(10, 20))
		
		# ========== 色設定セクション ==========
		color_section_label = ttk.Label(options_window, text="● 色設定", font=(self.ui_font_family, 10, "bold"))
		color_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# 近位カラー設定
		prox_frame = ttk.Frame(options_window)
		prox_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(prox_frame, text="近位カラー:").pack(side="left")
		prox_color_label = tk.Label(prox_frame, text="　　　", bg=self.prox_color, relief="solid", borderwidth=1, width=10)
		prox_color_label.pack(side="left", padx=10)
		
		def choose_prox_color():
			color = colorchooser.askcolor(initialcolor=self.prox_color, title="近位カラーを選択", parent=options_window)
			options_window.lift()
			options_window.focus_force()
			if color[1]:  # color[1]は16進数カラーコード
				self.prox_color = color[1]
				prox_color_label.config(bg=self.prox_color)
				self._save_state()
		
		ttk.Button(prox_frame, text="変更...", command=choose_prox_color).pack(side="left")
		
		# 遠位カラー設定
		dist_frame = ttk.Frame(options_window)
		dist_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(dist_frame, text="遠位カラー:").pack(side="left")
		dist_color_label = tk.Label(dist_frame, text="　　　", bg=self.dist_color, relief="solid", borderwidth=1, width=10)
		dist_color_label.pack(side="left", padx=10)
		
		def choose_dist_color():
			color = colorchooser.askcolor(initialcolor=self.dist_color, title="遠位カラーを選択", parent=options_window)
			options_window.lift()
			options_window.focus_force()
			if color[1]:  # color[1]は16進数カラーコード
				self.dist_color = color[1]
				dist_color_label.config(bg=self.dist_color)
				self._save_state()
		
		ttk.Button(dist_frame, text="変更...", command=choose_dist_color).pack(side="left")

		
		# オーバーラップ体積カラー設定
		overlap_frame = ttk.Frame(options_window)
		overlap_frame.pack(fill="x", padx=40, pady=5)

		ttk.Label(overlap_frame, text="オーバーラップ体積カラー:").pack(side="left")
		overlap_color_label = tk.Label(overlap_frame, text="　　　", bg=self.overlap_color, relief="solid", borderwidth=1, width=10)
		overlap_color_label.pack(side="left", padx=10)

		def choose_overlap_color():
			color = colorchooser.askcolor(initialcolor=self.overlap_color, title="オーバーラップ体積カラーを選択", parent=options_window)
			options_window.lift()
			options_window.focus_force()
			if color[1]:
				self.overlap_color = color[1]
				overlap_color_label.config(bg=self.overlap_color)
				self._save_state()

		ttk.Button(overlap_frame, text="変更...", command=choose_overlap_color).pack(side="left")
		
		# ========== 関節範囲の半径設定セクション ==========
		ttk.Separator(options_window, orient="horizontal").pack(fill="x", padx=20, pady=15)
		
		radius_section_label = ttk.Label(options_window, text="● 関節範囲の半径（球体抽出用）", font=(self.ui_font_family, 10, "bold"))
		radius_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# 近位側半径設定
		prox_radius_frame = ttk.Frame(options_window)
		prox_radius_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(prox_radius_frame, text="近位側半径 (mm):").pack(side="left")
		prox_radius_entry = ttk.Entry(prox_radius_frame, textvariable=self.prox_radius, width=10)
		prox_radius_entry.pack(side="left", padx=10)
		ttk.Label(prox_radius_frame, text="※寛骨臼領域を抽出", foreground="gray").pack(side="left")
		
		# 遠位側半径設定
		dist_radius_frame = ttk.Frame(options_window)
		dist_radius_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(dist_radius_frame, text="遠位側半径 (mm):").pack(side="left")
		dist_radius_entry = ttk.Entry(dist_radius_frame, textvariable=self.dist_radius, width=10)
		dist_radius_entry.pack(side="left", padx=10)
		ttk.Label(dist_radius_frame, text="※大腿骨頭領域を抽出", foreground="gray").pack(side="left")
		
		# セパレータ
		ttk.Separator(options_window, orient="horizontal").pack(fill="x", padx=20, pady=15)
		
		# O'原点位置オフセット設定
		offset_section_label = ttk.Label(options_window, text="● O'原点位置オフセット（ローカル座標系に沿った移動）", font=(self.ui_font_family, 10, "bold"))
		offset_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# 近位O'オフセット設定
		prox_offset_label = ttk.Label(options_window, text="近位O'オフセット (mm):", font=(self.ui_font_family, 9))
		prox_offset_label.pack(anchor="w", padx=40, pady=(5, 2))
		
		prox_offset_frame = ttk.Frame(options_window)
		prox_offset_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(prox_offset_frame, text="X:").pack(side="left")
		prox_offset_x_entry = ttk.Entry(prox_offset_frame, textvariable=self.prox_offset_x, width=8)
		prox_offset_x_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(prox_offset_frame, text="Y:").pack(side="left")
		prox_offset_y_entry = ttk.Entry(prox_offset_frame, textvariable=self.prox_offset_y, width=8)
		prox_offset_y_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(prox_offset_frame, text="Z:").pack(side="left")
		prox_offset_z_entry = ttk.Entry(prox_offset_frame, textvariable=self.prox_offset_z, width=8)
		prox_offset_z_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(prox_offset_frame, text="※抽出中心 = O + X*x軸 + Y*y軸 + Z*z軸", foreground="gray").pack(side="left")
		
		# 遠位O'オフセット設定
		dist_offset_label = ttk.Label(options_window, text="遠位O'オフセット (mm):", font=(self.ui_font_family, 9))
		dist_offset_label.pack(anchor="w", padx=40, pady=(10, 2))
		
		dist_offset_frame = ttk.Frame(options_window)
		dist_offset_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(dist_offset_frame, text="X:").pack(side="left")
		dist_offset_x_entry = ttk.Entry(dist_offset_frame, textvariable=self.dist_offset_x, width=8)
		dist_offset_x_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(dist_offset_frame, text="Y:").pack(side="left")
		dist_offset_y_entry = ttk.Entry(dist_offset_frame, textvariable=self.dist_offset_y, width=8)
		dist_offset_y_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(dist_offset_frame, text="Z:").pack(side="left")
		dist_offset_z_entry = ttk.Entry(dist_offset_frame, textvariable=self.dist_offset_z, width=8)
		dist_offset_z_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(dist_offset_frame, text="※抽出中心 = O + X*x軸 + Y*y軸 + Z*z軸", foreground="gray").pack(side="left")

		# セパレータ
		ttk.Separator(options_window, orient="horizontal").pack(fill="x", padx=20, pady=15)
		
		# その他設定
		misc_section_label = ttk.Label(options_window, text="● その他", font=(self.ui_font_family, 10, "bold"))
		misc_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# A-O線表示
		ao_frame = ttk.Frame(options_window)
		ao_frame.pack(fill="x", padx=40, pady=5)
		ttk.Checkbutton(ao_frame, text="骨盤A-O線と大腿骨A-Oを表示する", variable=self.show_ao_angle).pack(side="left")
		
		# 保存と閉じるボタン
		button_frame = ttk.Frame(options_window)
		button_frame.pack(pady=20)
		
		def save_and_close():
			self._save_state()
			options_window.destroy()
		
		ttk.Button(button_frame, text="保存して閉じる", command=save_and_close).pack(side="left", padx=5)
		ttk.Button(button_frame, text="閉じる", command=options_window.destroy).pack(side="left", padx=5)

	def on_check_transform(self) -> None:
		"""同次変換行列群の確認: xlsxファイルから変換行列を読み込んで表示"""
		transform_path = self.transform_group_path.get()
		if not transform_path:
			messagebox.showwarning("同次変換行列群の確認", "同次変換行列群ファイルを選択してください。")
			return
		
		try:
			# xlsxファイルを読み込み
			data = self._load_transform_matrices(transform_path)
			if not data:
				messagebox.showwarning("同次変換行列群の確認", "変換行列データを読み込めませんでした。")
				return
			
			# 確認ウィンドウを表示
			self._show_transform_matrices(data)
			
		except Exception as e:
			messagebox.showerror("エラー", f"同次変換行列の読み込みに失敗しました:\n{str(e)}")
	
	def _load_transform_matrices(self, filepath: str):
		"""xlsxまたはKKRファイルから同次変換行列データを読み込む
		
		Args:
			filepath: xlsxまたはKKRファイルのパス
			
		Returns:
			list of dict: [{"time": float, "matrix": np.ndarray(4,4)}, ...]
		"""
		# ファイル拡張子でフォーマットを判定
		file_ext = Path(filepath).suffix.lower()
		if file_ext in ['.kkr']:
			return self._load_transform_matrices_from_kkr(filepath)
		else:
			return self._load_transform_matrices_from_xlsx(filepath)
	
	def _load_transform_matrices_from_xlsx(self, filepath: str):
		"""xlsxファイルから同次変換行列データを読み込む
		
		Args:
			filepath: xlsxファイルのパス
			
		Returns:
			list of dict: [{"time": float, "matrix": np.ndarray(4,4)}, ...]
		"""
		try:
			# data_only=Falseで数式も読み取れるようにする
			wb = openpyxl.load_workbook(filepath, data_only=False)
			sheet = wb.active
			
			# 列のマッピング (AB=FE, AC=ML, AD=VV, AE=AP, AF=IE, AG=PD)
			# AB=28, AC=29, AD=30, AE=31, AF=32, AG=33
			col_indices = {
				"FE": 28,  # Z軸回転
				"ML": 29,  # Z軸変位
				"VV": 30,  # X軸回転
				"AP": 31,  # X軸変位
				"IE": 32,  # Y軸回転
				"PD": 33,  # Y軸変位
			}
			
			data = []
			# 16行目以降を読み取り
			for row_idx in range(16, sheet.max_row + 1):
				# A列からTime値を取得（列番号1）
				time_cell = sheet.cell(row=row_idx, column=1)
				time_val = time_cell.value
				
				if time_val is None:
					continue
				
				# Time値を数値に変換
				try:
					time_val = float(time_val)
				except (TypeError, ValueError):
					continue
				
				# 各列の値を取得
				try:
					fe = float(sheet.cell(row=row_idx, column=col_indices["FE"]).value or 0)  # Z軸回転
					ml = float(sheet.cell(row=row_idx, column=col_indices["ML"]).value or 0)  # Z軸変位
					vv = float(sheet.cell(row=row_idx, column=col_indices["VV"]).value or 0)  # X軸回転
					ap = float(sheet.cell(row=row_idx, column=col_indices["AP"]).value or 0)  # X軸変位
					ie = float(sheet.cell(row=row_idx, column=col_indices["IE"]).value or 0)  # Y軸回転
					pd = float(sheet.cell(row=row_idx, column=col_indices["PD"]).value or 0)  # Y軸変位
				except (TypeError, ValueError):
					continue
				
				# 同次変換行列を構築 (Z-X-Y オイラー角 + 変位)
				matrix = self._build_transform_matrix(fe, vv, ie, ml, ap, pd)
				
				data.append({
					"time": time_val,
					"matrix": matrix,
					"angles": (fe, vv, ie),
					"translations": (ml, ap, pd)
				})
			
			wb.close()
			
			# データを時間順にソート（昇順）
			data.sort(key=lambda x: x["time"])
			
			print(f"[データ読み込み] {len(data)}フレーム読み込み完了")
			if data:
				print(f"[データ読み込み] 時間範囲: {data[0]['time']:.3f}s ~ {data[-1]['time']:.3f}s")
			
			return data
			
		except Exception as e:
			raise Exception(f"xlsx読み込みエラー: {str(e)}")
	
	def _load_transform_matrices_from_kkr(self, filepath: str):
		"""KKRファイルから同次変換行列データを読み込む
		
		Args:
			filepath: KKRファイルのパス
			
		Returns:
			list of dict: [{"time": float, "matrix": np.ndarray(4,4)}, ...]
		"""
		try:
			import csv
			
			if not Path(filepath).exists():
				raise FileNotFoundError(f"ファイルが見つかりません: {filepath}")
				
			data = []
			
			with open(filepath, 'r', encoding='utf-8') as f:
				# メタデータ行をスキップして、実際のデータヘッダーを見つける
				lines = f.readlines()
				header_line_idx = None
				
				for i, line in enumerate(lines):
					if 'Time' in line and 'FE' in line and 'ML' in line:
						header_line_idx = i
						break
				
				if header_line_idx is None:
					raise Exception("KKRファイルにデータヘッダーが見つかりません")
				
				# ヘッダー行からデータ行までを再構築
				data_lines = lines[header_line_idx:]
				reader = csv.DictReader(data_lines)
				
				for row in reader:
					try:
						# Time列を取得
						time_val = float(row['Time'])
						
						# FE, ML, VV, AP, IE, PD列を取得
						fe = float(row['FE'])  # Z軸回転
						ml = float(row['ML'])  # Z軸変位（Medial/Lateral）
						vv = float(row['VV'])  # X軸回転
						ap = float(row['AP'])  # X軸変位（Anterior/Posterior）
						ie = float(row['IE'])  # Y軸回転
						pd = float(row['PD'])  # Y軸変位（Proximal/Distal）
					except (KeyError, TypeError, ValueError) as e:
						# 列が存在しない、または変換できない場合はスキップ
						continue
					
					# 同次変換行列を構築 (Z-X-Y オイラー角 + 変位)
					matrix = self._build_transform_matrix(fe, vv, ie, ml, ap, pd)
					
					data.append({
						"time": time_val,
						"matrix": matrix,
						"angles": (fe, vv, ie),
						"translations": (ml, ap, pd)
					})
			
			# データを時間順にソート（昇順）
			data.sort(key=lambda x: x["time"])
			
			print(f"[データ読み込み] {len(data)}フレーム読み込み完了")
			if data:
				print(f"[データ読み込み] 時間範囲: {data[0]['time']:.3f}s ~ {data[-1]['time']:.3f}s")
			
			return data
			
		except Exception as e:
			raise Exception(f"KKR読み込みエラー: {str(e)}")
	
	def _build_transform_matrix(self, rz: float, rx: float, ry: float, 
							  ml: float, ap: float, pd: float) -> np.ndarray:
		"""Z-X-Y オイラー角と変位から4x4同次変換行列を構築
		
		Args:
			rz, rx, ry: Z, X, Y軸周りの回転角度 (度) [FE, VV, IE]
			ml: Medial/Lateral変位 (mm)
			ap: Anterior/Posterior変位 (mm)
			pd: Proximal/Distal変位 (mm)
			
		Returns:
			4x4 同次変換行列
		"""
		# 度をラジアンに変換
		rz_rad = np.deg2rad(rz)
		rx_rad = np.deg2rad(rx)
		ry_rad = np.deg2rad(ry)
		
		# Z軸回転行列
		Rz = np.array([
			[np.cos(rz_rad), -np.sin(rz_rad), 0],
			[np.sin(rz_rad),  np.cos(rz_rad), 0],
			[0,               0,              1]
		])
		
		# X軸回転行列
		Rx = np.array([
			[1, 0,               0],
			[0, np.cos(rx_rad), -np.sin(rx_rad)],
			[0, np.sin(rx_rad),  np.cos(rx_rad)]
		])
		
		# Y軸回転行列
		Ry = np.array([
			[ np.cos(ry_rad), 0, np.sin(ry_rad)],
			[ 0,              1, 0],
			[-np.sin(ry_rad), 0, np.cos(ry_rad)]
		])
		
		# 合成回転行列 (Intrinsic: Y -> X -> Z)
		# R = Rz @ Rx @ Ry
		R = Rz @ Rx @ Ry

		# --- 指定数式による並進ベクトルの計算 ---
		# x = AP * cos(FE) - PD * sin(FE) * cos(VV)
		# y = AP * sin(FE) + PD * cos(FE) * cos(VV)
		# z = ML + PD * sin(VV)
		cf = np.cos(rz_rad)
		sf = np.sin(rz_rad)
		cv = np.cos(rx_rad)
		sv = np.sin(rx_rad)
		x = (ap * cf) - (pd * sf * cv)
		y = (ap * sf) + (pd * cf * cv)
		z = ml + (pd * sv)
		t_vec = np.array([x, y, z], dtype=float)

		# 4x4同次変換行列へセット
		T = np.eye(4, dtype=float)
		T[:3, :3] = R
		T[:3, 3] = t_vec

		return T
	
	def _show_transform_matrices(self, data):
		"""変位・姿勢変化データを表示するウィンドウ"""
		window = tk.Toplevel(self)
		window.title("変位・姿勢変化データの確認")
		window.geometry("600x500")
		
		# スクロール可能なテキストウィジェット
		frame = ttk.Frame(window)
		frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
		
		scrollbar = ttk.Scrollbar(frame)
		scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
		
		text_widget = tk.Text(frame, wrap=tk.NONE, yscrollcommand=scrollbar.set, font=(self.mono_font_family, 10))
		text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		scrollbar.config(command=text_widget.yview)
		
		# ヘッダー行を表示（各時刻に対して 4x4 行列を出力）
		text_widget.insert(tk.END, f"{'Time':>8}    {'Transform Matrix (4x4)'}\n")
		text_widget.insert(tk.END, "" + "=" * 70 + "\n")
		
		# データを整形して表示（各時刻ごとに行列を見やすく表示）
		for item in data:
			time = item["time"]
			matrix = item.get("matrix")
			if matrix is None:
				# 予備: 行列が未計算なら angles/translations から構築
				angles = item.get("angles", (0.0, 0.0, 0.0))
				trans = item.get("translations", (0.0, 0.0, 0.0))
				matrix = self._build_transform_matrix(angles[0], angles[1], angles[2], trans[0], trans[1], trans[2])
			
			# 時刻行
			text_widget.insert(tk.END, f"{time:8.3f}\n")
			# 行列を 3 桁整形で 4 行に分けて表示
			for r in range(4):
				row_vals = [matrix[r, c] for c in range(4)]
				row_str = '  '.join([f"{v:12.6f}" for v in row_vals])
				text_widget.insert(tk.END, f"    {row_str}\n")
			text_widget.insert(tk.END, "\n")
		
		text_widget.config(state=tk.DISABLED)
		
		# 閉じるボタン
		close_btn = ttk.Button(window, text="閉じる", command=window.destroy)
		close_btn.pack(pady=10)

	def _get_cache_hash(self, prox_mesh, dist_mesh, transform_data, simplify_mesh=False):
		"""キャッシュファイル用のハッシュを生成
		
		Args:
			prox_mesh: 近位メッシュ（整合性のため引数としては残すが、実際はパスを使用）
			dist_mesh: 遠位メッシュ（同上）
			transform_data: 変換データ
			simplify_mesh: メッシュ簡略化フラグ
			
		Returns:
			str: ハッシュ文字列
		"""
		# メッシュと変換データから一意のハッシュを生成
		hash_input = ["v5_robust_params"]  # バージョンを変更
		
		# ファイルパスを使用（ポイント配列より安定かつ確実）
		hash_input.append(f"pm:{self.prox_model_path.get()}")
		hash_input.append(f"dm:{self.dist_model_path.get()}")
		hash_input.append(f"tg:{self.transform_group_path.get()}")
		
		# PPファイル（座標系・変換に影響）
		hash_input.append(f"pp1:{self.prox_pp_abcd_path.get()}")
		hash_input.append(f"pp2:{self.prox_pp_olmn_path.get()}")
		hash_input.append(f"dp1:{self.dist_pp_abc_path.get()}")
		hash_input.append(f"dp2:{self.dist_pp_olmn_path.get()}")
		
		# パラメータ（抽出領域・ヒートマップに影響）
		hash_input.append(f"pr:{self.prox_radius.get()}")
		hash_input.append(f"dr:{self.dist_radius.get()}")
		hash_input.append(f"po:{self.prox_offset_x.get()}/{self.prox_offset_y.get()}/{self.prox_offset_z.get()}")
		hash_input.append(f"do:{self.dist_offset_x.get()}/{self.dist_offset_y.get()}/{self.dist_offset_z.get()}")
		
		# 変換データのフレーム数
		hash_input.append(str(len(transform_data)))
		
		# ※ transform_dataの中身（numpy配列）を文字列化してハッシュに含めると、
		#   浮動小数点の表示形式の違いなどでハッシュが不安定になる場合があるため、
		#   ファイルパスとフレーム数で識別する形に変更。
		
		# メッシュ簡略化フラグを追加
		hash_input.append(f"simplify:{simplify_mesh}")
		
		# ハッシュ生成
		hash_str = '|'.join(hash_input)
		return hashlib.md5(hash_str.encode()).hexdigest()

	def _get_cache_filepath(self, cache_hash):
		"""キャッシュファイルのパスを取得
		
		Args:
			cache_hash: ハッシュ文字列
			
		Returns:
			Path: キャッシュファイルのパス
		"""
		cache_dir = Path(__file__).parent / ".overlap_cache"
		cache_dir.mkdir(exist_ok=True)
		return cache_dir / f"overlap_{cache_hash}.pkl"

	def _save_overlap_cache(self, cache_filepath, overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes=None):
		"""オーバーラップデータとヒートマップをキャッシュファイルに保存
		
		Args:
			cache_filepath: キャッシュファイルのパス
			overlap_meshes: オーバーラップメッシュのリスト
			overlap_areas: オーバーラップ面積のリスト
			overlap_depths: 接触深度のリスト
			heatmap_meshes: ヒートマップメッシュのリスト（オプション）
		"""
		print(f"[CACHE DEBUG] Saving to: {cache_filepath}")
		print(f"[CACHE DEBUG] Overlap entries: {len(overlap_meshes)}")
		if heatmap_meshes:
			print(f"[CACHE DEBUG] Heatmap entries: {len(heatmap_meshes)}")

		try:
			# メッシュデータをシリアライズ可能な形式に変換
			cache_data = {
				'overlap': [],
				'heatmap': [],
				'heatmap_base': None # 最適化用共通ジオメトリ
			}
			
			# オーバーラップデータ
			for mesh, area, depth in zip(overlap_meshes, overlap_areas, overlap_depths):
				if mesh is not None and mesh.n_points > 0:
					cache_data['overlap'].append({
						'points': mesh.points,
						'faces': mesh.faces,
						'area': area,
						'depth': depth
					})
				else:
					cache_data['overlap'].append(None)
			
			# ヒートマップデータ（最適化：共通ジオメトリを抽出）
			if heatmap_meshes is not None:
				# ベースメッシュを探す（最初の非空メッシュ）
				base_mesh_info = None
				for mesh in heatmap_meshes:
					if mesh is not None and mesh.n_points > 0:
						base_mesh_info = {
							'n_points': mesh.n_points,
							'n_faces': mesh.n_cells,
							'points': mesh.points,
							'faces': mesh.faces
						}
						cache_data['heatmap_base'] = base_mesh_info
						print(f"[キャッシュ] ヒートマップ共通ジオメトリを抽出: {mesh.n_points}点, {mesh.n_cells}面")
						break
				
				for mesh in heatmap_meshes:
					if mesh is not None and mesh.n_points > 0:
						# ベースメッシュと一致するか確認
						is_same_topology = (base_mesh_info is not None and 
										  mesh.n_points == base_mesh_info['n_points'] and 
										  mesh.n_cells == base_mesh_info['n_faces'])
						
						heatmap_dict = {}
						if is_same_topology:
							# ジオメトリが同じならスカラーのみ保存（タイプ: ref）
							heatmap_dict['type'] = 'ref'
						else:
							# 異なるなら全保存（タイプ: full）
							heatmap_dict['type'] = 'full'
							heatmap_dict['points'] = mesh.points
							heatmap_dict['faces'] = mesh.faces if mesh.faces is not None and len(mesh.faces) > 0 else None
						
						# スカラーデータ（distance）を保存
						if 'distance' in mesh.array_names:
							heatmap_dict['distance'] = mesh['distance']
							
						cache_data['heatmap'].append(heatmap_dict)
					else:
						cache_data['heatmap'].append(None)
			
			# ファイルに保存
			with open(cache_filepath, 'wb') as f:
				pickle.dump(cache_data, f)
			
			print(f"[キャッシュ] 保存完了: {cache_filepath}")
			if heatmap_meshes is not None:
				print(f"[キャッシュ] ヒートマップデータも保存: {len(heatmap_meshes)}フレーム")
		except Exception as e:
			print(f"[キャッシュ] 保存失敗 (詳細): {e}")
			import traceback
			traceback.print_exc()
			messagebox.showerror("キャッシュ保存エラー", f"キャッシュの保存に失敗しました。\n\n詳細: {e}")

	def _load_overlap_cache(self, cache_filepath):
		"""キャッシュファイルからオーバーラップデータとヒートマップを読み込み
		
		Args:
			cache_filepath: キャッシュファイルのパス
			
		Returns:
			tuple: (オーバーラップメッシュのリスト, オーバーラップ面積のリスト, 接触深度のリスト, ヒートマップメッシュのリスト)、またはNone
		"""
		try:
			if not cache_filepath.exists():
				return None
			
			# 進捗ダイアログを作成
			progress_window = tk.Toplevel(self)
			progress_window.title("キャッシュ読み込み中")
			progress_window.geometry("450x150")
			progress_window.resizable(False, False)
			progress_window.attributes('-topmost', True)
			
			ttk.Label(progress_window, text="キャッシュファイルを読み込んでいます...", font=(self.ui_font_family, 11)).pack(pady=10)
			
			progress_label = ttk.Label(progress_window, text="0 / ? フレーム", font=(self.ui_font_family, 10))
			progress_label.pack(pady=5)
			
			progress_bar = ttk.Progressbar(progress_window, mode='determinate', length=400)
			progress_bar.pack(pady=10)
			
			cancel_var = tk.BooleanVar(value=False)
			ttk.Button(progress_window, text="キャンセル", command=lambda: cancel_var.set(True)).pack(pady=5)
			
			progress_window.update()
			
			# ファイルから読み込み
			print(f"[キャッシュ] pickle読み込み開始: {cache_filepath}")
			with open(cache_filepath, 'rb') as f:
				cache_data = pickle.load(f)
			print(f"[キャッシュ] pickle読み込み完了")
			
			# 新形式（辞書型）と旧形式（リスト型）の両方に対応
			if isinstance(cache_data, dict):
				# 新形式: オーバーラップとヒートマップを含む
				overlap_meshes = []
				overlap_areas = []
				overlap_depths = []
				heatmap_meshes = []
				
				# オーバーラップデータを再構築（高速化：最小限の処理）
				overlap_data = cache_data.get('overlap', [])
				for data in overlap_data:
					if data is None:
						overlap_meshes.append(pv.PolyData())
						overlap_areas.append(0.0)
						overlap_depths.append(0.0)
					else:
						# メッシュ作成をスキップ、データだけ保持（必要に応じて後で作成）
						overlap_meshes.append(None)  # プレースホルダー
						overlap_areas.append(data.get('area', 0.0))
						overlap_depths.append(data.get('depth', 0.0))
				
				# ヒートマップデータを再構築（進捗表示付き）
				heatmap_data = cache_data.get('heatmap', [])
				total_frames = len(heatmap_data)
				progress_bar['maximum'] = total_frames
				
				# 共通ジオメトリを事前に復元（最適化対応）
				base_mesh_info = cache_data.get('heatmap_base')
				base_mesh_cache = None
				if base_mesh_info:
					try:
						base_mesh_cache = pv.PolyData(base_mesh_info['points'], base_mesh_info['faces'])
						print(f"[キャッシュ] 共通ジオメトリを復元: {base_mesh_cache.n_points}点")
					except Exception as e:
						print(f"[キャッシュ] 共通ジオメトリ復元失敗: {e}")

				print(f"[キャッシュ] ヒートマップ再構築開始: {total_frames}フレーム")
				
				for i, data in enumerate(heatmap_data):
					# キャンセルチェック
					if cancel_var.get():
						progress_window.destroy()
						print("[キャッシュ] 読み込みがキャンセルされました")
						return None
					
					if data is None:
						heatmap_meshes.append(pv.PolyData())
					else:
						mesh = None
						
						# 新形式（最適化版）対応: 共通ジオメトリ利用
						if 'type' in data and data['type'] == 'ref' and base_mesh_cache is not None:
							# ベースメッシュの形状を再利用（メモリ効率・速度向上）
							# deep=Falseでポイント配列を共有、ただしスカラーは別
							mesh = base_mesh_cache.copy(deep=False)
						else:
							# 従来形式または個別保持（full）
							if 'faces' in data and data['faces'] is not None:
								mesh = pv.PolyData(data['points'], data['faces'])
							elif 'points' in data:
								mesh = pv.PolyData(data['points'])
						
						if mesh:
							# スカラーデータ（distance）も復元
							if 'distance' in data:
								mesh['distance'] = data['distance']
							heatmap_meshes.append(mesh)
						else:
							heatmap_meshes.append(pv.PolyData())
					
					# 100フレームごとに進捗更新（UI応答性確保）
					if i % 100 == 0 or i == total_frames - 1:
						progress_bar['value'] = i + 1
						progress_label.config(text=f"{i + 1} / {total_frames} フレーム ({(i+1)/total_frames*100:.1f}%)")
						progress_window.update()
				
				progress_window.destroy()
				
				print(f"[キャッシュ] 読み込み完了: {cache_filepath}")
				if heatmap_meshes:
					print(f"[キャッシュ] ヒートマップデータも読み込み: {len(heatmap_meshes)}フレーム")
				return overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes
			
			else:
				# 旧形式: オーバーラップのみ（後方互換性）
				overlap_meshes = []
				overlap_areas = []
				overlap_depths = []
				
				total_frames = len(cache_data)
				progress_bar['maximum'] = total_frames
				
				for i, data in enumerate(cache_data):
					# キャンセルチェック
					if cancel_var.get():
						progress_window.destroy()
						print("[キャッシュ] 読み込みがキャンセルされました")
						return None
					
					if data is None:
						overlap_meshes.append(None)
						overlap_areas.append(0.0)
						overlap_depths.append(0.0)
					else:
						overlap_meshes.append(None)  # プレースホルダー
						overlap_areas.append(data.get('area', 0.0))
						overlap_depths.append(data.get('depth', 0.0))
					
					# 100フレームごとに進捗更新
					if i % 100 == 0 or i == total_frames - 1:
						progress_bar['value'] = i + 1
						progress_label.config(text=f"{i + 1} / {total_frames} フレーム ({(i+1)/total_frames*100:.1f}%)")
						progress_window.update()
				
				progress_window.destroy()
				
				print(f"[キャッシュ] 読み込み完了（旧形式）: {cache_filepath}")
				return overlap_meshes, overlap_areas, overlap_depths, []
		
		except Exception as e:
			print(f"[キャッシュ] 読み込み失敗: {e}")
			# 進捗ウィンドウを閉じる
			try:
				progress_window.destroy()
			except:
				pass
			return None

	def _precompute_overlap_volumes(self, prox_mesh, dist_mesh, transform_data, progress_callback=None, use_cache=True, use_parallel=True, use_bbox_check=True, simplify_mesh=False):
		"""オーバーラップ体積機能は無効化されています。ヒートマップのみ後段で計算します。

		Returns:
		    tuple: 空の (overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes)
		"""
		print("[事前計算] オーバーラップ体積機能は無効化されています。ヒートマップのみ計算されます。")
		
		# キャッシュが有効な場合、キャッシュファイルの存在を確認
		if use_cache:
			cache_hash = self._get_cache_hash(prox_mesh, dist_mesh, transform_data, simplify_mesh)
			cache_filepath = self._get_cache_filepath(cache_hash)
			
			if cache_filepath.exists():
				# キャッシュファイルのサイズと更新日時を取得
				file_size_mb = cache_filepath.stat().st_size / (1024 * 1024)
				file_mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(cache_filepath.stat().st_mtime))
				
				# ユーザーに確認
				response = messagebox.askyesnocancel(
					"キャッシュファイルが見つかりました",
					f"条件に合うキャッシュファイルが見つかりました。\n\n"
					f"ファイル: {cache_filepath.name}\n"
					f"サイズ: {file_size_mb:.2f} MB\n"
					f"作成日時: {file_mtime}\n"
					f"フレーム数: {len(transform_data)}\n\n"
					f"このキャッシュを使用しますか？\n\n"
					f"「はい」: キャッシュを使用（高速）\n"
					f"「いいえ」: キャッシュを削除して再計算\n"
					f"「キャンセル」: 処理を中断"
				)
				
				if response is None:  # キャンセル
					print("[キャッシュ] 処理がキャンセルされました")
					return [], [], [], []
				elif response:  # はい（キャッシュを使用）
					print(f"[キャッシュ] キャッシュファイルを使用します: {cache_filepath}")
					try:
						result = self._load_overlap_cache(cache_filepath)
						if result:
							overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes = result
							print(f"[キャッシュ] 読み込み完了: {len(overlap_meshes)}フレーム")
							return overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes
						else:
							print("[キャッシュ] 読み込み失敗、再計算します")
					except Exception as e:
						print(f"[キャッシュ] 読み込みエラー: {e}")
						messagebox.showwarning("キャッシュエラー", f"キャッシュの読み込みに失敗しました。\n再計算します。\n\nエラー: {e}")
				else:  # いいえ（キャッシュを削除）
					try:
						cache_filepath.unlink()
						print(f"[キャッシュ] キャッシュファイルを削除しました: {cache_filepath}")
						messagebox.showinfo("キャッシュ削除", "キャッシュファイルを削除しました。\n新しく計算を開始します。")
					except Exception as e:
						print(f"[キャッシュ] 削除エラー: {e}")
						messagebox.showerror("削除エラー", f"キャッシュファイルの削除に失敗しました。\n\nエラー: {e}")
		
		return [], [], [], []

	def _compute_contact_depth(self, prox_mesh, dist_mesh, overlap_mesh=None):
		"""接触深度を計算
		
		接触深度は、2つのメッシュが重なっている領域において、どれだけ侵入しているかを表します。
		複数の方法で計算を試み、最も適切な結果を返します。
		
		方法1: オーバーラップメッシュを使用（最も正確）
		方法2: 遠位メッシュの頂点と近位メッシュの距離
		方法3: 両メッシュの最小距離
		
		Args:
			prox_mesh: 近位メッシュ（PyVistaメッシュ）
			dist_mesh: 遠位メッシュ（PyVistaメッシュ）
			overlap_mesh: オーバーラップメッシュ（オプション）
			
		Returns:
			float: 接触深度（mm）、接触がない場合は0.0
		"""
		try:
			import pyvista as pv
			
			# 方法1: オーバーラップメッシュが提供されている場合
			if overlap_mesh is not None and overlap_mesh.n_points > 0:
				# オーバーラップ領域の中心から両メッシュまでの平均距離を計算
				overlap_points = overlap_mesh.points
				
				# サンプリング（計算時間短縮のため、最大1000点）
				if len(overlap_points) > 1000:
					indices = np.random.choice(len(overlap_points), 1000, replace=False)
					sample_points = overlap_points[indices]
				else:
					sample_points = overlap_points
				
				# 各点から近位メッシュと遠位メッシュへの距離を計算
				prox_distances = []
				dist_distances = []
				
				for point in sample_points:
					# 近位メッシュへの距離
					try:
						idx = prox_mesh.find_closest_point(point)
						if isinstance(idx, (int, np.integer)) and idx >= 0:
							closest_pt = prox_mesh.points[idx]
							prox_dist = np.linalg.norm(point - closest_pt)
							prox_distances.append(prox_dist)
					except Exception:
						pass
					
					# 遠位メッシュへの距離
					try:
						idx = dist_mesh.find_closest_point(point)
						if isinstance(idx, (int, np.integer)) and idx >= 0:
							closest_pt = dist_mesh.points[idx]
							dist_dist = np.linalg.norm(point - closest_pt)
							dist_distances.append(dist_dist)
					except Exception:
						pass
				
				if len(prox_distances) > 0 and len(dist_distances) > 0:
					# 両方の平均距離の和を深度とする
					mean_depth = np.mean(prox_distances) + np.mean(dist_distances)
					print(f"[接触深度] 方法1（オーバーラップメッシュ）: {mean_depth:.4f} mm")
					return mean_depth
			
			# 方法2: 遠位メッシュの頂点から近位メッシュへの最小距離を計算
			# （オーバーラップがあれば、一部の点は非常に近いはず）
			dist_points = dist_mesh.points
			
			# サンプリング（計算時間短縮）
			if len(dist_points) > 2000:
				indices = np.random.choice(len(dist_points), 2000, replace=False)
				sample_dist_points = dist_points[indices]
			else:
				sample_dist_points = dist_points
			
			# 各点から近位メッシュへの距離を計算
			min_distances = []
			for point in sample_dist_points:
				try:
					idx = prox_mesh.find_closest_point(point)
					if isinstance(idx, (int, np.integer)) and idx >= 0:
						closest_pt = prox_mesh.points[idx]
						distance = np.linalg.norm(point - closest_pt)
						min_distances.append(distance)
				except Exception:
					pass
			
			if len(min_distances) > 0:
				# 最小距離の下位10%の平均を深度とする（最も近い点 = 最も侵入している点）
				sorted_distances = np.sort(min_distances)
				n_close = max(1, len(sorted_distances) // 10)
				close_distances = sorted_distances[:n_close]
				mean_depth = np.mean(close_distances)
				print(f"[接触深度] 方法2（最小距離法）: {mean_depth:.4f} mm")
				return mean_depth
			
			# 方法3: 両メッシュ間の最小距離を直接計算
			# これは計算コストが高いため、最後の手段
			try:
				# 近位メッシュの中心付近の点を取得
				prox_center = prox_mesh.center
				prox_points = prox_mesh.points
				
				# 中心に近い点を選択（上位20%）
				distances_to_center = np.linalg.norm(prox_points - prox_center, axis=1)
				n_select = max(100, len(prox_points) // 5)
				close_indices = np.argpartition(distances_to_center, n_select)[:n_select]
				selected_prox_points = prox_points[close_indices]
				
				# 選択された近位点から遠位メッシュへの最小距離
				min_dist_list = []
				for point in selected_prox_points:
					try:
						idx = dist_mesh.find_closest_point(point)
						if isinstance(idx, (int, np.integer)) and idx >= 0:
							closest_pt = dist_mesh.points[idx]
							distance = np.linalg.norm(point - closest_pt)
							min_dist_list.append(distance)
					except Exception:
						pass
				
				if len(min_dist_list) > 0:
					# 最小距離の下位10%の平均
					sorted_distances = np.sort(min_dist_list)
					n_close = max(1, len(sorted_distances) // 10)
					close_distances = sorted_distances[:n_close]
					mean_depth = np.mean(close_distances)
					print(f"[接触深度] 方法3（双方向最小距離）: {mean_depth:.4f} mm")
					return mean_depth
			except Exception as e:
				print(f"[接触深度] 方法3でエラー: {e}")
			
			return 0.0
		
		except Exception as e:
			print(f"[接触深度計算] エラー: {e}")
			import traceback
			traceback.print_exc()
			return 0.0

	def _show_precompute_dialog(self, total_frames, has_cartilage=False):
		"""事前計算の設定と進捗を表示するダイアログ

		Args:
			total_frames: 総フレーム数
			has_cartilage: FEM解析対象（関節領域または軟骨モデル）が利用可能かどうか

		Returns:
			tuple: (dialog_window, progress_var, status_var, cancel_var, options_dict, start_var)
		"""
		# プログレスダイアログを作成
		progress_window = tk.Toplevel(self)
		progress_window.title("事前計算")
		progress_window.geometry("550x750")
		progress_window.resizable(False, False)
		progress_window.grab_set()  # モーダルダイアログにする

		# プログレスバー
		progress_var = tk.DoubleVar()
		status_var = tk.StringVar(value="オプションを設定の上、「計算開始」をクリックしてください")
		cancel_var = tk.BooleanVar(value=False)
		start_var = tk.BooleanVar(value=False)  # 計算開始フラグ
		skip_var = tk.BooleanVar(value=False)  # スキップフラグ

		# 最適化オプション
		enable_precompute_var = tk.BooleanVar(value=True)  # 事前計算を行うかどうか
		use_parallel_var = tk.BooleanVar(value=True)
		use_cache_var = tk.BooleanVar(value=True)
		use_bbox_var = tk.BooleanVar(value=True)
		use_simplify_var = tk.BooleanVar(value=True)
		enable_fem_precompute_var = tk.BooleanVar(value=False)  # FEM事前計算
		
		# タイトル
		title_label = ttk.Label(
			progress_window,
			text="事前計算",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.pack(pady=10)
		
		# 説明文
		info_label = ttk.Label(
			progress_window,
			text=f"フレーム数: {total_frames}\n"
			     f"事前計算することで、アニメーション中の処理が高速化されます。",
			justify=tk.LEFT
		)
		info_label.pack(pady=5)
		
		# 事前計算有効化チェックボックス
		enable_frame = ttk.Frame(progress_window)
		enable_frame.pack(pady=5, padx=10, fill=tk.X)
		
		enable_check = ttk.Checkbutton(
			enable_frame,
			text="ヒートマップを事前計算する（オフの場合、ヒートマップ表示は簡易/無効になります）",
			variable=enable_precompute_var,
			command=lambda: update_option_states()
		)
		enable_check.pack(anchor=tk.W)
		
		# 最適化オプションフレーム
		options_frame = ttk.LabelFrame(progress_window, text="高速化オプション", padding=10, style="Bold.TLabelframe")
		options_frame.pack(pady=5, padx=10, fill=tk.X)
		
		parallel_check = ttk.Checkbutton(
			options_frame,
			text=f"並列処理を使用（{max(1, cpu_count()-1)}コア）",
			variable=use_parallel_var
		)
		parallel_check.pack(anchor=tk.W, pady=2)
		
		cache_check = ttk.Checkbutton(
			options_frame,
			text="キャッシュファイルを使用（次回以降高速化）",
			variable=use_cache_var
		)
		cache_check.pack(anchor=tk.W, pady=2)
		
		bbox_check = ttk.Checkbutton(
			options_frame,
			text="バウンディングボックスチェック（交差なしをスキップ）",
			variable=use_bbox_var
		)
		bbox_check.pack(anchor=tk.W, pady=2)
		
		simplify_check = ttk.Checkbutton(
			options_frame,
			text="メッシュ簡略化（精度低下、高速化）",
			variable=use_simplify_var
		)
		simplify_check.pack(anchor=tk.W, pady=2)

		# FEM事前計算フレーム
		fem_frame = ttk.LabelFrame(progress_window, text="FEM接触解析事前計算", padding=10, style="Bold.TLabelframe")
		fem_frame.pack(pady=5, padx=10, fill=tk.X)

		fem_check = ttk.Checkbutton(
			fem_frame,
			text="FEM接触解析を事前計算する（関節領域で計算・時間がかかります）",
			variable=enable_fem_precompute_var,
			state="normal" if (_HAS_FEM and has_cartilage) else "disabled"
		)
		fem_check.pack(anchor=tk.W, pady=2)

		if not _HAS_FEM:
			fem_check_note = ttk.Label(fem_frame, text="FEMモジュール (fem_contact_solver_2.py) が見つかりません")
			fem_check_note.pack(anchor=tk.W, padx=20, pady=2)
		elif not has_cartilage:
			fem_check_note = ttk.Label(fem_frame, text="関節領域が抽出できていません（半径設定を確認してください）")
			fem_check_note.pack(anchor=tk.W, padx=20, pady=2)

		# オプションの有効/無効を切り替える関数
		def update_option_states():
			state = "normal" if enable_precompute_var.get() else "disabled"
			parallel_check.config(state=state)
			cache_check.config(state=state)
			bbox_check.config(state=state)
			simplify_check.config(state=state)
		
		# 状態表示
		status_label = ttk.Label(progress_window, textvariable=status_var)
		status_label.pack(pady=5)
		
		# プログレスバー
		progress_bar = ttk.Progressbar(
			progress_window,
			variable=progress_var,
			maximum=100,
			length=500,
			mode='determinate'
		)
		progress_bar.pack(pady=10)
		
		# パーセンテージ表示
		percent_label = ttk.Label(progress_window, text="0%")
		percent_label.pack(pady=5)
		
		# ボタンフレーム
		button_frame = ttk.Frame(progress_window)
		button_frame.pack(pady=10)
		
		# 計算開始ボタン
		def start_computation():
			if not enable_precompute_var.get():
				# 事前計算をスキップする場合
				skip_var.set(True)
				progress_window.destroy()
			else:
				# 計算開始
				start_var.set(True)
				start_button.config(state="disabled")
				skip_button.config(state="disabled")
				# オプションを無効化
				enable_check.config(state="disabled")
				parallel_check.config(state="disabled")
				cache_check.config(state="disabled")
				bbox_check.config(state="disabled")
				simplify_check.config(state="disabled")
		
		start_button = ttk.Button(
			button_frame,
			text="計算開始",
			command=start_computation
		)
		start_button.pack(side=tk.LEFT, padx=5)
		
		# スキップボタン（計算せずに進む）
		def skip_computation():
			skip_var.set(True)
			progress_window.destroy()
		
		skip_button = ttk.Button(
			button_frame,
			text="計算せずに進む",
			command=skip_computation
		)
		skip_button.pack(side=tk.LEFT, padx=5)
		
		# キャンセルボタン
		def cancel_computation():
			cancel_var.set(True)
			progress_window.destroy()
		
		cancel_button = ttk.Button(
			button_frame,
			text="キャンセル",
			command=cancel_computation
		)
		cancel_button.pack(side=tk.LEFT, padx=5)
		
		# 進捗更新関数
		def update_progress(current, total, message):
			if cancel_var.get():
				return False  # キャンセルされた
			
			percentage = (current / total) * 100
			progress_var.set(percentage)
			status_var.set(message)
			percent_label.config(text=f"{percentage:.1f}%")
			progress_window.update()
			return True  # 継続
		
		# ウィンドウを中央に配置
		progress_window.transient(self)
		progress_window.update_idletasks()
		x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
		y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
		progress_window.geometry(f"+{x}+{y}")
		
		options_dict = {
			'use_parallel': use_parallel_var,
			'use_cache': use_cache_var,
			'use_bbox': use_bbox_var,
			'use_simplify': use_simplify_var,
			'enable_precompute': enable_precompute_var,
			'enable_fem_precompute': enable_fem_precompute_var
		}

		return progress_window, update_progress, cancel_var, options_dict, start_var, skip_var

	def show_fitting_help(self) -> None:
		"""フィッティングパラメータの説明を表示"""
		help_text = (
			"【パラメータの意味】\n\n"
			"■ RANSACパラメータ（粗い位置合わせ）\n"
			"  ・距離閾値: 対応点とみなす最大距離 (mm)\n"
			"  ・最大反復回数: ランダムサンプリングの試行回数\n"
			"  ・信頼度: 良い変換を見つける確率 (0.0～1.0)\n\n"
			"■ ICPパラメータ（精密な位置合わせ）\n"
			"  ・閾値: 対応点とみなす最大距離 (mm)\n"
			"  ・最大反復回数: 位置合わせの反復回数\n"
			"  ・サンプル点数: 計算に使う点の数（多いほど精密だが遅い）"
		)
		messagebox.showinfo("パラメータの説明", help_text)
	
	def reset_fitting_parameters(self) -> None:
		"""フィッティングパラメータをデフォルト値に戻す"""
		
		# RANSACパラメータ
		self.ransac_distance_threshold.set(1.0)
		self.ransac_max_iterations.set(1000)
		self.ransac_confidence.set(0.99)
		
		# ICPパラメータ
		self.icp_threshold.set(5.0)
		self.icp_max_iterations.set(2000)
		self.sample_points.set(10000)
		
		messagebox.showinfo("完了", "パラメータをデフォルト値に戻しました")

	def on_fitting_execute(self) -> None:
		"""
		【重要メソッド】フィッティング実行（Open3Dベース）
		
		RANSACとICPを使用して、親任意領域に子任意領域を位置合わせします。
		
		■ 処理フロー:
		  1. 親任意領域・子任意領域メッシュ読み込み
		  2. 主軸による事前アライメント（重心と主軸を一致）
		  3. RANSACによる粗い位置合わせ
		     - FPFHフィーチャー抽出
		     - ランダムサンプリングで最適変換を探索
		  4. ICPによる精密な位置合わせ
		     - 点対点ICP（Point-to-Point）
		     - 反復的に最適変換を収束
		  5. 結果の可視化と保存オプション表示
		  
		■ パラメータ:
		  - RANSACパラメータ（UI設定）:
		    * 距離閾値: 対応点とみなす最大距離 (mm)
		    * 最大反復回数: ランダムサンプリングの試行回数
		    * 信頼度: 良い変換を見つける確率 (0.0～1.0)
		  - ICPパラメータ（UI設定）:
		    * 閾値: 対応点とみなす最大距離 (mm)
		    * 最大反復回数: 位置合わせの反復回数
		    * サンプル点数: 計算に使う点の数
		  
		■ 変換行列の合成:
		  最終的な変換行列 = ICP変換 × RANSAC変換 × 主軸アライメント変換
		  この変換を子モデル全体に適用することで、フィッティング結果を得ます。
		  
		■ 改修ポイント:
		  - 行3180付近: RANSAC実装（FPFHフィーチャー計算）
		  - 行3250付近: ICP実装（反復最適化）
		  - 行3360付近: 結果メニュー表示
		  
		■ トラブルシューティング:
		  - RANSAC失敗: 距離閾値を大きく、または事前アライメントを確認
		  - ICP収束不良: 閾値を大きく、または反復回数を増やす
		  - メッシュ読み込み失敗: trimeshライブラリのインストールを確認
		"""
		# 必要なファイルがすべて選択されているか確認
		if not all([
			self.fitting_parent_region_path.get(),
			self.fitting_child_region_path.get()
		]):
			messagebox.showerror("エラー", "親任意領域と子任意領域を選択してください。")
			return
		
		try:
			import copy
			import trimesh # type: ignore
			
			# メッシュファイルを読み込む（Open3D形式）
			def load_mesh_o3d(file_path):
				"""メッシュファイルをOpen3Dで読み込む"""
				try:
					print(f"ファイル読み込み中: {Path(file_path).name}")
					# trimeshで読み込み
					trimesh_mesh = trimesh.load(file_path, force='mesh')
					vertices = np.array(trimesh_mesh.vertices)
					triangles = np.array(trimesh_mesh.faces)
					
					# Open3Dメッシュに変換
					mesh = o3d.geometry.TriangleMesh()
					mesh.vertices = o3d.utility.Vector3dVector(vertices)
					mesh.triangles = o3d.utility.Vector3iVector(triangles)
					mesh.compute_vertex_normals()
					
					print(f"  ✓ 成功: 頂点数={len(vertices)}, 三角形数={len(triangles)}")
					return mesh
				except Exception as e:
					print(f"  ✗ エラー: {e}")
					messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました:\n{Path(file_path).name}\n\n{str(e)}")
					return None
			
			# 2つの任意領域モデルを読み込み
			print("\n=== フィッティング処理開始 ===")
			print(f"親任意領域: {Path(self.fitting_parent_region_path.get()).name}")
			print(f"子任意領域: {Path(self.fitting_child_region_path.get()).name}")
			
			parent_region_mesh = load_mesh_o3d(self.fitting_parent_region_path.get())
			child_region_mesh = load_mesh_o3d(self.fitting_child_region_path.get())
			
			if not all([parent_region_mesh, child_region_mesh]):
				return
			
			# 色を設定
			parent_region_mesh.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_region_mesh.paint_uniform_color([0.0, 0.0, 1.0])  # 青
			
			# 初期状態を表示
			print("\n[初期状態] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_mesh)
			o3d.visualization.draw_geometries([parent_display, child_display],
										  window_name="初期状態: 親(赤) + 子(青)")
			# ウィンドウサイズを復元
			self._restore_window_geometry()
			
			# 主軸による事前アライメント
			print("\n=== 主軸による事前アライメント ===")
			
			# 主軸計算関数
			def get_principal_axis(mesh):
				"""メッシュの主軸（最大固有ベクトル）と重心を計算"""
				pcd = mesh.sample_points_uniformly(number_of_points=20000)
				pts = np.asarray(pcd.points)
				center = np.mean(pts, axis=0)
				cov = np.cov((pts - center).T)
				eigvals, eigvecs = np.linalg.eigh(cov)
				principal_axis = eigvecs[:, np.argmax(eigvals)]
				return center, principal_axis
			
			# 主軸回転関数
			def align_axis(mesh, src_axis, tgt_axis):
				"""主軸の向きを合わせる"""
				v = np.cross(src_axis, tgt_axis)
				c = np.dot(src_axis, tgt_axis)
				if np.linalg.norm(v) < 1e-8:
					R = np.eye(3)
				else:
					vx = np.array([[0, -v[2], v[1]],
								   [v[2], 0, -v[0]],
								   [-v[1], v[0], 0]])
					R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
				mesh_rot = copy.deepcopy(mesh)
				mesh_rot.rotate(R, center=(0, 0, 0))
				return mesh_rot, R
			
			# 主軸を計算
			c_parent, axis_parent = get_principal_axis(parent_region_mesh)
			c_child, axis_child = get_principal_axis(child_region_mesh)
			
			print(f"親の重心: {c_parent}")
			print(f"親の主軸: {axis_parent}")
			print(f"子の重心: {c_child}")
			print(f"子の主軸: {axis_child}")
			
			# 主軸の向きを一致させる
			if np.dot(axis_parent, axis_child) < 0:
				axis_child = -axis_child
				print("主軸の向きを反転しました")
			
			# 子メッシュの主軸を親メッシュに合わせる
			child_region_aligned, R_align = align_axis(child_region_mesh, axis_child, axis_parent)
			
			# 重心を一致させる
			c_child_aligned, _ = get_principal_axis(child_region_aligned)
			translation_vector = c_parent - c_child_aligned
			child_region_aligned.translate(translation_vector, relative=True)
			
			# 主軸アライメントの変換行列を構築（子の元の位置から変換後の位置への変換）
			# 1. 元の子の重心を原点に移動
			# 2. 回転を適用
			# 3. 新しい位置（親の重心）に移動
			axis_transform = np.eye(4)
			axis_transform[:3, :3] = R_align
			axis_transform[:3, 3] = c_parent - R_align @ c_child
			
			print(f"主軸アライメント変換行列:")
			print(axis_transform)
			
			# 主軸・重心一致後を表示
			print("\n[主軸アライメント後] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_aligned)
			parent_display.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_display.paint_uniform_color([0.0, 1.0, 0.0])  # 緑
			
			# 主軸を線で表示
			line_set = o3d.geometry.LineSet()
			points = [c_parent, c_parent + axis_parent*50, c_parent, c_parent + axis_child*50]
			lines = [[0, 1], [2, 3]]
			colors = [[1, 0, 0], [0, 0, 1]]
			line_set.points = o3d.utility.Vector3dVector(points)
			line_set.lines = o3d.utility.Vector2iVector(lines)
			line_set.colors = o3d.utility.Vector3dVector(colors)
			
			o3d.visualization.draw_geometries([parent_display, child_display, line_set],
										  window_name="主軸・重心一致後: 親(赤) + 子(緑)")
			# RANSACによる粗い位置合わせ
			print("\n=== RANSAC位置合わせ開始 ===")
			
			# 点群に変換
			parent_pcd = parent_region_mesh.sample_points_uniformly(number_of_points=10000)
			child_pcd = child_region_aligned.sample_points_uniformly(number_of_points=10000)
			
			print(f"点群サンプル数 - 親: {len(parent_pcd.points)}, 子: {len(child_pcd.points)}")
			
			# RANSACパラメータを取得
			# 信頼度は0.0-1.0のfloatとして扱う（UI値のスケーリングバグ修正）
			max_iterations = int(self.ransac_max_iterations.get())
			confidence = float(self.ransac_confidence.get())
			
			# voxel_sizeを点群から自動推定（平均最近傍距離に基づく）
			try:
				num_points = len(parent_pcd.points)
				sample_n = min(500, max(10, num_points))
				sample_indices = np.random.choice(num_points, sample_n, replace=False)
				kdt = o3d.geometry.KDTreeFlann(parent_pcd)
				nn_dists = []
				for idx in sample_indices:
					[_, idxs, dists] = kdt.search_knn_vector_3d(parent_pcd.points[idx], 2)
					if len(dists) >= 2:
						nn_dists.append(float(np.sqrt(dists[1])))
				if len(nn_dists) > 0:
					mean_nn = float(np.mean(nn_dists))
					voxel_size = max(mean_nn * 1.5, 1e-6)
				else:
					voxel_size = max(self.ransac_distance_threshold.get(), 1e-6)
			except Exception as e:
				voxel_size = max(self.ransac_distance_threshold.get(), 1e-6)
				print(f"[警告] voxel_size自動推定に失敗しました: {e}. UI値を使用します: {voxel_size}")
			
			print(f"RANSACパラメータ: voxel_size={voxel_size:.6f}, max_iterations={max_iterations}, confidence={confidence}")
			# デバッグ: バウンディングボックス情報
			minb = parent_region_mesh.get_min_bound()
			maxb = parent_region_mesh.get_max_bound()
			bbox_diag = np.linalg.norm(maxb - minb)
			print(f"親メッシュバウンディングボックス: min={minb}, max={maxb}, diag={bbox_diag:.3f}")
			
			# ボクセルダウンサンプリング
			parent_down = parent_pcd.voxel_down_sample(voxel_size)
			child_down = child_pcd.voxel_down_sample(voxel_size)
			
			print(f"ダウンサンプリング後 - 親: {len(parent_down.points)}, 子: {len(child_down.points)}")
			
			# 法線を計算
			radius_normal = voxel_size * 2
			parent_down.estimate_normals(
				search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			child_down.estimate_normals(
				search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			
			# FPFHフィーチャーを計算
			radius_feature = voxel_size * 5
			print(f"FPFH計算中... (radius: {radius_feature})")
			
			parent_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				parent_down,
				o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			child_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				child_down,
				o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			
			# RANSACによる位置合わせ
			distance_threshold = voxel_size * 1.5
			print(f"RANSAC実行中... (distance_threshold: {distance_threshold})")
			
			result_ransac = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
				child_down, parent_down, child_fpfh, parent_fpfh,
				mutual_filter=True,
				max_correspondence_distance=distance_threshold,
				estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPoint(False),
				ransac_n=4,
				checkers=[
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(0.9),
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(distance_threshold)
				],
				criteria=o3d.pipelines.registration.RANSACConvergenceCriteria(max_iterations, confidence))
			
			ransac_transform = result_ransac.transformation
			print(f"RANSAC完了 - Fitness: {result_ransac.fitness:.4f}, RMSE: {result_ransac.inlier_rmse:.4f}")
			print(f"変換行列:\n{ransac_transform}")
			
			# RANSAC結果をメッシュに適用
			child_region_ransac = copy.deepcopy(child_region_aligned)
			child_region_ransac.transform(ransac_transform)
			
			# RANSAC結果を可視化
			print("\n[RANSAC結果] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_ransac)
			parent_display.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_display.paint_uniform_color([0.0, 1.0, 1.0])  # シアン
			o3d.visualization.draw_geometries([parent_display, child_display],
										  window_name="RANSAC後: 親(赤) + 子(シアン)")
			# ウィンドウサイズを復元
			self._restore_window_geometry()
			
			# ICPによる精密な位置合わせ
			print("\n=== ICP位置合わせ開始 ===")
			threshold = self.icp_threshold.get()
			if threshold <= 0.0:
				print(f"[警告] ICP閾値が0以下({threshold})です。デフォルトの閾値(1e-6)を使用します。")
				threshold = 1e-6
			max_iterations = self.icp_max_iterations.get()
			
			print(f"ICPパラメータ: threshold={threshold}, max_iterations={max_iterations}")
			
			# 点群に変換
			parent_icp_pcd = parent_region_mesh.sample_points_uniformly(number_of_points=20000)
			child_icp_pcd = child_region_ransac.sample_points_uniformly(number_of_points=20000)
			
			print(f"ICP実行中... (点数: 親={len(parent_icp_pcd.points)}, 子={len(child_icp_pcd.points)})")
			
			# ICPによる位置合わせ
			result_icp = o3d.pipelines.registration.registration_icp(
				child_icp_pcd, parent_icp_pcd, threshold, np.eye(4),
				o3d.pipelines.registration.TransformationEstimationPointToPoint(),
				o3d.pipelines.registration.ICPConvergenceCriteria(max_iteration=max_iterations))
			
			icp_transform = result_icp.transformation
			print(f"ICP完了 - Fitness: {result_icp.fitness:.4f}, RMSE: {result_icp.inlier_rmse:.4f}")
			print(f"変換行列:\n{icp_transform}")
			
			# ICP結果をメッシュに適用
			child_region_final = copy.deepcopy(child_region_ransac)
			child_region_final.transform(icp_transform)
			
			# ICP結果を可視化
			print("\n[ICP結果(最終)] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_final)
			parent_display.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_display.paint_uniform_color([0.0, 0.0, 1.0])  # 青
			o3d.visualization.draw_geometries([parent_display, child_display],
										  window_name="ICP後(最終結果): 親(赤) + 子(青)")
			# ウィンドウサイズを復元
			self._restore_window_geometry()
			
			print("\n=== フィッティング処理完了 ===")
			
			# ユーザーに結果の確認を求める
			result = messagebox.askyesno(
				"フィッティング結果の確認",
				"フィッティング結果に満足しましたか？\n\n"
				"「はい」を選択すると、変換後の操作メニューを表示します。"
			)
			
			if not result:
				print("フィッティング処理をキャンセルしました")
				messagebox.showinfo("キャンセル", "フィッティング処理をキャンセルしました")
				return
			
			# 子モデルに変換を適用して新しい子モデルを作成
			print("\n=== 子モデルへの変換適用 ===")
			
			# 子モデル（フルモデル）を読み込み
			child_model_path = self.fitting_child_model_path.get()
			if not child_model_path:
				messagebox.showerror("エラー", "子モデルが選択されていません")
				return
			
			print(f"子モデル読み込み中: {Path(child_model_path).name}")
			child_model_mesh = load_mesh_o3d(child_model_path)
			print(f"  ✓ 頂点数={len(child_model_mesh.vertices)}, 三角形数={len(child_model_mesh.triangles)}")
			
			# 合成変換行列を計算（主軸アライメント + RANSAC + ICP）
			# フィッティングで求めた変換：子の任意領域 → 親の任意領域
			print("\n=== 変換行列の合成 ===")
			print("1. 主軸アライメント変換（既に計算済み）")
			print("2. RANSAC変換:")
			print(ransac_transform)
			print("3. ICP変換:")
			print(icp_transform)
			
			# 全変換を合成（子→親の変換）
			child_to_parent_transform = icp_transform @ ransac_transform @ axis_transform
			
			print("子→親の変換行列:")
			print(child_to_parent_transform)
			
			# 子モデル全体に変換を適用
			child_model_transformed = copy.deepcopy(child_model_mesh)
			child_model_transformed.transform(child_to_parent_transform)
			print("  ✓ 子モデルに変換を適用しました")
			
			# 統合GUIを表示
			self._show_fitting_result_menu(
				child_model_transformed,
				child_to_parent_transform,
				child_model_path,
				load_mesh_o3d
			)
			
		except Exception as e:
			messagebox.showerror("エラー", f"フィッティング処理中にエラーが発生しました:\n{str(e)}")
			print(f"エラー詳細: {e}")
			import traceback
			traceback.print_exc()
		finally:
			# メインウィンドウのサイズを復元
			self._restore_window_geometry()
	
	def _show_fitting_result_menu(self, child_model_transformed, child_to_parent_transform, child_model_path, load_mesh_o3d):
		"""フィッティング結果の操作メニューGUIを表示
		
		Args:
			child_model_transformed: 変換後の子モデル（Open3D）
			child_to_parent_transform: 変換行列
			child_model_path: 元の子モデルパス
			load_mesh_o3d: メッシュ読み込み関数
		"""
		# 新しいウィンドウを作成
		menu_window = tk.Toplevel(self)
		menu_window.title("フィッティング完了 - 操作メニュー")
		menu_window.geometry("800x850") # サイズ変更
		menu_window.resizable(True, True) # リサイズ可能に
		
		# ウィンドウを中央に配置
		menu_window.update_idletasks()
		x = (menu_window.winfo_screenwidth() // 2) - (800 // 2)
		y = (menu_window.winfo_screenheight() // 2) - (850 // 2)
		menu_window.geometry(f"+{x}+{y}")
		
		# メインフレーム
		main_frame = ttk.Frame(menu_window, padding=20)
		main_frame.pack(fill=tk.BOTH, expand=True)
		
		# タイトル
		title_label = ttk.Label(
			main_frame,
			text="フィッティング完了！",
			font=(self.ui_font_family, 14, "bold")
		)
		title_label.pack(pady=(0, 10))

		# 同時変換行列の表示（テキストエリア）
		matrix_frame = ttk.LabelFrame(main_frame, text="計算された同時変換行列 (Child ➙ Parent)", padding=10)
		matrix_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

		matrix_text = tk.Text(matrix_frame, height=14, width=65, font=(self.mono_font_family, 10))
		matrix_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		
		matrix_scrollbar = ttk.Scrollbar(matrix_frame, orient="vertical", command=matrix_text.yview)
		matrix_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
		matrix_text.config(yscrollcommand=matrix_scrollbar.set)
		
		# 行列文字列作成
		m = child_to_parent_transform
		
		# PosCheckerと同じフォーマットで解析情報を付与
		result_text = "【フィッティング結果変換行列 (Child ➙ Parent)】\n"
		result_text += "※ Child(移動側)からParent(基準側)への座標変換\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{m[0][0]:12.6f} {m[0][1]:12.6f} {m[0][2]:12.6f} {m[0][3]:12.6f}\n"
		result_text += f"{m[1][0]:12.6f} {m[1][1]:12.6f} {m[1][2]:12.6f} {m[1][3]:12.6f}\n"
		result_text += f"{m[2][0]:12.6f} {m[2][1]:12.6f} {m[2][2]:12.6f} {m[2][3]:12.6f}\n"
		result_text += f"{m[3][0]:12.6f} {m[3][1]:12.6f} {m[3][2]:12.6f} {m[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"
		
		# 解析 (Translation / Rotation)
		trans = m[:3, 3]
		norm_val = np.linalg.norm(trans)
		rot_mat = m[:3, :3]
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0
			
		result_text += "【解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Norm (mm)       : {norm_val:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"
		
		matrix_text.insert("1.0", result_text)
		matrix_text.config(state="disabled") # 編集不可

		# 行列保存ボタン
		def save_matrix_txt():
			matrix_path = filedialog.asksaveasfilename(
				title="同時変換行列を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if matrix_path:
				try:
					with open(matrix_path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"行列を保存しました:\n{matrix_path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(matrix_frame, text="行列をテキストとして保存", command=save_matrix_txt).pack(anchor="e", pady=(5, 0))

		# 説明
		info_label = ttk.Label(
			main_frame,
			text="変換が完了しました。以下のボタンから必要な操作を選択してください。",
			wraplength=750
		)
		info_label.pack(pady=(0, 20))
		
		# ボタンフレーム
		button_frame = ttk.Frame(main_frame)
		button_frame.pack(fill=tk.BOTH, expand=True)
		
		# 保存済みフラグ
		model_saved = [False]  # リストで包んで参照渡し
		pp_saved = [False]
		saved_model_path = [None]
		
		# ヒートマップデータキャッシュ
		heatmap_data_cache = {
			'distances': None,
			'points': None,
			'info': ""
		}
		
		# 1. 変換後のモデル保存
		def save_model():
			original_child_path = Path(child_model_path)
			init_dir = str(original_child_path.parent)
			init_name = original_child_path.stem + ".Fitted.stl"
			output_path = filedialog.asksaveasfilename(
				title="変換後の子モデルを保存",
				defaultextension=".stl",
				filetypes=[
					("STL files", "*.stl"),
					("OBJ files", "*.obj"),
					("PLY files", "*.ply"),
					("All files", "*.*")
				],
				initialfile=init_name,
				initialdir=init_dir
			)
			
			if output_path:
				print(f"\n変換後の子モデルを保存中: {Path(output_path).name}")
				o3d.io.write_triangle_mesh(output_path, child_model_transformed)
				print(f"  ✓ 保存完了: {output_path}")
				model_saved[0] = True
				saved_model_path[0] = output_path
				btn_save_model.config(text="✓ モデル保存完了", state="disabled")
				messagebox.showinfo("保存完了", f"モデルを保存しました:\n{output_path}")
		
		btn_save_model = ttk.Button(
			button_frame,
			text="変換後のモデルを保存",
			command=save_model
		)
		btn_save_model.pack(fill=tk.X, pady=5)
		
		# 2. 変換後のPP保存
		def save_pp():
			try:
				child_pp_path = self.fitting_child_pp_path.get()
				if not child_pp_path:
					child_pp_path = filedialog.askopenfilename(
						title="子PPファイルを選択",
						filetypes=[("PickedPoints/XML", "*.pp;*.xml;*.txt"), ("All files", "*.*")]
					)
					if not child_pp_path:
						return
					self.fitting_child_pp_path.set(child_pp_path)
				
				# PP読み込み
				points, labels = self._parse_pp_file(child_pp_path)
				filtered_points = np.array(points)
				filtered_labels = list(labels)
				
				# 変換適用
				if len(filtered_points) > 0:
					homo = np.hstack([filtered_points, np.ones((len(filtered_points), 1))])
					transformed = (child_to_parent_transform @ homo.T).T[:, :3]
				else:
					transformed = np.empty((0, 3))
				
				# 保存ダイアログ
				pp_dir = str(Path(child_pp_path).parent)
				base_name = Path(child_pp_path).stem
				out_name = base_name + ".Fitted_points.pp"
				save_pp_path = filedialog.asksaveasfilename(
					title="変換後の子PPを保存",
					defaultextension=".pp",
					filetypes=[("PickedPoints/XML", "*.pp;*.xml"), ("All files", "*.*")],
					initialdir=pp_dir,
					initialfile=out_name
				)
				
				if save_pp_path:
					save_child_pp_file(save_pp_path, transformed, filtered_labels)
					print(f"  ✓ 子PP保存完了: {save_pp_path}")
					pp_saved[0] = True
					btn_save_pp.config(text="✓ PP保存完了", state="disabled")
					messagebox.showinfo("保存完了", f"子PPを保存しました:\n{save_pp_path}")
			except Exception as e:
				print(f"[PP保存] エラー: {e}")
				messagebox.showerror("エラー", f"子PP保存でエラーが発生しました:\n{e}")
		
		btn_save_pp = ttk.Button(
			button_frame,
			text="変換後のPPを保存",
			command=save_pp
		)
		btn_save_pp.pack(fill=tk.X, pady=5)
		
		# 区切り線
		separator = ttk.Separator(button_frame, orient='horizontal')
		separator.pack(fill=tk.X, pady=(15, 5))

		# --- 可視化設定 (色設定) ---
		settings_frame = ttk.LabelFrame(button_frame, text="可視化の色設定", padding=10)
		settings_frame.pack(fill=tk.X, pady=(0, 15))
		
		# デフォルト色設定
		viz_colors = {
			'parent': '#FF0000',     # 赤
			'child': '#00FF00',      # 緑
			'cmap': 'turbo'          # ヒートマップ
		}
		
		# 色選択コールバック
		def pick_parent_color():
			c = colorchooser.askcolor(color=viz_colors['parent'], title="親モデルの色を選択")[1]
			if c:
				viz_colors['parent'] = c
				lbl_parent_color.config(bg=c)

		def pick_child_color():
			c = colorchooser.askcolor(color=viz_colors['child'], title="子モデルの色を選択")[1]
			if c:
				viz_colors['child'] = c
				lbl_child_color.config(bg=c)
				
		# 親モデル色
		f_colors = ttk.Frame(settings_frame)
		f_colors.pack(fill=tk.X)
		
		f_p = ttk.Frame(f_colors)
		f_p.pack(side=tk.LEFT, padx=5)
		ttk.Button(f_p, text="親モデル色", command=pick_parent_color, width=12).pack(side=tk.LEFT)
		lbl_parent_color = tk.Label(f_p, bg=viz_colors['parent'], width=4, relief="solid")
		lbl_parent_color.pack(side=tk.LEFT, padx=2)
		
		# 子モデル色
		f_c = ttk.Frame(f_colors)
		f_c.pack(side=tk.LEFT, padx=5)
		ttk.Button(f_c, text="子モデル色", command=pick_child_color, width=12).pack(side=tk.LEFT)
		lbl_child_color = tk.Label(f_c, bg=viz_colors['child'], width=4, relief="solid")
		lbl_child_color.pack(side=tk.LEFT, padx=2)

		# 3. 普通に可視化
		def visualize_simple():
			print("\n[可視化] Open3Dで表示中...")
			
			# hex to rgb helper
			def h2r(h):
				h = h.lstrip('#')
				return [int(h[i:i+2], 16)/255.0 for i in (0, 2, 4)]
			
			p_col = h2r(viz_colors['parent'])
			c_col = h2r(viz_colors['child'])
			
			geoms = []
			
			# 親モデル
			parent_model_path = self.fitting_parent_model_path.get()
			if parent_model_path:
				parent_model_mesh = load_mesh_o3d(parent_model_path)
				parent_model_mesh.paint_uniform_color(p_col)
				geoms.append(parent_model_mesh)
			
			# 変換後の子モデル
			child_display = copy.deepcopy(child_model_transformed)
			child_display.paint_uniform_color(c_col)
			geoms.append(child_display)
			
			o3d.visualization.draw_geometries(
				geoms,
				window_name="フィッティング結果: シンプル表示"
			)
			# ウィンドウサイズを復元
			self._restore_window_geometry()
		
		btn_visualize_simple = ttk.Button(
			button_frame,
			text="普通に可視化する",
			command=visualize_simple
		)
		btn_visualize_simple.pack(fill=tk.X, pady=5)
		
		# 4. ヒートマップありで可視化
		def visualize_heatmap():
			print("\n[ヒートマップ可視化] 準備中...")
			
			# 優先順位: 指定されたヒートマップ用モデル > (ダイアログ選択: 任意領域 vs 全体)
			heatmap_model_path_val = self.fitting_child_heatmap_model_path.get().strip()
			use_region_only = False
			
			# ヒートマップ表示用モデルが指定されていない場合のみ、ダイアログを表示
			if not heatmap_model_path_val:
				# 子任意領域のみを対象にするか確認（カスタムダイアログ）
				selection_var = tk.IntVar(value=-1) # 1: Region, 0: Full
				
				d = tk.Toplevel(menu_window)
				d.title("ヒートマップ対象選択")
				d.geometry("450x230")
				d.resizable(False, False)
				
				# 中央配置
				d.update_idletasks()
				dx = (d.winfo_screenwidth() // 2) - (450 // 2)
				dy = (d.winfo_screenheight() // 2) - (230 // 2)
				d.geometry(f"+{dx}+{dy}")
				
				ttk.Label(d, text="距離計算の基準メッシュを選択してください", font=(self.ui_font_family, 11, "bold")).pack(pady=(20, 10))
				ttk.Label(d, text="どちらのモデルに対して距離ヒートマップを表示しますか？", font=(self.ui_font_family, 10)).pack(pady=(0, 20))
				
				def select_region():
					selection_var.set(1)
					d.destroy()
					
				def select_full():
					selection_var.set(0)
					d.destroy()
					
				btn_region = ttk.Button(d, text="子任意領域のみ (高速)", command=select_region, width=30)
				btn_region.pack(pady=2)
				ttk.Label(d, text="※ フィッティングに使用した部分のみ", font=(self.ui_font_family, 9), foreground="gray").pack(pady=(0, 12))
				
				btn_full = ttk.Button(d, text="子モデル全体 (低速)", command=select_full, width=30)
				btn_full.pack(pady=2)
				ttk.Label(d, text="※ 全体の干渉を確認できます（計算に時間がかかります）", font=(self.ui_font_family, 9), foreground="gray").pack(pady=(0, 10))
				
				d.transient(menu_window)
				d.grab_set()
				menu_window.wait_window(d)
				
				if selection_var.get() == -1:
					print("ヒートマップ表示キャンセル")
					return

				use_region_only = (selection_var.get() == 1)
			
			# 親モデル読み込み
			parent_model_path = self.fitting_parent_model_path.get()
			if not parent_model_path:
				messagebox.showerror("エラー", "親モデルが選択されていません")
				return
			
			parent_model_mesh_pv = pv.read(parent_model_path)
			
			heatmap_mesh_to_display = None

			# ヒートマップ表示用モデルがある場合（子モデル側にヒートマップを表示）
			if heatmap_model_path_val:
				print(f"ヒートマップ表示用モデルを使用: {Path(heatmap_model_path_val).name}")
				# 指定モデルを読み込み、フィッティング変換を適用
				hm_o3d = load_mesh_o3d(heatmap_model_path_val)
				if hm_o3d is None:
					messagebox.showerror("エラー", "ヒートマップ表示用モデルの読み込みに失敗しました")
					return
				hm_transformed = copy.deepcopy(hm_o3d)
				hm_transformed.transform(child_to_parent_transform)
				verts = np.asarray(hm_transformed.vertices)
				tris = np.asarray(hm_transformed.triangles)
				faces_pv = np.hstack([np.full((len(tris), 1), 3), tris]) if len(tris) > 0 else np.empty((0, 4), dtype=int)
				
				# 計算対象（Subject）: ヒートマップ表示用モデル（子）
				subject_vertices = verts
				# メッシュが巨大な場合は簡略化（ユーザー要望: 軽量化）
				if len(verts) > 30000:
					print(f"  メッシュ簡略化中 ({len(verts)} points)...")
					tmp_mesh = pv.PolyData(verts, faces_pv)
					subject_mesh = tmp_mesh.decimate(target_reduction=0.5)
					subject_vertices = np.array(subject_mesh.points)
					print(f"  -> 簡略化後: {len(subject_vertices)} points")
				else:
					subject_mesh = pv.PolyData(verts, faces_pv)
				
				# 距離基準（Target）: 親モデル全体
				# ユーザー提案: バウンディングボックス（長方形）で親モデルを切り抜く（ROI抽出）
				print(f"  距離計算用ターゲット: 親モデル全体（ROI抽出）")
				
				# バウンディングボックス計算
				bounds = subject_mesh.bounds # (xmin, xmax, ymin, ymax, zmin, zmax)
				margin = 15.0 # mm (接触判定用なので余裕を持たせる。これ以上離れた点は精密計算しても意味が薄い)
				
				clip_box_bounds = [
					bounds[0] - margin, bounds[1] + margin,
					bounds[2] - margin, bounds[3] + margin,
					bounds[4] - margin, bounds[5] + margin
				]
				
				try:
					# PyVistaのclip_boxでROI抽出
					# invert=False: 箱の中身を残す
					print(f"  親モデルをクリッピング中 (ROI Margin: {margin}mm)...")
					target_mesh = parent_model_mesh_pv.clip_box(clip_box_bounds, invert=False)
					
					# クリップ結果が空（全く重なっていない）場合の対策
					if target_mesh.n_points == 0:
						print("  警告: ROI内に親モデルが存在しません。親モデル全体を使用します。")
						target_mesh = parent_model_mesh_pv
					else:
						print(f"  -> 親モデル要素数: {parent_model_mesh_pv.n_points} -> {target_mesh.n_points} points")
				
				except Exception as e:
					print(f"  クリッピング失敗: {e}")
					target_mesh = parent_model_mesh_pv
				
				calc_label = "距離ヒートマップを計算中（子ヒートマップモデル(簡略) → 親モデル(ROI)）..."
				
			# 指定がない場合（親モデル側にヒートマップを表示して、子との距離を見る）
			else:
				if use_region_only:
					print("子任意領域に変換を適用中...")
					child_region_path = self.fitting_child_region_path.get()
					child_region_mesh_o3d = load_mesh_o3d(child_region_path)
					child_region_transformed = copy.deepcopy(child_region_mesh_o3d)
					child_region_transformed.transform(child_to_parent_transform)
					region_vertices = np.asarray(child_region_transformed.vertices)
					region_triangles = np.asarray(child_region_transformed.triangles)
					faces_pv = np.hstack([np.full((len(region_triangles), 1), 3), region_triangles])
					reference_mesh = pv.PolyData(region_vertices, faces_pv)
				else:
					child_vertices = np.asarray(child_model_transformed.vertices)
					child_triangles = np.asarray(child_model_transformed.triangles)
					faces_pv = np.hstack([np.full((len(child_triangles), 1), 3), child_triangles])
					reference_mesh = pv.PolyData(child_vertices, faces_pv)
				
				# 計算対象（Subject）: 親モデル
				subject_vertices = np.array(parent_model_mesh_pv.points)
				subject_mesh = parent_model_mesh_pv.copy()
				
				# 距離基準（Target）: 子モデル（領域 or 全体）
				target_mesh = reference_mesh
				
				calc_label = "距離ヒートマップを計算中（親モデル → 子モデル）..."

			print(f"  計算対象頂点数: {len(subject_vertices)} 点")
			print(calc_label)
			
			# プログレスウィンドウを作成
			progress_win = tk.Toplevel(menu_window)
			progress_win.title("計算中...")
			progress_win.geometry("400x120")
			progress_win.resizable(False, False)
			progress_win.attributes('-topmost', True)
			
			# 中央配置
			progress_win.update_idletasks()
			px = (progress_win.winfo_screenwidth() // 2) - (400 // 2)
			py = (progress_win.winfo_screenheight() // 2) - (120 // 2)
			progress_win.geometry(f"+{px}+{py}")
			
			ttk.Label(progress_win, text=calc_label, font=(self.ui_font_family, 10)).pack(pady=(15, 5))
			
			progress_var = tk.DoubleVar(value=0)
			progress_bar = ttk.Progressbar(progress_win, variable=progress_var, maximum=100)
			progress_bar.pack(fill=tk.X, padx=20, pady=5)
			
			status_label = ttk.Label(progress_win, text="0%", font=(self.ui_font_family, 9), foreground="gray")
			status_label.pack(pady=2)
			
			progress_win.update()
			
			def update_progress(val):
				if val > 100:
					status_label.config(text="補間計算中...")
				else:
					progress_var.set(val)
					status_label.config(text=f"{val:.0f}%")
				progress_win.update()
			
			try:
				# 高速化のため、Open3DのPoint-to-Point距離計算を使用
				# 元々は「近傍点のみPyVistaの精密計算(Point-to-Surface)」をループで行っていたが、
				# 点数が多いと非常に時間がかかるため(Pythonループのオーバーヘッド)、
				# ターゲットメッシュの「頂点」に加えて「面の中心点」もサンプリングして密度を上げ、
				# Open3DのKDTree(C++)で一括計算する方式に変更。
				
				status_label.config(text="距離分布を計算中...")
				progress_win.update()
				
				# 計算対象（Subject）
				pcd_s = o3d.geometry.PointCloud()
				pcd_s.points = o3d.utility.Vector3dVector(subject_vertices)
				
				# 距離基準（Target）
				# 頂点だけでなく面の中心点も追加することで、メッシュ表面への距離をより正確に近似する
				tgt_pts = np.array(target_mesh.points)
				try:
					if target_mesh.n_cells > 0:
						status_label.config(text="メッシュ表面点をサンプリング中...")
						progress_win.update()
						centers = target_mesh.cell_centers().points
						tgt_pts = np.vstack([tgt_pts, centers])
				except Exception as e:
					print(f"面中心計算スキップ: {e}")
				
				pcd_t = o3d.geometry.PointCloud()
				pcd_t.points = o3d.utility.Vector3dVector(tgt_pts)
				
				status_label.config(text="最近傍探索を実行中(KDTree)...")
				progress_win.update()
				
				# 全点に対して高速計算
				distances = np.asarray(pcd_s.compute_point_cloud_distance(pcd_t))
				
				progress_var.set(100)
				status_label.config(text="計算完了")
				progress_win.update()
				
				# 3. 遠すぎる点の除外（分布を見やすくするため、閾値以上は丸めるか除外）
				# ここでは生の値を保持するが、可視化時にclimで調整される
				
			finally:
				progress_win.destroy()
			
			# 計算対象メッシュにヒートマップを適用
			heatmap_mesh_to_display = subject_mesh # 既にPolyData
			heatmap_mesh_to_display['distance'] = distances
			
			# キャッシュに保存
			heatmap_data_cache['distances'] = distances
			heatmap_data_cache['points'] = np.array(subject_mesh.points)
			heatmap_data_cache['info'] = calc_label
			
			# エクセル出力ボタンを有効化
			try:
				btn_export_heatmap.config(state="normal", text="ヒートマップ分布をExcelに出力 (データあり)")
			except Exception as e:
				print(f"ボタン有効化エラー: {e}")
			
			max_dist = np.max(distances)
			mean_dist = np.mean(distances)
			min_dist = np.min(distances)
			
			print(f"\n距離統計:")
			print(f"  最小距離: {min_dist:.4f} mm")
			print(f"  平均距離: {mean_dist:.4f} mm")
			print(f"  最大距離: {max_dist:.4f} mm")
			
			# 子モデル全体も準備
			child_vertices_full = np.asarray(child_model_transformed.vertices)
			child_triangles_full = np.asarray(child_model_transformed.triangles)
			faces_pv_full = np.hstack([np.full((len(child_triangles_full), 1), 3), child_triangles_full])
			child_model_pv = pv.PolyData(child_vertices_full, faces_pv_full)
			
			# PyVistaで可視化
			print("\n可視化ウィンドウを起動中...")
			self._show_fitting_result_with_heatmap(
				parent_model_mesh_pv,
				child_model_pv,
				heatmap_mesh_to_display,
				None,  # parent_mesh_o3d (不要)
				child_model_transformed,
				child_to_parent_transform,
				parent_color=viz_colors['parent'],
				child_color=viz_colors['child'],
				cmap='turbo'
			)
		
		btn_visualize_heatmap = ttk.Button(
			button_frame,
			text="ヒートマップありで可視化する（時間かかるよ）",
			command=visualize_heatmap
		)
		btn_visualize_heatmap.pack(fill=tk.X, pady=5)

		# 4.5 ヒートマップデータ出力
		def export_heatmap_data():
			try:
				import pandas as pd
				import numpy as np
				
				distances = heatmap_data_cache['distances']
				points = heatmap_data_cache['points']
				
				if distances is None:
					return
				
				# NaN除去
				valid_mask = ~np.isnan(distances)
				n_total = len(distances)
				distances = distances[valid_mask]
				points = points[valid_mask]
				n_valid = len(distances)
				
				if n_valid < n_total:
					print(f"  Note: NaNを含む {n_total - n_valid} 点を除外しました")

				# 保存ダイアログ
				default_name = "Heatmap_Distribution.xlsx"
				if self.fitting_child_model_path.get():
					default_name = Path(self.fitting_child_model_path.get()).stem + "_Heatmap.xlsx"

				out_path = filedialog.asksaveasfilename(
					title="ヒートマップ分布データの保存",
					defaultextension=".xlsx",
					filetypes=[("Excel files", "*.xlsx")],
					initialfile=default_name
				)
				
				if not out_path:
					return
				
				print(f"Excel保存開始: {out_path}")
				
				# 1. 基本統計量
				stats_df = pd.DataFrame({
					"Item": ["Min (mm)", "Max (mm)", "Mean (mm)", "Std (mm)", "Count"],
					"Value": [
						np.min(distances),
						np.max(distances),
						np.mean(distances),
						np.std(distances),
						len(distances)
					]
				})
				
				# 2. 分布（ヒストグラム） 0.1mm刻み
				# バケット作成
				max_dist = np.max(distances)
				max_val = np.ceil(max_dist * 10) / 10.0 # 0.1単位で切り上げ
				if max_val < 0.1: max_val = 0.1
				
				bins = np.arange(0, max_val + 0.1001, 0.1) # 0, 0.1, 0.2 ...
				
				# pd.cutでビン詰め
				cuts = pd.cut(distances, bins=bins, right=True, include_lowest=True)
				counts = cuts.value_counts().sort_index()
				
				dist_df = pd.DataFrame({
					"Range": counts.index.astype(str),
					"Count": counts.values,
					"Percentage": (counts.values / len(distances) * 100)
				})
				
				# 3. 生データ (座標 + 距離)
				# 100万行制限チェック
				if len(distances) < 1000000:
					raw_df = pd.DataFrame({
						"X": points[:, 0],
						"Y": points[:, 1],
						"Z": points[:, 2],
						"Distance_mm": distances
					})
					write_raw = True
				else:
					print("  データ数が多すぎるためRawDataシートはスキップします")
					write_raw = False
				
				# ExcelWriterで複数シートに書き出し
				with pd.ExcelWriter(out_path) as writer:
					stats_df.to_excel(writer, sheet_name="Summary", index=False)
					dist_df.to_excel(writer, sheet_name="Distribution", index=False)
					if write_raw:
						raw_df.to_excel(writer, sheet_name="RawData", index=False)
					
				messagebox.showinfo("保存完了", f"ヒートマップデータを保存しました:\n{out_path}")
				print("  ✓ 保存完了")
				
			except ImportError:
				messagebox.showerror("エラー", "pandasライブラリが見つかりません。\n'pip install pandas openpyxl'を実行してください。")
			except Exception as e:
				print(f"Excel保存エラー: {e}")
				print(traceback.format_exc())
				messagebox.showerror("エラー", f"Excel保存に失敗しました:\n{e}")

		btn_export_heatmap = ttk.Button(
			button_frame,
			text="ヒートマップ分布をExcelに出力",
			command=export_heatmap_data,
			state="disabled"
		)
		btn_export_heatmap.pack(fill=tk.X, pady=5)
		
		# 区切り線
		separator2 = ttk.Separator(button_frame, orient='horizontal')
		separator2.pack(fill=tk.X, pady=15)
		
		# 5. 閉じるボタン
		def close_menu():
			menu_window.destroy()
			if model_saved[0] or pp_saved[0]:
				messagebox.showinfo("完了", "フィッティング処理が完了しました。")
		
		btn_close = ttk.Button(
			button_frame,
			text="閉じる",
			command=close_menu
		)
		btn_close.pack(fill=tk.X, pady=5)
		
		# ウィンドウを最前面に
		menu_window.transient(self)
		menu_window.grab_set()
		self.wait_window(menu_window)
	
	def _show_fitting_result_with_heatmap(self, parent_mesh_pv, child_mesh_pv, child_heatmap_mesh, parent_mesh_o3d, child_mesh_o3d, transform, parent_color='red', child_color='green', cmap='turbo'):
		"""フィッティング結果を距離ヒートマップ付きで可視化
		
		Args:
			parent_mesh_pv: 親メッシュ（PyVista）
			child_mesh_pv: 子メッシュ全体（PyVista、通常表示用）
			child_heatmap_mesh: ヒートマップ用メッシュ（PyVista、距離データ付き）
			parent_mesh_o3d: 親メッシュ（Open3D）
			child_mesh_o3d: 子メッシュ（Open3D）
			transform: 変換行列
			parent_color: 親モデルの色 (default: 'red')
			child_color: 子モデルの色 (default: 'green')
			cmap: ヒートマップのカラーマップ (default: 'turbo')
		"""
		plotter = pv.Plotter()
		plotter.set_background('white')
		
		# アクター辞書（表示切り替え用）
		actors = {}
		
		# 親モデル（赤）
		if parent_mesh_pv is not None:
			actor_parent = plotter.add_mesh(
				parent_mesh_pv,
				color=parent_color,
				opacity=0.5,
				show_edges=False,
				label='親モデル'
			)
			actors['parent'] = actor_parent
		
		# 子モデル（緑、通常表示用）
		if child_mesh_pv is not None:
			# 通常表示用（緑）
			actor_child = plotter.add_mesh(
				child_mesh_pv,
				color=child_color,
				opacity=0.5,
				show_edges=False,
				label='子モデル'
			)
			actors['child'] = actor_child
		
		# ヒートマップ表示用（別メッシュ）
		if child_heatmap_mesh is not None and 'distance' in child_heatmap_mesh.array_names:
			actor_heatmap = plotter.add_mesh(
				child_heatmap_mesh,
				scalars='distance',
				cmap=cmap,  # カラーマップ
				clim=[0.0, 5.0],  # カラースケール: 0～5mm
				show_edges=False,
				opacity=1.0,
				label='距離ヒートマップ',
				scalar_bar_args={
					'title': '距離 (mm)',
					'vertical': True,
				'title_font_size': 16,
				'label_font_size': 14
			}
		)
		actors['heatmap'] = actor_heatmap
		# 初期状態は非表示に設定
		actor_heatmap.SetVisibility(False)
		def update_parent_opacity(value):
			if 'parent' in actors:
				actors['parent'].GetProperty().SetOpacity(value)
		
		plotter.add_slider_widget(
			update_parent_opacity,
			[0.0, 1.0],
			value=0.5,
			title='親モデル透明度',
			pointa=(0.65, 0.92),
			pointb=(0.95, 0.92),
			style='modern'
		)
		
		# 透明度スライダー（子モデル）
		def update_child_opacity(value):
			if 'child' in actors and actors['child'].GetVisibility():
				actors['child'].GetProperty().SetOpacity(value)
		
		plotter.add_slider_widget(
			update_child_opacity,
			[0.0, 1.0],
			value=0.5,
			title='子モデル透明度',
			pointa=(0.65, 0.87),
			pointb=(0.95, 0.87),
			style='modern'
		)
		
		# チェックボックス（親モデル表示切り替え）
		def toggle_parent(state):
			if 'parent' in actors:
				actors['parent'].SetVisibility(state)
		
		plotter.add_checkbox_button_widget(
			toggle_parent,
			value=True,
			position=(10, 150),
			size=30,
			border_size=2,
			color_on='red',
			color_off='grey',
			background_color='white'
		)
		
		# ラベル（親モデル）
		plotter.add_text('親モデル', position=(45, 155), font_size=10, color='black')
		
		# チェックボックス（子モデル表示切り替え）
		def toggle_child(state):
			if 'child' in actors:
				actors['child'].SetVisibility(state)
		
		plotter.add_checkbox_button_widget(
			toggle_child,
			value=True,
			position=(10, 110),
			size=30,
			border_size=2,
			color_on='green',
			color_off='grey',
			background_color='white'
		)
		
		# ラベル（子モデル）
		plotter.add_text('子モデル', position=(45, 115), font_size=10, color='black')
		
		# チェックボックス（ヒートマップ表示切り替え）
		def toggle_heatmap(state):
			if 'heatmap' in actors:
				actors['heatmap'].SetVisibility(state)
		
		plotter.add_checkbox_button_widget(
			toggle_heatmap,
			value=False,
			position=(10, 70),
			size=30,
			border_size=2,
			color_on='orange',
			color_off='grey',
			background_color='white'
		)
		
		# ラベル（ヒートマップ）
		plotter.add_text('距離ヒートマップ', position=(45, 75), font_size=10, color='black')
		
		# ヘルプテキスト
		help_text = (
			'フィッティング結果\n'
			'チェックボックス: 表示切り替え\n'
			'スライダー: 透明度調整\n'
			'距離ヒートマップ: 青(近い)→赤(遠い)'
		)
		plotter.add_text(
			help_text,
			position=(10, 10),
			font_size=9,
			color='black',
			font='arial'
		)
		
		print("\n[最終結果] 可視化ウィンドウを表示中...")
		print("  - 親モデル(赤)と子モデル(緑)の透明度を調整できます")
		print("  - チェックボックスで表示/非表示を切り替えられます")
		print("  - 距離ヒートマップで位置合わせ精度を確認できます")
		
		plotter.show()
	
	def _compute_distance_heatmap_fast(self, vertices: np.ndarray, parent_mesh: pv.PolyData, max_sample_points: int = 10000, progress_callback=None) -> np.ndarray:
		"""高速な距離ヒートマップ計算（ダウンサンプリング + 補間）
		
		Args:
			vertices: 子メッシュの頂点座標 (N, 3)
			parent_mesh: 親メッシュ（PyVista）
			max_sample_points: 最大サンプル点数
			progress_callback: 進捗更新用コールバック関数(0.0~100.0)
		
		Returns:
			距離配列 (N,)
		"""
		n_vertices = len(vertices)
		print(f"  元の頂点数: {n_vertices}")
		
		# サンプリング比率を計算
		if n_vertices > max_sample_points:
			sample_rate = max_sample_points / n_vertices
			sample_indices = np.random.choice(n_vertices, max_sample_points, replace=False)
			sample_vertices = vertices[sample_indices]
			print(f"  サンプリング: {max_sample_points} 点 ({sample_rate*100:.1f}%)")
		else:
			sample_indices = np.arange(n_vertices)
			sample_vertices = vertices
			print(f"  サンプリング不要")
		
		# 距離計算（逐次処理）
		print(f"  距離計算中...")
		sample_distances = np.zeros(len(sample_vertices))
		
		progress_step = max(1, len(sample_vertices) // 20)  # 5%ごとに進捗表示
		for i, vertex in enumerate(sample_vertices):
			_, closest_point = parent_mesh.find_closest_cell(vertex, return_closest_point=True)
			sample_distances[i] = np.linalg.norm(vertex - closest_point)
			
			if (i + 1) % progress_step == 0:
				progress = (i + 1) / len(sample_vertices) * 100
				print(f"    進行状況: {progress:.0f}%")
				if progress_callback:
					progress_callback(progress)
		
		# 100%通知
		if progress_callback:
			progress_callback(100.0)
		
		# 全頂点への補間（サンプリングした場合）
		if n_vertices > max_sample_points:
			print(f"  補間中...")
			if progress_callback:
				progress_callback(101.0) # 100%超えで補間中表現
				
			try:
				from scipy.interpolate import NearestNDInterpolator
				interpolator = NearestNDInterpolator(sample_vertices, sample_distances)
				distances = interpolator(vertices)
			except ImportError:
				print("  警告: scipyが見つかりません。サンプル点のみ使用します")
				# scipyがない場合は、最近傍を単純に探す
				distances = np.zeros(n_vertices)
				for i, vertex in enumerate(vertices):
					dists = np.linalg.norm(sample_vertices - vertex, axis=1)
					nearest_idx = np.argmin(dists)
					distances[i] = sample_distances[nearest_idx]
		else:
			distances = sample_distances
		
		print(f"  ✓ 距離計算完了")
		return distances
	
	def _voxel_downsample(self, points: np.ndarray, voxel_size: float) -> np.ndarray:
		"""ボクセルグリッドによるダウンサンプリング
		
		Args:
			points: 点群 (N, 3)
			voxel_size: ボクセルサイズ (mm)
			
		Returns:
			ダウンサンプリングされた点群 (M, 3)
		"""
		# 各点をボクセルグリッドにマッピング
		voxel_indices = np.floor(points / voxel_size).astype(np.int32)
		
		# 重複を除去（各ボクセルに1点のみ残す）
		unique_voxels, inverse_indices = np.unique(voxel_indices, axis=0, return_inverse=True)
		
		# 各ボクセルの代表点を計算（平均）
		downsampled = np.zeros((len(unique_voxels), 3))
		for i in range(len(unique_voxels)):
			mask = inverse_indices == i
			downsampled[i] = np.mean(points[mask], axis=0)
		
		return downsampled
	
	def _ransac_registration(self, source: np.ndarray, target: np.ndarray) -> np.ndarray:
		"""RANSACによる位置合わせ（対応点ベース）
		
		Args:
			source: 移動させる点群 (N, 3)
			target: ターゲット点群 (M, 3)
			
		Returns:
			4x4 変換行列、失敗時はNone
		"""
		from scipy.spatial import cKDTree
		
		distance_threshold = self.ransac_distance_threshold.get()
		max_iterations = self.ransac_max_iterations.get()
		confidence = self.ransac_confidence.get()
		
		print(f"RANSACパラメータ: 距離閾値={distance_threshold}mm, 最大反復={max_iterations}, 信頼度={confidence}")
		
		# ステップ1: 最近傍対応点を事前計算（KDTreeを使用）
		print("  [1/3] 最近傍対応点を計算中...")
		target_tree = cKDTree(target)
		distances, correspondences = target_tree.query(source)
		
		# 対応点ペアを作成（距離が近い順にソート）
		valid_pairs = []
		for i, (dist, target_idx) in enumerate(zip(distances, correspondences)):
			if dist < distance_threshold * 5:  # 初期対応は緩い閾値
				valid_pairs.append((i, target_idx, dist))
		
		valid_pairs.sort(key=lambda x: x[2])  # 距離でソート
		
		if len(valid_pairs) < 100:
			print(f"  警告: 初期対応点が少なすぎます（{len(valid_pairs)}組）")
			print(f"  ヒント: ボクセルサイズを小さくするか、2つのモデルを手動で近づけてください")
			return None
		
		print(f"  対応点候補: {len(valid_pairs)}組")
		
		# ステップ2: RANSACで最良の変換を探索
		print(f"  [2/3] RANSAC実行中...")
		best_transform = None
		best_inliers = 0
		estimated_inlier_ratio = 0.1
		
		for iteration in range(max_iterations):
			# 対応点ペアからランダムに3組選択
			if len(valid_pairs) < 3:
				return None
			
			sample_pairs = np.random.choice(len(valid_pairs), min(3, len(valid_pairs)), replace=False)
			
			source_indices = [valid_pairs[idx][0] for idx in sample_pairs]
			target_indices = [valid_pairs[idx][1] for idx in sample_pairs]
			
			source_sample = source[source_indices]
			target_sample = target[target_indices]
			
			# 3点対応から変換行列を推定
			transform = self._estimate_transform_from_points(source_sample, target_sample)
			
			if transform is None:
				continue
			
			# すべての点に変換を適用
			source_transformed = self._apply_transform_to_points(source, transform)
			
			# インライア数をカウント（厳密な閾値で）
			distances, _ = target_tree.query(source_transformed)
			inliers = np.sum(distances < distance_threshold)
			
			if inliers > best_inliers:
				best_inliers = inliers
				best_transform = transform
				current_ratio = inliers / len(source)
				
				if iteration % 100 == 0 or iteration < 10:
					print(f"    反復 {iteration}: インライア数={inliers}/{len(source)} ({100*current_ratio:.1f}%)")
				
				# 適応的な早期終了
				if current_ratio > estimated_inlier_ratio:
					estimated_inlier_ratio = current_ratio
					
					if estimated_inlier_ratio > 0.01:
						w_cubed = estimated_inlier_ratio ** 3
						if w_cubed < 0.9999:
							required_iterations = int(np.log(1 - confidence) / np.log(1 - w_cubed))
							
							if iteration >= required_iterations:
								print(f"    信頼度 {confidence} を満たす反復回数を達成")
								break
		
		# ステップ3: 結果の評価
		print(f"  [3/3] 結果評価...")
		if best_transform is not None:
			final_ratio = best_inliers / len(source)
			print(f"RANSAC完了: インライア数={best_inliers}/{len(source)} ({100*final_ratio:.1f}%)")
			
			if final_ratio < 0.3:
				print(f"  警告: インライア率が低いです（{100*final_ratio:.1f}%）")
				print(f"  ヒント: パラメータ調整が必要かもしれません")
				# それでも結果を返す（ICPで改善できる可能性がある）
		else:
			print("RANSACに失敗しました")
		
		return best_transform
	
	def _estimate_transform_from_points(self, source: np.ndarray, target: np.ndarray) -> np.ndarray:
		"""3点対応から変換行列を推定（SVDベース）
		
		Args:
			source: ソース点群 (3, 3)
			target: ターゲット点群 (3, 3)
			
		Returns:
			4x4 変換行列、失敗時はNone
		"""
		try:
			# 重心を計算
			source_center = np.mean(source, axis=0)
			target_center = np.mean(target, axis=0)
			
			# 中心化
			source_centered = source - source_center
			target_centered = target - target_center
			
			# 共分散行列
			H = source_centered.T @ target_centered
			
			# SVD
			U, S, Vt = np.linalg.svd(H)
			R = Vt.T @ U.T
			
			# 反射を防ぐ
			if np.linalg.det(R) < 0:
				Vt[-1, :] *= -1
				R = Vt.T @ U.T
			
			# 平行移動ベクトル
			t = target_center - R @ source_center
			
			# 4x4変換行列
			transform = np.eye(4)
			transform[:3, :3] = R
			transform[:3, 3] = t
			
			return transform
			
		except Exception:
			return None
	
	def _apply_transform_to_points(self, points: np.ndarray, transform: np.ndarray) -> np.ndarray:
		"""点群に変換行列を適用
		
		Args:
			points: 点群 (N, 3)
			transform: 4x4 変換行列
			
		Returns:
			変換後の点群 (N, 3)
		"""
		ones = np.ones((len(points), 1))
		points_homogeneous = np.hstack([points, ones])
		points_transformed = (transform @ points_homogeneous.T).T[:, :3]
		return points_transformed

	def on_visualize_fitting_parent(self) -> None:
		"""親モデルと親任意領域を同時に可視化"""
		parent_model_path = self.fitting_parent_model_path.get()
		parent_region_path = self.fitting_parent_region_path.get()
		
		if not parent_model_path and not parent_region_path:
			messagebox.showerror("エラー", "親モデルまたは親任意領域ファイルを選択してください。")
			return
		
		try:
			# ディスプレイサイズの取得
			screen_width = self.winfo_screenwidth()
			screen_height = self.winfo_screenheight()
			window_width = int(screen_width * 0.9)
			window_height = int(screen_height * 0.9)
			
			plotter = pv.Plotter(title="親モデルの可視化", window_size=(window_width, window_height))
			plotter.set_background("white")
			
			model_actor = None
			region_actor = None
			
			# 親モデル（若干透明）
			if parent_model_path:
				parent_model = pv.read(parent_model_path)
				model_actor = plotter.add_mesh(
					parent_model, 
					color='lightcoral', 
					opacity=0.4,
					smooth_shading=True,
					show_edges=False, 
					label='親モデル'
				)
			
			# 親任意領域（不透明）
			if parent_region_path:
				parent_region = pv.read(parent_region_path)
				region_actor = plotter.add_mesh(
					parent_region, 
					color='red', 
					opacity=1.0,
					smooth_shading=True,
					show_edges=False, 
					label='親任意領域'
				)
			
			# チェックボックスとスライダー
			def toggle_model(state):
				if model_actor:
					model_actor.SetVisibility(state)
			
			def toggle_region(state):
				if region_actor:
					region_actor.SetVisibility(state)
			
			def update_model_opacity(value):
				if model_actor:
					model_actor.GetProperty().SetOpacity(value)
			
			def update_region_opacity(value):
				if region_actor:
					region_actor.GetProperty().SetOpacity(value)
			
			# UI配置
			checkbox_y = 10
			if model_actor:
				plotter.add_text("親モデル", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_model, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="lightcoral", color_off="white")
				checkbox_y += 30
			
			if region_actor:
				plotter.add_text("親任意領域", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_region, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="red", color_off="white")
			
			# 透明度スライダー
			if model_actor:
				plotter.add_slider_widget(
					update_model_opacity,
					rng=[0.0, 1.0],
					value=0.4,
					title="親モデル ",
					pointa=(0.05, 0.92),
					pointb=(0.15, 0.92),
					style='modern',
					color="lightcoral"
				)
			
			if region_actor:
				plotter.add_slider_widget(
					update_region_opacity,
					rng=[0.0, 1.0],
					value=1.0,
					title="親任意領域 ",
					pointa=(0.05, 0.85),
					pointb=(0.15, 0.85),
					style='modern',
					color="red"
				)
			
			plotter.add_legend()
			plotter.show()
			
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_visualize_fitting_child(self) -> None:
		"""子モデル、子任意領域、子特徴点を同時に可視化"""
		child_model_path = self.fitting_child_model_path.get()
		child_region_path = self.fitting_child_region_path.get()
		child_pp_path = self.fitting_child_pp_path.get()
		
		if not child_model_path and not child_region_path and not child_pp_path:
			messagebox.showerror("エラー", "子モデル、子任意領域、または子特徴点ファイルを選択してください。")
			return
		
		try:
			# ディスプレイサイズの取得
			screen_width = self.winfo_screenwidth()
			screen_height = self.winfo_screenheight()
			window_width = int(screen_width * 0.9)
			window_height = int(screen_height * 0.9)
			
			plotter = pv.Plotter(title="子モデルの可視化", window_size=(window_width, window_height))
			plotter.set_background("white")
			
			model_actor = None
			region_actor = None
			points_actor = None
			
			# 子モデル（若干透明）
			if child_model_path:
				child_model = pv.read(child_model_path)
				model_actor = plotter.add_mesh(
					child_model, 
					color='lightblue', 
					opacity=0.4,
					smooth_shading=True,
					show_edges=False, 
					label='子モデル'
				)
			
			# 子任意領域（不透明）
			if child_region_path:
				child_region = pv.read(child_region_path)
				region_actor = plotter.add_mesh(
					child_region, 
					color='blue', 
					opacity=1.0,
					smooth_shading=True,
					show_edges=False, 
					label='子任意領域'
				)
			
			# 子特徴点（PP）- すべての点を表示（LMNも含む）
			if child_pp_path:
				points, labels = self._parse_pp_file(child_pp_path)
				
				if len(points) > 0:
					# 特徴点を表示
					points_actor = plotter.add_points(
						points,
						color='green',
						point_size=15,
						render_points_as_spheres=True,
						label='子特徴点'
					)
					
					# ラベル表示
					for point, label in zip(points, labels):
						plotter.add_point_labels(
							[point],
							[label],
							font_size=12,
							text_color='green',
							point_size=0,
							always_visible=True,
							shape_opacity=0.5,
							bold=True
						)
			
			# チェックボックスとスライダー
			def toggle_model(state):
				if model_actor:
					model_actor.SetVisibility(state)
			
			def toggle_region(state):
				if region_actor:
					region_actor.SetVisibility(state)
			
			def update_model_opacity(value):
				if model_actor:
					model_actor.GetProperty().SetOpacity(value)
			
			def update_region_opacity(value):
				if region_actor:
					region_actor.GetProperty().SetOpacity(value)
			
			# UI配置
			checkbox_y = 10
			if model_actor:
				plotter.add_text("子モデル", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_model, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="lightblue", color_off="white")
				checkbox_y += 30
			
			if region_actor:
				plotter.add_text("子任意領域", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_region, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="blue", color_off="white")
			
			# 透明度スライダー
			if model_actor:
				plotter.add_slider_widget(
					update_model_opacity,
					rng=[0.0, 1.0],
					value=0.4,
					title="子モデル ",
					pointa=(0.05, 0.92),
					pointb=(0.15, 0.92),
					style='modern',
					color="lightblue"
				)
			
			if region_actor:
				plotter.add_slider_widget(
					update_region_opacity,
					rng=[0.0, 1.0],
					value=1.0,
					title="子任意領域 ",
					pointa=(0.05, 0.85),
					pointb=(0.15, 0.85),
					style='modern',
					color="blue"
				)
			
			plotter.add_legend()
			plotter.show()
			
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_visualize_org(self) -> None:
		"""ORG: RobotScan モデルと座標ポイントを可視化"""
		model_path = self.org_model_path.get().strip()
		pp_path = self.org_pp_path.get().strip()
		
		if not model_path or not pp_path:
			messagebox.showerror("エラー", "モデルファイルと座標ポイントファイルを選択してください。")
			return
		
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista"
			)
			return
		
		# PP読み込み
		try:
			points, labels = self._parse_pp_file(pp_path)
		except Exception as e:
			messagebox.showerror("エラー", f"PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# CL点と座標系の計算
		# PPファイルの順序: L, M, N
		C_cl_origin = None
		C_cl_x_axis = None
		C_cl_y_axis = None
		C_cl_z_axis = None
		C_o_origin = None
		C_o_x_axis = None
		C_o_y_axis = None
		C_o_z_axis = None
		L_prime_point = None
		
		if len(points) >= 3:
			L_point = points[0]  # L点
			M_point = points[1]  # M点
			N_point = points[2]  # N点
			
			# MとNの中点を計算（CL点）
			CL_point = np.array([
				(M_point[0] + N_point[0]) / 2.0,
				(M_point[1] + N_point[1]) / 2.0,
				(M_point[2] + N_point[2]) / 2.0
			])
			
			# LMベクトルを作成
			LM_vector = M_point - L_point
			LM_unit = LM_vector / np.linalg.norm(LM_vector)
			
			# NからLM直線への垂線の足L'を計算
			# L'はLM上の点で、NL'がLMに垂直
			LN_vector = N_point - L_point
			projection_length = np.dot(LN_vector, LM_unit)
			L_prime_point = L_point + projection_length * LM_unit
			
			# C_cl座標系を構築
			# 原点: CL
			C_cl_origin = CL_point
			
			# X軸: M → L'
			x_vec = L_prime_point - M_point
			C_cl_x_axis = x_vec / np.linalg.norm(x_vec)
			
			# Z軸: N → L'
			z_vec = L_prime_point - N_point
			C_cl_z_axis = z_vec / np.linalg.norm(z_vec)
			
			# Y軸: Z × X
			C_cl_y_axis = np.cross(C_cl_z_axis, C_cl_x_axis)
			C_cl_y_axis = C_cl_y_axis / np.linalg.norm(C_cl_y_axis)
			
			# C_o座標系: C_clをY軸方向に250mm平行移動
			C_o_origin = C_cl_origin + -250.0 * C_cl_y_axis
			C_o_x_axis = C_cl_x_axis
			C_o_y_axis = C_cl_y_axis
			C_o_z_axis = C_cl_z_axis
			
			# CL点とL'点を追加
			points = np.vstack([points, CL_point, L_prime_point])
			labels = labels + ['CL', "L'"]
		
		# メッシュ読み込み
		try:
			mesh = pv.read(model_path)
		except Exception as e:
			messagebox.showerror("エラー", f"モデルの読み込みに失敗しました:\n{e}")
			return
		
		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)
		
		# プロッタ作成
		plotter = pv.Plotter(title="ORG - RobotScan データ可視化", window_size=(window_width, window_height))
		plotter.set_background("white")
		
		# メッシュ表示
		mesh_actor = plotter.add_mesh(
			mesh,
			color='lightgray',
			smooth_shading=True,
			show_edges=False,
			opacity=0.8,
			label='RobotScan モデル'
		)
		
		# 座標ポイント表示
		points_actor = plotter.add_points(
			points,
			color='red',
			point_size=15,
			render_points_as_spheres=True,
			label='座標ポイント'
		)
		
		# ラベル表示（各点の名前）
		for i, (point, label) in enumerate(zip(points, labels)):
			plotter.add_point_labels(
				[point],
				[label],
				font_size=12,
				text_color='black',
				point_size=0,
				always_visible=True,
				shape_opacity=0.5
			)
		
		# C_cl座標系の表示
		c_cl_actors = []
		if C_cl_origin is not None:
			axis_length = 50.0  # 座標軸の長さ
			
			# X軸（赤）
			x_line = pv.Line(C_cl_origin, C_cl_origin + axis_length * C_cl_x_axis)
			c_cl_actors.append(plotter.add_mesh(x_line, color='red', line_width=5, label='C_cl X'))
			
			# Y軸（緑）
			y_line = pv.Line(C_cl_origin, C_cl_origin + axis_length * C_cl_y_axis)
			c_cl_actors.append(plotter.add_mesh(y_line, color='green', line_width=5, label='C_cl Y'))
			
			# Z軸（青）
			z_line = pv.Line(C_cl_origin, C_cl_origin + axis_length * C_cl_z_axis)
			c_cl_actors.append(plotter.add_mesh(z_line, color='blue', line_width=5, label='C_cl Z'))
			
			# 原点（CL）に小さな球
			c_cl_actors.append(plotter.add_mesh(
				pv.Sphere(radius=3.0, center=C_cl_origin),
				color='yellow',
				label='C_cl原点'
			))
		
		# C_o座標系の表示
		c_o_actors = []
		if C_o_origin is not None:
			axis_length = 50.0
			
			# X軸（濃い赤）
			x_line = pv.Line(C_o_origin, C_o_origin + axis_length * C_o_x_axis)
			c_o_actors.append(plotter.add_mesh(x_line, color='darkred', line_width=5, label='C_o X'))
			
			# Y軸（濃い緑）
			y_line = pv.Line(C_o_origin, C_o_origin + axis_length * C_o_y_axis)
			c_o_actors.append(plotter.add_mesh(y_line, color='darkgreen', line_width=5, label='C_o Y'))
			
			# Z軸（濃い青）
			z_line = pv.Line(C_o_origin, C_o_origin + axis_length * C_o_z_axis)
			c_o_actors.append(plotter.add_mesh(z_line, color='darkblue', line_width=5, label='C_o Z'))
			
			# 原点に小さな球
			c_o_actors.append(plotter.add_mesh(
				pv.Sphere(radius=3.0, center=C_o_origin),
				color='orange',
				label='C_o原点'
			))
		
		# ORG座標をC_o座標系で表示
		org_actor = None
		if C_o_origin is not None and C_o_x_axis is not None:
			try:
				# ORG座標を取得（C_o座標系での座標）
				org_x = self.org_coord_x.get()
				org_y = self.org_coord_y.get()
				org_z = self.org_coord_z.get()
				
				# C_o座標系からワールド座標系に変換
				# ORG_world = C_o_origin + org_x * C_o_x_axis + org_y * C_o_y_axis + org_z * C_o_z_axis
				org_world = C_o_origin + org_x * C_o_x_axis + org_y * C_o_y_axis + org_z * C_o_z_axis
				
				# ORG点を表示（紫の大きな球）
				org_actor = plotter.add_mesh(
					pv.Sphere(radius=5.0, center=org_world),
					color='purple',
					label='ORG'
				)
				
				# ORGラベル
				plotter.add_point_labels(
					[org_world],
					['ORG'],
					font_size=14,
					text_color='purple',
					point_size=0,
					always_visible=True,
					shape_opacity=0.7,
					bold=True
				)
			except (ValueError, TypeError):
				pass  # 座標入力が不正な場合は無視
		
		# チェックボックス
		def toggle_mesh(state):
			mesh_actor.SetVisibility(state)
		
		def toggle_points(state):
			points_actor.SetVisibility(state)
		
		def toggle_c_cl(state):
			for actor in c_cl_actors:
				actor.SetVisibility(state)
		
		def toggle_c_o(state):
			for actor in c_o_actors:
				actor.SetVisibility(state)
		
		def update_mesh_opacity(value):
			mesh_actor.GetProperty().SetOpacity(value)
		
		plotter.add_text("Model", position=(10, 10), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_mesh, value=True, position=(100, 12), size=20, border_size=1, color_on="lightgray", color_off="white")
		
		plotter.add_text("Points", position=(10, 40), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_points, value=True, position=(100, 42), size=20, border_size=1, color_on="red", color_off="white")
		
		if c_cl_actors:
			plotter.add_text("C_cl", position=(10, 70), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_c_cl, value=True, position=(100, 72), size=20, border_size=1, color_on="yellow", color_off="white")
		
		if c_o_actors:
			plotter.add_text("C_o", position=(10, 100), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_c_o, value=True, position=(100, 102), size=20, border_size=1, color_on="orange", color_off="white")
		
		if org_actor:
			def toggle_org(state):
				org_actor.SetVisibility(state)
			
			plotter.add_text("ORG", position=(10, 130), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_org, value=True, position=(100, 132), size=20, border_size=1, color_on="purple", color_off="white")
		
		# 透明度スライダー
		plotter.add_slider_widget(
			update_mesh_opacity,
			rng=[0.0, 1.0],
			value=0.8,
			title="Model ",
			pointa=(0.05, 0.02),
			pointb=(0.15, 0.02),
			style='modern',
			color="lightgray"
		)
		
		plotter.add_legend()
		plotter.show()
		
		# ウィンドウを閉じた後、ORG座標の保存を確認
		if C_o_origin is not None and org_actor is not None:
			try:
				# ORG座標を取得
				org_x = self.org_coord_x.get()
				org_y = self.org_coord_y.get()
				org_z = self.org_coord_z.get()
				
				# C_o座標系からワールド座標系に変換
				org_world = C_o_origin + org_x * C_o_x_axis + org_y * C_o_y_axis + org_z * C_o_z_axis
				
				# 保存確認ダイアログ
				result = messagebox.askyesno(
					"ORG座標の保存",
					f"ORG座標をPPファイルに保存しますか？\n\n"
					f"C_o座標系: ({org_x:.2f}, {org_y:.2f}, {org_z:.2f})\n"
					f"ワールド座標: ({org_world[0]:.2f}, {org_world[1]:.2f}, {org_world[2]:.2f})"
				)
				
				if result:
					self._save_org_to_pp(pp_path, org_world)
			except (ValueError, TypeError):
				pass  # 座標が不正な場合は何もしない

	def _save_org_to_pp(self, pp_path: str, org_world: np.ndarray) -> None:
		"""ORG座標をPPファイルに追加保存（混在形式対応：XML + テキスト行を保持）"""
		try:
			import xml.etree.ElementTree as ET
			from datetime import datetime
			
			# ファイル全体を読み込み、XML部分とテキスト行部分を分離
			with open(pp_path, 'r', encoding='utf-8') as f:
				content = f.read()
			
			# XML終了タグを探す
			xml_end = content.find('</PickedPoints>')
			if xml_end == -1:
				raise ValueError("PickedPoints終了タグが見つかりません")
			
			xml_end += len('</PickedPoints>')
			xml_part = content[:xml_end]
			text_part = content[xml_end:].strip()
			
			# XML部分をパース
			root = ET.fromstring(xml_part)
			
			# O点が既に存在するか確認
			existing_o = None
			for point in root.findall('point'):
				if point.get('name') == 'O':
					existing_o = point
					break
			
			if existing_o is not None:
				# 既存のO点を更新
				existing_o.set('x', f"{org_world[0]:.5f}")
				existing_o.set('y', f"{org_world[1]:.5f}")
				existing_o.set('z', f"{org_world[2]:.5f}")
			else:
				# 新しいO点を追加
				new_point = ET.Element('point')
				new_point.set('x', f"{org_world[0]:.5f}")
				new_point.set('y', f"{org_world[1]:.5f}")
				new_point.set('z', f"{org_world[2]:.5f}")
				new_point.set('active', '1')
				new_point.set('name', 'O')
				root.append(new_point)
			
			# DocumentDataのDateTime要素を更新
			doc_data = root.find('DocumentData')
			if doc_data is not None:
				date_time = doc_data.find('DateTime')
				if date_time is not None:
					now = datetime.now()
					date_time.set('date', now.strftime('%Y-%m-%d'))
					date_time.set('time', now.strftime('%H:%M:%S'))
			
			# XML部分を文字列に変換
			tree = ET.ElementTree(root)
			xml_str = io.BytesIO()
			tree.write(xml_str, encoding='utf-8', xml_declaration=True)
			xml_output = xml_str.getvalue().decode('utf-8')
			
			# テキスト行部分を結合して書き戻し
			final_content = xml_output
			if text_part:
				final_content += '\n' + text_part + '\n'
			
			with open(pp_path, 'w', encoding='utf-8') as f:
				f.write(final_content)
			
			messagebox.showinfo(
				"保存完了",
				f"ORG座標をO点として保存しました:\n{pp_path}"
			)
		except Exception as e:
			messagebox.showerror("保存エラー", f"PPファイルへの保存に失敗しました:\n{e}")

	def on_visualize_a_b(self) -> None:
		"""A-B: 2つのモデルを同時に可視化"""
		model_a_path = self.a_b_model_a_path.get().strip()
		model_b_path = self.a_b_model_b_path.get().strip()

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path)
			mesh_b = pv.read(model_b_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="A-B Visualization")
			plotter.set_background("white")
			plotter.add_mesh(mesh_a, color="lightgray", opacity=0.6, show_edges=False, label="Model A")
			plotter.add_mesh(mesh_b, color="tomato", opacity=0.6, show_edges=False, label="Model B")
			plotter.add_legend()
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_execute_a_b_heatmap(self) -> None:
		"""A-B: AからBメッシュまでの距離をヒートマップ表示し、点群をExcel出力"""
		model_a_path = self.a_b_model_a_path.get().strip()
		model_b_path = self.a_b_model_b_path.get().strip()

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path).triangulate().clean()
			mesh_b = pv.read(model_b_path).triangulate().clean()
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		if mesh_a.n_points == 0 or mesh_b.n_points == 0:
			messagebox.showerror("エラー", "メッシュの頂点数が0です。ファイルを確認してください。")
			return

		try:
			max_points = int(self.sample_points.get()) if hasattr(self, "sample_points") else 10000
			distances = self._compute_distance_heatmap_fast(mesh_a.points, mesh_b, max_sample_points=max_points)
		except Exception as e:
			messagebox.showerror("計算失敗", f"距離計算に失敗しました:\n{e}")
			return

		try:
			heatmap_mesh = mesh_a.copy(deep=False)
			heatmap_mesh['distance'] = np.array(distances, dtype=float)
		except Exception as e:
			messagebox.showerror("エラー", f"ヒートマップ生成に失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="A-B Distance Heatmap")
			plotter.set_background("white")
			min_d = float(np.min(distances))
			max_d = float(np.max(distances))
			plotter.add_mesh(
				heatmap_mesh,
				scalars="distance",
				cmap="turbo",
				clim=[min_d, max_d],
				show_edges=False,
				scalar_bar_args={
					"title": "Distance (mm)",
					"n_labels": 5,
					"fmt": "%.2f",
				},
			)
			plotter.add_mesh(mesh_b, color="lightgray", opacity=0.35, show_edges=False, label="Model B")
			plotter.add_legend()
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")
			return

		# Excel出力
		if not messagebox.askyesno("Excel出力", "点群情報をExcelで保存しますか？"):
			return

		save_path = filedialog.asksaveasfilename(
			title="点群情報を保存",
			defaultextension=".xlsx",
			filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
		)
		if not save_path:
			return

		try:
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "A_to_B_Distance"
			ws.append(["index", "x", "y", "z", "distance_mm"])
			for i, (pt, d) in enumerate(zip(mesh_a.points, distances)):
				ws.append([i, float(pt[0]), float(pt[1]), float(pt[2]), float(d)])
			wb.save(save_path)
			messagebox.showinfo("保存完了", f"保存しました:\n{save_path}")
		except Exception as e:
			messagebox.showerror("保存失敗", f"Excel保存に失敗しました:\n{e}")

	def on_visualize_heatmap_models(self) -> None:
		"""HeatMap: 2つのモデルを同時に可視化"""
		model_a_path = self.heatmap_model_a_path.get().strip()
		model_b_path = self.heatmap_model_b_path.get().strip()
		show_model_a = bool(self.heatmap_show_model_a.get())
		show_model_b = bool(self.heatmap_show_model_b.get())

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		if not (show_model_a or show_model_b):
			messagebox.showwarning("表示設定", "表示するモデルを1つ以上選択してください。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path)
			mesh_b = pv.read(model_b_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="HeatMap Models Visualization")
			plotter.set_background("white")
			if show_model_a:
				plotter.add_mesh(mesh_a, color="lightgray", opacity=0.6, show_edges=False, label="Model A (reference)")
			if show_model_b:
				plotter.add_mesh(mesh_b, color="tomato", opacity=0.6, show_edges=False, label="Model B (target)")
			plotter.add_legend()
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_execute_heatmap_b_to_a(self) -> None:
		"""HeatMap: BからA表面までの距離をヒートマップ表示（クリックで距離表示）"""
		model_a_path = self.heatmap_model_a_path.get().strip()
		model_b_path = self.heatmap_model_b_path.get().strip()
		show_model_a = bool(self.heatmap_show_model_a.get())
		show_model_b = bool(self.heatmap_show_model_b.get())

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		if not show_model_b:
			messagebox.showwarning("表示設定", "ヒートマップ表示にはモデルBの表示が必要です。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path).triangulate().clean()
			mesh_b = pv.read(model_b_path).triangulate().clean()
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		if mesh_a.n_points == 0 or mesh_b.n_points == 0:
			messagebox.showerror("エラー", "メッシュの頂点数が0です。ファイルを確認してください。")
			return

		try:
			max_points = int(self.sample_points.get()) if hasattr(self, "sample_points") else 10000
			distances = self._compute_distance_heatmap_fast(mesh_b.points, mesh_a, max_sample_points=max_points)
		except Exception as e:
			messagebox.showerror("計算失敗", f"距離計算に失敗しました:\n{e}")
			return

		try:
			heatmap_mesh = mesh_b.copy(deep=False)
			heatmap_mesh["distance"] = np.array(distances, dtype=float)
		except Exception as e:
			messagebox.showerror("エラー", f"ヒートマップ生成に失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="HeatMap (B→A Distance)")
			plotter.set_background("white")
			min_d = float(np.min(distances))
			max_d = float(np.max(distances))
			plotter.add_mesh(
				heatmap_mesh,
				scalars="distance",
				cmap="turbo",
				clim=[min_d, max_d],
				show_edges=False,
				label="Model B (heatmap)",
				scalar_bar_args={
					"title": "Distance B→A (mm)",
					"n_labels": 5,
					"fmt": "%.2f",
				},
			)
			if show_model_a:
				plotter.add_mesh(mesh_a, color="lightgray", opacity=0.35, show_edges=False, label="Model A (reference)", pickable=False)
			plotter.add_legend()
			plotter.add_text("Click a point to see distance (mm)", position=(10, 10), font_size=9, color="black", name="pick_hint")
			pick_text_actor = plotter.add_text("Picked distance: --", position=(10, 28), font_size=10, color="black", name="pick_value")

			def _on_pick(*args):
				point = None
				if len(args) >= 1:
					candidate = args[0]
					if hasattr(candidate, "GetPickPosition"):
						try:
							pos = candidate.GetPickPosition()
							point = np.array([pos[0], pos[1], pos[2]], dtype=float)
						except Exception:
							point = None
					elif isinstance(candidate, (list, tuple, np.ndarray)) and len(candidate) >= 3:
						point = np.array(candidate[:3], dtype=float)
					elif isinstance(candidate, pv.PolyData):
						if candidate.n_points > 0:
							point = candidate.points[0]
				if point is None:
					return
				try:
					idx = heatmap_mesh.find_closest_point(point)
					d = float(heatmap_mesh["distance"][idx])
					pick_text_actor.SetInput(f"Picked distance: {d:.3f} mm")
					plotter.render()
				except Exception:
					pass

			plotter.enable_point_picking(callback=_on_pick, show_point=True, use_picker=True, show_message=False)
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")
			return

	def on_c_fix_visualize(self) -> None:
		"""C. Fixタブでの可視化：近位・遠位モデルと全特徴点を表示"""
		# ファイルパスを取得
		prox_model_path = self.prox_model_path.get().strip()
		prox_pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		prox_pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		dist_model_path = self.dist_model_path.get().strip()
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		
		if not all([prox_model_path, prox_pp_abcd_path, prox_pp_olmn_path,
					dist_model_path, dist_pp_abc_path, dist_pp_olmn_path]):
			messagebox.showwarning("入力不足", "Simulatorタブですべてのファイルを選択してください。")
			return
		
		# ライブラリの確認
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return
		
		# PP読み込み（近位）
		try:
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(prox_pp_abcd_path)
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(prox_pp_olmn_path)
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"近位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# PP読み込み（遠位）
		try:
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"遠位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# メッシュ読み込み（近位）
		try:
			prox_mesh = pv.read(prox_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"近位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# メッシュ読み込み（遠位）
		try:
			dist_mesh = pv.read(dist_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"遠位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# 近位CL座標系の構築
		prox_C_cl_origin = None
		prox_C_cl_x_axis = None
		prox_C_cl_y_axis = None
		prox_C_cl_z_axis = None
		prox_CL_point = None
		prox_L_prime_point = None
		
		# 近位: L, M, N点を探す
		prox_L_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'L'), None)
		prox_M_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'M'), None)
		prox_N_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'N'), None)
		
		if prox_L_idx is not None and prox_M_idx is not None and prox_N_idx is not None:
			prox_L_point = prox_points[prox_L_idx]
			prox_M_point = prox_points[prox_M_idx]
			prox_N_point = prox_points[prox_N_idx]
			
			# MとNの中点を計算（CL点）
			prox_CL_point = np.array([
				(prox_M_point[0] + prox_N_point[0]) / 2.0,
				(prox_M_point[1] + prox_N_point[1]) / 2.0,
				(prox_M_point[2] + prox_N_point[2]) / 2.0
			])
			
			# LMベクトルを作成
			prox_LM_vector = prox_M_point - prox_L_point
			prox_LM_unit = prox_LM_vector / np.linalg.norm(prox_LM_vector)
			
			# NからLM直線への垂線の足L'を計算
			prox_LN_vector = prox_N_point - prox_L_point
			prox_projection_length = np.dot(prox_LN_vector, prox_LM_unit)
			prox_L_prime_point = prox_L_point + prox_projection_length * prox_LM_unit
			
			# C_cl座標系を構築
			prox_C_cl_origin = prox_CL_point
			
			# X軸: M → L'
			prox_x_vec = prox_L_prime_point - prox_M_point
			prox_C_cl_x_axis = prox_x_vec / np.linalg.norm(prox_x_vec)
			
			# Z軸: N → L'
			prox_z_vec = prox_L_prime_point - prox_N_point
			prox_C_cl_z_axis = prox_z_vec / np.linalg.norm(prox_z_vec)
			
			# Y軸: Z × X
			prox_C_cl_y_axis = np.cross(prox_C_cl_z_axis, prox_C_cl_x_axis)
			prox_C_cl_y_axis = prox_C_cl_y_axis / np.linalg.norm(prox_C_cl_y_axis)
			
			# CL点とL'点を点群に追加
			prox_points = np.vstack([prox_points, prox_CL_point, prox_L_prime_point])
			prox_labels = prox_labels + ['CL', "L'"]
		
		# 修正量を取得（近位用）
		correction_x = self.c_fix_correction_x.get()
		correction_y = self.c_fix_correction_y.get()
		correction_z = self.c_fix_correction_z.get()
		
		# 近位のA'B'C'D'点を計算
		prox_corrected_points = None
		prox_corrected_labels = []
		if prox_C_cl_origin is not None:
			# A, B, C, D点を探す
			prox_ABCD_indices = []
			prox_ABCD_labels_found = []
			for target_label in ['A', 'B', 'C', 'D']:
				idx = next((i for i, lbl in enumerate(prox_labels) if lbl == target_label), None)
				if idx is not None:
					prox_ABCD_indices.append(idx)
					prox_ABCD_labels_found.append(target_label)
			
			if prox_ABCD_indices:
				# CL座標系に沿って修正量を適用
				correction_vector = (correction_x * prox_C_cl_x_axis + 
									 correction_y * prox_C_cl_y_axis + 
									 correction_z * prox_C_cl_z_axis)
				
				prox_corrected_points = []
				for idx, label in zip(prox_ABCD_indices, prox_ABCD_labels_found):
					original_point = prox_points[idx]
					corrected_point = original_point + correction_vector
					prox_corrected_points.append(corrected_point)
					prox_corrected_labels.append(f"{label}'")
				
				prox_corrected_points = np.array(prox_corrected_points)
		
		# 遠位CL座標系の構築
		dist_C_cl_origin = None
		dist_C_cl_x_axis = None
		dist_C_cl_y_axis = None
		dist_C_cl_z_axis = None
		dist_CL_point = None
		dist_L_prime_point = None
		
		# 遠位: L, M, N点を探す
		dist_L_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'L'), None)
		dist_M_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'M'), None)
		dist_N_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'N'), None)
		
		if dist_L_idx is not None and dist_M_idx is not None and dist_N_idx is not None:
			dist_L_point = dist_points[dist_L_idx]
			dist_M_point = dist_points[dist_M_idx]
			dist_N_point = dist_points[dist_N_idx]
			
			# MとNの中点を計算（CL点）
			dist_CL_point = np.array([
				(dist_M_point[0] + dist_N_point[0]) / 2.0,
				(dist_M_point[1] + dist_N_point[1]) / 2.0,
				(dist_M_point[2] + dist_N_point[2]) / 2.0
			])
			
			# LMベクトルを作成
			dist_LM_vector = dist_M_point - dist_L_point
			dist_LM_unit = dist_LM_vector / np.linalg.norm(dist_LM_vector)
			
			# NからLM直線への垂線の足L'を計算
			dist_LN_vector = dist_N_point - dist_L_point
			dist_projection_length = np.dot(dist_LN_vector, dist_LM_unit)
			dist_L_prime_point = dist_L_point + dist_projection_length * dist_LM_unit
			
			# C_cl座標系を構築
			dist_C_cl_origin = dist_CL_point
			
			# X軸: M → L'
			dist_x_vec = dist_L_prime_point - dist_M_point
			dist_C_cl_x_axis = dist_x_vec / np.linalg.norm(dist_x_vec)
			
			# Z軸: N → L'
			dist_z_vec = dist_L_prime_point - dist_N_point
			dist_C_cl_z_axis = dist_z_vec / np.linalg.norm(dist_z_vec)
			
			# Y軸: Z × X
			dist_C_cl_y_axis = np.cross(dist_C_cl_z_axis, dist_C_cl_x_axis)
			dist_C_cl_y_axis = dist_C_cl_y_axis / np.linalg.norm(dist_C_cl_y_axis)
			
			# CL点とL'点を点群に追加
			dist_points = np.vstack([dist_points, dist_CL_point, dist_L_prime_point])
			dist_labels = dist_labels + ['CL', "L'"]
		
		# 遠位のA'B'C'点を計算（遠位はA, B, Cのみ）
		dist_corrected_points = None
		dist_corrected_labels = []
		if dist_C_cl_origin is not None:
			# A, B, C点を探す
			dist_ABC_indices = []
			dist_ABC_labels_found = []
			for target_label in ['A', 'B', 'C']:
				idx = next((i for i, lbl in enumerate(dist_labels) if lbl == target_label), None)
				if idx is not None:
					dist_ABC_indices.append(idx)
					dist_ABC_labels_found.append(target_label)
			
			if dist_ABC_indices:
				# CL座標系に沿って修正量を適用
				correction_vector = (correction_x * dist_C_cl_x_axis + 
									 correction_y * dist_C_cl_y_axis + 
									 correction_z * dist_C_cl_z_axis)
				
				dist_corrected_points = []
				for idx, label in zip(dist_ABC_indices, dist_ABC_labels_found):
					original_point = dist_points[idx]
					corrected_point = original_point + correction_vector
					dist_corrected_points.append(corrected_point)
					dist_corrected_labels.append(f"{label}'")
				
				dist_corrected_points = np.array(dist_corrected_points)
		
		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)
		
		# プロッタ作成
		plotter = pv.Plotter(title="C. Fix - 座標系確認", window_size=(window_width, window_height))
		plotter.set_background("white")
		
		# 近位メッシュ（水色）
		prox_mesh_actor = plotter.add_mesh(prox_mesh, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=0.5)
		
		# 遠位メッシュ（ピンク）
		dist_mesh_actor = plotter.add_mesh(dist_mesh, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=0.5)
		
		# 近位点群（水色）
		prox_points_actor = plotter.add_points(prox_points, color=self.prox_color, point_size=15, render_points_as_spheres=True)
		
		# 近位点ラベル
		prox_label_actors = []
		for pt, label in zip(prox_points, prox_labels):
			label_actor = plotter.add_point_labels([pt], [f"Prox_{label}"], point_size=0, font_size=14, 
											   text_color=self.prox_color, bold=True, shadow=True, show_points=False)
		
		# 遠位点ラベル
		dist_label_actors = []
		for pt, label in zip(dist_points, dist_labels):
			label_actor = plotter.add_point_labels([pt], [f"Dist_{label}"], point_size=0, font_size=14, 
											   text_color=self.dist_color, bold=True, shadow=True, show_points=False)
		
		prox_corrected_actors = []
		prox_corrected_label_actors = []
		if prox_corrected_points is not None and len(prox_corrected_points) > 0:
			prox_corrected_points_actor = plotter.add_points(
				prox_corrected_points, 
				color='darkturquoise', 
				point_size=18, 
				render_points_as_spheres=True
			)
			prox_corrected_actors.append(prox_corrected_points_actor)
			
			# 修正点のラベル
			for pt, label in zip(prox_corrected_points, prox_corrected_labels):
				label_actor = plotter.add_point_labels(
					[pt], [f"Prox_{label}"], 
					point_size=0, font_size=16, 
					text_color='darkturquoise', 
					bold=True, shadow=True
				)
				prox_corrected_label_actors.append(label_actor)
		
		# 遠位修正点群（A'B'C'）- 濃いピンク
		dist_corrected_actors = []
		dist_corrected_label_actors = []
		if dist_corrected_points is not None and len(dist_corrected_points) > 0:
			dist_corrected_points_actor = plotter.add_points(
				dist_corrected_points, 
				color='deeppink', 
				point_size=18, 
				render_points_as_spheres=True
			)
			dist_corrected_actors.append(dist_corrected_points_actor)
			
			# 修正点のラベル
			for pt, label in zip(dist_corrected_points, dist_corrected_labels):
				label_actor = plotter.add_point_labels(
					[pt], [f"Dist_{label}"], 
					point_size=0, font_size=16, 
					text_color='deeppink', 
					bold=True, shadow=True
				)
				dist_corrected_label_actors.append(label_actor)
		
		# 近位CL座標系の可視化
		prox_cl_actors = []
		if prox_C_cl_origin is not None:
			axis_length = 50.0
			# X軸（赤）
			prox_cl_x_line = pv.Line(prox_C_cl_origin, prox_C_cl_origin + prox_C_cl_x_axis * axis_length)
			prox_cl_x_actor = plotter.add_mesh(prox_cl_x_line, color="red", line_width=4)
			prox_cl_actors.append(prox_cl_x_actor)
			# Y軸（緑）
			prox_cl_y_line = pv.Line(prox_C_cl_origin, prox_C_cl_origin + prox_C_cl_y_axis * axis_length)
			prox_cl_y_actor = plotter.add_mesh(prox_cl_y_line, color="green", line_width=4)
			prox_cl_actors.append(prox_cl_y_actor)
			# Z軸（青）
			prox_cl_z_line = pv.Line(prox_C_cl_origin, prox_C_cl_origin + prox_C_cl_z_axis * axis_length)
			prox_cl_z_actor = plotter.add_mesh(prox_cl_z_line, color="blue", line_width=4)
			prox_cl_actors.append(prox_cl_z_actor)
			# 原点（CL）に小さな球
			prox_cl_sphere = plotter.add_mesh(
				pv.Sphere(radius=3.0, center=prox_C_cl_origin),
				color='yellow',
				opacity=0.8
			)
			prox_cl_actors.append(prox_cl_sphere)
		
		# 遠位CL座標系の可視化
		dist_cl_actors = []
		if dist_C_cl_origin is not None:
			axis_length = 50.0
			# X軸（暗赤）
			dist_cl_x_line = pv.Line(dist_C_cl_origin, dist_C_cl_origin + dist_C_cl_x_axis * axis_length)
			dist_cl_x_actor = plotter.add_mesh(dist_cl_x_line, color="darkred", line_width=4)
			dist_cl_actors.append(dist_cl_x_actor)
			# Y軸（暗緑）
			dist_cl_y_line = pv.Line(dist_C_cl_origin, dist_C_cl_origin + dist_C_cl_y_axis * axis_length)
			dist_cl_y_actor = plotter.add_mesh(dist_cl_y_line, color="darkgreen", line_width=4)
			dist_cl_actors.append(dist_cl_y_actor)
			# Z軸（暗青）
			dist_cl_z_line = pv.Line(dist_C_cl_origin, dist_C_cl_origin + dist_C_cl_z_axis * axis_length)
			dist_cl_z_actor = plotter.add_mesh(dist_cl_z_line, color="darkblue", line_width=4)
			dist_cl_actors.append(dist_cl_z_actor)
			# 原点（CL）に小さな球
			dist_cl_sphere = plotter.add_mesh(
				pv.Sphere(radius=3.0, center=dist_C_cl_origin),
				color='orange',
				opacity=0.8
			)
			dist_cl_actors.append(dist_cl_sphere)
		
		# チェックボックスUI
		def toggle_prox_mesh(state):
			prox_mesh_actor.SetVisibility(state)
		
		def toggle_dist_mesh(state):
			dist_mesh_actor.SetVisibility(state)
		
		def toggle_prox_points(state):
			prox_points_actor.SetVisibility(state)
			for label_actor in prox_label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_dist_points(state):
			dist_points_actor.SetVisibility(state)
			for label_actor in dist_label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_prox_cl(state):
			for actor in prox_cl_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_cl(state):
			for actor in dist_cl_actors:
				actor.SetVisibility(state)
		
		def toggle_prox_corrected(state):
			for actor in prox_corrected_actors:
				actor.SetVisibility(state)
			for actor in prox_corrected_label_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_corrected(state):
			for actor in dist_corrected_actors:
				actor.SetVisibility(state)
			for actor in dist_corrected_label_actors:
				actor.SetVisibility(state)
		
		# チェックボックス配置
		plotter.add_text("Prox Mesh", position=(10, 10), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_prox_mesh, value=True, position=(120, 12), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		plotter.add_text("Dist Mesh", position=(10, 40), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_dist_mesh, value=True, position=(120, 42), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		plotter.add_text("Prox PP", position=(10, 70), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_prox_points, value=True, position=(120, 72), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		plotter.add_text("Dist PP", position=(10, 100), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_dist_points, value=True, position=(120, 102), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		# CL座標系のチェックボックス
		checkbox_y = 130
		if prox_cl_actors:
			plotter.add_text("Prox C_cl", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_prox_cl, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="yellow", color_off="white")
			checkbox_y += 30
		
		if dist_cl_actors:
			plotter.add_text("Dist C_cl", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_dist_cl, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="orange", color_off="white")
			checkbox_y += 30
		
		# 修正点のチェックボックス
		if prox_corrected_actors:
			plotter.add_text("Prox A'B'C'D'", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_prox_corrected, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="darkturquoise", color_off="white")
			checkbox_y += 30
		
		if dist_corrected_actors:
			plotter.add_text("Dist A'B'C'", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_dist_corrected, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="deeppink", color_off="white")
		
		# カメラを全体が見えるように自動調整
		plotter.reset_camera()
		plotter.show()

	def on_c_fix_transform_prox(self) -> None:
		"""C. Fix: 近位ABCD修正後の座標をPPファイルとして保存"""
		prox_pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		prox_pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		
		if not all([prox_pp_abcd_path, prox_pp_olmn_path]):
			messagebox.showwarning("入力不足", "近位のPPファイルを選択してください。")
			return
		
		try:
			# PP読み込み（近位）
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(prox_pp_abcd_path)
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(prox_pp_olmn_path)
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"近位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# 近位CL座標系の構築
		prox_L_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'L'), None)
		prox_M_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'M'), None)
		prox_N_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'N'), None)
		
		if prox_L_idx is None or prox_M_idx is None or prox_N_idx is None:
			messagebox.showerror("エラー", "近位のL, M, N点が見つかりません。")
			return
		
		prox_L_point = prox_points[prox_L_idx]
		prox_M_point = prox_points[prox_M_idx]
		prox_N_point = prox_points[prox_N_idx]
		
		prox_CL_point = (prox_M_point + prox_N_point) / 2.0
		prox_LM_vector = prox_M_point - prox_L_point
		prox_LM_unit = prox_LM_vector / np.linalg.norm(prox_LM_vector)
		prox_LN_vector = prox_N_point - prox_L_point
		prox_projection_length = np.dot(prox_LN_vector, prox_LM_unit)
		prox_L_prime_point = prox_L_point + prox_projection_length * prox_LM_unit
		
		prox_C_cl_origin = prox_CL_point
		prox_x_vec = prox_L_prime_point - prox_M_point
		prox_C_cl_x_axis = prox_x_vec / np.linalg.norm(prox_x_vec)
		prox_z_vec = prox_L_prime_point - prox_N_point
		prox_C_cl_z_axis = prox_z_vec / np.linalg.norm(prox_z_vec)
		prox_C_cl_y_axis = np.cross(prox_C_cl_z_axis, prox_C_cl_x_axis)
		prox_C_cl_y_axis = prox_C_cl_y_axis / np.linalg.norm(prox_C_cl_y_axis)
		
		# 修正量を取得
		correction_x = self.c_fix_correction_x.get()
		correction_y = self.c_fix_correction_y.get()
		correction_z = self.c_fix_correction_z.get()
		
		# 近位ABCD点を修正
		prox_ABCD_corrected = {}
		for target_label in ['A', 'B', 'C', 'D']:
			idx = next((i for i, lbl in enumerate(prox_labels) if lbl == target_label), None)
			if idx is not None:
				original_point = prox_points[idx]
				correction_vector = (correction_x * prox_C_cl_x_axis + 
									 correction_y * prox_C_cl_y_axis + 
									 correction_z * prox_C_cl_z_axis)
				corrected_point = original_point + correction_vector
				prox_ABCD_corrected[target_label] = corrected_point
		
		# 近位ABCDファイルの保存
		prox_abcd_save_path = filedialog.asksaveasfilename(
			title="近位ABCD修正後のPPファイルを保存",
			defaultextension=".pp",
			filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")],
			initialdir=str(Path(prox_pp_abcd_path).parent),
			initialfile=Path(prox_pp_abcd_path).stem + "_corrected.pp"
		)
		
		if not prox_abcd_save_path:
			messagebox.showinfo("キャンセル", "保存がキャンセルされました。")
			return
		
		try:
			# ABCDの順で保存
			prox_abcd_points_list = []
			prox_abcd_labels_list = []
			for label in ['A', 'B', 'C', 'D']:
				if label in prox_ABCD_corrected:
					prox_abcd_points_list.append(prox_ABCD_corrected[label])
					prox_abcd_labels_list.append(label)
			
			if prox_abcd_points_list:
				save_child_pp_file(prox_abcd_save_path, np.array(prox_abcd_points_list), prox_abcd_labels_list)
				messagebox.showinfo("完了", f"近位ABCD修正後のPPファイルを保存しました:\n{prox_abcd_save_path}")
		except Exception as e:
			messagebox.showerror("保存エラー", f"ファイルの保存に失敗しました:\n{e}")

	def on_c_fix_transform_dist(self) -> None:
		"""C. Fix: 遠位ABC修正後の座標をPPファイルとして保存"""
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		
		if not all([dist_pp_abc_path, dist_pp_olmn_path]):
			messagebox.showwarning("入力不足", "遠位のPPファイルを選択してください。")
			return
		
		try:
			# PP読み込み（遠位）
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"遠位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# 遠位CL座標系の構築
		dist_L_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'L'), None)
		dist_M_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'M'), None)
		dist_N_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'N'), None)
		
		if dist_L_idx is None or dist_M_idx is None or dist_N_idx is None:
			messagebox.showerror("エラー", "遠位のL, M, N点が見つかりません。")
			return
		
		dist_L_point = dist_points[dist_L_idx]
		dist_M_point = dist_points[dist_M_idx]
		dist_N_point = dist_points[dist_N_idx]
		
		dist_CL_point = (dist_M_point + dist_N_point) / 2.0
		dist_LM_vector = dist_M_point - dist_L_point
		dist_LM_unit = dist_LM_vector / np.linalg.norm(dist_LM_vector)
		dist_LN_vector = dist_N_point - dist_L_point
		dist_projection_length = np.dot(dist_LN_vector, dist_LM_unit)
		dist_L_prime_point = dist_L_point + dist_projection_length * dist_LM_unit
		
		dist_C_cl_origin = dist_CL_point
		dist_x_vec = dist_L_prime_point - dist_M_point
		dist_C_cl_x_axis = dist_x_vec / np.linalg.norm(dist_x_vec)
		dist_z_vec = dist_L_prime_point - dist_N_point
		dist_C_cl_z_axis = dist_z_vec / np.linalg.norm(dist_z_vec)
		dist_C_cl_y_axis = np.cross(dist_C_cl_z_axis, dist_C_cl_x_axis)
		dist_C_cl_y_axis = dist_C_cl_y_axis / np.linalg.norm(dist_C_cl_y_axis)
		
		# 修正量を取得
		correction_x = self.c_fix_correction_x.get()
		correction_y = self.c_fix_correction_y.get()
		correction_z = self.c_fix_correction_z.get()
		
		# 遠位ABC点を修正
		dist_ABC_corrected = {}
		for target_label in ['A', 'B', 'C']:
			idx = next((i for i, lbl in enumerate(dist_labels) if lbl == target_label), None)
			if idx is not None:
				original_point = dist_points[idx]
				correction_vector = (correction_x * dist_C_cl_x_axis + 
									 correction_y * dist_C_cl_y_axis + 
									 correction_z * dist_C_cl_z_axis)
				corrected_point = original_point + correction_vector
				dist_ABC_corrected[target_label] = corrected_point
		
		# 遠位ABCファイルの保存
		dist_abc_save_path = filedialog.asksaveasfilename(
			title="遠位ABC修正後のPPファイルを保存",
			defaultextension=".pp",
			filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")],
			initialdir=str(Path(dist_pp_abc_path).parent),
			initialfile=Path(dist_pp_abc_path).stem + "_corrected.pp"
		)
		
		if not dist_abc_save_path:
			messagebox.showinfo("キャンセル", "保存がキャンセルされました。")
			return
		
		try:
			# ABCの順で保存
			dist_abc_points_list = []
			dist_abc_labels_list = []
			for label in ['A', 'B', 'C']:
				if label in dist_ABC_corrected:
					dist_abc_points_list.append(dist_ABC_corrected[label])
					dist_abc_labels_list.append(label)
			
			if dist_abc_points_list:
				save_child_pp_file(dist_abc_save_path, np.array(dist_abc_points_list), dist_abc_labels_list)
				messagebox.showinfo("完了", f"遠位ABC修正後のPPファイルを保存しました:\n{dist_abc_save_path}")
		except Exception as e:
			messagebox.showerror("保存エラー", f"ファイルの保存に失敗しました:\n{e}")

	def on_animate(self) -> None:
		"""
		【最重要メソッド】アニメーション実行: 近位固定、遠位を同次変換行列に沿って動かす
		
		このメソッドは約1400行のコードで構成されており、以下の処理を行います:
		
		■ 処理フロー:
		  1. ファイル読み込み（近位・遠位モデル、PP、変換行列）
		  2. 座標系構築（C-RobotScan座標系とC-CL座標系）
		  3. ワールド座標系への統一変換
		  4. ヒートマップ事前計算（オプション、並列処理対応）
		  5. PyVistaプロッターUI構築
		  6. アニメーションループ（リアルタイム座標変換と可視化）
		  
		■ 座標系の種類:
		  - C-RobotScan: ロボット座標系（元データ）
		  - C-CL: 解剖学的座標系（L, M, N点から構築）
		  - ワールド座標系: 可視化用の統一座標系
		  
		■ 主要な内部関数:
		  - show_frame(frame_idx): 指定フレームを描画
		  - animation_loop(): アニメーション再生ループ
		  - on_playback_scale_change(frame_idx): スライダー操作でのフレーム移動
		  - export_to_csv(): 接触深度データをCSVエクスポート
		  
		■ 改修ポイント:
		  - 行5520付近: ヒートマップ計算処理（並列化可能）
		  - 行6230-6540: show_frame関数（座標変換ロジック）
		  - 行6541-6650: animation_loop関数（再生速度制御）
		  
		■ パフォーマンス最適化:
		  - キャッシュ機能: 2回目以降はヒートマップを再利用
		  - 並列処理: マルチコアCPUでヒートマップ計算を高速化
		  - ダウンサンプリング: 点群数を削減して計算を軽量化
		  
		■ トラブルシューティング:
		  - メッシュ読み込み失敗: ファイルパスとフォーマットを確認
		  - 座標系構築失敗: L, M, N点の存在を確認
		  - アニメーション動作不良: transform_dataの行列が正しいか確認
		"""
		# 必要なファイルがすべて選択されているか確認
		if not all([
			self.prox_model_path.get(),
			self.prox_pp_abcd_path.get(),
			self.prox_pp_olmn_path.get(),
			self.dist_model_path.get(),
			self.dist_pp_abc_path.get(),
			self.dist_pp_olmn_path.get(),
			self.transform_group_path.get()
		]):
			messagebox.showerror("エラー", "全てのファイルを選択してください。")
			return
		
		try:
			# 変位・姿勢変化データを読み込み
			transform_data = self._load_transform_matrices(self.transform_group_path.get())
			if not transform_data:
				messagebox.showerror("エラー", "変位・姿勢変化データが空です。")
				return
			
			# 近位モデルとPPを読み込み - 2つのPPファイルを統合
			prox_mesh = pv.read(self.prox_model_path.get())
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(self.prox_pp_abcd_path.get())
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(self.prox_pp_olmn_path.get())
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
			
			# 遠位モデルとPPを読み込み - 2つのPPファイルを統合
			dist_mesh = pv.read(self.dist_model_path.get())
			dist_points_abc, dist_labels_abc = self._parse_pp_file(self.dist_pp_abc_path.get())
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(self.dist_pp_olmn_path.get())
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
			
			# 近位座標系を構築（_build_coordinate_systemを使用）
			try:
				prox_origin, prox_x_axis, prox_y_axis, prox_z_axis = self._build_coordinate_system(prox_points, prox_labels)
				# 自動オフセット設定
				self._try_auto_set_offset(prox_points, prox_labels, prox_origin, prox_x_axis, prox_y_axis, prox_z_axis,
				                          self.prox_offset_x, self.prox_offset_y, self.prox_offset_z)
			except Exception as e:
				messagebox.showerror("近位座標系構築", f"近位座標系の構築に失敗しました:\n{e}")
				return
			
			# 遠位座標系を構築
			try:
				dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = self._build_coordinate_system_dist(dist_points, dist_labels)
				# 自動オフセット設定
				self._try_auto_set_offset(dist_points, dist_labels, dist_origin, dist_x_axis, dist_y_axis, dist_z_axis,
				                          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z)
			except Exception as e:
				messagebox.showerror("遠位座標系構築", f"遠位座標系の構築に失敗しました:\n{e}")
				return
			
			# --- Step 1: 近位座標系をワールド座標系に揃える変換 ---
			prox_rotation_matrix = np.column_stack([prox_x_axis, prox_y_axis, prox_z_axis])
			prox_to_world = np.eye(4)
			prox_to_world[:3, :3] = prox_rotation_matrix.T
			prox_to_world[:3, 3] = -prox_rotation_matrix.T @ prox_origin
			
			# 近位メッシュを変換
			prox_mesh_homo = np.hstack([prox_mesh.points, np.ones((prox_mesh.points.shape[0], 1))])
			prox_mesh.points = (prox_to_world @ prox_mesh_homo.T).T[:, :3]
			
			# 近位座標系をワールド座標系に変換
			prox_origin = np.array([0.0, 0.0, 0.0])
			prox_x_axis = np.array([1.0, 0.0, 0.0])
			prox_y_axis = np.array([0.0, 1.0, 0.0])
			prox_z_axis = np.array([0.0, 0.0, 1.0])
			
			# --- Step 2: 遠位メッシュと座標系に同じ変換を適用 ---
			dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
			dist_mesh.points = (prox_to_world @ dist_mesh_homo.T).T[:, :3]
			
			# 遠位座標系の原点と軸を変換
			dist_origin_homo = np.append(dist_origin, 1)
			dist_origin_in_world = (prox_to_world @ dist_origin_homo)[:3]
			
			dist_rotation_matrix = np.column_stack([dist_x_axis, dist_y_axis, dist_z_axis])
			dist_rotation_in_world = prox_rotation_matrix.T @ dist_rotation_matrix
			dist_x_axis_in_world = dist_rotation_in_world[:, 0]
			dist_y_axis_in_world = dist_rotation_in_world[:, 1]
			dist_z_axis_in_world = dist_rotation_in_world[:, 2]
			
			# --- Step 3: 遠位座標系をワールド座標系に揃える変換 ---
			dist_to_world = np.eye(4)
			dist_to_world[:3, :3] = dist_rotation_in_world.T
			dist_to_world[:3, 3] = -dist_rotation_in_world.T @ dist_origin_in_world
			
			# 遠位メッシュを変換
			dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
			dist_mesh.points = (dist_to_world @ dist_mesh_homo.T).T[:, :3]
			
			# 遠位座標系をワールド座標系に変換
			dist_origin = np.array([0.0, 0.0, 0.0])
			dist_x_axis = np.array([1.0, 0.0, 0.0])
			dist_y_axis = np.array([0.0, 1.0, 0.0])
			dist_z_axis = np.array([0.0, 0.0, 1.0])
			
			# --- 軟骨メッシュの読み込みと変換（任意） ---
			prox_cartilage_mesh = None
			dist_cartilage_mesh = None
			dist_cartilage_for_anim = None
			
			prox_cart_path = self.prox_cartilage_model_path.get().strip()
			dist_cart_path = self.dist_cartilage_model_path.get().strip()
			
			if prox_cart_path:
				try:
					prox_cartilage_mesh = pv.read(prox_cart_path)
					# 骨と同じ変換を適用: prox_to_world
					pc_homo = np.hstack([prox_cartilage_mesh.points, np.ones((prox_cartilage_mesh.points.shape[0], 1))])
					prox_cartilage_mesh.points = (prox_to_world @ pc_homo.T).T[:, :3]
					print(f"近位軟骨モデル読み込み: {Path(prox_cart_path).name} ({prox_cartilage_mesh.n_points}点)")
				except Exception as e:
					print(f"近位軟骨モデル読み込み失敗: {e}")
					prox_cartilage_mesh = None
			
			if dist_cart_path:
				try:
					dist_cartilage_mesh = pv.read(dist_cart_path)
					# 骨と同じ変換を適用: prox_to_world → dist_to_world
					dc_homo = np.hstack([dist_cartilage_mesh.points, np.ones((dist_cartilage_mesh.points.shape[0], 1))])
					dist_cartilage_mesh.points = (prox_to_world @ dc_homo.T).T[:, :3]
					dc_homo2 = np.hstack([dist_cartilage_mesh.points, np.ones((dist_cartilage_mesh.points.shape[0], 1))])
					dist_cartilage_mesh.points = (dist_to_world @ dc_homo2.T).T[:, :3]
					# アニメーション用に初期位置を保存
					dist_cartilage_for_anim = dist_cartilage_mesh.copy()
					print(f"遠位軟骨モデル読み込み: {Path(dist_cart_path).name} ({dist_cartilage_mesh.n_points}点)")
				except Exception as e:
					print(f"遠位軟骨モデル読み込み失敗: {e}")
					dist_cartilage_mesh = None

			# --- 関節領域の抽出 ---
			# O'原点位置を計算（ローカル座標系に沿ったオフセット）
			# Step 1, 3でメッシュと座標系はワールド原点(0,0,0)と軸(1,0,0等)に整列済み
			prox_origin_prime = np.array([
				self.prox_offset_x.get(),
				self.prox_offset_y.get(),
				self.prox_offset_z.get()
			])
			
			dist_origin_prime = np.array([
				self.dist_offset_x.get(),
				self.dist_offset_y.get(),
				self.dist_offset_z.get()
			])
			
			print(f"[抽出] 近位O: {prox_origin}, 近位O': {prox_origin_prime}, オフセット設定: [{self.prox_offset_x.get()}, {self.prox_offset_y.get()}, {self.prox_offset_z.get()}]")
			print(f"[抽出] 遠位O: {dist_origin}, 遠位O': {dist_origin_prime}, オフセット設定: [{self.dist_offset_x.get()}, {self.dist_offset_y.get()}, {self.dist_offset_z.get()}]")
			
			# 近位関節領域（寛骨臼）を球体で抽出（O'を中心に）
			prox_joint_region = self._extract_region_by_sphere(prox_mesh, prox_origin_prime, self.prox_radius.get())
			
			# 遠位関節領域（大腿骨頭）を球体で抽出（O'を中心に）
			dist_joint_region = self._extract_region_by_sphere(dist_mesh, dist_origin_prime, self.dist_radius.get())
			
			# 関節領域が空の場合は警告
			if prox_joint_region.n_points == 0:
				messagebox.showwarning("警告", "近位関節領域（寛骨臼）が抽出できませんでした。半径を確認してください。")
			if dist_joint_region.n_points == 0:
				messagebox.showwarning("警告", "遠位関節領域（大腿骨頭）が抽出できませんでした。半径を確認してください。")
			
			# --- オーバーラップ体積の事前計算 ---
			# 事前計算ダイアログを表示
			has_cartilage = (prox_cartilage_mesh is not None and dist_cartilage_for_anim is not None)
			has_joint_regions = (prox_joint_region is not None and prox_joint_region.n_points > 0
			                     and dist_joint_region is not None and dist_joint_region.n_points > 0)
			has_fem_target = has_cartilage or has_joint_regions
			progress_window, update_progress, cancel_var, options_dict, start_var, skip_var = self._show_precompute_dialog(len(transform_data), has_cartilage=has_fem_target)
			
			# ユーザーがボタンを押すまで待機（モーダルダイアログ）
			# ボタンが押されるとdialog_closedがTrueになる
			while progress_window.winfo_exists():
				if start_var.get() or skip_var.get() or cancel_var.get():
					break
				progress_window.update()
				time.sleep(0.01)
			
			# キャンセルされた場合は終了
			if cancel_var.get():
				try:
					progress_window.destroy()
				except:
					pass
				return
			
			# オーバーラップ計算を実行
			overlap_precomputed = []
			overlap_areas_precomputed = []  # 面積リストを追加
			overlap_depths_precomputed = []  # 深度リストを追加
			heatmap_precomputed = []  # 距離ヒートマップリストを追加
			
			if skip_var.get():
				# 事前計算をスキップ
				try:
					progress_window.destroy()
				except:
					pass
				messagebox.showinfo("スキップ", "事前計算をスキップしました。\nオーバーラップ表示は無効になります。")
			else:
				# 事前計算を実行する
				try:
					# 最適化オプションを取得
					use_parallel = options_dict['use_parallel'].get()
					use_cache = options_dict['use_cache'].get()
					use_bbox = options_dict['use_bbox'].get()
					use_simplify = options_dict['use_simplify'].get()
					
					# 事前計算を実行
					def progress_wrapper(current, total, msg):
						if cancel_var.get():
							return False
						return update_progress(current, total, msg)
					
					# オーバーラップ体積の計算
					result = self._precompute_overlap_volumes(
						prox_mesh, dist_mesh, transform_data, 
						progress_callback=progress_wrapper,
						use_cache=use_cache,
						use_parallel=use_parallel,
						use_bbox_check=use_bbox,
						simplify_mesh=use_simplify
					)
					
					# 戻り値を展開（ヒートマップを含む場合と含まない場合の両方に対応）
					if len(result) == 4:
						overlap_precomputed, overlap_areas_precomputed, overlap_depths_precomputed, heatmap_precomputed = result
						if heatmap_precomputed:
							print(f"[キャッシュ] ヒートマップがキャッシュから読み込まれました: {len(heatmap_precomputed)}フレーム")
							# 最初のフレームの情報を表示
							if len(heatmap_precomputed) > 0:
								first_heatmap = heatmap_precomputed[0]
								print(f"[キャッシュ] 最初のヒートマップ: {first_heatmap.n_points}点")
								print(f"[キャッシュ] スカラー配列: {first_heatmap.array_names}")
								if 'distance' in first_heatmap.array_names:
									distances = first_heatmap['distance']
									print(f"[キャッシュ] 距離範囲: min={distances.min():.2f}, max={distances.max():.2f}")
						else:
							print(f"[キャッシュ] ヒートマップは空でした")
					else:
						overlap_precomputed, overlap_areas_precomputed, overlap_depths_precomputed = result
						heatmap_precomputed = []  # ヒートマップがない場合は空リスト
						print(f"[キャッシュ] 旧形式のキャッシュ（ヒートマップなし）")
					
					if cancel_var.get():
						messagebox.showinfo("キャンセル", "事前計算がキャンセルされました。\nオーバーラップ表示は無効になります。")
						overlap_precomputed = []
						overlap_areas_precomputed = []
						overlap_depths_precomputed = []
						heatmap_precomputed = []
					else:
						# 新規計算フラグ（キャッシュ保存判定用）
						heatmap_newly_computed = False
						
						# 距離ヒートマップの事前計算（キャッシュから読み込まれていない場合のみ）
						if not heatmap_precomputed:
							print(f"[事前計算] 距離ヒートマップを{len(transform_data)}フレーム分計算します...")
							print(f"[事前計算] 近位関節領域: {prox_joint_region.n_points}点")
							print(f"[事前計算] 遠位関節領域: {dist_joint_region.n_points}点")
							
							if prox_joint_region.n_points == 0 or dist_joint_region.n_points == 0:
								print(f"[警告] 関節領域が空のため、ヒートマップ計算をスキップします")
								heatmap_precomputed = []
							else:
								# 事前最適化: 表面抽出を1回だけ実行（ループ外で）
								prox_surface = prox_joint_region.extract_surface() if hasattr(prox_joint_region, 'extract_surface') else prox_joint_region
								dist_surface = dist_joint_region.extract_surface() if hasattr(dist_joint_region, 'extract_surface') else dist_joint_region
								print(f"[最適化] 表面抽出完了: prox={prox_surface.n_points}点, dist={dist_surface.n_points}点")
								
								# Phase 3最適化: 並列処理は一旦無効化（安定性優先）
								use_parallel_heatmap = False  # use_parallel and len(transform_data) > 1
								if use_parallel_heatmap:
									print(f"[Phase 3] 並列処理を使用してヒートマップ計算を高速化します")
								else:
									print(f"[最適化] 逐次処理でヒートマップを計算します（安定性優先）")
								
								# Phase 2最適化: prox_points配列を事前抽出（不変なので1回だけ）
								prox_points = np.asarray(prox_joint_region.points)
								print(f"[最適化] prox_points配列を事前抽出: {len(prox_points)}点")
								
								# 事前最適化: VTKエラー出力を1回だけ設定（ループ外で）
								import vtk
								import os
								vtk_error_output = vtk.vtkFileOutputWindow()
								vtk_error_output.SetFileName("nul" if os.name == 'nt' else "/dev/null")
								vtk_original_output = vtk.vtkOutputWindow.GetInstance()
								vtk.vtkOutputWindow.SetInstance(vtk_error_output)
								
								try:
									if use_parallel_heatmap:
										# Phase 3: スレッド並列処理でヒートマップを計算（GIL制約あるが、I/O待機が多いため有効）
										from concurrent.futures import ThreadPoolExecutor, as_completed
										import multiprocessing
										
										max_workers = min(multiprocessing.cpu_count(), len(transform_data))
										print(f"[Phase 3] {max_workers}個のワーカーでスレッド並列処理を実行")
										
										# 並列処理用のヘルパー関数
										def compute_single_heatmap(args):
											i, tf_info, prox_region, dist_region, prox_surf, dist_surf, prox_pts = args
											try:
												# 遠位関節領域を変換
												dist_joint_transformed = self._apply_transform(dist_region.copy(), tf_info['matrix'])
												dist_surface_transformed = self._apply_transform(dist_surf.copy(), tf_info['matrix'])
												
												# ヒートマップを計算
												heatmap = self._compute_distance_heatmap(prox_region, dist_joint_transformed, prox_surf, dist_surface_transformed, prox_pts)
												return (i, heatmap, None)
											except Exception as e:
												return (i, None, str(e))
										
										# 並列タスクを準備
										tasks = [
											(i, tf_info, prox_joint_region, dist_joint_region, prox_surface, dist_surface, prox_points)
											for i, tf_info in enumerate(transform_data)
										]
										
										# 結果を格納する配列（インデックス順に保持）
										heatmap_results = [None] * len(transform_data)
										
										with ThreadPoolExecutor(max_workers=max_workers) as executor:
											futures = {executor.submit(compute_single_heatmap, task): task[0] for task in tasks}
											
											for future in as_completed(futures):
												if cancel_var.get():
													break
												
												i, heatmap, error = future.result()
												
												if error:
													print(f"[並列処理] フレーム {i}: 計算失敗 - {error}")
													# ダミーヒートマップを作成
													empty_heatmap = prox_joint_region.copy()
													if empty_heatmap.n_points > 0:
														empty_heatmap['distance'] = np.zeros(empty_heatmap.n_points)
													heatmap_results[i] = empty_heatmap
												else:
													heatmap_results[i] = heatmap
												
												# 進捗更新
												completed = sum(1 for h in heatmap_results if h is not None)
												if not update_progress(completed, len(transform_data), f"並列ヒートマップ計算中: {completed}/{len(transform_data)}"):
													break
										
										# 結果をリストに追加
										for heatmap in heatmap_results:
											if heatmap is not None:
												heatmap_precomputed.append(heatmap)
										
										# デバッグ: 最初のフレームの情報を表示
										if len(heatmap_precomputed) > 0:
											first_heatmap = heatmap_precomputed[0]
											print(f"[並列処理] フレーム0のヒートマップ: {first_heatmap.n_points}点, スカラー: {first_heatmap.array_names}")
											if 'distance' in first_heatmap.array_names:
												distances = first_heatmap['distance']
												print(f"[並列処理] 距離の範囲: min={distances.min():.2f}mm, max={distances.max():.2f}mm, mean={distances.mean():.2f}mm")
										
									else:
										# 従来の逐次処理
										for i, tf_info in enumerate(transform_data):
											if cancel_var.get():
												break
											
											try:
												# 遠位関節領域を変換（最適化: copyを最小限に）
												dist_joint_transformed = self._apply_transform(dist_joint_region.copy(), tf_info['matrix'])
												dist_surface_transformed = self._apply_transform(dist_surface.copy(), tf_info['matrix'])
												
												# 距離ヒートマップを計算（prox_joint_regionは変更されないため.copy()不要、表面は事前抽出済み）
												heatmap = self._compute_distance_heatmap(prox_joint_region, dist_joint_transformed, prox_surface, dist_surface_transformed, prox_points)
												heatmap_precomputed.append(heatmap)
												
												# デバッグログは100フレーム毎のみ（I/O負荷削減）
												if i % 100 == 0:
													print(f"[進捗] フレーム{i}/{len(transform_data)}完了")
												
												# デバッグ: 最初のフレームの情報を表示
												if i == 0:
													print(f"[事前計算] フレーム0のヒートマップ: {heatmap.n_points}点, スカラー: {heatmap.array_names}")
													if 'distance' in heatmap.array_names:
														distances = heatmap['distance']
														print(f"[事前計算] 距離の範囲: min={distances.min():.2f}mm, max={distances.max():.2f}mm, mean={distances.mean():.2f}mm")
												
												# 進捗更新を10フレーム毎にバッチ化（GUI更新負荷削減）
												if (i + 1) % 10 == 0 or (i + 1) == len(transform_data):
													if not update_progress(i + 1, len(transform_data), f"ヒートマップ計算中: {i+1}/{len(transform_data)}"):
														break
											
											except Exception as e:
												# import traceback # 既にグローバルでインポート済み
												print(f"[エラー] フレーム {i}: ヒートマップ計算失敗")
												print(f"[エラー詳細] {e}")
												print(f"[スタックトレース]\n{traceback.format_exc()}")
												# 空のメッシュにダミーデータを追加
												try:
													empty_heatmap = prox_joint_region.copy()
													if empty_heatmap.n_points > 0:
														empty_heatmap['distance'] = np.zeros(empty_heatmap.n_points)
													heatmap_precomputed.append(empty_heatmap)
												except Exception as e2:
													print(f"[エラー] ダミーヒートマップ作成も失敗: {e2}")
								finally:
									# VTKエラー出力を元に戻す
									vtk.vtkOutputWindow.SetInstance(vtk_original_output)
								
								# 新規計算完了フラグを立てる
								heatmap_newly_computed = True
						else:
							print(f"[キャッシュ] ヒートマップはキャッシュから読み込み済みです")
						
						if cancel_var.get():
							messagebox.showinfo("キャンセル", "事前計算がキャンセルされました。\nオーバーラップ表示は無効になります。")
							overlap_precomputed = []
							overlap_areas_precomputed = []
							overlap_depths_precomputed = []
							heatmap_precomputed = []
						else:
							# ヒートマップも含めてキャッシュに保存（新規計算した場合のみ）
							if use_cache and heatmap_newly_computed and heatmap_precomputed:
								cache_hash = self._get_cache_hash(prox_mesh, dist_mesh, transform_data, use_simplify)
								cache_filepath = self._get_cache_filepath(cache_hash)
								print(f"[キャッシュ] 新規計算したヒートマップをキャッシュに保存します...")
								self._save_overlap_cache(cache_filepath, overlap_precomputed, overlap_areas_precomputed, overlap_depths_precomputed, heatmap_precomputed)
							
							messagebox.showinfo("完了", f"事前計算が完了しました。\n{len(overlap_precomputed)}フレーム分のオーバーラップと\n{len(heatmap_precomputed)}フレーム分のヒートマップを生成しました。")

					# --- FEM接触解析の事前計算（関節領域で実行） ---
					fem_pressure_precomputed = []  # List of (pressure_array, results_summary_dict) per frame
					prox_fem_surface = None  # For mapping pressure results back to heatmap points
					if options_dict.get('enable_fem_precompute') and options_dict['enable_fem_precompute'].get():
						if prox_joint_region.n_points > 0 and dist_joint_region.n_points > 0 and _HAS_FEM:
							print(f"[FEM事前計算] 関節領域で{len(transform_data)}フレーム分のFEM解析を実行します...")
							print(f"[FEM事前計算] 近位関節領域: {prox_joint_region.n_points}点")
							print(f"[FEM事前計算] 遠位関節領域: {dist_joint_region.n_points}点")
							try:
								material = MaterialProperties(
									E=self.fem_cart_E.get(),
									nu=self.fem_cart_nu.get(),
									thickness=self.fem_cart_thickness.get(),
								)
								contact = ContactParameters(
									penalty_stiffness=self.fem_penalty_stiffness.get(),
									contact_tolerance=self.fem_contact_tolerance.get(),
								)
								# 近位関節領域の表面を事前抽出（FEM解析用）
								prox_fem_surface = prox_joint_region.extract_surface() if hasattr(prox_joint_region, 'extract_surface') else prox_joint_region
								print(f"[FEM事前計算] 近位FEM表面: {prox_fem_surface.n_points}点")
								for i, tf_info in enumerate(transform_data):
									if cancel_var.get():
										break
									try:
										# 遠位関節領域をこのフレームの姿勢に変換
										matrix = tf_info['matrix']
										dist_joint_transformed = self._apply_transform(dist_joint_region.copy(), matrix)
										dist_fem_surface = dist_joint_transformed.extract_surface() if hasattr(dist_joint_transformed, 'extract_surface') else dist_joint_transformed
										# FEM解析を実行
										solver = FEMContactSolver(material=material, contact=contact, verbose=False)
										results = solver.analyze(
											prox_fem_surface.copy(), dist_fem_surface.copy(),
											boundary_mode=self.fem_boundary_mode.get(),
											max_nodes=self.fem_max_nodes.get(),
										)
										# contact_pressure配列を保存（prox_fem_surface.n_points長）
										# ヒートマップのprox_joint_regionに対応させるため、必要に応じてマッピング
										if results.n_nodes == prox_fem_surface.n_points:
											pressure = results.contact_pressure.copy()
										else:
											# デシメーションが起きた場合、元の近位メッシュにマッピング
											from scipy.spatial import cKDTree
											tree = cKDTree(solver._nodes)
											_, idx = tree.query(prox_fem_surface.points, k=1)
											pressure = results.contact_pressure[idx]
										summary = {
											'contact_area': results.contact_area,
											'peak_contact_pressure': results.peak_contact_pressure,
											'n_contact_nodes': results.n_contact_nodes,
											'solve_time_sec': results.solve_time_sec,
										}
										fem_pressure_precomputed.append((pressure, summary))
									except Exception as e:
										print(f"[FEM事前計算] フレーム{i}でエラー: {e}")
										fem_pressure_precomputed.append(None)
									# 進捗更新
									if (i + 1) % 5 == 0 or (i + 1) == len(transform_data):
										update_progress(i + 1, len(transform_data), f"FEM解析: フレーム {i+1}/{len(transform_data)}")
								n_success = sum(1 for x in fem_pressure_precomputed if x is not None)
								print(f"[FEM事前計算] 完了: {n_success}/{len(transform_data)}フレーム成功")
							except Exception as e:
								print(f"[FEM事前計算] 初期化エラー: {e}")
								import traceback
								traceback.print_exc()
								fem_pressure_precomputed = []
						else:
							if not _HAS_FEM:
								print("[FEM事前計算] FEMモジュールが利用できません")
							else:
								print("[FEM事前計算] 関節領域が空のためスキップ")
					
				except Exception as e:
					messagebox.showerror("エラー", f"事前計算中にエラーが発生しました:\n{e}")
					overlap_precomputed = []
					overlap_areas_precomputed = []
					overlap_depths_precomputed = []
					heatmap_precomputed = []
				
				finally:
					# プログレスダイアログを閉じる
					try:
						progress_window.destroy()
					except:
						pass

			# --- プロッターの作成 ---
			# 空のメッシュをプロット可能にする設定（プロッター作成前に設定）
			pv.global_theme.allow_empty_mesh = True
			
			# ディスプレイサイズの取得
			screen_width = self.winfo_screenwidth()
			screen_height = self.winfo_screenheight()
			window_width = int(screen_width * 0.9)
			window_height = int(screen_height * 0.9)
			
			anim_plotter = pv.Plotter(title="アニメーション実行", window_size=(window_width, window_height))
			anim_plotter.set_background("white")
			
			# 近位モデルを追加（透明度0.5でフィッティングと合わせる）
			prox_actor = anim_plotter.add_mesh(prox_mesh, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=0.5)
			
			# 近位座標系の軸を表示
			axis_length = 50.0
			prox_x_line = pv.Line(prox_origin, prox_origin + prox_x_axis * axis_length)
			prox_y_line = pv.Line(prox_origin, prox_origin + prox_y_axis * axis_length)
			prox_z_line = pv.Line(prox_origin, prox_origin + prox_z_axis * axis_length)
			prox_x_actor = anim_plotter.add_mesh(prox_x_line, color='red', line_width=3, label='Prox X')
			prox_y_actor = anim_plotter.add_mesh(prox_y_line, color='green', line_width=3, label='Prox Y')
			prox_z_actor = anim_plotter.add_mesh(prox_z_line, color='blue', line_width=3, label='Prox Z')
			
			# 遠位モデルを追加（透明度0.5）
			dist_actor = anim_plotter.add_mesh(dist_mesh, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=0.3)
			
			# 軟骨モデルを追加（任意）
			prox_cart_actor = None
			dist_cart_actor = None
			if prox_cartilage_mesh is not None:
				prox_cart_actor = anim_plotter.add_mesh(prox_cartilage_mesh, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='近位軟骨')
			if dist_cartilage_mesh is not None:
				dist_cart_actor = anim_plotter.add_mesh(dist_cartilage_mesh, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='遠位軟骨')
			
			# 遠位関節領域（球体抽出部分）を強調表示（不透明、色違い）
			# これによりユーザーはオフセットの設定が正しいか視覚的に確認できる
			dist_region_actor_mesh = dist_joint_region.copy()
			dist_region_actor = anim_plotter.add_mesh(dist_region_actor_mesh, color="#00FFFF", smooth_shading=True, show_edges=False, opacity=1.0, label='Extracted Distal Region')
			dist_region_actor.SetVisibility(False)  # デフォルトでは非表示にする（ユーザーの要望）

			# 遠位座標系の軸を表示
			dist_x_line = pv.Line(dist_origin, dist_origin + dist_x_axis * axis_length)
			dist_y_line = pv.Line(dist_origin, dist_origin + dist_y_axis * axis_length)
			dist_z_line = pv.Line(dist_origin, dist_origin + dist_z_axis * axis_length)
			dist_x_actor = anim_plotter.add_mesh(dist_x_line, color='red', line_width=3, label='Dist X')
			dist_y_actor = anim_plotter.add_mesh(dist_y_line, color='green', line_width=3, label='Dist Y')
			dist_z_actor = anim_plotter.add_mesh(dist_z_line, color='blue', line_width=3, label='Dist Z')
			
			# オーバーラップ体積メッシュは廃止

			# --- 抽出領域（関節領域）の表示 ---
			# 近位関節領域と遠位関節領域は削除（ヒートマップで代替）
			
			# 距離ヒートマップメッシュを初期化
			# ヒートマップデータがある場合は最初のフレームで初期化、ない場合はダミー
			print(f"[デバッグ] ヒートマップ初期化: len={len(heatmap_precomputed)}")
			if len(heatmap_precomputed) > 0:
				print(f"[デバッグ] heatmap_precomputed[0]: {heatmap_precomputed[0].n_points}点")
				if heatmap_precomputed[0].n_points > 0:
					heatmap_mesh = heatmap_precomputed[0].copy()
					# スカラー 'distance' が無ければダミーで埋める（キャッシュ/計算の不整合対策）
					if 'distance' not in heatmap_mesh.array_names:
						print(f"[警告] 最初のヒートマップに 'distance' スカラーがありません。ダミー値で埋めます。")
						try:
							n = heatmap_mesh.n_points
							heatmap_mesh['distance'] = np.zeros(n)
						except Exception:
							# それでも失敗する場合は空のダミーメッシュに置き換え
							heatmap_mesh = pv.PolyData()
							heatmap_mesh.points = np.array([[0, 0, 0]])
							heatmap_mesh['distance'] = np.array([0.0])
					print(f"[デバッグ] ヒートマップメッシュを最初のフレームで初期化: {heatmap_mesh.n_points}点")
					print(f"[デバッグ] スカラー: {heatmap_mesh.array_names}")
				else:
					heatmap_mesh = pv.PolyData()
					heatmap_mesh.points = np.array([[0, 0, 0]])
					heatmap_mesh['distance'] = np.array([0.0])
					print(f"[デバッグ] ヒートマップメッシュをダミーで初期化（0点のため）")
			else:
				heatmap_mesh = pv.PolyData()
				# 空のメッシュにダミーのスカラーを追加（エラー回避）
				heatmap_mesh.points = np.array([[0, 0, 0]])
				heatmap_mesh['distance'] = np.array([0.0])
				print(f"[デバッグ] ヒートマップメッシュをダミーで初期化（データなし）")
			
			# カラースケール（符号付き距離対応: 0を中心に対称）
			if len(heatmap_precomputed) > 0 and 'distance' in heatmap_precomputed[0].array_names:
				try:
					# 全フレームの距離範囲を走査してスケールを安定化
					mins = []
					maxs = []
					for hm in heatmap_precomputed:
						if hm is not None and hm.n_points > 0 and 'distance' in hm.array_names:
							d = np.asarray(hm['distance'])
							if d.size:
								mins.append(np.nanmin(d))
								maxs.append(np.nanmax(d))
					min_val = np.nanmin(mins) if mins else 0.0
					max_val = np.nanmax(maxs) if maxs else 1.0
					max_abs = float(max(abs(min_val), abs(max_val)))
					if max_abs <= 1e-6:
						max_abs = 1.0
				except Exception:
					max_abs = 5.0
			else:
				max_abs = 5.0

			heatmap_actor = anim_plotter.add_mesh(
				heatmap_mesh,
				scalars='distance',
				cmap='jet_r',  # 符号付き距離向けの発散カラーマップ
				clim=[-5, 5],  # 0を中心に対称範囲
				show_edges=False,
				opacity=1.0,  # 1.0にしてフィッティング側と合わせる
				label='Heatmap',
			)
			
			# Z-fighting対策：マッパーの設定でPolygonOffsetを有効化
			heatmap_mapper = heatmap_actor.GetMapper()
			heatmap_mapper.SetResolveCoincidentTopologyToPolygonOffset()
			heatmap_mapper.SetRelativeCoincidentTopologyPolygonOffsetParameters(0, -5) # 負の値で手前に引き寄せる

			# 境界線（0mm等高線）を近位抽出領域メッシュ表面に張り付けて表示
			print(f"[初期化] 境界線（等高線）の初期化を開始...")
			boundary_actor = None
			# 近位抽出領域メッシュに距離スカラーをマッピングして等高線抽出
			if len(heatmap_precomputed) > 0 and 'distance' in heatmap_precomputed[0].array_names and prox_joint_region.n_points > 0:
				try:
					print(f"[初期化] 近位メッシュをコピー中...")
					# 近位抽出領域メッシュに距離値をマッピング
					prox_with_distance = prox_joint_region.copy()
					
					print(f"[初期化] ヒートマップデータを取得中...")
					# ヒートマップの点と距離値を取得
					heatmap_points = heatmap_precomputed[0].points
					heatmap_distances = heatmap_precomputed[0]['distance']
					
					print(f"[初期化] KDTreeを構築中（{len(heatmap_points)}点）...")
					# 近位メッシュの各頂点に最近傍の距離値を割り当て
					from scipy.spatial import cKDTree
					tree = cKDTree(heatmap_points)
					
					print(f"[初期化] 最近傍探索を実行中（{prox_with_distance.n_points}点）...")
					_, indices = tree.query(prox_with_distance.points, k=1)
					prox_with_distance['distance'] = heatmap_distances[indices]
					print(f"[初期化] 距離マッピング完了")
					
					# メッシュ情報を表示
					print(f"[デバッグ] 近位メッシュ: {prox_with_distance.n_points}点, {prox_with_distance.n_cells}面")
					print(f"[デバッグ] 距離範囲: min={prox_with_distance['distance'].min():.3f}, max={prox_with_distance['distance'].max():.3f}")
					
					print(f"[初期化] 等高線（0mm）を抽出中...")
					# PyVistaのcontour機能で距離=0.0mmの等高線を面メッシュから抽出
					contour_mesh = prox_with_distance.contour(isosurfaces=[0.0], scalars='distance')
					print(f"[初期化] 等高線抽出完了")
					
					if contour_mesh.n_points > 0:
						print(f"[デバッグ] 等高線抽出成功: {contour_mesh.n_points}点, {contour_mesh.n_lines}線")
						# 境界線メッシュを作成（後で更新用）
						boundary_mesh = contour_mesh.copy()
					else:
						print(f"[警告] 等高線が抽出されませんでした（0mm付近の点がメッシュ上にない可能性）")
						boundary_mesh = pv.PolyData()
				except Exception as e:
					print(f"[警告] 等高線抽出失敗: {e}")
					# import traceback # 既にグローバルでインポート済み
					traceback.print_exc()
					boundary_mesh = pv.PolyData()
			else:
				boundary_mesh = pv.PolyData()
			
			print(f"[初期化] 境界線actorを作成中...")
			# 境界線actorを一度だけ作成（以降はboundary_meshのポイントを更新するだけで点滅しない）
			boundary_actor = anim_plotter.add_mesh(
				boundary_mesh,
				color='white',
				line_width=5,
				render_lines_as_tubes=True,
				lighting=False
			)
			
			print(f"[初期化] 境界線actor作成完了: type={type(boundary_actor)}")
			
			print(f"[初期化] 表示設定を構成中...")
			# ヒートマップがある場合は初期表示ON、ない場合は非表示
			if len(heatmap_precomputed) > 0:
				print(f"[デバッグ] ヒートマップデータあり: {len(heatmap_precomputed)}フレーム")
				print(f"[デバッグ] 最初のフレームのヒートマップ: {heatmap_precomputed[0].n_points}点")
				if 'distance' in heatmap_precomputed[0].array_names:
					print(f"[デバッグ] distanceスカラーあり")
				heatmap_actor.SetVisibility(True)
				boundary_actor.SetVisibility(True)
				print(f"[デバッグ] ヒートマップ表示をONに設定")
				print(f"[デバッグ] 境界線表示をONに設定（visibility={boundary_actor.GetVisibility()}）")
			else:
				print(f"[デバッグ] ヒートマップデータなし")
				heatmap_actor.SetVisibility(False)
				boundary_actor.SetVisibility(False)

			# --- FEM接触圧メッシュの初期化（ヒートマップと同じ形状、異なるスカラー） ---
			print(f"[初期化] FEM接触圧メッシュの初期化を開始...")
			pressure_mesh = None
			pressure_actor = None
			heatmap_display_mode = [0]  # 0=distance, 1=contact_pressure

			if fem_pressure_precomputed and any(x is not None for x in fem_pressure_precomputed):
				try:
					# 最初の有効なフレームを探す
					first_pressure = None
					first_summary = None
					for fp in fem_pressure_precomputed:
						if fp is not None:
							first_pressure, first_summary = fp
							break

					if first_pressure is not None:
						# ヒートマップと同じジオメトリを持つメッシュを作成
						pressure_mesh = heatmap_mesh.copy()

						# 圧力データをマッピング
						if prox_fem_surface is not None and prox_fem_surface.n_points > 0:
							if prox_fem_surface.n_points == pressure_mesh.n_points:
								pressure_mesh['contact_pressure'] = first_pressure.copy()
							else:
								from scipy.spatial import cKDTree
								tree = cKDTree(prox_fem_surface.points)
								_, idx = tree.query(pressure_mesh.points, k=1)
								pressure_mesh['contact_pressure'] = first_pressure[idx]
						else:
							pressure_mesh['contact_pressure'] = first_pressure.copy()

						# 全フレームの圧力範囲を走査してスケールを決定
						p_max = 0.0
						for fp in fem_pressure_precomputed:
							if fp is not None:
								p_max = max(p_max, float(np.nanmax(fp[0])))
						if p_max < 1e-6:
							p_max = 1.0

						print(f"[初期化] 圧力範囲: 0 - {p_max:.2f} MPa")

						pressure_actor = anim_plotter.add_mesh(
							pressure_mesh,
							scalars='contact_pressure',
							cmap='hot_r',
							clim=[0, p_max],
							show_edges=False,
							opacity=1.0,
							scalar_bar_args={'title': '接触圧 [MPa]', 'color': 'black'},
							label='FEM接触圧'
						)

						# Z-fighting対策
						p_mapper = pressure_actor.GetMapper()
						p_mapper.SetResolveCoincidentTopologyToPolygonOffset()
						p_mapper.SetRelativeCoincidentTopologyPolygonOffsetParameters(0, -6)

						pressure_actor.SetVisibility(False)
						print(f"[初期化] FEM接触圧メッシュを初期化: {pressure_mesh.n_points}点, p_max={p_max:.2f}MPa")
				except Exception as e:
					print(f"[警告] FEM接触圧メッシュの初期化失敗: {e}")
					traceback.print_exc()
					pressure_mesh = None
					pressure_actor = None
			else:
				print(f"[初期化] FEM事前計算データがないため、接触圧メッシュは作成されません")

			print(f"[初期化] ウィジェット（表示制御）を構成中...")
			# オーバーラップメッシュの枠線表示状態
			overlap_edges_visible = [True]  # デフォルトで枠線表示
			
			# --- ウィジェット: 表示/非表示と透明度 ---
			# 近位モデルの表示/非表示
			def toggle_prox_model(state):
				prox_actor.SetVisibility(state)
			
			# 近位座標系の表示/非表示
			def toggle_prox_axes(state):
				prox_x_actor.SetVisibility(state)
				prox_y_actor.SetVisibility(state)
				prox_z_actor.SetVisibility(state)
			
			# 遠位モデルの表示/非表示
			def toggle_dist_model(state):
				dist_actor.SetVisibility(state)
			
			# 遠位座標系の表示/非表示
			def toggle_dist_axes(state):
				dist_x_actor.SetVisibility(state)
				dist_y_actor.SetVisibility(state)
				dist_z_actor.SetVisibility(state)

			# ヒートマップの表示/非表示
			def toggle_heatmap(state):
				if heatmap_precomputed:  # 事前計算済みの場合のみ有効
					heatmap_actor.SetVisibility(state)
				else:
					heatmap_actor.SetVisibility(False)
			
			# 境界線（0mm等高線）の表示/非表示
			def toggle_boundary(state):
				if boundary_actor is not None:
					boundary_actor.SetVisibility(state)

			# FEM接触圧の表示切り替え
			def toggle_pressure_display(state):
				if state:
					# 圧力表示ON: distanceをOFF、pressureをON
					heatmap_actor.SetVisibility(False)
					if pressure_actor is not None:
						pressure_actor.SetVisibility(True)
					heatmap_display_mode[0] = 1
				else:
					# 圧力表示OFF: distanceをON（ヒートマップが事前計算されている場合）
					if heatmap_precomputed:
						heatmap_actor.SetVisibility(True)
					if pressure_actor is not None:
						pressure_actor.SetVisibility(False)
					heatmap_display_mode[0] = 0

			# 近位モデルの透明度
			def update_prox_opacity(value):
				prox_actor.GetProperty().SetOpacity(value)
			
			# 遠位モデルの透明度
			def update_dist_opacity(value):
				dist_actor.GetProperty().SetOpacity(value)
			
			# オーバーラップ体積の透明度制御は削除
			
			# チェックボックスとラベル - 近位
			anim_plotter.add_text("Prox Model", position=(10, 10), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_prox_model, value=True, position=(130, 12), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			
			anim_plotter.add_text("Prox Axes", position=(10, 40), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_prox_axes, value=True, position=(130, 42), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			
			# チェックボックスとラベル - 遠位
			anim_plotter.add_text("Dist Model", position=(10, 70), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_dist_model, value=True, position=(130, 72), size=20, border_size=1, color_on=self.dist_color, color_off="white")
			
			anim_plotter.add_text("Dist Axes", position=(10, 100), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_dist_axes, value=True, position=(130, 102), size=20, border_size=1, color_on=self.dist_color, color_off="white")

			# チェックボックスとラベル - ヒートマップ（事前計算されている場合のみ表示）
			heatmap_enabled = len(heatmap_precomputed) > 0
			if heatmap_enabled:
				anim_plotter.add_text("Heatmap", position=(10, 130), font_size=10, color="black")
				anim_plotter.add_checkbox_button_widget(
					toggle_heatmap, 
					value=True,  # 初期表示ON
					position=(130, 132), 
					size=20, 
					border_size=1, 
					color_on="orange", 
					color_off="white"
				)
				
				# チェックボックスとラベル - 境界線（0mm等高線）
				anim_plotter.add_text("Boundary", position=(10, 160), font_size=10, color="black")
				anim_plotter.add_checkbox_button_widget(
					toggle_boundary,
					value=True,  # 初期表示ON
					position=(130, 162),
					size=20,
					border_size=1,
					color_on="white",
					color_off="gray"
				)

			# チェックボックスとラベル - FEM接触圧（事前計算されている場合のみ表示）
			if fem_pressure_precomputed and any(x is not None for x in fem_pressure_precomputed):
				anim_plotter.add_text("FEM圧", position=(10, 190), font_size=10, color="black")
				anim_plotter.add_checkbox_button_widget(
					toggle_pressure_display,
					value=False,  # デフォルトで非表示
					position=(130, 192),
					size=20,
					border_size=1,
					color_on="purple",
					color_off="white"
				)
				next_y = 220
			else:
				next_y = 190
			
			# Overlap UIは削除済み
			
			# 透明度スライダー - 近位
			anim_plotter.add_slider_widget(
				update_prox_opacity,
				rng=[0.0, 1.0],
				value=0.5,
				title=" ",
				pointa=(0.10, 0.025),
				pointb=(0.20, 0.025),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.prox_color,
			)
			
			# 透明度スライダー - 遠位
			anim_plotter.add_slider_widget(
				update_dist_opacity,
				rng=[0.0, 1.0],
				value=0.5,
				title=" ",
				pointa=(0.10, 0.085),
				pointb=(0.20, 0.085),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.dist_color,
			)
			
			# 重複体積の透明度スライダーは削除
			
			# --- アニメーション設定 ---
			# 遠位メッシュの初期状態を保存（スライダーと同じロジック用）
			dist_mesh_for_anim = dist_mesh.copy()
			dist_origin_initial = np.array([0.0, 0.0, 0.0])  # 変換後の遠位座標系原点
			dist_x_axis_initial = np.array([1.0, 0.0, 0.0])
			dist_y_axis_initial = np.array([0.0, 1.0, 0.0])
			dist_z_axis_initial = np.array([0.0, 0.0, 1.0])

			# 遠位関節領域の初期状態を保存（メッシュと初期点群）
			dist_region_for_anim = dist_joint_region.copy()  # dist_region_meshではなくdist_joint_regionから
			
			# 現在のフレーム番号
			current_frame = [0]  # リストで包んで参照を保持（見かけのフレーム: Excelデータのインデックス）
			actual_frame_counter = [0]  # 実際のフレームカウンタ（5ms間隔で増加）
			animation_start_time = [None]  # アニメーション開始時刻（実時間）
			playback_speed = [1.0]  # 再生速度（1.0 = 等倍速）
			last_scale_update_time = [0.0]  # 最後に再生バーを更新した時刻
			scale_update_interval = [0.2]  # 再生バー更新間隔（秒）- YouTubeスタイル
			print(f"[初期化] アニメーション制御変数を初期化中...")
			after_id = [None]  # Tkinter afterのID（アニメーション用）
			check_id = [None]  # Tkinter afterのID（ウィンドウチェック用）
			is_animation_active = [True]  # アニメーションが有効かどうか
			is_paused = [False]  # 一時停止フラグ
			is_seeking = [False]  # シーク中フラグ（再生バー操作中）
			is_programmatic_update = [False]  # プログラムから再生バーを更新中かどうか
			user_is_dragging = [False]  # ユーザーがドラッグ中かどうか
			
			print(f"[初期化] 再生コントロールウィンドウを作成中...")
			# データ総数をログ出力
			print(f"[アニメーション] データ総数: {len(transform_data)} フレーム")
			
			# CSVの時間データを取得（各フレームのTime値）
			frame_times = [frame_data["time"] for frame_data in transform_data]
			print(f"[アニメーション] Time範囲: {frame_times[0]:.3f}s ~ {frame_times[-1]:.3f}s")
			
			# データが昇順に並んでいるか確認
			is_sorted = all(frame_times[i] <= frame_times[i+1] for i in range(len(frame_times)-1))
			print(f"[アニメーション] データは時間順: {is_sorted}")
			if not is_sorted:
				print("[警告] データが時間順に並んでいません！")
			
			# 最大時間を取得（ループ用）
			max_time = frame_times[-1]
			
			# 再生コントロール用のTkinterウィンドウを作成
			control_window = tk.Toplevel(self)
			control_window.title("再生コントロール")
			control_window.geometry("850x360")
			control_window.resizable(False, False)
			control_window.attributes('-topmost', True)  # 常に最前面に表示
			
			# 情報表示フレーム
			info_frame = ttk.Frame(control_window)
			info_frame.pack(pady=5, padx=10, fill=tk.X)
			
			# フレーム情報表示（横並び）
			visible_frame_label = ttk.Label(
				info_frame,
				text=f"Frame: 0/{len(transform_data)-1} | Time: {frame_times[0]:.3f}s",
				font=(self.ui_font_family, 10)
			)
			visible_frame_label.pack(side=tk.LEFT, padx=5)
			
			actual_time_label = ttk.Label(
				info_frame,
				text=f"Actual: 0.000s",
				font=(self.ui_font_family, 10),
				foreground="blue"
			)
			actual_time_label.pack(side=tk.LEFT, padx=5)
			
			# 再生速度表示
			speed_label = ttk.Label(
				info_frame,
				text=f"Speed: 1.0x",
				font=(self.ui_font_family, 10),
				foreground="darkgreen"
			)
			speed_label.pack(side=tk.LEFT, padx=5)
			
			# オーバーラップ面積表示（新規追加）
			overlap_area_label = ttk.Label(
				info_frame,
				text=f"Area: 0.00 mm²",
				font=(self.ui_font_family, 10),
				foreground="red"
			)
			if len(overlap_areas_precomputed) > 0:
				overlap_area_label.pack(side=tk.LEFT, padx=5)
			
			# 接触深度表示（新規追加）
			overlap_depth_label = ttk.Label(
				info_frame,
				text=f"Depth: 0.00 mm",
				font=(self.ui_font_family, 10),
				foreground="purple"
			)
			if len(overlap_depths_precomputed) > 0:
				overlap_depth_label.pack(side=tk.LEFT, padx=5)

			# 最大めり込み量表示（新規追加）
			max_penetration_label = ttk.Label(
				info_frame,
				text=f"Max Pent: 0.00 mm",
				font=(self.ui_font_family, 10),
				foreground="red"
			)
			max_penetration_label.pack(side=tk.LEFT, padx=5)
			
			# 再生バー（スライダー）
			# YouTubeスタイル: マウスイベントを追跡
			def on_scale_button_press(event):
				"""マウスボタンが押された（ドラッグ開始）"""
				user_is_dragging[0] = True
			
			def on_scale_button_release(event):
				"""マウスボタンが離された（ドラッグ終了）"""
				user_is_dragging[0] = False
			
			playback_scale = tk.Scale(
				control_window,
				from_=0,
				to=len(transform_data)-1,
				orient=tk.HORIZONTAL,
				length=810,
				label="Frame Position",
				command=lambda val: on_playback_scale_change(int(val))
			)
			playback_scale.pack(pady=5, padx=10)
			
			# マウスイベントをバインド
			playback_scale.bind("<ButtonPress-1>", on_scale_button_press)
			playback_scale.bind("<ButtonRelease-1>", on_scale_button_release)
			
			# 再生速度スライダー
			def on_speed_change(val):
				"""再生速度が変更されたときの処理"""
				speed = float(val)
				playback_speed[0] = speed
				speed_label.config(text=f"Speed: {speed:.2f}x")
				# 再生中の場合、開始時刻を調整
				if animation_start_time[0] is not None and not is_paused[0]:
					actual_time = actual_frame_counter[0] * 0.005
					animation_start_time[0] = time.time() - (actual_time / speed)
			
			speed_scale = tk.Scale(
				control_window,
				from_=0.25,
				to=10.0,
				resolution=0.25,
				orient=tk.HORIZONTAL,
				length=810,
				label="Playback Speed (0.25x - 10x)",
				command=on_speed_change
			)
			speed_scale.set(1.0)  # 初期値: 等倍速
			speed_scale.pack(pady=5, padx=10)
			
			# 一時停止/再生ボタン
			button_frame = ttk.Frame(control_window)
			button_frame.pack(pady=5)
			
			pause_button = ttk.Button(button_frame, text="一時停止", width=15)
			pause_button.pack(side=tk.LEFT, padx=5)
			
			# CSVエクスポートボタン（新規追加）
			def export_to_csv():
				"""オーバーラップ面積と接触深度データをCSVにエクスポート"""
				if len(overlap_areas_precomputed) == 0:
					messagebox.showwarning("エクスポート", "オーバーラップデータがありません。\n事前計算を実行してください。")
					return
				
				# ファイル保存ダイアログ
				filepath = filedialog.asksaveasfilename(
					title="CSVファイルを保存",
					defaultextension=".csv",
					filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
				)
				
				if not filepath:
					return  # キャンセルされた
				
				try:
					import csv
					with open(filepath, 'w', newline='', encoding='utf-8') as f:
						writer = csv.writer(f)
						# ヘッダー行
						writer.writerow(["IE (deg)", "Overlap Area (mm^2)", "Contact Depth (mm)"])
						
						# データ行
						for i, frame_data in enumerate(transform_data):
							ie_angle = frame_data["angles"][2]  # IE角度（Y軸回転）
							area = overlap_areas_precomputed[i] if i < len(overlap_areas_precomputed) else 0.0
							depth = overlap_depths_precomputed[i] if i < len(overlap_depths_precomputed) else 0.0
							writer.writerow([f"{ie_angle:.6f}", f"{area:.6f}", f"{depth:.6f}"])
					
					messagebox.showinfo("エクスポート完了", f"CSVファイルを保存しました:\n{filepath}")
				
				except Exception as e:
					messagebox.showerror("エクスポート失敗", f"CSVファイルの保存に失敗しました:\n{e}")
			
			export_button = ttk.Button(button_frame, text="CSV出力", width=15, command=export_to_csv)
			if len(overlap_areas_precomputed) > 0:
				export_button.pack(side=tk.LEFT, padx=5)

			# --- このモデルを出力 ボタン (新規機能) ---
			def export_current_frame_model():
				"""現在のフレームの近位・遠位（変換後）モデルを1つのファイルに出力"""
				# 一時停止してから保存処理に入る
				current_was_paused = is_paused[0]
				is_paused[0] = True
				pause_button.config(text="再生")
				
				try:
					# 現在のフレームインデックス
					idx = current_frame[0]
					
					# 保存先を選択
					save_path = filedialog.asksaveasfilename(
						title=f"現在のモデルを出力 (Frame {idx})",
						defaultextension=".stl",
						filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj")],
						initialfile=f"model_frame_{idx}.stl"
					)
					
					if not save_path:
						# キャンセルされたら再生状態を戻す
						if not current_was_paused:
							toggle_pause()
						return
					
					# 1. 近位メッシュのコピー（変更不要）
					prox_copy = prox_mesh.copy()
					
					# 2. 遠位メッシュは既に変換済み（anim_plotter内で変換されている）
					# しかし、PyVista上のMeshは `dist_mesh.points` が `show_frame` で更新されているので
					# 現在の `dist_mesh` をそのまま使えばよい
					dist_copy = dist_mesh.copy()
					
					# 3. 2つのメッシュを結合
					merged_mesh = prox_copy + dist_copy
					
					# 4. 保存
					merged_mesh.save(save_path)
					
					messagebox.showinfo("出力完了", f"モデルを出力しました:\nFrame: {idx}\nPath: {save_path}")
					
				except Exception as e:
					messagebox.showerror("エラー", f"モデル出力中にエラーが発生しました:\n{e}\n{traceback.format_exc()}")
				
				# 再生状態を戻す（ただし、明示的に止めたい場合もあるので、ユーザーが再開ボタンを押す運用でもよいが、ここでは元の状態に戻す）
				if not current_was_paused:
					toggle_pause()

			export_model_button = ttk.Button(button_frame, text="このモデルを出力", width=18, command=export_current_frame_model)
			export_model_button.pack(side=tk.LEFT, padx=5)

			# --- FEM接触解析ボタン（関節領域がある場合に表示）---
			if _HAS_FEM and prox_joint_region.n_points > 0 and dist_joint_region.n_points > 0:
				def run_fem_current_frame():
					"""現在フレームの姿勢でFEM接触解析を実行（関節領域で計算）"""
					was_paused = is_paused[0]
					is_paused[0] = True
					pause_button.config(text="再生")
					# 現在フレームの遠位関節領域を変換
					frame_idx = current_frame[0] if hasattr(current_frame, '__getitem__') else 0
					matrix = transform_data[frame_idx]['matrix']
					dist_joint_current = self._apply_transform(dist_joint_region.copy(), matrix)
					prox_fem = prox_joint_region.extract_surface() if hasattr(prox_joint_region, 'extract_surface') else prox_joint_region
					dist_fem = dist_joint_current.extract_surface() if hasattr(dist_joint_current, 'extract_surface') else dist_joint_current
					self._run_fem_on_current_frame(
						prox_fem, dist_fem,
						prox_bone_mesh=prox_mesh, dist_bone_mesh=dist_mesh,
					)
					if not was_paused:
						toggle_pause()

				fem_anim_btn = ttk.Button(button_frame, text="FEM解析（関節領域）", width=18, command=run_fem_current_frame)
				fem_anim_btn.pack(side=tk.LEFT, padx=5)
			# --------------------------------------------
			
			def on_playback_scale_change(frame_idx):
				"""再生バースライダーが操作されたときの処理（YouTubeスタイル）"""
				# プログラムからの更新の場合は何もしない（ユーザー操作のみ処理）
				if is_programmatic_update[0]:
					return
				
				# ユーザーがドラッグ中の場合のみシークする
				if user_is_dragging[0]:
					is_seeking[0] = True
					current_frame[0] = frame_idx
					show_frame(frame_idx, force_render=True)
					# 見かけのフレーム表示を更新
					time_val = frame_times[frame_idx]
					visible_frame_label.config(text=f"Frame: {frame_idx}/{len(transform_data)-1} | Time: {time_val:.3f}s")
					# シーク時に実際の時間もリセット
					actual_frame_counter[0] = int(time_val / 0.005)
					actual_time = actual_frame_counter[0] * 0.005
					actual_time_label.config(text=f"Actual: {actual_time:.3f}s")
					# アニメーション開始時刻もリセット（再生中の場合、再生速度を考慮）
					if animation_start_time[0] is not None:
						animation_start_time[0] = time.time() - (actual_time / playback_speed[0])
					is_seeking[0] = False
			
			def toggle_pause():
				"""一時停止/再生を切り替え"""
				is_paused[0] = not is_paused[0]
				if is_paused[0]:
					pause_button.config(text="再生")
				else:
					pause_button.config(text="一時停止")
					# 再生再開時に実時間をリセット（再生速度を考慮）
					actual_time = actual_frame_counter[0] * 0.005
					animation_start_time[0] = time.time() - (actual_time / playback_speed[0])
			
			pause_button.config(command=toggle_pause)
			
			# ウィンドウが閉じられたときの処理
			def on_control_window_close():
				is_animation_active[0] = False
				control_window.destroy()
			
			control_window.protocol("WM_DELETE_WINDOW", on_control_window_close)
			
			print(f"[初期化] アニメーション関数を定義中...")
			# 再生バー用の変数（PyVistaスライダーは削除）
			frame_text_actor = [None]  # 見かけのフレーム（CSVのTime値）
			actual_frame_text_actor = [None]  # 実際のフレーム（0.005秒ごと）
			six_axis_text_actor = [None]  # 6軸データ表示用
			
			def show_frame(frame_idx, force_render=False):
				"""指定されたフレームを表示"""
				nonlocal boundary_actor, boundary_mesh  # 外側のboundary_actorとboundary_meshを参照
				
				# アニメーションが無効化されているかチェック（force_render時は無視）
				if not force_render and not is_animation_active[0]:
					return
				
				# ウィンドウが閉じられていないかチェック
				if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
					is_animation_active[0] = False
					return
				
				# render windowが有効かチェック
				if hasattr(anim_plotter, 'render_window'):
					try:
						if not anim_plotter.render_window:
							is_animation_active[0] = False
							return
					except:
						is_animation_active[0] = False
						return
				
				try:
					if frame_idx >= len(transform_data):
						frame_idx = 0  # ループ再生
					
					# 現在のフレームのデータを取得
					try:
						frame_data = transform_data[frame_idx]
						time_val = frame_data["time"]
						angles = frame_data["angles"]  # [FE, VV, IE]
						trans = frame_data["translations"]  # [ML, AP, PD]
					except Exception as data_err:
						print(f"[エラー] フレーム{frame_idx}のデータ取得失敗: {data_err}")
						return
					
					# 6軸データをPyVistaウィンドウに表示（右上）
					six_axis_text = (
						f"FE: {angles[0]:+7.2f}°\n"
						f"VV: {angles[1]:+7.2f}°\n"
						f"IE: {angles[2]:+7.2f}°\n"
						f"ML: {trans[0]:+7.2f} mm\n"
						f"AP: {trans[1]:+7.2f} mm\n"
						f"PD: {trans[2]:+7.2f} mm"
					)
					
					# 初回のみactor作成、以降はテキスト更新のみ（点滅防止）
					if six_axis_text_actor[0] is None:
						try:
							six_axis_text_actor[0] = anim_plotter.add_text(
								six_axis_text,
								position='upper_right',
								font_size=12,
								color='black',
								font='courier'
							)
						except:
							pass
					else:
						# actorが既に存在する場合、テキストだけ更新
						try:
							six_axis_text_actor[0].SetText(3, six_axis_text)
						except:
							pass
					
					# 同次変換行列を構築（スライダーと同じロジック）
					matrix = self._build_transform_matrix(
						rz=angles[0],  # FE
						rx=angles[1],  # VV
						ry=angles[2],  # IE
						ml=trans[0],   # ML
						ap=trans[1],   # AP
						pd=trans[2]    # PD
					)
					
					# 遠位メッシュを変換（スライダーと同じロジック）
					# 1. 遠位座標系の原点を中心に配置
					centered_points = dist_mesh_for_anim.points.copy() - dist_origin_initial
					
					# 2. 同次座標に変換
					ones = np.ones((centered_points.shape[0], 1))
					points_homogeneous = np.hstack([centered_points, ones])
					
					# 3. 同次変換行列を適用
					transformed_homogeneous = (matrix @ points_homogeneous.T).T
					transformed_points = transformed_homogeneous[:, :3]
					
					# 4. 原点に戻す
					transformed_points = transformed_points + dist_origin_initial
					
					# 5. メッシュの頂点を更新
					dist_mesh.points = transformed_points

					# 5b. 遠位軟骨メッシュも同じ変換を適用
					if dist_cartilage_for_anim is not None and dist_cartilage_mesh is not None:
						cart_centered = dist_cartilage_for_anim.points.copy() - dist_origin_initial
						cart_ones = np.ones((cart_centered.shape[0], 1))
						cart_homo = np.hstack([cart_centered, cart_ones])
						cart_trans = (matrix @ cart_homo.T).T[:, :3] + dist_origin_initial
						dist_cartilage_mesh.points = cart_trans

					# --- 遠位関節領域（抽出球）の更新（オフセット確認用）---
					# 初期位置から計算
					if dist_region_for_anim.n_points > 0:
						region_centered = dist_region_for_anim.points.copy() - dist_origin_initial
						region_homo = np.hstack([region_centered, np.ones((region_centered.shape[0], 1))])
						region_trans_homo = (matrix @ region_homo.T).T
						region_trans_points = region_trans_homo[:, :3] + dist_origin_initial
						dist_region_actor_mesh.points = region_trans_points

					# 遠位座標系の軸を更新
					# 原点を変換
					origin_centered = np.array([0, 0, 0, 1])
					transformed_origin = (matrix @ origin_centered)[:3]
					transformed_origin = transformed_origin + dist_origin_initial
					
					# 軸を変換
					rotation_matrix = matrix[:3, :3]
					transformed_x_axis = rotation_matrix @ dist_x_axis_initial
					transformed_y_axis = rotation_matrix @ dist_y_axis_initial
					transformed_z_axis = rotation_matrix @ dist_z_axis_initial
					
					axis_length = 50.0
					dist_x_line.points = np.array([transformed_origin, transformed_origin + transformed_x_axis * axis_length])
					dist_y_line.points = np.array([transformed_origin, transformed_origin + transformed_y_axis * axis_length])
					dist_z_line.points = np.array([transformed_origin, transformed_origin + transformed_z_axis * axis_length])
					
					# 重複体積の更新ロジックは削除
					
					# ヒートマップの更新（事前計算済みの場合、可視性に関係なく常に更新）
					if heatmap_precomputed and frame_idx < len(heatmap_precomputed):
						try:
							heatmap_data = heatmap_precomputed[frame_idx]
						except Exception as e:
							print(f"[エラー] フレーム{frame_idx}: ヒートマップデータ取得失敗: {e}")
							return
						
						if heatmap_data is not None and heatmap_data.n_points > 0:
							# distanceスカラーの存在確認
							if 'distance' in heatmap_data.array_names:
								# デバッグ: 最初のフレームのみログ出力
								if frame_idx == 0:
									print(f"[デバッグ] フレーム0のヒートマップ更新: {heatmap_data.n_points}点")
									distances = heatmap_data['distance']
									print(f"[デバッグ] 距離範囲: min={distances.min():.2f}, max={distances.max():.2f}")
								
								# メッシュを安全に更新
								try:
									# shallow_copyを使用してメッシュを更新
									if hasattr(heatmap_mesh, 'copy_structure'):
										heatmap_mesh.copy_structure(heatmap_data)
										heatmap_mesh.point_data['distance'] = heatmap_data['distance']
									else:
										# copy_structureが使えない場合は直接代入
										heatmap_mesh.points = heatmap_data.points.copy()
										heatmap_mesh['distance'] = heatmap_data['distance'].copy()
										if hasattr(heatmap_data, 'faces') and heatmap_data.faces is not None:
											if hasattr(heatmap_mesh, 'faces'):
												heatmap_mesh.faces = heatmap_data.faces.copy()
										
								except Exception as hm_err:
									if frame_idx == 0:
										print(f"[エラー] フレーム{frame_idx}: メッシュ更新失敗: {hm_err}")
									# 更新失敗時はダミーデータ
									try:
										heatmap_mesh.points = np.array([[0, 0, 0]])
										heatmap_mesh['distance'] = np.array([0.0])
									except:
										pass
								
								# 境界線（0mm等高線）を毎フレーム更新（actorの可視性制御で点滅防止）
								try:
									# 近位抽出領域メッシュに距離値をマッピング
									prox_with_distance = prox_joint_region.copy()
									
									# ヒートマップの点と距離値を取得
									heatmap_points = heatmap_data.points
									heatmap_distances = heatmap_data['distance']
									
									# 近位メッシュの各頂点に最近傍の距離値を割り当て
									from scipy.spatial import cKDTree
									tree = cKDTree(heatmap_points)
									_, indices = tree.query(prox_with_distance.points, k=1)
									prox_with_distance['distance'] = heatmap_distances[indices]
									
									# PyVistaのcontour機能で距離=0.0mmの等高線を抽出
									contour_mesh = prox_with_distance.contour(isosurfaces=[0.0], scalars='distance')
									
									if contour_mesh.n_points > 0:
										# boundary_meshを更新（actorはそのまま、メッシュだけ更新で点滅防止）
										try:
											# shallow_copyまたはcopy_structureで安全に更新
											if hasattr(boundary_mesh, 'copy_structure'):
												boundary_mesh.copy_structure(contour_mesh)
											else:
												boundary_mesh.points = contour_mesh.points.copy()
												if hasattr(contour_mesh, 'lines') and contour_mesh.lines is not None:
													boundary_mesh.lines = contour_mesh.lines.copy()
											# actorを表示
											if boundary_actor is not None:
												try:
													boundary_actor.SetVisibility(True)
												except:
													pass
										except Exception as update_err:
											if frame_idx == 0:
												print(f"[エラー] フレーム{frame_idx}: boundary_mesh更新失敗: {update_err}")
											# 更新失敗時はactorを非表示に
											if boundary_actor is not None:
												try:
													boundary_actor.SetVisibility(False)
												except:
													pass
									else:
										# 等高線がない場合はactorを非表示に（メッシュは触らない）
										if boundary_actor is not None:
											try:
												boundary_actor.SetVisibility(False)
											except:
												pass
								except Exception as e:
									if frame_idx == 0:
										print(f"[エラー] フレーム{frame_idx}: 境界線抽出失敗: {e}")
									# 等高線抽出失敗時はactorを非表示に
									if boundary_actor is not None:
										try:
											boundary_actor.SetVisibility(False)
										except:
											pass
							else:
								# distanceスカラーがない場合
								if frame_idx == 0:
									print(f"[警告] フレーム0のヒートマップにdistanceスカラーがありません: {heatmap_data.array_names}")
								# 空のメッシュの場合はダミーデータで初期化
								try:
									heatmap_mesh.points = np.array([[0, 0, 0]])
									heatmap_mesh['distance'] = np.array([0.0])
								except:
									pass
								# 境界線を非表示に
								if boundary_actor is not None:
									try:
										boundary_actor.SetVisibility(False)
									except:
										pass
						else:
							# 空のメッシュの場合はダミーデータで初期化
							try:
								heatmap_mesh.points = np.array([[0, 0, 0]])
								heatmap_mesh['distance'] = np.array([0.0])
							except:
								pass
							# 境界線を非表示に
							if boundary_actor is not None:
								try:
									boundary_actor.SetVisibility(False)
								except:
									pass
					else:
						# ヒートマップデータがない場合
						if frame_idx == 0:
							print(f"[デバッグ] フレーム0: heatmap_precomputedなし（len={len(heatmap_precomputed) if heatmap_precomputed else 0}）")
						# 境界線を非表示に
						if boundary_actor is not None:
							try:
								boundary_actor.SetVisibility(False)
							except:
								pass
					
					# FEM接触圧メッシュの更新（毎フレーム更新）
					if fem_pressure_precomputed and frame_idx < len(fem_pressure_precomputed) and pressure_mesh is not None:
						try:
							fem_data = fem_pressure_precomputed[frame_idx]
							if fem_data is not None:
								pressure_arr, summary = fem_data
								# pressure_meshはheatmap_meshと同じ点を持つので、点を更新
								pressure_mesh.points = heatmap_mesh.points.copy()
								# 圧力値をマッピング
								if prox_fem_surface is not None:
									if len(pressure_arr) == pressure_mesh.n_points:
										pressure_mesh['contact_pressure'] = pressure_arr.copy()
									else:
										from scipy.spatial import cKDTree
										tree = cKDTree(prox_fem_surface.points)
										_, idx = tree.query(pressure_mesh.points, k=1)
										pressure_mesh['contact_pressure'] = pressure_arr[idx]
								else:
									pressure_mesh['contact_pressure'] = pressure_arr.copy()
						except Exception as e:
							if frame_idx == 0:
								print(f"[エラー] フレーム{frame_idx}: FEM接触圧更新失敗: {e}")
				
					# 面積表示を更新（新規追加）
					if len(overlap_areas_precomputed) > 0 and frame_idx < len(overlap_areas_precomputed):
						current_area = overlap_areas_precomputed[frame_idx]
						try:
							overlap_area_label.config(text=f"Area: {current_area:.2f} mm²")
						except:
							pass  # ウィンドウが閉じられている場合
					
					# 深度表示を更新（新規追加）
					if len(overlap_depths_precomputed) > 0 and frame_idx < len(overlap_depths_precomputed):
						current_depth = overlap_depths_precomputed[frame_idx]
						try:
							overlap_depth_label.config(text=f"Depth: {current_depth:.2f} mm")
						except:
							pass  # ウィンドウが閉じられている場合

					# 最大めり込み量表示を更新（ヒートマップから計算）
					if heatmap_precomputed and frame_idx < len(heatmap_precomputed):
						try:
							hm = heatmap_precomputed[frame_idx]
							if hm is not None and hm.n_points > 0 and 'distance' in hm.array_names:
								d = hm['distance']
								# 符号付き距離: 負の値がめり込み
								# 最小値（負の最大値）を取得し、負の場合のみ絶対値を表示
								min_dist = np.nanmin(d)
								max_penetration = abs(min_dist) if min_dist < 0 else 0.0
								max_penetration_label.config(text=f"Max Pent: {max_penetration:.2f} mm")
							else:
								max_penetration_label.config(text=f"Max Pent: -- mm")
						except Exception as e:
							print(f"[警告] めり込み量計算エラー: {e}")
							try:
								max_penetration_label.config(text=f"Max Pent: -- mm")
							except:
								pass
					

					# PyVistaの描画を更新（強制レンダリングまたはアクティブ時）
					if force_render or is_animation_active[0]:
						try:
							if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
								is_animation_active[0] = False
								return
							if hasattr(anim_plotter, 'render_window') and anim_plotter.render_window:
								try:
									anim_plotter.render()
								except Exception as render_err:
									# OpenGLエラーを無視（ログのみ出力）
									if frame_idx == 0:
										print(f"[エラー] フレーム{frame_idx}: レンダリング失敗: {render_err}")
									if not force_render:
										is_animation_active[0] = False
										if after_id[0] is not None:
											try:
												self.after_cancel(after_id[0])
											except:
												pass
											after_id[0] = None
							else:
								# render_windowが無効になった
								is_animation_active[0] = False
								return
						except Exception as check_err:
							# チェック自体が失敗
							is_animation_active[0] = False
							return
				except Exception as e:
					# エラー内容を詳細に表示
					print(f"[致命的エラー] フレーム{frame_idx}でshow_frame()がクラッシュ: {e}")
					import traceback
					traceback.print_exc()
					if not force_render:
						is_animation_active[0] = False
						if after_id[0] is not None:
							try:
								self.after_cancel(after_id[0])
							except:
								pass
							after_id[0] = None
			
			def animation_loop():
				"""アニメーションループ（実時間に同期して5ms間隔で進める）"""
				# デバッグ: ループが呼ばれていることを確認
				# print("[DEBUG] animation_loop() called")
				
				# アニメーションが無効化されているかチェック
				if not is_animation_active[0]:
					print("[アニメーション] is_animation_active is False, stopping")
					if after_id[0] is not None:
						self.after_cancel(after_id[0])
						after_id[0] = None
					return
				
				# ウィンドウが閉じられたかチェック
				if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
					is_animation_active[0] = False
					if after_id[0] is not None:
						self.after_cancel(after_id[0])
						after_id[0] = None
					print("[アニメーション] ウィンドウが閉じられました")
					return
				
				# 一時停止中またはシーク中は更新しない
				if not is_paused[0] and not is_seeking[0]:
					# アニメーション開始時刻を記録（初回のみ）
					if animation_start_time[0] is None:
						animation_start_time[0] = time.time()
					
					# 実際の経過時間を計算し、再生速度を適用
					elapsed_real_time = time.time() - animation_start_time[0]
					elapsed_animation_time = elapsed_real_time * playback_speed[0]  # 再生速度を適用
					
					# 経過時間から実際のフレームカウンタを計算（5ms単位）
					actual_frame_counter[0] = int(elapsed_animation_time / 0.005)
					actual_time = actual_frame_counter[0] * 0.005
					
					# デバッグ: 時間の詳細を表示（最初の10秒間のみ）
					# if elapsed_real_time < 10.0 and actual_frame_counter[0] % 20 == 0:
					#     print(f"[TIME] real={elapsed_real_time:.3f}s, anim={elapsed_animation_time:.3f}s, actual={actual_time:.3f}s, speed={playback_speed[0]:.2f}x")
					
					# 最大時間を超えたらループ
					if actual_time > max_time:
						actual_frame_counter[0] = 0
						actual_time = 0.0
						animation_start_time[0] = time.time()  # リセット
						print("[アニメーション] ループ再生")
					
					# 実際の時間表示を更新
					try:
						actual_time_label.config(text=f"Actual: {actual_time:.3f}s")
					except:
						pass
					
					# 実際の時間に最も近いExcelデータのフレームを見つける（改善版）
					# 逆走を防ぐため、現在のフレームより前には戻らないようにする
					target_frame = current_frame[0]  # デフォルトは現在のフレーム
					
					# 現在のフレームから先を検索（前進のみ）
					for i in range(current_frame[0], len(frame_times)):
						if frame_times[i] <= actual_time:
							target_frame = i
						else:
							# 次のフレームの時間を超えたら停止
							break
					
					# もし適切なフレームが見つからなかった場合（actual_timeが最初のフレームより前）
					# その場合のみ、最初から検索
					if target_frame == current_frame[0] and current_frame[0] > 0:
						if actual_time < frame_times[current_frame[0]]:
							# actual_timeが現在のフレームより前の場合のみ、0から検索
							target_frame = 0
							for i in range(len(frame_times)):
								if frame_times[i] <= actual_time:
									target_frame = i
								else:
									break
					
					# デバッグ: 逆走が発生した場合のみログ出力
					if target_frame < current_frame[0]:
						print(f"[WARNING] 逆走検出: Frame {current_frame[0]} -> {target_frame}")
						print(f"  actual_time={actual_time:.3f}s, current_frame_time={frame_times[current_frame[0]]:.3f}s, target_frame_time={frame_times[target_frame]:.3f}s")
						print(f"  elapsed_real_time={elapsed_real_time:.3f}s, speed={playback_speed[0]:.2f}x")
					
					# フレームが変わった場合のみ更新
					if target_frame != current_frame[0]:
						current_frame[0] = target_frame
						show_frame(current_frame[0])
						# 見かけのフレーム表示を更新
						try:
							time_val = frame_times[current_frame[0]]
							visible_frame_label.config(text=f"Frame: {current_frame[0]}/{len(transform_data)-1} | Time: {time_val:.3f}s")
						except:
							pass  # ウィンドウが閉じられている場合
					
					# 再生バーは200msごとにのみ更新（YouTubeスタイル: ドラッグ中は更新しない）
					current_time = time.time()
					if not user_is_dragging[0] and (current_time - last_scale_update_time[0] >= scale_update_interval[0]):
						try:
							is_programmatic_update[0] = True
							playback_scale.set(current_frame[0])
							is_programmatic_update[0] = False
							last_scale_update_time[0] = current_time
						except:
							pass
				
				# 次のチェックを予約（5msごと）
				after_id[0] = self.after(5, animation_loop)
			
			def stop_animation():
				"""アニメーションを停止"""
				print("[アニメーション] 停止します")
				is_animation_active[0] = False
				
				# タイマーをキャンセル
				if after_id[0] is not None:
					try:
						self.after_cancel(after_id[0])
					except:
						pass
					after_id[0] = None
				
				# ウィンドウを閉じる（安全に）
				try:
					if hasattr(anim_plotter, 'close'):
						# 既に閉じられていないかチェック
						if not (hasattr(anim_plotter, 'closed') and anim_plotter.closed):
							# render_windowが有効かチェック
							if hasattr(anim_plotter, 'render_window') and anim_plotter.render_window:
								anim_plotter.close()
				except Exception as e:
					print(f"[アニメーション] クローズエラー（無視）: {e}")
			
			# ウィンドウのクローズコールバックを設定
			def on_window_close():
				"""ウィンドウが閉じられたときの処理"""
				print("[アニメーション] ウィンドウクローズイベント")
				is_animation_active[0] = False
				if after_id[0] is not None:
					try:
						self.after_cancel(after_id[0])
					except:
						pass
					after_id[0] = None
				if check_id[0] is not None:
					try:
						self.after_cancel(check_id[0])
					except:
						pass
					check_id[0] = None
			
			# カメラ位置の設定
			anim_plotter.camera_position = 'iso'
			anim_plotter.reset_camera()
			
			# 最初のフレームを表示
			show_frame(0)
			
			# 一時停止/再生ボタン（Pキー）
			anim_plotter.add_key_event('p', toggle_pause)
			anim_plotter.add_key_event('P', toggle_pause)
			
			# アニメーション開始
			print("[アニメーション] 開始します（実時間に同期、5ms間隔で更新）")
			print("[アニメーション] ウィンドウの×ボタンまたはQキーで終了してください")
			print("[アニメーション] Pキーで一時停止/再生")
			
			# ウィンドウクローズ時のコールバックを設定
			def on_window_close():
				"""ウィンドウが閉じられた時に呼ばれるコールバック"""
				print("[アニメーション] ウィンドウが閉じられました。タイマーを停止します。")
				is_animation_active[0] = False
				if after_id[0] is not None:
					try:
						self.after_cancel(after_id[0])
					except:
						pass
					after_id[0] = None
				if check_id[0] is not None:
					try:
						self.after_cancel(check_id[0])
					except:
						pass
					check_id[0] = None
				# 再生コントロールウィンドウも閉じる
				try:
					control_window.destroy()
				except:
					pass
			
			# ウィンドウが閉じられた時のコールバックを登録
			try:
				anim_plotter.iren.add_observer('ExitEvent', lambda obj, event: on_window_close())
			except:
				pass  # コールバック登録に失敗しても続行
			
			print(f"[初期化] PyVistaウィンドウを表示中...")
			# ウィンドウを表示（ノンブロッキング）してからアニメーション開始
			try:
				anim_plotter.show(auto_close=False, interactive_update=True)
			except TypeError:
				anim_plotter.show(auto_close=False)
			
			print(f"[初期化完了] アニメーションループを開始します")
			# ウィンドウが表示されたのでアニメーションループを開始
			after_id[0] = self.after(5, animation_loop)
			
		except FileNotFoundError as e:
			messagebox.showerror("ファイルエラー", f"ファイルが見つかりません:\n{e}")
		except Exception as e:
			messagebox.showerror("エラー", f"アニメーション実行中にエラーが発生しました:\n{e}\n\n{traceback.format_exc()}")

	# ----- State control -----
	def update_button_states(self) -> None:
		# 可視化条件
		all_files_selected = (
			bool(self.prox_model_path.get())
			and bool(self.prox_pp_abcd_path.get())
			and bool(self.prox_pp_olmn_path.get())
			and bool(self.dist_model_path.get())
			and bool(self.dist_pp_abc_path.get())
			and bool(self.dist_pp_olmn_path.get())
		)
		joint_selected = self.joint_var.get() in (1, 2)
		# 近位のみ
		prox_ready = joint_selected and bool(self.prox_model_path.get()) and bool(self.prox_pp_abcd_path.get()) and bool(self.prox_pp_olmn_path.get())
		self.visualize_prox_button.state(["!disabled"] if prox_ready else ["disabled"])
		# 遠位のみ
		dist_ready = joint_selected and bool(self.dist_model_path.get()) and bool(self.dist_pp_abc_path.get()) and bool(self.dist_pp_olmn_path.get())
		self.visualize_dist_button.state(["!disabled"] if dist_ready else ["disabled"])
		# 全体
		all_ready = joint_selected and all_files_selected
		self.visualize_all_button.state(["!disabled"] if all_ready else ["disabled"])

		# 同次変換行列群の確認: 同ファイル選択済み
		check_enabled = bool(self.transform_group_path.get())
		self.check_tf_button.state(["!disabled"] if check_enabled else ["disabled"])

		# アニメーション実行: 4ファイル + 同次変換行列群 選択済み
		animate_enabled = all_files_selected and check_enabled
		self.animate_button.state(["!disabled"] if animate_enabled else ["disabled"])

	def update_fitting_button_states(self) -> None:
		"""Fittingタブのボタン状態を更新"""
		# 各ファイルの選択状態
		parent_model_selected = bool(self.fitting_parent_model_path.get())
		parent_region_selected = bool(self.fitting_parent_region_path.get())
		child_model_selected = bool(self.fitting_child_model_path.get())
		child_region_selected = bool(self.fitting_child_region_path.get())
		
		# 統合可視化ボタン（どちらかのファイルが選択されていればOK）
		parent_viz_enabled = parent_model_selected or parent_region_selected
		self.fitting_viz_parent_button.state(["!disabled"] if parent_viz_enabled else ["disabled"])
		
		child_viz_enabled = child_model_selected or child_region_selected
		self.fitting_viz_child_button.state(["!disabled"] if child_viz_enabled else ["disabled"])

		# フィッティング実行: 4つのファイルすべて選択済み
		all_fitting_files_selected = (
			parent_model_selected
			and parent_region_selected
			and child_model_selected
			and child_region_selected
		)
		self.fitting_execute_button.state(["!disabled"] if all_fitting_files_selected else ["disabled"])

	def update_org_button_states(self) -> None:
		"""ORGタブのボタン状態を更新"""
		model_selected = bool(self.org_model_path.get())
		pp_selected = bool(self.org_pp_path.get())
		
		# 両方のファイルが選択されている場合のみ可視化ボタンを有効化
		viz_enabled = model_selected and pp_selected
		self.org_visualize_button.state(["!disabled"] if viz_enabled else ["disabled"])

	def update_a_b_button_states(self) -> None:
		"""A-Bタブのボタン状態を更新"""
		model_a_selected = bool(self.a_b_model_a_path.get())
		model_b_selected = bool(self.a_b_model_b_path.get())
		viz_enabled = model_a_selected and model_b_selected
		self.a_b_visualize_button.state(["!disabled"] if viz_enabled else ["disabled"])
		self.a_b_diff_button.state(["!disabled"] if viz_enabled else ["disabled"])

	def update_heatmap_button_states(self) -> None:
		"""HeatMapタブのボタン状態を更新"""
		model_a_selected = bool(self.heatmap_model_a_path.get())
		model_b_selected = bool(self.heatmap_model_b_path.get())
		viz_enabled = model_a_selected and model_b_selected
		self.heatmap_visualize_button.state(["!disabled"] if viz_enabled else ["disabled"])
		self.heatmap_execute_button.state(["!disabled"] if viz_enabled else ["disabled"])

	def update_c_fix_button_states(self) -> None:
		"""C fixタブのボタン状態を更新"""
		# Simulatorタブのファイルがすべて選択されているか確認
		all_files_selected = (
			bool(self.prox_model_path.get())
			and bool(self.prox_pp_abcd_path.get())
			and bool(self.prox_pp_olmn_path.get())
			and bool(self.dist_model_path.get())
			and bool(self.dist_pp_abc_path.get())
			and bool(self.dist_pp_olmn_path.get())
		)
		self.c_fix_visualize_button.state(["!disabled"] if all_files_selected else ["disabled"])
		
		# 近位ファイルが選択されているか確認
		prox_files_selected = (
			bool(self.prox_pp_abcd_path.get())
			and bool(self.prox_pp_olmn_path.get())
		)
		self.c_fix_transform_prox_button.state(["!disabled"] if prox_files_selected else ["disabled"])
		
		# 遠位ファイルが選択されているか確認
		dist_files_selected = (
			bool(self.dist_pp_abc_path.get())
			and bool(self.dist_pp_olmn_path.get())
		)
		self.c_fix_transform_dist_button.state(["!disabled"] if dist_files_selected else ["disabled"])

	# ----- Persistence -----
	def _state_file_path(self) -> Path:
		return Path(__file__).with_name("frs2015_gui_state.json")

	def _load_state(self) -> None:
		try:
			p = self._state_file_path()
			if not p.exists():
				return
			data = json.load(p.open("r", encoding="utf-8"))
		except Exception:
			return

		# 反映（妥当性チェックは軽め）
		try:
			j = int(data.get("joint", 0))
			self.joint_var.set(j if j in (1, 2) else 0)
		except Exception:
			self.joint_var.set(0)

		self.prox_model_path.set(str(data.get("prox_model", "")))
		self.prox_pp_abcd_path.set(str(data.get("prox_pp_abcd", "")))
		self.prox_pp_olmn_path.set(str(data.get("prox_pp_olmn", "")))
		self.dist_model_path.set(str(data.get("dist_model", "")))
		self.dist_pp_abc_path.set(str(data.get("dist_pp_abc", "")))
		self.dist_pp_olmn_path.set(str(data.get("dist_pp_olmn", "")))
		self.transform_group_path.set(str(data.get("transform_group", "")))
		
		# Fittingファイルの復元
		self.fitting_parent_model_path.set(str(data.get("fitting_parent_model", "")))
		self.fitting_parent_region_path.set(str(data.get("fitting_parent_region", "")))
		self.fitting_child_model_path.set(str(data.get("fitting_child_model", "")))
		self.fitting_child_region_path.set(str(data.get("fitting_child_region", "")))
		self.fitting_child_pp_path.set(str(data.get("fitting_child_pp", "")))
		
		# ORGファイルの復元
		self.org_model_path.set(str(data.get("org_model", "")))
		self.org_pp_path.set(str(data.get("org_pp", "")))

		# A-Bファイルの復元
		self.a_b_model_a_path.set(str(data.get("a_b_model_a", "")))
		self.a_b_model_b_path.set(str(data.get("a_b_model_b", "")))

		# HeatMapファイルの復元
		self.heatmap_model_a_path.set(str(data.get("heatmap_model_a", "")))
		self.heatmap_model_b_path.set(str(data.get("heatmap_model_b", "")))
		try:
			self.heatmap_show_model_a.set(bool(data.get("heatmap_show_model_a", True)))
		except Exception:
			self.heatmap_show_model_a.set(True)
		try:
			self.heatmap_show_model_b.set(bool(data.get("heatmap_show_model_b", True)))
		except Exception:
			self.heatmap_show_model_b.set(True)
		
		# ORG座標の復元
		try:
			self.org_coord_x.set(float(data.get("org_coord_x", 0.0)))
		except (ValueError, TypeError):
			self.org_coord_x.set(0.0)
		
		try:
			self.org_coord_y.set(float(data.get("org_coord_y", 0.0)))
		except (ValueError, TypeError):
			self.org_coord_y.set(0.0)
		
		try:
			self.org_coord_z.set(float(data.get("org_coord_z", 0.0)))
		except (ValueError, TypeError):
			self.org_coord_z.set(0.0)
		
		# Fittingパラメータの復元
		# RANSACパラメータ
		try:
			self.ransac_distance_threshold.set(float(data.get("ransac_distance_threshold", 1.0)))
		except (ValueError, TypeError):
			self.ransac_distance_threshold.set(1.0)
		
		try:
			self.ransac_max_iterations.set(int(data.get("ransac_max_iterations", 1000)))
		except (ValueError, TypeError):
			self.ransac_max_iterations.set(1000)
		
		try:
			self.ransac_confidence.set(float(data.get("ransac_confidence", 0.99)))
		except (ValueError, TypeError):
			self.ransac_confidence.set(0.99)
		
		# ICPパラメータ
		try:
			self.icp_threshold.set(float(data.get("icp_threshold", 5.0)))
		except (ValueError, TypeError):
			self.icp_threshold.set(5.0)
		
		try:
			self.icp_max_iterations.set(int(data.get("icp_max_iterations", 2000)))
		except (ValueError, TypeError):
			self.icp_max_iterations.set(2000)
		
		try:
			self.sample_points.set(int(data.get("sample_points", 10000)))
		except (ValueError, TypeError):
			self.sample_points.set(10000)
		
		# カラー設定の復元
		self.prox_color = data.get("prox_color", "#B0C4DE")
		self.dist_color = data.get("dist_color", "#FFB6C1")
		self.overlap_color = data.get("overlap_color", "#BC0051")
		
		# 関節範囲の半径の復元
		try:
			self.prox_radius.set(float(data.get("prox_radius", 30.0)))
		except (ValueError, TypeError):
			self.prox_radius.set(30.0)
		
		try:
			self.dist_radius.set(float(data.get("dist_radius", 30.0)))
		except (ValueError, TypeError):
			self.dist_radius.set(30.0)
		
		# O'原点位置オフセットの復元
		try:
			self.prox_offset_x.set(float(data.get("prox_offset_x", 0.0)))
		except (ValueError, TypeError):
			self.prox_offset_x.set(0.0)
		
		try:
			self.prox_offset_y.set(float(data.get("prox_offset_y", 0.0)))
		except (ValueError, TypeError):
			self.prox_offset_y.set(0.0)
		
		try:
			self.prox_offset_z.set(float(data.get("prox_offset_z", 0.0)))
		except (ValueError, TypeError):
			self.prox_offset_z.set(0.0)
		
		try:
			self.dist_offset_x.set(float(data.get("dist_offset_x", 0.0)))
		except (ValueError, TypeError):
			self.dist_offset_x.set(0.0)
		
		try:
			self.dist_offset_y.set(float(data.get("dist_offset_y", 0.0)))
		except (ValueError, TypeError):
			self.dist_offset_y.set(0.0)
		
		try:
			self.dist_offset_z.set(float(data.get("dist_offset_z", 0.0)))
		except (ValueError, TypeError):
			self.dist_offset_z.set(0.0)

		# 軟骨分離ファイルの復元（近位）
		self.cs_prox_model1_whole_path.set(str(data.get("cs_model1_whole", data.get("cs_prox_model1_whole", ""))))
		self.cs_prox_model1_region_path.set(str(data.get("cs_model1_region", data.get("cs_prox_model1_region", ""))))
		self.cs_prox_model2_whole_path.set(str(data.get("cs_model2_whole", data.get("cs_prox_model2_whole", ""))))
		self.cs_prox_model2_region_path.set(str(data.get("cs_model2_region", data.get("cs_prox_model2_region", ""))))

		# 軟骨分離ファイルの復元（遠位）
		self.cs_dist_model1_whole_path.set(str(data.get("cs_dist_model1_whole", "")))
		self.cs_dist_model1_region_path.set(str(data.get("cs_dist_model1_region", "")))
		self.cs_dist_model2_whole_path.set(str(data.get("cs_dist_model2_whole", "")))
		self.cs_dist_model2_region_path.set(str(data.get("cs_dist_model2_region", "")))

		# Simulator 軟骨モデルの復元
		self.prox_cartilage_model_path.set(str(data.get("prox_cartilage_model", "")))
		self.dist_cartilage_model_path.set(str(data.get("dist_cartilage_model", "")))

		# 軟骨分離パラメータの復元
		try:
			self.cs_ransac_distance_threshold.set(float(data.get("cs_ransac_distance_threshold", 1.0)))
		except (ValueError, TypeError):
			self.cs_ransac_distance_threshold.set(1.0)
		try:
			self.cs_ransac_max_iterations.set(int(data.get("cs_ransac_max_iterations", 1000)))
		except (ValueError, TypeError):
			self.cs_ransac_max_iterations.set(1000)
		try:
			self.cs_ransac_confidence.set(float(data.get("cs_ransac_confidence", 0.99)))
		except (ValueError, TypeError):
			self.cs_ransac_confidence.set(0.99)
		try:
			self.cs_icp_threshold.set(float(data.get("cs_icp_threshold", 5.0)))
		except (ValueError, TypeError):
			self.cs_icp_threshold.set(5.0)
		try:
			self.cs_icp_max_iterations.set(int(data.get("cs_icp_max_iterations", 2000)))
		except (ValueError, TypeError):
			self.cs_icp_max_iterations.set(2000)
		try:
			self.cs_sample_points.set(int(data.get("cs_sample_points", 10000)))
		except (ValueError, TypeError):
			self.cs_sample_points.set(10000)
		try:
			self.cs_distance_threshold.set(float(data.get("cs_distance_threshold", 0.5)))
		except (ValueError, TypeError):
			self.cs_distance_threshold.set(0.5)
		try:
			self.cs_enable_scaling.set(bool(data.get("cs_enable_scaling", False)))
		except (ValueError, TypeError):
			self.cs_enable_scaling.set(False)

		# 軟骨分離カラー・透明度の復元
		self.cs_bone_color = data.get("cs_bone_color", "#F5DEB3")
		self.cs_cartilage_color = data.get("cs_cartilage_color", "#00CED1")
		try:
			self.cs_bone_opacity.set(float(data.get("cs_bone_opacity", 0.4)))
		except (ValueError, TypeError):
			self.cs_bone_opacity.set(0.4)
		try:
			self.cs_cartilage_opacity.set(float(data.get("cs_cartilage_opacity", 1.0)))
		except (ValueError, TypeError):
			self.cs_cartilage_opacity.set(1.0)

		# FEM解析パラメータの復元
		self.fem_prox_bone_path.set(str(data.get("fem_prox_bone", "")))
		self.fem_prox_cartilage_path.set(str(data.get("fem_prox_cartilage", "")))
		self.fem_dist_bone_path.set(str(data.get("fem_dist_bone", "")))
		self.fem_dist_cartilage_path.set(str(data.get("fem_dist_cartilage", "")))
		try: self.fem_cart_E.set(float(data.get("fem_cart_E", 10.0)))
		except (ValueError, TypeError): self.fem_cart_E.set(10.0)
		try: self.fem_cart_nu.set(float(data.get("fem_cart_nu", 0.45)))
		except (ValueError, TypeError): self.fem_cart_nu.set(0.45)
		try: self.fem_cart_thickness.set(float(data.get("fem_cart_thickness", 2.0)))
		except (ValueError, TypeError): self.fem_cart_thickness.set(2.0)
		try: self.fem_bone_E.set(float(data.get("fem_bone_E", 17000.0)))
		except (ValueError, TypeError): self.fem_bone_E.set(17000.0)
		try: self.fem_bone_nu.set(float(data.get("fem_bone_nu", 0.3)))
		except (ValueError, TypeError): self.fem_bone_nu.set(0.3)
		try: self.fem_penalty_stiffness.set(float(data.get("fem_penalty_stiffness", 500.0)))
		except (ValueError, TypeError): self.fem_penalty_stiffness.set(500.0)
		try: self.fem_contact_tolerance.set(float(data.get("fem_contact_tolerance", 2.0)))
		except (ValueError, TypeError): self.fem_contact_tolerance.set(2.0)
		self.fem_boundary_mode.set(str(data.get("fem_boundary_mode", "auto")))
		try: self.fem_max_nodes.set(int(data.get("fem_max_nodes", 30000)))
		except (ValueError, TypeError): self.fem_max_nodes.set(30000)
		self.fem_scalar_name.set(str(data.get("fem_scalar_name", "contact_pressure")))
		self.fem_cmap.set(str(data.get("fem_cmap", "jet")))
		try: self.fem_show_dist_mesh.set(bool(data.get("fem_show_dist_mesh", True)))
		except (ValueError, TypeError): self.fem_show_dist_mesh.set(True)
		try: self.fem_show_bone.set(bool(data.get("fem_show_bone", True)))
		except (ValueError, TypeError): self.fem_show_bone.set(True)
		try: self.show_fem_analysis.set(bool(data.get("show_fem_analysis", False)))
		except (ValueError, TypeError): self.show_fem_analysis.set(False)

	def _save_state(self) -> None:
		data = {
			"joint": self.joint_var.get(),
			"prox_model": self.prox_model_path.get(),
			"prox_pp_abcd": self.prox_pp_abcd_path.get(),
			"prox_pp_olmn": self.prox_pp_olmn_path.get(),
			"dist_model": self.dist_model_path.get(),
			"dist_pp_abc": self.dist_pp_abc_path.get(),
			"dist_pp_olmn": self.dist_pp_olmn_path.get(),
			"transform_group": self.transform_group_path.get(),
			"fitting_parent_model": self.fitting_parent_model_path.get(),
			"fitting_parent_region": self.fitting_parent_region_path.get(),
			"fitting_child_model": self.fitting_child_model_path.get(),
			"fitting_child_region": self.fitting_child_region_path.get(),
			"fitting_child_pp": self.fitting_child_pp_path.get(),
			"org_model": self.org_model_path.get(),
			"org_pp": self.org_pp_path.get(),
			"a_b_model_a": self.a_b_model_a_path.get(),
			"a_b_model_b": self.a_b_model_b_path.get(),
			"heatmap_model_a": self.heatmap_model_a_path.get(),
			"heatmap_model_b": self.heatmap_model_b_path.get(),
			"heatmap_show_model_a": self.heatmap_show_model_a.get(),
			"heatmap_show_model_b": self.heatmap_show_model_b.get(),
			"org_coord_x": self.org_coord_x.get(),
			"org_coord_y": self.org_coord_y.get(),
			"org_coord_z": self.org_coord_z.get(),
			"ransac_distance_threshold": self.ransac_distance_threshold.get(),
			"ransac_max_iterations": self.ransac_max_iterations.get(),
			"ransac_confidence": self.ransac_confidence.get(),
			"icp_threshold": self.icp_threshold.get(),
			"icp_max_iterations": self.icp_max_iterations.get(),
			"sample_points": self.sample_points.get(),
			"prox_color": self.prox_color,
			"dist_color": self.dist_color,
			"overlap_color": self.overlap_color,
			"prox_radius": self.prox_radius.get(),
			"dist_radius": self.dist_radius.get(),
			"prox_offset_x": self.prox_offset_x.get(),
			"prox_offset_y": self.prox_offset_y.get(),
			"prox_offset_z": self.prox_offset_z.get(),
			"dist_offset_x": self.dist_offset_x.get(),
			"dist_offset_y": self.dist_offset_y.get(),
			"dist_offset_z": self.dist_offset_z.get(),
			# 軟骨分離（近位）
			"cs_prox_model1_whole": self.cs_prox_model1_whole_path.get(),
			"cs_prox_model1_region": self.cs_prox_model1_region_path.get(),
			"cs_prox_model2_whole": self.cs_prox_model2_whole_path.get(),
			"cs_prox_model2_region": self.cs_prox_model2_region_path.get(),
			# 軟骨分離（遠位）
			"cs_dist_model1_whole": self.cs_dist_model1_whole_path.get(),
			"cs_dist_model1_region": self.cs_dist_model1_region_path.get(),
			"cs_dist_model2_whole": self.cs_dist_model2_whole_path.get(),
			"cs_dist_model2_region": self.cs_dist_model2_region_path.get(),
			# 軟骨分離パラメータ
			"cs_ransac_distance_threshold": self.cs_ransac_distance_threshold.get(),
			"cs_ransac_max_iterations": self.cs_ransac_max_iterations.get(),
			"cs_ransac_confidence": self.cs_ransac_confidence.get(),
			"cs_icp_threshold": self.cs_icp_threshold.get(),
			"cs_icp_max_iterations": self.cs_icp_max_iterations.get(),
			"cs_sample_points": self.cs_sample_points.get(),
			"cs_distance_threshold": self.cs_distance_threshold.get(),
			"cs_enable_scaling": self.cs_enable_scaling.get(),
			"cs_bone_color": self.cs_bone_color,
			"cs_cartilage_color": self.cs_cartilage_color,
			"cs_bone_opacity": self.cs_bone_opacity.get(),
			"cs_cartilage_opacity": self.cs_cartilage_opacity.get(),
			# Simulator 軟骨モデル
			"prox_cartilage_model": self.prox_cartilage_model_path.get(),
			"dist_cartilage_model": self.dist_cartilage_model_path.get(),
			# FEM解析
			"fem_prox_bone": self.fem_prox_bone_path.get(),
			"fem_prox_cartilage": self.fem_prox_cartilage_path.get(),
			"fem_dist_bone": self.fem_dist_bone_path.get(),
			"fem_dist_cartilage": self.fem_dist_cartilage_path.get(),
			"fem_cart_E": self.fem_cart_E.get(),
			"fem_cart_nu": self.fem_cart_nu.get(),
			"fem_cart_thickness": self.fem_cart_thickness.get(),
			"fem_bone_E": self.fem_bone_E.get(),
			"fem_bone_nu": self.fem_bone_nu.get(),
			"fem_penalty_stiffness": self.fem_penalty_stiffness.get(),
			"fem_contact_tolerance": self.fem_contact_tolerance.get(),
			"fem_boundary_mode": self.fem_boundary_mode.get(),
			"fem_max_nodes": self.fem_max_nodes.get(),
			"fem_scalar_name": self.fem_scalar_name.get(),
			"fem_cmap": self.fem_cmap.get(),
			"fem_show_dist_mesh": self.fem_show_dist_mesh.get(),
			"fem_show_bone": self.fem_show_bone.get(),
			"show_fem_analysis": self.show_fem_analysis.get(),
		}
		try:
			p = self._state_file_path()
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
		except Exception:
			# 保存失敗は致命的でないため無視
			pass

	def _on_close(self) -> None:
		self._save_state()
		self.destroy()
	
	def _save_initial_geometry(self) -> None:
		"""初期ウィンドウサイズを保存"""
		try:
			self.update_idletasks()
			self._initial_window_geometry = self.geometry()
		except Exception:
			pass
	
	def _restore_window_geometry(self) -> None:
		"""ウィンドウサイズを復元"""
		try:
			if self._initial_window_geometry:
				self.geometry(self._initial_window_geometry)
				self.update_idletasks()
		except Exception:
			pass

	# ----- Utils -----
	def _apply_transform(self, mesh: pv.PolyData, transform_matrix: np.ndarray) -> pv.PolyData:
		"""メッシュに同次変換行列を適用
		
		Args:
			mesh: 変換するメッシュ
			transform_matrix: 4x4同次変換行列
			
		Returns:
			pv.PolyData: 変換されたメッシュ
		"""
		if mesh.n_points == 0:
			return mesh
		
		# 同次座標に変換
		points_homo = np.hstack([mesh.points, np.ones((mesh.points.shape[0], 1))])
		
		# 変換を適用
		transformed_points = (transform_matrix @ points_homo.T).T[:, :3]
		
		# 新しいメッシュを作成
		transformed_mesh = mesh.copy()
		transformed_mesh.points = transformed_points
		
		return transformed_mesh
	
	def _compute_distance_heatmap(self, prox_joint_region: pv.PolyData, dist_joint_region: pv.PolyData, prox_surface: pv.PolyData = None, dist_surface: pv.PolyData = None, prox_points: np.ndarray = None) -> pv.PolyData:
		"""
		近位関節領域の各頂点に対して、遠位関節領域表面までの符号付き距離を計算してヒートマップを作成
		
		このメソッドは接触状態の可視化に使用されます。距離が小さいほど接触に近い状態を示します。
		
		Args:
			prox_joint_region: 近位関節領域（寛骨臼）のメッシュ
			dist_joint_region: 遠位関節領域（大腿骨頭）のメッシュ
			prox_surface: 事前抽出された近位表面（オプション、高速化用）
			dist_surface: 事前抽出された遠位表面（オプション、高速化用）
			prox_points: 事前抽出された近位点群（オプション、高速化用）
			
		Returns:
			pv.PolyData: "distance" スカラーに符号付き距離を持つ点群
			            - 正の値: 遠位が近位の外側にある（非接触）
			            - 0付近: 接触状態
			            - 負の値: 侵入状態（オーバーラップ）
			
		■ 処理フロー:
		  1. バウンディングボックスチェック（10mm以上離れている場合は簡易計算）
		  2. 近位表面抽出（prox_surfaceが指定されていない場合）
		  3. 遠位表面抽出（dist_surfaceが指定されていない場合）
		  4. 近位点群の各点について遠位表面への最近接点を検索
		  5. 距離計算と符号決定（法線ベクトルの内積で判定）
		  
		■ 最適化手法:
		  - バウンディングボックスによる早期リターン
		  - 表面抽出の再利用（prox_surface, dist_surface指定時）
		  - 点群のダウンサンプリング（prox_points指定時）
		  - KD-Tree高速近傍探索（PyVistaのfind_closest_cell使用）
		  
		■ 改修ポイント:
		  - 行7260付近: 表面抽出ロジック
		  - 行7300付近: 距離計算と符号決定
		  
		■ パフォーマンス注意:
		  - 大規模メッシュ（10万点以上）では処理時間が長くなる可能性あり
		  - 並列処理での呼び出しを推奨（_precompute_overlap_volumes経由）
		"""
		if prox_joint_region.n_points == 0 or dist_joint_region.n_points == 0:
			return prox_joint_region.copy() if prox_joint_region.n_points > 0 else pv.PolyData()
		
		# バウンディングボックスの事前チェック（交差判定の前に実施）
		try:
			prox_bounds = prox_joint_region.bounds
			dist_bounds = dist_joint_region.bounds
			
			# バウンディングボックスが遠く離れている場合（閾値: 10mm以上）
			# X, Y, Z軸それぞれで距離をチェック
			x_gap = max(0, dist_bounds[0] - prox_bounds[1], prox_bounds[0] - dist_bounds[1])
			y_gap = max(0, dist_bounds[2] - prox_bounds[3], prox_bounds[2] - dist_bounds[3])
			z_gap = max(0, dist_bounds[4] - prox_bounds[5], prox_bounds[4] - dist_bounds[5])
			
			# いずれかの軸で10mm以上離れている場合は、簡易計算のみ実行
			if x_gap > 10 or y_gap > 10 or z_gap > 10:
				# 中心間距離を全点の距離として設定（概算）
				prox_center = prox_joint_region.center
				dist_center = dist_joint_region.center
				approx_distance = np.linalg.norm(prox_center - dist_center)
				# メッシュ構造を保持して結果を返す
				result = prox_joint_region.copy(deep=False)
				result['distance'] = np.full(prox_joint_region.n_points, approx_distance)
				return result
		except Exception as e:
			# バウンディングボックスチェックに失敗した場合は通常計算を続行
			pass
		
		# 正確な符号付き距離計算：boolean_intersection で実際の重複を判定
		try:
			# Phase 2最適化: 事前抽出された配列があれば使用
			pts = prox_points if prox_points is not None else np.asarray(prox_joint_region.points)
			if pts.size == 0:
				return pv.PolyData(pts)

			# まず最近傍距離を計算
			from scipy.spatial import cKDTree
			dist_points = np.asarray(dist_joint_region.points)
			if dist_points.shape[0] == 0:
				if prox_points is None or len(prox_points) == prox_joint_region.n_points:
					result = prox_joint_region.copy(deep=False)
				else:
					result = pv.PolyData(pts)
				result['distance'] = np.full(len(pts), 100.0)  # 遠位メッシュがない場合は大きな値
				return result
			
			# KDTreeで各近位点から遠位メッシュへの最小距離を計算（最近傍インデックスも保存）
			tree = cKDTree(dist_points)
			d_vals, closest_indices = tree.query(pts, k=1)
			
			# 最適化: 最近傍距離で早期スキップ判定（結果は変わらない）
			min_distance = np.min(d_vals)
			if min_distance > 0.5:
				# 全ての点が0.5m以上離れている場合、boolean_intersectionは不要
				# （0.5m未満の点のみを負にする処理なので、結果は同じ）
				print(f"[最適化] 最小距離{min_distance:.2f}mm > 0.5m のためboolean_intersectionをスキップ")
				
				# 元のメッシュ構造を保持できるか確認
				if prox_points is None or len(prox_points) == prox_joint_region.n_points:
					result = prox_joint_region.copy(deep=False)
				else:
					result = pv.PolyData(pts)
					
				result['distance'] = np.array(d_vals, dtype=float)
				return result
			
			# 【超高速化】Signed Distance で埋め込み検出（10-50倍高速）
			try:
				# 事前抽出された表面を使用（高速化）、なければその場で抽出
				if dist_surface is not None:
					dist_mesh_for_implicit = dist_surface
				else:
					dist_mesh_for_implicit = dist_joint_region
					if hasattr(dist_joint_region, 'extract_surface'):
						dist_mesh_for_implicit = dist_joint_region.extract_surface()
				
				# PyVistaのimplicit distance（signed distance）を計算
				# 遠位メッシュを暗黙的関数として、近位点からの符号付き距離を計算
				# 負 = 遠位メッシュの内部、正 = 外部
				prox_point_cloud = pv.PolyData(pts)
				signed_distances = prox_point_cloud.compute_implicit_distance(dist_mesh_for_implicit, inplace=False)
				
				if 'implicit_distance' in signed_distances.point_data:
					implicit_dist = signed_distances.point_data['implicit_distance']
					
					# 埋め込み判定: implicit_distance が負 = 遠位メッシュの内部にある
					embedded_mask = implicit_dist < 0
					
					# デバッグ: 統計表示
					n_embedded = np.sum(embedded_mask)
					if n_embedded > 0:
						print(f"[Signed Distance] 埋め込み検出: {n_embedded}個 / {len(pts)}点")
						embedded_implicit = implicit_dist[embedded_mask]
						print(f"[Signed Distance] 埋め込み深さ: min={embedded_implicit.min():.3f}mm, max={embedded_implicit.max():.3f}mm, mean={embedded_implicit.mean():.3f}mm")
					
					# 埋め込まれている点: implicit_distanceを使用（既に符号付き）
					# KDTree距離（d_vals）と組み合わせる: 埋め込みなら負、外なら正
					# d_valsの絶対値を保持しつつ、埋め込みの場合のみ負にする
					d_vals[embedded_mask] = -np.abs(d_vals[embedded_mask])
					
				else:
					print(f"[警告] implicit_distance計算失敗 - フォールバック判定")
					# フォールバック: 距離が0.2mm未満なら埋め込みとみなす
					very_close_mask = d_vals < 0.2
					if np.any(very_close_mask):
						d_vals[very_close_mask] = -d_vals[very_close_mask]
				
			except Exception as e:
				# フォールバック: 距離が極めて小さい点を負にする（0.2mm未満）
				inside_mask = d_vals < 0.2
				if np.any(inside_mask):
					d_vals[inside_mask] = -d_vals[inside_mask]
			
			try:
				# 結果メッシュを作成（元のメッシュ構造を保持）
				# prox_pointsが指定されている場合はそれがprox_joint_regionの頂点に対応すると仮定
				# もしprox_pointsがダウンサンプリングされている場合は、元のメッシュ構造を使えないためPoint Cloudになる
				if prox_points is None or len(prox_points) == prox_joint_region.n_points:
					result = prox_joint_region.copy(deep=False)
				else:
					# ダウンサンプリングされている場合はPoint Cloudとして作成せざるを得ない
					result = pv.PolyData(pts)
				
				# 距離配列を設定
				result['distance'] = np.array(d_vals, dtype=float)
				return result
				
			except Exception as e2:
				print(f"[警告] 結果メッシュ生成エラー: {e2}")
				result = pv.PolyData(pts)
				result['distance'] = np.array(d_vals, dtype=float)
				return result

		except Exception as e:
			print(f"[警告] 距離ヒートマップ計算エラー: {e}")
			# フォールバックは元のメッシュ構造を維持
			fallback = prox_joint_region.copy(deep=False) if prox_joint_region.n_points > 0 else pv.PolyData()
			if fallback.n_points > 0:
				fallback['distance'] = np.full(fallback.n_points, 100.0)  # エラー時は全点を遠くに
			else:
				# 点がない場合はダミー
				fallback = pv.PolyData(np.array([[0.0, 0.0, 0.0]]))
				fallback['distance'] = np.array([100.0])
			return fallback

		
		# 変換を適用
		transformed_points = (transform_matrix @ points_homo.T).T[:, :3]
		
		# 新しいメッシュを作成
		transformed_mesh = mesh.copy()
		transformed_mesh.points = transformed_points
		
		return transformed_mesh
	
	# ----- Utils: Sphere Extraction -----
	def _extract_region_by_sphere(self, mesh: pv.PolyData, origin: np.ndarray, radius: float) -> pv.PolyData:
		"""球体の内部にあるメッシュ部分を抽出
		
		Args:
			mesh: 元のメッシュ
			origin: 座標系の原点（球の中心）
			radius: 球の半径（mm）
			
		Returns:
			pv.PolyData: 球内に含まれるメッシュ部分（表面メッシュとして有効なもの）
		"""
		if mesh is None or mesh.n_points == 0:
			return pv.PolyData()
		
		try:
			# メッシュの全頂点を取得
			points = mesh.points
			
			# 各点から原点までの距離を計算
			distances = np.linalg.norm(points - origin, axis=1)
			
			# 半径以内の点のマスクを作成
			mask = distances <= radius
			
			if not np.any(mask):
				return pv.PolyData()
				
			# extract_pointsで点を抽出（adjacent_cells=Trueでセルも含める）
			# PyVistaのバージョンによってはUnstructuredGridが返る
			extracted = mesh.extract_points(mask, adjacent_cells=True)
			
			# UnstructuredGridの場合はPolyData（表面）に変換して、面情報を正しく保持する
			# これにより点群ではなくメッシュとして描画されるようになる
			if hasattr(extracted, 'extract_surface'):
				surface = extracted.extract_surface()
				return surface
			elif isinstance(extracted, pv.PolyData):
				return extracted.clean()
			else:
				# 変換できない場合はクリーニングだけ試みる
				if hasattr(extracted, 'clean'):
					return extracted.clean()
				return extracted
				
		except Exception as e:
			print(f"領域抽出エラー: {e}")
			return pv.PolyData()

	# ----- Utils: Coordinate System Builder -----
	def _build_coordinate_system(self, points, labels):
		"""A, B, C, D, O点から座標系を構築
		
		原点: O
		Z軸: B→Aの方向
		X軸: Zに直交かつ、CとDの中点からAB線への垂線
		Y軸: Z×X軸
		
		Returns:
			origin, x_axis, y_axis, z_axis (各軸は単位ベクトル)
		"""
		import numpy as np
		
		# ラベルから点を取得
		point_dict = {}
		for i, label in enumerate(labels):
			label_upper = label.strip().upper()
			if label_upper in ['A', 'B', 'C', 'D', 'O']:
				point_dict[label_upper] = points[i]
		
		# 必要な点が揃っているか確認
		required = ['A', 'B', 'C', 'D', 'O']
		missing = [p for p in required if p not in point_dict]
		if missing:
			raise ValueError(f"必要な点が見つかりません: {', '.join(missing)}")
		
		A = np.array(point_dict['A'])
		B = np.array(point_dict['B'])
		C = np.array(point_dict['C'])
		D = np.array(point_dict['D'])
		O = np.array(point_dict['O'])
		
		# 原点
		origin = O
		
		# Z軸: B→Aの方向
		z_axis = A - B
		z_axis = z_axis / np.linalg.norm(z_axis)  # 正規化
		
		# CとDの中点
		mid_CD = (C + D) / 2.0
		
		# AB線上の点でmid_CDに最も近い点を求める
		# AB線: P = B + t*(A-B)
		# mid_CDからPへのベクトルがAB線に垂直になるtを求める
		AB = A - B
		t = np.dot(mid_CD - B, AB) / np.dot(AB, AB)
		closest_point = B + t * AB
		
		# mid_CDからAB線への垂線ベクトル
		perp = closest_point - mid_CD
		if np.linalg.norm(perp) < 1e-10:
			raise ValueError("CとDの中点がAB線上にあるため、X軸を定義できません")
		
		# X軸: 垂線方向でZ軸に直交
		x_axis = perp - np.dot(perp, z_axis) * z_axis  # Z軸成分を除去
		x_axis = x_axis / np.linalg.norm(x_axis)  # 正規化
		
		# Y軸: Z×X軸
		y_axis = np.cross(z_axis, x_axis)
		y_axis = y_axis / np.linalg.norm(y_axis)  # 正規化
		
		return origin, x_axis, y_axis, z_axis

	def _build_coordinate_system_dist(self, points, labels):
		"""A, B, O点から遠位用座標系を構築
		
		原点: O
		Z軸: B→Aの方向
		Y軸: BとAの中点からOに向かうベクトル（Z軸に直交）
		X軸: Y×Z軸
		
		Returns:
			origin, x_axis, y_axis, z_axis (各軸は単位ベクトル)
		"""
		import numpy as np
		
		# ラベルから点を取得
		point_dict = {}
		for i, label in enumerate(labels):
			label_upper = label.strip().upper()
			if label_upper in ['A', 'B', 'O']:
				point_dict[label_upper] = points[i]
		
		# 必要な点が揃っているか確認
		required = ['A', 'B', 'O']
		missing = [p for p in required if p not in point_dict]
		if missing:
			raise ValueError(f"必要な点が見つかりません: {', '.join(missing)}")
		
		A = np.array(point_dict['A'])
		B = np.array(point_dict['B'])
		O = np.array(point_dict['O'])
		
		# 原点
		origin = O
		
		# Z軸: B→Aの方向
		z_axis = A - B
		z_axis = z_axis / np.linalg.norm(z_axis)  # 正規化
		
		# BとAの中点
		mid_AB = (A + B) / 2.0
		
		# 中点からOに向かうベクトル
		vec_to_O = O - mid_AB
		if np.linalg.norm(vec_to_O) < 1e-10:
			raise ValueError("BとAの中点とOが一致しているため、Y軸を定義できません")
		
		# Y軸: ベクトルをZ軸に直交させる
		y_axis = vec_to_O - np.dot(vec_to_O, z_axis) * z_axis  # Z軸成分を除去
		if np.linalg.norm(y_axis) < 1e-10:
			raise ValueError("OがAB線上にあるため、Y軸を定義できません")
		y_axis = y_axis / np.linalg.norm(y_axis)  # 正規化
		
		# X軸: Y×Z軸
		x_axis = np.cross(y_axis, z_axis)
		x_axis = x_axis / np.linalg.norm(x_axis)  # 正規化
		
		return origin, x_axis, y_axis, z_axis

	def _try_auto_set_offset(self, points, labels, origin, x_axis, y_axis, z_axis, off_x_var, off_y_var, off_z_var):
		"""O'点が存在し、かつオフセットが未設定(0)の場合、自動的にオフセットを設定する"""
		import numpy as np
		# 現在のオフセット値を確認
		try:
			curr_x = off_x_var.get()
			curr_y = off_y_var.get()
			curr_z = off_z_var.get()
		except:
			return

		# 既に設定されている場合は何もしない（0.001mm以上の絶対値がある場合）
		if abs(curr_x) > 1e-3 or abs(curr_y) > 1e-3 or abs(curr_z) > 1e-3:
			return

		# O'を探す
		op_idx = -1
		target_labels = ["O'", "O_PRIME", "OPRIME", "O_DASH", "ODASH"]
		
		for i, label in enumerate(labels):
			lab = label.strip().upper()
			if lab in target_labels:
				op_idx = i
				break
		
		if op_idx != -1:
			O_prime = points[op_idx]
			
			# Global Offset Vector
			diff = O_prime - origin
			
			# Project to Local Coordinate System (x, y, z axes are normalized)
			# Local Offset = Dot(Diff, Axis)
			local_off_x = np.dot(diff, x_axis)
			local_off_y = np.dot(diff, y_axis)
			local_off_z = np.dot(diff, z_axis)
			
			# Update Variables
			off_x_var.set(float(local_off_x))
			off_y_var.set(float(local_off_y))
			off_z_var.set(float(local_off_z))
			
			print(f"[自動設定] O'点を検出しました。オフセットを自動設定します: [{local_off_x:.3f}, {local_off_y:.3f}, {local_off_z:.3f}]")

	# ----- Utils: PP loader -----
	def _parse_pp_file(self, path: str):
		"""PPファイルから点群を抽出。
		対応形式:
		- GOM Inspect 等の PickedPoints XML 形式 (<PickedPoints><point x="" y="" z="" name="" active="1"/> ...)
		- テキスト行形式: "x y z" / "name x y z" / CSV(3〜4列)
		- 混在形式: XMLブロックの後にテキスト行が続く場合にも対応
		- 行頭が#の行はコメント扱い
		"""
		p = Path(path)
		if not p.exists():
			raise FileNotFoundError(f"PPファイルが存在しません: {path}")

		# ファイル全文を読み込む
		full_text = p.read_text(encoding="utf-8", errors="ignore")
		pts = []
		labels = []

		# まずXMLブロック（先頭〜</PickedPoints>まで）を安全に抽出してパース
		xml_end_idx = full_text.lower().find('</pickedpoints>')
		if full_text.lstrip().startswith('<') and xml_end_idx != -1:
			xml_block = full_text[:xml_end_idx + len('</PickedPoints>')]
			try:
				import xml.etree.ElementTree as ET
				root = ET.fromstring(xml_block)
				if root.tag.lower() == 'pickedpoints':
					for pt in root.findall('.//point'):
						active = pt.get('active')
						if active is not None and str(active) not in ('1', 'true', 'True'):
							continue
						name = pt.get('name') or f"P{len(pts)+1}"
						x = float(pt.get('x'))
						y = float(pt.get('y'))
						z = float(pt.get('z'))
						pts.append([x, y, z])
						labels.append(name)
			except Exception:
				# XML抽出/パース失敗時は後続のテキスト処理へ
				pass

		# XMLの後に続くテキスト部分を個別にパース（混在形式対応）
		text_tail = full_text[xml_end_idx + len('</PickedPoints>'):] if xml_end_idx != -1 else full_text
		with io.StringIO(text_tail) as f:
			for line in f:
				s = line.strip()
				if not s or s.startswith('#'):
					continue
				# カンマ or 空白区切り
				parts = [t for t in s.replace(',', ' ').split() if t]
				if len(parts) < 3:
					continue
				# 後ろ3つをx,y,zとして解釈
				try:
					x, y, z = float(parts[-3]), float(parts[-2]), float(parts[-1])
					pts.append([x, y, z])
					name = ' '.join(parts[:-3]) if len(parts) > 3 else f"P{len(pts)}"
					labels.append(name)
				except Exception:
					continue
		import numpy as np
		if not pts:
			raise ValueError("PPファイルから座標を抽出できませんでした。")
		return np.asarray(pts, dtype=float), labels

	def _uniform_sample(self, points: np.ndarray, n_samples: int) -> np.ndarray:
		"""点群から均一にサンプリング
		
		Args:
			points: 点群 (N, 3)
			n_samples: サンプリング点数
		
		Returns:
			サンプリングされた点群 (n_samples, 3)
		"""
		if len(points) <= n_samples:
			return points
		
		# ランダムにサンプリング
		indices = np.random.choice(len(points), n_samples, replace=False)
		return points[indices]


# グローバル関数（multiprocessing用）
def _compute_overlap_for_frame(args):
	"""単一フレームのオーバーラップ体積を計算（並列処理用）
	
	Args:
		args: (prox_mesh_points, prox_mesh_faces, dist_mesh_points_initial, 
		       dist_mesh_faces, dist_origin_initial, matrix, frame_idx, use_bbox_check)
	
	Returns:
		tuple: (frame_idx, overlap_mesh_points, overlap_mesh_faces, overlap_area, contact_depth)
	"""
	try:
		import pyvista as pv
		import numpy as np
		
		(prox_mesh_points, prox_mesh_faces, dist_mesh_points_initial, 
		 dist_mesh_faces, dist_origin_initial, matrix, frame_idx, use_bbox_check) = args
		
		# 近位メッシュを再構築
		prox_mesh = pv.PolyData(prox_mesh_points, prox_mesh_faces)
		
		# 遠位メッシュを変換
		centered_points = dist_mesh_points_initial - dist_origin_initial
		ones = np.ones((centered_points.shape[0], 1))
		points_homogeneous = np.hstack([centered_points, ones])
		transformed_homogeneous = (matrix @ points_homogeneous.T).T
		transformed_points = transformed_homogeneous[:, :3] + dist_origin_initial
		
		# 変換された遠位メッシュを作成
		dist_mesh = pv.PolyData(transformed_points, dist_mesh_faces)
		
		# バウンディングボックスチェック（オプション）
		if use_bbox_check:
			prox_bounds = prox_mesh.bounds
			dist_bounds = dist_mesh.bounds
			
			# バウンディングボックスが交差していない場合はスキップ
			if (prox_bounds[1] < dist_bounds[0] or prox_bounds[0] > dist_bounds[1] or
				prox_bounds[3] < dist_bounds[2] or prox_bounds[2] > dist_bounds[3] or
				prox_bounds[5] < dist_bounds[4] or prox_bounds[4] > dist_bounds[5]):
				return (frame_idx, None, None, 0.0, 0.0)
		
		# 距離ベースの事前チェック：最小距離が閾値以上なら交差なし
		# 近位メッシュからサンプル点を取得（高速化のため全点は使わない）
		sample_size = min(1000, prox_mesh.n_points)
		if prox_mesh.n_points > sample_size:
			sample_indices = np.random.choice(prox_mesh.n_points, sample_size, replace=False)
			sample_points = prox_mesh.points[sample_indices]
		else:
			sample_points = prox_mesh.points
		
		# 各サンプル点から遠位メッシュへの最小距離を計算
		min_distances = []
		for point in sample_points:
			_, closest_point = dist_mesh.find_closest_cell(point, return_closest_point=True)
			distance = np.linalg.norm(point - closest_point)
			min_distances.append(distance)
		
		# 最小距離が0.1mm以上なら接触なしと判定
		min_dist = min(min_distances) if min_distances else float('inf')
		if min_dist > 0.1:  # 閾値: 0.1mm
			return (frame_idx, None, None, 0.0, 0.0)
		
		# オーバーラップ体積を計算
		# VTKの警告・エラーログを一時的に抑制
		import vtk
		import os
		error_output = vtk.vtkFileOutputWindow()
		error_output.SetFileName("nul" if os.name == 'nt' else "/dev/null")
		vtk.vtkOutputWindow.SetInstance(error_output)
		
		overlap_result = prox_mesh.boolean_intersection(dist_mesh)
		
		# ログ出力を元に戻す
		vtk.vtkOutputWindow.SetInstance(None)
		
		if overlap_result.n_points > 0:
			# メッシュデータと面積を返す
			area = overlap_result.area
			# 面積が極端に小さい場合は数値誤差とみなす
			if area < 0.01:  # 0.01 mm² 以下は無視
				return (frame_idx, None, None, 0.0, 0.0)
			
			# 接触深度を計算（オーバーラップメッシュを渡す）
			depth = _compute_contact_depth_global(prox_mesh, dist_mesh, overlap_result)
			return (frame_idx, overlap_result.points, overlap_result.faces, area, depth)
		else:
			return (frame_idx, None, None, 0.0, 0.0)
	
	except Exception as e:
		print(f"[並列計算] フレーム {frame_idx}: エラー - {e}")
		return (frame_idx, None, None, 0.0, 0.0)


def _compute_contact_depth_global(prox_mesh, dist_mesh, overlap_mesh=None):
	"""接触深度を計算（グローバル関数・並列処理用）
	
	Args:
		prox_mesh: 近位メッシュ（PyVistaメッシュ）
		dist_mesh: 遠位メッシュ（PyVistaメッシュ）
		overlap_mesh: オーバーラップメッシュ（オプション）
		
	Returns:
		float: 接触深度（mm）、接触がない場合は0.0
	"""
	try:
		import pyvista as pv
		import numpy as np
		
		# 方法1: オーバーラップメッシュが提供されている場合
		# オーバーラップメッシュの厚み（バウンディングボックスの最小辺）を深度とする
		if overlap_mesh is not None and overlap_mesh.n_points > 0:
			try:
				# オーバーラップ領域のバウンディングボックスを取得
				bounds = overlap_mesh.bounds
				# X, Y, Z方向の幅を計算
				x_width = bounds[1] - bounds[0]
				y_width = bounds[3] - bounds[2]
				z_width = bounds[5] - bounds[4]
				# 最小幅を深度とする（貫入深さの近似）
				min_width = min(x_width, y_width, z_width)
				return max(0.0, min_width)  # 負の値は返さない
			except Exception:
				pass
		
		# 方法2: 遠位メッシュの頂点から近位メッシュへの最小距離を計算
		dist_points = dist_mesh.points
		
		# サンプリング（計算時間短縮）
		if len(dist_points) > 2000:
			indices = np.random.choice(len(dist_points), 2000, replace=False)
			sample_dist_points = dist_points[indices]
		else:
			sample_dist_points = dist_points
		
		# 各点から近位メッシュへの距離を計算
		min_distances = []
		for point in sample_dist_points:
			try:
				idx = prox_mesh.find_closest_point(point)
				if isinstance(idx, (int, np.integer)) and idx >= 0:
					closest_pt = prox_mesh.points[idx]
					distance = np.linalg.norm(point - closest_pt)
					min_distances.append(distance)
			except Exception:
				pass
		
		if len(min_distances) > 0:
			# 最小距離の下位10%の平均を深度とする
			sorted_distances = np.sort(min_distances)
			n_close = max(1, len(sorted_distances) // 10)
			close_distances = sorted_distances[:n_close]
			mean_depth = np.mean(close_distances)
			return mean_depth
		
		return 0.0
	
	except Exception as e:
		print(f"[接触深度計算] エラー: {e}")
		return 0.0

	def _uniform_sample(self, points: np.ndarray, n_samples: int) -> np.ndarray:
		"""点群から均一にサンプリング
		
		Args:
			points: 点群 (N, 3)
			n_samples: サンプリング点数
		
		Returns:
			サンプリングされた点群 (n_samples, 3)
		"""
		if len(points) <= n_samples:
			return points
		
		# ランダムにサンプリング
		indices = np.random.choice(len(points), n_samples, replace=False)
		return points[indices]
	
	def _icp_registration(self, source_points: np.ndarray, target_points: np.ndarray) -> Optional[np.ndarray]:
		"""ICP（Iterative Closest Point）による位置合わせ
		
		Args:
			source_points: 子の点群 (N, 3)
			target_points: 親の点群 (M, 3)
		
		Returns:
			変換行列 (4, 4) または None（失敗時）
		"""
		max_iterations = self.icp_max_iterations.get()
		threshold = self.icp_threshold.get()
		
		print(f"ICP設定: 最大反復回数={max_iterations}, 距離閾値={threshold}mm")
		
		# KDTreeを構築（親点群）
		tree = cKDTree(target_points)
		
		# 現在の変換（単位行列から開始）
		current_transform = np.eye(4)
		prev_error = float('inf')
		
		transformed = source_points.copy()
		
		for iteration in range(max_iterations):
			# 最近傍点を探索
			distances, indices = tree.query(transformed)
			
			# 距離閾値以下の点のみ使用
			valid_mask = distances < threshold
			valid_source = transformed[valid_mask]
			valid_target = target_points[indices[valid_mask]]
			
			if len(valid_source) < 3:
				print(f"反復{iteration+1}: 対応点が不足（{len(valid_source)}点）")
				break
			
			# 平均誤差を計算
			mean_error = np.mean(distances[valid_mask])
			
			if iteration % 50 == 0 or iteration == max_iterations - 1:
				print(f"反復{iteration+1}: 平均誤差={mean_error:.4f}mm, 対応点数={len(valid_source)}")
			
			# 収束判定
			if abs(prev_error - mean_error) < 0.0001:
				print(f"収束しました（反復{iteration+1}）")
				break
			
			prev_error = mean_error
			
			# 変換行列を推定
			transform = self._estimate_transform_svd(valid_source, valid_target)
			
			# 変換を適用
			transformed = self._apply_transform_to_points(transformed, transform)
			
			# 累積変換を更新
			current_transform = transform @ current_transform
		
		# 最終的な対応点数と誤差を計算
		distances, _ = tree.query(transformed)
		valid_mask = distances < threshold
		final_inliers = np.sum(valid_mask)
		final_error = np.mean(distances[valid_mask]) if final_inliers > 0 else float('inf')
		
		print(f"\nICP完了:")
		print(f"  最終平均誤差: {final_error:.4f}mm")
		print(f"  対応点数: {final_inliers} / {len(source_points)}")
		
		return current_transform
	
	def _estimate_transform_svd(self, source: np.ndarray, target: np.ndarray) -> np.ndarray:
		"""SVDを使用して剛体変換を推定（重心ベース）
		
		Args:
			source: 子の点群 (N, 3)
			target: 親の点群 (N, 3)
		
		Returns:
			変換行列 (4, 4)
		"""
		# 重心を計算
		centroid_source = np.mean(source, axis=0)
		centroid_target = np.mean(target, axis=0)
		
		# 重心を原点に移動
		source_centered = source - centroid_source
		target_centered = target - centroid_target
		
		# 共分散行列を計算
		H = source_centered.T @ target_centered
		
		# SVD分解
		U, S, Vt = np.linalg.svd(H)
		
		# 回転行列を計算
		R = Vt.T @ U.T
		
		# 特殊な反射を修正
		if np.linalg.det(R) < 0:
			Vt[-1, :] *= -1
			R = Vt.T @ U.T
		
		# 平行移動を計算
		t = centroid_target - R @ centroid_source
		
		# 4x4変換行列を構築
		transform = np.eye(4)
		transform[:3, :3] = R
		transform[:3, 3] = t
		
		return transform
	
	def _save_transformed_model(self, transform: np.ndarray) -> None:
		"""変換済みの子モデル全体を保存
		
		Args:
			transform: 変換行列 (4, 4)
		"""
		try:
			# 子モデル全体を読み込み
			child_model_mesh = pv.read(self.fitting_child_model_path.get())
			
			print(f"\n子モデル全体を変換中...")
			print(f"  頂点数: {child_model_mesh.n_points}")
			print(f"  三角形数: {child_model_mesh.n_cells}")
			
			# 変換を適用
			child_model_transformed = child_model_mesh.copy()
			child_model_transformed.transform(transform, inplace=True)
			
			# 保存先を選択
			original_path = Path(self.fitting_child_model_path.get())
			default_filename = original_path.stem + "_fitted" + original_path.suffix
			
			save_path = filedialog.asksaveasfilename(
				title="変換済みモデルを保存",
				defaultextension=original_path.suffix,
				initialfile=default_filename,
				filetypes=[
					("STL files", "*.stl"),
					("OBJ files", "*.obj"),
					("All files", "*.*")
				]
			)
			
			if save_path:
				child_model_transformed.save(save_path)
				print(f"\n変換済みモデルを保存しました: {Path(save_path).name}")

				# 子PPも同じ変換で保存（存在する場合）
				# ユーザーにPP保存の意志確認
				want_save_pp = messagebox.askyesno(
					"確認",
					"子特徴点 (PP) も同じ変換で保存しますか？"
				)
				child_pp_path = self.fitting_child_pp_path.get().strip()
				if want_save_pp and not child_pp_path:
					# 未選択ならここで選択を促す
					messagebox.showinfo("情報", "子PPファイルが未選択です。選択してください。")
					self.choose_fitting_child_pp()
					child_pp_path = self.fitting_child_pp_path.get().strip()
				if want_save_pp and child_pp_path:
					try:
						pp_points, pp_labels = self._parse_pp_file(child_pp_path)
						# L/M/Nは保存対象から除外する（仕様に応じて変更可）
						filtered_points = []
						filtered_labels = []
						for p, lb in zip(pp_points, pp_labels):
							if lb not in ['L', 'M', 'N']:
								filtered_points.append(p)
								filtered_labels.append(lb)
						if not filtered_points:
							filtered_points = pp_points
							filtered_labels = pp_labels
						filtered_points = np.asarray(filtered_points, dtype=float)
						# 変換を適用
						ones = np.ones((filtered_points.shape[0], 1))
						points_h = np.hstack([filtered_points, ones])
						points_t = (transform @ points_h.T).T[:, :3]
						# 保存先提案（元モデルの隣に .pp）
						pp_default = Path(save_path).with_suffix('').with_name(Path(save_path).stem + "_points.pp")
						pp_save_path = filedialog.asksaveasfilename(
							title="変換済み子特徴点 (PP) を保存",
							defaultextension=".pp",
							initialfile=pp_default.name,
							filetypes=[("PP files", "*.pp"), ("All files", "*.*")]
						)
						if pp_save_path:
							self.save_child_pp_file(pp_save_path, points_t, filtered_labels)
							print(f"変換済み子PPを保存しました: {Path(pp_save_path).name}")
						else:
							print("子PPの保存はキャンセルされました。")
					except Exception as e:
						print(f"子PPの保存に失敗: {e}")
				
				# 確認のため可視化
				response = messagebox.askyesno("確認", 
					f"変換済みモデルを保存しました。\n"
					f"保存先: {Path(save_path).name}\n\n"
					f"親モデルと変換済み子モデルを表示しますか？")
				
				if response:
					# 親モデルを読み込み
					parent_model_mesh = pv.read(self.fitting_parent_model_path.get())
					
					# 最終結果を可視化
					plotter = pv.Plotter()
					plotter.set_background('white')
					
					plotter.add_mesh(parent_model_mesh, color='red', opacity=0.5, show_edges=False, label='親モデル')
					plotter.add_mesh(child_model_transformed, color='cyan', opacity=0.5, show_edges=False, label='子モデル（フィッティング後）')
					
					plotter.add_axes()
					plotter.add_legend()
					
					print("\n[最終結果] 親モデル + 変換済み子モデルを表示中...")
					plotter.show()
			
		except Exception as e:
			messagebox.showerror("エラー", f"モデル保存中にエラーが発生しました:\n{str(e)}")
			print(f"エラー詳細: {e}")
			import traceback
			traceback.print_exc()
	
	def _get_principal_axis(self, mesh: pv.PolyData) -> tuple:
		"""メッシュの主軸（最大固有ベクトル）と重心を計算
		
		Args:
			mesh: PyVistaメッシュ
		
		Returns:
			tuple: (重心, 主軸の単位ベクトル)
		"""
		points = mesh.points
		
		# 重心を計算
		center = np.mean(points, axis=0)
		
		# 共分散行列を計算
		centered = points - center
		cov = np.cov(centered.T)
		
		# 固有値・固有ベクトルを計算
		eigvals, eigvecs = np.linalg.eigh(cov)
		
		# 最大固有値に対応する固有ベクトル（主軸）
		principal_axis = eigvecs[:, np.argmax(eigvals)]
		
		return center, principal_axis
	
	def _align_axis(self, mesh: pv.PolyData, src_axis: np.ndarray, tgt_axis: np.ndarray) -> tuple:
		"""主軸をターゲット軸に合わせる回転を適用
		
		Args:
			mesh: PyVistaメッシュ
			src_axis: ソース軸（単位ベクトル）
			tgt_axis: ターゲット軸（単位ベクトル）
		
		Returns:
			tuple: (回転後のメッシュ, 回転行列)
		"""
		# 外積で回転軸を計算
		v = np.cross(src_axis, tgt_axis)
		c = np.dot(src_axis, tgt_axis)
		
		# すでに平行な場合
		if np.linalg.norm(v) < 1e-8:
			return mesh.copy(), np.eye(3)
		
		# Rodriguesの回転公式で回転行列を計算
		vx = np.array([
			[0, -v[2], v[1]],
			[v[2], 0, -v[0]],
			[-v[1], v[0], 0]
		])
		R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
		# メッシュに回転適用
		rotated = mesh.copy()
		rotated.points = (R @ rotated.points.T).T
		return rotated, R

	def save_child_model_cache(self, cache_path: str, child_model: object) -> None:
		"""子モデルのキャッシュ保存（例: pickle形式）"""
		import pickle
		with open(cache_path, 'wb') as f:
			pickle.dump(child_model, f)

	def choose_save_child_pp_directory(self) -> str:
		"""子新PPファイル保存用ディレクトリ選択ダイアログ"""
		dir_path = filedialog.askdirectory(title="子新PPファイルの保存先ディレクトリを選択")
		return dir_path

	def save_child_pp_file(self, save_path: str, points: np.ndarray, labels: list) -> None:
		"""子特徴点（PP）をXML形式で保存する
		<PickedPoints><point x="..." y="..." z="..." active="1" name="..."/></PickedPoints>
		"""
		import xml.etree.ElementTree as ET
		root = ET.Element('PickedPoints')
		for pt, label in zip(points, labels):
			point_elem = ET.SubElement(root, 'point')
			point_elem.set('x', f"{pt[0]:.8f}")
			point_elem.set('y', f"{pt[1]:.8f}")
			point_elem.set('z', f"{pt[2]:.8f}")
			point_elem.set('active', "1")
			point_elem.set('name', label)
		tree = ET.ElementTree(root)
		tree.write(save_path, encoding="utf-8", xml_declaration=True)
		
		R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
		
		# 4x4変換行列に変換
		transform = np.eye(4)
		transform[:3, :3] = R
		
		# メッシュに適用
		mesh_rotated = mesh.copy()
		mesh_rotated.transform(transform, inplace=True)
		
		return mesh_rotated, R
# endregion MainMenuGUIクラス


# region メインエントリーポイント
def main():
	"""メインエントリーポイント"""
	# --- プラットフォーム固有の初期化 ---
	if IS_WINDOWS:
		# Windows: Open3DによるDPI設定の強制変更（ウィンドウ縮小現象）への対策
		# アプリケーション起動時から高DPI対応(System DPI Aware)としてマークすることで、
		# 途中で設定が切り替わるのを防ぎ、サイズ変化を抑制する。
		try:
			import ctypes
			ctypes.windll.shcore.SetProcessDpiAwareness(1)
		except Exception:
			pass
	elif IS_MACOS:
		# macOS: Retinaディスプレイ対応とアプリケーション名の設定
		try:
			# macOSでTkinterアプリをDockで正しく表示させる
			import subprocess
			# PyObjCが利用可能な場合、追加の最適化
		except Exception:
			pass

	app = MainMenuGUI()

	# macOS固有のウィンドウ設定
	if IS_MACOS:
		try:
			# macOSでウィンドウを前面に持ってくる
			app.lift()
			app.attributes('-topmost', True)
			app.after(100, lambda: app.attributes('-topmost', False))
		except Exception:
			pass

	app.mainloop()


if __name__ == "__main__":
	# macOS/Windows両対応: spawnモードでの並列処理に必要
	from multiprocessing import freeze_support
	freeze_support()
	main()
# endregion メインエントリーポイント

# ============================================================================
# ファイル終端
# ============================================================================
# FRS.Simulator.py 1.0
# 
# このファイルは約8100行のPythonコードで構成されています。
# 
# ■ ファイル構成:
#   - 目次: 行1-70
#   - インポート: 行71-97
#   - グローバル関数: 行98-123
#   - MainMenuGUIクラス: 行124-8098
#   - main()関数: 行8099-8112
#
# ■ 今後の改善案:
#   1. on_animate()の分割: 1400行→複数の小関数に分割
#   2. キャッシュシステムの強化: SQLiteベースのDBキャッシュ
#   3. ユニットテスト追加: 座標変換・幾何計算の検証
#   4. 設定ファイル外部化: パラメータをYAML/TOMLで管理
#   5. ログ機能強化: loggingモジュール導入
#
# ============================================================================
