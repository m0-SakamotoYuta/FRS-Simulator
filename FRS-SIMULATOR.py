# ============================================================================
# FRS.Simulator - 関節力学試験ロボットシステム FRS-2015 可視化プログラム
# ============================================================================
# バージョン: 1.0
# 作成日: 2024
# 説明: 股関節/膝関節の3Dモデル可視化、アニメーション、フィッティング
# ============================================================================
VERSION = "2.4"  # ここでバージョンを変更してください
# ============================================================================

# ============================================================================
# 【目次 - Table of Contents】
# ============================================================================
# このファイルは単一ファイルとして実装されており、約7900行のコードで構成されています。
# 以下の目次を使用して、必要なセクションに素早くジャンプできます。
#
# ■ 主要セクション
#   1. インポートと初期設定                           (行 30-50)
#   2. グローバルヘルパー関数                         (行 52-70)
#   3. MainMenuGUIクラス                              (行 72-7950)
#       3.1 初期化・UI構築                            (行 72-500)
#       3.2 ファイル選択メソッド                      (行 775-950)
#       3.3 可視化メソッド（近位・遠位・全体）        (行 951-2100)
#       3.4 変換行列管理                              (行 2101-2370)
#       3.5 キャッシュ管理                            (行 2369-2815)
#       3.6 フィッティング処理【重要】                (行 3043-4300)
#       3.7 ORG可視化                                 (行 4302-4670)
#       3.8 C Fix機能                                 (行 4671-5280)
#       3.9 アニメーション処理【最重要・1500行】      (行 5279-6750)
#       3.10 ボタン状態管理                           (行 6753-6845)
#       3.11 状態保存・復元                           (行 6846-7045)
#       3.12 座標系・幾何処理                         (行 7045-7440)
#
# ■ VSCode region/endregion マーカー
#   このファイルはVSCodeのregion/endregion機能に対応しています。
#   エディタの折りたたみ機能を活用して、セクションごとに表示/非表示を切り替えられます。
#
# ■ コード検索Tips
#   - メソッド定義: "def method_name"で検索
#   - 座標変換: "transform"または"matrix"で検索
#   - PyVistaプロッター: "pv.Plotter"で検索
#   - Open3D処理: "o3d."で検索
#   - ボタンコールバック: "on_"で検索（例: on_visualize_prox）
#
# ============================================================================

# region インポートと初期設定
import sys
import platform
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser, simpledialog
from tkinter import font as tkfont
import urllib.parse

# プラットフォーム判定
IS_WINDOWS = sys.platform == 'win32'
IS_MACOS = sys.platform == 'darwin'
IS_LINUX = sys.platform.startswith('linux')


def _mousewheel_units(event) -> int:
	"""<MouseWheel>/<Button-4/5> イベントを yview_scroll("units") 用の行数に変換する。

	Windows/Linux(X11): event.delta は ±120 の倍数 → /120 で行数
	macOS(Aqua):        event.delta は ±1..±数 の小さい生の値 → そのまま行数として扱う
	X11 の物理ホイール: <Button-4> = 上, <Button-5> = 下
	戻り値は yview_scroll に渡す符号付き行数（上方向がマイナス）。
	"""
	num = getattr(event, 'num', 0)
	if num == 4:
		return -1
	if num == 5:
		return 1
	delta = getattr(event, 'delta', 0) or 0
	if not delta:
		return 0
	if IS_MACOS:
		# Aqua の delta は既に「行に近い単位」。0にならないよう符号を保つ。
		step = -int(delta) if int(delta) != 0 else (-1 if delta > 0 else 1)
		return step
	# Windows / X11(<MouseWheel>): 120 単位
	step = int(-delta / 120)
	if step == 0:
		step = -1 if delta > 0 else 1
	return step

# ドラッグ&ドロップ対応（tkinterdnd2）
try:
	from tkinterdnd2 import TkinterDnD, DND_FILES
	_HAS_DND = True
except ImportError:	
	_HAS_DND = False
import json
from pathlib import Path
import openpyxl
import numpy as np
import pyvista as pv
import open3d as o3d # type: ignore
import traceback
import threading
import time
import datetime
from multiprocessing import Pool, cpu_count
import os
import pickle
import hashlib
import io
import xml.etree.ElementTree as ET
import copy
from typing import Optional

# FEM接触解析ソルバー
try:
	from fem_contact_solver_2 import (
		FEMContactSolver, MaterialProperties, ContactParameters,
		FEMResults, apply_fem_results_to_mesh,
	)
	_HAS_FEM = True
except ImportError:
	_HAS_FEM = False

# レポート自動生成モジュール (v2)
try:
	from fem_report_generator_v2 import generate_report_from_animation, detect_key_frames
	_HAS_REPORT = True
except ImportError:
	_HAS_REPORT = False

# 共有キャッシュマネージャ（NAS + ローカル2層キャッシュ）
try:
	from shared_cache_manager import SharedCacheManager, compute_content_hash
	_HAS_SHARED_CACHE = True
except ImportError:
	_HAS_SHARED_CACHE = False

# Open3Dの警告を抑制
o3d.utility.set_verbosity_level(o3d.utility.VerbosityLevel.Error)
# endregion インポートと初期設定

# region グローバルヘルパー関数
def save_child_pp_file(save_path: str, points: np.ndarray, labels: list) -> None:
	"""子特徴点（PP）をXML形式で保存する
	<PickedPoints><point x="..." y="..." z="..." active="1" name="..."/></PickedPoints>
	モジュールレベルのヘルパー（クラス外）
	"""
	root = ET.Element('PickedPoints')
	for pt, label in zip(points, labels):
		point_elem = ET.SubElement(root, 'point')
		point_elem.set('x', f"{pt[0]:.8f}")
		point_elem.set('y', f"{pt[1]:.8f}")
		point_elem.set('z', f"{pt[2]:.8f}")
		point_elem.set('active', "1")
		point_elem.set('name', label)
	tree = ET.ElementTree(root)
	tree.write(save_path, encoding="utf-8", xml_declaration=True)
# endregion グローバルヘルパー関数


class _LazyHeatmapList:
	"""ヒートマップ(全フレーム同一形状・距離のみ異なる)の遅延生成リスト。

	全フレームが共通ジオメトリ(base)を持ち、距離配列だけがフレームごとに異なる場合、
	19000超のPolyDataを先に全部生成すると読込後に十数秒待たされる。本クラスは
	__getitem__ で必要になったフレームのメッシュだけをその場で生成し、この待ちを無くす。
	距離配列は self.distances で直接参照でき、最小距離などはメッシュ生成なしで計算できる。
	"""

	def __init__(self, base_mesh, distances):
		self._base = base_mesh            # pv.PolyData or None（共通ジオメトリ）
		self.distances = list(distances)  # list[np.ndarray | None]

	def __len__(self):
		return len(self.distances)

	def __getitem__(self, i):
		d = self.distances[i]
		if self._base is None or d is None:
			return pv.PolyData()
		m = self._base.copy(deep=False)
		m['distance'] = np.asarray(d)
		return m

	def __iter__(self):
		for i in range(len(self.distances)):
			yield self[i]

	def min_distance(self):
		"""全フレームの距離最小値(=最深めり込み, 負)。メッシュ生成なしで計算。"""
		m = None
		for d in self.distances:
			if d is not None and len(d):
				dm = float(np.nanmin(d))
				m = dm if m is None else min(m, dm)
		return m if m is not None else -1.0


# region MainMenuGUIクラス
# tkinterdnd2がある場合はTkinterDnD.Tkを基底に、なければtk.Tkを使用
_BaseWindow = TkinterDnD.Tk if _HAS_DND else tk.Tk

class MainMenuGUI(_BaseWindow):
	"""
	FRS-2015 可視化プログラムのメインGUIクラス
	
	このクラスは以下の主要機能を提供します：
	- Simulator: 近位・遠位モデルの可視化とアニメーション
	- Fitting: RANSAC/ICPによる3Dモデル位置合わせ
	- ORG: RobotScanデータの可視化
	- C Fix: 座標系修正ツール
	
	【改修ポイント】
	- on_animate(): アニメーション処理（約1500行）
	- on_fitting_execute(): フィッティング処理
	- 各種可視化メソッド: on_visualize_prox(), on_visualize_dist(), on_visualize_all()
	"""
	
	# region 初期化とUI構築
	def __init__(self) -> None:
		super().__init__()
		# Window
		self.title(f"関節力学試験ロボットシステム FRS-2015：可視化プログラム Ver.{VERSION}")
		self.minsize(820, 680)

		# macOS固有のウィンドウ設定
		if IS_MACOS:
			try:
				# macOSでのネイティブメニューバー対応
				self.createcommand('tk::mac::ShowPreferences', lambda: None)
				# Retinaディスプレイでのスケーリング
				self.tk.call('tk', 'scaling', self.winfo_fpixels('1i') / 72.0)
			except Exception:
				pass

		# ウィンドウサイズを保存（後で復元用）
		self._initial_window_geometry = None

		# フォントを環境に合わせて統一
		self._setup_fonts()

		# Variables (Simulator)
		self.joint_var = tk.IntVar(value=0)  # 0: 未選択, 1: 股関節, 2: 膝関節
		self.prox_model_path = tk.StringVar(value="")
		self.prox_pp_abcd_path = tk.StringVar(value="")  # 近位ABCD特徴点
		self.prox_pp_olmn_path = tk.StringVar(value="")  # 近位OLMN特徴点
		self.dist_model_path = tk.StringVar(value="")
		self.dist_pp_abc_path = tk.StringVar(value="")   # 遠位ABC特徴点
		self.dist_pp_olmn_path = tk.StringVar(value="")  # 遠位OLMN特徴点
		self.transform_group_path = tk.StringVar(value="")
		
		# 可視化オプション
		self.show_ao_angle = tk.BooleanVar(value=True) # 骨盤A-O線と大腿骨A-O線の表示
		# 「関節全体モデル/特徴点を可視化」ウィンドウの表示/非表示チェックボックス（永続化）
		self.viz_show_prox_model = tk.BooleanVar(value=True)
		self.viz_show_prox_pp    = tk.BooleanVar(value=True)
		self.viz_show_prox_axes  = tk.BooleanVar(value=True)
		self.viz_show_dist_model = tk.BooleanVar(value=True)
		self.viz_show_dist_pp    = tk.BooleanVar(value=True)
		self.viz_show_dist_axes  = tk.BooleanVar(value=True)

		# Variables (Fitting)
		self.fitting_parent_model_path = tk.StringVar(value="")
		self.fitting_parent_region_path = tk.StringVar(value="")
		self.fitting_child_model_path = tk.StringVar(value="")
		self.fitting_child_region_path = tk.StringVar(value="")
		self.fitting_child_pp_path = tk.StringVar(value="")
		# ヒートマップ表示用モデル（任意）
		self.fitting_child_heatmap_model_path = tk.StringVar(value="")
		
		# Variables (ORG)
		self.org_model_path = tk.StringVar(value="")
		self.org_pp_path = tk.StringVar(value="")
		self.org_coord_x = tk.DoubleVar(value=0.0)
		self.org_coord_y = tk.DoubleVar(value=0.0)
		self.org_coord_z = tk.DoubleVar(value=0.0)

		# Variables (A-B)
		self.a_b_model_a_path = tk.StringVar(value="")
		self.a_b_model_b_path = tk.StringVar(value="")
		
		# Variables (HeatMap)
		self.heatmap_model_a_path = tk.StringVar(value="")
		self.heatmap_model_b_path = tk.StringVar(value="")
		self.heatmap_show_model_a = tk.BooleanVar(value=True)
		self.heatmap_show_model_b = tk.BooleanVar(value=True)
		
		# Variables (Randomizer)
		self.randomizer_input_path = tk.StringVar(value="")
		self.random_trans_range = tk.DoubleVar(value=50.0)
		self.random_rot_range = tk.DoubleVar(value=180.0)

		# Variables (Cartilage Separation / 軟骨分離)
		# 近位（骨盤）
		self.cs_prox_model1_whole_path = tk.StringVar(value="")   # 骨盤: 骨+軟骨 全体
		self.cs_prox_model1_region_path = tk.StringVar(value="")  # 骨盤: 骨+軟骨 任意領域
		self.cs_prox_model2_whole_path = tk.StringVar(value="")   # 骨盤: 骨のみ 全体
		self.cs_prox_model2_region_path = tk.StringVar(value="")  # 骨盤: 骨のみ 任意領域
		# 遠位（大腿骨）
		self.cs_dist_model1_whole_path = tk.StringVar(value="")   # 大腿骨: 骨+軟骨 全体
		self.cs_dist_model1_region_path = tk.StringVar(value="")  # 大腿骨: 骨+軟骨 任意領域
		self.cs_dist_model2_whole_path = tk.StringVar(value="")   # 大腿骨: 骨のみ 全体
		self.cs_dist_model2_region_path = tk.StringVar(value="")  # 大腿骨: 骨のみ 任意領域

		# 軟骨分離 専用RANSACパラメータ
		self.cs_ransac_distance_threshold = tk.DoubleVar(value=1.0)
		self.cs_ransac_max_iterations = tk.IntVar(value=1000)
		self.cs_ransac_confidence = tk.DoubleVar(value=0.99)

		# 軟骨分離 専用ICPパラメータ
		self.cs_icp_threshold = tk.DoubleVar(value=5.0)
		self.cs_icp_max_iterations = tk.IntVar(value=2000)
		self.cs_sample_points = tk.IntVar(value=10000)

		# 軟骨分離 距離閾値（セグメンテーション用）
		self.cs_distance_threshold = tk.DoubleVar(value=0.5)  # mm
		# 軟骨分離 スケール補正（3Dスキャナ vs CT のスケール差自動補正）
		self.cs_enable_scaling = tk.BooleanVar(value=False)

		# 軟骨分離 可視化オプション
		self.cs_bone_color = "#F5DEB3"       # 骨: 小麦色
		self.cs_cartilage_color = "#00CED1"  # 軟骨: ダークターコイズ
		self.cs_bone_opacity = tk.DoubleVar(value=0.4)
		self.cs_cartilage_opacity = tk.DoubleVar(value=1.0)

		# Simulator用: 軟骨モデルパス（任意・分離後に設定）
		self.prox_cartilage_model_path = tk.StringVar(value="")  # 近位軟骨モデル
		self.dist_cartilage_model_path = tk.StringVar(value="")  # 遠位軟骨モデル

		# Variables (FEM解析)
		# 入力メッシュパス（複合モデル: 骨+軟骨）
		self.fem_prox_bone_path = tk.StringVar(value="")       # 近位骨モデル
		self.fem_prox_cartilage_path = tk.StringVar(value="")  # 近位軟骨モデル
		self.fem_dist_bone_path = tk.StringVar(value="")       # 遠位骨モデル
		self.fem_dist_cartilage_path = tk.StringVar(value="")  # 遠位軟骨モデル
		# 材料パラメータ（軟骨）
		self.fem_cart_E = tk.DoubleVar(value=10.0)        # ヤング率 [MPa]
		self.fem_cart_nu = tk.DoubleVar(value=0.45)       # ポアソン比
		self.fem_cart_thickness = tk.DoubleVar(value=2.0)  # 軟骨厚さ [mm]
		# 材料パラメータ（骨 — 参考表示）
		self.fem_bone_E = tk.DoubleVar(value=17000.0)     # ヤング率 [MPa]
		self.fem_bone_nu = tk.DoubleVar(value=0.3)        # ポアソン比
		# 接触パラメータ
		self.fem_penalty_stiffness = tk.DoubleVar(value=500.0)  # ペナルティ剛性 [MPa/mm]
		self.fem_contact_tolerance = tk.DoubleVar(value=2.0)    # 接触判定閾値 [mm]
		# 解析オプション
		self.fem_boundary_mode = tk.StringVar(value="auto")     # 境界条件モード
		self.fem_max_nodes = tk.IntVar(value=30000)             # 最大節点数
		# 可視化オプション
		self.fem_scalar_name = tk.StringVar(value="contact_pressure")
		self.fem_cmap = tk.StringVar(value="jet")
		self.fem_show_dist_mesh = tk.BooleanVar(value=True)
		self.fem_show_bone = tk.BooleanVar(value=True)
		# Simulator連携
		self.show_fem_analysis = tk.BooleanVar(value=False)
		# ランタイム状態（非永続）
		self.fem_last_results = None
		self.fem_last_prox_cart_mesh = None
		self.fem_last_dist_cart_mesh = None
		self.fem_last_prox_bone_mesh = None
		self.fem_last_dist_bone_mesh = None
		self.fem_is_running = False

		# 共有キャッシュ設定（NAS + ローカル2層）
		self.cache_nas_path = tk.StringVar(value="")  # NASキャッシュパス（空=NAS無効）
		self.cache_enabled = tk.BooleanVar(value=True)  # キャッシュ有効/無効
		self._shared_fem_cache = None       # FEM用SharedCacheManager
		self._shared_overlap_cache = None   # Overlap/Heatmap用SharedCacheManager
		# ※ _init_shared_cache() は _load_state() でNASパスを復元した後に呼ぶ

		# Fittingパラメータ
		# RANSACパラメータ
		self.ransac_distance_threshold = tk.DoubleVar(value=1.0)  # 距離閾値 (mm)
		self.ransac_max_iterations = tk.IntVar(value=1000)  # 最大反復回数
		self.ransac_confidence = tk.DoubleVar(value=0.99)  # 信頼度
		
		# ICPパラメータ
		self.icp_threshold = tk.DoubleVar(value=5.0)  # 閾値 (mm)
		self.icp_max_iterations = tk.IntVar(value=2000)  # 最大反復回数
		self.sample_points = tk.IntVar(value=10000)  # サンプル点数
		
		# 可視化オプション（色設定）
		self.prox_color = "#B0C4DE"  # デフォルト: 水色
		self.dist_color = "#FFB6C1"  # デフォルト: ピンク
		self.overlap_color = "#BC0051"  # デフォルト: 黄色
		
		# 関節範囲の半径（球体抽出用）
		self.prox_radius = tk.DoubleVar(value=30.0)  # 近位側の半径（mm）
		self.dist_radius = tk.DoubleVar(value=30.0)  # 遠位側の半径（mm）
		
		# O'原点位置オフセット（ローカル座標系に沿った移動量）
		self.prox_offset_x = tk.DoubleVar(value=0.0)  # 近位O' X方向オフセット（mm）
		self.prox_offset_y = tk.DoubleVar(value=0.0)  # 近位O' Y方向オフセット（mm）
		self.prox_offset_z = tk.DoubleVar(value=0.0)  # 近位O' Z方向オフセット（mm）
		self.dist_offset_x = tk.DoubleVar(value=0.0)  # 遠位O' X方向オフセット（mm）
		self.dist_offset_y = tk.DoubleVar(value=0.0)  # 遠位O' Y方向オフセット（mm）
		self.dist_offset_z = tk.DoubleVar(value=0.0)  # 遠位O' Z方向オフセット（mm）

		# ============================================================
		# Variables (Knee Simulator / FRS-2015 ロボット準拠 膝関節)
		# ------------------------------------------------------------
		# 座標系定義: 木村圭 学位論文 第3章 / 木村・藤江 機械学会論文集 2015 /
		#             藤江「動作原理と改良」2024。
		#  - 大腿骨座標系 Cf: 原点=MCL/LCL付着部の中点, Z=LCL→MCL(右膝で外側正),
		#    Y=大腿骨骨軸(≒基準座標系Coのy軸=L,M,N由来のCcl姿勢)をZ⊥面へ投影(近位正),
		#    X=Y×Z
		#  - 脛骨座標系 Ct: 伸展位でCfに一致。動物膝はCfをZ軸まわりに屈曲角φ(=W_scan)回転
		# 特徴点(L,M,N,MCL,LCL)は「ロボット取付け初期状態(組立)3Dスキャン」上にある。
		# 大腿骨/脛骨モデルは試験後の別スキャンで、ICP/RANSACで初期スキャンへ位置合わせする。
		# ============================================================
		self.knee_initial_scan_path = tk.StringVar(value="")   # 初期状態(組立)スキャン
		self.knee_initial_pp_path = tk.StringVar(value="")     # 初期スキャン特徴点(L,M,N,MCL,LCL)
		self.knee_femur_model_path = tk.StringVar(value="")    # 大腿骨モデル(試験後・別スキャン)
		self.knee_tibia_model_path = tk.StringVar(value="")    # 脛骨モデル(試験後・別スキャン)
		self.knee_transform_path = tk.StringVar(value="")      # 変位・姿勢変化データ(xlsx/kkr)
		self.knee_side_var = tk.IntVar(value=1)                # 1: 右膝, 2: 左膝
		self.knee_w_scan_deg = tk.DoubleVar(value=0.0)         # 伸展位屈曲角 φ (=スキャン時W軸角度, deg)
		# 表示色（hipの可視化オプションとは独立。タブごとにスナップショット保存される）
		self.knee_femur_color_var = tk.StringVar(value="#DEB887")  # 大腿骨: 既定=肌色
		self.knee_tibia_color_var = tk.StringVar(value="#ADD8E6")  # 脛骨: 既定=水色
		# ICP/RANSAC 位置合わせ用の領域(任意領域)ファイル
		self.knee_reg_femur_src_path = tk.StringVar(value="")  # 大腿骨モデル側 位置合わせ領域
		self.knee_reg_femur_tgt_path = tk.StringVar(value="")  # 初期スキャン側 大腿骨領域
		self.knee_reg_tibia_src_path = tk.StringVar(value="")  # 脛骨モデル側 位置合わせ領域
		self.knee_reg_tibia_tgt_path = tk.StringVar(value="")  # 初期スキャン側 脛骨領域
		# 位置合わせ結果(4x4変換行列; 各モデル→初期スキャン座標系)。未計算はNone
		self._knee_femur_reg_T = None
		self._knee_tibia_reg_T = None
		# 位置合わせパラメータ（大腿骨・脛骨で別々に保持。プリセットで保存/呼出可）
		self.knee_reg_pvars = {}        # {bone: {key: tkVar}}
		self.knee_reg_method_var = {}   # {bone: StringVar} "ransac"/"pca"/"manual"
		self.knee_reg_scaling_var = {}  # {bone: BooleanVar}
		self.knee_reg_preview_var = {}  # {bone: BooleanVar}
		self.knee_preset_sel = {}       # {bone: StringVar} 選択中プリセット名
		for bone in ("femur", "tibia"):
			d = {}
			for key, kind, default, _lbl in self._knee_param_spec():
				d[key] = tk.IntVar(value=int(default)) if kind == "int" else tk.DoubleVar(value=float(default))
			self.knee_reg_pvars[bone] = d
			self.knee_reg_method_var[bone] = tk.StringVar(value="ransac")
			self.knee_reg_scaling_var[bone] = tk.BooleanVar(value=False)
			self.knee_reg_preview_var[bone] = tk.BooleanVar(value=True)
			self.knee_preset_sel[bone] = tk.StringVar(value="")
		self.knee_preset_combo = {}     # {bone: ttk.Combobox}（UI構築時に格納）
		self._knee_presets = self._load_knee_presets()  # 名前→パラメータdict

		# 複数「試験」タブ（各タブ=独立した状態スナップショット。UI/コードは共通で全タブに反映）
		self._knee_default_snap = self._knee_snapshot_current()  # 既定値スナップショット（新規タブ用）
		self._knee_tabs = []            # [{'name': str, 'snapshot': dict}]
		self._knee_active_tab = 0
		self._knee_tabbar_frame = None  # UI構築時に格納
		self._knee_tab_buttons = []     # タブバー上のボタン（並び順=タブ順、ドラッグ入替の座標判定用）
		self._knee_tab_drag = None      # ドラッグ中の入替情報 {'from', 'x', 'moved'}

		# ============================================================
		# ankle simulator（足関節・ArUcoマーカートラッキング方式）
		# 特徴: 膝/股関節と異なり「近位/遠位2骨」ではなく、任意N本(脛骨,距骨,他)の
		# 動きをRGB-Dカメラ+骨に刺したピン+ArUcoマーカーで追跡する。
		# データフロー:
		#   1) 骨にM3ピンを刺しArUcoプレートを固定 → 初期状態スキャン
		#   2) RGB-D動画で背屈/底屈を撮影
		#   3) ArUco検出→PnPで各マーカーの6DOFを時系列取得
		#   4) 試験後解剖: 各骨を個別スキャンし「マーカー座標系→骨座標系」の
		#      剛体変換(marker_to_bone_T)を算出
		#   5) マーカー時系列に marker_to_bone_T を掛けて骨の動きを復元
		#   6) 骨ペア間の接触面ヒートマップ(膝と同じLUT)
		# ============================================================
		self.ankle_initial_scan_path = tk.StringVar(value="")  # 初期状態(ピン+マーカー装着後)スキャン
		self.ankle_video_path = tk.StringVar(value="")         # RGBビデオ(mp4/avi)
		self.ankle_depth_path = tk.StringVar(value="")         # 深度データ(bag/npz/mkv等)
		self.ankle_camera_intrinsics_path = tk.StringVar(value="")  # カメラ内部パラメータ(json/yaml)
		self.ankle_aruco_dict_var = tk.StringVar(value="DICT_4X4_50")  # ArUco辞書
		self.ankle_marker_size_mm = tk.DoubleVar(value=20.0)   # マーカー実寸(mm)
		self.ankle_pose_series_path = tk.StringVar(value="")   # 事前計算済み姿勢時系列(任意, npz/csv)
		# RealSense D405 ライブ撮影設定 (友人 make_date_movie_D405.py の設定を既定値化)
		self.ankle_rs_resolution = tk.StringVar(value="1280x720@15")   # 解像度@fps
		self.ankle_rs_discard_frames = tk.IntVar(value=45)             # 録画開始前の破棄フレーム
		self.ankle_rs_manual_exposure = tk.BooleanVar(value=False)     # 深度手動露光
		self.ankle_rs_exposure_val = tk.IntVar(value=5000)             # 手動露光値
		# カラー露光 (ArUco のモーションブラーを決める最重要パラメータ)
		# 実測: ブラー 4px で回転誤差 2.1倍、8px で 8.3倍かつ検出率 29% に激減。
		# 露光[ms] × マーカーの画面上の速度[px/s] = ブラー[px] なので、露光を直接短くするのが効く。
		self.ankle_rs_color_manual_exposure = tk.BooleanVar(value=True)
		self.ankle_rs_color_exposure_us = tk.IntVar(value=5000)        # マイクロ秒 (5000us = 5ms)
		self.ankle_rs_color_gain = tk.IntVar(value=64)                 # 露光を詰めた分ゲインで補う
		self.ankle_rs_status = tk.StringVar(value="(未接続)")
		# タブ共通の位置合わせパラメータ (膝と同構造・全骨で共有)
		self.ankle_reg_ransac_distance = tk.DoubleVar(value=1.0)     # RANSAC距離(mm)
		self.ankle_reg_ransac_max_iter = tk.IntVar(value=1000)       # RANSAC反復
		self.ankle_reg_ransac_confidence = tk.DoubleVar(value=0.99)  # RANSAC信頼度
		self.ankle_reg_icp_threshold = tk.DoubleVar(value=5.0)       # ICP距離(mm)
		self.ankle_reg_icp_max_iter = tk.IntVar(value=2000)          # ICP反復
		self.ankle_reg_sample_points = tk.IntVar(value=100000)       # サンプル点数
		self.ankle_reg_voxel_size = tk.DoubleVar(value=0.0)          # ボクセル(mm, 0=自動)
		self.ankle_reg_preview_var = tk.BooleanVar(value=True)       # 位置合わせ経過表示
		# N本可変の骨リスト: 各要素は dict
		#   {name, aruco_id, model_path, post_scan_path, color,
		#    method('ransac'|'pca'|'manual'), enable_scaling(bool),
		#    marker_to_bone_T(4x4 list or None), reg_T(4x4 list or None)}
		self.ankle_bones = []
		self._ankle_selected_bone = 0  # 現在選択中の骨index
		# 骨ヒートマップ骨対: 骨ラベル文字列で保持 (「(未選択)」または「N. 骨名」)
		self.ankle_heatmap_prox_var = tk.StringVar(value="(未選択)")
		self.ankle_heatmap_dist_var = tk.StringVar(value="(未選択)")
		# 骨リスト編集UIの参照（構築時に格納）
		self._ankle_bones_listbox = None
		self._ankle_bone_editor_widgets = {}
		# 位置合わせ結果はankle_bonesの各要素内 reg_T に保持
		# ArUco/PnP検出パラメータ
		self.ankle_ref_frame = tk.IntVar(value=0)         # 参照フレーム t=0 (Stage 5で使用)
		self.ankle_detect_stride = tk.IntVar(value=1)     # フレーム間引き (1=全フレーム)
		self.ankle_detection_status = tk.StringVar(value="(未実行)")
		# 可視化オプション
		self.ankle_smooth_enable = tk.BooleanVar(value=True)   # 平滑化 ON/OFF
		# カットオフ [Hz]: 15fps 録画なら Nyquist 7.5Hz。関節試験は低速なので 2〜3Hz が目安
		self.ankle_smooth_cutoff_hz = tk.DoubleVar(value=2.5)
		self.ankle_reject_outliers = tk.BooleanVar(value=True)  # Hampel 外れ値除去
		self.ankle_show_markers = tk.BooleanVar(value=True)    # ArUcoマーカー軸を可視化
		# 診断モード: 骨重心をマーカー位置に強制配置 (キャリブを無視する切り分け用。既定OFF)
		self.ankle_recenter_meshes = tk.BooleanVar(value=False)
		# depth_scale の自動補正 (SDK報告値が実データと乖離する場合)
		self.ankle_depth_scale_autofix = tk.BooleanVar(value=True)
		# 平面フィット領域の拡張倍率 (マーカーの外側=台座まで使って法線精度を上げる)
		self.ankle_plane_expand = tk.DoubleVar(value=2.2)
		# 姿勢推定手法: "rgb" | "fusion" | "depth-corners"
		# 実測 (2026-08-29) では rgb が 融合より 7〜14倍 安定だったため rgb を既定にする。
		# 深度は「IPPEの表裏判定」と「深度スケール検証」にのみ使う。
		self.ankle_pose_method = tk.StringVar(value="rgb  ★推奨 (最も安定)")
		# 手法比較を行うフレーム数 (0 で無効)。冒頭だけ3手法を計算して比較レポートを出す。
		self.ankle_compare_frames = tk.IntVar(value=300)
		# 動作モード:
		#   "self_pose" = 新プラン (マーカー付き骨単独スキャンから T_L←Mk 自己決定 → 絶対姿勢アニメ)  【デフォルト】
		#   "original"  = 原プラン (組立スキャン + 位置合わせで T_W←L 決定 → 相対運動アニメ)
		#   "simple"    = 簡易版 (3Dスキャン無し, クランプArUcoで関節座標系Cj校正 → 骨マーカー動揺性)
		self.ankle_workflow_mode = tk.StringVar(value="self_pose")
		# 新プラン用: W 座標系の選択  "camera_ref" (カメラref) or "boneA_ref" (骨A基準)
		self.ankle_self_pose_world = tk.StringVar(value="camera_ref")
		# 簡易版用: クランプ (関節座標系校正用) ArUco ID
		self.ankle_clamp_aruco_id = tk.IntVar(value=100)
		# 簡易版校正の記録時間 (各軸)
		self.ankle_axis_calib_seconds = tk.DoubleVar(value=4.0)
		# 簡易版で得られた関節座標系 Cj: 4x4 (R + origin)。 None なら未校正
		self._ankle_joint_frame_Cj = None
		# 各軸校正の記録 (最終) と姿勢時系列
		self._ankle_axis_calib_results = {}   # {"ML": {axis_dir(3), positions(N,3), residual}, ...}
		self.ankle_axis_calib_status = tk.StringVar(value="(未校正)")
		# 姿勢時系列キャッシュ: {tab_name: {frame_count, timestamps, intrinsics, marker_size_mm,
		#   aruco_dict, source, bones: {aruco_id: {poses(N,4,4), detected(N,), reproj_err(N,)}}}}
		self._ankle_pose_cache = {}
		# 検出処理のキャンセルフラグ (progress dialog)
		self._ankle_detect_cancel = False

		# 既定値スナップショット（新規タブ用）を確保
		self._ankle_default_snap = self._ankle_snapshot_current()
		self._ankle_tabs = []            # [{'name': str, 'snapshot': dict}]
		self._ankle_active_tab = 0
		self._ankle_tabbar_frame = None
		self._ankle_tab_buttons = []
		self._ankle_tab_drag = None

		# 関節種別ごとに切替えるUIウィジェットの参照（ラベル変更用）
		self._joint_widgets = {}

		# 状態復元（可能なら）
		self._load_state()
		# knee simulator の状態復元（hipの関節ラジオと独立）
		self._load_knee_state()
		# ankle simulator の状態復元（膝/股と独立）
		self._load_ankle_state()
		# 直前の関節種別を記録（関節切替時の差分判定用）
		self._prev_joint = self.joint_var.get()

		# 可視化チェックボックスの状態は、可視化ウィンドウを閉じる時に _save_state で
		# 永続化する（trace_add で即時保存すると VTK のcallback 内で再入が起きて
		# PyVista ウィンドウが落ちることがあるため）。
		# NASパスを含む状態復元後にキャッシュを初期化
		self._init_shared_cache()

		# スタイル設定（LabelFrameのタイトルを太字に）
		style = ttk.Style()
		style.configure("Bold.TLabelframe.Label", font=(self.ui_font_family, 9, "bold"))

		# Layout root container
		main_container = ttk.Frame(self, padding=12)
		main_container.grid(row=0, column=0, sticky="nsew")
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		main_container.columnconfigure(0, weight=1)
		main_container.rowconfigure(1, weight=1)

		# Title
		title_label = ttk.Label(
			main_container,
			text="関節力学試験ロボットシステム FRS-2015：可視化プログラム",
			font=(self.ui_font_family, 13, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", pady=(0, 12))

		# タブコントロールを作成
		self.notebook = ttk.Notebook(main_container)
		self.notebook.grid(row=1, column=0, sticky="nsew")

		# タブ1: hip simulator（旧「Simulator」。股関節用可視化）
		self.simulator_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.simulator_tab, text="hip simulator")

		# タブ1b: knee simulator（FRS-2015 ロボット準拠 膝関節用可視化）
		self.knee_simulator_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.knee_simulator_tab, text="knee simulator")

		# タブ1c: ankle simulator（足関節・ArUcoマーカートラッキング方式）
		self.ankle_simulator_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.ankle_simulator_tab, text="ankle simulator")

		# タブ2: Cache Settings
		self.cache_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.cache_tab, text="Cache Settings")

		# タブ3: Fitting
		self.fitting_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.fitting_tab, text="Fitting")

		# タブ4: 軟骨分離 (Cartilage Separation)
		self.cartilage_sep_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.cartilage_sep_tab, text="軟骨分離")

		# タブ5: FEM解析
		self.fem_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.fem_tab, text="FEM解析")

		# タブ6: Utilities (サブタブを含むコンテナ)
		self.utilities_tab = ttk.Frame(self.notebook, padding=12)
		self.notebook.add(self.utilities_tab, text="Utilities")
		
		# Utilities内のサブノートブック
		self.utilities_notebook = ttk.Notebook(self.utilities_tab)
		self.utilities_notebook.pack(fill="both", expand=True)

		# Utilities > ORG
		self.org_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.org_tab, text="ORG")

		# Utilities > A-B
		self.a_b_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.a_b_tab, text="A-B")

		# Utilities > HeatMap
		self.heatmap_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.heatmap_tab, text="HeatMap")

		# Utilities > C fix
		self.c_fix_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.c_fix_tab, text="C fix")

		# Utilities > STL2ASC
		self.stl2asc_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.stl2asc_tab, text="STL2ASC")

		# Utilities > Randomizer
		self.randomizer_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.randomizer_tab, text="Randomizer")

		# Utilities > PosChecker
		self.pos_checker_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.pos_checker_tab, text="PosChecker")

		# Utilities > Coordinate Creater
		self.coordinate_creater_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.coordinate_creater_tab, text="Coordinate Creater")

		# Utilities > pp merge
		self.pp_merge_tab = ttk.Frame(self.utilities_notebook, padding=12)
		self.utilities_notebook.add(self.pp_merge_tab, text="pp merge")


		# Simulatタブ (hip simulator) のコンテンツを作成
		self._create_simulator_tab()

		# knee simulatorタブのコンテンツを作成
		self._create_knee_simulator_tab()

		# ankle simulatorタブのコンテンツを作成
		self._create_ankle_simulator_tab()

		# Fittingタブのコンテンツを作成
		self._create_fitting_tab()

		# 軟骨分離タブのコンテンツを作成
		self._create_cartilage_separation_tab()

		# FEM解析タブのコンテンツを作成
		self._create_fem_tab()

		# ORGタブのコンテンツを作成
		# self._create_org_tab() は存在しないためダミー実装を呼び出し
		self._create_org_tab()

		# A-Bタブのコンテンツを作成
		self._create_a_b_tab()

		# HeatMapタブのコンテンツを作成
		self._create_heatmap_tab()

		# C fixタブのコンテンツを作成
		self._create_c_fix_tab()

		# STL2ASCタブのコンテンツを作成
		self._create_stl2asc_tab()

		# Randomizerタブのコンテンツを作成
		self._create_randomizer_tab()

		# PosCheckerタブのコンテンツを作成
		self._create_pos_checker_tab()

		# Coordinate Createrタブのコンテンツを作成
		self._create_coordinate_creater_tab()

		# pp mergeタブのコンテンツを作成
		self._create_pp_merge_tab()

		# Cache Settingsタブのコンテンツを作成
		self._create_cache_tab()

		# 初期状態のボタン活性制御
		self.update_button_states()

		# Variable trace for dynamic state updates
		self.prox_model_path.trace_add("write", lambda *_: self.update_button_states())
		self.prox_pp_abcd_path.trace_add("write", lambda *_: self.update_button_states())
		self.prox_pp_olmn_path.trace_add("write", lambda *_: self.update_button_states())
		self.dist_model_path.trace_add("write", lambda *_: self.update_button_states())
		self.dist_pp_abc_path.trace_add("write", lambda *_: self.update_button_states())
		self.dist_pp_olmn_path.trace_add("write", lambda *_: self.update_button_states())
		self.transform_group_path.trace_add("write", lambda *_: self.update_button_states())

		# 終了時に状態保存
		self.protocol("WM_DELETE_WINDOW", self._on_close)
		
		# ウィンドウが完全に表示された後にサイズを保存
		self.after(100, self._save_initial_geometry)
	# endregion 初期化とUI構築

	# region 共有キャッシュ初期化
	def _init_shared_cache(self) -> None:
		"""SharedCacheManagerを初期化する。

		NASパスが設定されていればNAS+ローカル2層、
		未設定ならローカルのみのキャッシュを構築。
		"""
		try:
			nas_path = self.cache_nas_path.get().strip() or None
			base_dir = Path(__file__).parent

			# FEM用キャッシュ
			self._shared_fem_cache = SharedCacheManager(
				nas_dir=nas_path + "/fem_cache" if nas_path else None,
				namespace="fem",
				max_nas_gb=10.0,
			)

			# Overlap/Heatmap用キャッシュ
			self._shared_overlap_cache = SharedCacheManager(
				nas_dir=nas_path + "/overlap_cache" if nas_path else None,
				namespace="overlap",
				max_nas_gb=10.0,
			)

			if nas_path:
				print(f"[SharedCache] NASパス設定: {nas_path}")
				# NAS接続チェック（ログ目的のみ）をバックグラウンドで実行（UIをブロックしない）
				fem_cache_ref = self._shared_fem_cache
				overlap_cache_ref = self._shared_overlap_cache

				def _check_nas_log():
					if fem_cache_ref.is_nas_available():
						print("[SharedCache] NAS接続OK (FEM)")
					else:
						print("[SharedCache] NAS接続失敗 → ローカルフォールバック (FEM)")
					if overlap_cache_ref.is_nas_available():
						print("[SharedCache] NAS接続OK (Overlap)")
					else:
						print("[SharedCache] NAS接続失敗 → ローカルフォールバック (Overlap)")

				threading.Thread(target=_check_nas_log, daemon=True).start()
			else:
				print("[SharedCache] NAS未設定 → ローカルキャッシュのみ")
		except Exception as e:
			print(f"[SharedCache] 初期化エラー: {e}")
			self._shared_fem_cache = None
			self._shared_overlap_cache = None

	def _reinit_shared_cache(self) -> None:
		"""NASパス変更時にキャッシュを再初期化する"""
		self._init_shared_cache()
	# endregion 共有キャッシュ初期化

	# region UI構築メソッド
	def _create_simulator_tab(self) -> None:
		"""Simulatorタブのコンテンツを作成"""
		# スクロール可能なメインフレーム
		canvas = tk.Canvas(self.simulator_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.simulator_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)

		scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
		)

		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)

		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")

		# マウスホイールでスクロール（Mac/Windows/Linux 両対応）
		def _on_mousewheel(event):
			step = _mousewheel_units(event)
			if step:
				canvas.yview_scroll(step, "units")
		canvas.bind_all("<MouseWheel>", _on_mousewheel)
		canvas.bind_all("<Button-4>", _on_mousewheel)
		canvas.bind_all("<Button-5>", _on_mousewheel)

		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# Section: 使用する関節
		joint_frame = ttk.LabelFrame(container, text="使用する関節", style="Bold.TLabelframe")
		joint_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 12))
		joint_frame.columnconfigure(0, weight=1)

		rb1 = ttk.Radiobutton(
			joint_frame, text="1. 股関節", value=1, variable=self.joint_var, command=self._on_joint_changed
		)
		rb2 = ttk.Radiobutton(
			joint_frame, text="2. 膝関節 (ISB仕様)", value=2, variable=self.joint_var, command=self._on_joint_changed
		)
		# 左詰め・上下配置
		rb1.grid(row=0, column=0, sticky="w", padx=12, pady=6)
		rb2.grid(row=1, column=0, sticky="w", padx=12, pady=6)

		# Section: 近位ファイル選択
		prox_file_frame = ttk.LabelFrame(container, text=self._joint_label("prox_file_frame"), style="Bold.TLabelframe")
		self._joint_widgets["prox_file_frame"] = prox_file_frame
		prox_file_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			prox_file_frame.columnconfigure(i, weight=1)

		# 近位モデル (OBJ/STL)
		self._add_file_row(
			parent=prox_file_frame,
			row=0,
			label_text="近位モデル (OBJ/STL)",
			textvariable=self.prox_model_path,
			command=self.choose_prox_model, 
		)

		# 近位特徴点ABCD (PP)
		self._joint_widgets["prox_pp_abcd_label"] = self._add_file_row(
			parent=prox_file_frame,
			row=1,
			label_text=self._joint_label("prox_pp_abcd_label"),
			textvariable=self.prox_pp_abcd_path,
			command=self.choose_prox_pp_abcd,
		)

		# 近位特徴点OLMN (PP)
		self._joint_widgets["prox_pp_olmn_label"] = self._add_file_row(
			parent=prox_file_frame,
			row=2,
			label_text=self._joint_label("prox_pp_olmn_label"),
			textvariable=self.prox_pp_olmn_path,
			command=self.choose_prox_pp_olmn,
		)

		# 近位軟骨モデル（任意）
		self._add_file_row(
			parent=prox_file_frame,
			row=3,
			label_text="近位軟骨モデル (STL/OBJ) ※任意",
			textvariable=self.prox_cartilage_model_path,
			command=self.choose_prox_cartilage_model,
		)

		# 特徴点の意味を表示する小ラベル
		prox_legend = ttk.Label(prox_file_frame, text=self._joint_label("prox_legend"), foreground="gray", font=(self.ui_font_family, 8))
		prox_legend.grid(row=4, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 6))
		self._joint_widgets["prox_legend"] = prox_legend

		# Section: 遠位ファイル選択
		dist_file_frame = ttk.LabelFrame(container, text=self._joint_label("dist_file_frame"), style="Bold.TLabelframe")
		self._joint_widgets["dist_file_frame"] = dist_file_frame
		dist_file_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 12))
		for i in range(3):
			dist_file_frame.columnconfigure(i, weight=1)

		# 遠位モデル (OBJ/STL)
		self._add_file_row(
			parent=dist_file_frame,
			row=0,
			label_text="遠位モデル (OBJ/STL)",
			textvariable=self.dist_model_path,
			command=self.choose_dist_model,
		)

		# 遠位特徴点ABC/ABCD (PP) - 関節種別でラベル切替
		self._joint_widgets["dist_pp_abc_label"] = self._add_file_row(
			parent=dist_file_frame,
			row=1,
			label_text=self._joint_label("dist_pp_abc_label"),
			textvariable=self.dist_pp_abc_path,
			command=self.choose_dist_pp_abc,
		)

		# 遠位特徴点OLMN (PP)
		self._joint_widgets["dist_pp_olmn_label"] = self._add_file_row(
			parent=dist_file_frame,
			row=2,
			label_text=self._joint_label("dist_pp_olmn_label"),
			textvariable=self.dist_pp_olmn_path,
			command=self.choose_dist_pp_olmn,
		)

		# 遠位軟骨モデル（任意）
		self._add_file_row(
			parent=dist_file_frame,
			row=3,
			label_text="遠位軟骨モデル (STL/OBJ) ※任意",
			textvariable=self.dist_cartilage_model_path,
			command=self.choose_dist_cartilage_model,
		)

		# 特徴点の意味を表示する小ラベル
		dist_legend = ttk.Label(dist_file_frame, text=self._joint_label("dist_legend"), foreground="gray", font=(self.ui_font_family, 8))
		dist_legend.grid(row=4, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 6))
		self._joint_widgets["dist_legend"] = dist_legend

		# 操作ボタン段（可視化）
		action_frame = ttk.Frame(container)
		action_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.visualize_prox_button = ttk.Button(action_frame, text="近位モデル/特徴点を可視化", command=self.on_visualize_prox)
		self.visualize_prox_button.grid(row=0, column=0, sticky="w")

		self.visualize_dist_button = ttk.Button(action_frame, text="遠位モデル/特徴点を可視化", command=self.on_visualize_dist)
		self.visualize_dist_button.grid(row=1, column=0, sticky="w", pady=(6, 0))

		self.visualize_all_button = ttk.Button(action_frame, text="関節全体モデル/特徴点を可視化", command=self.on_visualize_all)
		self.visualize_all_button.grid(row=2, column=0, sticky="w", pady=(6, 0))

		self.visualize_options_button = ttk.Button(action_frame, text="可視化オプション", command=self.on_visualize_options)
		self.visualize_options_button.grid(row=3, column=0, sticky="w", pady=(6, 0))

		# 同次変換行列群 入力 & 確認
		tf_frame = ttk.LabelFrame(container, text="変位・姿勢変化データファイルの選択", style="Bold.TLabelframe")
		tf_frame.grid(row=4, column=0, sticky="nsew", pady=(0, 12))
		for i in range(3):
			tf_frame.columnconfigure(i, weight=1)

		ttk.Label(tf_frame, text="変位・姿勢変化データファイルを入力").grid(row=0, column=0, sticky="w", padx=12, pady=6)
		tf_entry = ttk.Entry(tf_frame, textvariable=self.transform_group_path, state="readonly")
		tf_entry.grid(row=0, column=1, sticky="ew", padx=12, pady=6)
		self._register_dnd(tf_entry, self.transform_group_path)
		ttk.Button(tf_frame, text="参照...", command=self.choose_transform_group).grid(row=0, column=2, sticky="e", padx=12, pady=6)

		self.check_tf_button = ttk.Button(tf_frame, text="変位・姿勢変化データファイルの確認", command=self.on_check_transform)
		self.check_tf_button.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 6))

		# シミュレーション実行（ヒートマップ事前計算→アニメーション）
		self.animate_button = ttk.Button(tf_frame, text="シミュレーション実行", command=self.on_animate)
		self.animate_button.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 6))

	# region knee simulator (FRS-2015 ロボット準拠 膝関節)
	def _create_knee_simulator_tab(self) -> None:
		"""knee simulatorタブのUIを構築（FRS-2015 ロボット準拠 膝関節）。

		座標系: 木村圭 学位論文 第3章 / 木村・藤江 機械学会論文集 2015 / 藤江「動作原理と改良」2024。
		入力: 初期状態(組立)スキャン + 特徴点(L,M,N,MCL,LCL) / 大腿骨・脛骨モデル(試験後別スキャン) /
		      左右・W_scan / 変位データ。大腿骨/脛骨モデルはICP/RANSACで初期スキャンへ位置合わせする。
		"""
		# 試験タブバー（スクロール外・常時表示）: 各タブ=独立した試験の状態。＋で追加/右クリックで削除・改名
		tabbar_row = tk.Frame(self.knee_simulator_tab)
		tabbar_row.pack(side="top", fill="x", padx=4, pady=(4, 0))
		tk.Label(tabbar_row, text="試験タブ:", font=(self.ui_font_family, 9)).pack(side="left", padx=(0, 4))
		self._knee_tabbar_frame = tk.Frame(tabbar_row)
		self._knee_tabbar_frame.pack(side="left", fill="x")

		# スクロール可能なメインフレーム
		canvas = tk.Canvas(self.knee_simulator_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.knee_simulator_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)
		scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")

		def _on_mousewheel(event):
			step = _mousewheel_units(event)
			if step:
				canvas.yview_scroll(step, "units")
		canvas.bind("<Enter>", lambda e: (canvas.bind_all("<MouseWheel>", _on_mousewheel),
		                                   canvas.bind_all("<Button-4>", _on_mousewheel),
		                                   canvas.bind_all("<Button-5>", _on_mousewheel)))
		canvas.bind("<Leave>", lambda e: (canvas.unbind_all("<MouseWheel>"),
		                                   canvas.unbind_all("<Button-4>"),
		                                   canvas.unbind_all("<Button-5>")))

		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		ttk.Label(container, text="knee simulator（FRS-2015 ロボット準拠 膝関節）",
		          font=(self.ui_font_family, 12, "bold")).grid(row=0, column=0, sticky="w", padx=4, pady=(4, 2))
		ttk.Label(container,
		          text="大腿骨座標系Cf(木村論文): 原点=MCL/LCL付着部の中点, Z=LCL→MCL(右膝で外側+), "
		               "Y=骨軸(L,M,N由来の基準方向)をZ⊥面へ投影(近位+), X=Y×Z。脛骨座標系Ctは伸展位でCfに一致。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=1, column=0, sticky="w", padx=4, pady=(0, 8))

		# Section: 初期状態スキャン（座標系の基準）
		init_frame = ttk.LabelFrame(container, text="① 初期状態スキャン（座標系の基準）", style="Bold.TLabelframe")
		init_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			init_frame.columnconfigure(i, weight=[0, 1, 0][i])
		self._add_file_row(init_frame, 0, "初期状態スキャン (STL/OBJ)", self.knee_initial_scan_path,
		                   lambda: self._knee_choose(self.knee_initial_scan_path, "初期状態スキャンを選択", "model"))
		self._add_file_row(init_frame, 1, "初期スキャン特徴点 (PP)", self.knee_initial_pp_path,
		                   lambda: self._knee_choose(self.knee_initial_pp_path, "初期スキャン特徴点(L,M,N,MCL,LCL)を選択", "pp"))
		ttk.Label(init_frame, text="特徴点ラベル: L, M, N (大腿骨クランプパネル頂点), MCL, LCL (側副靭帯 大腿骨付着部)",
		          foreground="gray", font=(self.ui_font_family, 8)).grid(row=2, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 4))
		# 左右・W_scan
		opt = ttk.Frame(init_frame)
		opt.grid(row=3, column=0, columnspan=3, sticky="w", padx=12, pady=(2, 6))
		ttk.Label(opt, text="関節側:").grid(row=0, column=0, sticky="w")
		ttk.Radiobutton(opt, text="右膝", value=1, variable=self.knee_side_var).grid(row=0, column=1, sticky="w", padx=(6, 4))
		ttk.Radiobutton(opt, text="左膝", value=2, variable=self.knee_side_var).grid(row=0, column=2, sticky="w", padx=(0, 16))
		ttk.Label(opt, text="W_scan (伸展位屈曲角φ, deg):").grid(row=0, column=3, sticky="w")
		ttk.Entry(opt, textvariable=self.knee_w_scan_deg, width=8).grid(row=0, column=4, sticky="w", padx=(6, 4))
		ttk.Label(opt, text="ヒト膝=0 / 動物膝=実測(30〜40)", foreground="gray",
		          font=(self.ui_font_family, 8)).grid(row=0, column=5, sticky="w")

		ttk.Button(init_frame, text="特徴点/座標系を可視化", command=self.on_knee_visualize_coords
		           ).grid(row=4, column=0, sticky="w", padx=12, pady=(0, 8))

		# Section: 大腿骨・脛骨モデル（試験後スキャン）
		model_frame = ttk.LabelFrame(container, text="② 大腿骨・脛骨モデル（試験後の別スキャン）", style="Bold.TLabelframe")
		model_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			model_frame.columnconfigure(i, weight=[0, 1, 0][i])
		self._add_file_row(model_frame, 0, "大腿骨モデル (STL/OBJ)", self.knee_femur_model_path,
		                   lambda: self._knee_choose(self.knee_femur_model_path, "大腿骨モデルを選択", "model"))
		self._add_file_row(model_frame, 1, "脛骨モデル (STL/OBJ)", self.knee_tibia_model_path,
		                   lambda: self._knee_choose(self.knee_tibia_model_path, "脛骨モデルを選択", "model"))
		ttk.Label(model_frame, text="※ これらのモデルには特徴点は不要（幾何のみ）。③の位置合わせで初期スキャン座標系へ整列します。",
		          foreground="gray", font=(self.ui_font_family, 8)).grid(row=2, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 4))

		# Section: 位置合わせ（大腿骨・脛骨を個別に）
		reg_frame = ttk.LabelFrame(container, text="③ 位置合わせ（大腿骨・脛骨を個別に → 初期スキャン座標系へ）", style="Bold.TLabelframe")
		reg_frame.grid(row=4, column=0, sticky="nsew", pady=(0, 8))
		reg_frame.columnconfigure(0, weight=1)
		self._build_knee_reg_panel(
			reg_frame, 0, "femur", "大腿骨",
			self.knee_reg_femur_src_path, self.knee_reg_femur_tgt_path,
			"大腿骨モデル側 位置合わせ領域を選択", "初期スキャンの大腿骨領域を選択")
		self._build_knee_reg_panel(
			reg_frame, 1, "tibia", "脛骨",
			self.knee_reg_tibia_src_path, self.knee_reg_tibia_tgt_path,
			"脛骨モデル側 位置合わせ領域を選択", "初期スキャンの脛骨領域を選択")
		ttk.Button(reg_frame, text="位置合わせ結果を確認（大腿骨＋脛骨）", command=self.on_knee_preview_registration
		           ).grid(row=2, column=0, sticky="w", padx=12, pady=(2, 6))
		ttk.Label(reg_frame,
		          text="※ 領域はモデル/初期スキャンから同一部位を切り出したもの。"
		               "方式: RANSAC=特徴量 / 主軸PCA=向き粗合わせ / 手動3点=対応点クリック → いずれもICPで仕上げ。"
		               "パラメータは骨ごとに設定・プリセット保存できます。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=3, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 6))
		self._knee_refresh_preset_combos()

		# Section: 変位データ・シミュレーション
		sim_frame = ttk.LabelFrame(container, text="④ 変位・姿勢変化データ / シミュレーション", style="Bold.TLabelframe")
		sim_frame.grid(row=5, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			sim_frame.columnconfigure(i, weight=[0, 1, 0][i])
		self._add_file_row(sim_frame, 0, "変位・姿勢変化データ (xlsx/kkr)", self.knee_transform_path,
		                   lambda: self._knee_choose(self.knee_transform_path, "変位・姿勢変化データを選択", "transform"))
		# 表示色（hipの可視化オプションとは独立。タブごとに保存され、可視化/シミュレーションに反映）
		colf = ttk.Frame(sim_frame)
		colf.grid(row=1, column=0, columnspan=3, sticky="w", padx=12, pady=(4, 2))
		ttk.Label(colf, text="表示色:").grid(row=0, column=0, sticky="w")
		ttk.Label(colf, text="大腿骨").grid(row=0, column=1, sticky="w", padx=(8, 2))
		self.knee_femur_color_btn = tk.Button(colf, width=3, relief="ridge",
		                                      command=lambda: self._knee_choose_color("femur"))
		self.knee_femur_color_btn.grid(row=0, column=2, sticky="w")
		ttk.Label(colf, text="脛骨").grid(row=0, column=3, sticky="w", padx=(12, 2))
		self.knee_tibia_color_btn = tk.Button(colf, width=3, relief="ridge",
		                                      command=lambda: self._knee_choose_color("tibia"))
		self.knee_tibia_color_btn.grid(row=0, column=4, sticky="w")
		ttk.Label(colf, text="（試験タブごとに保存。hip simulatorの可視化オプションとは独立）",
		          foreground="gray", font=(self.ui_font_family, 8)).grid(row=0, column=5, sticky="w", padx=(10, 0))
		self._knee_update_color_buttons()
		ttk.Button(sim_frame, text="関節全体を可視化", command=self.on_knee_visualize_all
		           ).grid(row=2, column=0, sticky="w", padx=12, pady=(2, 4))
		ttk.Button(sim_frame, text="シミュレーション実行", command=self.on_knee_animate
		           ).grid(row=3, column=0, sticky="w", padx=12, pady=(0, 8))

		# Section: シーン共有（別PC・別OS＝解析サーバーやMacで計算/再現するため）
		scene_frame = ttk.LabelFrame(container, text="⑤ シーン共有（別PC・別OSで計算/動作再現）", style="Bold.TLabelframe")
		scene_frame.grid(row=6, column=0, sticky="nsew", pady=(0, 8))
		scene_frame.columnconfigure(0, weight=1)
		sbtn = ttk.Frame(scene_frame)
		sbtn.grid(row=0, column=0, sticky="w", padx=12, pady=(4, 2))
		ttk.Button(sbtn, text="このタブのシーンを保存", command=self.on_knee_scene_save
		           ).grid(row=0, column=0, sticky="w", padx=(0, 8))
		ttk.Button(sbtn, text="全タブのシーンを保存", command=self.on_knee_scene_save_all
		           ).grid(row=0, column=1, sticky="w", padx=(0, 8))
		ttk.Button(sbtn, text="シーンを読込（タブ名で同期）", command=self.on_knee_scene_load
		           ).grid(row=0, column=2, sticky="w")
		ttk.Label(scene_frame,
		          text="試験タブ名ごとに 位置合わせ(reg_T)・W_scan・左右・ファイル内容ハッシュを cache/knee_scene.json へ保存します\n"
		               "（小さいのでgit共有可）。読込はタブ名で同期: 同名タブは更新・無ければ新規作成＝タブ名も別PCへ引き継げます。\n"
		               "運用A（サーバーで計算）: Windowsで位置合わせ→保存→push → サーバーでpull→読込→シミュレーション実行。\n"
		               "運用B（Mac等で動作再現）: シーンは git pull、キャッシュ(cache/overlap/*.pkl)は大容量のためgit対象外\n"
		               "　→ Box/USB等でコピーして配置。同じ内容のモデル/変位データを選んで実行するとキャッシュが即ヒットし、\n"
		               "　再計算なしで動作再現できます（キャッシュは内容ハッシュ方式・OS非依存）。\n"
		               "※ファイルパス自体は引き継がれません。各PCで同内容のファイルを選び直してください（一致は読込時に自動チェック）。",
		          foreground="gray", font=(self.ui_font_family, 8), justify="left"
		          ).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 6))

		# 試験タブを初期化（未初期化なら現状態を試験1として作成）してタブバーを描画
		self._knee_init_tabs()

	def _knee_choose(self, var: tk.StringVar, title: str, kind: str) -> None:
		"""knee simulator用の共通ファイル選択ダイアログ。kind: 'model' | 'pp' | 'transform'."""
		if kind == "pp":
			ft = [("特徴点ファイル", "*.pp"), ("すべてのファイル", "*.*")]
		elif kind == "transform":
			ft = [("変位データ", "*.xlsx *.kkr *.KKR"), ("Excelファイル", "*.xlsx"),
			      ("KKRファイル", "*.KKR *.kkr"), ("すべてのファイル", "*.*")]
		else:
			ft = [("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")]
		path = filedialog.askopenfilename(title=title, filetypes=ft)
		if path:
			var.set(path)

	# ----- Knee simulator 表示色（hipの可視化オプションとは独立・タブごとに保存） -----
	def _knee_color(self, bone: str) -> str:
		"""kneeの表示色（femur/tibia）を返す。不正値は既定色にフォールバック。"""
		var = self.knee_femur_color_var if bone == "femur" else self.knee_tibia_color_var
		default = "#DEB887" if bone == "femur" else "#ADD8E6"
		try:
			c = str(var.get()).strip()
		except Exception:
			return default
		if len(c) == 7 and c[0] == "#" and all(ch in "0123456789abcdefABCDEF" for ch in c[1:]):
			return c
		return default

	def _knee_choose_color(self, bone: str) -> None:
		"""④の表示色（大腿骨/脛骨）をカラーピッカーで変更する。"""
		var = self.knee_femur_color_var if bone == "femur" else self.knee_tibia_color_var
		title = "大腿骨の表示色を選択" if bone == "femur" else "脛骨の表示色を選択"
		color = colorchooser.askcolor(initialcolor=self._knee_color(bone), title=title)
		if color and color[1]:
			var.set(color[1])
			self._knee_update_color_buttons()

	def _knee_update_color_buttons(self) -> None:
		"""④の色見本ボタンの背景を現在の設定色に合わせる（タブ切替・シーン読込時にも呼ぶ）。"""
		for bone, btn_name in (("femur", "knee_femur_color_btn"), ("tibia", "knee_tibia_color_btn")):
			btn = getattr(self, btn_name, None)
			if btn is None:
				continue
			try:
				c = self._knee_color(bone)
				btn.configure(bg=c, activebackground=c)
			except Exception:
				pass

	# ----- Knee simulator 複数タブ（各タブ=独立スナップショット、UI/コードは共通） -----
	def _knee_snapshot_current(self) -> dict:
		"""現在の knee 状態（全入力・パラメータ・位置合わせ結果）をdictに保存する。"""
		snap = {}
		for key, (var, _t) in self._knee_state_vars().items():
			try:
				snap[key] = var.get()
			except Exception:
				pass
		snap['_femur_reg_T'] = self._knee_femur_reg_T.tolist() if self._knee_femur_reg_T is not None else None
		snap['_tibia_reg_T'] = self._knee_tibia_reg_T.tolist() if self._knee_tibia_reg_T is not None else None
		return snap

	def _knee_restore_snapshot(self, snap: dict) -> None:
		"""dict の状態を現在の knee UI（ライブ変数）へ反映する。
		スナップショットに無いキーは既定値へ戻す（後から追加した変数が旧タブに残留しないように）。"""
		if not isinstance(snap, dict):
			return
		default = self._knee_default_snap if isinstance(getattr(self, "_knee_default_snap", None), dict) else {}
		for key, (var, typ) in self._knee_state_vars().items():
			if key in snap:
				v = snap[key]
			elif key in default:
				v = default[key]
			else:
				continue
			try:
				if typ is bool:
					var.set(bool(v))
				elif typ is int:
					var.set(int(v))
				elif typ is float:
					var.set(float(v))
				else:
					var.set(str(v))
			except Exception:
				pass
		self._knee_femur_reg_T = np.array(snap['_femur_reg_T'], dtype=float) if snap.get('_femur_reg_T') is not None else None
		self._knee_tibia_reg_T = np.array(snap['_tibia_reg_T'], dtype=float) if snap.get('_tibia_reg_T') is not None else None
		self._knee_update_color_buttons()

	def _knee_rebuild_tabbar(self) -> None:
		"""タブバー（試験ごとのボタン＋「＋」）を再描画する。"""
		fr = self._knee_tabbar_frame
		if fr is None:
			return
		for w in fr.winfo_children():
			w.destroy()
		self._knee_tab_buttons = []
		for i, tab in enumerate(self._knee_tabs):
			active = (i == self._knee_active_tab)
			b = tk.Button(
				fr, text=tab.get('name', f"試験{i+1}"),
				relief=('sunken' if active else 'raised'),
				bg=('#cfe3ff' if active else '#f0f0f0'),
				font=(self.ui_font_family, 9, 'bold' if active else 'normal'),
				command=lambda i=i: self.on_knee_tab_select(i), padx=8, pady=2)
			b.pack(side='left', padx=2)
			b.bind("<Button-3>", lambda e, i=i: self._knee_tab_context_menu(e, i))
			# ドラッグで並べ替え（クリック=切替はしきい値未満のときだけ Button の command に任せる）
			b.bind("<ButtonPress-1>", lambda e, i=i: self._knee_tab_drag_start(e, i))
			b.bind("<B1-Motion>", self._knee_tab_drag_motion)
			b.bind("<ButtonRelease-1>", self._knee_tab_drag_release)
			self._knee_tab_buttons.append(b)
		plus = tk.Button(fr, text="＋", command=self.on_knee_tab_add, padx=6, pady=2)
		plus.pack(side='left', padx=(8, 2))

	def _knee_tab_context_menu(self, event, i: int) -> None:
		menu = tk.Menu(self, tearoff=0)
		menu.add_command(label="名前変更", command=lambda: self.on_knee_tab_rename(i))
		menu.add_command(label="削除", command=lambda: self.on_knee_tab_delete(i))
		menu.add_separator()
		menu.add_command(label="← 左へ移動", command=lambda: self._knee_tab_move(i, i - 1),
		                 state=("normal" if i > 0 else "disabled"))
		menu.add_command(label="→ 右へ移動", command=lambda: self._knee_tab_move(i, i + 1),
		                 state=("normal" if i < len(self._knee_tabs) - 1 else "disabled"))
		try:
			menu.tk_popup(event.x_root, event.y_root)
		finally:
			menu.grab_release()

	# ----- 試験タブの並べ替え（ドラッグ＆右クリックメニュー） -----
	def _knee_tab_move(self, i: int, j: int) -> None:
		"""試験タブ i を位置 j へ移動する（アクティブタブは移動に追従）。"""
		n = len(self._knee_tabs)
		if not (0 <= i < n and 0 <= j < n) or i == j:
			return
		tab = self._knee_tabs.pop(i)
		self._knee_tabs.insert(j, tab)
		a = self._knee_active_tab
		if a == i:
			a = j
		else:
			if a > i:
				a -= 1
			if a >= j:
				a += 1
		self._knee_active_tab = a
		self._knee_rebuild_tabbar()

	def _knee_tab_drag_start(self, event, i: int) -> None:
		self._knee_tab_drag = {"from": i, "x": event.x_root, "moved": False}

	def _knee_tab_drag_motion(self, event) -> None:
		d = self._knee_tab_drag
		if d is None:
			return
		if not d["moved"] and abs(event.x_root - d["x"]) > 12:
			d["moved"] = True
			try:
				event.widget.configure(cursor="sb_h_double_arrow")
			except Exception:
				pass

	def _knee_tab_drag_release(self, event):
		d = self._knee_tab_drag
		self._knee_tab_drag = None
		if not d or not d["moved"]:
			return None  # ただのクリック → Button の command（タブ切替）に任せる
		try:
			event.widget.configure(cursor="")
		except Exception:
			pass
		j = self._knee_tab_index_at(event.x_root)
		if j is None or j == d["from"]:
			return None
		self._knee_tab_move(d["from"], j)
		return "break"  # 並べ替えた場合はクリック（タブ切替）を発火させない

	def _knee_tab_index_at(self, x_root: int):
		"""タブバー上の x_root 座標に対応する試験タブのindexを返す（隙間は左隣、範囲外は端）。"""
		btns = self._knee_tab_buttons
		if not btns:
			return None
		best = 0
		for idx, b in enumerate(btns):
			try:
				bx = b.winfo_rootx()
				bw = b.winfo_width()
			except Exception:
				continue
			if x_root >= bx:
				best = idx
			if bx <= x_root < bx + bw:
				return idx
		return best

	def on_knee_tab_select(self, i: int) -> None:
		if i < 0 or i >= len(self._knee_tabs) or i == self._knee_active_tab:
			return
		# 現在のタブへ保存してから切替
		self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
		self._knee_active_tab = i
		self._knee_restore_snapshot(self._knee_tabs[i]['snapshot'])
		self._knee_rebuild_tabbar()

	# ④（変位・姿勢変化データ）に属するキー。タブ追加時の複製から除外する
	_KNEE_TAB_NO_COPY_KEYS = ("knee_transform",)

	def on_knee_tab_add(self) -> None:
		"""タブ追加: 一つ左（末尾）のタブを複製する（④変位データのみ既定値に戻す）。"""
		if self._knee_tabs:
			self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
			# 新タブの一つ左＝現在の末尾タブを複製（モデルパス・位置合わせ結果reg_T等を引き継ぐ）
			snap = copy.deepcopy(self._knee_tabs[-1].get('snapshot') or {})
			for key in self._KNEE_TAB_NO_COPY_KEYS:
				if isinstance(self._knee_default_snap, dict) and key in self._knee_default_snap:
					snap[key] = self._knee_default_snap[key]
				else:
					snap.pop(key, None)
		else:
			snap = copy.deepcopy(self._knee_default_snap or {})
		self._knee_tabs.append({'name': self._knee_unique_tab_name(), 'snapshot': snap})
		self._knee_active_tab = len(self._knee_tabs) - 1
		self._knee_restore_snapshot(self._knee_tabs[self._knee_active_tab]['snapshot'])
		self._knee_rebuild_tabbar()

	def _knee_unique_tab_name(self, base: str = "試験") -> str:
		"""既存タブと重複しない「試験N」を返す（シーン共有がタブ名キーのため重複回避）。"""
		names = {t.get('name', '') for t in self._knee_tabs}
		n = len(self._knee_tabs) + 1
		while f"{base}{n}" in names:
			n += 1
		return f"{base}{n}"

	def on_knee_tab_delete(self, i: int) -> None:
		if len(self._knee_tabs) <= 1:
			messagebox.showinfo("タブ削除", "最後のタブは削除できません。")
			return
		if not messagebox.askyesno("タブ削除", f"タブ「{self._knee_tabs[i].get('name','')}」を削除しますか？"):
			return
		# 現在のアクティブ状態を保存してから削除
		self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
		del self._knee_tabs[i]
		if self._knee_active_tab == i:
			self._knee_active_tab = min(i, len(self._knee_tabs) - 1)
		elif self._knee_active_tab > i:
			self._knee_active_tab -= 1
		self._knee_restore_snapshot(self._knee_tabs[self._knee_active_tab]['snapshot'])
		self._knee_rebuild_tabbar()

	def on_knee_tab_rename(self, i: int) -> None:
		cur = self._knee_tabs[i].get('name', '')
		new = simpledialog.askstring("タブ名変更", "タブ名:", initialvalue=cur, parent=self)
		if new and new.strip():
			new = new.strip()
			others = {t.get('name', '') for k, t in enumerate(self._knee_tabs) if k != i}
			if new in others:
				messagebox.showwarning("タブ名変更",
					f"「{new}」は既に存在します。シーン共有はタブ名で対応付けるため、別の名前にしてください。")
				return
			self._knee_tabs[i]['name'] = new
			self._knee_rebuild_tabbar()

	def _knee_init_tabs(self) -> None:
		"""UI構築後: タブが未初期化なら現在の状態を試験1として1つ作る。"""
		if not self._knee_tabs:
			self._knee_tabs = [{'name': '試験1', 'snapshot': self._knee_snapshot_current()}]
			self._knee_active_tab = 0
		if self._knee_active_tab >= len(self._knee_tabs):
			self._knee_active_tab = 0
		self._knee_restore_snapshot(self._knee_tabs[self._knee_active_tab]['snapshot'])
		self._knee_rebuild_tabbar()

	# ----- Knee registration params / presets (骨ごと) -----
	def _knee_param_spec(self):
		"""位置合わせ数値パラメータの仕様: (key, kind, default, ラベル)。"""
		return [
			("ransac_distance", "float", 1.0, "RANSAC距離(mm)"),
			("ransac_max_iter", "int", 1000, "RANSAC反復"),
			("ransac_confidence", "float", 0.99, "RANSAC信頼度"),
			("icp_threshold", "float", 5.0, "ICP距離(mm)"),
			("icp_max_iter", "int", 2000, "ICP反復"),
			("sample_points", "int", 100000, "サンプル点数"),
			("voxel_size", "float", 0.0, "ボクセル(mm,0=自動)"),
		]

	def _knee_get_params(self, bone: str) -> dict:
		"""指定した骨の現在の位置合わせパラメータを plain dict で返す。"""
		p = {}
		for key, kind, default, _lbl in self._knee_param_spec():
			try:
				v = self.knee_reg_pvars[bone][key].get()
				p[key] = int(v) if kind == "int" else float(v)
			except Exception:
				p[key] = default
		p["method"] = self.knee_reg_method_var[bone].get()
		p["enable_scaling"] = bool(self.knee_reg_scaling_var[bone].get())
		p["preview"] = bool(self.knee_reg_preview_var[bone].get())
		return p

	def _knee_set_params(self, bone: str, d: dict) -> None:
		"""dict の値を指定した骨のパラメータUIに反映する。"""
		for key, kind, default, _lbl in self._knee_param_spec():
			if key in d:
				try:
					self.knee_reg_pvars[bone][key].set(int(d[key]) if kind == "int" else float(d[key]))
				except Exception:
					pass
		if "method" in d:
			self.knee_reg_method_var[bone].set(str(d["method"]))
		if "enable_scaling" in d:
			self.knee_reg_scaling_var[bone].set(bool(d["enable_scaling"]))
		if "preview" in d:
			self.knee_reg_preview_var[bone].set(bool(d["preview"]))

	def _knee_presets_file(self) -> Path:
		import platform
		fn = "frs2015_knee_reg_presets.json"
		if platform.system() == "Darwin":
			dd = Path.home() / ".frs_simulator"
			dd.mkdir(parents=True, exist_ok=True)
			return dd / fn
		return Path(__file__).with_name(fn)

	def _load_knee_presets(self) -> dict:
		try:
			p = self._knee_presets_file()
			if p.exists():
				data = json.load(p.open("r", encoding="utf-8"))
				if isinstance(data, dict):
					return data
		except Exception as e:
			print(f"[knee preset] 読込失敗: {e}")
		return {}

	def _save_knee_presets(self) -> None:
		try:
			p = self._knee_presets_file()
			with p.open("w", encoding="utf-8") as f:
				json.dump(self._knee_presets, f, ensure_ascii=False, indent=2)
		except Exception as e:
			print(f"[knee preset] 保存失敗: {e}")

	def _knee_refresh_preset_combos(self) -> None:
		names = sorted(self._knee_presets.keys())
		for bone, combo in self.knee_preset_combo.items():
			try:
				combo["values"] = names
			except Exception:
				pass

	def on_knee_preset_apply(self, bone: str) -> None:
		name = self.knee_preset_sel[bone].get().strip()
		if not name or name not in self._knee_presets:
			messagebox.showwarning("プリセット", "適用するプリセットを選択してください。")
			return
		self._knee_set_params(bone, self._knee_presets[name])
		messagebox.showinfo("プリセット", f"『{name}』を{('大腿骨' if bone == 'femur' else '脛骨')}に適用しました。")

	def on_knee_preset_save(self, bone: str) -> None:
		default = self.knee_preset_sel[bone].get().strip() or "preset1"
		name = simpledialog.askstring("プリセット保存", "プリセット名:", initialvalue=default, parent=self)
		if name is None:
			return
		name = name.strip()
		if not name:
			return
		if name in self._knee_presets and not messagebox.askyesno("上書き確認", f"『{name}』を上書きしますか？"):
			return
		self._knee_presets[name] = self._knee_get_params(bone)
		self._save_knee_presets()
		self._knee_refresh_preset_combos()
		self.knee_preset_sel[bone].set(name)
		messagebox.showinfo("プリセット", f"『{name}』を保存しました。")

	def on_knee_preset_rename(self, bone: str) -> None:
		old = self.knee_preset_sel[bone].get().strip()
		if not old or old not in self._knee_presets:
			messagebox.showwarning("プリセット", "名前を変更するプリセットを選択してください。")
			return
		new = simpledialog.askstring("名前変更", "新しい名前:", initialvalue=old, parent=self)
		if new is None:
			return
		new = new.strip()
		if not new or new == old:
			return
		if new in self._knee_presets and not messagebox.askyesno("上書き確認", f"『{new}』は既にあります。上書きしますか？"):
			return
		self._knee_presets[new] = self._knee_presets.pop(old)
		self._save_knee_presets()
		self._knee_refresh_preset_combos()
		for b in ("femur", "tibia"):
			if self.knee_preset_sel[b].get() == old:
				self.knee_preset_sel[b].set(new)

	def on_knee_preset_delete(self, bone: str) -> None:
		name = self.knee_preset_sel[bone].get().strip()
		if not name or name not in self._knee_presets:
			messagebox.showwarning("プリセット", "削除するプリセットを選択してください。")
			return
		if not messagebox.askyesno("削除確認", f"プリセット『{name}』を削除しますか？"):
			return
		del self._knee_presets[name]
		self._save_knee_presets()
		self._knee_refresh_preset_combos()
		for b in ("femur", "tibia"):
			if self.knee_preset_sel[b].get() == name:
				self.knee_preset_sel[b].set("")

	def _build_knee_reg_panel(self, parent, row, bone, label, src_var, tgt_var, src_title, tgt_title) -> None:
		"""1つの骨(大腿骨/脛骨)の位置合わせパネルを構築する。"""
		panel = ttk.LabelFrame(parent, text=f"{label}の位置合わせ", style="Bold.TLabelframe")
		panel.grid(row=row, column=0, sticky="nsew", padx=8, pady=(4, 4))
		for i in range(3):
			panel.columnconfigure(i, weight=[0, 1, 0][i])
		# 領域
		self._add_file_row(panel, 0, f"{label}: モデル側 領域", src_var,
		                   lambda: self._knee_choose(src_var, src_title, "model"))
		self._add_file_row(panel, 1, f"{label}: 初期スキャン側 領域", tgt_var,
		                   lambda: self._knee_choose(tgt_var, tgt_title, "model"))
		# 方式
		mf = ttk.Frame(panel)
		mf.grid(row=2, column=0, columnspan=3, sticky="w", padx=12, pady=(2, 0))
		ttk.Label(mf, text="方式:").grid(row=0, column=0, sticky="w")
		mv = self.knee_reg_method_var[bone]
		ttk.Radiobutton(mf, text="RANSAC", value="ransac", variable=mv).grid(row=0, column=1, sticky="w", padx=(6, 6))
		ttk.Radiobutton(mf, text="主軸PCA", value="pca", variable=mv).grid(row=0, column=2, sticky="w", padx=(0, 6))
		ttk.Radiobutton(mf, text="手動3点", value="manual", variable=mv).grid(row=0, column=3, sticky="w", padx=(0, 6))
		# パラメータ（3列レイアウト）
		pf = ttk.Frame(panel)
		pf.grid(row=3, column=0, columnspan=3, sticky="w", padx=12, pady=(4, 2))
		pvars = self.knee_reg_pvars[bone]
		col = 0
		r = 0
		for key, kind, default, lbl in self._knee_param_spec():
			ttk.Label(pf, text=lbl + ":").grid(row=r, column=col * 2, sticky="w", pady=(2, 0))
			ttk.Entry(pf, textvariable=pvars[key], width=8).grid(row=r, column=col * 2 + 1, sticky="w", padx=(4, 12), pady=(2, 0))
			col += 1
			if col >= 3:
				col = 0
				r += 1
		# scaling / preview
		cf = ttk.Frame(panel)
		cf.grid(row=4, column=0, columnspan=3, sticky="w", padx=12, pady=(2, 0))
		ttk.Checkbutton(cf, text="スケール補正", variable=self.knee_reg_scaling_var[bone]).grid(row=0, column=0, sticky="w", padx=(0, 12))
		ttk.Checkbutton(cf, text="経過を表示", variable=self.knee_reg_preview_var[bone]).grid(row=0, column=1, sticky="w")
		# プリセット
		prf = ttk.Frame(panel)
		prf.grid(row=5, column=0, columnspan=3, sticky="w", padx=12, pady=(4, 2))
		ttk.Label(prf, text="プリセット:").grid(row=0, column=0, sticky="w")
		combo = ttk.Combobox(prf, textvariable=self.knee_preset_sel[bone], width=16, state="readonly")
		combo.grid(row=0, column=1, sticky="w", padx=(4, 6))
		self.knee_preset_combo[bone] = combo
		ttk.Button(prf, text="適用", width=5, command=lambda b=bone: self.on_knee_preset_apply(b)).grid(row=0, column=2, padx=2)
		ttk.Button(prf, text="保存", width=5, command=lambda b=bone: self.on_knee_preset_save(b)).grid(row=0, column=3, padx=2)
		ttk.Button(prf, text="名前変更", width=8, command=lambda b=bone: self.on_knee_preset_rename(b)).grid(row=0, column=4, padx=2)
		ttk.Button(prf, text="削除", width=5, command=lambda b=bone: self.on_knee_preset_delete(b)).grid(row=0, column=5, padx=2)
		# 実行
		ttk.Button(panel, text=f"{label}を位置合わせ", command=lambda b=bone: self.on_knee_register_bone(b)
		           ).grid(row=6, column=0, sticky="w", padx=12, pady=(4, 6))

	# ----- Knee coordinate system builders (木村論文準拠) -----
	def _knee_named_points(self, points, labels) -> dict:
		"""特徴点配列とラベルから {ラベル大文字: np.array([x,y,z])} を作る。"""
		d = {}
		for p, lab in zip(points, labels):
			key = str(lab).strip().upper().replace(" ", "")
			d[key] = np.asarray(p, dtype=float)
		return d

	def _knee_reference_shaft_dir(self, pts: dict) -> np.ndarray:
		"""L, M, N クランプパネル頂点から、大腿骨骨軸(≒基準座標系Coのy軸=近位方向)の
		参照方向ベクトルを求める。パネル面の法線を用い、原点→パネル重心の向き(近位)に符号を合わせる。
		"""
		L, M, N = pts["L"], pts["M"], pts["N"]
		normal = np.cross(M - L, N - L)
		nn = np.linalg.norm(normal)
		if nn < 1e-9:
			raise ValueError("L,M,N が同一直線上にあり、パネル法線を定義できません")
		normal = normal / nn
		# 近位方向のヒント: 膝(MCL/LCL中点)からクランプ重心(L,M,N重心=近位側)へ向かうベクトル
		mcl = pts.get("MCL"); lcl = pts.get("LCL")
		if mcl is not None and lcl is not None:
			origin = (mcl + lcl) / 2.0
			proximal_hint = (L + M + N) / 3.0 - origin
			if np.dot(normal, proximal_hint) < 0:
				normal = -normal
		return normal

	def _build_knee_femur_frame(self, points, labels, side: int):
		"""大腿骨座標系 Cf を構築（木村論文 第3章）。

		Returns: (origin, x_axis, y_axis, z_axis) すべて np.array。
		side: 1=右膝(Z+=外側=LCL方向), 2=左膝(符号反転)。
		"""
		pts = self._knee_named_points(points, labels)
		for k in ("MCL", "LCL", "L", "M", "N"):
			if k not in pts:
				raise ValueError(f"必要な特徴点が見つかりません: {k}（必要: L,M,N,MCL,LCL）")
		MCL, LCL = pts["MCL"], pts["LCL"]
		origin = (MCL + LCL) / 2.0
		# Z軸: LCL→MCL を結ぶ直線。右膝でZ+=外側(=LCL方向)なので LCL-MCL 方向を正にとる
		z = LCL - MCL
		if np.linalg.norm(z) < 1e-9:
			raise ValueError("MCL と LCL が一致しており、Z軸を定義できません")
		z = z / np.linalg.norm(z)
		if side == 2:  # 左膝は内外側の符号反転
			z = -z
		# 骨軸参照方向(近位)を Z⊥面へ投影して Y軸(近位+)
		shaft = self._knee_reference_shaft_dir(pts)
		y = shaft - np.dot(shaft, z) * z
		if np.linalg.norm(y) < 1e-9:
			raise ValueError("骨軸参照方向がZ軸と平行で、Y軸を定義できません")
		y = y / np.linalg.norm(y)
		# X軸 = Y×Z
		x = np.cross(y, z)
		x = x / np.linalg.norm(x)
		return origin, x, y, z

	def _build_knee_tibia_frame(self, cf_origin, cf_x, cf_y, cf_z, w_scan_deg: float):
		"""脛骨座標系 Ct を構築。伸展位で Cf に一致。動物膝は Cf を Z軸まわりに屈曲角 φ 回転
		(木村論文 3.1.2: 「Z軸の負方向に回転」)。φ=w_scan_deg。

		Returns: (origin, x_axis, y_axis, z_axis)。
		"""
		phi = np.deg2rad(w_scan_deg)
		# Cf の Z軸まわりに -φ 回転（ロドリゲス）
		k = cf_z / np.linalg.norm(cf_z)
		ang = -phi

		def rot(v):
			return (v * np.cos(ang)
			        + np.cross(k, v) * np.sin(ang)
			        + k * np.dot(k, v) * (1 - np.cos(ang)))

		x = rot(cf_x); y = rot(cf_y); z = cf_z.copy()
		return cf_origin.copy(), x, y, z

	def on_knee_visualize_coords(self) -> None:
		"""初期状態スキャン上に特徴点と大腿骨座標系Cf(・脛骨座標系Ct)を可視化する。"""
		scan_path = self.knee_initial_scan_path.get().strip()
		pp_path = self.knee_initial_pp_path.get().strip()
		if not scan_path or not pp_path:
			messagebox.showwarning("入力不足", "初期状態スキャンと特徴点(PP)を選択してください。")
			return
		try:
			points, labels = self._parse_pp_file(pp_path)
		except Exception as e:
			messagebox.showerror("特徴点読み込み失敗", f"PPファイルの読み込みに失敗しました:\n{e}")
			return
		try:
			origin, x_axis, y_axis, z_axis = self._build_knee_femur_frame(points, labels, self.knee_side_var.get())
		except Exception as e:
			messagebox.showerror("大腿骨座標系構築エラー", f"Cfの構築に失敗しました:\n{e}")
			return
		w_scan = 0.0
		try:
			w_scan = float(self.knee_w_scan_deg.get())
		except Exception:
			pass
		t_origin, t_x, t_y, t_z = self._build_knee_tibia_frame(origin, x_axis, y_axis, z_axis, w_scan)

		try:
			mesh = pv.read(scan_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"初期状態スキャンの読み込みに失敗しました:\n{e}")
			return

		# 軸長: モデルサイズに応じて決定
		try:
			diag = float(np.linalg.norm(np.array(mesh.bounds[1::2]) - np.array(mesh.bounds[0::2])))
			axis_len = max(diag * 0.15, 10.0)
		except Exception:
			axis_len = 40.0

		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="knee: 特徴点/座標系の可視化",
		                     window_size=(int(sw * 0.9), int(sh * 0.9)))
		plotter.set_background("white")
		plotter.add_mesh(mesh, color="lightgray", opacity=0.55, smooth_shading=True, show_edges=False)

		# 特徴点表示
		pts_named = self._knee_named_points(points, labels)
		colors = {"MCL": "red", "LCL": "blue", "L": "green", "M": "orange", "N": "purple"}
		for name, col in colors.items():
			if name in pts_named:
				p = pts_named[name]
				plotter.add_mesh(pv.Sphere(radius=max(axis_len * 0.05, 1.0), center=p), color=col)
				plotter.add_point_labels([p], [name], font_size=14, text_color=col,
				                         point_size=1, show_points=False, always_visible=True)

		def draw_frame(o, ax, ay, az, prefix, solid=True):
			specs = [(ax, "red", f"{prefix}X"), (ay, "green", f"{prefix}Y"), (az, "blue", f"{prefix}Z")]
			for vec, col, name in specs:
				end = o + vec * axis_len
				plotter.add_mesh(pv.Line(o, end), color=col, line_width=(5 if solid else 2))
				plotter.add_point_labels([end], [name], font_size=13, text_color=col,
				                         point_size=1, show_points=False, always_visible=True)

		# Cf（実線）
		draw_frame(origin, x_axis, y_axis, z_axis, "Cf_", solid=True)
		# Ct（W_scan≠0 のとき破線的に細く表示）
		if abs(w_scan) > 1e-6:
			draw_frame(t_origin, t_x, t_y, t_z, "Ct_", solid=False)

		plotter.add_text("Cf: X=red Y=green Z=blue  | 点: MCL=red LCL=blue L=green M=orange N=purple",
		                 position="upper_left", font_size=10, color="black")
		self._show_plotter_coop(plotter)

	def _knee_load_o3d_pcd(self, path: str, n_points: int):
		"""STL/OBJ/PLY を VTK/PyVista で読み込み、Open3D の点群として返す。

		Open3D の read_triangle_mesh は Windows の非ASCII(日本語)パスや一部STLで
		空メッシュを返すことがあるため、堅牢な VTK(pyvista) 経由で読み込んでから変換する。
		面がある場合は表面から一様サンプリング、点群のみの場合は頂点をそのまま使う。
		"""
		try:
			pm = pv.read(path)
		except Exception as e:
			raise ValueError(f"メッシュを読み込めません: {Path(path).name}\n{e}")
		if pm is None or pm.n_points == 0:
			raise ValueError(f"メッシュを読み込めません（頂点なし）: {Path(path).name}")

		verts = np.asarray(pm.points, dtype=np.float64)
		# 面(三角形)を取り出す
		tris = None
		try:
			surf = pm
			if not isinstance(surf, pv.PolyData):
				surf = surf.extract_surface()
			surf = surf.triangulate()
			verts = np.asarray(surf.points, dtype=np.float64)
			faces = np.asarray(surf.faces)
			if faces.size >= 4:
				tris = faces.reshape(-1, 4)[:, 1:4].astype(np.int32)
		except Exception:
			tris = None

		if tris is not None and len(tris) > 0:
			mesh = o3d.geometry.TriangleMesh()
			mesh.vertices = o3d.utility.Vector3dVector(verts)
			mesh.triangles = o3d.utility.Vector3iVector(tris)
			pcd = mesh.sample_points_uniformly(number_of_points=max(int(n_points), 1000))
		else:
			# 面が無い（点群）場合は頂点をそのまま使用
			pcd = o3d.geometry.PointCloud()
			pcd.points = o3d.utility.Vector3dVector(verts)
		if len(pcd.points) == 0:
			raise ValueError(f"点群を生成できません: {Path(path).name}")
		return pcd

	def _rigid_from_correspondences(self, src, tgt, allow_scale: bool = False) -> np.ndarray:
		"""対応点 src(n,3)→tgt(n,3) から剛体(必要ならスケール付)変換 4x4 を求める(Umeyama/Kabsch)。"""
		src = np.asarray(src, dtype=float)
		tgt = np.asarray(tgt, dtype=float)
		sc = src.mean(axis=0)
		tc = tgt.mean(axis=0)
		S = src - sc
		T = tgt - tc
		H = S.T @ T
		U, D, Vt = np.linalg.svd(H)
		d = np.sign(np.linalg.det(Vt.T @ U.T))
		Dd = np.diag([1.0, 1.0, d])
		R = Vt.T @ Dd @ U.T
		scale = 1.0
		if allow_scale:
			var = float((S ** 2).sum())
			scale = float((D * np.array([1.0, 1.0, d])).sum() / var) if var > 1e-12 else 1.0
		t = tc - scale * R @ sc
		M = np.eye(4)
		M[:3, :3] = scale * R
		M[:3, 3] = t
		return M

	def _pca_frame(self, points):
		"""点群の重心と主軸(固有値降順、右手系)を返す。columns=主軸。"""
		points = np.asarray(points, dtype=float)
		c = points.mean(axis=0)
		X = points - c
		cov = (X.T @ X) / max(len(X) - 1, 1)
		w, V = np.linalg.eigh(cov)
		order = np.argsort(w)[::-1]
		R = V[:, order]
		if np.linalg.det(R) < 0:
			R[:, 2] *= -1.0
		return c, R

	def _pca_align_candidates(self, child_points, parent_points):
		"""child→parent の主軸整列 4x4 候補を、主軸の符号反転4通り(右手系維持)で返す。"""
		cc, Rc = self._pca_frame(child_points)
		cp, Rp = self._pca_frame(parent_points)
		flips = [np.diag([1.0, 1.0, 1.0]), np.diag([1.0, -1.0, -1.0]),
		         np.diag([-1.0, 1.0, -1.0]), np.diag([-1.0, -1.0, 1.0])]
		cands = []
		for F in flips:
			R = Rp @ F @ Rc.T
			t = cp - R @ cc
			M = np.eye(4)
			M[:3, :3] = R
			M[:3, 3] = t
			cands.append(M)
		return cands

	def _knee_pick_points(self, mesh, n: int, title: str, texture=None) -> np.ndarray:
		"""PyVistaでメッシュ表面の点を n 個クリックさせ、(m,3) 配列で返す。

		texture: pv.Texture or None。指定時はテクスチャ付きレンダリング。
		"""
		picked = []
		try:
			b = np.array(mesh.bounds).reshape(3, 2)
			r = max(float(np.linalg.norm(b[:, 1] - b[:, 0])) * 0.02, 0.5)
		except Exception:
			r = 1.0
		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		p = pv.Plotter(title=title, window_size=(int(sw * 0.7), int(sh * 0.7)))
		p.set_background("white")
		if texture is not None:
			try:
				p.add_mesh(mesh, texture=texture, show_edges=False)
			except Exception as e:
				print(f"[pick] テクスチャ描画失敗 → 灰色にフォールバック: {e}")
				p.add_mesh(mesh, color="lightgray", show_edges=False)
		else:
			p.add_mesh(mesh, color="lightgray", show_edges=False)

		def cb(pt, *args):
			if pt is None or len(picked) >= n:
				return
			pt = np.asarray(pt, dtype=float)
			picked.append(pt)
			p.add_mesh(pv.Sphere(radius=r, center=pt), color="red")
			p.add_point_labels([pt], [str(len(picked))], font_size=16, text_color="red",
			                   show_points=False, always_visible=True)

		p.add_text(f"{title}\n対応点を {n} 個、順番にクリック → ウィンドウを閉じる",
		           position="upper_left", font_size=11, color="black")
		try:
			p.enable_surface_point_picking(callback=cb, show_point=False)
		except Exception:
			try:
				p.enable_point_picking(callback=cb, use_picker=True, show_point=False)
			except Exception as e:
				print(f"[knee pick] 点ピッキング初期化に失敗: {e}")
		p.show()
		return np.array(picked) if picked else np.zeros((0, 3))

	def _knee_register_region(self, src_path: str, tgt_path: str, label: str,
	                          params: dict, manual_src_pts=None, manual_tgt_pts=None):
		"""領域どうしを位置合わせし、src→tgt の4x4変換行列を返す。

		params: 骨ごとのパラメータ dict（_knee_get_params 参照。method/enable_scaling/preview含む）。
		src = 動かす側(モデル側 領域, child), tgt = 合わせる先(初期スキャン側 領域, parent)。
		Returns: (transform 4x4 np.ndarray, fitness float, rmse float)
		"""
		method = params.get("method", "ransac")
		n = max(int(params.get("sample_points", 100000)), 1000)
		child_pcd = self._knee_load_o3d_pcd(src_path, n)
		parent_pcd = self._knee_load_o3d_pcd(tgt_path, n)

		# voxel_size: 手動指定(>0)があれば優先。0なら近傍距離の中央値から自動推定。
		manual_voxel = 0.0
		try:
			manual_voxel = float(params.get("voxel_size", 0.0))
		except Exception:
			manual_voxel = 0.0
		ransac_dist = float(params.get("ransac_distance", 1.0))
		if manual_voxel > 1e-9:
			voxel_size = manual_voxel
		else:
			try:
				num_points = len(parent_pcd.points)
				sample_n = min(500, max(10, num_points))
				sample_indices = np.random.choice(num_points, sample_n, replace=False)
				kdt = o3d.geometry.KDTreeFlann(parent_pcd)
				nn_dists = []
				for idx in sample_indices:
					[_, idxs, dists] = kdt.search_knn_vector_3d(parent_pcd.points[idx], 2)
					if len(dists) >= 2:
						nn_dists.append(float(np.sqrt(dists[1])))
				voxel_size = max(float(np.mean(nn_dists)) * 1.5, 1e-6) if nn_dists else max(ransac_dist, 1e-6)
			except Exception:
				voxel_size = max(ransac_dist, 1e-6)

		parent_down = parent_pcd.voxel_down_sample(voxel_size)
		child_down = child_pcd.voxel_down_sample(voxel_size)
		print(f"[knee reg:{label}] サンプル={len(parent_pcd.points)}/{len(child_pcd.points)}点, "
		      f"voxel={voxel_size:.4f}mm{'(手動)' if manual_voxel>1e-9 else '(自動)'}, "
		      f"ダウンサンプル後={len(parent_down.points)}/{len(child_down.points)}点")
		radius_normal = voxel_size * 2
		parent_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
		child_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
		scaling = bool(params.get("enable_scaling", False))

		# --- 初期変換の候補を方式ごとに用意 ---
		init_transforms = []
		if method == "manual":
			if manual_src_pts is None or manual_tgt_pts is None or len(manual_src_pts) < 3 or len(manual_tgt_pts) < 3:
				raise ValueError("手動位置合わせには、モデル側・初期スキャン側それぞれ3点以上の対応点が必要です。")
			m = min(len(manual_src_pts), len(manual_tgt_pts))
			init_transforms = [self._rigid_from_correspondences(
				np.asarray(manual_src_pts)[:m], np.asarray(manual_tgt_pts)[:m], allow_scale=scaling)]
			print(f"[knee reg:{label}] 手動3点対応から初期変換を計算")
		elif method == "pca":
			init_transforms = self._pca_align_candidates(np.asarray(child_down.points), np.asarray(parent_down.points))
			print(f"[knee reg:{label}] 主軸(PCA)整列: {len(init_transforms)}候補")
		else:  # ransac
			radius_feature = voxel_size * 5
			parent_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				parent_down, o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			child_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				child_down, o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			dist_ransac = ransac_dist
			result_ransac = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
				child_down, parent_down, child_fpfh, parent_fpfh,
				mutual_filter=True,
				max_correspondence_distance=dist_ransac,
				estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPoint(scaling),
				ransac_n=3,
				checkers=[
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(0.9),
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(dist_ransac),
				],
				criteria=o3d.pipelines.registration.RANSACConvergenceCriteria(
					int(params.get("ransac_max_iter", 1000)), float(params.get("ransac_confidence", 0.99))),
			)
			print(f"[knee reg:{label}] RANSAC fitness={result_ransac.fitness:.6f}, RMSE={result_ransac.inlier_rmse:.6f}")
			init_transforms = [np.array(result_ransac.transformation, dtype=float)]

		# --- 各初期変換から ICP 精密化し、最良(fitness最大)を採用 ---
		icp_thresh = float(params.get("icp_threshold", 5.0))
		icp_iter = int(params.get("icp_max_iter", 2000))
		best_fit = -1.0
		best_rmse = float("inf")
		T = np.eye(4)
		for T0 in init_transforms:
			res = o3d.pipelines.registration.registration_icp(
				child_down, parent_down, icp_thresh, np.asarray(T0, dtype=float),
				o3d.pipelines.registration.TransformationEstimationPointToPoint(scaling),
				o3d.pipelines.registration.ICPConvergenceCriteria(max_iteration=icp_iter),
			)
			if res.fitness > best_fit:
				best_fit = float(res.fitness)
				best_rmse = float(res.inlier_rmse)
				T = np.array(res.transformation, dtype=float)
		print(f"[knee reg:{label}] ICP(最良) fitness={best_fit:.6f}, RMSE={best_rmse:.6f} "
		      f"(method={method}, {len(init_transforms)}候補から採用)")

		# 経過表示（任意）: 点群で重ね合わせを表示
		if bool(params.get("preview", True)):
			try:
				r_parent = copy.deepcopy(parent_pcd); r_parent.paint_uniform_color([1.0, 0.0, 0.0])
				r_child = copy.deepcopy(child_pcd); r_child.transform(T); r_child.paint_uniform_color([0.0, 1.0, 0.0])
				o3d.visualization.draw_geometries(
					[r_parent, r_child],
					window_name=f"位置合わせ後 [{label}]: 初期スキャン領域(赤) + モデル領域(緑)")
			except Exception as e:
				print(f"[knee reg:{label}] 経過表示に失敗: {e}")

		return T, best_fit, best_rmse

	def on_knee_register_bone(self, bone: str) -> None:
		"""指定した骨(大腿骨/脛骨)だけを、その骨のパラメータ・方式で位置合わせする。"""
		if bone == "femur":
			model = self.knee_femur_model_path.get().strip()
			src = self.knee_reg_femur_src_path.get().strip()
			tgt = self.knee_reg_femur_tgt_path.get().strip()
			label = "大腿骨"
		else:
			model = self.knee_tibia_model_path.get().strip()
			src = self.knee_reg_tibia_src_path.get().strip()
			tgt = self.knee_reg_tibia_tgt_path.get().strip()
			label = "脛骨"

		if not (model and src and tgt):
			messagebox.showwarning(
				"入力不足",
				f"{label}: モデルと位置合わせ領域（モデル側・初期スキャン側）を選択してください。")
			return

		p = self._knee_get_params(bone)
		method = p.get("method", "ransac")
		try:
			if method == "manual":
				messagebox.showinfo(
					"手動3点位置合わせ",
					f"[{label}] まず『モデル側 領域』で対応点を3つクリックし、ウィンドウを閉じてください。\n"
					f"次に『初期スキャン側 領域』で、同じ順番・同じ部位の3点をクリックしてください。")
				src_pts = self._knee_pick_points(pv.read(src), 3, f"{label}: モデル側 領域（対応点3つ）")
				if len(src_pts) < 3:
					raise ValueError(f"{label}: モデル側の対応点が3つ未満です（{len(src_pts)}点）。")
				tgt_pts = self._knee_pick_points(pv.read(tgt), 3, f"{label}: 初期スキャン側 領域（同順で3つ）")
				if len(tgt_pts) < 3:
					raise ValueError(f"{label}: 初期スキャン側の対応点が3つ未満です（{len(tgt_pts)}点）。")
				T, fit, rmse = self._knee_register_region(src, tgt, label, p,
				                                          manual_src_pts=src_pts, manual_tgt_pts=tgt_pts)
			else:
				T, fit, rmse = self._knee_register_region(src, tgt, label, p)
		except Exception as e:
			messagebox.showerror("位置合わせエラー", f"{label}の位置合わせに失敗しました:\n{e}")
			return

		if bone == "femur":
			self._knee_femur_reg_T = T
		else:
			self._knee_tibia_reg_T = T

		messagebox.showinfo(
			f"{label} 位置合わせ完了",
			f"{label}の位置合わせが完了しました。\n\n"
			f"fitness={fit:.4f}（1に近いほど良好）, RMSE={rmse:.4f} mm（小さいほど良好）\n\n"
			"「位置合わせ結果を確認（大腿骨＋脛骨）」で重ね合わせを確認できます。")

	def on_knee_preview_registration(self) -> None:
		"""位置合わせ結果（初期スキャン＋整列済み大腿骨・脛骨モデル）を重ねて表示する。"""
		scan_path = self.knee_initial_scan_path.get().strip()
		if not scan_path:
			messagebox.showwarning("入力不足", "初期状態スキャンを選択してください。")
			return
		if self._knee_femur_reg_T is None and self._knee_tibia_reg_T is None:
			messagebox.showwarning("未実行", "先に「位置合わせ実行」を行ってください。")
			return
		try:
			scan_mesh = pv.read(scan_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"初期状態スキャンの読み込みに失敗しました:\n{e}")
			return

		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="knee: 位置合わせ結果の確認", window_size=(int(sw * 0.9), int(sh * 0.9)))
		plotter.set_background("white")
		plotter.add_mesh(scan_mesh, color="lightgray", opacity=0.4, smooth_shading=True)

		def add_registered(model_path, T, color):
			if not model_path or T is None:
				return
			try:
				m = pv.read(model_path)
				pts = np.hstack([m.points, np.ones((m.points.shape[0], 1))])
				m.points = (np.asarray(T) @ pts.T).T[:, :3]
				plotter.add_mesh(m, color=color, opacity=1.0, smooth_shading=True)
			except Exception as e:
				print(f"[knee preview] {color} モデル表示失敗: {e}")

		add_registered(self.knee_femur_model_path.get().strip(), self._knee_femur_reg_T, "burlywood")
		add_registered(self.knee_tibia_model_path.get().strip(), self._knee_tibia_reg_T, "lightblue")
		plotter.add_text("初期スキャン(灰) + 大腿骨(整列済,肌色) + 脛骨(整列済,水色)",
		                 position="upper_left", font_size=10, color="black")
		plotter.show()

	# ----- Knee scene sharing (別PC=解析サーバーとキャッシュ整合させるため) -----
	def _knee_scene_file(self) -> Path:
		"""シーン共有ファイル cache/knee_scene.json のパス（git追跡）。"""
		d = Path(__file__).parent / "cache"
		d.mkdir(parents=True, exist_ok=True)
		return d / "knee_scene.json"

	def _file_content_hash(self, path: str) -> str:
		"""ファイル内容(バイト列)のSHA-256短縮ハッシュ。未存在/失敗は空文字。"""
		try:
			p = Path(path)
			if not p.exists():
				return ""
			h = hashlib.sha256()
			with p.open("rb") as f:
				for chunk in iter(lambda: f.read(1 << 20), b""):
					h.update(chunk)
			return h.hexdigest()[:16]
		except Exception:
			return ""

	def _knee_read_scene_file(self) -> dict:
		"""シーン共有ファイルを読み {タブ名: シーン} に正規化して返す（旧v1単一形式も互換）。"""
		p = self._knee_scene_file()
		if not p.exists():
			return {}
		data = json.load(p.open("r", encoding="utf-8"))
		if isinstance(data, dict) and isinstance(data.get("scenes"), dict):
			return data["scenes"]
		if isinstance(data, dict) and ("femur_reg_T" in data or "tibia_reg_T" in data):
			return {"旧シーン": data}  # v1（単一シーン・名前なし）→ 名前付きへ移行
		return {}

	def _knee_write_scene_file(self, scenes: dict) -> Path:
		p = self._knee_scene_file()
		with p.open("w", encoding="utf-8") as f:
			json.dump({"version": 2, "scenes": scenes}, f, ensure_ascii=False, indent=2)
		return p

	def _knee_scene_for_tab(self, snap: dict) -> dict:
		"""タブのスナップショットからシーン共有エントリを作る（パス非依存・内容ハッシュ＋ファイル名ヒント）。"""
		def h(key):
			return self._file_content_hash(str(snap.get(key, "")).strip())
		def bn(key):
			pth = str(snap.get(key, "")).strip()
			return Path(pth).name if pth else ""
		return {
			"w_scan_deg": float(snap.get("knee_w_scan_deg", 0.0) or 0.0),
			"side": int(snap.get("knee_side", 1) or 1),
			"femur_reg_T": snap.get("_femur_reg_T"),
			"tibia_reg_T": snap.get("_tibia_reg_T"),
			"femur_color": str(snap.get("knee_femur_color", "") or ""),
			"tibia_color": str(snap.get("knee_tibia_color", "") or ""),
			# 内容一致確認用（別PCで同じファイルかを警告するため。パスではなく中身のハッシュ）
			"hash_initial_pp": h("knee_initial_pp"),
			"hash_femur_model": h("knee_femur_model"),
			"hash_tibia_model": h("knee_tibia_model"),
			"hash_transform": h("knee_transform"),
			# 別PCでファイルを選び直す際のヒント（ファイル名のみ）
			"file_hints": {
				"initial_scan": bn("knee_initial_scan"),
				"initial_pp": bn("knee_initial_pp"),
				"femur_model": bn("knee_femur_model"),
				"tibia_model": bn("knee_tibia_model"),
				"transform": bn("knee_transform"),
			},
			"saved_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
		}

	def _knee_apply_scene_to_snap(self, snap: dict, sc: dict) -> None:
		"""シーンエントリ（reg_T・W_scan・左右）をタブのスナップショットへ反映する。"""
		try:
			snap["knee_w_scan_deg"] = float(sc.get("w_scan_deg", snap.get("knee_w_scan_deg", 0.0)))
		except Exception:
			pass
		s = sc.get("side")
		if s in (1, 2):
			snap["knee_side"] = int(s)
		for ckey, skey in (("femur_color", "knee_femur_color"), ("tibia_color", "knee_tibia_color")):
			c = sc.get(ckey)
			if c:
				snap[skey] = str(c)
		# 厳密再現のため無条件で同期（保存側で未登録ならNoneになる）
		snap["_femur_reg_T"] = sc.get("femur_reg_T")
		snap["_tibia_reg_T"] = sc.get("tibia_reg_T")

	def _knee_scene_check_files(self, snap: dict, sc: dict) -> list:
		"""タブのファイル内容とシーンの内容ハッシュを照合し、警告文リストを返す。"""
		hints = sc.get("file_hints") or {}
		out = []
		for label, hkey, pkey, hintkey in (
			("特徴点PP", "hash_initial_pp", "knee_initial_pp", "initial_pp"),
			("大腿骨", "hash_femur_model", "knee_femur_model", "femur_model"),
			("脛骨", "hash_tibia_model", "knee_tibia_model", "tibia_model"),
			("変位データ", "hash_transform", "knee_transform", "transform"),
		):
			saved_h = sc.get(hkey, "")
			if not saved_h:
				continue
			path = str(snap.get(pkey, "")).strip()
			cur_h = self._file_content_hash(path) if path else ""
			if not cur_h:
				hint = hints.get(hintkey, "")
				out.append(f"{label}: 未選択" + (f"（→ {hint}）" if hint else ""))
			elif cur_h != saved_h:
				out.append(f"{label}: 内容不一致")
		return out

	def on_knee_scene_save(self) -> None:
		"""アクティブな試験タブのシーンを、タブ名をキーに共有ファイルへ書き出す。"""
		if self._knee_femur_reg_T is None and self._knee_tibia_reg_T is None:
			messagebox.showwarning("シーン保存", "先に位置合わせを行ってください（reg_Tがありません）。")
			return
		self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
		name = self._knee_tabs[self._knee_active_tab].get('name', '試験1')
		try:
			scenes = self._knee_read_scene_file()
		except Exception:
			scenes = {}
		scenes[name] = self._knee_scene_for_tab(self._knee_tabs[self._knee_active_tab]['snapshot'])
		try:
			p = self._knee_write_scene_file(scenes)
		except Exception as e:
			messagebox.showerror("シーン保存", f"保存に失敗しました:\n{e}")
			return
		messagebox.showinfo(
			"シーン保存",
			f"シーン「{name}」を保存しました:\n{p}\n\n"
			"このファイルを git commit/push すると、別PC（サーバー/Mac等）で「シーンを読込」して\n"
			"同じタブ名・同じ位置合わせ結果で計算/再現できます。")

	def on_knee_scene_save_all(self) -> None:
		"""全試験タブのシーンをまとめて共有ファイルへ書き出す（位置合わせ未実施のタブはスキップ）。"""
		if self._knee_tabs:
			self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
		try:
			scenes = self._knee_read_scene_file()
		except Exception:
			scenes = {}
		saved, skipped = [], []
		for t in self._knee_tabs:
			snap = t.get('snapshot') or {}
			name = t.get('name', '?')
			if snap.get('_femur_reg_T') is None and snap.get('_tibia_reg_T') is None:
				skipped.append(name)
				continue
			scenes[name] = self._knee_scene_for_tab(snap)
			saved.append(name)
		if not saved:
			messagebox.showwarning("シーン保存", "reg_T を持つタブがありません。先に位置合わせを行ってください。")
			return
		try:
			p = self._knee_write_scene_file(scenes)
		except Exception as e:
			messagebox.showerror("シーン保存", f"保存に失敗しました:\n{e}")
			return
		msg = f"{len(saved)}タブのシーンを保存しました: {', '.join(saved)}\n{p}"
		if skipped:
			msg += f"\n（位置合わせ未実施のためスキップ: {', '.join(skipped)}）"
		msg += "\n\ngit commit/push すると別PCで「シーンを読込」でき、タブ名ごと引き継がれます。"
		messagebox.showinfo("シーン保存", msg)

	def on_knee_scene_load(self) -> None:
		"""シーン共有ファイルをタブ名で同期する（同名タブ=reg_T等を更新 / 無ければ新規タブ作成）。"""
		p = self._knee_scene_file()
		if not p.exists():
			messagebox.showwarning("シーン読込", f"シーン共有ファイルが見つかりません:\n{p}\n\n"
			                        "先に位置合わせ済みのPCで「シーンを保存」→git共有してください。")
			return
		try:
			scenes = self._knee_read_scene_file()
		except Exception as e:
			messagebox.showerror("シーン読込", f"読み込みに失敗しました:\n{e}")
			return
		if not scenes:
			messagebox.showwarning("シーン読込", "シーン共有ファイルに有効なシーンがありません。")
			return
		# アクティブタブの現状を保存してから同期
		self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
		updated, created, warn_lines = [], [], []
		for name, sc in scenes.items():
			if not isinstance(sc, dict):
				continue
			idx = next((k for k, t in enumerate(self._knee_tabs) if t.get('name') == name), None)
			if idx is None:
				snap = copy.deepcopy(self._knee_default_snap or {})
				self._knee_apply_scene_to_snap(snap, sc)
				self._knee_tabs.append({'name': name, 'snapshot': snap})
				created.append(name)
			else:
				snap = self._knee_tabs[idx].get('snapshot') or {}
				self._knee_apply_scene_to_snap(snap, sc)
				self._knee_tabs[idx]['snapshot'] = snap
				updated.append(name)
			ws = self._knee_scene_check_files(snap, sc)
			if ws:
				warn_lines.append(f"・{name}: " + " / ".join(ws))
		# アクティブタブへ反映（同期で変わった可能性があるため）
		self._knee_restore_snapshot(self._knee_tabs[self._knee_active_tab]['snapshot'])
		self._knee_rebuild_tabbar()
		msg = "シーンを読み込みました（タブ名で同期）。\n"
		if updated:
			msg += f"更新: {', '.join(updated)}\n"
		if created:
			msg += f"新規タブ: {', '.join(created)}\n"
		if warn_lines:
			msg += ("\n⚠ ファイル内容が保存時と一致しない/未選択のタブがあります"
			        "（キャッシュが当たらない可能性）:\n" + "\n".join(warn_lines)
			        + "\n\n各タブで同じ内容のファイルを選択してください（→はファイル名ヒント）。")
		else:
			msg += "\nモデル/変位データの内容一致を確認しました。このまま計算/表示できます。"
		messagebox.showinfo("シーン読込", msg)

	def _knee_apply_T(self, mesh, T):
		"""pvメッシュに4x4同次変換を適用（in-place）してメッシュを返す。"""
		if mesh.n_points == 0:
			return mesh
		pts = np.hstack([mesh.points, np.ones((mesh.points.shape[0], 1))])
		mesh.points = (np.asarray(T, dtype=float) @ pts.T).T[:, :3]
		return mesh

	def _knee_prepare_scene(self, require_models: bool = True) -> dict:
		"""初期スキャン特徴点から Cf/Ct を構築し、ワールド座標系(Cf=原点)へ整列した
		大腿骨・脛骨モデルと座標系を準備する。both 静的可視化 と アニメーション で共用。

		Returns dict: A(Cf→world), w_scan, points_world, labels, cf_world, ct_world,
		              femur_world, tibia_world (require_models=Trueのとき)。
		エラーは ValueError(メッセージ) で送出。
		"""
		scan_path = self.knee_initial_scan_path.get().strip()
		pp_path = self.knee_initial_pp_path.get().strip()
		if not pp_path:
			raise ValueError("初期スキャン特徴点(PP)を選択してください。")
		points, labels = self._parse_pp_file(pp_path)
		origin, cfx, cfy, cfz = self._build_knee_femur_frame(points, labels, self.knee_side_var.get())
		try:
			w_scan = float(self.knee_w_scan_deg.get())
		except Exception:
			w_scan = 0.0
		t_o, t_x, t_y, t_z = self._build_knee_tibia_frame(origin, cfx, cfy, cfz, w_scan)

		# A: Cf座標系 → ワールド座標系（Cfを原点・単位軸に整列）
		Rf = np.column_stack([cfx, cfy, cfz])
		A = np.eye(4)
		A[:3, :3] = Rf.T
		A[:3, 3] = -Rf.T @ origin

		# 特徴点をワールドへ
		pw = np.hstack([points, np.ones((points.shape[0], 1))])
		points_world = (A @ pw.T).T[:, :3]

		# Ct をワールドへ（原点=0、軸は回転のみ）
		ct_axes_world = (Rf.T @ np.column_stack([t_x, t_y, t_z]))
		ct_world = (np.array([0.0, 0.0, 0.0]), ct_axes_world[:, 0], ct_axes_world[:, 1], ct_axes_world[:, 2])
		cf_world = (np.array([0.0, 0.0, 0.0]), np.array([1.0, 0, 0]), np.array([0, 1.0, 0]), np.array([0, 0, 1.0]))

		result = {"A": A, "w_scan": w_scan, "points_world": points_world, "labels": labels,
		          "cf_world": cf_world, "ct_world": ct_world}

		if require_models:
			femur_path = self.knee_femur_model_path.get().strip()
			tibia_path = self.knee_tibia_model_path.get().strip()
			if not femur_path or not tibia_path:
				raise ValueError("大腿骨モデルと脛骨モデルを選択してください。")
			if self._knee_femur_reg_T is None or self._knee_tibia_reg_T is None:
				raise ValueError("先に③「位置合わせ実行」で大腿骨・脛骨を初期スキャンへ整列してください。")
			femur = pv.read(femur_path)
			tibia = pv.read(tibia_path)
			# モデル → 初期スキャン座標系(reg) → ワールド(A)
			self._knee_apply_T(femur, A @ np.asarray(self._knee_femur_reg_T))
			self._knee_apply_T(tibia, A @ np.asarray(self._knee_tibia_reg_T))
			result["femur_world"] = femur
			result["tibia_world"] = tibia
		return result

	def _knee_draw_frame(self, plotter, o, ax, ay, az, prefix, axis_len, solid=True):
		"""PyVistaに座標系(3軸)を描画する。"""
		for vec, col, name in [(ax, "red", f"{prefix}X"), (ay, "green", f"{prefix}Y"), (az, "blue", f"{prefix}Z")]:
			end = o + np.asarray(vec) * axis_len
			plotter.add_mesh(pv.Line(o, end), color=col, line_width=(5 if solid else 2))
			plotter.add_point_labels([end], [name], font_size=13, text_color=col,
			                         point_size=1, show_points=False, always_visible=True)

	def on_knee_visualize_all(self) -> None:
		"""大腿骨・脛骨モデル（位置合わせ済み）と Cf/Ct を全体可視化する。"""
		try:
			scene = self._knee_prepare_scene(require_models=True)
		except Exception as e:
			messagebox.showwarning("全体可視化", str(e))
			return

		femur = scene["femur_world"]; tibia = scene["tibia_world"]
		try:
			b = np.array(femur.bounds).reshape(3, 2)
			diag = float(np.linalg.norm(b[:, 1] - b[:, 0]))
			axis_len = max(diag * 0.15, 10.0)
		except Exception:
			axis_len = 40.0

		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="knee: 関節全体（位置合わせ済み + Cf/Ct）",
		                     window_size=(int(sw * 0.9), int(sh * 0.9)))
		plotter.set_background("white")
		plotter.add_mesh(femur, color=self._knee_color("femur"), opacity=1.0, smooth_shading=True, show_edges=False)
		plotter.add_mesh(tibia, color=self._knee_color("tibia"), opacity=1.0, smooth_shading=True, show_edges=False)

		# 特徴点
		named = self._knee_named_points(scene["points_world"], scene["labels"])
		colors = {"MCL": "red", "LCL": "blue", "L": "green", "M": "orange", "N": "purple"}
		for name, col in colors.items():
			if name in named:
				plotter.add_mesh(pv.Sphere(radius=max(axis_len * 0.04, 1.0), center=named[name]), color=col)

		cfo, cfx, cfy, cfz = scene["cf_world"]
		self._knee_draw_frame(plotter, cfo, cfx, cfy, cfz, "Cf_", axis_len, solid=True)
		if abs(scene["w_scan"]) > 1e-6:
			cto, ctx, cty, ctz = scene["ct_world"]
			self._knee_draw_frame(plotter, cto, ctx, cty, ctz, "Ct_", axis_len, solid=False)

		plotter.add_text("大腿骨 + 脛骨（色は④の表示色設定） / Cf: X=red Y=green Z=blue（Ctは細線）",
		                 position="upper_left", font_size=10, color="black")
		plotter.show()

	def on_knee_animate(self) -> None:
		"""変位・姿勢変化データに沿って脛骨を動かすシミュレーション（共通コアを利用）。

		大腿骨(Cf)固定・脛骨を fmTtb(Θ(t)) で駆動する。脛骨メッシュの初期配置からは
		スキャン時屈曲 fmTtb(FE=W_scan) を差し引く（ヒト膝 W=0 で恒等）。
		実際のアニメ/ヒートマップ描画は hip と共通の on_animate エンジンを再利用する。
		"""
		transform_path = self.knee_transform_path.get().strip()
		if not transform_path:
			messagebox.showwarning("シミュレーション", "変位・姿勢変化データ(xlsx/kkr)を選択してください。")
			return
		try:
			scene = self._knee_prepare_scene(require_models=True)
		except Exception as e:
			messagebox.showwarning("シミュレーション", str(e))
			return
		try:
			transform_data = self._load_transform_matrices(transform_path)
			if not transform_data:
				messagebox.showerror("エラー", "変位・姿勢変化データが空です。")
				return
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"変位・姿勢変化データの読み込みに失敗しました:\n{e}")
			return

		femur = scene["femur_world"]
		tibia = scene["tibia_world"]

		# 脛骨の初期配置から W_scan ぶんの屈曲を差し引く: fmTtb(FE=W_scan)^-1
		# （transform_data の絶対屈曲 FE(t) がそのまま脛骨姿勢になる。ヒト膝 W=0 で恒等。
		#   後で on_animate が同じ fmTtb を再適用するので符号規約に依存せず自己整合する）
		w_scan = scene["w_scan"]
		if abs(w_scan) > 1e-9:
			T_wscan = self._build_transform_matrix(w_scan, 0.0, 0.0, 0.0, 0.0, 0.0)
			self._knee_apply_T(tibia, np.linalg.inv(T_wscan))

		# --- hip 共通の on_animate エンジンをそのまま再利用（アダプタ方式・hip側は無改変） ---
		# ワールド整列済みの大腿骨(prox)/脛骨(dist)を一時STLに書き出し、hipの座標系ビルダーが
		# 「単位座標系」を返す特徴点(pp)を合成して与える。こうすると on_animate 内の座標変換が
		# 恒等になり、準備済みメッシュをそのまま変位データで動かせる（=同じ描画/ヒートマップ資産）。
		import tempfile, shutil
		tmpdir = Path(tempfile.mkdtemp(prefix="frs_knee_anim_"))

		def _radius_of(mesh):
			try:
				b = np.array(mesh.bounds).reshape(3, 2)
				return max(float(np.linalg.norm(b[:, 1] - b[:, 0])) * 0.5, 20.0)
			except Exception:
				return 60.0

		try:
			femur_stl = tmpdir / "femur_world.stl"
			tibia_stl = tmpdir / "tibia_world.stl"
			femur.triangulate().save(str(femur_stl))
			tibia.triangulate().save(str(tibia_stl))
			# 単位座標系を生む特徴点（hipビルダーで identity を返すことを検証済み）
			prox_abcd = tmpdir / "prox_abcd.pp"; prox_olmn = tmpdir / "prox_olmn.pp"
			dist_abc = tmpdir / "dist_abc.pp"; dist_olmn = tmpdir / "dist_olmn.pp"
			self._knee_write_pp(prox_abcd, [((0, 0, 1), "A"), ((0, 0, 0), "B"), ((-1, 0, 0.5), "C"), ((-1, 0, 0.5), "D")])
			self._knee_write_pp(prox_olmn, [((0, 0, 0), "O"), ((1, 0, 0), "L"), ((0, 1, 0), "M"), ((0, 0, 1), "N")])
			self._knee_write_pp(dist_abc, [((0, -0.001, 1), "A"), ((0, -0.001, 0), "B")])
			self._knee_write_pp(dist_olmn, [((0, 0, 0), "O"), ((1, 0, 0), "L"), ((0, 1, 0), "M"), ((0, 0, 1), "N")])
		except Exception as e:
			try:
				shutil.rmtree(tmpdir)
			except Exception:
				pass
			messagebox.showerror("準備失敗", f"アニメーション用データの準備に失敗しました:\n{e}")
			return

		# 現在の hip 側状態を退避（on_animate 実行中だけ差し替え、終了後に必ず復元）
		saved = {}
		save_names = (
			"joint_var", "prox_model_path", "prox_pp_abcd_path", "prox_pp_olmn_path",
			"dist_model_path", "dist_pp_abc_path", "dist_pp_olmn_path", "transform_group_path",
			"prox_cartilage_model_path", "dist_cartilage_model_path",
			"prox_offset_x", "prox_offset_y", "prox_offset_z",
			"dist_offset_x", "dist_offset_y", "dist_offset_z",
			"prox_radius", "dist_radius", "show_ao_angle",
		)
		for nm in save_names:
			try:
				saved[nm] = getattr(self, nm).get()
			except Exception:
				pass
		saved_prox_color = getattr(self, "prox_color", None)
		saved_dist_color = getattr(self, "dist_color", None)
		try:
			self.joint_var.set(1)  # hip 座標系ビルダーを使用
			self.prox_model_path.set(str(femur_stl))
			self.prox_pp_abcd_path.set(str(prox_abcd))
			self.prox_pp_olmn_path.set(str(prox_olmn))
			self.dist_model_path.set(str(tibia_stl))
			self.dist_pp_abc_path.set(str(dist_abc))
			self.dist_pp_olmn_path.set(str(dist_olmn))
			self.transform_group_path.set(transform_path)
			self.prox_cartilage_model_path.set("")
			self.dist_cartilage_model_path.set("")
			for v in (self.prox_offset_x, self.prox_offset_y, self.prox_offset_z,
			          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z):
				v.set(0.0)
			self.prox_radius.set(_radius_of(femur))
			self.dist_radius.set(_radius_of(tibia))
			self.show_ao_angle.set(False)
			self.prox_color = self._knee_color("femur")  # 大腿骨（④の表示色設定、既定=肌色）
			self.dist_color = self._knee_color("tibia")  # 脛骨（④の表示色設定、既定=水色）
			# hip 共通エンジンを起動（モーダル：ウィンドウを閉じるまでブロック）
			self.on_animate()
		finally:
			for nm, val in saved.items():
				try:
					getattr(self, nm).set(val)
				except Exception:
					pass
			if saved_prox_color is not None:
				self.prox_color = saved_prox_color
			if saved_dist_color is not None:
				self.dist_color = saved_dist_color
			try:
				shutil.rmtree(tmpdir)
			except Exception:
				pass

	def _knee_write_pp(self, path, entries) -> None:
		"""特徴点を PickedPoints XML 形式で書き出す。entries: [((x,y,z), label), ...]。"""
		root = ET.Element('PickedPoints')
		for (x, y, z), lab in entries:
			e = ET.SubElement(root, 'point')
			e.set('x', f"{float(x):.8f}")
			e.set('y', f"{float(y):.8f}")
			e.set('z', f"{float(z):.8f}")
			e.set('active', "1")
			e.set('name', str(lab))
		ET.ElementTree(root).write(str(path), encoding="utf-8", xml_declaration=True)
	# endregion knee simulator

	# ================================================================
	# region ankle simulator（足関節・ArUcoマーカートラッキング方式）
	# ================================================================
	# データフロー（膝/股と異なる: 近位/遠位2骨ではなく N骨）:
	#   ① 初期状態スキャン (ピン+マーカー装着後)
	#   ② RGB-D計測データ (video + depth + camera intrinsics + ArUco設定)
	#   ③ 骨リスト (N本可変) — 骨ごとに ArUco ID / 試験後スキャン /
	#       マーカー-骨キャリブレーション / 位置合わせ / 表示色
	#   ④ ArUco検出 → PnP → 姿勢時系列を計算/キャッシュ
	#   ⑤ 可視化・アニメーション・骨対ヒートマップ (膝のLUT流用)

	_ANKLE_DEFAULT_BONE_COLORS = (
		"#DEB887", "#ADD8E6", "#98FB98", "#FFB6C1", "#DDA0DD",
		"#FFD700", "#87CEFA", "#F4A460", "#B0E0E6", "#FFA07A")

	_ANKLE_ARUCO_DICT_CHOICES = (
		"DICT_4X4_50", "DICT_4X4_100", "DICT_4X4_250", "DICT_4X4_1000",
		"DICT_5X5_50", "DICT_5X5_100", "DICT_5X5_250", "DICT_5X5_1000",
		"DICT_6X6_50", "DICT_6X6_100", "DICT_6X6_250", "DICT_6X6_1000",
		"DICT_APRILTAG_36h11")

	def _ankle_default_bone(self, index: int) -> dict:
		"""新規骨のデフォルト値。"""
		color = self._ANKLE_DEFAULT_BONE_COLORS[index % len(self._ANKLE_DEFAULT_BONE_COLORS)]
		return {
			"name": f"骨{index + 1}",
			"aruco_id": index,
			"model_path": "",            # 可視化用フルモデル (試験後スキャン)
			"calib_model_path": "",      # キャリブ用モデル (省略時 model_path を使う、ArUco切り出し推奨)
			"post_scan_path": "",
			"color": color,
			"method": "ransac",         # 'ransac' | 'pca' | 'manual'
			"enable_scaling": False,
			"marker_to_bone_T": None,
			"reg_T": None,
			"fixed": False,             # True: この骨を固定して他骨の相対姿勢を表示 (骨リスト中で1本のみ)
		}

	# ---- テクスチャ対応 メッシュ読込ヘルパ ----
	def _ankle_load_mesh_and_texture(self, path: str):
		"""メッシュ + テクスチャを読み込む。テクスチャ取得できなければ None を返す。

		対応形式:
		- PLY: mesh.textures に埋め込みテクスチャがあれば使用
		- OBJ: 同名の .mtl から map_Kd 行を探して画像パスを解決
		- 同名の .png/.jpg/.jpeg/.tif/.bmp を fallback で探索

		Returns: (mesh: pv.PolyData, texture: pv.Texture or None)
		"""
		p = Path(path)
		mesh = pv.read(str(p))
		if mesh is None:
			return None, None
		tex = None
		# 1. mesh.textures に既に入っている場合 (PLY埋込など)
		try:
			if hasattr(mesh, 'textures') and mesh.textures:
				first_key = next(iter(mesh.textures))
				tex = mesh.textures[first_key]
		except Exception:
			tex = None
		# 2. OBJ の場合: 同名 .mtl から map_Kd を解析
		if tex is None and p.suffix.lower() == '.obj':
			mtl_path = p.with_suffix('.mtl')
			if mtl_path.exists():
				try:
					with mtl_path.open('r', encoding='utf-8', errors='ignore') as f:
						for line in f:
							s = line.strip()
							if not s or s.startswith('#'):
								continue
							# map_Kd texture.png  or  map_Kd -options texture.png
							low = s.lower()
							if low.startswith('map_kd'):
								parts = s.split()
								if len(parts) >= 2:
									tex_name = parts[-1]  # 最後の要素 = ファイル名
									tex_p = Path(tex_name)
									if not tex_p.is_absolute():
										tex_p = p.parent / tex_p
									if tex_p.exists():
										try:
											tex = pv.read_texture(str(tex_p))
											print(f"[tex] OBJ+MTL からテクスチャ読込: {tex_p.name}")
											break
										except Exception as e:
											print(f"[tex] {tex_p.name} 読込失敗: {e}")
				except Exception as e:
					print(f"[tex] MTL 解析失敗: {e}")
		# 3. 同名の画像ファイルを探す (最終手段)
		if tex is None:
			for ext in ('.png', '.jpg', '.jpeg', '.tif', '.tiff', '.bmp'):
				candidate = p.with_suffix(ext)
				if candidate.exists():
					try:
						tex = pv.read_texture(str(candidate))
						print(f"[tex] 同名画像 fallback: {candidate.name}")
						break
					except Exception:
						continue
		if tex is None:
			print(f"[tex] テクスチャ見つからず: {p.name}")
		return mesh, tex

	def _ankle_get_calib_model_path(self, bone: dict) -> str:
		"""キャリブ用モデルパスを返す。calib_model_path が空なら model_path を fallback。"""
		p = str(bone.get("calib_model_path", "") or "").strip()
		if p:
			return p
		return str(bone.get("model_path", "") or "").strip()

	def _create_ankle_simulator_tab(self) -> None:
		"""ankle simulator タブのUIを構築（ArUcoマーカートラッキング方式）。"""
		tabbar_row = tk.Frame(self.ankle_simulator_tab)
		tabbar_row.pack(side="top", fill="x", padx=4, pady=(4, 0))
		tk.Label(tabbar_row, text="試験タブ:", font=(self.ui_font_family, 9)).pack(side="left", padx=(0, 4))
		self._ankle_tabbar_frame = tk.Frame(tabbar_row)
		self._ankle_tabbar_frame.pack(side="left", fill="x")

		canvas = tk.Canvas(self.ankle_simulator_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.ankle_simulator_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)
		scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")

		def _on_mousewheel(event):
			step = _mousewheel_units(event)
			if step:
				canvas.yview_scroll(step, "units")
		canvas.bind("<Enter>", lambda e: (canvas.bind_all("<MouseWheel>", _on_mousewheel),
		                                   canvas.bind_all("<Button-4>", _on_mousewheel),
		                                   canvas.bind_all("<Button-5>", _on_mousewheel)))
		canvas.bind("<Leave>", lambda e: (canvas.unbind_all("<MouseWheel>"),
		                                   canvas.unbind_all("<Button-4>"),
		                                   canvas.unbind_all("<Button-5>")))

		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		ttk.Label(container, text="ankle simulator（足関節・ArUcoマーカートラッキング）",
		          font=(self.ui_font_family, 12, "bold")).grid(row=0, column=0, sticky="w", padx=4, pady=(4, 2))
		ttk.Label(container,
		          text="骨に刺したM3ピン+ArUcoマーカーをRGB-Dカメラで追跡し、任意N本(脛骨,距骨,他)の"
		               "6自由度姿勢を復元。試験後解剖時に取得したマーカー-骨変換で骨自体の動きに変換します。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=1, column=0, sticky="w", padx=4, pady=(0, 8))

		# --- 動作モード選択 (新プラン / 原プラン / 簡易版) ---
		mode_frame = ttk.LabelFrame(container, text="動作モード (座標系の決め方)", style="Bold.TLabelframe")
		mode_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 10))
		mode_frame.columnconfigure(0, weight=1)
		mf = ttk.Frame(mode_frame)
		mf.grid(row=0, column=0, sticky="w", padx=12, pady=(6, 2))
		ttk.Radiobutton(mf, text="🟢 新プラン【推奨】(マーカー付き骨単独スキャンから T_L←Mk 自己決定)",
		                value="self_pose", variable=self.ankle_workflow_mode
		                ).grid(row=0, column=0, sticky="w", padx=(0, 0))
		ttk.Radiobutton(mf, text="⚪ 原プラン (①組立スキャン + ③位置合わせで T_W←L 決定)",
		                value="original", variable=self.ankle_workflow_mode
		                ).grid(row=1, column=0, sticky="w", padx=(0, 0))
		ttk.Radiobutton(mf, text="⚡ 簡易版 (3Dスキャン無し, クランプArUcoで関節座標系Cj校正 → 骨マーカー動揺性)",
		                value="simple", variable=self.ankle_workflow_mode
		                ).grid(row=2, column=0, sticky="w", padx=(0, 0))
		wf = ttk.Frame(mode_frame)
		wf.grid(row=1, column=0, sticky="w", padx=12, pady=(4, 6))
		ttk.Label(wf, text="新プランのワールド系 W:", foreground="#666"
		          ).grid(row=0, column=0, sticky="w", padx=(0, 6))
		ttk.Radiobutton(wf, text="カメラ参照フレーム", value="camera_ref",
		                variable=self.ankle_self_pose_world
		                ).grid(row=0, column=1, sticky="w", padx=(0, 8))
		ttk.Radiobutton(wf, text="骨A基準 (ヒートマップ骨A)", value="boneA_ref",
		                variable=self.ankle_self_pose_world
		                ).grid(row=0, column=2, sticky="w")
		ttk.Label(mode_frame,
		          text="🟢 新プラン: ①+③位置合わせは不要。各骨モデルに「🔍 スキャンから自動キャリブ」または手動キャリブで T_L←Mk が必要。組立スキャン省略で工数減。\n"
		               "⚪ 原プラン: ①+③位置合わせが必要。マーカー-骨キャリブは不要 (剛結合仮定)。\n"
		               "⚡ 簡易版: 3Dスキャン不要。クランプにArUco貼付→ロボットML/AP/PDで軸方向を計測→関節座標系Cjを校正→骨マーカー動揺をCj系で出力。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=2, column=0, sticky="w", padx=12, pady=(0, 6))

		# --- ⚡ 関節座標系校正 (簡易版モードで使用) ---
		cj_frame = ttk.LabelFrame(container, text="⚡ 関節座標系校正 (Cj) — 簡易版モードで使用",
		                            style="Bold.TLabelframe")
		cj_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 8))
		cj_frame.columnconfigure(0, weight=1)
		self._ankle_cj_frame = cj_frame
		cj_row1 = ttk.Frame(cj_frame)
		cj_row1.grid(row=0, column=0, sticky="w", padx=12, pady=(6, 2))
		ttk.Label(cj_row1, text="クランプArUco ID:").grid(row=0, column=0, sticky="w")
		ttk.Spinbox(cj_row1, from_=0, to=999, textvariable=self.ankle_clamp_aruco_id, width=6
		            ).grid(row=0, column=1, sticky="w", padx=(4, 12))
		ttk.Label(cj_row1, text="各軸校正の記録秒数:").grid(row=0, column=2, sticky="w")
		ttk.Entry(cj_row1, textvariable=self.ankle_axis_calib_seconds, width=6
		          ).grid(row=0, column=3, sticky="w", padx=(4, 12))
		cj_row2 = ttk.Frame(cj_frame)
		cj_row2.grid(row=1, column=0, sticky="w", padx=12, pady=(4, 2))
		ttk.Button(cj_row2, text="🔴 ML軸校正 (内-外)",
		           command=lambda: self.on_ankle_calibrate_axis("ML")
		           ).grid(row=0, column=0, padx=(0, 4))
		ttk.Button(cj_row2, text="🟢 AP軸校正 (前-後)",
		           command=lambda: self.on_ankle_calibrate_axis("AP")
		           ).grid(row=0, column=1, padx=(0, 4))
		ttk.Button(cj_row2, text="🔵 PD軸校正 (近-遠)",
		           command=lambda: self.on_ankle_calibrate_axis("PD")
		           ).grid(row=0, column=2, padx=(0, 12))
		ttk.Button(cj_row2, text="関節座標系Cjを確定",
		           command=self.on_ankle_finalize_joint_frame
		           ).grid(row=0, column=3, padx=(0, 4))
		ttk.Button(cj_row2, text="Cjクリア",
		           command=self.on_ankle_clear_joint_frame
		           ).grid(row=0, column=4, padx=(0, 4))
		ttk.Label(cj_frame, textvariable=self.ankle_axis_calib_status,
		          foreground="#005580", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=2, column=0, sticky="w", padx=12, pady=(2, 2))
		cj_row3 = ttk.Frame(cj_frame)
		cj_row3.grid(row=3, column=0, sticky="w", padx=12, pady=(2, 6))
		ttk.Button(cj_row3, text="骨マーカーをCj系で分析 (CSV+グラフ)",
		           command=self.on_ankle_analyze_in_cj
		           ).grid(row=0, column=0, padx=(0, 4))
		ttk.Label(cj_frame,
		          text="手順: (1) クランプにArUco貼付 → (2) 各軸校正ボタンで各方向にロボットを動かして記録\n"
		               "(3) 3軸そろったら「関節座標系Cjを確定」→ (4) ⓪で本試験録画 → ④で検出 → (5) 「Cj系で分析」でCSV/グラフ出力",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=4, column=0, sticky="w", padx=12, pady=(0, 6))

		# ⓪ RealSense D405 ライブ撮影 (①より前 = 実試験で最初に使うため)
		rs_frame = ttk.LabelFrame(container, text="⓪ RealSense D405 ライブ撮影 (.db3を直接生成)", style="Bold.TLabelframe")
		rs_frame.grid(row=4, column=0, sticky="nsew", pady=(0, 8))
		rs_frame.columnconfigure(0, weight=1)
		rf1 = ttk.Frame(rs_frame)
		rf1.grid(row=0, column=0, sticky="w", padx=12, pady=(6, 2))
		ttk.Label(rf1, text="解像度@fps:").grid(row=0, column=0, sticky="w")
		ttk.Combobox(rf1, textvariable=self.ankle_rs_resolution, width=16, state="readonly",
		             values=["848x480@60", "1280x720@30", "1280x720@15", "640x480@30", "640x480@15"]
		             ).grid(row=0, column=1, sticky="w", padx=(4, 12))
		ttk.Label(rf1, text="開始前 破棄フレーム:").grid(row=0, column=2, sticky="w")
		ttk.Spinbox(rf1, from_=0, to=200, textvariable=self.ankle_rs_discard_frames, width=6
		            ).grid(row=0, column=3, sticky="w", padx=(4, 12))
		ttk.Checkbutton(rf1, text="深度手動露光", variable=self.ankle_rs_manual_exposure
		                ).grid(row=0, column=4, sticky="w", padx=(0, 4))
		ttk.Entry(rf1, textvariable=self.ankle_rs_exposure_val, width=8
		          ).grid(row=0, column=5, sticky="w")
		# --- カラー露光: ArUco のブラーを決めるので独立して制御する ---
		rf1b = ttk.Frame(rs_frame)
		rf1b.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 2))
		ttk.Checkbutton(rf1b, text="カラー手動露光 (ArUco精度に直結)",
		                variable=self.ankle_rs_color_manual_exposure
		                ).grid(row=0, column=0, sticky="w")
		ttk.Label(rf1b, text="露光:").grid(row=0, column=1, sticky="w", padx=(12, 2))
		ttk.Spinbox(rf1b, from_=100, to=33000, increment=500,
		            textvariable=self.ankle_rs_color_exposure_us, width=8
		            ).grid(row=0, column=2, sticky="w")
		ttk.Label(rf1b, text="μs").grid(row=0, column=3, sticky="w", padx=(2, 12))
		ttk.Label(rf1b, text="ゲイン:").grid(row=0, column=4, sticky="w", padx=(0, 2))
		ttk.Spinbox(rf1b, from_=16, to=248, increment=8,
		            textvariable=self.ankle_rs_color_gain, width=6
		            ).grid(row=0, column=5, sticky="w")
		ttk.Label(rs_frame,
		          text="※ モーションブラーが姿勢精度の最大の悪化要因（実測で最悪 8.3 倍）。"
		               "露光 5000μs=5ms なら、骨が 20mm/s で動いても 220mm 距離でブラーは約 0.3px に収まる。"
		               "露光を詰めると暗くなるので、照明を明るくするかゲインを上げて補うこと。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=2, column=0, sticky="w", padx=12, pady=(0, 4))
		rf2 = ttk.Frame(rs_frame)
		rf2.grid(row=3, column=0, sticky="w", padx=12, pady=(2, 4))
		ttk.Button(rf2, text="接続確認", command=self.on_ankle_rs_test_connection
		           ).grid(row=0, column=0, padx=(0, 8))
		ttk.Button(rf2, text="プレビュー+録画 (.db3保存)", command=self.on_ankle_rs_capture
		           ).grid(row=0, column=1, padx=(0, 8))
		ttk.Label(rs_frame, textvariable=self.ankle_rs_status,
		          foreground="#005580", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=4, column=0, sticky="w", padx=12, pady=(0, 2))
		ttk.Label(rs_frame,
		          text="接続確認: pipeline起動+カメラ情報表示 / "
		               "録画: プレビュー→スペースで開始→スペース(またはESC)で停止 → .db3 を ankle_depth_path に自動セット。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=5, column=0, sticky="w", padx=12, pady=(0, 6))

		# ① 初期状態スキャン
		init_frame = ttk.LabelFrame(container, text="① 初期状態スキャン（ピン+マーカー装着後）— 原プランのみ",
		                              style="Bold.TLabelframe")
		init_frame.grid(row=5, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			init_frame.columnconfigure(i, weight=[0, 1, 0][i])
		self._ankle_init_frame = init_frame   # トグル用参照
		self._add_file_row(init_frame, 0, "初期状態スキャン (STL/OBJ)", self.ankle_initial_scan_path,
		                   lambda: self._ankle_choose(self.ankle_initial_scan_path, "初期状態スキャンを選択", "model"))
		self._ankle_init_note = ttk.Label(init_frame,
		          text="※ 原プランのみ使用。全骨がピン+マーカーを装着した状態でスキャンされたもの。"
		               "新プラン/簡易版では不要 (空欄でOK)。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left")
		self._ankle_init_note.grid(row=1, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 4))

		# ② 計測データ
		meas_frame = ttk.LabelFrame(container, text="② RGB-D計測データ / ArUco設定", style="Bold.TLabelframe")
		meas_frame.grid(row=6, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3):
			meas_frame.columnconfigure(i, weight=[0, 1, 0][i])
		self._add_file_row(meas_frame, 0, "RGBビデオ (mp4/avi)", self.ankle_video_path,
		                   lambda: self._ankle_choose(self.ankle_video_path, "RGBビデオを選択", "video"))
		self._add_file_row(meas_frame, 1, "深度データ (bag/npz/mkv)", self.ankle_depth_path,
		                   lambda: self._ankle_choose(self.ankle_depth_path, "深度データを選択", "depth"))
		self._add_file_row(meas_frame, 2, "カメラ内部パラメータ (json/yaml)", self.ankle_camera_intrinsics_path,
		                   lambda: self._ankle_choose(self.ankle_camera_intrinsics_path, "カメラ内部パラメータを選択", "intrinsics"))
		af = ttk.Frame(meas_frame)
		af.grid(row=3, column=0, columnspan=3, sticky="w", padx=12, pady=(4, 4))
		ttk.Label(af, text="ArUco辞書:").grid(row=0, column=0, sticky="w")
		ttk.Combobox(af, textvariable=self.ankle_aruco_dict_var, width=18, state="readonly",
		             values=list(self._ANKLE_ARUCO_DICT_CHOICES)).grid(row=0, column=1, sticky="w", padx=(4, 12))
		ttk.Label(af, text="マーカー実寸 (mm):").grid(row=0, column=2, sticky="w")
		ttk.Entry(af, textvariable=self.ankle_marker_size_mm, width=8).grid(row=0, column=3, sticky="w", padx=(4, 12))
		ttk.Checkbutton(af, text="深度スケールを自動補正",
		                variable=self.ankle_depth_scale_autofix
		                ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(4, 0))
		ttk.Label(af, text="※ SDKが報告する depth_scale が実データと10倍等ズレる場合、"
		               "ArUcoの実寸基準と照合して自動補正します。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=700, justify="left"
		          ).grid(row=2, column=0, columnspan=4, sticky="w", padx=(20, 0))
		ttk.Label(af, text="姿勢推定手法:").grid(row=3, column=0, sticky="w", pady=(8, 0))
		ttk.Combobox(af, textvariable=self.ankle_pose_method, width=30, state="readonly",
		             values=["rgb  ★推奨 (最も安定)",
		                     "depth-corners  (4〜10倍 悪化)",
		                     "fusion  (6〜14倍 悪化)"]
		             ).grid(row=3, column=1, sticky="w", pady=(8, 0))
		ttk.Label(af, text="比較する冒頭フレーム数:").grid(row=3, column=2, sticky="w", pady=(8, 0))
		ttk.Spinbox(af, from_=0, to=2000, increment=100,
		            textvariable=self.ankle_compare_frames, width=6
		            ).grid(row=3, column=3, sticky="w", pady=(8, 0))
		ttk.Label(af,
		          text="※ rgb = RGB画像の角だけで姿勢を出す。深度は表裏判定と深度スケール検証にのみ使用。\n"
		               "   depth-corners = 角の位置の深度を読んで3D化。角は境界なので深度が最も荒れる。\n"
		               "   fusion = マーカー面の深度に平面をあて、RGBの視線と交差させる。平面の法線が約3°ばらつく。\n"
		               "   実機計測では rgb が最良。この設定は ④検出の実行時に効くので、変えたら ④ を押し直すこと。\n"
		               "   冒頭の指定フレームだけ3手法を計算し、実行後に比較レポートを出します (0 で無効)。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=700, justify="left"
		          ).grid(row=4, column=0, columnspan=4, sticky="w", padx=(20, 0))
		ttk.Label(af, text="平面フィット範囲 (fusion 用・マーカー幅の倍率):").grid(
			row=5, column=0, columnspan=2, sticky="w", pady=(6, 0))
		ttk.Spinbox(af, from_=1.0, to=5.0, increment=0.1,
		            textvariable=self.ankle_plane_expand, width=6
		            ).grid(row=5, column=2, sticky="w", pady=(6, 0))
		self._add_file_row(meas_frame, 4, "姿勢時系列 (事前計算済, 任意)", self.ankle_pose_series_path,
		                   lambda: self._ankle_choose(self.ankle_pose_series_path, "姿勢時系列ファイル(npz/csv)を選択", "pose"))

		# ③ 骨リスト
		bones_frame = ttk.LabelFrame(container, text="③ 骨リスト（N本可変：脛骨・距骨・他）", style="Bold.TLabelframe")
		bones_frame.grid(row=7, column=0, sticky="nsew", pady=(0, 8))
		bones_frame.columnconfigure(0, weight=0)
		bones_frame.columnconfigure(1, weight=1)

		left = ttk.Frame(bones_frame)
		left.grid(row=0, column=0, sticky="ns", padx=(8, 4), pady=6)
		ttk.Label(left, text="骨リスト", font=(self.ui_font_family, 9, "bold")).grid(row=0, column=0, columnspan=3, sticky="w")
		self._ankle_bones_listbox = tk.Listbox(left, height=8, width=22, exportselection=False,
		                                        font=(self.ui_font_family, 9))
		self._ankle_bones_listbox.grid(row=1, column=0, columnspan=3, sticky="ns", pady=(2, 4))
		self._ankle_bones_listbox.bind("<<ListboxSelect>>", lambda e: self._ankle_on_bone_select())
		self._ankle_bones_listbox.bind("<Double-Button-1>", lambda e: self.on_ankle_bone_rename())
		btnrow = ttk.Frame(left)
		btnrow.grid(row=2, column=0, columnspan=3, sticky="w")
		ttk.Button(btnrow, text="＋追加", width=6, command=self.on_ankle_bone_add).grid(row=0, column=0, padx=1)
		ttk.Button(btnrow, text="複製", width=5, command=self.on_ankle_bone_duplicate).grid(row=0, column=1, padx=1)
		ttk.Button(btnrow, text="削除", width=5, command=self.on_ankle_bone_delete).grid(row=0, column=2, padx=1)
		btnrow2 = ttk.Frame(left)
		btnrow2.grid(row=3, column=0, columnspan=3, sticky="w", pady=(2, 0))
		ttk.Button(btnrow2, text="↑", width=3, command=lambda: self.on_ankle_bone_move(-1)).grid(row=0, column=0, padx=1)
		ttk.Button(btnrow2, text="↓", width=3, command=lambda: self.on_ankle_bone_move(1)).grid(row=0, column=1, padx=1)
		ttk.Button(btnrow2, text="名前変更", width=9, command=self.on_ankle_bone_rename).grid(row=0, column=2, padx=1)
		btnrow3 = ttk.Frame(left)
		btnrow3.grid(row=4, column=0, columnspan=3, sticky="w", pady=(2, 0))
		ttk.Button(btnrow3, text="🔒 固定/解除", width=15, command=self.on_ankle_bone_toggle_fixed).grid(row=0, column=0, padx=1)

		editor = ttk.LabelFrame(bones_frame, text="選択中の骨", style="Bold.TLabelframe")
		editor.grid(row=0, column=1, sticky="nsew", padx=(4, 8), pady=6)
		for i in range(3):
			editor.columnconfigure(i, weight=[0, 1, 0][i])
		row = 0
		ttk.Label(editor, text="骨名:").grid(row=row, column=0, sticky="w", padx=(8, 4), pady=2)
		name_var = tk.StringVar(value="")
		ttk.Entry(editor, textvariable=name_var, width=18).grid(row=row, column=1, sticky="w", pady=2)
		ttk.Button(editor, text="適用", width=5,
		           command=lambda: self._ankle_apply_editor_field("name", name_var.get())
		           ).grid(row=row, column=2, sticky="w", padx=4)
		self._ankle_bone_editor_widgets["name_var"] = name_var

		row += 1
		ttk.Label(editor, text="ArUco ID:").grid(row=row, column=0, sticky="w", padx=(8, 4), pady=2)
		aruco_var = tk.IntVar(value=0)
		ttk.Spinbox(editor, from_=0, to=999, textvariable=aruco_var, width=6).grid(row=row, column=1, sticky="w", pady=2)
		ttk.Button(editor, text="適用", width=5,
		           command=lambda: self._ankle_apply_editor_field("aruco_id", int(aruco_var.get()))
		           ).grid(row=row, column=2, sticky="w", padx=4)
		self._ankle_bone_editor_widgets["aruco_var"] = aruco_var

		row += 1
		ttk.Label(editor, text="表示色:").grid(row=row, column=0, sticky="w", padx=(8, 4), pady=2)
		color_btn = tk.Button(editor, width=4, relief="ridge",
		                       command=self._ankle_choose_bone_color)
		color_btn.grid(row=row, column=1, sticky="w", pady=2)
		self._ankle_bone_editor_widgets["color_btn"] = color_btn

		row += 1
		ttk.Label(editor, text="骨モデル (可視化用フル):").grid(row=row, column=0, sticky="w", padx=(8, 4), pady=2)
		model_var = tk.StringVar(value="")
		ttk.Entry(editor, textvariable=model_var, state="readonly", width=40).grid(row=row, column=1, sticky="ew", pady=2)
		ttk.Button(editor, text="参照", width=5,
		           command=lambda: self._ankle_pick_bone_file("model_path", "骨モデル(可視化用フル)を選択", "model")
		           ).grid(row=row, column=2, sticky="w", padx=4)
		self._ankle_bone_editor_widgets["model_var"] = model_var

		# キャリブ用モデル (省略時は上のフルモデルを使う, ArUco部分だけ切り出したもの推奨)
		row += 1
		ttk.Label(editor, text="キャリブ用モデル (省略可):").grid(row=row, column=0, sticky="w", padx=(8, 4), pady=2)
		calib_model_var = tk.StringVar(value="")
		ttk.Entry(editor, textvariable=calib_model_var, state="readonly", width=40
		          ).grid(row=row, column=1, sticky="ew", pady=2)
		ttk.Button(editor, text="参照", width=5,
		           command=lambda: self._ankle_pick_bone_file("calib_model_path",
		                                                       "キャリブ用モデル(ArUco部分切り出し)を選択", "model")
		           ).grid(row=row, column=2, sticky="w", padx=4)
		self._ankle_bone_editor_widgets["calib_model_var"] = calib_model_var
		row += 1
		ttk.Label(editor,
		          text="※ キャリブ用モデルは、フルモデルからArUco貼付部分だけを切り出したもの (座標系はフルと同じであること)。\n"
		               "   省略時はフルモデルをそのままキャリブに使います。テクスチャ (OBJ+MTL / PLY埋込 / 同名png/jpg) が必要。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=row, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 4))

		# --- 位置合わせ関連 (原プランのみで使用、他モードではグレーアウト) ---
		row += 1
		post_row_frame = ttk.Frame(editor)
		post_row_frame.grid(row=row, column=0, columnspan=3, sticky="ew")
		for i in range(3):
			post_row_frame.columnconfigure(i, weight=[0, 1, 0][i])
		ttk.Label(post_row_frame, text="初期スキャン側 骨領域 (位置合わせ先):"
		          ).grid(row=0, column=0, sticky="w", padx=(8, 4), pady=2)
		post_var = tk.StringVar(value="")
		ttk.Entry(post_row_frame, textvariable=post_var, state="readonly", width=40
		          ).grid(row=0, column=1, sticky="ew", pady=2)
		ttk.Button(post_row_frame, text="参照", width=5,
		           command=lambda: self._ankle_pick_bone_file("post_scan_path", "初期スキャン側の骨領域を選択", "model")
		           ).grid(row=0, column=2, sticky="w", padx=4)
		self._ankle_bone_editor_widgets["post_var"] = post_var
		self._ankle_reg_widgets = [post_row_frame]

		# 位置合わせ方式・スケーリング(骨ごと)
		row += 1
		mf = ttk.Frame(editor)
		mf.grid(row=row, column=0, columnspan=3, sticky="w", padx=(8, 4), pady=(2, 2))
		ttk.Label(mf, text="位置合わせ方式:").grid(row=0, column=0, sticky="w")
		method_var = tk.StringVar(value="ransac")
		ttk.Radiobutton(mf, text="RANSAC", value="ransac", variable=method_var,
		                command=lambda: self._ankle_apply_editor_field("method", method_var.get())
		                ).grid(row=0, column=1, sticky="w", padx=(6, 6))
		ttk.Radiobutton(mf, text="主軸PCA", value="pca", variable=method_var,
		                command=lambda: self._ankle_apply_editor_field("method", method_var.get())
		                ).grid(row=0, column=2, sticky="w", padx=(0, 6))
		ttk.Radiobutton(mf, text="手動3点", value="manual", variable=method_var,
		                command=lambda: self._ankle_apply_editor_field("method", method_var.get())
		                ).grid(row=0, column=3, sticky="w", padx=(0, 12))
		scaling_var = tk.BooleanVar(value=False)
		ttk.Checkbutton(mf, text="スケール補正", variable=scaling_var,
		                command=lambda: self._ankle_apply_editor_field("enable_scaling", bool(scaling_var.get()))
		                ).grid(row=0, column=4, sticky="w")
		self._ankle_bone_editor_widgets["method_var"] = method_var
		self._ankle_bone_editor_widgets["scaling_var"] = scaling_var
		self._ankle_reg_widgets.append(mf)

		row += 1
		status_lbl = ttk.Label(editor, text="", foreground="gray", font=(self.ui_font_family, 8))
		status_lbl.grid(row=row, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 4))
		self._ankle_bone_editor_widgets["status_lbl"] = status_lbl

		row += 1
		opsf = ttk.Frame(editor)
		opsf.grid(row=row, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 6))
		ttk.Button(opsf, text="🔍 スキャンから自動キャリブ",
		           command=self.on_ankle_auto_calibrate_from_mesh).grid(row=0, column=0, padx=(0, 4))
		ttk.Button(opsf, text="マーカー-骨キャリブ (手動4点)",
		           command=self.on_ankle_calibrate_marker_to_bone).grid(row=0, column=1, padx=(0, 4))
		self._ankle_register_bone_btn = ttk.Button(opsf, text="初期スキャンへ位置合わせ",
		           command=self.on_ankle_register_bone)
		self._ankle_register_bone_btn.grid(row=0, column=2, padx=(0, 4))
		self._ankle_reg_widgets.append(self._ankle_register_bone_btn)
		row += 1
		opsf2 = ttk.Frame(editor)
		opsf2.grid(row=row, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 6))
		ttk.Button(opsf2, text="キャリブをクリア",
		           command=lambda: self._ankle_apply_editor_field("marker_to_bone_T", None)
		           ).grid(row=0, column=0, padx=(0, 4))
		self._ankle_clear_reg_btn = ttk.Button(opsf2, text="位置合わせをクリア",
		           command=lambda: self._ankle_apply_editor_field("reg_T", None))
		self._ankle_clear_reg_btn.grid(row=0, column=1, padx=(0, 4))
		self._ankle_reg_widgets.append(self._ankle_clear_reg_btn)

		# 位置合わせパラメータ (タブ共通、原プランのみで使用)
		param_frame = ttk.LabelFrame(bones_frame, text="位置合わせパラメータ (原プランのみ・全骨共有)",
		                              style="Bold.TLabelframe")
		param_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=(8, 8), pady=(0, 6))
		self._ankle_reg_param_frame = param_frame
		self._ankle_reg_widgets.append(param_frame)
		pf = ttk.Frame(param_frame)
		pf.grid(row=0, column=0, sticky="w", padx=8, pady=(4, 2))
		ank_reg_spec = [
			("RANSAC距離(mm):", self.ankle_reg_ransac_distance),
			("RANSAC反復:", self.ankle_reg_ransac_max_iter),
			("RANSAC信頼度:", self.ankle_reg_ransac_confidence),
			("ICP距離(mm):", self.ankle_reg_icp_threshold),
			("ICP反復:", self.ankle_reg_icp_max_iter),
			("サンプル点数:", self.ankle_reg_sample_points),
			("ボクセル(mm,0=自動):", self.ankle_reg_voxel_size),
		]
		col = 0; r = 0
		for lbl, var in ank_reg_spec:
			ttk.Label(pf, text=lbl).grid(row=r, column=col * 2, sticky="w", padx=(0, 2), pady=1)
			ttk.Entry(pf, textvariable=var, width=8).grid(row=r, column=col * 2 + 1, sticky="w", padx=(0, 10), pady=1)
			col += 1
			if col >= 3:
				col = 0; r += 1
		ttk.Checkbutton(param_frame, text="経過を表示 (点群重ね合わせプレビュー)",
		                variable=self.ankle_reg_preview_var
		                ).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 4))
		bulkf = ttk.Frame(param_frame)
		bulkf.grid(row=2, column=0, sticky="w", padx=12, pady=(2, 6))
		ttk.Button(bulkf, text="全骨を一括位置合わせ",
		           command=self.on_ankle_register_all_bones).grid(row=0, column=0, padx=(0, 8))
		ttk.Button(bulkf, text="位置合わせ結果を確認 (プレビュー)",
		           command=self.on_ankle_preview_registration).grid(row=0, column=1, padx=(0, 8))

		# 🔨 マーカー準備 (印刷) — 事前準備用: ③のID/名前と②の辞書/実寸から印刷ファイル生成
		prep_frame = ttk.LabelFrame(container, text="🔨 マーカー準備 (印刷用ファイル生成)",
		                             style="Bold.TLabelframe")
		prep_frame.grid(row=8, column=0, sticky="nsew", pady=(0, 8))
		prep_frame.columnconfigure(0, weight=1)
		prep_btn = ttk.Frame(prep_frame)
		prep_btn.grid(row=0, column=0, sticky="w", padx=12, pady=(6, 2))
		ttk.Button(prep_btn, text="マーカー印刷用PDF (実寸・推奨)",
		           command=self.on_ankle_save_marker_pdf).grid(row=0, column=0, padx=(0, 8))
		ttk.Button(prep_btn, text="マーカー画像を保存 (PNG)",
		           command=self.on_ankle_save_marker_images).grid(row=0, column=1, padx=(0, 8))
		ttk.Label(prep_frame,
		          text="事前準備用: ③のArUco IDと②の辞書・マーカー実寸から印刷ファイルを生成します。\n"
		               "PDFは「実際のサイズ / 100%」で印刷 → ノギスで実寸確認 → ②に反映。\n"
		               "印刷後、骨にマーカーを装着 → ①で初期スキャン → ⓪で撮影、の順で本試験に進みます。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=1, column=0, sticky="w", padx=12, pady=(2, 6))

		# ④ ArUco/PnP 実行
		det_frame = ttk.LabelFrame(container, text="④ ArUco検出 / PnPで6DOF姿勢時系列を計算", style="Bold.TLabelframe")
		det_frame.grid(row=9, column=0, sticky="nsew", pady=(0, 8))
		det_frame.columnconfigure(0, weight=1)
		# 検出パラメータ
		pf = ttk.Frame(det_frame)
		pf.grid(row=0, column=0, sticky="w", padx=12, pady=(6, 2))
		ttk.Label(pf, text="フレーム間引き:").grid(row=0, column=0, sticky="w")
		ttk.Spinbox(pf, from_=1, to=100, textvariable=self.ankle_detect_stride, width=6
		            ).grid(row=0, column=1, sticky="w", padx=(4, 12))
		ttk.Label(pf, text="参照フレーム t=0:").grid(row=0, column=2, sticky="w")
		ttk.Spinbox(pf, from_=0, to=100000, textvariable=self.ankle_ref_frame, width=8
		            ).grid(row=0, column=3, sticky="w", padx=(4, 0))
		ttk.Label(pf, text="(Stage 5 でアニメ基準に使用)",
		          foreground="gray", font=(self.ui_font_family, 8)
		          ).grid(row=0, column=4, sticky="w", padx=(6, 0))
		# 実行ボタン
		dbtn = ttk.Frame(det_frame)
		dbtn.grid(row=1, column=0, sticky="w", padx=12, pady=(4, 4))
		ttk.Button(dbtn, text="ArUco検出+PnP実行", command=self.on_ankle_detect_markers).grid(row=0, column=0, padx=(0, 8))
		ttk.Button(dbtn, text="姿勢時系列を保存", command=self.on_ankle_save_pose_series).grid(row=0, column=1, padx=(0, 8))
		ttk.Button(dbtn, text="姿勢時系列を読込", command=self.on_ankle_load_pose_series).grid(row=0, column=2, padx=(0, 8))
		ttk.Button(dbtn, text="マーカー軌跡を可視化 (骨モデル不要)",
		           command=self.on_ankle_visualize_pose_series).grid(row=0, column=3, padx=(16, 8))
		# 検出結果ステータス
		ttk.Label(det_frame, textvariable=self.ankle_detection_status,
		          foreground="#005580", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=2, column=0, sticky="w", padx=12, pady=(2, 4))
		ttk.Label(det_frame,
		          text="対応データ形式: RealSense .bag (自動で色/深度/内部パラメータを抽出) / "
		               "汎用 mp4+npz深度+jsonパラメータ。SOLVEPNP_IPPE_SQUAREで両姿勢解を計算し、"
		               "深度で表裏を選択。要 opencv-contrib-python (+.bag使用時 pyrealsense2)。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=3, column=0, sticky="w", padx=12, pady=(0, 6))

		# ⑤ 可視化・アニメ
		vis_frame = ttk.LabelFrame(container, text="⑤ 可視化・アニメーション (N骨マルチヒートマップ)", style="Bold.TLabelframe")
		vis_frame.grid(row=10, column=0, sticky="nsew", pady=(0, 8))
		vis_frame.columnconfigure(0, weight=1)
		vbtn = ttk.Frame(vis_frame)
		vbtn.grid(row=0, column=0, sticky="w", padx=12, pady=(6, 4))
		ttk.Button(vbtn, text="初期状態を可視化", command=self.on_ankle_visualize_initial).grid(row=0, column=0, padx=(0, 8))
		ttk.Button(vbtn, text="全骨を可視化", command=self.on_ankle_visualize_all).grid(row=0, column=1, padx=(0, 8))
		ttk.Button(vbtn, text="シミュレーション実行", command=self.on_ankle_animate).grid(row=0, column=2, padx=(0, 8))
		ttk.Label(vis_frame,
		          text="※ ③骨リストに登録された全骨が自動的に可視化されます。各骨のヒートマップは"
		               " 他全骨との signed distance の最小値 (接触=緑・めり込み=赤・離間=骨本来の色)。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 6))
		# 可視化オプション行
		opt = ttk.Frame(vis_frame)
		opt.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 6))
		ttk.Checkbutton(opt, text="時系列平滑化",
		                variable=self.ankle_smooth_enable
		                ).grid(row=0, column=0, sticky="w")
		ttk.Label(opt, text="カットオフ周波数:").grid(row=0, column=1, sticky="w", padx=(12, 4))
		ttk.Spinbox(opt, from_=0.2, to=15.0, increment=0.1,
		            textvariable=self.ankle_smooth_cutoff_hz,
		            width=6).grid(row=0, column=2, sticky="w")
		ttk.Label(opt, text="Hz").grid(row=0, column=3, sticky="w", padx=(2, 12))
		ttk.Checkbutton(opt, text="外れ値除去 (Hampel)",
		                variable=self.ankle_reject_outliers
		                ).grid(row=0, column=4, sticky="w", padx=(4, 12))
		ttk.Checkbutton(opt, text="ArUcoマーカー軸を表示",
		                variable=self.ankle_show_markers
		                ).grid(row=0, column=5, sticky="w", padx=(4, 0))
		ttk.Label(opt,
		          text="※ 小さいほど滑らか (動きは鈍る)。15fps 録画なら 1.5〜3 Hz が目安。"
		               "実測フレームレートはログに出ます。",
		          foreground="gray", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=1, column=0, columnspan=6, sticky="w", padx=(20, 0), pady=(2, 0))
		opt2 = ttk.Frame(vis_frame)
		opt2.grid(row=3, column=0, sticky="w", padx=12, pady=(0, 6))
		ttk.Checkbutton(opt2, text="🔧 診断モード: 骨重心をマーカー位置に強制配置 (通常はOFF)",
		                variable=self.ankle_recenter_meshes
		                ).grid(row=0, column=0, sticky="w")
		ttk.Label(opt2,
		          text="※ 骨が変な位置に出るときの原因切り分け用。キャリブ (T_L←Mk) を無視し、"
		               "骨の重心をArUcoが検出したマーカー位置にそのまま置きます。\n"
		               "   ONにして骨が正常な位置に来る → キャリブが原因。変なまま → ArUco検出が原因。\n"
		               "   骨の向きは意味を持たなくなるため、定量解析では必ずOFFにしてください。",
		          foreground="#8B0000", font=(self.ui_font_family, 8), wraplength=760, justify="left"
		          ).grid(row=1, column=0, sticky="w", padx=(20, 0))

		self._ankle_init_tabs()
		# モード変化に応じて①の enable/disable を切替
		self.ankle_workflow_mode.trace_add("write", self._ankle_on_mode_change)
		self._ankle_on_mode_change()  # 初期状態反映

	def _ankle_on_mode_change(self, *args) -> None:
		"""動作モード変化時、モード別に不要ウィジェットを enable/disable する。

		- 原プラン (original):
		    ①初期状態スキャン    → 有効
		    ③位置合わせ関連     → 有効
		    ⚡ 関節座標系Cj      → 無効
		- 新プラン (self_pose):
		    ①                    → 無効 (不要)
		    ③位置合わせ関連     → 無効 (ArUcoで自己位置決定)
		    ⚡ Cj                 → 無効
		- 簡易版 (simple):
		    ①                    → 無効
		    ③位置合わせ関連     → 無効 (骨モデル自体不要だが編集は許可)
		    ⚡ Cj                 → 有効
		"""
		try:
			mode = str(self.ankle_workflow_mode.get())
		except Exception:
			mode = "self_pose"

		def _set_state(w, state):
			try: w.configure(state=state)
			except Exception: pass
			for c in w.winfo_children():
				_set_state(c, state)

		# ①初期状態スキャン (原プランのみ)
		init_enabled = (mode == "original")
		init_frame = getattr(self, "_ankle_init_frame", None)
		if init_frame is not None:
			_set_state(init_frame, "normal" if init_enabled else "disabled")
			try:
				if init_enabled:
					init_frame.configure(text="① 初期状態スキャン（ピン+マーカー装着後）— 原プランのみ")
				else:
					init_frame.configure(text="① 初期状態スキャン (このモードでは不要 — スキップ)")
			except Exception: pass

		# ③位置合わせ関連 (原プランのみ) — post_scan/method/register/param
		reg_enabled = (mode == "original")
		for w in getattr(self, "_ankle_reg_widgets", []):
			_set_state(w, "normal" if reg_enabled else "disabled")
		# パラメータフレームのタイトルも更新
		rpf = getattr(self, "_ankle_reg_param_frame", None)
		if rpf is not None:
			try:
				if reg_enabled:
					rpf.configure(text="位置合わせパラメータ (原プランのみ・全骨共有)")
				else:
					rpf.configure(text="位置合わせパラメータ (このモードでは不要 — スキップ)")
			except Exception: pass

		# ⚡ 関節座標系Cj (簡易版のみ)
		cj_enabled = (mode == "simple")
		cj_frame = getattr(self, "_ankle_cj_frame", None)
		if cj_frame is not None:
			_set_state(cj_frame, "normal" if cj_enabled else "disabled")
			try:
				if cj_enabled:
					cj_frame.configure(text="⚡ 関節座標系校正 (Cj) — 簡易版モードで使用")
				else:
					cj_frame.configure(text="⚡ 関節座標系校正 (Cj) — このモードでは不要 — スキップ")
			except Exception: pass

	# ---- ankle: ファイル/カラーピッカー ----
	def _ankle_choose(self, var: tk.StringVar, title: str, kind: str) -> None:
		if kind == "model":
			ft = [("3Dモデル", "*.obj *.stl *.ply"), ("すべてのファイル", "*.*")]
		elif kind == "video":
			ft = [("動画", "*.mp4 *.avi *.mov *.mkv"), ("すべてのファイル", "*.*")]
		elif kind == "depth":
			ft = [("深度データ", "*.db3 *.bag *.npz *.mkv *.tiff *.tif"), ("すべてのファイル", "*.*")]
		elif kind == "intrinsics":
			ft = [("カメラパラメータ", "*.json *.yaml *.yml"), ("すべてのファイル", "*.*")]
		elif kind == "pose":
			ft = [("姿勢時系列", "*.npz *.csv *.json"), ("すべてのファイル", "*.*")]
		else:
			ft = [("すべてのファイル", "*.*")]
		path = filedialog.askopenfilename(title=title, filetypes=ft)
		if path:
			var.set(path)

	def _ankle_color_of(self, i: int) -> str:
		if 0 <= i < len(self.ankle_bones):
			c = str(self.ankle_bones[i].get("color", "") or "").strip()
			if len(c) == 7 and c[0] == "#" and all(ch in "0123456789abcdefABCDEF" for ch in c[1:]):
				return c
		return self._ANKLE_DEFAULT_BONE_COLORS[max(0, i) % len(self._ANKLE_DEFAULT_BONE_COLORS)]

	# ---- ankle: 骨リスト管理 ----
	def _ankle_current_bone(self):
		i = self._ankle_selected_bone
		if 0 <= i < len(self.ankle_bones):
			return self.ankle_bones[i]
		return None

	def _ankle_refresh_bone_listbox(self) -> None:
		lb = self._ankle_bones_listbox
		if lb is None:
			return
		lb.delete(0, tk.END)
		for i, b in enumerate(self.ankle_bones):
			name = b.get("name", f"骨{i+1}")
			aid = b.get("aruco_id", "?")
			mk = "M" if b.get("marker_to_bone_T") is not None else "-"
			rg = "R" if b.get("reg_T") is not None else "-"
			lock = "🔒 " if b.get("fixed") else "   "
			lb.insert(tk.END, f"{lock}{i+1}. {name} [ID={aid}] {mk}{rg}")
		if self.ankle_bones:
			idx = max(0, min(self._ankle_selected_bone, len(self.ankle_bones) - 1))
			self._ankle_selected_bone = idx
			try:
				lb.selection_clear(0, tk.END)
				lb.selection_set(idx)
				lb.see(idx)
			except Exception:
				pass
		self._ankle_load_editor_from_bone()
		self._ankle_refresh_heatmap_combos()

	def _ankle_refresh_heatmap_combos(self) -> None:
		labels = [f"{i+1}. {b.get('name','')}" for i, b in enumerate(self.ankle_bones)]
		for combo in (getattr(self, "_ankle_heatmap_prox_combo", None),
		              getattr(self, "_ankle_heatmap_dist_combo", None)):
			if combo is not None:
				try:
					combo["values"] = ["(未選択)"] + labels
				except Exception:
					pass

	def _ankle_on_bone_select(self) -> None:
		lb = self._ankle_bones_listbox
		if lb is None:
			return
		sel = lb.curselection()
		if not sel:
			return
		self._ankle_selected_bone = int(sel[0])
		self._ankle_load_editor_from_bone()

	def _ankle_load_editor_from_bone(self) -> None:
		b = self._ankle_current_bone()
		w = self._ankle_bone_editor_widgets
		if not w:
			return
		if b is None:
			for key in ("name_var", "model_var", "post_var"):
				if key in w:
					try:
						w[key].set("")
					except Exception:
						pass
			if "aruco_var" in w:
				try:
					w["aruco_var"].set(0)
				except Exception:
					pass
			if "status_lbl" in w:
				w["status_lbl"].configure(text="(骨が未登録です)")
			if "color_btn" in w:
				try:
					w["color_btn"].configure(bg="#cccccc", activebackground="#cccccc")
				except Exception:
					pass
			return
		try:
			w["name_var"].set(b.get("name", ""))
			w["aruco_var"].set(int(b.get("aruco_id", 0)))
			w["model_var"].set(b.get("model_path", "") or "")
			if "calib_model_var" in w:
				w["calib_model_var"].set(b.get("calib_model_path", "") or "")
			w["post_var"].set(b.get("post_scan_path", "") or "")
			if "method_var" in w:
				w["method_var"].set(str(b.get("method", "ransac")))
			if "scaling_var" in w:
				w["scaling_var"].set(bool(b.get("enable_scaling", False)))
			c = self._ankle_color_of(self._ankle_selected_bone)
			w["color_btn"].configure(bg=c, activebackground=c)
			mk_ok = b.get("marker_to_bone_T") is not None
			reg_ok = b.get("reg_T") is not None
			w["status_lbl"].configure(
				text=f"マーカー-骨キャリブ: {'済' if mk_ok else '未'} / 位置合わせ: {'済' if reg_ok else '未'}")
		except Exception:
			pass

	def _ankle_apply_editor_field(self, key: str, value) -> None:
		b = self._ankle_current_bone()
		if b is None:
			return
		b[key] = value
		self._ankle_refresh_bone_listbox()

	def _ankle_pick_bone_file(self, key: str, title: str, kind: str) -> None:
		b = self._ankle_current_bone()
		if b is None:
			messagebox.showinfo("骨選択", "先に骨を選択してください。")
			return
		if kind == "model":
			ft = [("3Dモデル", "*.obj *.stl *.ply"), ("すべてのファイル", "*.*")]
		else:
			ft = [("すべてのファイル", "*.*")]
		path = filedialog.askopenfilename(title=title, filetypes=ft)
		if path:
			b[key] = path
			self._ankle_load_editor_from_bone()

	def _ankle_choose_bone_color(self) -> None:
		b = self._ankle_current_bone()
		if b is None:
			return
		i = self._ankle_selected_bone
		color = colorchooser.askcolor(initialcolor=self._ankle_color_of(i), title=f"{b.get('name','')}の表示色")
		if color and color[1]:
			b["color"] = color[1]
			self._ankle_load_editor_from_bone()
			self._ankle_refresh_bone_listbox()

	def on_ankle_bone_add(self) -> None:
		new_bone = self._ankle_default_bone(len(self.ankle_bones))
		existing = {b.get("name", "") for b in self.ankle_bones}
		n = len(self.ankle_bones) + 1
		while new_bone["name"] in existing:
			n += 1
			new_bone["name"] = f"骨{n}"
		existing_ids = {int(b.get("aruco_id", -1)) for b in self.ankle_bones}
		aid = new_bone["aruco_id"]
		while aid in existing_ids:
			aid += 1
		new_bone["aruco_id"] = aid
		self.ankle_bones.append(new_bone)
		self._ankle_selected_bone = len(self.ankle_bones) - 1
		self._ankle_refresh_bone_listbox()

	def on_ankle_bone_duplicate(self) -> None:
		b = self._ankle_current_bone()
		if b is None:
			messagebox.showinfo("骨複製", "先に骨を選択してください。")
			return
		nb = copy.deepcopy(b)
		existing = {x.get("name", "") for x in self.ankle_bones}
		base = nb.get("name", "骨")
		k = 2
		nb["name"] = f"{base}({k})"
		while nb["name"] in existing:
			k += 1
			nb["name"] = f"{base}({k})"
		existing_ids = {int(x.get("aruco_id", -1)) for x in self.ankle_bones}
		aid = int(nb.get("aruco_id", 0))
		while aid in existing_ids:
			aid += 1
		nb["aruco_id"] = aid
		self.ankle_bones.insert(self._ankle_selected_bone + 1, nb)
		self._ankle_selected_bone += 1
		self._ankle_refresh_bone_listbox()

	def on_ankle_bone_delete(self) -> None:
		b = self._ankle_current_bone()
		if b is None:
			return
		if not messagebox.askyesno("骨削除", f"骨「{b.get('name','')}」を削除しますか？"):
			return
		del self.ankle_bones[self._ankle_selected_bone]
		if self._ankle_selected_bone >= len(self.ankle_bones):
			self._ankle_selected_bone = max(0, len(self.ankle_bones) - 1)
		self._ankle_refresh_bone_listbox()

	def on_ankle_bone_rename(self) -> None:
		b = self._ankle_current_bone()
		if b is None:
			return
		cur = b.get("name", "")
		new = simpledialog.askstring("骨名変更", "新しい骨名:", initialvalue=cur, parent=self)
		if not new:
			return
		new = new.strip()
		if not new:
			return
		existing = {x.get("name", "") for i, x in enumerate(self.ankle_bones) if i != self._ankle_selected_bone}
		if new in existing:
			messagebox.showwarning("骨名変更", f"「{new}」は既に使われています。")
			return
		b["name"] = new
		self._ankle_refresh_bone_listbox()

	def on_ankle_bone_move(self, delta: int) -> None:
		i = self._ankle_selected_bone
		j = i + int(delta)
		if not (0 <= i < len(self.ankle_bones) and 0 <= j < len(self.ankle_bones)) or i == j:
			return
		self.ankle_bones[i], self.ankle_bones[j] = self.ankle_bones[j], self.ankle_bones[i]
		self._ankle_selected_bone = j
		self._ankle_refresh_bone_listbox()

	def on_ankle_bone_toggle_fixed(self) -> None:
		"""選択中の骨を「固定」対象にトグルする。固定は1骨のみで排他 (他骨の fixed は自動解除)。

		固定された骨は on_ankle_animate で常に identity 姿勢になり、他の骨は固定骨の座標系
		での相対姿勢として表示される (= 固定骨から見た他骨の運動を観察できる)。
		"""
		i = self._ankle_selected_bone
		if not (0 <= i < len(self.ankle_bones)):
			messagebox.showwarning("骨固定", "骨リストで固定/解除する骨を選択してください。")
			return
		current = bool(self.ankle_bones[i].get("fixed", False))
		# 他骨の fixed を全て解除 (排他)
		for j, b in enumerate(self.ankle_bones):
			if j != i:
				b["fixed"] = False
		# トグル
		self.ankle_bones[i]["fixed"] = not current
		self._ankle_refresh_bone_listbox()
		# 選択状態を維持
		try:
			lb = self._ankle_bones_listbox
			lb.selection_clear(0, tk.END)
			lb.selection_set(i)
			lb.see(i)
		except Exception:
			pass
		name = self.ankle_bones[i].get("name", f"骨{i+1}")
		if self.ankle_bones[i]["fixed"]:
			print(f"[ankle] {name} を固定に設定 (他骨は解除)")
		else:
			print(f"[ankle] {name} の固定を解除")

	# ---- ankle: 多試験タブ ----
	def _ankle_state_vars(self) -> dict:
		return {
			"ankle_initial_scan": (self.ankle_initial_scan_path, str),
			"ankle_video": (self.ankle_video_path, str),
			"ankle_depth": (self.ankle_depth_path, str),
			"ankle_camera_intrinsics": (self.ankle_camera_intrinsics_path, str),
			"ankle_aruco_dict": (self.ankle_aruco_dict_var, str),
			"ankle_marker_size_mm": (self.ankle_marker_size_mm, float),
			"ankle_pose_series": (self.ankle_pose_series_path, str),
			"ankle_heatmap_prox": (self.ankle_heatmap_prox_var, str),
			"ankle_heatmap_dist": (self.ankle_heatmap_dist_var, str),
			# 位置合わせパラメータ (タブ共通)
			"ankle_reg_ransac_distance": (self.ankle_reg_ransac_distance, float),
			"ankle_reg_ransac_max_iter": (self.ankle_reg_ransac_max_iter, int),
			"ankle_reg_ransac_confidence": (self.ankle_reg_ransac_confidence, float),
			"ankle_reg_icp_threshold": (self.ankle_reg_icp_threshold, float),
			"ankle_reg_icp_max_iter": (self.ankle_reg_icp_max_iter, int),
			"ankle_reg_sample_points": (self.ankle_reg_sample_points, int),
			"ankle_reg_voxel_size": (self.ankle_reg_voxel_size, float),
			"ankle_reg_preview_var": (self.ankle_reg_preview_var, bool),
			# ArUco/PnP検出パラメータ
			"ankle_ref_frame": (self.ankle_ref_frame, int),
			"ankle_detect_stride": (self.ankle_detect_stride, int),
			# D405 ライブ撮影
			"ankle_rs_resolution": (self.ankle_rs_resolution, str),
			"ankle_rs_discard_frames": (self.ankle_rs_discard_frames, int),
			"ankle_rs_manual_exposure": (self.ankle_rs_manual_exposure, bool),
			"ankle_rs_exposure_val": (self.ankle_rs_exposure_val, int),
			"ankle_rs_color_manual_exposure": (self.ankle_rs_color_manual_exposure, bool),
			"ankle_rs_color_exposure_us": (self.ankle_rs_color_exposure_us, int),
			"ankle_rs_color_gain": (self.ankle_rs_color_gain, int),
			# 動作モード (原プラン/新プラン)
			"ankle_workflow_mode": (self.ankle_workflow_mode, str),
			"ankle_self_pose_world": (self.ankle_self_pose_world, str),
		}

	def _ankle_snapshot_current(self) -> dict:
		snap = {}
		for key, (var, _t) in self._ankle_state_vars().items():
			try:
				snap[key] = var.get()
			except Exception:
				pass
		snap["_bones"] = copy.deepcopy(self.ankle_bones)
		snap["_selected_bone"] = int(self._ankle_selected_bone)
		return snap

	def _ankle_restore_snapshot(self, snap: dict) -> None:
		if not isinstance(snap, dict):
			return
		default = self._ankle_default_snap if isinstance(getattr(self, "_ankle_default_snap", None), dict) else {}
		for key, (var, typ) in self._ankle_state_vars().items():
			if key in snap:
				v = snap[key]
			elif key in default:
				v = default[key]
			else:
				continue
			try:
				if typ is bool:
					var.set(bool(v))
				elif typ is int:
					var.set(int(v))
				elif typ is float:
					var.set(float(v))
				else:
					var.set(str(v))
			except Exception:
				pass
		bones = snap.get("_bones", None)
		if isinstance(bones, list):
			self.ankle_bones = copy.deepcopy(bones)
		else:
			self.ankle_bones = []
		try:
			self._ankle_selected_bone = int(snap.get("_selected_bone", 0))
		except Exception:
			self._ankle_selected_bone = 0
		if self.ankle_bones and self._ankle_selected_bone >= len(self.ankle_bones):
			self._ankle_selected_bone = 0
		self._ankle_refresh_bone_listbox()

	def _ankle_rebuild_tabbar(self) -> None:
		fr = self._ankle_tabbar_frame
		if fr is None:
			return
		for w in fr.winfo_children():
			w.destroy()
		self._ankle_tab_buttons = []
		for i, tab in enumerate(self._ankle_tabs):
			active = (i == self._ankle_active_tab)
			b = tk.Button(
				fr, text=tab.get('name', f"試験{i+1}"),
				relief=('sunken' if active else 'raised'),
				bg=('#cfe3ff' if active else '#f0f0f0'),
				font=(self.ui_font_family, 9, 'bold' if active else 'normal'),
				command=lambda i=i: self.on_ankle_tab_select(i), padx=8, pady=2)
			b.pack(side='left', padx=2)
			b.bind("<Button-3>", lambda e, i=i: self._ankle_tab_context_menu(e, i))
			b.bind("<ButtonPress-1>", lambda e, i=i: self._ankle_tab_drag_start(e, i))
			b.bind("<B1-Motion>", self._ankle_tab_drag_motion)
			b.bind("<ButtonRelease-1>", self._ankle_tab_drag_release)
			self._ankle_tab_buttons.append(b)
		plus = tk.Button(fr, text="＋", command=self.on_ankle_tab_add, padx=6, pady=2)
		plus.pack(side='left', padx=(8, 2))

	def _ankle_tab_context_menu(self, event, i: int) -> None:
		menu = tk.Menu(self, tearoff=0)
		menu.add_command(label="名前変更", command=lambda: self.on_ankle_tab_rename(i))
		menu.add_command(label="削除", command=lambda: self.on_ankle_tab_delete(i))
		menu.add_separator()
		menu.add_command(label="← 左へ移動", command=lambda: self._ankle_tab_move(i, i - 1),
		                 state=("normal" if i > 0 else "disabled"))
		menu.add_command(label="→ 右へ移動", command=lambda: self._ankle_tab_move(i, i + 1),
		                 state=("normal" if i < len(self._ankle_tabs) - 1 else "disabled"))
		try:
			menu.tk_popup(event.x_root, event.y_root)
		finally:
			menu.grab_release()

	def _ankle_tab_move(self, i: int, j: int) -> None:
		n = len(self._ankle_tabs)
		if not (0 <= i < n and 0 <= j < n) or i == j:
			return
		tab = self._ankle_tabs.pop(i)
		self._ankle_tabs.insert(j, tab)
		a = self._ankle_active_tab
		if a == i:
			a = j
		else:
			if a > i:
				a -= 1
			if a >= j:
				a += 1
		self._ankle_active_tab = a
		self._ankle_rebuild_tabbar()

	def _ankle_tab_drag_start(self, event, i: int) -> None:
		self._ankle_tab_drag = {"from": i, "x": event.x_root, "moved": False}

	def _ankle_tab_drag_motion(self, event) -> None:
		d = self._ankle_tab_drag
		if d is None:
			return
		if not d["moved"] and abs(event.x_root - d["x"]) > 12:
			d["moved"] = True
			try:
				event.widget.configure(cursor="sb_h_double_arrow")
			except Exception:
				pass

	def _ankle_tab_drag_release(self, event):
		d = self._ankle_tab_drag
		self._ankle_tab_drag = None
		if not d or not d["moved"]:
			return None
		try:
			event.widget.configure(cursor="")
		except Exception:
			pass
		j = self._ankle_tab_index_at(event.x_root)
		if j is None or j == d["from"]:
			return None
		self._ankle_tab_move(d["from"], j)
		return "break"

	def _ankle_tab_index_at(self, x_root: int):
		btns = self._ankle_tab_buttons
		if not btns:
			return None
		best = 0
		for idx, b in enumerate(btns):
			try:
				bx = b.winfo_rootx()
				bw = b.winfo_width()
			except Exception:
				continue
			if x_root >= bx:
				best = idx
			if bx <= x_root < bx + bw:
				return idx
		return best

	def on_ankle_tab_select(self, i: int) -> None:
		if i < 0 or i >= len(self._ankle_tabs) or i == self._ankle_active_tab:
			return
		self._ankle_tabs[self._ankle_active_tab]['snapshot'] = self._ankle_snapshot_current()
		self._ankle_active_tab = i
		self._ankle_restore_snapshot(self._ankle_tabs[i]['snapshot'])
		self._ankle_rebuild_tabbar()
		self._ankle_update_detection_status()

	# ④(計測データ)のキー: 新規タブは末尾タブから複製、これらだけリセット
	_ANKLE_TAB_NO_COPY_KEYS = ("ankle_video", "ankle_depth", "ankle_pose_series")

	def on_ankle_tab_add(self) -> None:
		if self._ankle_tabs:
			self._ankle_tabs[self._ankle_active_tab]['snapshot'] = self._ankle_snapshot_current()
			snap = copy.deepcopy(self._ankle_tabs[-1].get('snapshot') or {})
			for key in self._ANKLE_TAB_NO_COPY_KEYS:
				if isinstance(self._ankle_default_snap, dict) and key in self._ankle_default_snap:
					snap[key] = self._ankle_default_snap[key]
				else:
					snap.pop(key, None)
		else:
			snap = copy.deepcopy(self._ankle_default_snap or {})
		self._ankle_tabs.append({'name': self._ankle_unique_tab_name(), 'snapshot': snap})
		self._ankle_active_tab = len(self._ankle_tabs) - 1
		self._ankle_restore_snapshot(self._ankle_tabs[self._ankle_active_tab]['snapshot'])
		self._ankle_rebuild_tabbar()

	def _ankle_unique_tab_name(self, base: str = "試験") -> str:
		names = {t.get('name', '') for t in self._ankle_tabs}
		n = len(self._ankle_tabs) + 1
		while f"{base}{n}" in names:
			n += 1
		return f"{base}{n}"

	def on_ankle_tab_delete(self, i: int) -> None:
		if len(self._ankle_tabs) <= 1:
			messagebox.showinfo("タブ削除", "最後のタブは削除できません。")
			return
		if not messagebox.askyesno("タブ削除", f"タブ「{self._ankle_tabs[i].get('name','')}」を削除しますか？"):
			return
		self._ankle_tabs[self._ankle_active_tab]['snapshot'] = self._ankle_snapshot_current()
		# 削除されるタブの姿勢キャッシュも破棄
		old_name = self._ankle_tabs[i].get('name', '')
		if old_name:
			self._ankle_pose_cache.pop(old_name, None)
		del self._ankle_tabs[i]
		if self._ankle_active_tab == i:
			self._ankle_active_tab = min(i, len(self._ankle_tabs) - 1)
		elif self._ankle_active_tab > i:
			self._ankle_active_tab -= 1
		self._ankle_restore_snapshot(self._ankle_tabs[self._ankle_active_tab]['snapshot'])
		self._ankle_rebuild_tabbar()
		self._ankle_update_detection_status()

	def on_ankle_tab_rename(self, i: int) -> None:
		cur = self._ankle_tabs[i].get('name', '')
		new = simpledialog.askstring("タブ名変更", "タブ名:", initialvalue=cur, parent=self)
		if new and new.strip():
			new = new.strip()
			others = {t.get('name', '') for k, t in enumerate(self._ankle_tabs) if k != i}
			if new in others:
				messagebox.showwarning("タブ名変更", f"「{new}」は既に存在します。")
				return
			# 姿勢キャッシュのキーを新名に付け替え
			if cur in self._ankle_pose_cache:
				self._ankle_pose_cache[new] = self._ankle_pose_cache.pop(cur)
			self._ankle_tabs[i]['name'] = new
			self._ankle_rebuild_tabbar()

	def _ankle_init_tabs(self) -> None:
		if not self._ankle_tabs:
			self._ankle_tabs = [{'name': '試験1', 'snapshot': self._ankle_snapshot_current()}]
			self._ankle_active_tab = 0
		if self._ankle_active_tab >= len(self._ankle_tabs):
			self._ankle_active_tab = 0
		self._ankle_restore_snapshot(self._ankle_tabs[self._ankle_active_tab]['snapshot'])
		self._ankle_rebuild_tabbar()

	# ---- ankle: 状態永続化 ----
	def _ankle_state_file_path(self) -> Path:
		import platform
		filename = "frs2015_gui_state_ankle_sim.json"
		if platform.system() == "Darwin":
			state_dir = Path.home() / ".frs_simulator"
			state_dir.mkdir(parents=True, exist_ok=True)
			return state_dir / filename
		return Path(__file__).with_name(filename)

	def _save_ankle_state(self) -> None:
		# 姿勢時系列も一緒に保存しておく (④の結果を次回起動へ引き継ぐ)
		try:
			self._ankle_autosave_all_pose_caches()
		except Exception:
			pass
		try:
			if getattr(self, "_ankle_tabs", None):
				self._ankle_tabs[self._ankle_active_tab]['snapshot'] = self._ankle_snapshot_current()
		except Exception:
			pass
		data = {
			"tabs": getattr(self, "_ankle_tabs", []),
			"active": getattr(self, "_ankle_active_tab", 0),
		}
		try:
			p = self._ankle_state_file_path()
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
			print(f"[ankle状態保存] {p}（{len(data['tabs'])}タブ）")
		except Exception as e:
			print(f"[ankle状態保存] 失敗: {e}")

	def _load_ankle_state(self) -> None:
		try:
			p = self._ankle_state_file_path()
			if not p.exists():
				return
			data = json.load(p.open("r", encoding="utf-8"))
		except Exception as e:
			print(f"[ankle状態復元] 失敗: {e}")
			return
		if isinstance(data, dict) and isinstance(data.get("tabs"), list) and data["tabs"]:
			self._ankle_tabs = data["tabs"]
			try:
				self._ankle_active_tab = int(data.get("active", 0))
			except Exception:
				self._ankle_active_tab = 0
			if self._ankle_active_tab < 0 or self._ankle_active_tab >= len(self._ankle_tabs):
				self._ankle_active_tab = 0
			self._ankle_restore_snapshot(self._ankle_tabs[self._ankle_active_tab].get("snapshot", {}))
		# 姿勢時系列 (④の結果) も自動復元する
		self._ankle_autoload_pose_caches()

	# ---- ankle: Stage 2〜5 プレースホルダ ----
	def _ankle_not_impl(self, feature: str) -> None:
		messagebox.showinfo(
			"未実装",
			f"「{feature}」は現在Stage 1(UI骨組み)のみ実装済みです。\n"
			f"次のステージで実装予定:\n"
			f"  Stage 2 — 骨モデル読込・位置合わせ・静的可視化\n"
			f"  Stage 3 — ArUco検出+PnPパイプライン\n"
			f"  Stage 4 — マーカー-骨キャリブレーション\n"
			f"  Stage 5 — アニメーション・骨対ヒートマップ")

	# ---- Stage 4: マーカー-骨キャリブレーション ----
	def _ankle_visualize_calibration(self, mesh, picked, T, marker_size_mm: float, title: str,
	                                   texture=None) -> None:
		"""キャリブ結果を可視化: 骨(テクスチャまたは灰) + ピック点(赤) + 復元マーカー4隅(青枠) + Mk座標軸(RGB)."""
		import numpy as np
		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title=title, window_size=(int(sw * 0.8), int(sh * 0.8)))
		plotter.set_background("white")
		# 骨とスケール
		try:
			b = np.array(mesh.bounds).reshape(3, 2)
			diag = float(np.linalg.norm(b[:, 1] - b[:, 0]))
			r_sphere = max(diag * 0.005, 0.5)
		except Exception:
			r_sphere = 1.0
		if texture is not None:
			try:
				# テクスチャ付きは完全不透明で表示 (半透明だとテクスチャが薄れる)
				plotter.add_mesh(mesh, texture=texture, opacity=1.0, smooth_shading=True)
			except Exception as e:
				print(f"[calib viz] テクスチャ描画失敗: {e}")
				plotter.add_mesh(mesh, color="lightgray", opacity=0.55, smooth_shading=True)
		else:
			plotter.add_mesh(mesh, color="lightgray", opacity=0.55, smooth_shading=True)
		# ピック点 (赤球+ラベル)
		for i, pt in enumerate(picked):
			plotter.add_mesh(pv.Sphere(radius=r_sphere, center=np.asarray(pt)), color="red")
			plotter.add_point_labels([np.asarray(pt)], [str(i + 1)],
			                          font_size=16, text_color="red",
			                          show_points=False, always_visible=True)
		# 復元マーカー4隅を line で結ぶ
		obj_pts = self._ankle_marker_obj_points(marker_size_mm)
		homog = np.hstack([obj_pts, np.ones((4, 1))])
		corners_L = (np.asarray(T, dtype=float) @ homog.T).T[:, :3]
		loop = np.vstack([corners_L, corners_L[0:1]])
		for i in range(4):
			plotter.add_mesh(pv.Line(loop[i], loop[i + 1]), color="blue", line_width=4)
		# Mk座標軸 (原点=T[:3,3], X=red, Y=green, Z=blue)
		axis_len = marker_size_mm * 0.6
		origin = np.asarray(T[:3, 3], dtype=float)
		Rmat = np.asarray(T[:3, :3], dtype=float)
		for j, col in enumerate(("red", "green", "blue")):
			direction = Rmat[:, j] * axis_len
			plotter.add_mesh(pv.Arrow(start=origin, direction=direction,
			                           scale='auto', tip_length=0.15, tip_radius=0.04, shaft_radius=0.015),
			                 color=col)
		plotter.add_text(
			"骨モデル(灰) / ピック点(赤球+番号) / 復元マーカー4隅(青枠) / Mk座標軸(X=R, Y=G, Z=B)",
			position="upper_left", font_size=10, color="black")
		plotter.show()

	# ---- Stage 4 新機能: テクスチャ付き骨スキャンからArUcoを自動検出 → T_L←Mk ----
	def _ankle_render_mesh_and_detect(self, mesh, cam_pos, look_at, up, W, H,
	                                    detector, dictionary, params, use_new_api,
	                                    expected_id: int, texture=None):
		"""1視点でメッシュをオフスクリーン描画→cv2.arucoで指定IDを探す。

		texture: pv.Texture (優先使用) or None。None時は mesh.textures を試す。
		Returns: (found: bool, corners_2d(4,2), area, cam_pos, look_at, up, W, H, view_angle_deg)
		"""
		import numpy as np
		import cv2
		try:
			plotter = pv.Plotter(off_screen=True, window_size=(W, H))
			plotter.set_background("white")
			# 明示的に渡されたテクスチャを優先
			tex = texture
			if tex is None:
				# メッシュ内蔵テクスチャを探す (fallback)
				try:
					if hasattr(mesh, 'textures') and mesh.textures:
						first_key = next(iter(mesh.textures))
						tex = mesh.textures[first_key]
				except Exception:
					tex = None
			try:
				if tex is not None:
					plotter.add_mesh(mesh, texture=tex, show_edges=False)
				else:
					plotter.add_mesh(mesh, show_edges=False, color="lightgray")
			except Exception:
				plotter.add_mesh(mesh, show_edges=False, color="lightgray")
			plotter.camera_position = [tuple(cam_pos), tuple(look_at), tuple(up)]
			view_angle = float(plotter.camera.view_angle)
			plotter.render()
			img_rgb = plotter.screenshot(return_img=True)
		except Exception as e:
			print(f"[ankle auto calib] render失敗: {e}")
			return False, None, 0.0, None, None, None, W, H, 30.0
		finally:
			try: plotter.close()
			except Exception: pass
		gray = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2GRAY)
		try:
			if use_new_api and detector is not None:
				corners, ids, _ = detector.detectMarkers(gray)
			else:
				corners, ids, _ = cv2.aruco.detectMarkers(gray, dictionary, parameters=params)
		except Exception as e:
			print(f"[ankle auto calib] detectMarkers失敗: {e}")
			return False, None, 0.0, cam_pos, look_at, up, W, H, view_angle
		if ids is None:
			return False, None, 0.0, cam_pos, look_at, up, W, H, view_angle
		for corner, mid_arr in zip(corners, ids):
			try:
				mid = int(mid_arr[0] if hasattr(mid_arr, '__len__') else mid_arr)
			except Exception:
				continue
			if mid != expected_id:
				continue
			c2d = corner.reshape(-1, 2).astype(np.float64)
			# 面積 = 4角形の外周
			try:
				area = float(cv2.contourArea(c2d.astype(np.float32)))
			except Exception:
				area = 0.0
			return True, c2d, area, cam_pos, look_at, up, W, H, view_angle
		return False, None, 0.0, cam_pos, look_at, up, W, H, view_angle

	def _ankle_auto_calibrate_from_mesh_impl(self, bone: dict, progress=None) -> tuple:
		"""テクスチャ付きOBJから ArUco 検出→T_L←Mk 推定。

		戦略:
		1. メッシュ中心を焦点に、球面上に多視点をサンプル
		2. 各視点でオフスクリーンレンダリング→cv2.aruco検出
		3. 該当IDが最大面積で見えた視点を選ぶ
		4. その視点で cv2.solvePnP を実行 (仮想カメラ内部パラメータを計算)
		5. T_cv2cam←Mk を T_L←Mk に変換

		Returns: (T (4,4) list, rmse mm, best_area px^2, n_hits, id)
		"""
		import numpy as np
		import cv2
		# キャリブ用モデル優先 (無ければ フルモデル)
		calib_path = self._ankle_get_calib_model_path(bone)
		if not calib_path or not Path(calib_path).exists():
			raise ValueError("キャリブ用モデル(or 骨モデル)が未設定 or ファイル無し")
		try:
			expected_id = int(bone.get("aruco_id", -1))
		except Exception:
			expected_id = -1
		if expected_id < 0:
			raise ValueError("骨のArUco IDが未設定")
		try:
			marker_size_mm = float(self.ankle_marker_size_mm.get())
		except Exception:
			marker_size_mm = 20.0
		if marker_size_mm <= 0:
			raise ValueError("マーカー実寸が不正")
		aruco_dict_name = str(self.ankle_aruco_dict_var.get())
		detector, dictionary, params, use_new_api = self._ankle_make_detector(aruco_dict_name)

		# メッシュ + テクスチャ読込
		mesh, texture = self._ankle_load_mesh_and_texture(calib_path)
		if mesh is None or mesh.n_points == 0:
			raise ValueError("メッシュ空")
		if texture is None:
			print("[auto calib] 警告: テクスチャが検出できませんでした。"
			      "灰色レンダリングになり、ArUcoは検出できません。"
			      "OBJの場合は .mtl と画像ファイルが同じフォルダにあるか、"
			      "PLYなら埋込テクスチャがあるか、同名の .png/.jpg が横にあるか確認してください。")
		center = np.array(mesh.center, dtype=float)
		diag = float(mesh.length)
		radius = diag * 1.2

		# 多視点サンプリング (球面, azimuth 8 × elevation 5 = 40視点)
		hits = []
		W, H = 1200, 1200
		up_default = np.array([0.0, 0.0, 1.0])
		up_alt = np.array([0.0, 1.0, 0.0])
		azimuths = np.linspace(0, 2 * np.pi, 8, endpoint=False)
		elevations = np.linspace(np.deg2rad(-70), np.deg2rad(70), 5)
		total = len(azimuths) * len(elevations)
		done = 0
		for phi in azimuths:
			for theta in elevations:
				cam_pos = center + radius * np.array([
					np.cos(theta) * np.cos(phi),
					np.cos(theta) * np.sin(phi),
					np.sin(theta),
				])
				# up ベクトルは cam→焦点方向にほぼ平行なら別軸に切替
				look_dir = center - cam_pos; look_dir /= np.linalg.norm(look_dir)
				up = up_default if abs(np.dot(look_dir, up_default)) < 0.95 else up_alt
				found, c2d, area, cp, la, u, w, h, va = \
					self._ankle_render_mesh_and_detect(
						mesh, cam_pos, center, up, W, H,
						detector, dictionary, params, use_new_api, expected_id,
						texture=texture)
				if found:
					hits.append({"c2d": c2d, "area": area, "cam_pos": cp,
					              "look_at": la, "up": u, "W": w, "H": h, "view_angle": va})
				done += 1
				if progress is not None:
					progress(done, total, len(hits))
		if not hits:
			raise ValueError(f"どの視点からもID={expected_id}が検出されませんでした")
		# 最大面積の視点を採用
		best = max(hits, key=lambda x: x["area"])
		c2d = best["c2d"]
		W_r, H_r = best["W"], best["H"]
		view_angle_deg = best["view_angle"]
		fov_y = np.deg2rad(view_angle_deg)
		fy = H_r / (2 * np.tan(fov_y / 2))
		fx = fy
		cx_p = W_r / 2.0
		cy_p = H_r / 2.0
		K = np.array([[fx, 0, cx_p], [0, fy, cy_p], [0, 0, 1]], dtype=np.float64)
		dist_c = np.zeros(5, dtype=np.float64)
		obj_pts = np.asarray(self._ankle_marker_obj_points(marker_size_mm), dtype=np.float64)
		try:
			ok, rvec, tvec = cv2.solvePnP(obj_pts, c2d, K, dist_c, flags=cv2.SOLVEPNP_IPPE)
		except Exception as e:
			raise ValueError(f"solvePnP失敗: {e}")
		if not ok:
			raise ValueError("solvePnP解が得られませんでした")
		R_cv, _ = cv2.Rodrigues(rvec)
		T_cv = np.eye(4); T_cv[:3, :3] = R_cv; T_cv[:3, 3] = tvec.flatten()

		# ワールド (mesh L) 系での PyVistaカメラ姿勢
		cam_pos = np.array(best["cam_pos"], dtype=float)
		look_at = np.array(best["look_at"], dtype=float)
		up_v = np.array(best["up"], dtype=float)
		z_pv = cam_pos - look_at
		z_pv /= np.linalg.norm(z_pv)
		x_pv = np.cross(up_v, z_pv); x_pv /= np.linalg.norm(x_pv)
		y_pv = np.cross(z_pv, x_pv)
		R_L_pv = np.column_stack([x_pv, y_pv, z_pv])
		T_L_pv = np.eye(4); T_L_pv[:3, :3] = R_L_pv; T_L_pv[:3, 3] = cam_pos
		# cv2カメラ (X右, Y下, Z奥) → PyVistaカメラ (X右, Y上, Z手前) の変換
		R_pv_cv = np.diag([1.0, -1.0, -1.0])
		T_pv_cv = np.eye(4); T_pv_cv[:3, :3] = R_pv_cv
		T_L_from_Mk = T_L_pv @ T_pv_cv @ T_cv

		# 残差 (対応点の再構成誤差, mm)
		homog = np.hstack([obj_pts, np.ones((4, 1))])
		pred_L = (T_L_from_Mk @ homog.T).T[:, :3]
		# cv2 側で予測した3D位置 (cv2カメラ系) → L系
		pred_cv2 = (T_cv @ homog.T).T[:, :3]
		pred_pv = (T_pv_cv @ np.hstack([pred_cv2, np.ones((4, 1))]).T).T[:, :3]
		# 再投影誤差 (pixel)
		try:
			img_pts_reproj, _ = cv2.projectPoints(obj_pts, rvec, tvec, K, dist_c)
			img_pts_reproj = img_pts_reproj.reshape(-1, 2)
			reproj_rmse = float(np.sqrt(np.mean(np.sum((img_pts_reproj - c2d) ** 2, axis=1))))
		except Exception:
			reproj_rmse = float("nan")
		return (T_L_from_Mk.tolist(), reproj_rmse, best["area"], len(hits), expected_id)

	def on_ankle_auto_calibrate_from_mesh(self) -> None:
		"""スキャンから自動でマーカー-骨キャリブを実行 (新プラン)。"""
		import numpy as np
		b = self._ankle_current_bone()
		if b is None:
			messagebox.showinfo("自動キャリブ", "先に骨を選択してください。"); return
		name = b.get("name", "?")
		model_path = str(b.get("model_path", "") or "").strip()
		if not model_path:
			messagebox.showwarning("自動キャリブ",
				"骨モデル (試験後スキャン、マーカー付き) が未設定です。③で選択してください。")
			return
		try:
			aid = int(b.get("aruco_id", -1))
		except Exception:
			aid = -1
		if aid < 0:
			messagebox.showwarning("自動キャリブ", "骨のArUco IDが未設定です。③で設定してください。")
			return

		# 進捗ダイアログ
		update_cb, close_cb, cancel_cb = self._ankle_open_progress(
			f"スキャンからArUco検出中 [{name}, ID={aid}]")
		def _progress(done, total, hits):
			try:
				update_cb(done, total, 0.0)
			except Exception:
				pass
		try:
			T_list, reproj_rmse, area, n_hits, det_id = \
				self._ankle_auto_calibrate_from_mesh_impl(b, progress=_progress)
		except Exception as e:
			close_cb()
			messagebox.showerror("自動キャリブ 失敗", f"{name}: {e}\n\n"
				"うまくいかない場合、手動「マーカー-骨キャリブ (手動4点)」を試してください。")
			return
		close_cb()

		# 保存
		b["marker_to_bone_T"] = T_list
		self._ankle_refresh_bone_listbox()

		# 品質評価
		if reproj_rmse < 1.0:
			verdict = "極めて良好"
		elif reproj_rmse < 2.0:
			verdict = "良好"
		elif reproj_rmse < 5.0:
			verdict = "許容範囲 (手動で確認推奨)"
		else:
			verdict = "要再取得または手動ピック"
		msg = (f"[{name}] マーカー-骨変換 T_L←Mk を自動保存しました。\n\n"
		       f"検出成功視点数 = {n_hits} / 40 視点\n"
		       f"採用視点のマーカー面積 = {area:.0f} px²\n"
		       f"再投影誤差 RMSE = {reproj_rmse:.3f} px\n"
		       f"評価         = {verdict}\n\n"
		       f"次のダイアログで確認画面を表示します。")
		messagebox.showinfo(f"{name} 自動キャリブ完了", msg)

		# 確認可視化 (既存の手動キャリブと同じフォーマット、テクスチャ付き)
		try:
			calib_path = self._ankle_get_calib_model_path(b)
			mesh, texture = self._ankle_load_mesh_and_texture(calib_path)
			marker_size_mm = float(self.ankle_marker_size_mm.get())
			T_arr = np.asarray(T_list, dtype=float)
			# 4隅の3D位置を obj_pts + T から生成
			obj_pts = np.asarray(self._ankle_marker_obj_points(marker_size_mm), dtype=float)
			homog = np.hstack([obj_pts, np.ones((4, 1))])
			corners_L = (T_arr @ homog.T).T[:, :3]
			self._ankle_visualize_calibration(
				mesh, corners_L, T_arr, marker_size_mm,
				title=f"{name}: 自動キャリブ確認 (青枠=復元, 赤=角)",
				texture=texture)
		except Exception as e:
			print(f"[ankle auto calib] 確認可視化失敗: {e}")

	def on_ankle_calibrate_marker_to_bone(self) -> None:
		"""選択中の骨のモデル上でマーカー4隅をクリック → T_L←Mk を計算し bone['marker_to_bone_T'] に保存。

		前提: 骨モデル(model_path)は試験後スキャンで、マーカー(ArUcoプレート)が骨に取り付いた状態。
		方式: マーカー4隅の実座標(mm)をユーザーがピック → マーカー座標系での既知4隅と剛体対応付け。
		結果は Stage 3で使うマーカー座標系→骨ローカル系変換 T_L←Mk (4x4)。
		"""
		import numpy as np
		b = self._ankle_current_bone()
		if b is None:
			messagebox.showinfo("マーカー-骨キャリブ", "先に骨を選択してください。")
			return
		# キャリブ用モデル優先 (無ければ フルモデル)
		calib_path = self._ankle_get_calib_model_path(b)
		if not calib_path:
			messagebox.showwarning("マーカー-骨キャリブ",
				"キャリブ用モデル or 骨モデルが未設定です。③で選択してください。")
			return
		try:
			marker_size_mm = float(self.ankle_marker_size_mm.get())
		except Exception:
			marker_size_mm = 20.0
		if marker_size_mm <= 0:
			messagebox.showwarning("マーカー-骨キャリブ", "マーカー実寸(mm)を正しく設定してください。")
			return

		name = b.get("name", "?")
		using_calib_only = bool(str(b.get("calib_model_path", "") or "").strip())
		src_label = "キャリブ用モデル (切り出し)" if using_calib_only else "骨モデル (フル)"
		messagebox.showinfo(
			"マーカー-骨キャリブ",
			f"[{name}] {src_label} 上のArUcoマーカーの4隅を順にクリックしてください。\n\n"
			f"クリック順序: 左上(TL) → 右上(TR) → 右下(BR) → 左下(BL)\n"
			f"（マーカー表面を正面から見た向きで）\n\n"
			f"※ マーカー実寸 = {marker_size_mm:.2f} mm を使用します。\n"
			f"※ 4点全てクリックしたらウィンドウを閉じてください。")

		# テクスチャ付きで読込
		try:
			mesh, texture = self._ankle_load_mesh_and_texture(calib_path)
		except Exception as e:
			messagebox.showerror("マーカー-骨キャリブ",
				f"モデル読込失敗:\n{e}"); return
		if mesh is None or mesh.n_points == 0:
			messagebox.showerror("マーカー-骨キャリブ", "モデルが空です。"); return
		if texture is None:
			print(f"[calib] テクスチャなし → 灰色レンダリング。ArUcoが視認しにくい可能性")

		# 膝の点ピッキングを再利用 (PyVistaの surface point picker)、テクスチャ渡す
		picked = self._knee_pick_points(mesh, 4, f"{name}: マーカー4隅 (TL→TR→BR→BL)",
		                                  texture=texture)
		if len(picked) < 4:
			messagebox.showwarning("マーカー-骨キャリブ",
				f"4点必要ですが {len(picked)} 点しかクリックされていません。やり直してください。")
			return

		# マーカー座標系での既知4隅 (Stage 3と同じ順 TL,TR,BR,BL, Y下向き)
		obj_pts_mk = self._ankle_marker_obj_points(marker_size_mm)  # (4,3) float32

		# Umeyama/Kabsch (obj_pts_mk → picked が T_L←Mk)
		T = self._rigid_from_correspondences(
			np.asarray(obj_pts_mk, dtype=float),
			np.asarray(picked, dtype=float),
			allow_scale=False)

		# 残差 (品質指標)
		homog = np.hstack([np.asarray(obj_pts_mk, dtype=float), np.ones((4, 1))])
		predicted_L = (T @ homog.T).T[:, :3]
		residuals = np.linalg.norm(predicted_L - np.asarray(picked, dtype=float), axis=1)
		rmse = float(np.sqrt(np.mean(residuals ** 2)))
		max_r = float(residuals.max())

		# 保存
		b["marker_to_bone_T"] = T.tolist()
		self._ankle_refresh_bone_listbox()

		# 品質メッセージ
		if rmse < 0.5:
			verdict = "極めて良好"
		elif rmse < 1.0:
			verdict = "良好"
		elif rmse < 2.5:
			verdict = "許容範囲 (推奨: 再ピッキング)"
		else:
			verdict = "要再取得 (RMSEが大きすぎます)"
		msg = (f"[{name}] マーカー-骨変換 T_L←Mk を保存しました。\n\n"
		       f"4隅の残差 RMSE = {rmse:.3f} mm\n"
		       f"最大残差     = {max_r:.3f} mm\n"
		       f"評価         = {verdict}\n\n"
		       f"次のダイアログで確認画面を表示します。")
		messagebox.showinfo(f"{name} キャリブ完了", msg)

		# 確認可視化
		try:
			self._ankle_visualize_calibration(
				mesh, picked, T, marker_size_mm,
				title=f"{name}: マーカー-骨キャリブ確認",
				texture=texture)
		except Exception as e:
			print(f"[ankle calib] 確認可視化失敗: {e}")

	# ---- Stage 2: 位置合わせ (膝の _knee_register_region を再利用) ----
	def _ankle_reg_params(self, bone: dict) -> dict:
		"""指定骨のパラメータdict(タブ共通 + 骨ごとの method/scaling)を返す。"""
		try:
			return {
				"ransac_distance": float(self.ankle_reg_ransac_distance.get()),
				"ransac_max_iter": int(self.ankle_reg_ransac_max_iter.get()),
				"ransac_confidence": float(self.ankle_reg_ransac_confidence.get()),
				"icp_threshold": float(self.ankle_reg_icp_threshold.get()),
				"icp_max_iter": int(self.ankle_reg_icp_max_iter.get()),
				"sample_points": int(self.ankle_reg_sample_points.get()),
				"voxel_size": float(self.ankle_reg_voxel_size.get()),
				"method": str(bone.get("method", "ransac")),
				"enable_scaling": bool(bone.get("enable_scaling", False)),
				"preview": bool(self.ankle_reg_preview_var.get()),
			}
		except Exception:
			return {"method": str(bone.get("method", "ransac"))}

	def _ankle_register_one(self, bone: dict) -> tuple:
		"""単骨の位置合わせを実行し (T, fit, rmse) を返す。エラーは例外。"""
		label = bone.get("name", "骨")
		src = str(bone.get("model_path", "") or "").strip()   # 骨モデル(試験後スキャン, 動かす側)
		tgt = str(bone.get("post_scan_path", "") or "").strip()  # 初期スキャン側の該当骨領域(合わせる先)
		if not src:
			raise ValueError(f"{label}: 骨モデルが選択されていません。")
		if not tgt:
			raise ValueError(f"{label}: 位置合わせ先(初期スキャン側 骨領域)が選択されていません。")
		p = self._ankle_reg_params(bone)
		method = p.get("method", "ransac")
		if method == "manual":
			messagebox.showinfo(
				"手動3点位置合わせ",
				f"[{label}] まず『骨モデル』で対応点を3つクリックし、ウィンドウを閉じてください。\n"
				f"次に『初期スキャン側 骨領域』で、同じ順番・同じ部位の3点をクリックしてください。")
			src_pts = self._knee_pick_points(pv.read(src), 3, f"{label}: 骨モデル (対応点3つ)")
			if len(src_pts) < 3:
				raise ValueError(f"{label}: 骨モデル側の対応点が3つ未満です ({len(src_pts)}点)。")
			tgt_pts = self._knee_pick_points(pv.read(tgt), 3, f"{label}: 初期スキャン側 骨領域 (同順で3つ)")
			if len(tgt_pts) < 3:
				raise ValueError(f"{label}: 初期スキャン側の対応点が3つ未満です ({len(tgt_pts)}点)。")
			return self._knee_register_region(src, tgt, label, p,
			                                   manual_src_pts=src_pts, manual_tgt_pts=tgt_pts)
		return self._knee_register_region(src, tgt, label, p)

	def on_ankle_register_bone(self) -> None:
		"""選択中の骨だけを位置合わせする。結果は bone['reg_T'] に保存。"""
		b = self._ankle_current_bone()
		if b is None:
			messagebox.showinfo("位置合わせ", "先に骨を選択してください。")
			return
		try:
			T, fit, rmse = self._ankle_register_one(b)
		except Exception as e:
			messagebox.showerror("位置合わせエラー", f"{b.get('name','?')}の位置合わせに失敗しました:\n{e}")
			return
		b["reg_T"] = np.asarray(T, dtype=float).tolist()
		self._ankle_refresh_bone_listbox()
		messagebox.showinfo(
			f"{b.get('name','')} 位置合わせ完了",
			f"fitness={fit:.4f} (1に近いほど良好), RMSE={rmse:.4f} mm\n\n"
			"「位置合わせ結果を確認 (プレビュー)」で重ね合わせを確認できます。")

	def on_ankle_register_all_bones(self) -> None:
		"""全骨を順に位置合わせ。既に成功した骨はスキップ確認する。"""
		if not self.ankle_bones:
			messagebox.showinfo("一括位置合わせ", "骨リストが空です。まず骨を追加してください。")
			return
		targets = []
		for b in self.ankle_bones:
			label = b.get("name", "骨")
			if not (b.get("model_path") and b.get("post_scan_path")):
				print(f"[ankle reg-all] skip {label}: 入力ファイル未設定")
				continue
			if b.get("reg_T") is not None:
				if not messagebox.askyesno("上書き確認",
					f"骨「{label}」は既に位置合わせ済です。再計算しますか？"):
					continue
			targets.append(b)
		if not targets:
			messagebox.showinfo("一括位置合わせ", "実行対象の骨がありません。")
			return
		results = []
		for b in targets:
			label = b.get("name", "骨")
			try:
				T, fit, rmse = self._ankle_register_one(b)
				b["reg_T"] = np.asarray(T, dtype=float).tolist()
				results.append((label, True, fit, rmse, ""))
			except Exception as e:
				results.append((label, False, 0.0, 0.0, str(e)))
		self._ankle_refresh_bone_listbox()
		msg = "\n".join(
			f"{'✓' if ok else '×'} {name}: "
			+ (f"fitness={fit:.4f}, RMSE={rmse:.4f}mm" if ok else f"失敗 — {err}")
			for name, ok, fit, rmse, err in results)
		messagebox.showinfo("一括位置合わせ 結果", msg)

	def on_ankle_preview_registration(self) -> None:
		"""初期スキャン(灰半透明) + 位置合わせ済み全骨を重ねて表示する。"""
		scan_path = self.ankle_initial_scan_path.get().strip()
		if not scan_path:
			messagebox.showwarning("プレビュー", "①初期状態スキャンを選択してください。")
			return
		registered = [b for b in self.ankle_bones if b.get("reg_T") is not None and b.get("model_path")]
		if not registered:
			messagebox.showwarning("プレビュー", "位置合わせ済みの骨がありません。先に位置合わせを実行してください。")
			return
		try:
			scan_mesh = pv.read(scan_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"初期状態スキャンの読み込みに失敗しました:\n{e}")
			return
		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="ankle: 位置合わせ結果 プレビュー",
		                     window_size=(int(sw * 0.9), int(sh * 0.9)))
		plotter.set_background("white")
		plotter.add_mesh(scan_mesh, color="lightgray", opacity=0.35, smooth_shading=True)
		names = []
		for b in registered:
			try:
				m = pv.read(b["model_path"])
				T = np.asarray(b["reg_T"], dtype=float)
				self._knee_apply_T(m, T)
				plotter.add_mesh(m, color=self._ankle_color_of(self.ankle_bones.index(b)),
				                 opacity=1.0, smooth_shading=True)
				names.append(b.get("name", "?"))
			except Exception as e:
				print(f"[ankle preview] {b.get('name','?')} 表示失敗: {e}")
		plotter.add_text(f"初期スキャン(灰・半透明) + 位置合わせ済み: {', '.join(names)}",
		                 position="upper_left", font_size=10, color="black")
		plotter.show()

	def on_ankle_register(self) -> None:
		# 後方互換: プレースホルダーAPI (未使用)
		self.on_ankle_register_bone()

	# ---- Stage 3: ArUco検出 + PnP パイプライン ----
	def _ankle_check_cv2(self) -> bool:
		"""cv2/aruco が使えるか確認。無ければ案内。"""
		try:
			import cv2  # noqa: F401
			import cv2.aruco  # noqa: F401
			return True
		except ImportError:
			messagebox.showerror(
				"OpenCV未導入",
				"ArUco検出には opencv-contrib-python が必要です。\n\n"
				"venv で以下を実行してください:\n"
				"  pip install opencv-contrib-python")
			return False

	def _ankle_check_rs(self) -> bool:
		"""pyrealsense2 (RealSense .bag読込) が使えるか確認。"""
		try:
			import pyrealsense2  # noqa: F401
			return True
		except ImportError:
			messagebox.showerror(
				"pyrealsense2未導入",
				"RealSense .bag ファイルの読込には pyrealsense2 が必要です。\n\n"
				"venv で以下を実行してください:\n"
				"  pip install pyrealsense2\n\n"
				"※ .bag以外(mp4+depth+json)なら pyrealsense2 は不要です。")
			return False

	def _ankle_resolve_aruco_dict(self, name: str):
		"""'DICT_4X4_50' のような名前を cv2 の辞書オブジェクトへ解決。"""
		import cv2
		const = getattr(cv2.aruco, name, None)
		if const is None:
			raise ValueError(f"未対応のArUco辞書名: {name}")
		return cv2.aruco.getPredefinedDictionary(const)

	def _ankle_marker_obj_points(self, marker_size_mm: float):
		"""マーカー座標系での4隅座標 (mm単位)。

		cv2.aruco.detectMarkers の返すコーナー順 (画像座標系で TL→TR→BR→BL, Y下向き)
		と一致させるため、マーカー座標系も Y下向き (=画像座標系互換) にする。
		結果: solvePnPで得られる rvec,tvec は「マーカー→カメラ」変換。
		"""
		s = float(marker_size_mm) / 2.0
		import numpy as np
		return np.array([
			[-s, -s, 0.0],   # TL
			[ s, -s, 0.0],   # TR
			[ s,  s, 0.0],   # BR
			[-s,  s, 0.0],   # BL
		], dtype=np.float32)

	@staticmethod
	def _ankle_safe_print(msg) -> None:
		"""コンソールのエンコーディングで表示できない文字があっても例外を出さない print。

		Windows の cp932 コンソールでは絵文字や一部記号が UnicodeEncodeError を起こす。
		計算処理の途中でログ出力が例外になると結果を失うため、安全側に倒す。
		"""
		try:
			print(msg)
		except Exception:
			try:
				print(str(msg).encode("ascii", "replace").decode("ascii"))
			except Exception:
				pass

	@staticmethod
	def _ankle_imwrite(path, img) -> bool:
		"""日本語パス対応の cv2.imwrite。

		Windows の OpenCV は cv2.imwrite に非ASCIIパスを渡すと黙って失敗する
		(戻り値 False, 例外なし)。imencode でメモリ上にエンコードしてから
		numpy.tofile で書き出すことで回避する。
		"""
		import cv2
		import numpy as np
		try:
			ext = Path(str(path)).suffix or ".png"
			ok, buf = cv2.imencode(ext, img)
			if not ok:
				return False
			buf.tofile(str(path))
			return True
		except Exception as e:
			print(f"[imwrite] 保存失敗 {path}: {e}")
			return False

	def _ankle_auto_correct_depth_scale(self, gray, depth_arr, K, dist, obj_pts,
	                                     target_ids, detector, dictionary, params,
	                                     use_new_api, depth_scale_mm: float):
		"""SDKが報告する depth_scale が実データと合わない場合に自動補正する。

		【背景】一部の RealSense SDK / .db3 記録では depth_scale が誤って報告される
		(例: D405 の真値 0.0001 m/unit に対し 0.001 m/unit を返す = 10倍過大)。
		この場合、深度から復元する 3D 座標がすべて 10倍の距離になり、
		ArUco の solvePnP 結果 (マーカー実寸から計算・スケール的に信頼できる) と乖離する。

		【手法】ArUco を検出できるフレームで
		    ratio = solvePnP の tvec_Z / (マーカー4隅の raw depth 中央値 × 現行 scale)
		を求める。ratio が 1.0 付近なら補正不要。10 や 1/10 のような綺麗な倍率に
		近い場合のみ、その倍率で depth_scale_mm を補正する
		(中途半端な値は物理的なズレの可能性があるため補正しない)。

		Returns:
			(corrected_depth_scale_mm: float, info: dict)
			補正しなかった場合は入力値をそのまま返す。
		"""
		import cv2
		import numpy as np
		info = {"corrected": False, "ratio": None, "factor": 1.0,
		        "original_mm": float(depth_scale_mm)}
		try:
			if use_new_api and detector is not None:
				corners, ids, _ = detector.detectMarkers(gray)
			else:
				corners, ids, _ = cv2.aruco.detectMarkers(gray, dictionary, parameters=params)
			if ids is None:
				return depth_scale_mm, info
			obj_pts_f64 = np.asarray(obj_pts, dtype=np.float64)
			H, W = depth_arr.shape[:2]
			ratios = []
			for corner, mid_arr in zip(corners, ids):
				mid = int(mid_arr[0] if hasattr(mid_arr, '__len__') else mid_arr)
				if mid not in target_ids:
					continue
				img_pts = corner.reshape(-1, 2).astype(np.float64)
				# solvePnP の Z (マーカー実寸に基づく距離。スケール的に信頼できる)
				try:
					ok, rvec, tvec = cv2.solvePnP(obj_pts_f64, img_pts, K, dist,
					                               flags=cv2.SOLVEPNP_ITERATIVE)
					if not ok:
						continue
					pnp_z = float(np.asarray(tvec).flatten()[2])
				except Exception:
					continue
				if pnp_z <= 0:
					continue
				# マーカー4隅の raw depth (5x5窓の中央値)
				raws = []
				for pt in img_pts:
					u_c, v_c = int(round(pt[0])), int(round(pt[1]))
					patch = []
					for du in range(-2, 3):
						for dv in range(-2, 3):
							u, v = u_c + du, v_c + dv
							if 0 <= u < W and 0 <= v < H:
								d_raw = float(depth_arr[v, u])
								if d_raw > 0:
									patch.append(d_raw)
					if patch:
						raws.append(float(np.median(patch)))
				if len(raws) < 3:
					continue
				depth_z = float(np.median(raws)) * depth_scale_mm
				if depth_z <= 0:
					continue
				ratios.append(pnp_z / depth_z)
			if not ratios:
				return depth_scale_mm, info
			ratio = float(np.median(ratios))
			info["ratio"] = ratio
			# 1.0 付近 (0.7〜1.4) なら補正不要
			if 0.7 <= ratio <= 1.4:
				return depth_scale_mm, info
			# 10 の整数乗に近いか判定 (log10 が整数に近い = 単位系の取り違え)
			log10_ratio = float(np.log10(ratio))
			nearest_pow = round(log10_ratio)
			if nearest_pow == 0 or abs(log10_ratio - nearest_pow) > 0.08:
				# 綺麗な倍率でない → 物理的なズレの可能性。補正しない
				self._ankle_safe_print(
					f"[深度スケール] 不一致率 {ratio:.2f}x を検出しましたが、"
					f"10の整数乗ではないため自動補正しません "
					f"(マーカー実寸設定 or 深度品質を確認してください)")
				return depth_scale_mm, info
			factor = float(10.0 ** nearest_pow)
			corrected = depth_scale_mm * factor
			info.update({"corrected": True, "factor": factor,
			             "corrected_mm": corrected})
			# 【重要】print の失敗 (cp932 で表示できない文字など) で補正結果を失わないよう、
			# 表示は安全な print ヘルパ経由で行い、例外を出さない
			self._ankle_safe_print("=" * 70)
			self._ankle_safe_print(
				f"[深度スケール] [警告] SDK報告の depth_scale が実データと {ratio:.2f}倍 乖離")
			self._ankle_safe_print("  solvePnP (マーカー実寸 基準) と 深度値 を比較した結果、")
			self._ankle_safe_print(
				f"  depth_scale_mm を {depth_scale_mm:.6f} -> {corrected:.6f} mm/unit に自動補正します"
				f" (x{factor:g})")
			self._ankle_safe_print("  ※ 補正後は深度から復元する 3D 座標が solvePnP と整合します")
			self._ankle_safe_print("=" * 70)
			return corrected, info
		except Exception as e:
			self._ankle_safe_print(
				f"[深度スケール] 自動補正の判定に失敗 ({e}) - SDK報告値を使用します")
			return depth_scale_mm, info

	@staticmethod
	def _ankle_fit_plane_ransac(pts, thresh_mm: float = 1.5, iters: int = 120, seed: int = 0):
		"""3D点群に平面をロバストフィットする。

		RANSAC で外れ値 (flying pixel など) を除いてから、inlier に SVD をかけて最終化する。

		Returns:
			(n, d, inlier_mask, rms) — 平面は n·p + d = 0 (|n|=1)。失敗時は (None, None, None, None)
		"""
		import numpy as np
		pts = np.asarray(pts, dtype=np.float64)
		n_pts = len(pts)
		if n_pts < 3:
			return None, None, None, None
		rng = np.random.default_rng(seed)
		best_mask = None
		best_cnt = 0
		for _ in range(iters):
			idx = rng.choice(n_pts, 3, replace=False)
			p1, p2, p3 = pts[idx]
			nv = np.cross(p2 - p1, p3 - p1)
			nn = float(np.linalg.norm(nv))
			if nn < 1e-9:
				continue
			nv = nv / nn
			dd = -float(nv @ p1)
			inl = np.abs(pts @ nv + dd) < thresh_mm
			cnt = int(inl.sum())
			if cnt > best_cnt:
				best_cnt = cnt
				best_mask = inl
		if best_mask is None or best_cnt < 3:
			return None, None, None, None
		# inlier で最小二乗リファイン (SVD)
		inl_pts = pts[best_mask]
		centroid = inl_pts.mean(axis=0)
		try:
			_, _, Vt = np.linalg.svd(inl_pts - centroid, full_matrices=False)
		except np.linalg.LinAlgError:
			return None, None, None, None
		n = Vt[2]
		nn = float(np.linalg.norm(n))
		if nn < 1e-12:
			return None, None, None, None
		n = n / nn
		d = -float(n @ centroid)
		rms = float(np.sqrt(np.mean((inl_pts @ n + d) ** 2)))
		return n, d, best_mask, rms

	def _ankle_pose_from_rgb_depth_fusion(self, img_pts, obj_pts, depth_arr, K, dist,
	                                       depth_scale_mm, shrink: float = 0.80,
	                                       ransac_thresh_mm: float = 1.5,
	                                       min_plane_points: int = 60,
	                                       expand: float = 2.2,
	                                       use_measured_size: bool = True):
		"""RGB の面内精度 と Depth の面外精度 を融合して ArUco 姿勢を推定する (推奨手法)。

		【なぜこれが必要か】
		マーカー4隅の深度だけで姿勢を組むと、20mm マーカーに対し深度ノイズ 0.5〜1mm が
		そのまま角度誤差 atan(1/10) ≈ 3° として乗る。しかも角はマーカーと背景の境界で、
		深度が最も荒れる場所 (flying pixel) なので実際にはさらに悪化する。
		これがフレーム毎のガタつきの主因になる。

		【手法】RGB と Depth の得意分野だけを使う:
		  - Depth: マーカー *内部* の数百〜数千画素に平面をフィット → 法線が √N で高精度化
		           (4点 → 2000点で理論上 40倍以上の改善)
		  - RGB:   サブピクセル角から歪み補正済みの視線ベクトルを作る (面内位置・回転が高精度)
		  - 融合:  各視線を平面と交差させ、両者に矛盾しない 3D 角を求める
		  - 姿勢:  得られた 4 隅で Umeyama 剛体フィット

		角の深度を一切読まないため、境界の flying pixel の影響を受けない。
		また 3D 点が直接得られるので IPPE の表裏 2 解の曖昧性も原理的に発生しない。

		Args:
			img_pts: (4,2) サブピクセル角 (整数化しないこと)
			obj_pts: (4,3) マーカー座標系の 4 隅 [mm]
			depth_arr: (H,W) 深度画像 (raw units)
			K, dist: カメラ内部パラメータ・歪み係数
			depth_scale_mm: raw → mm
			shrink: 角に向かって縮める率。0.8 なら面積 64% の内側だけを平面フィットに使う
			ransac_thresh_mm: 平面 inlier しきい値
			min_plane_points: 平面フィットに必要な最小点数

		Returns:
			(T: 4x4, ok: bool, info: dict)
		"""
		import cv2
		import numpy as np
		info = {"method": "rgbd_fusion", "n_plane_points": 0, "plane_rms_mm": None,
		        "inlier_ratio": None, "fit_rmse_mm": None, "reason": None,
		        "plane_radius_mm": None, "expanded": False}
		try:
			H, W = depth_arr.shape[:2]
			img_pts = np.asarray(img_pts, dtype=np.float64)
			obj_pts_f = np.asarray(obj_pts, dtype=np.float64)
			c2d = img_pts.mean(axis=0)

			def _collect(scale):
				"""重心を中心に scale 倍した多角形内の有効深度画素を 3D 化して返す。"""
				poly_i = np.round(c2d + (img_pts - c2d) * float(scale)).astype(np.int32)
				m = np.zeros((H, W), dtype=np.uint8)
				cv2.fillPoly(m, [poly_i], 1)
				yy, xx = np.nonzero(m)
				if len(xx) == 0:
					return None
				zz = depth_arr[yy, xx].astype(np.float64) * float(depth_scale_mm)
				g = zz > 0
				if int(g.sum()) == 0:
					return None
				xx = xx[g].astype(np.float64); yy = yy[g].astype(np.float64); zz = zz[g]
				pix = np.stack([xx, yy], axis=1).reshape(-1, 1, 2)
				und = cv2.undistortPoints(pix, K, dist).reshape(-1, 2)
				return np.stack([und[:, 0] * zz, und[:, 1] * zz, zz], axis=1)

			# --- 1. アンカー: マーカー内部だけで平面を仮フィット ---
			# ここは確実にマーカー面なので、正しい平面を掴むための足がかりになる。
			anchor_pts = _collect(shrink)
			if anchor_pts is None or len(anchor_pts) < min_plane_points:
				info["reason"] = f"マーカー内部の有効深度が少ない ({0 if anchor_pts is None else len(anchor_pts)})"
				return np.eye(4), False, info
			n0, d0, _, _ = self._ankle_fit_plane_ransac(
				anchor_pts, thresh_mm=ransac_thresh_mm, iters=120)
			if n0 is None:
				info["reason"] = "アンカー平面フィット失敗"
				return np.eye(4), False, info

			# --- 2. 台座まで領域を広げる ---
			# 法線精度は σ/r で決まり (深度ノイズは空間相関があるので √N では稼げない)、
			# r を大きくすることが唯一の効く手。マーカーが貼られた平らな台座は
			# 印刷模様が無く深度が綺麗なので、そこまで使う。
			# アンカー平面から離れた点 (骨・背景) は距離しきい値で自動的に外れる。
			pts3d = anchor_pts
			if expand and expand > shrink:
				wide = _collect(expand)
				if wide is not None and len(wide) > len(anchor_pts):
					keep = np.abs(wide @ n0 + d0) < (ransac_thresh_mm * 2.0)
					if int(keep.sum()) >= len(anchor_pts):
						pts3d = wide[keep]
						info["expanded"] = True
			info["n_plane_points"] = int(len(pts3d))

			# --- 3. 拡張した点群で平面を最終フィット ---
			n_vec, d_val, inl, rms = self._ankle_fit_plane_ransac(
				pts3d, thresh_mm=ransac_thresh_mm, iters=120)
			if n_vec is None:
				info["reason"] = "平面フィット失敗"
				return np.eye(4), False, info
			info["plane_rms_mm"] = float(rms)
			info["inlier_ratio"] = float(inl.sum()) / float(len(pts3d))
			if info["inlier_ratio"] < 0.4:
				info["reason"] = f"平面 inlier 率が低い ({info['inlier_ratio']:.2f})"
				return np.eye(4), False, info
			# 法線精度の指標: 点群の重心まわりの RMS 半径 (大きいほど法線が安定)
			try:
				inl_pts = pts3d[inl]
				ctr = inl_pts.mean(axis=0)
				info["plane_radius_mm"] = float(np.sqrt(np.mean(
					np.sum((inl_pts - ctr) ** 2, axis=1) - ((inl_pts - ctr) @ n_vec) ** 2)))
			except Exception:
				pass

			# --- 4. RGB のサブピクセル角 → 視線ベクトル (歪み補正込み) ---
			und_c = cv2.undistortPoints(img_pts.reshape(-1, 1, 2), K, dist).reshape(-1, 2)
			rays = np.concatenate([und_c, np.ones((len(und_c), 1))], axis=1)  # (4,3)

			# --- 5. 視線 × 平面 の交点 = 深度ノイズに汚されていない 3D 角 ---
			denom = rays @ n_vec
			if np.any(np.abs(denom) < 1e-9):
				info["reason"] = "視線が平面と平行"
				return np.eye(4), False, info
			t_scale = -d_val / denom
			if np.any(t_scale <= 0):
				info["reason"] = "交点がカメラ後方"
				return np.eye(4), False, info
			corners3d = rays * t_scale[:, None]

			# --- 6. 実測マーカーサイズ ---
			# 視線×平面の交点は「実物の角の3D位置」なので、その辺長 = 実際のマーカー寸法。
			# ②の設定と食い違う (印刷スケールのズレ) と、剛体フィットに残差が乗って
			# 姿勢がわずかに歪む。実測値に合わせた obj_pts を使えばこの誤差は消える。
			nominal = float(np.linalg.norm(obj_pts_f[0] - obj_pts_f[1]))
			obj_fit = obj_pts_f
			try:
				edges = [float(np.linalg.norm(corners3d[i] - corners3d[(i + 1) % 4]))
				         for i in range(4)]
				measured = float(np.mean(edges))
				info["measured_marker_mm"] = measured
				info["marker_edge_spread_mm"] = float(np.max(edges) - np.min(edges))
				if use_measured_size and nominal > 1e-6:
					ratio = measured / nominal
					# 極端な値は誤検出とみなして採用しない
					if 0.8 < ratio < 1.25:
						obj_fit = obj_pts_f * ratio
						info["size_ratio_applied"] = float(ratio)
			except Exception:
				pass

			# --- 7. Umeyama 剛体フィット ---
			T = self._rigid_from_correspondences(obj_fit, corners3d, allow_scale=False)
			pred = (T[:3, :3] @ obj_fit.T).T + T[:3, 3]
			resid = np.linalg.norm(pred - corners3d, axis=1)
			info["fit_rmse_mm"] = float(np.sqrt(np.mean(resid ** 2)))
			return T, True, info
		except Exception as e:
			info["reason"] = f"例外: {e}"
			return np.eye(4), False, info

	def _ankle_make_detector(self, aruco_dict_name: str):
		"""ArUco Detectorインスタンスを返す (新旧API対応)。

		【高精度化設定 (Option C)】検出コーナーのサブピクセル精度を向上させる:
		- CORNER_REFINE_APRILTAG (contribの新精度化アルゴ) を優先。
		  内部でエッジ勾配から放物線フィットでピーク位置を求めるためピクセル 1/10 精度が出る。
		- 未対応環境では CORNER_REFINE_SUBPIX にフォールバック (window size を 5→7 拡大、
		  反復回数 30→50 増、収束精度 0.01→0.001 に厳格化)。
		- 検出パラメータも高精度寄りに調整 (adaptive threshold / polygon approx / error corr)。
		"""
		import cv2
		dictionary = self._ankle_resolve_aruco_dict(aruco_dict_name)
		try:
			params = cv2.aruco.DetectorParameters()
			# --- コーナーサブピクセル精緻化 (最高精度優先) ---
			refined = False
			try:
				apriltag_flag = getattr(cv2.aruco, 'CORNER_REFINE_APRILTAG', None)
				if apriltag_flag is not None:
					params.cornerRefinementMethod = apriltag_flag
					# APRILTAG 用パラメータ (放物線フィット近傍サイズ, deglitch)
					try:
						params.aprilTagQuadDecimate = 1.0  # 1.0 = decimate 無し (精度最優先)
						params.aprilTagQuadSigma = 0.0
						params.aprilTagMinClusterPixels = 5
						params.aprilTagMaxNmaxima = 10
						params.aprilTagCriticalRad = 10.0 * np.pi / 180.0
						params.aprilTagMaxLineFitMse = 10.0
						params.aprilTagMinWhiteBlackDiff = 5
						params.aprilTagDeglitch = 0
					except Exception:
						pass
					refined = True
			except Exception:
				refined = False
			if not refined:
				# フォールバック: SUBPIX (window/iter を拡大して精度向上)
				try:
					params.cornerRefinementMethod = cv2.aruco.CORNER_REFINE_SUBPIX
					params.cornerRefinementWinSize = 7                # 5 → 7 拡大
					params.cornerRefinementMaxIterations = 50         # 30 → 50 増加
					params.cornerRefinementMinAccuracy = 0.001        # 0.01 → 0.001 厳格化
				except Exception:
					pass
			# 実際に何が有効になったかを明示する。
			# 角検出精度が姿勢精度を直接決めるので、意図せず低精度モードに落ちていないか確認する。
			try:
				actual = int(params.cornerRefinementMethod)
				names = {}
				for nm in ("CORNER_REFINE_NONE", "CORNER_REFINE_SUBPIX",
				            "CORNER_REFINE_CONTOUR", "CORNER_REFINE_APRILTAG"):
					v = getattr(cv2.aruco, nm, None)
					if v is not None:
						names[int(v)] = nm
				label = names.get(actual, f"不明({actual})")
				extra = ""
				if label == "CORNER_REFINE_SUBPIX":
					extra = (f" win={getattr(params, 'cornerRefinementWinSize', '?')}"
					         f" iter={getattr(params, 'cornerRefinementMaxIterations', '?')}")
				self._ankle_safe_print(f"[ArUco] コーナー精緻化 = {label}{extra}")
				if label in ("CORNER_REFINE_NONE",):
					self._ankle_safe_print(
						"  [警告] 精緻化が無効です。角がピクセル単位でしか出ないため姿勢が大きく荒れます")
			except Exception:
				pass
			# --- 検出パラメータの高精度寄り調整 ---
			try:
				# 適応的閾値の細かいスイープ (窓サイズを 3→23 の間で走査)
				params.adaptiveThreshWinSizeMin = 3
				params.adaptiveThreshWinSizeMax = 23
				params.adaptiveThreshWinSizeStep = 4
				params.adaptiveThreshConstant = 7
				# 小さいマーカーでも検出可
				params.minMarkerPerimeterRate = 0.02
				params.maxMarkerPerimeterRate = 4.0
				# 多角形近似を厳格化 (角検出精度向上)
				params.polygonalApproxAccuracyRate = 0.03
				# ボーダー誤検出許容を厳格化 (誤検出削減)
				params.maxErroneousBitsInBorderRate = 0.15
				# エラー訂正を強化 (ID誤読減)
				params.errorCorrectionRate = 0.6
			except Exception:
				pass
			return cv2.aruco.ArucoDetector(dictionary, params), dictionary, params, True
		except AttributeError:
			# 旧API
			params = cv2.aruco.DetectorParameters_create()
			return None, dictionary, params, False

	def _ankle_pose_from_depth_corners(self, img_pts, obj_pts, depth_arr, K, depth_scale_mm,
	                                    patch_radius: int = 2):
		"""ArUcoの4隅ピクセルで深度を直接サンプルし、3D点から姿勢を推定する (mono PnP に依存しない)。

		【この関数の意義】
		mono PnP は marker_size_mm 設定と実物寸法が一致していないと tvec Z を大幅に誤る。
		本関数は「RGB で検出したコーナー座標 → その位置の深度」で 3D 点を得るため、
		marker_size_mm には依存しない (obj_pts の相対形状にのみ依存)。
		D405 の実測深度精度 (~0.1mm/unit at 30cm) をフル活用できる。

		Args:
			img_pts: (4, 2) RGB で検出された 4隅ピクセル座標 (u, v)
			obj_pts: (4, 3) マーカー物体座標系の 4隅 (mm)
			depth_arr: (H, W) 深度画像 (raw units)
			K: (3, 3) カメラ内部行列
			depth_scale_mm: depth 値 → mm 変換係数
			patch_radius: 各コーナーで深度をサンプルする窓半径 (2 → 5x5窓の中央値)

		Returns:
			(T_C_Mk: 4x4, ok: bool, info: dict)
			ok=False の場合 T_C_Mk は identity。info には診断情報。
		"""
		import numpy as np
		H, W = depth_arr.shape[:2]
		fx, fy = K[0, 0], K[1, 1]
		cx, cy = K[0, 2], K[1, 2]
		corners_3d = []
		corner_depths_mm = []
		for pt in img_pts:
			u_c, v_c = int(round(pt[0])), int(round(pt[1]))
			# コーナー中心の周辺で depth をサンプル (単一ピクセル depth は 0 が出やすいので中央値)
			depths = []
			for du in range(-patch_radius, patch_radius + 1):
				for dv in range(-patch_radius, patch_radius + 1):
					u, v = u_c + du, v_c + dv
					if 0 <= u < W and 0 <= v < H:
						d_raw = float(depth_arr[v, u])
						if d_raw > 0:
							depths.append(d_raw * depth_scale_mm)
			if len(depths) < 3:
				corners_3d.append(None)
				corner_depths_mm.append(0.0)
				continue
			# メジアン (外れ値除去)
			z_mm = float(np.median(depths))
			corner_depths_mm.append(z_mm)
			# u, v から 3D 座標 (バックプロジェクション)
			x_mm = (u_c - cx) * z_mm / fx
			y_mm = (v_c - cy) * z_mm / fy
			corners_3d.append([x_mm, y_mm, z_mm])
		valid = [c for c in corners_3d if c is not None]
		if len(valid) < 3:
			return np.eye(4), False, {
				"reason": f"valid corners = {len(valid)} < 3",
				"corner_depths_mm": corner_depths_mm,
			}
		# 有効なコーナーだけで剛体変換 (obj_pts (marker frame) → cam_pts (camera frame))
		obj_valid = np.asarray([obj_pts[i] for i in range(4) if corners_3d[i] is not None], dtype=np.float64)
		cam_valid = np.asarray([corners_3d[i] for i in range(4) if corners_3d[i] is not None], dtype=np.float64)
		try:
			T = self._rigid_from_correspondences(obj_valid, cam_valid, allow_scale=False)
		except Exception as e:
			return np.eye(4), False, {
				"reason": f"rigid fit failed: {e}",
				"corner_depths_mm": corner_depths_mm,
			}
		# 残差 (mm) — 各コーナーの再投影誤差
		cam_pred = (T[:3, :3] @ obj_valid.T).T + T[:3, 3]
		residuals = np.linalg.norm(cam_pred - cam_valid, axis=1)
		rmse = float(np.sqrt(np.mean(residuals ** 2)))
		return T, True, {
			"rmse_mm": rmse,
			"corner_depths_mm": corner_depths_mm,
			"n_valid_corners": int(len(valid)),
		}

	def _ankle_refine_pose_with_depth(self, T_orig, img_pts, depth_arr, K, depth_scale_mm,
	                                    roi_pad_ratio: float = 0.15,
	                                    min_inliers: int = 20,
	                                    ransac_thresh_mm: float = 1.0,
	                                    n_ransac_iter: int = 200):
		"""検出済み ArUco pose を深度画像の平面フィットで精緻化する (Option A)。

		【アルゴリズム】
		1. マーカー4隅の bounding box を pad で拡張した ROI 内の深度画素を 3D 化
		2. RANSAC で平面 n·p + d = 0 を推定 (外れ値耐性)
		3. inlier 集合で SVD 最小二乗リファイン (平面パラメータの精度向上)
		4. T_orig の位置 t を平面上に投影 (Z誤差消し込み)
		5. T_orig のマーカー Z軸を平面法線に揃える回転 R_align を適用 (法線ズレ修正)

		【効果】mono PnP の Z軸誤差 (0.3-1mm) を depth 平面フィット (0.05-0.2mm) に置換。
		D405 の 0.1mm/unit 深度精度が最大限活かされる。

		Args:
			T_orig: 4x4 mono PnP + depth-flip 済みの初期 pose (未精緻化)
			img_pts: (4, 2) 検出されたマーカー4隅のピクセル座標
			depth_arr: (H, W) 深度画像 (raw units)
			K: (3, 3) カメラ内部行列
			depth_scale_mm: depth 値 → mm 変換係数
			roi_pad_ratio: マーカー bounding box の pad 比率
			min_inliers: 平面フィット inlier 最小数 (下回れば精緻化スキップ)
			ransac_thresh_mm: RANSAC inlier 判定閾値 [mm]
			n_ransac_iter: RANSAC 反復回数

		Returns:
			(T_refined: 4x4, refined: bool, info: dict)
			refined=False の場合 T_orig をそのまま返す
		"""
		import numpy as np
		H, W = depth_arr.shape[:2]
		xs = img_pts[:, 0]; ys = img_pts[:, 1]
		xmin, xmax = float(xs.min()), float(xs.max())
		ymin, ymax = float(ys.min()), float(ys.max())
		w = xmax - xmin; h = ymax - ymin
		px = int(round(w * roi_pad_ratio))
		py = int(round(h * roi_pad_ratio))
		u0 = max(0, int(round(xmin)) - px)
		u1 = min(W, int(round(xmax)) + px + 1)
		v0 = max(0, int(round(ymin)) - py)
		v1 = min(H, int(round(ymax)) + py + 1)
		if u1 <= u0 or v1 <= v0:
			return T_orig, False, {"reason": "empty ROI"}
		depth_roi = depth_arr[v0:v1, u0:u1].astype(np.float32) * depth_scale_mm
		valid = depth_roi > 0
		if valid.sum() < min_inliers:
			return T_orig, False, {"reason": f"too few depth pixels ({valid.sum()})"}
		us = np.arange(u0, u1)
		vs = np.arange(v0, v1)
		uu, vv = np.meshgrid(us, vs)
		zz = depth_roi
		fx, fy = K[0, 0], K[1, 1]
		cx, cy = K[0, 2], K[1, 2]
		xx = (uu - cx) * zz / fx
		yy = (vv - cy) * zz / fy
		pts = np.stack([xx, yy, zz], axis=-1)[valid].astype(np.float32)
		if len(pts) < min_inliers:
			return T_orig, False, {"reason": f"too few 3D points ({len(pts)})"}
		# --- RANSAC 平面フィット ---
		best_n = None; best_d = None; best_inliers = 0
		rng = np.random.default_rng(seed=42)
		for _ in range(n_ransac_iter):
			idx = rng.choice(len(pts), 3, replace=False)
			p1, p2, p3 = pts[idx]
			v1 = p2 - p1; v2 = p3 - p1
			n = np.cross(v1, v2)
			norm = np.linalg.norm(n)
			if norm < 1e-6:
				continue
			n = n / norm
			d = -float(n @ p1)
			dists = np.abs(pts @ n + d)
			inliers = int((dists < ransac_thresh_mm).sum())
			if inliers > best_inliers:
				best_inliers = inliers; best_n = n; best_d = d
		if best_n is None or best_inliers < min_inliers:
			return T_orig, False, {"reason": f"RANSAC failed (inliers={best_inliers})"}
		# --- 最小二乗リファイン (inlier 集合) ---
		dists = np.abs(pts @ best_n + best_d)
		inlier_pts = pts[dists < ransac_thresh_mm]
		centroid = inlier_pts.mean(axis=0)
		try:
			_, _, Vt = np.linalg.svd(inlier_pts - centroid, full_matrices=False)
		except np.linalg.LinAlgError:
			return T_orig, False, {"reason": "SVD failed"}
		n_refined = Vt[2]
		nn = float(np.linalg.norm(n_refined))
		if nn < 1e-9:
			return T_orig, False, {"reason": "degenerate normal"}
		n_refined = n_refined / nn
		d_refined = -float(n_refined @ centroid)
		# T_orig の Z軸 (マーカー法線) と向きを合わせる
		R_orig = T_orig[:3, :3].astype(np.float64)
		t_orig = T_orig[:3, 3].astype(np.float64)
		z_marker = R_orig[:, 2]
		if float(n_refined @ z_marker) < 0:
			n_refined = -n_refined
			d_refined = -d_refined
		# (1) 位置補正: t_orig を平面上に投影
		dist_to_plane = float(n_refined @ t_orig + d_refined)
		t_new = t_orig - dist_to_plane * n_refined
		# (2) 姿勢補正: マーカー Z軸を平面法線に一致させる回転
		axis = np.cross(z_marker, n_refined)
		axis_norm = float(np.linalg.norm(axis))
		delta_angle_deg = 0.0
		if axis_norm < 1e-9:
			R_align = np.eye(3)
		else:
			axis = axis / axis_norm
			cos_a = float(np.clip(z_marker @ n_refined, -1.0, 1.0))
			angle = float(np.arccos(cos_a))
			delta_angle_deg = float(np.degrees(angle))
			K_mat = np.array([[0, -axis[2], axis[1]],
			                   [axis[2], 0, -axis[0]],
			                   [-axis[1], axis[0], 0]], dtype=np.float64)
			R_align = np.eye(3) + np.sin(angle) * K_mat + (1 - np.cos(angle)) * (K_mat @ K_mat)
		R_new = R_align @ R_orig
		# 直交化 (数値誤差対策)
		U, _, Vt2 = np.linalg.svd(R_new)
		R_new = U @ Vt2
		if float(np.linalg.det(R_new)) < 0:
			R_new = U @ np.diag([1.0, 1.0, -1.0]) @ Vt2
		T_new = np.eye(4)
		T_new[:3, :3] = R_new
		T_new[:3, 3] = t_new
		return T_new, True, {
			"inliers": int(len(inlier_pts)),
			"total": int(len(pts)),
			"delta_t_mm": float(abs(dist_to_plane)),
			"delta_angle_deg": delta_angle_deg,
		}

	def _ankle_detect_markers_in_frame(self, gray, depth_arr, K, dist,
	                                   obj_pts, marker_size_mm,
	                                   depth_scale_mm, target_ids, detector, dictionary, params, use_new_api):
		"""1フレームでArUcoを検出し、target_ids に含まれるIDだけ姿勢を返す。

		Returns dict {aruco_id: {"pose": 4x4, "reproj_err": float}}
		"""
		import cv2
		import numpy as np
		if use_new_api and detector is not None:
			corners, ids, _ = detector.detectMarkers(gray)
		else:
			corners, ids, _ = cv2.aruco.detectMarkers(gray, dictionary, parameters=params)
		out = {}
		if ids is None:
			return out
		for corner, mid_arr in zip(corners, ids):
			mid = int(mid_arr[0] if hasattr(mid_arr, '__len__') else mid_arr)
			if mid not in target_ids:
				continue
			img_pts = corner.reshape(-1, 2).astype(np.float64)
			obj_pts_f64 = np.asarray(obj_pts, dtype=np.float64)
			# solvePnPGeneric + IPPE で平面マーカーの両姿勢解を取得
			# 注: cv2 5.0.0 の SOLVEPNP_IPPE_SQUARE には結果が不正になる不具合がある。
			# 汎用の SOLVEPNP_IPPE は同じく2解を返し、正常動作する。
			try:
				retval, rvecs, tvecs, reproj = cv2.solvePnPGeneric(
					obj_pts_f64, img_pts, K, dist, flags=cv2.SOLVEPNP_IPPE)
				num_sols = int(retval) if retval else len(rvecs)
			except Exception:
				ok, rvec, tvec = cv2.solvePnP(obj_pts_f64, img_pts, K, dist, flags=cv2.SOLVEPNP_ITERATIVE)
				if not ok:
					continue
				rvecs = [rvec]; tvecs = [tvec]; reproj = np.zeros((1, 1), dtype=np.float32); num_sols = 1
			# 深度で表裏を選ぶ
			best_idx = 0
			best_score = float("inf")
			for si in range(num_sols):
				R_mat, _ = cv2.Rodrigues(rvecs[si])
				tv = np.asarray(tvecs[si], dtype=np.float64).flatten()
				corners_cam = (R_mat @ obj_pts.T).T + tv   # (4,3) mm
				pred_depth = corners_cam[:, 2]              # mm
				observed = []
				H, W = depth_arr.shape[:2]
				for pt in img_pts:
					u, v = int(round(pt[0])), int(round(pt[1]))
					if 0 <= u < W and 0 <= v < H:
						d_raw = float(depth_arr[v, u])
						if d_raw > 0:
							observed.append(d_raw * depth_scale_mm)
							continue
					observed.append(np.nan)
				observed = np.array(observed)
				mask = ~np.isnan(observed)
				if mask.sum() >= 2:
					score = float(np.sqrt(np.mean((pred_depth[mask] - observed[mask]) ** 2)))
				else:
					score = float(reproj[si].item()) if hasattr(reproj[si], 'item') else float(reproj[si])
				if score < best_score:
					best_score = score
					best_idx = si
			R_mat, _ = cv2.Rodrigues(rvecs[best_idx])
			tv = np.asarray(tvecs[best_idx], dtype=np.float64).flatten()
			T = np.eye(4)
			T[:3, :3] = R_mat
			T[:3, 3] = tv
			try:
				err_val = float(reproj[best_idx].item())
			except Exception:
				err_val = float(reproj[best_idx])
			pnp_Z = float(tv[2])
			# solvePnP 素の結果。深度は既に「IPPEの表裏どちらを採るか」の判定に使われており、
			# これは深度の粗い使い方なのでノイズに強い。
			T_pnp_only = T.copy()

			# 実測 (2026-08-29) で RGB のみが融合より 7〜14倍 安定だったため既定は "rgb"。
			# 表示ラベル ("rgb  ★推奨 (最も安定)" 等) から手法名だけを取り出す
			try:
				method = str(self.ankle_pose_method.get()).strip().split()[0]
			except Exception:
				method = "rgb"
			if method not in ("rgb", "fusion", "depth-corners"):
				method = "rgb"
			# 冒頭の一定フレームだけ 3手法を計算して比較レポートに回す
			try:
				cmp_limit = int(self.ankle_compare_frames.get())
			except Exception:
				cmp_limit = 300
			store = getattr(self, "_ankle_method_compare", None) or {}
			done = len(store.get(int(mid), {}).get("pnp", [])) if store else 0
			comparing = (cmp_limit > 0 and done < cmp_limit)

			need_fusion = (method == "fusion") or comparing
			need_corners = (method == "depth-corners") or comparing

			T_fuse, ok_fuse, info_fuse = None, False, {}
			if need_fusion:
				try:
					_expand = float(self.ankle_plane_expand.get())
				except Exception:
					_expand = 2.2
				T_fuse, ok_fuse, info_fuse = self._ankle_pose_from_rgb_depth_fusion(
					img_pts, obj_pts, depth_arr, K, dist, depth_scale_mm, expand=_expand)
			use_fusion = (ok_fuse and info_fuse.get("fit_rmse_mm", 999.0) < 3.0)

			T_depth, ok_depth, info_depth = None, False, {}
			depth_Z_median = 0.0
			if need_corners:
				T_depth, ok_depth, info_depth = self._ankle_pose_from_depth_corners(
					img_pts, obj_pts, depth_arr, K, depth_scale_mm)
				try:
					valid_ds = [d for d in info_depth.get("corner_depths_mm", []) if d > 0]
					depth_Z_median = float(np.median(valid_ds)) if valid_ds else 0.0
				except Exception:
					pass

			ok_corners = (ok_depth and info_depth.get("rmse_mm", 999.0) < 5.0
			              and info_depth.get("n_valid_corners", 0) >= 3)

			if method == "rgb":
				# RGB のみ。深度は表裏判定と深度スケール検証にのみ使う。
				pose_source = "rgb-solvePnP"
			elif method == "depth-corners" and ok_corners:
				T = T_depth
				pose_source = "depth-corners"
			elif method == "fusion" and use_fusion:
				T = T_fuse
				pose_source = "rgbd-fusion"
			elif use_fusion:
				T = T_fuse
				pose_source = "rgbd-fusion"
			elif ok_corners:
				T = T_depth
				pose_source = "depth-corners"
			else:
				pose_source = "rgb-solvePnP"

			# 採用手法の集計 (実行後にまとめて報告する)
			if not hasattr(self, "_ankle_pose_source_counts"):
				self._ankle_pose_source_counts = {}
			self._ankle_pose_source_counts[pose_source] = \
				self._ankle_pose_source_counts.get(pose_source, 0) + 1

			# 診断ログ (ID ごとに最初のフレームだけ)
			if not hasattr(self, "_ankle_depth_diag_shown"):
				self._ankle_depth_diag_shown = set()
			diag_key = f"{mid}_first"
			if diag_key not in self._ankle_depth_diag_shown:
				self._ankle_depth_diag_shown.add(diag_key)
				if use_fusion:
					meas = info_fuse.get("measured_marker_mm")
					rad = info_fuse.get("plane_radius_mm")
					rms_p = info_fuse.get("plane_rms_mm") or 0.0
					self._ankle_safe_print(
						f"[ArUco診断 ID={mid}] 採用=rgbd-fusion  "
						f"平面フィット点数={info_fuse.get('n_plane_points')} "
						f"(inlier {100.0 * (info_fuse.get('inlier_ratio') or 0):.0f}%"
						f"{', 台座まで拡張' if info_fuse.get('expanded') else ', マーカー内部のみ'}), "
						f"平面平坦度RMS={rms_p:.3f}mm, "
						f"角フィットRMSE={info_fuse.get('fit_rmse_mm', float('nan')):.3f}mm, "
						f"距離={float(T[2, 3]):.1f}mm (solvePnP {pnp_Z:.1f}mm)")
					if rad:
						# 深度ノイズは空間相関があるので √N では稼げない。法線精度 ≈ σ/r。
						est_deg = np.degrees(np.arctan2(rms_p, max(rad, 1e-6)))
						self._ankle_safe_print(
							f"    平面フィット半径 = {rad:.1f} mm "
							f"→ 法線精度の目安 {est_deg:.2f}° (= 平坦度RMS ÷ 半径)。"
							f"  半径を大きくするほど回転が安定します")
					if meas:
						diff_pct = 100.0 * (meas - float(marker_size_mm)) / max(float(marker_size_mm), 1e-6)
						note = ""
						if abs(diff_pct) > 2.0:
							note = ("  [注意] ②の設定値と 2% 以上ずれています。"
							        "印刷スケールを確認し、実測値を②に入れてください")
						self._ankle_safe_print(
							f"    実測マーカー寸法 = {meas:.2f} mm "
							f"(②の設定 {float(marker_size_mm):.2f} mm, 差 {diff_pct:+.1f}%), "
							f"辺長ばらつき {info_fuse.get('marker_edge_spread_mm', float('nan')):.2f} mm{note}")
				else:
					extra = ""
					if need_fusion:
						extra = f" (rgbd-fusion 不可: {info_fuse.get('reason')})"
					self._ankle_safe_print(
						f"[ArUco診断 ID={mid}] 採用={pose_source}{extra}  "
						f"距離={float(T[2, 3]):.1f}mm, 再投影誤差={err_val:.3f}px")
					meas = info_fuse.get("measured_marker_mm")
					if meas:
						diff_pct = 100.0 * (meas - float(marker_size_mm)) / max(float(marker_size_mm), 1e-6)
						note = ""
						if abs(diff_pct) > 2.0:
							note = ("  [注意] ②の設定値と 2% 以上ずれています。"
							        "RGBのみの手法では距離が直接この比率でずれるため、"
							        "実測値を②に入れてください")
						self._ankle_safe_print(
							f"    実測マーカー寸法 (深度から) = {meas:.2f} mm "
							f"(②の設定 {float(marker_size_mm):.2f} mm, 差 {diff_pct:+.1f}%){note}")

			out[mid] = {"pose": T, "reproj_err": err_val, "depth_score": best_score,
			             "pose_source": pose_source,
			             "depth_Z_median": depth_Z_median,
			             "pnp_Z": pnp_Z,
			             "plane_rms_mm": info_fuse.get("plane_rms_mm"),
			             "fit_rmse_mm": info_fuse.get("fit_rmse_mm"),
			             # --- 手法比較用 (どれが実データで一番安定かを後で実測する) ---
			             "pose_pnp": T_pnp_only,
			             "pose_corners": (T_depth if ok_depth else None),
			             "pose_fusion": (T_fuse if ok_fuse else None)}
		return out

	def _ankle_detect_from_bag(self, bag_path: str, aruco_dict_name: str,
	                            marker_size_mm: float, target_ids: set,
	                            stride: int, progress_cb, cancel_check) -> dict:
		"""RealSense .bag から色/深度を読み、ArUco姿勢時系列を返す。"""
		import cv2
		import numpy as np
		import pyrealsense2 as rs
		obj_pts = self._ankle_marker_obj_points(marker_size_mm)
		detector, dictionary, params, use_new_api = self._ankle_make_detector(aruco_dict_name)

		pipeline = rs.pipeline()
		config = rs.config()
		rs.config.enable_device_from_file(config, bag_path, repeat_playback=False)
		config.enable_stream(rs.stream.color)
		config.enable_stream(rs.stream.depth)
		profile = pipeline.start(config)
		try:
			playback = profile.get_device().as_playback()
			playback.set_real_time(False)
		except Exception:
			playback = None

		color_profile = profile.get_stream(rs.stream.color)
		vsp = color_profile.as_video_stream_profile()
		intr = vsp.get_intrinsics()
		K = np.array([[intr.fx, 0, intr.ppx], [0, intr.fy, intr.ppy], [0, 0, 1]], dtype=np.float64)
		dist = np.array(intr.coeffs, dtype=np.float64)
		depth_sensor = profile.get_device().first_depth_sensor()
		depth_scale = float(depth_sensor.get_depth_scale())
		depth_scale_mm = depth_scale * 1000.0

		# 総フレーム数の推定 (再生時間 × fps / stride)。progress 表示のため。
		estimated_total = 0
		try:
			bag_fps = float(vsp.fps())
			duration_ns = playback.get_duration().total_seconds() if playback is not None else 0.0
			# .total_seconds() は timedelta のメソッド。pyrealsense2 の get_duration() は
			# timedelta 相当を返すことが多いが、環境により int (ns) の場合もある。
			if hasattr(duration_ns, 'total_seconds'):
				duration_ns = duration_ns.total_seconds()
			duration_s = float(duration_ns) if duration_ns else 0.0
			if bag_fps > 0 and duration_s > 0:
				estimated_total = max(1, int(bag_fps * duration_s / max(stride, 1)))
				print(f"[ankle detect] 総フレーム数を推定: {estimated_total} (duration={duration_s:.2f}s × fps={bag_fps:.1f} / stride={stride})")
		except Exception as e:
			print(f"[ankle detect] 総フレーム数推定失敗 (進捗率は表示されません): {e}")
			estimated_total = 0

		# --- Color / Depth の解像度と intrinsics を出力 (アライメント診断) ---
		try:
			depth_profile = profile.get_stream(rs.stream.depth).as_video_stream_profile()
			d_intr = depth_profile.get_intrinsics()
			print(f"[ankle detect] Color intrinsics: fx={intr.fx:.1f}, fy={intr.fy:.1f}, "
			      f"cx={intr.ppx:.1f}, cy={intr.ppy:.1f}, {intr.width}x{intr.height}")
			print(f"[ankle detect] Depth intrinsics: fx={d_intr.fx:.1f}, fy={d_intr.fy:.1f}, "
			      f"cx={d_intr.ppx:.1f}, cy={d_intr.ppy:.1f}, {d_intr.width}x{d_intr.height}")
			print(f"[ankle detect] Depth scale: {depth_scale} m/unit = {depth_scale_mm} mm/unit")
			if abs(intr.width - d_intr.width) > 0 or abs(intr.height - d_intr.height) > 0:
				print(f"[ankle detect] ⚠️ Color と Depth の解像度差 → align 必須 "
				      f"(align.process() で depth を color 系にリサンプル予定)")
		except Exception as e:
			print(f"[ankle detect] intrinsics 診断失敗: {e}")

		# 推定値を progress_cb に通知 (最初の update 呼び出しで反映される)
		if progress_cb and estimated_total > 0:
			try:
				progress_cb(0, 0, 0.0, id_counts_running={},
				            total_frames_known=estimated_total,
				            extra_msg=f"総フレーム数 (推定) = {estimated_total} — 検出処理開始")
			except TypeError:
				pass

		align = rs.align(rs.stream.color)

		poses = {aid: [] for aid in target_ids}
		detected = {aid: [] for aid in target_ids}
		reproj = {aid: [] for aid in target_ids}
		timestamps = []
		frame_idx = 0
		processed = 0
		try:
			while True:
				if cancel_check and cancel_check():
					break
				try:
					frames = pipeline.wait_for_frames(timeout_ms=2000)
				except RuntimeError:
					break  # EOF
				aligned = align.process(frames)
				color = aligned.get_color_frame()
				depth = aligned.get_depth_frame()
				if not color or not depth:
					break
				if frame_idx % max(stride, 1) != 0:
					frame_idx += 1
					continue
				rgb = np.asanyarray(color.get_data())
				depth_arr = np.asanyarray(depth.get_data())
				timestamps.append(color.get_timestamp() / 1000.0)
				gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY) if rgb.ndim == 3 else rgb

				# --- 【最初のフレームで詳細診断】depth アライメントと scale を検証 ---
				if processed == 0:
					# depth_scale の自動補正 (SDK報告値が実データと乖離する場合)
					try:
						auto_fix = bool(self.ankle_depth_scale_autofix.get())
					except Exception:
						auto_fix = True
					if auto_fix:
						depth_scale_mm, _dsinfo = self._ankle_auto_correct_depth_scale(
							gray, depth_arr, K, dist, obj_pts, target_ids,
							detector, dictionary, params, use_new_api, depth_scale_mm)
					try:
						print(f"[深度診断] color shape: {rgb.shape}, depth shape: {depth_arr.shape}")
						print(f"[深度診断] 使用する depth_scale_mm = {depth_scale_mm:.6f} mm/unit "
						      f"(SDK報告値: {depth_scale * 1000.0:.6f})")
						# 全体の depth 統計 (補正後のスケールで)
						valid_mask = depth_arr > 0
						if valid_mask.any():
							valid_d = depth_arr[valid_mask].astype(np.float32) * depth_scale_mm
							print(f"[深度診断] depth 全体 (有効値, mm): "
							      f"min={float(valid_d.min()):.1f}, "
							      f"median={float(np.median(valid_d)):.1f}, "
							      f"max={float(valid_d.max()):.1f}, "
							      f"有効率={100.0 * valid_mask.sum() / valid_mask.size:.1f}%")
						# 画像中心の depth を参考出力
						h, w = depth_arr.shape[:2]
						cd_raw = int(depth_arr[h // 2, w // 2])
						print(f"[深度診断] 画像中心 depth: raw={cd_raw}, mm={cd_raw * depth_scale_mm:.1f}")
						# アライメント問題の警告
						if rgb.shape[:2] != depth_arr.shape[:2]:
							print(f"[深度診断] ⚠️ color と depth の解像度が不一致! "
							      f"align.process() が正しく機能していない可能性")
						# デバッグ画像出力 (color, colored depth) — 日本語パス対応の _ankle_imwrite を使用
						try:
							debug_dir = Path(bag_path).parent / "debug_frame0"
							debug_dir.mkdir(parents=True, exist_ok=True)
							saved = []
							# color
							color_bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR) if rgb.ndim == 3 else rgb
							if self._ankle_imwrite(debug_dir / "color.png", color_bgr):
								saved.append("color.png")
							# depth colormap (mm 単位で正規化 → JET)
							depth_mm_img = depth_arr.astype(np.float32) * depth_scale_mm
							# 100mm〜1000mm を 0〜255 にマッピング (D405 動作範囲を強調)
							depth_vis = np.clip((depth_mm_img - 100.0) / 900.0 * 255.0, 0, 255).astype(np.uint8)
							depth_vis[depth_arr == 0] = 0  # 無効ピクセル黒
							depth_color = cv2.applyColorMap(depth_vis, cv2.COLORMAP_JET)
							if self._ankle_imwrite(debug_dir / "depth_colored.png", depth_color):
								saved.append("depth_colored.png")
							# ArUco 検出結果を color に重ね、各マーカーの深度値を注記した画像
							try:
								overlay = (cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR).copy() if rgb.ndim == 3
								           else cv2.cvtColor(rgb, cv2.COLOR_GRAY2BGR))
								corners_det, ids_det, _ = (detector.detectMarkers(gray) if use_new_api
								                            else cv2.aruco.detectMarkers(gray, dictionary, parameters=params))
								if ids_det is not None:
									cv2.aruco.drawDetectedMarkers(overlay, corners_det, ids_det)
									# 各マーカー中心に深度値を描画 (アライメント確認用)
									for c_det, i_det in zip(corners_det, ids_det):
										pts = c_det.reshape(-1, 2)
										cu, cv_ = int(round(pts[:, 0].mean())), int(round(pts[:, 1].mean()))
										if 0 <= cu < depth_arr.shape[1] and 0 <= cv_ < depth_arr.shape[0]:
											dmm = float(depth_arr[cv_, cu]) * depth_scale_mm
											cv2.putText(overlay, f"ID{int(i_det[0])}: {dmm:.0f}mm",
											            (cu - 40, cv_ - 12), cv2.FONT_HERSHEY_SIMPLEX,
											            0.5, (0, 0, 255), 2)
									if self._ankle_imwrite(debug_dir / "aruco_overlay.png", overlay):
										saved.append("aruco_overlay.png")
							except Exception:
								pass
							if saved:
								print(f"[深度診断] デバッグ画像を保存: {debug_dir}")
								print(f"  → {' / '.join(saved)} を確認してください")
							else:
								print(f"[深度診断] ⚠️ デバッグ画像を1つも保存できませんでした: {debug_dir}")
						except Exception as e_img:
							print(f"[深度診断] デバッグ画像保存失敗: {e_img}")
					except Exception as e:
						print(f"[深度診断] 診断出力失敗: {e}")

				det = self._ankle_detect_markers_in_frame(
					gray, depth_arr, K, dist, obj_pts, marker_size_mm,
					depth_scale_mm, target_ids, detector, dictionary, params, use_new_api)
				for aid in target_ids:
					if aid in det:
						poses[aid].append(det[aid]["pose"])
						detected[aid].append(True)
						reproj[aid].append(det[aid]["reproj_err"])
					else:
						poses[aid].append(np.full((4, 4), np.nan))
						detected[aid].append(False)
						reproj[aid].append(float("nan"))
				self._ankle_accumulate_method_compare(det, target_ids)
				processed += 1
				frame_idx += 1
				# 進捗更新 (5フレームに1回でUI負荷を抑える。running カウントを渡す)
				if progress_cb and (processed % 5 == 0 or processed == 1):
					id_counts_running = {aid: int(sum(detected[aid])) for aid in target_ids}
					try:
						progress_cb(processed, frame_idx,
						            timestamps[-1] if timestamps else 0.0,
						            id_counts_running=id_counts_running)
					except TypeError:
						# 旧シグネチャ互換
						progress_cb(processed, frame_idx,
						            timestamps[-1] if timestamps else 0.0)
		finally:
			try:
				pipeline.stop()
			except Exception:
				pass

		return {
			"frame_count": len(timestamps),
			"timestamps": np.array(timestamps, dtype=np.float64),
			"intrinsics": {"fx": float(intr.fx), "fy": float(intr.fy),
			                "cx": float(intr.ppx), "cy": float(intr.ppy),
			                "width": int(intr.width), "height": int(intr.height),
			                "dist": [float(x) for x in intr.coeffs],
			                "depth_scale_mm": float(depth_scale_mm)},
			"marker_size_mm": float(marker_size_mm),
			"aruco_dict": aruco_dict_name,
			"source": str(bag_path),
			"bones": {aid: {"poses": np.array(poses[aid]),
			                 "detected": np.array(detected[aid]),
			                 "reproj_err": np.array(reproj[aid])}
			           for aid in target_ids},
		}

	def _ankle_load_intrinsics_json(self, path: str) -> tuple:
		"""JSON/YAML から (K, dist, depth_scale_mm) を読み込む。
		期待キー: fx, fy, cx, cy, dist(list), depth_scale_mm (mm/unit)"""
		import numpy as np
		p = Path(path)
		ext = p.suffix.lower()
		with p.open("r", encoding="utf-8") as f:
			if ext in (".yaml", ".yml"):
				try:
					import yaml
					data = yaml.safe_load(f)
				except ImportError:
					raise ValueError("YAML読込には pyyaml が必要です。")
			else:
				data = json.load(f)
		K = np.array([[float(data["fx"]), 0, float(data["cx"])],
		              [0, float(data["fy"]), float(data["cy"])],
		              [0, 0, 1]], dtype=np.float64)
		dist = np.array([float(x) for x in data.get("dist", [0, 0, 0, 0, 0])], dtype=np.float64)
		depth_scale_mm = float(data.get("depth_scale_mm", 1.0))  # デフォルト: raw値=mm
		return K, dist, depth_scale_mm

	def _ankle_detect_from_video(self, video_path: str, depth_path: str, intrinsics_path: str,
	                              aruco_dict_name: str, marker_size_mm: float, target_ids: set,
	                              stride: int, progress_cb, cancel_check) -> dict:
		"""mp4/avi + 深度npz + 内部パラメータjson から検出する汎用パス。

		深度npz は 'frames' キーで (N, H, W) 配列(uint16 raw or float32 mm)。
		"""
		import cv2
		import numpy as np
		K, dist, depth_scale_mm = self._ankle_load_intrinsics_json(intrinsics_path)
		obj_pts = self._ankle_marker_obj_points(marker_size_mm)
		detector, dictionary, params, use_new_api = self._ankle_make_detector(aruco_dict_name)

		cap = cv2.VideoCapture(video_path)
		if not cap.isOpened():
			raise ValueError(f"動画を開けません: {video_path}")
		fps = float(cap.get(cv2.CAP_PROP_FPS) or 30.0)

		depth_npz = np.load(depth_path)
		if "frames" not in depth_npz.files:
			raise ValueError(f"npzに 'frames' キーがありません: {depth_path}")
		depth_frames = depth_npz["frames"]

		poses = {aid: [] for aid in target_ids}
		detected = {aid: [] for aid in target_ids}
		reproj = {aid: [] for aid in target_ids}
		timestamps = []
		frame_idx = 0
		processed = 0
		while True:
			if cancel_check and cancel_check():
				break
			ok, rgb = cap.read()
			if not ok:
				break
			if frame_idx >= len(depth_frames):
				break
			if frame_idx % max(stride, 1) != 0:
				frame_idx += 1
				continue
			depth_arr = depth_frames[frame_idx]
			gray = cv2.cvtColor(rgb, cv2.COLOR_BGR2GRAY)
			timestamps.append(frame_idx / fps)
			# 最初のフレームで depth_scale の自動補正を試みる
			if processed == 0:
				try:
					auto_fix_v = bool(self.ankle_depth_scale_autofix.get())
				except Exception:
					auto_fix_v = True
				if auto_fix_v:
					depth_scale_mm, _ = self._ankle_auto_correct_depth_scale(
						gray, depth_arr, K, dist, obj_pts, target_ids,
						detector, dictionary, params, use_new_api, depth_scale_mm)
			det = self._ankle_detect_markers_in_frame(
				gray, depth_arr, K, dist, obj_pts, marker_size_mm,
				depth_scale_mm, target_ids, detector, dictionary, params, use_new_api)
			for aid in target_ids:
				if aid in det:
					poses[aid].append(det[aid]["pose"])
					detected[aid].append(True)
					reproj[aid].append(det[aid]["reproj_err"])
				else:
					poses[aid].append(np.full((4, 4), np.nan))
					detected[aid].append(False)
					reproj[aid].append(float("nan"))
			self._ankle_accumulate_method_compare(det, target_ids)
			processed += 1
			frame_idx += 1
			if progress_cb and (processed % 5 == 0 or processed == 1):
				id_counts_running = {aid: int(sum(detected[aid])) for aid in target_ids}
				try:
					progress_cb(processed, frame_idx, timestamps[-1],
					            id_counts_running=id_counts_running)
				except TypeError:
					progress_cb(processed, frame_idx, timestamps[-1])
		cap.release()

		return {
			"frame_count": len(timestamps),
			"timestamps": np.array(timestamps, dtype=np.float64),
			"intrinsics": {"fx": float(K[0, 0]), "fy": float(K[1, 1]),
			                "cx": float(K[0, 2]), "cy": float(K[1, 2]),
			                "width": int(rgb.shape[1]) if rgb is not None else 0,
			                "height": int(rgb.shape[0]) if rgb is not None else 0,
			                "dist": [float(x) for x in dist],
			                "depth_scale_mm": float(depth_scale_mm)},
			"marker_size_mm": float(marker_size_mm),
			"aruco_dict": aruco_dict_name,
			"source": str(video_path),
			"bones": {aid: {"poses": np.array(poses[aid]),
			                 "detected": np.array(detected[aid]),
			                 "reproj_err": np.array(reproj[aid])}
			           for aid in target_ids},
		}

	# ---- 進捗ダイアログ ----
	def _ankle_open_progress(self, title: str, total_frames: int = 0):
		"""検出用の進捗ダイアログを開き、更新用closureを返す (リッチ表示版)。

		Args:
			title: ウィンドウタイトル
			total_frames: 総フレーム数 (既知の場合)。0 なら indeterminate mode。

		Returns:
			(update, close, cancel_check) の3タプル
			update(processed, frame_idx, t_sec, detected_ids=None, extra_msg=None):
			    processed: これまでに処理したフレーム数
			    frame_idx: 元のフレーム番号 (stride 考慮)
			    t_sec: 再生時刻 [s]
			    detected_ids: このフレームで検出された ID のリスト (per-ID カウント用)
			    extra_msg: 追加ステップ表示 (例 "深度平面フィット中")
		"""
		import time as _time
		win = tk.Toplevel(self)
		win.title(title)
		win.transient(self)
		# 【重要】geometry でサイズを明示指定 (Windows で表示されない問題を回避)
		x = self.winfo_rootx() + 80
		y = self.winfo_rooty() + 80
		win.geometry(f"560x340+{x}+{y}")
		win.resizable(True, True)
		win.minsize(500, 300)
		try:
			win.attributes('-topmost', True)
		except Exception:
			pass

		# タイトル
		ttk.Label(win, text=title,
		          font=(self.ui_font_family, 11, "bold")).pack(padx=12, pady=(10, 6))

		# メインステータス
		status_var = tk.StringVar(value="開始中… (処理準備)")
		ttk.Label(win, textvariable=status_var, anchor="w",
		          font=(self.ui_font_family, 10, "bold")
		          ).pack(padx=12, pady=(2, 4), fill=tk.X)

		# 進捗バー
		mode = "determinate" if total_frames > 0 else "indeterminate"
		pbar = ttk.Progressbar(win, mode=mode, length=520,
		                        maximum=max(total_frames, 100))
		pbar.pack(padx=12, pady=(4, 6))
		if mode == "indeterminate":
			pbar.start(80)

		percent_var = tk.StringVar(value="")
		ttk.Label(win, textvariable=percent_var,
		          font=(self.ui_font_family, 9),
		          foreground="#005580").pack(padx=12, pady=(0, 4))

		# 詳細情報 (経過時間, 速度, ETA)
		detail_var = tk.StringVar(value="")
		ttk.Label(win, textvariable=detail_var, anchor="w", justify=tk.LEFT,
		          font=(self.ui_font_family, 9),
		          foreground="#333333"
		          ).pack(padx=12, pady=(2, 4), fill=tk.X)

		# 現在の処理内容
		step_var = tk.StringVar(value="待機中")
		ttk.Label(win, text="■ 現在:", anchor="w", justify=tk.LEFT,
		          font=(self.ui_font_family, 9, "bold")
		          ).pack(padx=12, pady=(6, 0), anchor="w")
		ttk.Label(win, textvariable=step_var, anchor="w", justify=tk.LEFT,
		          font=(self.ui_font_family, 9), foreground="#005580", wraplength=520
		          ).pack(padx=20, pady=(0, 4), fill=tk.X)

		# 検出済 ID 別カウント
		per_id_var = tk.StringVar(value="(検出待機中…)")
		ttk.Label(win, text="■ マーカー検出状況:", anchor="w", justify=tk.LEFT,
		          font=(self.ui_font_family, 9, "bold")
		          ).pack(padx=12, pady=(6, 0), anchor="w")
		ttk.Label(win, textvariable=per_id_var, anchor="w", justify=tk.LEFT,
		          font=("Courier New", 9)
		          ).pack(padx=20, pady=(0, 4), fill=tk.X)

		self._ankle_detect_cancel = False
		start_time = _time.time()
		id_counts = {}       # id -> 検出できたフレーム数
		id_seen = set()      # このフレームで検出済みの id (重複防止用に呼び側で clear)

		def _cancel():
			self._ankle_detect_cancel = True
			status_var.set("[キャンセル要求] 完了を待機中…")

		ttk.Button(win, text="キャンセル", command=_cancel, width=15).pack(padx=12, pady=(8, 10))
		win.protocol("WM_DELETE_WINDOW", _cancel)

		# 初期レイアウトを強制 (widget が表示されない問題を回避)
		try:
			win.update_idletasks()
			win.update()
		except Exception:
			pass

		# 総フレーム数は後から更新可能 (bag ファイルは開いてみないと分からない)
		total_frames_ref = [int(total_frames)]

		def update(processed, frame_idx, t_sec, detected_ids=None, extra_msg=None,
		           id_counts_running=None, total_frames_known=None):
			"""進捗を更新。

			id_counts_running が渡されれば直接そのカウントを表示 (推奨・最新機能)。
			total_frames_known: 途中で総フレーム数が判明した場合、渡すと determinate に切替
			detected_ids はレガシー用 (呼び時にのみ +1 加算、5フレーム毎呼びだと過小カウント)。
			"""
			# 総フレーム数の後追い更新
			if total_frames_known and total_frames_known > total_frames_ref[0]:
				total_frames_ref[0] = int(total_frames_known)
				try:
					pbar.stop()
					pbar.configure(mode="determinate", maximum=int(total_frames_known))
				except Exception:
					pass
			total = total_frames_ref[0]
			elapsed = _time.time() - start_time
			fps = processed / max(elapsed, 1e-3)
			# メインステータス
			if total > 0:
				pct = 100.0 * processed / max(total, 1)
				status_var.set(f"[処理中] フレーム {processed} / {total} 相当 ({pct:.1f}%)")
				percent_var.set(f"{pct:.1f}% 完了")
				try:
					pbar['value'] = min(processed, total)
				except Exception:
					pass
			else:
				status_var.set(f"[処理中] フレーム {processed} を処理済")
				percent_var.set("(総フレーム数は未確定 — 進捗率は表示できません)")
			# 詳細
			eta_str = ""
			if total > 0 and processed > 0 and fps > 0:
				eta = max(0.0, (total - processed) / fps)
				eta_str = f"   残り予想: {int(eta // 60):02d}:{int(eta % 60):02d}"
			detail_var.set(
				f"経過時間: {int(elapsed // 60):02d}:{int(elapsed % 60):02d}"
				f"   処理速度: {fps:.1f} frames/s"
				f"   再生時刻: t≈{t_sec:.2f}s"
				f"{eta_str}"
			)
			# 現在の処理内容
			if extra_msg:
				step_var.set(extra_msg)
			else:
				step_var.set(f"元フレーム {frame_idx} を処理: ArUco検出 → SUBPIX精緻化 → 深度平面フィット")
			# ID 別カウント
			if id_counts_running is not None:
				# 呼び側の running カウントをそのまま反映 (推奨)
				id_counts.clear()
				id_counts.update({int(k): int(v) for k, v in id_counts_running.items()})
			elif detected_ids is not None:
				# レガシー: +1 加算 (呼び頻度に依存するので不正確)
				for id_val in set(int(i) for i in detected_ids):
					id_counts[id_val] = id_counts.get(id_val, 0) + 1
			if id_counts:
				lines = []
				for id_val in sorted(id_counts.keys()):
					cnt = id_counts[id_val]
					rate = 100.0 * cnt / max(processed, 1)
					lines.append(f"  ID={id_val:>3d}: {cnt:>5d}件 (検出率 {rate:5.1f}%)")
				per_id_var.set("\n".join(lines))
			try:
				win.update_idletasks()
				win.update()
			except Exception:
				pass

		def close():
			try:
				pbar.stop()
				win.destroy()
			except Exception:
				pass

		return update, close, lambda: self._ankle_detect_cancel

	# ---- キャッシュ (per-tab) ヘルパ ----
	def _ankle_current_tab_key(self) -> str:
		if not self._ankle_tabs:
			return ""
		return self._ankle_tabs[self._ankle_active_tab].get("name", "")

	def _ankle_get_current_cache(self):
		return self._ankle_pose_cache.get(self._ankle_current_tab_key())

	def _ankle_set_current_cache(self, cache) -> None:
		key = self._ankle_current_tab_key()
		if not key:
			return
		if cache is None:
			self._ankle_pose_cache.pop(key, None)
		else:
			self._ankle_pose_cache[key] = cache

	def _ankle_cache_status_text(self, cache) -> str:
		if not cache:
			return "(未実行)"
		N = int(cache.get("frame_count", 0))
		bones = cache.get("bones", {}) or {}
		lines = [f"検出完了: {N}フレーム / ソース={Path(str(cache.get('source',''))).name}"]
		for aid, b in sorted(bones.items()):
			det = b.get("detected", None)
			if det is None or len(det) == 0:
				continue
			rate = 100.0 * float(det.sum()) / len(det)
			re = b.get("reproj_err", None)
			re_med = ""
			if re is not None and len(re) > 0:
				import numpy as np
				valid = re[~np.isnan(re)] if hasattr(re, 'dtype') else np.array([r for r in re if r == r])
				if len(valid) > 0:
					re_med = f", 再投影誤差(中央値)={float(np.median(valid)):.2f}px"
			lines.append(f"  ID={aid}: 検出率 {rate:.1f}%{re_med}")
		return "\n".join(lines)

	def _ankle_update_detection_status(self) -> None:
		try:
			self.ankle_detection_status.set(self._ankle_cache_status_text(self._ankle_get_current_cache()))
		except Exception:
			pass

	# ---- UIアクション ----
	def on_ankle_detect_markers(self) -> None:
		"""現在の入力に基づき ArUco検出+PnPを実行する。"""
		if not self._ankle_check_cv2():
			return
		# 診断ログのマーカー ID 履歴をクリア (実行ごとに最初のフレーム診断を出す)
		self._ankle_depth_diag_shown = set()
		self._ankle_pose_source_counts = {}
		self._ankle_method_compare = {}
		# 骨リストからArUco IDを収集
		target_ids = set()
		for b in self.ankle_bones:
			try:
				target_ids.add(int(b.get("aruco_id", -1)))
			except Exception:
				pass
		target_ids.discard(-1)
		if not target_ids:
			messagebox.showwarning("検出", "骨リストが空、またはArUco IDが未設定です。③でIDを設定してください。")
			return
		aruco_dict_name = self.ankle_aruco_dict_var.get()
		try:
			marker_size_mm = float(self.ankle_marker_size_mm.get())
		except Exception:
			marker_size_mm = 20.0
		if marker_size_mm <= 0:
			messagebox.showwarning("検出", "マーカー実寸(mm)を正しく入力してください。")
			return
		try:
			stride = max(1, int(self.ankle_detect_stride.get()))
		except Exception:
			stride = 1

		# --- ソース選択ダイアログ (キャンセルなら②の設定を使用) ---
		# .db3/.bag なら 単独ソース、.mp4等なら②の深度+内部パラメータと組み合わせ
		cur_depth = self.ankle_depth_path.get().strip()
		cur_video = self.ankle_video_path.get().strip()
		_initial = cur_depth or cur_video
		if _initial:
			_initdir = str(Path(_initial).parent)
			_initfile = Path(_initial).name
		else:
			_initdir = str(Path(__file__).parent / "cache")
			_initfile = ""
		picked = filedialog.askopenfilename(
			title="ArUco検出のソースを選択 (キャンセルで②の設定を使用)",
			initialdir=_initdir,
			initialfile=_initfile,
			filetypes=[
				("すべての対応形式", "*.db3 *.bag *.mp4 *.avi *.mov *.mkv"),
				("RealSense録画 (単独ソース)", "*.db3 *.bag"),
				("動画のみ (深度+内部パラメータは②で指定)", "*.mp4 *.avi *.mov *.mkv"),
				("すべてのファイル", "*.*"),
			])
		if picked:
			_ext = Path(picked).suffix.lower()
			if _ext in (".db3", ".bag"):
				# RealSense録画 → 深度パスにセット (色/深度/内部パラメータ全部内包)
				self.ankle_depth_path.set(picked)
				print(f"[ankle detect] ソース更新: 深度データ (.db3/.bag) = {picked}")
			elif _ext in (".mp4", ".avi", ".mov", ".mkv"):
				self.ankle_video_path.set(picked)
				print(f"[ankle detect] ソース更新: RGBビデオ = {picked}")
			else:
				# 不明な拡張子: とりあえず深度データに入れる (ユーザー判断)
				self.ankle_depth_path.set(picked)
				print(f"[ankle detect] ソース更新: 深度データ (未知拡張子) = {picked}")

		video = self.ankle_video_path.get().strip()
		depth = self.ankle_depth_path.get().strip()
		intr = self.ankle_camera_intrinsics_path.get().strip()

		# 入力形式の判定 (.bag / .db3 なら pyrealsense2、それ以外は mp4+npz+json)
		use_bag = False
		bag_path = ""
		_RS_EXTS = (".bag", ".db3")
		if depth.lower().endswith(_RS_EXTS):
			use_bag = True; bag_path = depth
		elif video.lower().endswith(_RS_EXTS):
			use_bag = True; bag_path = video

		if use_bag:
			if not self._ankle_check_rs():
				return
			if not Path(bag_path).exists():
				messagebox.showwarning("検出", f"ファイルが見つかりません: {bag_path}"); return
			update_cb, close_cb, cancel_cb = self._ankle_open_progress(f"ArUco検出中 (bag): {Path(bag_path).name}")
			try:
				cache = self._ankle_detect_from_bag(
					bag_path, aruco_dict_name, marker_size_mm, target_ids, stride, update_cb, cancel_cb)
			except Exception as e:
				close_cb()
				messagebox.showerror("検出エラー", f"検出処理でエラーが発生しました:\n{e}")
				return
			close_cb()
		else:
			# 汎用パス: 動画 + 深度npz + 内部パラメータjson
			missing = [n for n, v in (("RGBビデオ", video), ("深度データ", depth), ("内部パラメータ", intr)) if not v]
			if missing:
				messagebox.showwarning("検出", "以下が未指定です: " + ", ".join(missing)); return
			for label, p in (("ビデオ", video), ("深度", depth), ("内部パラメータ", intr)):
				if not Path(p).exists():
					messagebox.showwarning("検出", f"{label}ファイルが見つかりません: {p}"); return
			update_cb, close_cb, cancel_cb = self._ankle_open_progress(f"ArUco検出中: {Path(video).name}")
			try:
				cache = self._ankle_detect_from_video(
					video, depth, intr, aruco_dict_name, marker_size_mm, target_ids, stride, update_cb, cancel_cb)
			except Exception as e:
				close_cb()
				messagebox.showerror("検出エラー", f"検出処理でエラーが発生しました:\n{e}")
				return
			close_cb()

		# 採用した姿勢推定手法の内訳を報告 (品質の目安になる)
		counts = getattr(self, "_ankle_pose_source_counts", {}) or {}
		src_lines = []
		if counts:
			total_det = sum(counts.values())
			self._ankle_safe_print("[姿勢推定手法の内訳]")
			for k, v in sorted(counts.items(), key=lambda kv: -kv[1]):
				pct = 100.0 * v / max(total_det, 1)
				self._ankle_safe_print(f"    {k}: {v} 件 ({pct:.1f}%)")
				src_lines.append(f"  {k}: {pct:.1f}%")
			fusion_pct = 100.0 * counts.get("rgbd-fusion", 0) / max(total_det, 1)
			if fusion_pct < 80.0:
				self._ankle_safe_print(
					"    [注意] rgbd-fusion の採用率が低いです。マーカー面の深度が"
					"欠けている可能性があります (映り込み・距離・角度を確認してください)")

		# キャッシュ保存 + ステータス更新
		self._ankle_set_current_cache(cache)
		self._ankle_autosave_pose_cache()      # 再起動しても引き継げるよう自動保存
		self._ankle_update_detection_status()
		msg = self._ankle_cache_status_text(cache)
		if src_lines:
			msg += "\n\n姿勢推定手法:\n" + "\n".join(src_lines)
		messagebox.showinfo("検出完了", msg)

	# ---- 姿勢時系列 保存/読込 (.npz) ----
	# ---- 姿勢時系列の自動永続化 (再起動しても ④ の結果を引き継ぐ) ----
	def _ankle_pose_cache_dir(self):
		"""姿勢キャッシュの保存先ディレクトリ。無ければ作る。"""
		d = Path(__file__).parent / "cache" / "ankle_pose"
		d.mkdir(parents=True, exist_ok=True)
		return d

	def _ankle_pose_cache_path(self, tab_key: str):
		"""試験タブ名から自動保存ファイルのパスを作る。

		タブ名は日本語や記号を含みうるので、安全な文字だけ残した名前 + 短いハッシュにする。
		(名前だけだと衝突しうるため、元の名前のハッシュを添える)
		"""
		import hashlib
		key = str(tab_key or "default")
		safe = "".join(ch if (ch.isalnum() or ch in "-_") else "_" for ch in key)[:40]
		h = hashlib.md5(key.encode("utf-8")).hexdigest()[:8]
		return self._ankle_pose_cache_dir() / f"pose_{safe}_{h}.npz"

	def _ankle_autosave_pose_cache(self, tab_key: str = None) -> None:
		"""現在タブの姿勢時系列を自動保存する。④の検出完了時と終了時に呼ばれる。"""
		try:
			key = tab_key if tab_key is not None else self._ankle_current_tab_key()
			if not key:
				return
			cache = self._ankle_pose_cache.get(key)
			path = self._ankle_pose_cache_path(key)
			if not cache:
				# キャッシュが消えていれば、古い保存も消して整合を保つ
				if path.exists():
					try:
						path.unlink()
					except Exception:
						pass
				return
			self._ankle_save_pose_cache_npz(str(path), cache)
			print(f"[ankle姿勢 自動保存] {path.name} "
			      f"({int(cache.get('frame_count', 0))}フレーム)")
		except Exception as e:
			print(f"[ankle姿勢 自動保存] 失敗: {e}")

	def _ankle_autosave_all_pose_caches(self) -> None:
		"""全タブぶんを自動保存する (終了時)。"""
		try:
			for key in list(getattr(self, "_ankle_pose_cache", {}).keys()):
				self._ankle_autosave_pose_cache(key)
		except Exception as e:
			print(f"[ankle姿勢 自動保存] 一括保存に失敗: {e}")

	def _ankle_autoload_pose_caches(self) -> None:
		"""起動時に、各試験タブの姿勢時系列を自動復元する。

		復元後に「今の②設定と食い違っていないか」を検査し、違えば警告を出す。
		(マーカー実寸や辞書を変えたのに古い検出結果を使い続ける事故を防ぐ)
		"""
		try:
			tabs = getattr(self, "_ankle_tabs", []) or []
			restored = []
			for tab in tabs:
				key = tab.get("name", "")
				if not key:
					continue
				path = self._ankle_pose_cache_path(key)
				if not path.exists():
					continue
				try:
					cache = self._ankle_load_pose_cache_npz(str(path))
				except Exception as e:
					print(f"[ankle姿勢 自動復元] {path.name} 読込失敗: {e}")
					continue
				self._ankle_pose_cache[key] = cache
				restored.append((key, int(cache.get("frame_count", 0)), cache))
			if restored:
				print("=" * 70)
				print(f"[ankle姿勢 自動復元] {len(restored)} タブぶんの検出結果を復元しました")
				for key, n, cache in restored:
					src = Path(str(cache.get("source", ""))).name
					print(f"    {key}: {n} フレーム  (元データ: {src})")
				print("  ※ ②の設定を変えた場合は ④ を再実行してください")
				print("=" * 70)
			self._ankle_warn_if_cache_stale()
		except Exception as e:
			print(f"[ankle姿勢 自動復元] 失敗: {e}")

	def _ankle_warn_if_cache_stale(self) -> None:
		"""復元したキャッシュが現在の②設定と食い違っていないか検査して警告する。"""
		try:
			cache = self._ankle_get_current_cache()
			if not cache:
				return
			issues = []
			try:
				cur_size = float(self.ankle_marker_size_mm.get())
				old_size = float(cache.get("marker_size_mm", cur_size))
				if abs(cur_size - old_size) > 1e-6:
					issues.append(f"マーカー実寸 {old_size} mm → 現在 {cur_size} mm")
			except Exception:
				pass
			try:
				cur_dict = str(self.ankle_aruco_dict_var.get())
				old_dict = str(cache.get("aruco_dict", cur_dict))
				if cur_dict != old_dict:
					issues.append(f"ArUco辞書 {old_dict} → 現在 {cur_dict}")
			except Exception:
				pass
			try:
				src = str(cache.get("source", ""))
				cur_src = self.ankle_depth_path.get().strip() or self.ankle_video_path.get().strip()
				if src and cur_src and Path(src).name != Path(cur_src).name:
					issues.append(f"元データ {Path(src).name} → 現在 {Path(cur_src).name}")
			except Exception:
				pass
			if issues:
				print("[ankle姿勢] [警告] 復元した検出結果は現在の設定と異なります:")
				for it in issues:
					print(f"    - {it}")
				print("    → ④ ArUco検出+PnP実行 を押し直すことを推奨します")
		except Exception:
			pass

	def _ankle_save_pose_cache_npz(self, path: str, cache: dict) -> None:
		"""キャッシュを .npz に保存。"""
		import numpy as np
		payload = {}
		payload["timestamps"] = cache["timestamps"]
		meta = {k: v for k, v in cache.items() if k not in ("timestamps", "bones")}
		payload["meta_json"] = np.array(json.dumps(meta, ensure_ascii=False), dtype=object)
		for aid, b in cache["bones"].items():
			prefix = f"id_{int(aid)}_"
			payload[prefix + "poses"] = b["poses"]
			payload[prefix + "detected"] = b["detected"]
			payload[prefix + "reproj_err"] = b["reproj_err"]
		np.savez_compressed(path, **payload)

	def _ankle_load_pose_cache_npz(self, path: str) -> dict:
		"""'.npz' から姿勢キャッシュを読み込む。"""
		import numpy as np
		data = np.load(path, allow_pickle=True)
		meta = json.loads(str(data["meta_json"].item()))
		cache = dict(meta)
		cache["timestamps"] = data["timestamps"]
		cache["bones"] = {}
		ids = set()
		for key in data.files:
			if key.startswith("id_") and key.endswith("_poses"):
				aid = int(key.split("_")[1])
				ids.add(aid)
		for aid in sorted(ids):
			prefix = f"id_{aid}_"
			cache["bones"][aid] = {
				"poses": data[prefix + "poses"],
				"detected": data[prefix + "detected"],
				"reproj_err": data[prefix + "reproj_err"],
			}
		return cache

	# ---- 簡易版: 関節座標系 Cj 校正 ----
	def on_ankle_calibrate_axis(self, axis_name: str) -> None:
		"""指定軸 (ML/AP/PD) の校正: 短時間D405録画 → クランプArUco軌跡 → PCA1軸抽出。"""
		import numpy as np
		if axis_name not in ("ML", "AP", "PD"):
			messagebox.showerror("軸校正", f"不正な軸名: {axis_name}"); return
		try:
			clamp_id = int(self.ankle_clamp_aruco_id.get())
		except Exception:
			clamp_id = 100
		try:
			seconds = float(self.ankle_axis_calib_seconds.get())
		except Exception:
			seconds = 4.0
		try:
			marker_size_mm = float(self.ankle_marker_size_mm.get())
		except Exception:
			marker_size_mm = 20.0
		aruco_dict_name = str(self.ankle_aruco_dict_var.get())
		if not self._ankle_check_rs():
			return
		if not self._ankle_check_cv2():
			return
		try:
			self.focus_set()
		except Exception:
			pass
		# 録画先
		import datetime as _dt
		default_dir = Path(__file__).parent / "cache"
		try: default_dir.mkdir(parents=True, exist_ok=True)
		except Exception: pass
		tag = f"axis_{axis_name}_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.db3"
		bag_path = filedialog.asksaveasfilename(
			title=f"{axis_name}軸校正: 録画先 .db3",
			initialdir=str(default_dir),
			initialfile=tag,
			defaultextension=".db3",
			filetypes=[("RealSense recording", "*.db3"), ("すべてのファイル", "*.*")])
		if not bag_path:
			return
		messagebox.showinfo(
			f"{axis_name}軸校正",
			f"クランプに ArUco (ID={clamp_id}) を貼り、以下の順で操作してください:\n\n"
			f"1. プレビューウィンドウで クランプマーカーが見えることを確認\n"
			f"2. コントロールパネルの「録画開始」を押す\n"
			f"3. **ロボットで {axis_name} 方向にのみ動かす** (往復可)\n"
			f"4. 十分な移動範囲を確保したら「録画停止」")
		ok, frame_count, err = self._ankle_rs_run_preview_and_record(bag_path)
		if err:
			messagebox.showerror("軸校正", f"録画失敗: {err}"); return
		if not ok or frame_count < 5:
			messagebox.showwarning("軸校正",
				f"録画フレームが少なすぎます (frames={frame_count})。もう一度お試しください。")
			return
		# 検出 (クランプIDだけ)
		update_cb, close_cb, cancel_cb = self._ankle_open_progress(
			f"{axis_name}軸: クランプ検出中")
		try:
			cache = self._ankle_detect_from_bag(
				bag_path, aruco_dict_name, marker_size_mm, {clamp_id}, 1, update_cb, cancel_cb)
		except Exception as e:
			close_cb()
			messagebox.showerror("軸校正 検出失敗", f"{e}"); return
		close_cb()
		bones_cache = cache.get("bones", {}) or {}
		if clamp_id not in bones_cache:
			messagebox.showerror("軸校正",
				f"クランプ ID={clamp_id} が検出されませんでした。ID・辞書・実寸を確認してください。")
			return
		poses = np.asarray(bones_cache[clamp_id]["poses"], dtype=float)
		detected = np.asarray(bones_cache[clamp_id]["detected"], dtype=bool)
		positions = poses[detected, :3, 3]  # (N_det, 3)  カメラ系のマーカー中心位置
		if len(positions) < 5:
			messagebox.showwarning("軸校正",
				f"クランプの検出フレーム数が少なすぎます: {len(positions)}件。"); return
		# PCA: 第1主成分 = 軸方向
		centroid = positions.mean(axis=0)
		centered = positions - centroid
		U, S, Vt = np.linalg.svd(centered, full_matrices=False)
		axis_dir = Vt[0] / np.linalg.norm(Vt[0])
		# 移動範囲 (第1主成分投影の最大幅)
		proj = centered @ axis_dir
		span_mm = float(proj.max() - proj.min())
		# 残差 (第2/3主成分の大きさ) → 直線からのバラつき
		residual_mm = float(np.sqrt(np.mean(centered @ Vt[1] * centered @ Vt[1]
		                                     + centered @ Vt[2] * centered @ Vt[2])))
		self._ankle_axis_calib_results[axis_name] = {
			"axis_dir": axis_dir.tolist(),
			"centroid": centroid.tolist(),
			"span_mm": span_mm,
			"residual_mm": residual_mm,
			"n_positions": int(len(positions)),
		}
		self._ankle_update_axis_calib_status()
		messagebox.showinfo(
			f"{axis_name}軸校正 完了",
			f"軸方向 (カメラ系, 単位ベクトル):\n"
			f"  [{axis_dir[0]:+.4f}, {axis_dir[1]:+.4f}, {axis_dir[2]:+.4f}]\n\n"
			f"移動範囲: {span_mm:.2f} mm ({len(positions)}点)\n"
			f"直線からの残差RMSE: {residual_mm:.3f} mm\n\n"
			+ ("良好" if residual_mm < 2.0 else "残差やや大 (要確認)"))

	def _ankle_update_axis_calib_status(self) -> None:
		lines = ["軸校正状態:"]
		for ax in ("ML", "AP", "PD"):
			r = self._ankle_axis_calib_results.get(ax)
			if r:
				d = r["axis_dir"]
				lines.append(f"  {ax}: [{d[0]:+.3f}, {d[1]:+.3f}, {d[2]:+.3f}] "
				             f"span={r['span_mm']:.1f}mm, 残差={r['residual_mm']:.2f}mm")
			else:
				lines.append(f"  {ax}: (未校正)")
		if self._ankle_joint_frame_Cj is not None:
			lines.append("→ Cj確定済み")
		try:
			self.ankle_axis_calib_status.set("\n".join(lines))
		except Exception:
			pass

	def on_ankle_finalize_joint_frame(self) -> None:
		"""3軸から関節座標系 Cj を確定 (SVDで直交化)。"""
		import numpy as np
		results = self._ankle_axis_calib_results
		missing = [ax for ax in ("ML", "AP", "PD") if ax not in results]
		if missing:
			messagebox.showwarning("Cj確定", f"未校正軸: {', '.join(missing)}"); return
		# 3軸を列としてまとめる
		M = np.column_stack([
			np.asarray(results["ML"]["axis_dir"], dtype=float),
			np.asarray(results["AP"]["axis_dir"], dtype=float),
			np.asarray(results["PD"]["axis_dir"], dtype=float),
		])
		# SVDで最近直交回転行列に (M ≈ U V^T, R = U V^T の変形)
		U, S, Vt = np.linalg.svd(M)
		R_Cj = U @ Vt
		# 右手系保証
		if np.linalg.det(R_Cj) < 0:
			U[:, -1] *= -1
			R_Cj = U @ Vt
		# 原点: 3軸校正の重心平均
		centroids = [np.asarray(results[ax]["centroid"], dtype=float) for ax in ("ML","AP","PD")]
		origin = np.mean(centroids, axis=0)
		# Cj: カメラ系 → Cj系 の変換。 R_Cj の列がカメラ系での Cj軸方向
		# T_Cj←C 構築: p_Cj = R_Cj^T (p_C - origin)  → 4x4形式で
		T = np.eye(4)
		T[:3, :3] = R_Cj.T   # C→Cj の回転
		T[:3, 3] = -R_Cj.T @ origin
		self._ankle_joint_frame_Cj = T.tolist()
		# 直交化残差 (元の3軸との角度差)
		angles = []
		for i, ax in enumerate(("ML","AP","PD")):
			orig = np.asarray(results[ax]["axis_dir"], dtype=float)
			new = R_Cj[:, i]
			cos_a = float(np.clip(np.dot(orig, new), -1, 1))
			angles.append(np.rad2deg(np.arccos(cos_a)))
		self._ankle_update_axis_calib_status()
		messagebox.showinfo(
			"関節座標系 Cj 確定",
			f"3軸から Cj を SVD直交化しました。\n\n"
			f"直交化での各軸ずれ (deg):\n"
			f"  ML: {angles[0]:.2f}°\n"
			f"  AP: {angles[1]:.2f}°\n"
			f"  PD: {angles[2]:.2f}°\n\n"
			f"最大ずれ < 5° なら3軸がおおむね直交しており良好。\n"
			f"それ以上なら3軸校正の動きが直交していなかった可能性があります。")

	def on_ankle_clear_joint_frame(self) -> None:
		if not messagebox.askyesno("Cjクリア", "3軸校正結果と Cj を全て削除しますか?"):
			return
		self._ankle_axis_calib_results = {}
		self._ankle_joint_frame_Cj = None
		self._ankle_update_axis_calib_status()

	def on_ankle_analyze_in_cj(self) -> None:
		"""現在の姿勢時系列キャッシュを、Cj系での骨マーカー動揺として出力 (CSV + matplotlibグラフ)。"""
		import numpy as np
		if self._ankle_joint_frame_Cj is None:
			messagebox.showwarning("Cj分析", "先に関節座標系 Cj を確定してください。"); return
		cache = self._ankle_get_current_cache()
		if not cache:
			messagebox.showwarning("Cj分析",
				"姿勢時系列がありません。⓪で本試験録画 → ④でArUco検出+PnP実行 してください。")
			return
		bones_cache = cache.get("bones", {}) or {}
		if not bones_cache:
			messagebox.showwarning("Cj分析", "検出データがありません。"); return
		T_Cj_C = np.asarray(self._ankle_joint_frame_Cj, dtype=float)
		R_Cj_C = T_Cj_C[:3, :3]
		# クランプIDを分析対象から除外 (静止基準として扱う)
		try:
			clamp_id = int(self.ankle_clamp_aruco_id.get())
		except Exception:
			clamp_id = -1
		# ID→(骨名, 色) 対応
		id_to_bone = {int(b.get("aruco_id", -1)): (b.get("name", ""), self._ankle_color_of(i))
		              for i, b in enumerate(self.ankle_bones)}
		# 出力先
		default_dir = Path(__file__).parent / "cache"
		try: default_dir.mkdir(parents=True, exist_ok=True)
		except Exception: pass
		import datetime as _dt
		out_dir = filedialog.askdirectory(
			title="Cj系分析結果 保存先フォルダ",
			initialdir=str(default_dir))
		if not out_dir:
			return
		ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
		timestamps = np.asarray(cache.get("timestamps", []), dtype=float)
		# matplotlib
		try:
			import matplotlib
			matplotlib.use("Agg")
			import matplotlib.pyplot as plt
		except ImportError:
			plt = None
		results_summary = []
		for aid, b in bones_cache.items():
			aid = int(aid)
			if aid == clamp_id:
				continue
			name, color = id_to_bone.get(aid, (f"ID{aid}", "#888888"))
			poses = np.asarray(b["poses"], dtype=float)  # (N,4,4)  T_C←Mk
			detected = np.asarray(b["detected"], dtype=bool)
			N = len(poses)
			# 各フレームの Cj系での位置と回転
			pos_cj = np.full((N, 3), np.nan)
			rot_cj_euler = np.full((N, 3), np.nan)  # ML, AP, PD 回りの回転
			for t in range(N):
				if not detected[t]:
					continue
				T_C_Mk = poses[t]
				# T_Cj←Mk = T_Cj←C · T_C←Mk
				T_Cj_Mk = T_Cj_C @ T_C_Mk
				pos_cj[t] = T_Cj_Mk[:3, 3]
				# Euler: ML(x), AP(y), PD(z) 回転
				R = T_Cj_Mk[:3, :3]
				# xyz順 Euler
				sy = np.sqrt(R[0,0]**2 + R[1,0]**2)
				if sy > 1e-6:
					x_rot = np.arctan2(R[2,1], R[2,2])
					y_rot = np.arctan2(-R[2,0], sy)
					z_rot = np.arctan2(R[1,0], R[0,0])
				else:
					x_rot = np.arctan2(-R[1,2], R[1,1])
					y_rot = np.arctan2(-R[2,0], sy)
					z_rot = 0.0
				rot_cj_euler[t] = np.rad2deg([x_rot, y_rot, z_rot])
			# 参照フレーム (t=0 の最初の検出) を減算 → 「動揺」
			ref_idx = None
			for t in range(N):
				if detected[t]:
					ref_idx = t; break
			if ref_idx is not None:
				pos_ref = pos_cj[ref_idx].copy()
				rot_ref = rot_cj_euler[ref_idx].copy()
				pos_delta = pos_cj - pos_ref[None, :]
				rot_delta = rot_cj_euler - rot_ref[None, :]
			else:
				pos_delta = pos_cj
				rot_delta = rot_cj_euler
			# CSV書き出し
			safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(name))
			csv_path = Path(out_dir) / f"ankle_cj_bone_id{aid:03d}_{safe_name}_{ts}.csv"
			try:
				with csv_path.open("w", encoding="utf-8") as f:
					f.write("frame,time_s,detected,pos_ML_mm,pos_AP_mm,pos_PD_mm,"
					        "rot_ML_deg,rot_AP_deg,rot_PD_deg,"
					        "delta_pos_ML_mm,delta_pos_AP_mm,delta_pos_PD_mm,"
					        "delta_rot_ML_deg,delta_rot_AP_deg,delta_rot_PD_deg\n")
					for t in range(N):
						ts_v = timestamps[t] if t < len(timestamps) else t / 15.0
						row = [t, f"{ts_v:.4f}", int(detected[t])]
						for v in (*pos_cj[t], *rot_cj_euler[t], *pos_delta[t], *rot_delta[t]):
							row.append(f"{v:.4f}" if not np.isnan(v) else "")
						f.write(",".join(str(x) for x in row) + "\n")
			except Exception as e:
				print(f"[cj分析] CSV書出失敗 ({name}): {e}")
			# グラフ
			if plt is not None:
				try:
					fig, axes = plt.subplots(2, 1, figsize=(9, 6), sharex=True)
					x = timestamps if len(timestamps) == N else np.arange(N) / 15.0
					axes[0].plot(x, pos_delta[:, 0], label="ΔML", color="red")
					axes[0].plot(x, pos_delta[:, 1], label="ΔAP", color="green")
					axes[0].plot(x, pos_delta[:, 2], label="ΔPD", color="blue")
					axes[0].set_ylabel("Δposition (mm)")
					axes[0].set_title(f"Bone motion in Cj — ID={aid} ({name})")
					axes[0].legend(); axes[0].grid(True, alpha=0.3)
					axes[1].plot(x, rot_delta[:, 0], label="ΔroundML", color="red")
					axes[1].plot(x, rot_delta[:, 1], label="ΔroundAP", color="green")
					axes[1].plot(x, rot_delta[:, 2], label="ΔroundPD", color="blue")
					axes[1].set_xlabel("time (s)"); axes[1].set_ylabel("Δrotation (deg)")
					axes[1].legend(); axes[1].grid(True, alpha=0.3)
					png_path = Path(out_dir) / f"ankle_cj_bone_id{aid:03d}_{safe_name}_{ts}.png"
					fig.tight_layout()
					fig.savefig(png_path, dpi=120)
					plt.close(fig)
				except Exception as e:
					print(f"[cj分析] グラフ失敗 ({name}): {e}")
			# サマリー
			det_rate = float(np.sum(detected)) / max(N, 1) * 100
			results_summary.append(
				f"  ID={aid} ({name}): 検出率{det_rate:.1f}%, "
				f"ΔML範囲={float(np.nanmax(pos_delta[:,0]) - np.nanmin(pos_delta[:,0])):.2f}mm, "
				f"ΔAP範囲={float(np.nanmax(pos_delta[:,1]) - np.nanmin(pos_delta[:,1])):.2f}mm, "
				f"ΔPD範囲={float(np.nanmax(pos_delta[:,2]) - np.nanmin(pos_delta[:,2])):.2f}mm")
		messagebox.showinfo(
			"Cj系分析 完了",
			f"骨マーカー動揺を Cj 系で分析しました。\n\n"
			f"保存先: {out_dir}\n"
			f"タイムスタンプ: {ts}\n\n"
			+ ("\n".join(results_summary) if results_summary else "分析対象骨なし"))

	def on_ankle_visualize_pose_series(self) -> None:
		"""検出済み姿勢時系列 T_C←Mk(t) を3D軌跡としてPyVistaで表示 (骨モデル・スキャン不要)。

		用途: スキャナ無しでの追跡動作確認。マーカーがカメラ座標系でどう動いたかが見える。
		"""
		import numpy as np
		cache = self._ankle_get_current_cache()
		if not cache:
			messagebox.showwarning("マーカー軌跡",
				"姿勢時系列がありません。先に「ArUco検出+PnP実行」または「姿勢時系列を読込」してください。")
			return
		bones_cache = cache.get("bones", {}) or {}
		if not bones_cache:
			messagebox.showwarning("マーカー軌跡", "検出データがありません。")
			return

		# 骨リストから ID → 名前/色 の対応表を作る
		id_to_idx = {}
		for i, b in enumerate(self.ankle_bones):
			try:
				id_to_idx[int(b.get("aruco_id", -1))] = i
			except Exception:
				pass

		# マーカー位置範囲を把握 (軸長さ推定用)
		all_pts = []
		trajectories = []  # [(aid, name, color, pts_all, detected)]
		for aid_raw, b in bones_cache.items():
			aid = int(aid_raw)
			poses = np.asarray(b.get("poses", []))
			detected = np.asarray(b.get("detected", []))
			if poses.size == 0 or len(detected) == 0:
				continue
			pts = poses[:, :3, 3]   # (N, 3)  未検出フレームはNaNのはず
			pts_det = pts[detected]
			if len(pts_det) == 0:
				continue
			idx = id_to_idx.get(aid)
			if idx is not None and 0 <= idx < len(self.ankle_bones):
				name = self.ankle_bones[idx].get("name", f"ID{aid}")
				color = self._ankle_color_of(idx)
			else:
				name = f"ID{aid}"
				color = self._ANKLE_DEFAULT_BONE_COLORS[aid % len(self._ANKLE_DEFAULT_BONE_COLORS)]
			trajectories.append((aid, name, color, pts, detected))
			all_pts.append(pts_det)

		if not trajectories:
			messagebox.showwarning("マーカー軌跡", "検出フレームが1つもありません。検出率0%です。")
			return

		all_arr = np.vstack(all_pts)
		bbox_min = all_arr.min(axis=0); bbox_max = all_arr.max(axis=0)
		diag = float(np.linalg.norm(bbox_max - bbox_min))
		axis_len = max(diag * 0.15, 30.0)
		sphere_r = max(diag * 0.005, 1.5)

		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="ankle: マーカー軌跡 (T_C←Mk, カメラ系, mm)",
		                     window_size=(int(sw * 0.85), int(sh * 0.85)))
		plotter.set_background("white")

		# カメラ座標軸 (原点=カメラ)
		origin = np.zeros(3)
		for j, col in enumerate(("red", "green", "blue")):
			direction = np.zeros(3); direction[j] = axis_len
			plotter.add_mesh(
				pv.Arrow(start=origin, direction=direction, scale='auto',
				         tip_length=0.15, tip_radius=0.04, shaft_radius=0.015),
				color=col)
		plotter.add_mesh(pv.Sphere(radius=sphere_r * 1.5, center=origin), color="black")
		plotter.add_point_labels([origin], ["Camera"], font_size=14,
		                          text_color="black", show_points=False, always_visible=True)

		summary_lines = ["マーカー軌跡 (カメラ座標系, mm)", "軸: X=赤, Y=緑, Z=青(奥)"]
		# 各マーカーの軌跡
		for aid, name, color, pts, detected in trajectories:
			pts_det = pts[detected]
			if len(pts_det) < 2:
				# 単一点だけならスフィアだけ
				plotter.add_mesh(pv.Sphere(radius=sphere_r, center=pts_det[0]), color=color)
				continue
			line = pv.lines_from_points(pts_det)
			plotter.add_mesh(line, color=color, line_width=2)
			# 始点/終点
			plotter.add_mesh(pv.Sphere(radius=sphere_r, center=pts_det[0]), color=color)
			plotter.add_mesh(pv.Sphere(radius=sphere_r * 1.3, center=pts_det[-1]), color=color)
			plotter.add_point_labels([pts_det[0]], [f"{name} (start)"], font_size=11,
			                          text_color=color, show_points=False, always_visible=True)
			# 統計
			rate = 100.0 * float(np.sum(detected)) / len(detected)
			span = float(np.linalg.norm(pts_det.max(axis=0) - pts_det.min(axis=0)))
			summary_lines.append(f"  ID={aid} ({name}): 検出率{rate:.1f}%, 移動範囲={span:.1f}mm")

		plotter.add_text("\n".join(summary_lines), position="upper_left",
		                 font_size=10, color="black")
		plotter.show()

	def on_ankle_save_marker_pdf(self) -> None:
		"""骨リストのマーカーを A4 PDF (物理実寸で印刷可能) にまとめて保存する。

		PDFはページに物理サイズが埋め込まれるので、PDFビューアで「実際のサイズ / 100%」
		で印刷すれば必ず指定mmになる。PNGの「ページに合わせる」自動拡大バグを回避。
		"""
		import numpy as np
		if not self._ankle_check_cv2():
			return
		if not self.ankle_bones:
			messagebox.showinfo("マーカーPDF", "骨リストが空です。③でまず骨を追加してください。")
			return
		items = []
		for b in self.ankle_bones:
			try:
				aid = int(b.get("aruco_id", -1))
			except Exception:
				continue
			if aid < 0:
				continue
			items.append((aid, str(b.get("name", "")).strip()))
		if not items:
			messagebox.showinfo("マーカーPDF", "有効なArUco IDがありません。")
			return
		aruco_dict_name = self.ankle_aruco_dict_var.get()
		try:
			marker_size_mm = float(self.ankle_marker_size_mm.get())
		except Exception:
			marker_size_mm = 20.0
		if marker_size_mm <= 0 or marker_size_mm > 200:
			messagebox.showwarning("マーカーPDF", "マーカー実寸は0〜200mmの範囲で指定してください。")
			return
		try:
			import matplotlib
			matplotlib.use('Agg')
			import matplotlib.pyplot as plt
			from matplotlib.backends.backend_pdf import PdfPages
		except ImportError:
			messagebox.showerror("マーカーPDF",
				"matplotlibが必要です。venvで pip install matplotlib を実行してください。")
			return
		import cv2
		try:
			dictionary = self._ankle_resolve_aruco_dict(aruco_dict_name)
		except Exception as e:
			messagebox.showerror("マーカーPDF", f"辞書解決失敗: {e}"); return
		path = filedialog.asksaveasfilename(
			title="マーカーPDF 保存先",
			defaultextension=".pdf",
			initialfile=f"aruco_markers_{marker_size_mm:.0f}mm_{aruco_dict_name}.pdf",
			filetypes=[("PDF", "*.pdf"), ("すべてのファイル", "*.*")])
		if not path:
			return
		# A4 (mm) / インチ変換
		A4_MM = (210.0, 297.0)
		A4_INCH = (A4_MM[0] / 25.4, A4_MM[1] / 25.4)
		# 各マーカー領域: quiet zone(マーカー20%相当) + ラベル
		quiet_zone_mm = max(marker_size_mm * 0.2, 3.0)
		cell_mm = marker_size_mm + 2 * quiet_zone_mm
		label_h_mm = 6.0
		unit_w_mm = cell_mm
		unit_h_mm = cell_mm + label_h_mm
		margin_mm = 12.0
		usable_w = A4_MM[0] - 2 * margin_mm
		usable_h = A4_MM[1] - 2 * margin_mm - 8   # 上部の説明分
		n_col = max(1, int(usable_w // unit_w_mm))
		n_row = max(1, int(usable_h // unit_h_mm))
		per_page = n_col * n_row
		# マーカー描画解像度 (高いほうがエッジきれい)
		render_px = 600
		n_pages = 0
		try:
			with PdfPages(path) as pdf:
				idx = 0
				while idx < len(items):
					n_pages += 1
					fig = plt.figure(figsize=A4_INCH)
					# ページ上部の説明
					fig.text(0.5, 0.98,
					         f"ArUco {aruco_dict_name} — 実寸 {marker_size_mm:.1f}mm — "
					         f"「実際のサイズ / 100%」で印刷 ({n_pages}ページ)",
					         ha='center', va='top', fontsize=9)
					for pos in range(per_page):
						if idx >= len(items):
							break
						aid, name = items[idx]
						mimg = cv2.aruco.generateImageMarker(dictionary, aid, render_px)
						row = pos // n_col
						col = pos % n_col
						# A4座標系 (左下原点, mm)
						x_mm = margin_mm + col * unit_w_mm
						y_top_mm = A4_MM[1] - margin_mm - 8 - row * unit_h_mm
						y_mm = y_top_mm - unit_h_mm
						# マーカー本体 (quiet zoneの内側に配置)
						mx = x_mm + quiet_zone_mm
						my = y_mm + label_h_mm + quiet_zone_mm
						ax = fig.add_axes([
							mx / A4_MM[0],
							my / A4_MM[1],
							marker_size_mm / A4_MM[0],
							marker_size_mm / A4_MM[1],
						])
						ax.imshow(mimg, cmap='gray', vmin=0, vmax=255, interpolation='nearest')
						ax.axis('off')
						# 外枠 (切り取り目安)
						cut_ax = fig.add_axes([
							x_mm / A4_MM[0],
							y_mm / A4_MM[1],
							cell_mm / A4_MM[0],
							cell_mm / A4_MM[1] + label_h_mm / A4_MM[1] * 0.3,
						])
						cut_ax.axis('off')
						cut_ax.set_xlim(0, 1); cut_ax.set_ylim(0, 1)
						cut_ax.plot([0, 1, 1, 0, 0], [0, 0, 1, 1, 0],
						            color='#cccccc', linewidth=0.5, linestyle=(0, (2, 2)))
						# ラベル
						label = f"ID={aid}"
						if name:
							label += f"  ({name})"
						label += f"  {marker_size_mm:.1f}mm"
						label_ax = fig.add_axes([
							x_mm / A4_MM[0],
							y_mm / A4_MM[1],
							cell_mm / A4_MM[0],
							label_h_mm / A4_MM[1],
						])
						label_ax.axis('off')
						label_ax.text(0.5, 0.5, label, ha='center', va='center', fontsize=7)
						idx += 1
					pdf.savefig(fig)
					plt.close(fig)
		except Exception as e:
			messagebox.showerror("マーカーPDF", f"PDF生成失敗: {e}"); return
		messagebox.showinfo(
			"マーカーPDF 保存完了",
			f"{len(items)} 個のマーカーを {n_pages} ページの PDF に保存しました。\n\n"
			f"ファイル: {path}\n"
			f"辞書:     {aruco_dict_name}\n"
			f"実寸:     {marker_size_mm:.1f} mm\n"
			f"配置:     {n_col} × {n_row} = {per_page} 個/ページ\n\n"
			f"【印刷方法 — 重要】\n"
			f"1. PDFを Adobe Reader / Chrome / Edge で開く\n"
			f"2. 印刷設定で「実際のサイズ」または「カスタムスケール 100%」を選択\n"
			f"   ・「ページに合わせる」「フィット」は絶対にOFF\n"
			f"3. レーザープリンタ + マット紙 で印刷\n"
			f"4. 印刷後、ノギスでマーカー実寸を確認 → ②の値に反映\n"
			f"5. quiet zone (外枠内の白余白) は切らずに残す")

	def on_ankle_save_marker_images(self) -> None:
		"""骨リストのArUco IDに対応する印刷用マーカー画像(PNG)をまとめて保存する。"""
		import numpy as np
		if not self._ankle_check_cv2():
			return
		if not self.ankle_bones:
			messagebox.showinfo("マーカー画像",
				"骨リストが空です。③でまず骨を追加してArUco IDを設定してください。")
			return
		# ID + 骨名を収集
		items = []
		for b in self.ankle_bones:
			try:
				aid = int(b.get("aruco_id", -1))
			except Exception:
				continue
			if aid < 0:
				continue
			items.append((aid, str(b.get("name", "")).strip()))
		if not items:
			messagebox.showinfo("マーカー画像", "有効なArUco IDが設定された骨がありません。")
			return
		aruco_dict_name = self.ankle_aruco_dict_var.get()
		try:
			marker_size_mm = float(self.ankle_marker_size_mm.get())
		except Exception:
			marker_size_mm = 20.0
		out_dir = filedialog.askdirectory(title="マーカー画像の保存先を選択")
		if not out_dir:
			return
		import cv2
		try:
			dictionary = self._ankle_resolve_aruco_dict(aruco_dict_name)
		except Exception as e:
			messagebox.showerror("マーカー画像", f"辞書解決失敗: {e}"); return
		# マーカー本体600px + 白余白(quiet zone)100px → 全体800px
		marker_px = 600
		pad = 100
		total_px = marker_px + 2 * pad
		saved = []
		errs = []
		for aid, name in items:
			try:
				mimg = cv2.aruco.generateImageMarker(dictionary, aid, marker_px)
				canvas = np.full((total_px, total_px), 255, dtype=np.uint8)
				canvas[pad:pad + marker_px, pad:pad + marker_px] = mimg
				name_safe = ("_" + "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in name)) if name else ""
				fname = f"aruco_{aruco_dict_name}_id{aid:03d}{name_safe}.png"
				fpath = Path(out_dir) / fname
				cv2.imwrite(str(fpath), canvas)
				saved.append((aid, name, fname))
			except Exception as e:
				errs.append((aid, str(e)))
		lines = [f"{len(saved)} 個のマーカー画像を保存しました。",
		         "",
		         f"保存先: {out_dir}",
		         f"辞書:   {aruco_dict_name}",
		         f"目標実寸: {marker_size_mm:.2f} mm",
		         "",
		         "保存ファイル:"]
		for a, n, f in saved:
			lines.append(f"  ID={a:3d}{('  ('+n+')') if n else ''}: {f}")
		if errs:
			lines.append("")
			lines.append("失敗:")
			for a, e in errs:
				lines.append(f"  ID={a}: {e}")
		lines += [
			"",
			"【印刷手順 / 重要】",
			"1. レーザープリンタ + マット紙 (光沢紙はNG)",
			f"2. マーカー実寸 = {marker_size_mm:.2f} mm になるよう印刷スケールを調整",
			"   (PNG外寸 800px、マーカー本体 600px / 800px)",
			"3. 印刷後にノギスで実測 → ②の「マーカー実寸(mm)」に反映",
			"4. マーカー外周の白部分 (quiet zone) は切らずに残す",
			"5. 剥がれ防止に透明フィルムでラミネート推奨"]
		messagebox.showinfo("マーカー画像 保存完了", "\n".join(lines))

	def on_ankle_save_pose_series(self) -> None:
		cache = self._ankle_get_current_cache()
		if not cache:
			messagebox.showinfo("保存", "保存する姿勢時系列がありません。先に「ArUco検出+PnP実行」してください。")
			return
		path = filedialog.asksaveasfilename(
			title="姿勢時系列を保存 (.npz)",
			defaultextension=".npz",
			filetypes=[("NPZファイル", "*.npz"), ("すべてのファイル", "*.*")])
		if not path:
			return
		try:
			self._ankle_save_pose_cache_npz(path, cache)
		except Exception as e:
			messagebox.showerror("保存エラー", f"保存に失敗しました:\n{e}"); return
		self.ankle_pose_series_path.set(path)
		messagebox.showinfo("保存完了", f"保存しました:\n{path}")

	def on_ankle_load_pose_series(self) -> None:
		path = self.ankle_pose_series_path.get().strip() or filedialog.askopenfilename(
			title="姿勢時系列を読込 (.npz)",
			filetypes=[("NPZファイル", "*.npz"), ("すべてのファイル", "*.*")])
		if not path:
			return
		if not Path(path).exists():
			messagebox.showwarning("読込", f"ファイルが見つかりません: {path}"); return
		try:
			cache = self._ankle_load_pose_cache_npz(path)
		except Exception as e:
			messagebox.showerror("読込エラー", f"読込に失敗しました:\n{e}"); return
		self._ankle_set_current_cache(cache)
		self._ankle_autosave_pose_cache()      # 手動読込も次回起動に引き継ぐ
		self.ankle_pose_series_path.set(path)
		self._ankle_update_detection_status()
		messagebox.showinfo("読込完了", self._ankle_cache_status_text(cache))

	def on_ankle_visualize_initial(self) -> None:
		"""初期状態スキャンだけをPyVistaで表示する。"""
		scan_path = self.ankle_initial_scan_path.get().strip()
		if not scan_path:
			messagebox.showwarning("可視化", "①初期状態スキャンを選択してください。")
			return
		try:
			scan_mesh = pv.read(scan_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"初期状態スキャンの読み込みに失敗しました:\n{e}")
			return
		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="ankle: 初期状態スキャン",
		                     window_size=(int(sw * 0.85), int(sh * 0.85)))
		plotter.set_background("white")
		plotter.add_mesh(scan_mesh, color="lightgray", smooth_shading=True)
		plotter.add_text(f"初期状態スキャン: {Path(scan_path).name}",
		                 position="upper_left", font_size=10, color="black")
		plotter.show()

	def on_ankle_visualize_all(self) -> None:
		"""位置合わせ済みの全骨(モデル→W系)と、任意で初期スキャン(灰半透明)を重ねて表示。"""
		registered = [b for b in self.ankle_bones if b.get("reg_T") is not None and b.get("model_path")]
		if not registered:
			messagebox.showwarning("全骨可視化",
				"位置合わせ済みの骨がありません。まず③で位置合わせを実行してください。\n"
				"(参考: 「初期状態を可視化」なら位置合わせ不要です)")
			return
		sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
		plotter = pv.Plotter(title="ankle: 全骨(位置合わせ済み・W系)",
		                     window_size=(int(sw * 0.9), int(sh * 0.9)))
		plotter.set_background("white")
		scan_path = self.ankle_initial_scan_path.get().strip()
		if scan_path:
			try:
				plotter.add_mesh(pv.read(scan_path), color="lightgray", opacity=0.25, smooth_shading=True)
			except Exception as e:
				print(f"[ankle vis-all] 初期スキャン表示失敗: {e}")
		names = []
		for b in registered:
			try:
				m = pv.read(b["model_path"])
				T = np.asarray(b["reg_T"], dtype=float)
				self._knee_apply_T(m, T)
				plotter.add_mesh(m, color=self._ankle_color_of(self.ankle_bones.index(b)),
				                 opacity=1.0, smooth_shading=True)
				names.append(b.get("name", "?"))
			except Exception as e:
				print(f"[ankle vis-all] {b.get('name','?')} 表示失敗: {e}")
		plotter.add_text(f"全骨(W系, 位置合わせ済み): {', '.join(names)}",
		                 position="upper_left", font_size=10, color="black")
		plotter.show()

	# ---- Stage 5: 骨アニメーション + 骨対ヒートマップ ----
	def _ankle_heatmap_bone_indices(self):
		"""ヒートマップ骨対 (骨A, 骨B) の ankle_bones 上インデックスを返す。

		現行の可視化は N骨マルチヒートマップに移行済みで、UI ではペア指定を廃止。
		本メソッドは残存参照 (self_pose の boneA_ref 世界フレーム基準) の互換維持用に
		「骨リスト先頭2本」を既定として返す。骨が1本のみなら (0, None)、空なら (None, None)。
		"""
		n = len(self.ankle_bones)
		if n == 0:
			return None, None
		if n == 1:
			return 0, None
		return 0, 1

	def _ankle_pick_reference_pose(self, poses, detected, ref_frame: int):
		"""参照フレーム(t=0)の姿勢を選ぶ。指定frameが検出失敗なら最近傍の検出済みfromへフォールバック。"""
		import numpy as np
		N = len(detected)
		if N == 0:
			return None, -1
		ref = max(0, min(int(ref_frame), N - 1))
		if bool(detected[ref]):
			return np.asarray(poses[ref], dtype=float), ref
		# 前後探索
		for d in range(1, N):
			for k in (ref - d, ref + d):
				if 0 <= k < N and bool(detected[k]):
					return np.asarray(poses[k], dtype=float), k
		return None, -1

	def _ankle_build_bone_transforms(self, cache: dict):
		"""キャッシュ+骨リストから、各骨の T_W←bone(t) 時系列を構築する。

		動作モードで分岐:
		- "original": T_W←bone(t) = ΔMk(t) · T_W←L    (要 reg_T)
		- "self_pose": T_W←bone(t) = T_W←C · T_C←Mk(t) · inv(T_L←Mk)    (要 marker_to_bone_T)

		Returns:
		    animatable: list of (bone_idx, mesh_L(pv.PolyData), T_series(N,4,4))
		    frame_count: int (共通フレーム数)
		    warnings: list[str]
		"""
		import numpy as np
		N = int(cache.get("frame_count", 0))
		if N == 0:
			return [], 0, ["キャッシュにフレームがありません"]
		try:
			mode = str(self.ankle_workflow_mode.get())
		except Exception:
			mode = "original"
		ref = int(self.ankle_ref_frame.get())
		warnings = []
		animatable = []
		bones_cache = cache.get("bones", {}) or {}

		if mode == "self_pose":
			# --- 新プラン: T_C←bone(t) = T_C←Mk(t) · inv(T_L←Mk) ---
			try:
				world_choice = str(self.ankle_self_pose_world.get())
			except Exception:
				world_choice = "camera_ref"
			# W = camera_ref: 参照フレームのカメラ姿勢を W とする → T_W←C = inv(T_C←C(0)) = I
			#                 つまり T_W←bone(t) = T_C←bone(t) をそのまま使用 (W = C(0))
			# W = boneA_ref: ヒートマップの骨A の参照フレーム姿勢を W とする
			#                T_W←bone(t) = inv(T_C←boneA(0)) · T_C←bone(t)
			# まず各骨の T_C←bone(t) を計算
			bone_data = []  # (idx, name, mesh_L, T_C_bone_series, ref_valid)
			for idx, bone in enumerate(self.ankle_bones):
				name = bone.get("name", f"骨{idx+1}")
				aid = int(bone.get("aruco_id", -1))
				if aid not in bones_cache:
					warnings.append(f"{name} (ID={aid}): 姿勢時系列にIDなし → スキップ")
					continue
				if bone.get("marker_to_bone_T") is None:
					warnings.append(f"{name}: マーカー-骨キャリブ未実行 (新プラン必須) → スキップ")
					continue
				if not bone.get("model_path"):
					warnings.append(f"{name}: 骨モデル未設定 → スキップ")
					continue
				try:
					mesh_L = pv.read(bone["model_path"])
				except Exception as e:
					warnings.append(f"{name}: モデル読込失敗 ({e}) → スキップ")
					continue
				T_L_Mk = np.asarray(bone["marker_to_bone_T"], dtype=float)
				try:
					T_Mk_L = np.linalg.inv(T_L_Mk)
				except np.linalg.LinAlgError:
					warnings.append(f"{name}: T_L←Mk が特異 → スキップ")
					continue
				b = bones_cache[aid]
				poses = np.asarray(b["poses"], dtype=float)     # (N,4,4) T_C←Mk(t)
				detected = np.asarray(b["detected"], dtype=bool)
				if len(poses) != N or len(detected) != N:
					warnings.append(f"{name}: 姿勢時系列のフレーム数不一致 → スキップ")
					continue
				# T_C←bone(t) = T_C←Mk(t) · T_Mk←L
				T_C_bone = np.zeros((N, 4, 4), dtype=float)
				last_valid = None
				for t in range(N):
					if detected[t]:
						T_C_bone[t] = poses[t] @ T_Mk_L
						last_valid = T_C_bone[t]
					else:
						T_C_bone[t] = last_valid if last_valid is not None else np.eye(4)
				bone_data.append((idx, name, mesh_L, T_C_bone, detected))
			if not bone_data:
				return [], N, warnings
			# W 座標系の決定
			T_W_C = np.eye(4)  # 既定: W = C
			if world_choice == "camera_ref":
				# W = 参照フレームでのカメラ姿勢 → T_W←C = C(ref)→C(ref) = I
				# 実際は各骨が動くだけなので、W=C(0) と等価。何もしない。
				T_W_C = np.eye(4)
			elif world_choice == "boneA_ref":
				# ヒートマップ骨A の参照フレーム姿勢を W とする
				prox_i, _ = self._ankle_heatmap_bone_indices()
				# bone_data の中で prox_i を探す
				target = None
				for (idx, name, mesh_L, T_C_bone, detected) in bone_data:
					if idx == prox_i:
						target = (name, T_C_bone, detected); break
				if target is None:
					warnings.append("W=骨A基準 指定だが骨Aが未選択 or 未検出 → W=C(0) にフォールバック")
					T_W_C = np.eye(4)
				else:
					name_A, T_C_A, det_A = target
					T_A_ref, ref_used = self._ankle_pick_reference_pose(T_C_A, det_A, ref)
					if T_A_ref is None:
						warnings.append(f"W=骨A基準: {name_A} の参照フレームが検出無し → W=C(0)")
						T_W_C = np.eye(4)
					else:
						# W = 骨A の参照姿勢 (in C) → T_W←C = inv(T_C←W) = inv(T_A_ref)
						try:
							T_W_C = np.linalg.inv(T_A_ref)
						except np.linalg.LinAlgError:
							warnings.append(f"W=骨A基準: T_A_ref 特異 → W=C(0)")
							T_W_C = np.eye(4)
			# 最終 T_W←bone(t) = T_W←C · T_C←bone(t)
			for (idx, name, mesh_L, T_C_bone, detected) in bone_data:
				Ts = np.zeros((N, 4, 4), dtype=float)
				for t in range(N):
					Ts[t] = T_W_C @ T_C_bone[t]
				animatable.append((idx, mesh_L, Ts))
			return animatable, N, warnings

		# --- 原プラン: 従来の相対運動アニメ ---
		for idx, bone in enumerate(self.ankle_bones):
			name = bone.get("name", f"骨{idx+1}")
			aid = int(bone.get("aruco_id", -1))
			if aid not in bones_cache:
				warnings.append(f"{name} (ID={aid}): 姿勢時系列にIDなし → スキップ")
				continue
			if bone.get("reg_T") is None:
				warnings.append(f"{name}: 位置合わせ未実行 → スキップ")
				continue
			if not bone.get("model_path"):
				warnings.append(f"{name}: 骨モデル未設定 → スキップ")
				continue
			try:
				mesh_L = pv.read(bone["model_path"])
			except Exception as e:
				warnings.append(f"{name}: モデル読込失敗 ({e}) → スキップ")
				continue
			reg_T = np.asarray(bone["reg_T"], dtype=float)
			b = bones_cache[aid]
			poses = np.asarray(b["poses"], dtype=float)  # (N,4,4) 未検出はNaN
			detected = np.asarray(b["detected"], dtype=bool)
			if len(poses) != N or len(detected) != N:
				warnings.append(f"{name}: 姿勢時系列のフレーム数不一致 → スキップ")
				continue
			T0, ref_used = self._ankle_pick_reference_pose(poses, detected, ref)
			if T0 is None:
				warnings.append(f"{name}: 参照フレームで検出無し(全フレーム未検出) → スキップ")
				continue
			try:
				T0_inv = np.linalg.inv(T0)
			except np.linalg.LinAlgError:
				warnings.append(f"{name}: 参照姿勢が特異 → スキップ")
				continue
			if ref_used != ref:
				warnings.append(f"{name}: 参照フレーム{ref}未検出 → 代替フレーム{ref_used}を使用")
			# 各フレームの T_W←bone(t) を計算 (未検出は直前保持)
			Ts = np.zeros((N, 4, 4), dtype=float)
			last_valid = reg_T.copy()   # 参照姿勢では delta=I なので T = reg_T
			for t in range(N):
				if detected[t]:
					delta = poses[t] @ T0_inv
					Ts[t] = delta @ reg_T
					last_valid = Ts[t]
				else:
					Ts[t] = last_valid
			animatable.append((idx, mesh_L, Ts))
		return animatable, N, warnings

	def _ankle_apply_T_to_mesh(self, mesh, T):
		"""pvメッシュに4x4同次変換を適用してW系配置したコピーを返す。"""
		import numpy as np
		m = mesh.copy()
		if m.n_points > 0:
			pts = np.hstack([m.points, np.ones((m.n_points, 1))])
			m.points = (np.asarray(T, dtype=float) @ pts.T).T[:, :3]
		return m

	def _ankle_signed_distance_A_to_B(self, mesh_A_W, mesh_B_W):
		"""A の各頂点から B の三角メッシュへの符号付き距離 (mm)。
		open3d BVH を毎フレーム構築する簡易実装 (MVP)。"""
		import numpy as np
		try:
			import open3d as o3d
		except ImportError:
			return None
		mb = mesh_B_W
		if not isinstance(mb, pv.PolyData):
			mb = mb.extract_surface()
		mb = mb.triangulate()
		verts = np.asarray(mb.points, dtype=np.float32)
		faces_raw = np.asarray(mb.faces)
		if faces_raw.size < 4:
			return None
		faces = faces_raw.reshape(-1, 4)[:, 1:4].astype(np.int32)
		scene = o3d.t.geometry.RaycastingScene()
		scene.add_triangles(o3d.t.geometry.TriangleMesh(
			o3d.core.Tensor(verts, o3d.core.float32),
			o3d.core.Tensor(faces, o3d.core.int32)))
		pts_A = np.asarray(mesh_A_W.points, dtype=np.float32)
		sd = scene.compute_signed_distance(o3d.core.Tensor(pts_A, o3d.core.float32)).numpy()
		return np.asarray(sd, dtype=np.float32)

	# ================================================================
	# 共通シミュレーション表示エンジン (_sim_view_*)
	# hip/knee の on_animate と同等の UI/UX/操作感を N骨に汎用化した実装。
	# ankle が使用中。将来 hip/knee も同じヘルパに載せ替え可 (TODO)。
	# ================================================================

	@staticmethod
	def _sim_view_heatmap_cmap():
		"""接触ヒートマップの共通カラーマップ (-10mm=赤 → 0mm=緑)。hip と揃える。"""
		try:
			from matplotlib.colors import LinearSegmentedColormap
			return LinearSegmentedColormap.from_list(
				'contact_pen',
				[(0.0, (0.75, 0.0, 0.0)),
				 (0.5, (1.0, 0.6, 0.0)),
				 (1.0, (0.1, 0.75, 0.1))])
		except Exception:
			return 'RdYlGn'

	def _sim_view_precompute_multi_heatmap(self, bones_data, update_progress, cancel_var):
		"""N骨のマルチヒートマップを事前計算 (膝/股と同じ高速エンジン _precompute_heatmaps_o3d を流用)。

		各フレーム t / 各骨 i について
		    scalars[t][i][v] = min over j!=i of signed_distance(bone_i vertex_v (t), bone_j surface (t))
		を返す。負値=めり込み・0=接触・正=離間。

		内部実装: 骨ペア (i, j) ごとに膝/股の `_precompute_heatmaps_o3d` を呼び出す。
		- 固定側 (dist_surface) = 骨 j のローカル姿勢メッシュ (BVH シーンをペアごと 1回だけ構築)
		- 近位点 (prox_points) = 骨 i のローカル頂点
		- transform_data[t] = {'matrix': Trel_ij[t]}  ここで Trel_ij[t] = inv(Ts_j[t]) @ Ts_i[t]
		  (骨 j ローカル系から見た骨 i の姿勢)
		- 剛体変換の不変性を利用し、近位点を逆変換して固定シーンで一括問い合わせ

		Args:
			bones_data: list of (bone_idx, name, mesh_L, T_series (N_frames,4,4))
			update_progress: fn(current, total, msg) -> bool (False で中断)
			cancel_var: tk.BooleanVar (True で中断)

		Returns:
			list of dict{bone_idx: np.ndarray(N_verts_i,)}  長さ N_frames
			失敗/中断時: 空配列
		"""
		import numpy as np
		if not bones_data:
			return []
		N = len(bones_data[0][3])
		n_bones = len(bones_data)
		# 骨 i のローカル頂点 (mesh_L.points)
		local_pts = {}
		for (idx, name, mesh_L, Ts) in bones_data:
			local_pts[idx] = np.asarray(mesh_L.points, dtype=float)

		# 骨 i × 骨 j ごとの距離配列時系列 (list of numpy(N_verts_i,)) を計算
		# per_pair_dist[(i, j)][t] = 骨 i の頂点 v から 骨 j 表面への signed distance (t フレーム目)
		per_pair_dist = {}
		n_pairs = n_bones * (n_bones - 1)
		pair_count = 0
		for (idx_i, name_i, mesh_L_i, Ts_i) in bones_data:
			for (idx_j, name_j, mesh_L_j, Ts_j) in bones_data:
				if idx_i == idx_j:
					continue
				pair_count += 1
				if cancel_var.get():
					return []
				# ペア (i, j) の transform_data を構築
				#   Trel_ij[t] = inv(Ts_j[t]) @ Ts_i[t] (骨 j ローカル系から見た骨 i の姿勢)
				transform_data_pair = []
				for t in range(N):
					try:
						Tj_inv = np.linalg.inv(Ts_j[t])
					except np.linalg.LinAlgError:
						Tj_inv = np.eye(4)
					transform_data_pair.append({'matrix': Tj_inv @ Ts_i[t]})

				# ペア進捗ラップ (全ペア中の進捗を通算)
				def _pair_progress(cur, tot, msg, _pi=pair_count, _np=n_pairs, _ni=name_i, _nj=name_j):
					# 全体進捗 = (前ペア完了分 + 当ペア内進捗) / 全ペア
					overall = ((_pi - 1) + cur / max(tot, 1)) / max(_np, 1) * 100
					try:
						update_progress(int(overall * 100), 10000,
						                f"[{_pi}/{_np}] {_ni} → {_nj}: {cur}/{tot} frames")
					except Exception:
						pass
					return not cancel_var.get()

				# 膝/股の高速エンジンを呼び出し (BVH 1回構築 + フレームチャンク一括)
				try:
					prox_region = pv.PolyData(local_pts[idx_i])
					lazy = self._precompute_heatmaps_o3d(
						prox_joint_region=prox_region,
						dist_surface=mesh_L_j,
						prox_points=local_pts[idx_i],
						transform_data=transform_data_pair,
						update_progress=_pair_progress,
						cancel_var=cancel_var)
					if cancel_var.get():
						return []
					# lazy は _LazyHeatmapList もしくは list of PolyData 相当
					distances_list = []
					if hasattr(lazy, 'distances'):
						distances_list = [np.asarray(d, dtype=np.float32) if d is not None else None
						                  for d in lazy.distances]
					else:
						# fallback: list of PolyData
						for hm in lazy:
							if hm is None or 'distance' not in getattr(hm, 'array_names', []):
								distances_list.append(None)
							else:
								distances_list.append(np.asarray(hm['distance'], dtype=np.float32))
					per_pair_dist[(idx_i, idx_j)] = distances_list
				except Exception as e:
					print(f"[multi-heatmap] ペア ({name_i} → {name_j}) 計算失敗: {e}")
					per_pair_dist[(idx_i, idx_j)] = [None] * N

		if cancel_var.get():
			return []

		# 各フレーム t / 各骨 i について min over j!=i を取る
		result = []
		for t in range(N):
			frame_map = {}
			for (idx_i, name_i, mesh_L_i, Ts_i) in bones_data:
				n_v = len(local_pts[idx_i])
				min_sd = np.full(n_v, np.inf, dtype=np.float32)
				for (idx_j, name_j, mesh_L_j, Ts_j) in bones_data:
					if idx_j == idx_i:
						continue
					dists = per_pair_dist.get((idx_i, idx_j))
					if dists is None or t >= len(dists) or dists[t] is None:
						continue
					if len(dists[t]) == n_v:
						min_sd = np.minimum(min_sd, dists[t])
				frame_map[idx_i] = min_sd
			result.append(frame_map)
		return result

	def _sim_view_create_control_panel(self, n_frames, frame_times, callbacks, features=None):
		"""hip と同じ再生コントロールウィンドウを作成 (N骨汎用)。

		Args:
			n_frames: 総フレーム数
			frame_times: 各フレームの時刻 [sec] (長さ n_frames)
			callbacks: dict {
				'on_frame_seek': fn(frame_idx)  ユーザーがバーをドラッグしたとき,
				'on_pause_toggle': fn(is_paused)  一時停止/再生切替後,
				'on_speed_change': fn(speed)  速度スライダー変更後,
				'on_csv_export': fn()  CSV出力ボタン (features['csv']=True のみ),
				'on_screenshot': fn()  スクリーンショット,
				'on_export_model': fn()  現在フレームのモデルを出力,
				'on_close': fn()  ウィンドウが閉じられたとき,
			}
			features: dict {'csv': bool, 'export_model': bool, 'screenshot': bool}

		Returns:
			dict {
				'window': tk.Toplevel,
				'frame_label': ttk.Label,     # "Frame: N/M | Time: X.XXXs"
				'actual_label': ttk.Label,    # "Actual: X.XXXs"
				'speed_label': ttk.Label,     # "Speed: X.XXx"
				'max_pent_label': ttk.Label,  # "Max Pent: X.XX mm"
				'playback_scale': tk.Scale,
				'speed_scale': tk.Scale,
				'pause_button': ttk.Button,
				'user_is_dragging': [bool],   # ドラッグ中フラグ (参照用)
				'is_programmatic_update': [bool],
			}
		"""
		if features is None:
			features = {'csv': True, 'export_model': True, 'screenshot': True}

		control_window = tk.Toplevel(self)
		control_window.title("再生コントロール")
		control_window.geometry("850x260")
		control_window.resizable(True, True)
		control_window.minsize(400, 200)
		control_window.attributes('-topmost', True)

		user_is_dragging = [False]
		is_programmatic_update = [False]

		# 情報表示行
		info_frame = ttk.Frame(control_window)
		info_frame.pack(pady=5, padx=10, fill=tk.X)
		frame_label = ttk.Label(info_frame,
			text=f"Frame: 0/{max(n_frames-1,0)} | Time: {frame_times[0]:.3f}s"
			     if n_frames > 0 else "Frame: 0/0",
			font=(self.ui_font_family, 10))
		frame_label.pack(side=tk.LEFT, padx=5)
		actual_label = ttk.Label(info_frame, text="Actual: 0.000s",
			font=(self.ui_font_family, 10), foreground="blue")
		actual_label.pack(side=tk.LEFT, padx=5)
		speed_label = ttk.Label(info_frame, text="Speed: 1.0x",
			font=(self.ui_font_family, 10), foreground="darkgreen")
		speed_label.pack(side=tk.LEFT, padx=5)
		max_pent_label = ttk.Label(info_frame, text="Max Pent: -- mm",
			font=(self.ui_font_family, 10), foreground="red")
		max_pent_label.pack(side=tk.LEFT, padx=5)

		# 再生バー
		def _on_press(event):
			user_is_dragging[0] = True

		def _on_release(event):
			user_is_dragging[0] = False

		def _on_scale_change(val):
			if is_programmatic_update[0]:
				return
			if user_is_dragging[0]:
				cb = callbacks.get('on_frame_seek')
				if cb is not None:
					try:
						cb(int(float(val)))
					except Exception as e:
						print(f"[sim_view] on_frame_seek 失敗: {e}")

		playback_scale = tk.Scale(
			control_window, from_=0, to=max(n_frames - 1, 0),
			orient=tk.HORIZONTAL, label="Frame Position",
			command=_on_scale_change)
		playback_scale.pack(pady=5, padx=10, fill=tk.X)
		playback_scale.bind("<ButtonPress-1>", _on_press)
		playback_scale.bind("<ButtonRelease-1>", _on_release)

		# 速度スライダー
		def _on_speed(val):
			try:
				speed = float(val)
				speed_label.config(text=f"Speed: {speed:.2f}x")
				cb = callbacks.get('on_speed_change')
				if cb is not None:
					cb(speed)
			except Exception as e:
				print(f"[sim_view] on_speed 失敗: {e}")

		speed_scale = tk.Scale(control_window, from_=0.25, to=10.0, resolution=0.25,
			orient=tk.HORIZONTAL, label="Playback Speed (0.25x - 10x)",
			command=_on_speed)
		speed_scale.set(1.0)
		speed_scale.pack(pady=5, padx=10, fill=tk.X)

		# ボタン行
		button_frame = ttk.Frame(control_window)
		button_frame.pack(pady=5)
		pause_button = ttk.Button(button_frame, text="一時停止", width=15)
		pause_button.pack(side=tk.LEFT, padx=5)

		def _toggle_pause():
			# 呼び側が実際の state を管理 (True/False は callback に任せる)
			cb = callbacks.get('on_pause_toggle')
			if cb is not None:
				try:
					new_paused = cb()
					pause_button.config(text="再生" if new_paused else "一時停止")
				except Exception as e:
					print(f"[sim_view] on_pause_toggle 失敗: {e}")

		pause_button.config(command=_toggle_pause)

		if features.get('csv') and callbacks.get('on_csv_export'):
			ttk.Button(button_frame, text="CSV出力", width=15,
			           command=callbacks['on_csv_export']).pack(side=tk.LEFT, padx=5)
		if features.get('export_model') and callbacks.get('on_export_model'):
			ttk.Button(button_frame, text="このモデルを出力", width=18,
			           command=callbacks['on_export_model']).pack(side=tk.LEFT, padx=5)
		if features.get('screenshot') and callbacks.get('on_screenshot'):
			ttk.Button(button_frame, text="スクリーンショット", width=18,
			           command=callbacks['on_screenshot']).pack(side=tk.LEFT, padx=5)

		def _on_close():
			cb = callbacks.get('on_close')
			if cb is not None:
				try:
					cb()
				except Exception:
					pass
			try:
				control_window.destroy()
			except Exception:
				pass

		control_window.protocol("WM_DELETE_WINDOW", _on_close)

		return {
			'window': control_window,
			'frame_label': frame_label,
			'actual_label': actual_label,
			'speed_label': speed_label,
			'max_pent_label': max_pent_label,
			'playback_scale': playback_scale,
			'speed_scale': speed_scale,
			'pause_button': pause_button,
			'user_is_dragging': user_is_dragging,
			'is_programmatic_update': is_programmatic_update,
		}

	def _ankle_accumulate_method_compare(self, det: dict, target_ids) -> None:
		"""各フレームで3手法の姿勢を貯めておく (実データでどれが一番安定かを後で実測するため)。

		runtime 専用。npz には保存しない。
		"""
		import numpy as np
		store = getattr(self, "_ankle_method_compare", None)
		if store is None:
			store = {}
			self._ankle_method_compare = store
		for aid in target_ids:
			slot = store.setdefault(int(aid), {"pnp": [], "corners": [], "fusion": []})
			d = det.get(aid)
			for key, field in (("pnp", "pose_pnp"), ("corners", "pose_corners"),
			                    ("fusion", "pose_fusion")):
				val = None if d is None else d.get(field)
				slot[key].append(np.asarray(val, dtype=float) if val is not None
				                 else np.full((4, 4), np.nan))

	@staticmethod
	def _ankle_rot_jitter(poses, fps: float = 15.0, lp_hz: float = 1.0):
		"""姿勢時系列の高周波回転ジッターを測る。

		強めに平滑化したものを「真の動き」とみなし、生との角度差をジッターとする。
		Returns: (median_deg, p95_deg, n_used) — 測れない場合は (None, None, 0)
		"""
		import numpy as np
		try:
			from scipy.signal import butter, filtfilt
			from scipy.spatial.transform import Rotation as _Rot
		except Exception:
			return None, None, 0
		P = np.asarray(poses, dtype=float)
		if P.ndim != 3 or len(P) < 60:
			return None, None, 0
		ok = np.all(np.isfinite(P.reshape(len(P), -1)), axis=1)
		vidx = np.where(ok)[0]
		if len(vidx) < 60:
			return None, None, 0
		quats = np.zeros((len(vidx), 4))
		for k, t in enumerate(vidx):
			Rm = P[t][:3, :3]
			try:
				U, _, Vt = np.linalg.svd(Rm)
				Rm = U @ Vt
				if np.linalg.det(Rm) < 0:
					Rm = U @ np.diag([1.0, 1.0, -1.0]) @ Vt
				quats[k] = _Rot.from_matrix(Rm).as_quat()
			except Exception:
				return None, None, 0
			if k > 0 and float(np.dot(quats[k], quats[k - 1])) < 0:
				quats[k] = -quats[k]
		nyq = max(fps / 2.0, 1e-6)
		wn = float(np.clip(lp_hz / nyq, 1e-3, 0.99))
		try:
			b, a = butter(2, wn, btype='low')
			pad = min(len(quats) - 1, 3 * max(len(a), len(b)))
			qs = np.stack([filtfilt(b, a, quats[:, c], padtype='odd', padlen=pad)
			               for c in range(4)], axis=1)
		except Exception:
			return None, None, 0
		qs /= np.maximum(np.linalg.norm(qs, axis=1, keepdims=True), 1e-12)
		dR = (_Rot.from_quat(quats) * _Rot.from_quat(qs).inv()).as_rotvec()
		jit = np.degrees(np.linalg.norm(dR, axis=1))
		return float(np.median(jit)), float(np.percentile(jit, 95)), int(len(vidx))

	def _ankle_report_method_comparison(self, bones_data, fps: float = 15.0):
		"""3手法 (solvePnPのみ / 4隅深度 / RGB+Depth融合) のジッターを実データで比較する。

		ハードを作り直す前に「深度が本当に効いているのか」を実測で確かめるための診断。
		"""
		store = getattr(self, "_ankle_method_compare", None)
		if not store:
			return
		self._ankle_safe_print("=" * 70)
		self._ankle_safe_print("[手法比較] 同じ実データで3手法の回転ジッターを実測")
		self._ankle_safe_print("=" * 70)
		labels = {"pnp": "solvePnP のみ (RGBのみ)",
		          "corners": "4隅深度",
		          "fusion": "RGB+Depth融合 (現行)"}
		for (idx, name, mesh_L, Ts) in bones_data:
			bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
			aid = int(bone.get("aruco_id", -1))
			slot = store.get(aid)
			if not slot:
				continue
			self._ankle_safe_print(f"  {name} (ID={aid}):")
			results = {}
			for key in ("pnp", "corners", "fusion"):
				med, p95, n = self._ankle_rot_jitter(slot.get(key, []), fps=fps)
				if med is None:
					self._ankle_safe_print(f"      {labels[key]:26s}: 測定不可")
					continue
				results[key] = med
				self._ankle_safe_print(
					f"      {labels[key]:26s}: 中央値 {med:6.3f}° / p95 {p95:6.3f}°  (n={n})")
			if len(results) >= 2:
				best = min(results, key=results.get)
				cur = results.get("fusion")
				if best != "fusion" and cur is not None:
					self._ankle_safe_print(
						f"      → この骨では [{labels[best]}] の方が "
						f"{cur / max(results[best], 1e-9):.1f}倍 安定しています")
				else:
					self._ankle_safe_print("      → 現行の RGB+Depth融合 が最良です")
		self._ankle_safe_print("=" * 70)

	def _ankle_report_pose_quality(self, cache, bones_data, fps: float = 15.0):
		"""実データからマーカーの角度ジッターを実測し、骨先端での振れ幅に換算して報告する。

		【なぜ必要か】
		マーカーは小さく (20mm)、骨は大きい (400mm超)。
		マーカーの角度誤差 δ は、マーカーから距離 L の骨表面で δ×L の変位に増幅される。
		20mm マーカーで 400mm の骨を制御するのは てこ比 15〜20:1 なので、
		わずかな角度ジッターが目に見えるガタつきになる。
		推測ではなく実測値で議論するための診断。

		【ジッターの測り方】
		生の回転を強めに平滑化したものを「真の動き」とみなし、
		生との差の高周波成分をジッターとする (静止区間が無くても測れる)。
		"""
		import numpy as np
		try:
			from scipy.signal import butter, filtfilt
			from scipy.spatial.transform import Rotation as _Rot
		except Exception:
			return
		cache_bones = cache.get("bones", {}) or {}
		self._ankle_safe_print("=" * 70)
		self._ankle_safe_print("[姿勢品質レポート] マーカー角度ジッター → 骨での振れ幅")
		self._ankle_safe_print("=" * 70)

		# 強めの low-pass (真の動きの推定用)。ロボット試験は低速なので 1Hz で十分。
		nyq = max(fps / 2.0, 1e-6)
		wn = float(np.clip(1.0 / nyq, 1e-3, 0.99))
		try:
			b, a = butter(2, wn, btype='low')
		except Exception:
			return

		for (idx, name, mesh_L, Ts) in bones_data:
			bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
			aid = int(bone.get("aruco_id", -1))
			if aid not in cache_bones:
				continue
			try:
				poses = np.asarray(cache_bones[aid]["poses"], dtype=float)
				det = np.asarray(cache_bones[aid]["detected"], dtype=bool)
				vidx = np.where(det)[0]
				if len(vidx) < 60:
					continue
				# 検出済みフレームの回転をクォータニオン化 (半球を揃える)
				quats = np.zeros((len(vidx), 4))
				for k, t in enumerate(vidx):
					Rm = poses[t][:3, :3]
					U, _, Vt = np.linalg.svd(Rm)
					Rm = U @ Vt
					if np.linalg.det(Rm) < 0:
						Rm = U @ np.diag([1.0, 1.0, -1.0]) @ Vt
					quats[k] = _Rot.from_matrix(Rm).as_quat()
					if k > 0 and float(np.dot(quats[k], quats[k - 1])) < 0:
						quats[k] = -quats[k]
				# 平滑化して「真の動き」を推定
				pad = min(len(quats) - 1, 3 * max(len(a), len(b)))
				qs = np.stack([filtfilt(b, a, quats[:, c], padtype='odd', padlen=pad)
				               for c in range(4)], axis=1)
				qs /= np.maximum(np.linalg.norm(qs, axis=1, keepdims=True), 1e-12)
				# 生との角度差 = ジッター
				R_raw = _Rot.from_quat(quats)
				R_smooth = _Rot.from_quat(qs)
				dR = (R_raw * R_smooth.inv()).as_rotvec()
				jit_deg = np.degrees(np.linalg.norm(dR, axis=1))
				jit_med = float(np.median(jit_deg))
				jit_p95 = float(np.percentile(jit_deg, 95))
				# 並進ジッター
				tr = poses[vidx][:, :3, 3]
				trs = np.stack([filtfilt(b, a, tr[:, c], padtype='odd', padlen=pad)
				                for c in range(3)], axis=1)
				jt_med = float(np.median(np.linalg.norm(tr - trs, axis=1)))

				# てこ比: マーカーから骨表面までの距離
				lever = 0.0
				T_L_Mk = bone.get("marker_to_bone_T")
				if T_L_Mk is not None and mesh_L is not None:
					try:
						t_mk = np.asarray(T_L_Mk, dtype=float)[:3, 3]
						pv_pts = np.asarray(mesh_L.points)
						if len(pv_pts) > 20000:
							pv_pts = pv_pts[::max(1, len(pv_pts) // 20000)]
						dists = np.linalg.norm(pv_pts - t_mk, axis=1)
						lever = float(np.percentile(dists, 95))
					except Exception:
						lever = 0.0
				swing_med = np.deg2rad(jit_med) * lever + jt_med
				swing_p95 = np.deg2rad(jit_p95) * lever + jt_med
				self._ankle_safe_print(
					f"  {name} (ID={aid}): 角度ジッター 中央値 {jit_med:.3f}° / p95 {jit_p95:.3f}°, "
					f"並進ジッター {jt_med:.3f} mm")

				# --- 光学的な限界との比較 ---
				# 平面マーカーの姿勢ジッターは、コーナー検出ノイズ sigma_px と
				# 画面上のマーカー見かけ半径 r_px から
				#     jitter ≈ COND * sigma_px / r_px
				# で決まる。COND は平面PnPの条件数 (傾きが弱い自由度である分の悪化) で、
				# 合成実験では 2.6 で安定していた。
				# ここでは実測ジッターから逆算して sigma_px を求め、
				# サブピクセル検出の実力上限 (0.03〜0.05px) と比べる。
				COND = 2.6
				SUBPIX_FLOOR_PX = 0.05
				try:
					K_i = cache.get("intrinsics", {}) or {}
					fx = float(K_i.get("fx", 645.9))
					msize = float(cache.get("marker_size_mm", 20.0))
					dist_mm = float(np.linalg.norm(poses[vidx[0]][:3, 3]))
					r_px = (msize / 2.0) * np.sqrt(2.0) * fx / max(dist_mm, 1e-6)
					sigma_px = np.tan(np.deg2rad(jit_med / COND)) * r_px
					floor_deg = COND * np.degrees(np.arctan2(SUBPIX_FLOOR_PX, max(r_px, 1e-6)))
					self._ankle_safe_print(
						f"      角検出: マーカー見かけ半径 {r_px:.1f} px "
						f"(実寸 {msize:.1f}mm @ {dist_mm:.0f}mm) "
						f"→ 実効コーナー精度 {sigma_px:.3f} px")
					if sigma_px <= SUBPIX_FLOOR_PX * 1.5:
						self._ankle_safe_print(
							f"      [評価] サブピクセル検出の実力上限 (~{SUBPIX_FLOOR_PX} px) に到達済み。"
							f"この構成での下限は {floor_deg:.3f}° で、ソフトでの改善余地はありません")
					else:
						self._ankle_safe_print(
							f"      [評価] 実力上限 {SUBPIX_FLOOR_PX} px なら {floor_deg:.3f}° まで下げられます "
							f"(現状 {jit_med:.3f}°)。照明・ピント・モーションブラーを確認してください")
					# 物理的に何を変えればどこまで行けるか (角度は倍率に反比例)
					self._ankle_safe_print("      物理的な改善案 (角度精度は見かけ半径に反比例):")
					opts = [("カメラを 150mm まで近づける", dist_mm / 150.0),
					        ("マーカーを 40mm にする", 40.0 / max(msize, 1e-6)),
					        ("ChArUco 5x5 40mm (16交点)", (40.0 / max(msize, 1e-6)) * np.sqrt(16.0 / 4.0))]
					for label, gain in opts:
						if gain <= 1.01:
							continue
						self._ankle_safe_print(
							f"        ・{label}: {jit_med / gain:.3f}° "
							f"→ 骨での振れ {np.deg2rad(jit_med / gain) * lever:.2f} mm")
				except Exception:
					pass
				if lever > 0:
					self._ankle_safe_print(
						f"      てこ長 {lever:.0f} mm (マーカー→骨表面 p95) "
						f"→ 骨での振れ 中央値 {swing_med:.2f} mm / p95 {swing_p95:.2f} mm")
					if swing_p95 > 3.0:
						need = np.degrees(np.arctan(1.0 / max(lever, 1e-6)))
						self._ankle_safe_print(
							f"      [評価] 目に見えるガタつきです。骨先端で 1mm 以内にするには "
							f"角度ジッターを {need:.2f}° 以下にする必要があります。"
							f"現状の {jit_p95:.2f}° からは マーカー大型化 or ChArUco 化 が必要です")
			except Exception:
				continue
		self._ankle_safe_print("=" * 70)

	@staticmethod
	def _ankle_hampel_inliers(x, window: int = 9, n_sigma: float = 3.0):
		"""Hampel フィルタで外れ値を検出し、inlier の bool マスクを返す。

		各点について、周囲 window 個の中央値・MAD を求め、
		|x - median| > n_sigma * 1.4826 * MAD なら外れ値とする。
		移動平均と違い、外れ値自身に引きずられないので誤検出の除去に向く。
		"""
		import numpy as np
		x = np.asarray(x, dtype=float)
		n = len(x)
		mask = np.ones(n, dtype=bool)
		if n < window or window < 3:
			return mask
		half = window // 2
		for i in range(n):
			lo = max(0, i - half)
			hi = min(n, i + half + 1)
			seg = x[lo:hi]
			med = float(np.median(seg))
			mad = float(np.median(np.abs(seg - med)))
			sigma = 1.4826 * mad
			if sigma > 1e-12 and abs(x[i] - med) > n_sigma * sigma:
				mask[i] = False
		return mask

	def _ankle_smooth_pose_series(self, bones_data, cutoff_hz: float = 5.0,
	                               frame_rate_hz: float = 30.0,
	                               detected_map=None,
	                               reject_outliers: bool = True,
	                               outlier_sigma: float = 3.0):
		"""骨のpose時系列を頑健に平滑化してガタつきを除去する。

		処理は3段構成:
		  1. **外れ値除去** — 並進3成分と回転(クォータニオン)4成分に Hampel フィルタをかけ、
		     突発的に飛んだフレームを「未検出」扱いに落とす。
		     (単純な low-pass だと外れ値が前後に滲んで波打つため、先に取り除く)
		  2. **欠損補間** — 未検出/外れ値のフレームを、前後の有効フレームから補間する。
		     並進は線形、回転は SLERP。従来の「直前値を保持」は階段状の段差を生むため廃止。
		  3. **平滑化** — 並進は各成分に Butterworth、回転はクォータニオン4成分に
		     Butterworth をかけてから正規化。
		     Rodrigues 軸角は ±π 付近で不連続になり回転軸の符号も反転しうるため、
		     半球を揃えたクォータニオンで扱う。
		     filtfilt (双方向) なので位相遅れは生じない。

		Args:
			bones_data: list of (idx, name, mesh_L, T_series)
			cutoff_hz: low-pass カットオフ周波数 [Hz]
			frame_rate_hz: 実フレームレート [Hz]。Nyquist の算出に使うので実測値を渡すこと
			detected_map: {bone_idx: bool配列} 検出済みフレームのマスク (None なら全フレーム有効)
			reject_outliers: True で Hampel による外れ値除去を行う
			outlier_sigma: 外れ値判定のしきい値 (小さいほど厳しい)

		Returns:
			(smoothed_bones_data, stats)
			stats: {bone_idx: {"outliers": int, "interpolated": int, "n": int}}
		"""
		import numpy as np
		from scipy.signal import butter, filtfilt
		from scipy.spatial.transform import Rotation as _Rot, Slerp

		nyq = max(frame_rate_hz / 2.0, 1e-6)
		wn = float(np.clip(cutoff_hz / nyq, 1e-3, 0.99))
		b, a = butter(N=4, Wn=wn, btype='low')
		min_len = 3 * max(len(a), len(b)) + 1

		def _lowpass(x):
			if len(x) <= min_len:
				return x
			try:
				return filtfilt(b, a, x, padtype='odd',
				                padlen=min(len(x) - 1, 3 * max(len(a), len(b))))
			except Exception:
				return x

		smoothed = []
		stats = {}
		for (idx, name, mesh_L, Ts) in bones_data:
			Ts = np.asarray(Ts, dtype=float)
			N = len(Ts)
			st = {"outliers": 0, "interpolated": 0, "n": N}
			if N < 5:
				smoothed.append((idx, name, mesh_L, Ts))
				stats[idx] = st
				continue

			# --- 有効フレームの初期マスク ---
			if detected_map is not None and idx in detected_map:
				valid = np.asarray(detected_map[idx], dtype=bool).copy()
				if len(valid) != N:
					valid = np.ones(N, dtype=bool)
			else:
				valid = np.ones(N, dtype=bool)
			# 姿勢が NaN/特異なフレームも無効化
			for t in range(N):
				if not np.all(np.isfinite(Ts[t])):
					valid[t] = False
			if valid.sum() < 4:
				smoothed.append((idx, name, mesh_L, Ts))
				stats[idx] = st
				continue

			# --- 並進とクォータニオンを取り出す ---
			trans = Ts[:, :3, 3].copy()
			quats = np.zeros((N, 4), dtype=float)   # (x, y, z, w)
			for t in range(N):
				if not valid[t]:
					continue
				try:
					Rm = Ts[t][:3, :3]
					# 数値誤差で非直交になっている場合に備えて直交化してから変換
					U, _, Vt = np.linalg.svd(Rm)
					Rm = U @ Vt
					if np.linalg.det(Rm) < 0:
						Rm = U @ np.diag([1.0, 1.0, -1.0]) @ Vt
					quats[t] = _Rot.from_matrix(Rm).as_quat()
				except Exception:
					valid[t] = False
			if valid.sum() < 4:
				smoothed.append((idx, name, mesh_L, Ts))
				stats[idx] = st
				continue

			# --- 半球を揃える (q と -q は同じ回転。符号が飛ぶと平滑化が壊れる) ---
			vidx = np.where(valid)[0]
			for k in range(1, len(vidx)):
				if float(np.dot(quats[vidx[k]], quats[vidx[k - 1]])) < 0.0:
					quats[vidx[k]] = -quats[vidx[k]]

			# --- 1. 外れ値除去 (Hampel) ---
			if reject_outliers and len(vidx) >= 9:
				bad = np.zeros(N, dtype=bool)
				for c in range(3):
					m = self._ankle_hampel_inliers(trans[vidx, c], window=9, n_sigma=outlier_sigma)
					bad[vidx[~m]] = True
				for c in range(4):
					m = self._ankle_hampel_inliers(quats[vidx, c], window=9, n_sigma=outlier_sigma)
					bad[vidx[~m]] = True
				# 有効フレームが極端に減る場合は除去しない (誤検出の暴走を防ぐ)
				if bad.sum() < 0.3 * len(vidx):
					st["outliers"] = int(bad.sum())
					valid[bad] = False
					vidx = np.where(valid)[0]

			if len(vidx) < 4:
				smoothed.append((idx, name, mesh_L, Ts))
				stats[idx] = st
				continue

			# --- 2. 欠損補間 (並進=線形 / 回転=SLERP) ---
			st["interpolated"] = int(N - len(vidx))
			all_t = np.arange(N, dtype=float)
			trans_full = np.stack(
				[np.interp(all_t, vidx.astype(float), trans[vidx, c]) for c in range(3)], axis=1)
			try:
				rots = _Rot.from_quat(quats[vidx])
				slerp = Slerp(vidx.astype(float), rots)
				# Slerp は補間区間外を扱えないので、両端は最近傍の有効値でクランプ
				t_clamped = np.clip(all_t, float(vidx[0]), float(vidx[-1]))
				quats_full = slerp(t_clamped).as_quat()
			except Exception:
				quats_full = np.stack(
					[np.interp(all_t, vidx.astype(float), quats[vidx, c]) for c in range(4)], axis=1)
			# 補間後も半球を揃え直す
			for t in range(1, N):
				if float(np.dot(quats_full[t], quats_full[t - 1])) < 0.0:
					quats_full[t] = -quats_full[t]

			# --- 3. 平滑化 ---
			trans_s = np.stack([_lowpass(trans_full[:, c]) for c in range(3)], axis=1)
			quats_s = np.stack([_lowpass(quats_full[:, c]) for c in range(4)], axis=1)
			norms = np.linalg.norm(quats_s, axis=1, keepdims=True)
			norms[norms < 1e-12] = 1.0
			quats_s = quats_s / norms

			# --- 復元 ---
			Ts_new = np.zeros_like(Ts)
			try:
				Rs = _Rot.from_quat(quats_s).as_matrix()
			except Exception:
				Rs = np.repeat(np.eye(3)[None, :, :], N, axis=0)
			for t in range(N):
				Ts_new[t] = np.eye(4)
				Ts_new[t, :3, :3] = Rs[t]
				Ts_new[t, :3, 3] = trans_s[t]
			smoothed.append((idx, name, mesh_L, Ts_new))
			stats[idx] = st
		return smoothed, stats

	def on_ankle_animate(self) -> None:
		"""ArUco姿勢時系列と骨キャリブから、N骨のアニメーション + マルチヒートマップを表示。

		hip の on_animate と同等の再生コントロール (Play/Pause/速度/フレームバー/CSV/スクショ) を持つ。
		描画エンジンは共通ヘルパ _sim_view_* を使用。
		"""
		import numpy as np
		# --- 1. 前提チェック ---
		cache = self._ankle_get_current_cache()
		if not cache:
			messagebox.showwarning("シミュレーション",
				"姿勢時系列がありません。④で「ArUco検出+PnP実行」または「姿勢時系列を読込」してください。")
			return
		if not self.ankle_bones:
			messagebox.showwarning("シミュレーション", "骨リストが空です。")
			return

		# --- 2. 骨ごとに W系姿勢時系列を構築 ---
		animatable, N, warns = self._ankle_build_bone_transforms(cache)
		if not animatable:
			messagebox.showwarning("シミュレーション",
				"アニメ可能な骨がありません。\n\n" + "\n".join(warns))
			return
		if warns:
			print("[ankle animate] 警告:\n  " + "\n  ".join(warns))

		# bones_data: list of (idx, name, mesh_L, T_series)
		bones_data = []
		for (idx, mesh_L, Ts) in animatable:
			name = self.ankle_bones[idx].get("name", f"骨{idx+1}") if idx < len(self.ankle_bones) else f"骨{idx+1}"
			bones_data.append((idx, name, mesh_L, Ts))
		n_bones = len(bones_data)

		# --- 【診断ログ】 T_L←Mk (marker_to_bone_T) の平行移動量を出力 ---
		# 骨があり得ないくらい離れる場合、この translation が異常値になっていないか確認する
		print("=" * 70)
		print("[ankle animate] キャリブレーション診断")
		print("=" * 70)
		for (idx, name, mesh_L, Ts) in bones_data:
			bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
			T_L_Mk = bone.get("marker_to_bone_T")
			if T_L_Mk is None:
				print(f"  {name}: T_L←Mk = None (原プラン or 未キャリブ)")
				continue
			T = np.asarray(T_L_Mk, dtype=float)
			tr = T[:3, 3]
			t_norm = float(np.linalg.norm(tr))
			# メッシュの bounding box 対角長 = 骨のスケール
			try:
				b = np.asarray(mesh_L.bounds).reshape(3, 2)
				bone_diag = float(np.linalg.norm(b[:, 1] - b[:, 0]))
			except Exception:
				bone_diag = 0.0
			warn = ""
			if t_norm > bone_diag * 2 and bone_diag > 0:
				warn = "  ⚠️ 警告: マーカーが骨サイズの2倍以上離れている → キャリブ異常の可能性!"
			print(f"  {name}: T_L←Mk translation = ({tr[0]:+.2f}, {tr[1]:+.2f}, {tr[2]:+.2f}) mm, "
			      f"norm = {t_norm:.2f} mm, 骨サイズ = {bone_diag:.1f} mm{warn}")
			# 最終フレーム t=0 の T_C←bone の位置も出力
			try:
				pos0 = Ts[0][:3, 3]
				print(f"    → t=0 での骨位置 (W系): ({pos0[0]:+.1f}, {pos0[1]:+.1f}, {pos0[2]:+.1f}) mm")
			except Exception:
				pass
		# 骨間距離の診断
		if n_bones >= 2:
			for i in range(n_bones):
				for j in range(i+1, n_bones):
					try:
						p_i = bones_data[i][3][0][:3, 3]
						p_j = bones_data[j][3][0][:3, 3]
						d = float(np.linalg.norm(p_i - p_j))
						warn = ""
						if d > 300:
							warn = "  ⚠️ 300mm 超は解剖学的にありえません (キャリブ or W選択の問題)"
						elif d < 5:
							warn = "  ⚠️ 5mm 以下は近すぎ (キャリブ or マーカー重複の可能性)"
						print(f"  骨間距離 [{bones_data[i][1]} ↔ {bones_data[j][1]}] "
						      f"@ t=0: {d:.1f} mm{warn}")
					except Exception:
						pass
		# マーカーのカメラ距離診断 (D405 推奨動作距離: ~7-50cm)
		cache_bones_diag = cache.get("bones", {}) or {}
		for (idx, name, mesh_L, Ts) in bones_data:
			bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
			aid = int(bone.get("aruco_id", -1))
			if aid not in cache_bones_diag:
				continue
			try:
				poses_C = np.asarray(cache_bones_diag[aid]["poses"], dtype=float)
				det = np.asarray(cache_bones_diag[aid]["detected"], dtype=bool)
				# 最初の検出済みフレームの marker 位置を取得
				valid_idx = np.where(det)[0]
				if len(valid_idx) == 0:
					continue
				pos = poses_C[valid_idx[0]][:3, 3]
				dist_from_cam = float(np.linalg.norm(pos))
				warn = ""
				if dist_from_cam > 500:
					warn = "  ⚠️ D405 推奨動作距離 (~50cm) を大きく超える → 精度大幅劣化"
				elif dist_from_cam < 70:
					warn = "  ⚠️ D405 最短距離 (~7cm) より近い → 深度取得不可"
				print(f"  {name} マーカー @ frame {valid_idx[0]}: "
				      f"位置 = ({pos[0]:+.1f}, {pos[1]:+.1f}, {pos[2]:+.1f}) mm, "
				      f"カメラからの距離 = {dist_from_cam:.1f} mm{warn}")
			except Exception:
				pass
		print("=" * 70)

		# --- 【診断モード】 メッシュ重心をマーカー位置に強制 ---
		# 【問題】mesh 座標系の原点が骨実体から遠い + キャリブが不正確な場合、
		# T_C←bone · v_mesh の変換で geometry が大きく離れて表示される。
		# 【診断】T_L←Mk を「centroid → 単位変換」で上書きし、mesh 重心が
		# 直接マーカー位置に来るように強制する (元のキャリブ情報は無視)。
		# これで骨の見た目位置 = マーカーの物理位置 になり、
		# ArUco 検出そのものの妥当性を確認できる。
		try:
			force_centroid_to_marker = bool(self.ankle_recenter_meshes.get())
		except Exception:
			force_centroid_to_marker = True
		if force_centroid_to_marker:
			print("[ankle animate] 🔧 診断モード: 骨重心をマーカー位置に強制配置 ON")
			print("  ※ 元のキャリブ情報 (T_L←Mk) は無視されます。マーカー検出の妥当性検証用。")
			cache_bones_force = cache.get("bones", {}) or {}
			forced = []
			try:
				mode = str(self.ankle_workflow_mode.get())
			except Exception:
				mode = "self_pose"
			ref = int(self.ankle_ref_frame.get()) if hasattr(self, 'ankle_ref_frame') else 0
			for (idx, name, mesh_L, Ts) in bones_data:
				bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
				aid = int(bone.get("aruco_id", -1))
				if aid not in cache_bones_force:
					forced.append((idx, name, mesh_L, Ts))
					continue
				try:
					b_cache = cache_bones_force[aid]
					poses_C = np.asarray(b_cache["poses"], dtype=float)   # T_C←Mk(t)
					det = np.asarray(b_cache["detected"], dtype=bool)
					# 重心をマーカー位置に対応させる、新しい T_L←Mk_new
					c = np.asarray(mesh_L.points).mean(axis=0)
					T_L_Mk_new = np.eye(4)
					T_L_Mk_new[:3, 3] = c  # 重心が「新しい bone-local 座標系での marker 位置」
					T_Mk_L_new = np.linalg.inv(T_L_Mk_new)
					# 各フレーム: T_C←bone = T_C←Mk · inv(T_L←Mk_new)
					N_f = len(Ts)
					Ts_new = np.zeros((N_f, 4, 4), dtype=float)
					last_valid = None
					for t in range(N_f):
						if t < len(det) and bool(det[t]):
							Ts_new[t] = poses_C[t] @ T_Mk_L_new
							last_valid = Ts_new[t]
						else:
							Ts_new[t] = last_valid if last_valid is not None else np.eye(4)
					forced.append((idx, name, mesh_L, Ts_new))
					print(f"  {name}: 重心 ({c[0]:+.1f}, {c[1]:+.1f}, {c[2]:+.1f}) mm を マーカー位置に強制配置")
				except Exception as e:
					print(f"  {name}: 強制配置失敗 ({e}) — 元 Ts のまま")
					forced.append((idx, name, mesh_L, Ts))
			bones_data = forced
			# 再度骨間距離を確認 (これは物理マーカー位置の差 = ArUco 検出精度の指標)
			if n_bones >= 2:
				for i in range(n_bones):
					for j in range(i+1, n_bones):
						try:
							p_i = bones_data[i][3][0][:3, 3]
							p_j = bones_data[j][3][0][:3, 3]
							d = float(np.linalg.norm(p_i - p_j))
							note = ""
							if d > 300:
								note = "  ⚠️ マーカー間距離自体が大きい → ArUco検出の Z軸精度低下の可能性 (D405 は近距離推奨)"
							print(f"  [強制配置後] マーカー距離 [{bones_data[i][1]} ↔ {bones_data[j][1]}] @ t=0: {d:.1f} mm{note}")
						except Exception:
							pass
			print("=" * 70)

		# --- 時系列平滑化 (外れ値除去 → 欠損補間 → low-pass) — ガタガタ削減 ---
		try:
			smooth_enabled = bool(self.ankle_smooth_enable.get())
		except Exception:
			smooth_enabled = True  # 既定 ON
		try:
			smooth_cutoff_hz = float(self.ankle_smooth_cutoff_hz.get())
		except Exception:
			smooth_cutoff_hz = 3.0
		try:
			reject_outliers = bool(self.ankle_reject_outliers.get())
		except Exception:
			reject_outliers = True

		# 実フレームレートをタイムスタンプから推定 (15fps 録画を 30fps と誤認すると
		# Nyquist が2倍になり、指定カットオフの半分しか効かない)
		fps_est = 30.0
		try:
			if len(frame_times) > 2:
				dts = np.diff(np.asarray(frame_times, dtype=float))
				dts = dts[(dts > 1e-6) & np.isfinite(dts)]
				if len(dts) > 0:
					fps_est = float(np.clip(1.0 / float(np.median(dts)), 1.0, 240.0))
		except Exception:
			fps_est = 30.0

		# 骨ごとの検出マスク (未検出フレームは補間対象にする)
		detected_map = {}
		try:
			cache_bones_sm = cache.get("bones", {}) or {}
			for (idx, name, mesh_L, Ts) in bones_data:
				bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
				aid = int(bone.get("aruco_id", -1))
				if aid in cache_bones_sm:
					det_arr = np.asarray(cache_bones_sm[aid].get("detected"), dtype=bool)
					if len(det_arr) == len(Ts):
						detected_map[idx] = det_arr
		except Exception:
			detected_map = {}

		# 平滑化の前に、生データの品質を実測して報告する
		try:
			self._ankle_report_pose_quality(cache, bones_data, fps=fps_est)
		except Exception as e:
			print(f"[ankle animate] 姿勢品質レポート失敗: {e}")
		try:
			self._ankle_report_method_comparison(bones_data, fps=fps_est)
		except Exception as e:
			print(f"[ankle animate] 手法比較レポート失敗: {e}")

		if smooth_enabled and N >= 20:
			try:
				bones_data, sm_stats = self._ankle_smooth_pose_series(
					bones_data,
					cutoff_hz=smooth_cutoff_hz,
					frame_rate_hz=fps_est,
					detected_map=detected_map,
					reject_outliers=reject_outliers)
				print(f"[ankle animate] 平滑化 適用: cutoff={smooth_cutoff_hz} Hz, "
				      f"実測フレームレート={fps_est:.1f} fps "
				      f"(Nyquist={fps_est/2:.1f} Hz)")
				for (idx, name, mesh_L, Ts) in bones_data:
					s = sm_stats.get(idx, {})
					if s:
						print(f"    {name}: 外れ値除去 {s.get('outliers', 0)} 件 / "
						      f"補間 {s.get('interpolated', 0)} フレーム / 全 {s.get('n', 0)} フレーム")
				if smooth_cutoff_hz > fps_est / 2.0:
					print(f"    [警告] カットオフ {smooth_cutoff_hz} Hz が Nyquist "
					      f"{fps_est/2:.1f} Hz を超えています → 平滑化はほぼ効きません。"
					      f"{max(1.0, fps_est/6.0):.1f} Hz 程度まで下げてください")
			except Exception as e:
				print(f"[ankle animate] 平滑化失敗 ({e}) — 生データで続行")
				import traceback; traceback.print_exc()

		# --- 骨固定モード: 選択骨があれば全骨に inv(T_fixed(t)) を前掛け ---
		# 結果: fixed 骨は常に identity (静止)、他の骨は fixed 骨の座標系での相対姿勢
		fixed_idx = None
		for (idx, name, mesh_L, Ts) in bones_data:
			bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
			if bone.get("fixed"):
				fixed_idx = idx
				break
		if fixed_idx is not None:
			# 固定骨の T_series を取得
			T_fixed_series = None
			for (idx, _, _, Ts) in bones_data:
				if idx == fixed_idx:
					T_fixed_series = Ts
					break
			if T_fixed_series is not None:
				N_frames = len(T_fixed_series)
				T_fixed_inv = np.zeros((N_frames, 4, 4), dtype=float)
				for t in range(N_frames):
					try:
						T_fixed_inv[t] = np.linalg.inv(T_fixed_series[t])
					except np.linalg.LinAlgError:
						T_fixed_inv[t] = np.eye(4)
				# 全骨に前掛け
				new_bones_data = []
				for (idx, name, mesh_L, Ts) in bones_data:
					Ts_new = np.zeros_like(Ts)
					for t in range(N_frames):
						Ts_new[t] = T_fixed_inv[t] @ Ts[t]
					new_bones_data.append((idx, name, mesh_L, Ts_new))
				bones_data = new_bones_data
				fixed_name = self.ankle_bones[fixed_idx].get("name", f"骨{fixed_idx+1}")
				print(f"[ankle animate] 骨固定モード: '{fixed_name}' を基準に他骨の相対運動を表示")

		# フレーム時刻 (キャッシュにあれば使う、なければ 1/30s 刻み)
		ts_cache = cache.get("timestamps", None)
		if ts_cache is not None and len(ts_cache) >= N:
			frame_times = [float(ts_cache[t]) for t in range(N)]
		else:
			frame_times = [t / 30.0 for t in range(N)]

		# --- 3. 事前計算ダイアログ (膝/股と共通のダイアログを流用) ---
		heatmap_data = []  # list of dict {bone_idx: distances} per frame
		if n_bones >= 2:
			# 膝/股と同じ _show_precompute_dialog を流用 (has_cartilage=False で FEMオプション無効)
			progress_window, update_progress, cancel_var, options_dict, start_var, skip_var = \
				self._show_precompute_dialog(N, has_cartilage=False)
			# モーダル待機 (膝/股と同じパターン)
			while progress_window.winfo_exists():
				if start_var.get() or skip_var.get() or cancel_var.get():
					break
				try:
					progress_window.update()
				except Exception:
					break
				time.sleep(0.01)
			if cancel_var.get():
				try:
					progress_window.destroy()
				except Exception:
					pass
				return
			if start_var.get() and options_dict.get('enable_precompute', tk.BooleanVar(value=True)).get():
				# --- メッシュ簡略化 (膝/股の use_simplify オプションと同等の挙動) ---
				# 膝/股は関節領域 (10-20k点) を使うため高速。ankle は骨モデル全体 (数十万〜数百万点)
				# を使うと _precompute_heatmaps_o3d のチャンクサイズが小さくなり大幅減速する。
				# → use_simplify=True (既定) なら ~15k点に間引く。視覚的にはほぼ変わらず、計算は
				#   膝/股と同スケールになる。
				use_simplify = True
				try:
					use_simplify = bool(options_dict['use_simplify'].get())
				except Exception:
					pass
				TARGET_VERTS = 15000 if use_simplify else 60000
				simplified_bones_data = []
				for (idx, name, mesh_L, Ts) in bones_data:
					mesh_use = mesh_L
					try:
						# 表面抽出→三角形化→頂点数が多ければ decimate
						surf = mesh_L.extract_surface() if hasattr(mesh_L, 'extract_surface') else mesh_L
						surf = surf.triangulate()
						n_v = surf.n_points
						if n_v > TARGET_VERTS:
							reduction = 1.0 - (TARGET_VERTS / n_v)
							try:
								mesh_use = surf.decimate(target_reduction=float(reduction), preserve_topology=False)
								mesh_use = mesh_use.triangulate()
								print(f"[ankle animate] {name}: {n_v} → {mesh_use.n_points} 頂点 に間引き (reduction={reduction:.3f})")
							except Exception as de:
								print(f"[ankle animate] {name}: decimate失敗 ({de}) → 表面のみ使用")
								mesh_use = surf
						else:
							mesh_use = surf
							print(f"[ankle animate] {name}: {n_v} 頂点 (間引き不要)")
					except Exception as e:
						print(f"[ankle animate] {name}: 簡略化失敗 ({e}) → 元メッシュ使用")
					simplified_bones_data.append((idx, name, mesh_use, Ts))
				# 元の bones_data を差し替え (以降のアクター作成・表示にも簡略化メッシュを使用)
				bones_data = simplified_bones_data

				# 計算実行 (膝/股と同じ高速エンジン _precompute_heatmaps_o3d をペア毎に流用)
				try:
					heatmap_data = self._sim_view_precompute_multi_heatmap(
						bones_data,
						update_progress=update_progress,
						cancel_var=cancel_var)
					print(f"[ankle animate] マルチヒートマップ計算完了: {len(heatmap_data)} frames")
				except Exception as e:
					print(f"[ankle animate] ヒートマップ計算失敗: {e}")
					import traceback; traceback.print_exc()
					heatmap_data = []
			try:
				progress_window.destroy()
			except Exception:
				pass

		# --- 4. プロッター作成 (hip と同じスタイル) ---
		pv.global_theme.allow_empty_mesh = True
		sw = self.winfo_screenwidth()
		sh = self.winfo_screenheight()
		anim_plotter = pv.Plotter(title="ankle: シミュレーション",
		                          window_size=(int(sw * 0.9), int(sh * 0.9)))
		anim_plotter.set_background("white")

		# 初期スキャン (半透明・参考表示)
		scan_path = self.ankle_initial_scan_path.get().strip()
		if scan_path:
			try:
				anim_plotter.add_mesh(pv.read(scan_path), color="lightgray",
				                       opacity=0.15, smooth_shading=True, name='initial_scan')
			except Exception as e:
				print(f"[ankle animate] 初期スキャン表示失敗: {e}")

		# --- 5. 各骨のアクター作成 (ヒートマップあり=scalars で色付け、なし=単色) ---
		heatmap_enabled = bool(heatmap_data)
		hm_cmap = self._sim_view_heatmap_cmap()
		CLIM_LO, CLIM_HI = -10.0, 0.0

		actors = {}                 # bone_idx -> actor (骨本体、SetUserMatrix で駆動)
		mesh_L_map = {}             # bone_idx -> mesh_L (ローカル)
		anim_first_T_inv = {}       # bone_idx -> inv(Ts[0])
		bone_colors = {}            # bone_idx -> hex color

		for (idx, name, mesh_L, Ts) in bones_data:
			mesh_L_map[idx] = mesh_L
			color = self._ankle_color_of(idx)
			bone_colors[idx] = color
			# 初期姿勢メッシュ (Ts[0] を適用)
			m0 = self._ankle_apply_T_to_mesh(mesh_L, Ts[0])
			if heatmap_enabled:
				# scalars を持たせる (初期フレームの distances)
				d0 = heatmap_data[0].get(idx)
				if d0 is None or len(d0) != m0.n_points:
					d0 = np.full(m0.n_points, np.inf, dtype=np.float32)
				m0['distance'] = d0
				actor = anim_plotter.add_mesh(m0, scalars='distance', cmap=hm_cmap,
				                              clim=[CLIM_LO, CLIM_HI], opacity=1.0,
				                              smooth_shading=True, show_edges=False,
				                              name=f'bone_{idx}',
				                              scalar_bar_args={'title': 'distance [mm]', 'color': 'black'} if idx == bones_data[0][0] else None)
				# 離間 (>0) は骨本来の色、めり込み (< -10) は 濃赤
				try:
					def _hex_to_rgb01(h):
						h = h.lstrip('#')
						return (int(h[0:2], 16) / 255.0,
						        int(h[2:4], 16) / 255.0,
						        int(h[4:6], 16) / 255.0)
					r, g, b = _hex_to_rgb01(color)
					lut = actor.GetMapper().GetLookupTable()
					lut.SetUseAboveRangeColor(True)
					lut.SetAboveRangeColor(r, g, b, 1.0)
					lut.SetUseBelowRangeColor(True)
					lut.SetBelowRangeColor(0.5, 0.0, 0.0, 1.0)
				except Exception:
					pass
				# 距離が inf (他骨無し) は透明
				try:
					mp = actor.GetMapper()
					mp.SetResolveCoincidentTopologyToPolygonOffset()
					mp.SetRelativeCoincidentTopologyPolygonOffsetParameters(-2.0, -100.0)
				except Exception:
					pass
			else:
				actor = anim_plotter.add_mesh(m0, color=color, opacity=1.0,
				                              smooth_shading=True, show_edges=False,
				                              name=f'bone_{idx}', label=name)
			actors[idx] = actor
			anim_first_T_inv[idx] = np.linalg.inv(Ts[0])

		# --- 6. アクターに SetUserMatrix を適用するヘルパ ---
		def _apply_matrix(actor, T4):
			import vtk
			m = vtk.vtkMatrix4x4()
			for i in range(4):
				for j in range(4):
					m.SetElement(i, j, float(T4[i, j]))
			try:
				actor.SetUserMatrix(m)
			except Exception:
				pass

		# --- 7. アニメーション状態変数 (hip と同じ形) ---
		current_frame = [0]
		actual_frame_counter = [0]
		animation_start_time = [None]
		playback_speed = [1.0]
		last_scale_update_time = [0.0]
		scale_update_interval = [0.2]
		after_id = [None]
		is_animation_active = [True]
		is_paused = [False]
		is_seeking = [False]

		# 6軸情報テキスト (upper_right)
		six_axis_text_actor = [None]

		# --- 8. show_frame: 1フレーム描画 ---
		def show_frame(frame_idx, force_render=False):
			if not force_render and not is_animation_active[0]:
				return
			if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
				is_animation_active[0] = False
				return
			try:
				fi = int(frame_idx) % max(N, 1)
				# 各骨アクターに (T(t) * inv(T(0))) を適用
				for (idx, name, mesh_L, Ts) in bones_data:
					dT = Ts[fi] @ anim_first_T_inv[idx]
					_apply_matrix(actors[idx], dT)
				# ヒートマップ scalars 更新
				max_pent = 0.0
				if heatmap_enabled and fi < len(heatmap_data):
					frame_hm = heatmap_data[fi]
					for (idx, name, mesh_L, Ts) in bones_data:
						d = frame_hm.get(idx)
						if d is None:
							continue
						actor = actors[idx]
						try:
							pd_in = actor.GetMapper().GetInput()
							arr = pd_in.GetPointData().GetScalars()
							if arr is not None and arr.GetNumberOfTuples() == len(d):
								for j in range(len(d)):
									arr.SetValue(j, float(d[j]))
								arr.Modified()
						except Exception:
							pass
						# 最大めり込み量 (負値の絶対値)
						try:
							dmin = float(np.nanmin(d))
							if dmin < 0 and abs(dmin) > max_pent:
								max_pent = abs(dmin)
						except Exception:
							pass
				# 情報ラベル更新 (制御パネル)
				try:
					widgets = ctrl_widgets  # closure から参照
					widgets['frame_label'].config(
						text=f"Frame: {fi}/{max(N-1,0)} | Time: {frame_times[fi]:.3f}s")
					widgets['max_pent_label'].config(
						text=f"Max Pent: {max_pent:.2f} mm" if heatmap_enabled else "Max Pent: -- mm")
				except Exception:
					pass
				# ArUco マーカー軸を毎フレーム更新
				try:
					_update_markers(fi)
				except Exception:
					pass
				# 6軸表示 (骨1のフレームに対する骨2の相対姿勢を表示、N>=2 のとき)
				try:
					if n_bones >= 2:
						T_A = bones_data[0][3][fi]
						T_B = bones_data[1][3][fi]
						T_rel = np.linalg.inv(T_A) @ T_B
						R = T_rel[:3, :3]
						t_vec = T_rel[:3, 3]
						# ZYX Euler (deg)
						sy = float(np.sqrt(R[0,0]**2 + R[1,0]**2))
						if sy > 1e-6:
							rz = float(np.degrees(np.arctan2(R[1,0], R[0,0])))
							ry = float(np.degrees(np.arctan2(-R[2,0], sy)))
							rx = float(np.degrees(np.arctan2(R[2,1], R[2,2])))
						else:
							rz = float(np.degrees(np.arctan2(-R[0,1], R[1,1])))
							ry = float(np.degrees(np.arctan2(-R[2,0], sy)))
							rx = 0.0
						txt = (f"骨A→B 相対姿勢\n"
						       f"Rx: {rx:+7.2f}°\n"
						       f"Ry: {ry:+7.2f}°\n"
						       f"Rz: {rz:+7.2f}°\n"
						       f"tx: {t_vec[0]:+7.2f} mm\n"
						       f"ty: {t_vec[1]:+7.2f} mm\n"
						       f"tz: {t_vec[2]:+7.2f} mm")
						if six_axis_text_actor[0] is None:
							six_axis_text_actor[0] = anim_plotter.add_text(
								txt, position='upper_right', font_size=11,
								color='black', font='courier')
						else:
							try:
								six_axis_text_actor[0].SetText(3, txt)
							except Exception:
								pass
				except Exception:
					pass
				# 描画
				if force_render or is_animation_active[0]:
					try:
						if hasattr(anim_plotter, 'render_window') and anim_plotter.render_window:
							anim_plotter.render()
					except Exception:
						pass
			except Exception as e:
				print(f"[ankle animate] フレーム{frame_idx} 描画失敗: {e}")
				traceback.print_exc()

		# --- 9. 制御パネルのコールバック ---
		def _cb_frame_seek(fi):
			is_seeking[0] = True
			current_frame[0] = fi
			show_frame(fi, force_render=True)
			try:
				time_val = frame_times[fi]
				actual_frame_counter[0] = int(time_val / 0.005)
				actual_time = actual_frame_counter[0] * 0.005
				ctrl_widgets['actual_label'].config(text=f"Actual: {actual_time:.3f}s")
				if animation_start_time[0] is not None:
					animation_start_time[0] = time.time() - (actual_time / playback_speed[0])
			except Exception:
				pass
			is_seeking[0] = False

		def _cb_pause_toggle():
			is_paused[0] = not is_paused[0]
			if not is_paused[0]:
				actual_time = actual_frame_counter[0] * 0.005
				animation_start_time[0] = time.time() - (actual_time / playback_speed[0])
			return is_paused[0]

		def _cb_speed_change(speed):
			playback_speed[0] = speed
			if animation_start_time[0] is not None and not is_paused[0]:
				actual_time = actual_frame_counter[0] * 0.005
				animation_start_time[0] = time.time() - (actual_time / speed)

		def _cb_csv_export():
			"""マルチヒートマップの Max Penetration とフレーム時刻を CSV に出力。"""
			if not heatmap_data:
				messagebox.showwarning("CSV出力", "ヒートマップ事前計算がありません。")
				return
			fp = filedialog.asksaveasfilename(
				title="CSV出力", defaultextension=".csv",
				filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")])
			if not fp:
				return
			try:
				import csv
				with open(fp, 'w', newline='', encoding='utf-8-sig') as f:
					w = csv.writer(f)
					header = ["Frame", "Time [s]"]
					for (idx, name, _, _) in bones_data:
						header.append(f"{name}_min_dist [mm]")
						header.append(f"{name}_max_pen [mm]")
					w.writerow(header)
					for t in range(N):
						row = [t, f"{frame_times[t]:.4f}"]
						for (idx, name, _, _) in bones_data:
							d = heatmap_data[t].get(idx) if t < len(heatmap_data) else None
							if d is None or len(d) == 0:
								row.extend(["", ""])
							else:
								finite = d[np.isfinite(d)]
								if finite.size == 0:
									row.extend(["", ""])
								else:
									dmin = float(np.min(finite))
									row.append(f"{dmin:.4f}")
									row.append(f"{abs(dmin):.4f}" if dmin < 0 else "0.0000")
						w.writerow(row)
				messagebox.showinfo("CSV出力完了", f"保存しました:\n{fp}")
			except Exception as e:
				messagebox.showerror("CSV出力失敗", f"CSV保存に失敗しました:\n{e}")

		def _cb_export_model():
			"""現在フレームの各骨 (変換後) を1つのメッシュに結合して保存。"""
			was_paused = is_paused[0]
			is_paused[0] = True
			try:
				fp = filedialog.asksaveasfilename(
					title=f"現在のモデルを出力 (Frame {current_frame[0]})",
					defaultextension=".stl",
					filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj")],
					initialfile=f"ankle_frame_{current_frame[0]}.stl")
				if not fp:
					return
				fi = current_frame[0]
				merged = None
				for (idx, name, mesh_L, Ts) in bones_data:
					m_W = self._ankle_apply_T_to_mesh(mesh_L, Ts[fi])
					merged = m_W if merged is None else (merged + m_W)
				if merged is None:
					messagebox.showwarning("出力", "出力するメッシュがありません。")
					return
				merged.save(fp)
				messagebox.showinfo("出力完了", f"Frame {fi} のモデルを出力しました:\n{fp}")
			except Exception as e:
				messagebox.showerror("出力失敗", f"モデル出力に失敗しました:\n{e}")
			finally:
				if not was_paused:
					is_paused[0] = False
					ctrl_widgets['pause_button'].config(text="一時停止")
					actual_time = actual_frame_counter[0] * 0.005
					animation_start_time[0] = time.time() - (actual_time / playback_speed[0])

		def _cb_screenshot():
			fp = filedialog.asksaveasfilename(
				title="スクリーンショット保存", defaultextension=".png",
				filetypes=[("PNG image", "*.png"), ("すべてのファイル", "*.*")],
				initialfile=f"ankle_frame_{current_frame[0]}.png")
			if not fp:
				return
			try:
				anim_plotter.screenshot(fp)
				messagebox.showinfo("スクリーンショット", f"保存しました:\n{fp}")
			except Exception as e:
				messagebox.showerror("スクリーンショット失敗", f"{e}")

		def _cb_close():
			is_animation_active[0] = False
			if after_id[0] is not None:
				try:
					self.after_cancel(after_id[0])
				except Exception:
					pass
				after_id[0] = None
			try:
				if hasattr(anim_plotter, 'close') and not (hasattr(anim_plotter, 'closed') and anim_plotter.closed):
					anim_plotter.close()
			except Exception:
				pass

		# --- 10. 制御パネル生成 ---
		ctrl_widgets = self._sim_view_create_control_panel(
			n_frames=N,
			frame_times=frame_times,
			callbacks={
				'on_frame_seek': _cb_frame_seek,
				'on_pause_toggle': _cb_pause_toggle,
				'on_speed_change': _cb_speed_change,
				'on_csv_export': _cb_csv_export,
				'on_export_model': _cb_export_model,
				'on_screenshot': _cb_screenshot,
				'on_close': _cb_close,
			},
			features={'csv': heatmap_enabled, 'export_model': True, 'screenshot': True})

		# --- 11. 再生ループ (hip と同じ 5ms/実時間同期) ---
		max_time = frame_times[-1] if frame_times else 0.0

		def animation_loop():
			if not is_animation_active[0]:
				return
			if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
				is_animation_active[0] = False
				return
			if not is_paused[0] and not is_seeking[0]:
				if animation_start_time[0] is None:
					animation_start_time[0] = time.time()
				elapsed_real = time.time() - animation_start_time[0]
				elapsed_anim = elapsed_real * playback_speed[0]
				actual_frame_counter[0] = int(elapsed_anim / 0.005)
				actual_time = actual_frame_counter[0] * 0.005
				if actual_time > max_time and max_time > 0:
					actual_frame_counter[0] = 0
					actual_time = 0.0
					animation_start_time[0] = time.time()
				try:
					ctrl_widgets['actual_label'].config(text=f"Actual: {actual_time:.3f}s")
				except Exception:
					pass
				# 現在フレームから前方検索
				target = current_frame[0]
				for i in range(current_frame[0], N):
					if frame_times[i] <= actual_time:
						target = i
					else:
						break
				if target == current_frame[0] and current_frame[0] > 0:
					if actual_time < frame_times[current_frame[0]]:
						target = 0
						for i in range(N):
							if frame_times[i] <= actual_time:
								target = i
							else:
								break
				if target != current_frame[0]:
					current_frame[0] = target
					show_frame(target)
				# 再生バーは 200ms 毎に更新 (ドラッグ中は更新しない)
				now = time.time()
				if (not ctrl_widgets['user_is_dragging'][0]
				    and now - last_scale_update_time[0] >= scale_update_interval[0]):
					try:
						ctrl_widgets['is_programmatic_update'][0] = True
						ctrl_widgets['playback_scale'].set(current_frame[0])
						ctrl_widgets['is_programmatic_update'][0] = False
						last_scale_update_time[0] = now
					except Exception:
						pass
			after_id[0] = self.after(5, animation_loop)

		# --- 11.5. ArUco マーカー軸の可視化 (診断用) ---
		# 各マーカーの位置に小さな軸線 (RGB=XYZ) を描画。SetUserMatrix で毎フレーム動かす
		marker_actors = {}  # bone_idx -> (x_actor, y_actor, z_actor)
		show_markers = False
		try:
			show_markers = bool(self.ankle_show_markers.get())
		except Exception:
			pass
		if show_markers:
			marker_axis_len = 15.0  # 15mm 軸
			cache_bones = cache.get("bones", {}) or {}
			for (idx, name, mesh_L, Ts) in bones_data:
				bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
				aid = int(bone.get("aruco_id", -1))
				if aid not in cache_bones:
					continue
				# マーカーは常にフレーム0姿勢を基準にした線分アクター (SetUserMatrix で駆動)
				origin = np.array([0.0, 0.0, 0.0])
				xL = pv.Line(origin, [marker_axis_len, 0, 0])
				yL = pv.Line(origin, [0, marker_axis_len, 0])
				zL = pv.Line(origin, [0, 0, marker_axis_len])
				xa = anim_plotter.add_mesh(xL, color="red", line_width=3, name=f"mk_{idx}_x")
				ya = anim_plotter.add_mesh(yL, color="green", line_width=3, name=f"mk_{idx}_y")
				za = anim_plotter.add_mesh(zL, color="blue", line_width=3, name=f"mk_{idx}_z")
				marker_actors[idx] = (xa, ya, za)

		# フレーム毎マーカー軸更新 (show_frame 内で呼ばれる)
		def _update_markers(fi):
			if not marker_actors:
				return
			import vtk
			cache_bones2 = cache.get("bones", {}) or {}
			for (idx, name, mesh_L, Ts) in bones_data:
				actors_xyz = marker_actors.get(idx)
				if actors_xyz is None:
					continue
				bone = self.ankle_bones[idx] if idx < len(self.ankle_bones) else {}
				aid = int(bone.get("aruco_id", -1))
				if aid not in cache_bones2:
					continue
				b_cache = cache_bones2[aid]
				poses_C = np.asarray(b_cache["poses"], dtype=float)  # (N, 4, 4) T_C←Mk
				if fi >= len(poses_C):
					continue
				# W = C (camera_ref) を前提。他モードでも近似的に妥当
				T_marker_W = poses_C[fi]
				m = vtk.vtkMatrix4x4()
				for i in range(4):
					for j in range(4):
						m.SetElement(i, j, float(T_marker_W[i, j]))
				for act in actors_xyz:
					try:
						act.SetUserMatrix(m)
					except Exception:
						pass

		# --- 12. カメラ・キー ---
		anim_plotter.camera_position = 'iso'
		anim_plotter.reset_camera()
		anim_plotter.add_key_event('p', lambda: ctrl_widgets['pause_button'].invoke())
		anim_plotter.add_key_event('P', lambda: ctrl_widgets['pause_button'].invoke())

		# 最初のフレーム
		show_frame(0, force_render=True)
		_update_markers(0)

		# 全アクターの変換後 bounds を含むように camera を再フィット
		# (SetUserMatrix 後の位置に基づいて画面全体に収める)
		try:
			all_bounds = []
			for (idx, name, mesh_L, Ts) in bones_data:
				m_W = self._ankle_apply_T_to_mesh(mesh_L, Ts[0])
				b = m_W.bounds
				if b is not None:
					all_bounds.append(b)
			if all_bounds:
				bnd = np.asarray(all_bounds)
				xmin, xmax = float(bnd[:, 0].min()), float(bnd[:, 1].max())
				ymin, ymax = float(bnd[:, 2].min()), float(bnd[:, 3].max())
				zmin, zmax = float(bnd[:, 4].min()), float(bnd[:, 5].max())
				anim_plotter.reset_camera(bounds=[xmin, xmax, ymin, ymax, zmin, zmax])
				print(f"[ankle animate] カメラをフィット: X[{xmin:.0f},{xmax:.0f}], "
				      f"Y[{ymin:.0f},{ymax:.0f}], Z[{zmin:.0f},{zmax:.0f}] mm")
		except Exception as e:
			print(f"[ankle animate] カメラフィット失敗: {e}")

		# ウィンドウ close コールバック
		try:
			anim_plotter.iren.add_observer('ExitEvent', lambda obj, ev: _cb_close())
		except Exception:
			pass

		# ノンブロッキング表示
		try:
			anim_plotter.show(auto_close=False, interactive_update=True)
		except TypeError:
			anim_plotter.show(auto_close=False)

		# ループ開始
		after_id[0] = self.after(5, animation_loop)

	# ---- RealSense D405 ライブ撮影 (友人スクリプトのUXを踏襲、出力は .bag) ----
	def _ankle_rs_parse_resolution(self):
		"""'1280x720@15' → (W, H, fps) を返す。"""
		s = str(self.ankle_rs_resolution.get()).strip()
		try:
			wh, fps = s.split("@")
			w, h = wh.split("x")
			return int(w), int(h), int(fps)
		except Exception:
			return 1280, 720, 15

	def _ankle_rs_configure_depth_exposure(self, depth_sensor):
		"""深度センサーの露光を設定 (友人 make_date_movie_D405.py と同じロジック)。"""
		import pyrealsense2 as rs
		try:
			if bool(self.ankle_rs_manual_exposure.get()):
				if depth_sensor.supports(rs.option.enable_auto_exposure):
					depth_sensor.set_option(rs.option.enable_auto_exposure, 0)
				if depth_sensor.supports(rs.option.exposure):
					depth_sensor.set_option(rs.option.exposure, float(self.ankle_rs_exposure_val.get()))
					print(f"[ankle rs] Depth Exposure = {int(self.ankle_rs_exposure_val.get())} (手動)")
				else:
					print("[ankle rs] 警告: 手動露光がサポートされていません")
			elif depth_sensor.supports(rs.option.enable_auto_exposure):
				depth_sensor.set_option(rs.option.enable_auto_exposure, 1)
				print("[ankle rs] Depth Auto Exposure ON")
		except Exception as e:
			print(f"[ankle rs] 露光設定エラー: {e}")

	def _ankle_rs_configure_color_exposure(self, profile):
		"""カラーセンサーの露光/ゲインを設定する。

		【なぜ重要か】ArUco の姿勢精度を最も悪化させるのはモーションブラー。
		実測 (合成, 20mm@220mm) ではブラー 4px で回転誤差 2.1倍、
		8px で 8.3倍かつ検出率 29% まで低下した。
		ブラー[px] = マーカーの画面上の速度[px/s] x 露光[s] なので、
		fps を上げるよりも **露光を直接短くする** のが本筋。
		(fps は露光の上限を決めるだけで、露光そのものではない)

		露光を詰めると暗くなるため、ゲインで補うか照明を明るくする。
		"""
		import pyrealsense2 as rs
		try:
			color_sensor = None
			for sensor in profile.get_device().query_sensors():
				try:
					if sensor.get_info(rs.camera_info.name).lower().find("rgb") >= 0:
						color_sensor = sensor
						break
				except Exception:
					continue
			if color_sensor is None:
				# 名前で見つからない場合は color ストリームを持つセンサーを探す
				for sensor in profile.get_device().query_sensors():
					try:
						if any(sp.stream_type() == rs.stream.color
						       for sp in sensor.get_stream_profiles()[:50]):
							color_sensor = sensor
							break
					except Exception:
						continue
			if color_sensor is None:
				print("[ankle rs] カラーセンサーが見つかりません - 露光は自動のまま")
				return

			if bool(self.ankle_rs_color_manual_exposure.get()):
				exp_us = int(self.ankle_rs_color_exposure_us.get())
				gain = int(self.ankle_rs_color_gain.get())
				if color_sensor.supports(rs.option.enable_auto_exposure):
					color_sensor.set_option(rs.option.enable_auto_exposure, 0)
				if color_sensor.supports(rs.option.exposure):
					rng = color_sensor.get_option_range(rs.option.exposure)
					exp_us = int(max(rng.min, min(rng.max, exp_us)))
					color_sensor.set_option(rs.option.exposure, float(exp_us))
					print(f"[ankle rs] Color Exposure = {exp_us} us ({exp_us/1000.0:.2f} ms) [手動]")
					print(f"           マーカーが 60 px/s で動く場合のブラー目安: "
					      f"{60.0 * exp_us / 1e6:.2f} px")
				else:
					print("[ankle rs] 警告: カラーの手動露光がサポートされていません")
				if color_sensor.supports(rs.option.gain):
					rng = color_sensor.get_option_range(rs.option.gain)
					gain = int(max(rng.min, min(rng.max, gain)))
					color_sensor.set_option(rs.option.gain, float(gain))
					print(f"[ankle rs] Color Gain = {gain}")
			else:
				if color_sensor.supports(rs.option.enable_auto_exposure):
					color_sensor.set_option(rs.option.enable_auto_exposure, 1)
				print("[ankle rs] Color Auto Exposure ON "
				      "(暗いと露光が伸びてモーションブラーが乗ります)")
		except Exception as e:
			print(f"[ankle rs] カラー露光設定エラー: {e}")

	def on_ankle_rs_test_connection(self) -> None:
		"""D405の接続を確認し、シリアル/内部パラメータ/深度スケールを表示。"""
		if not self._ankle_check_rs():
			return
		import pyrealsense2 as rs
		import numpy as np  # noqa: F401
		w, h, fps = self._ankle_rs_parse_resolution()
		self.ankle_rs_status.set(f"接続中… ({w}x{h}@{fps})")
		self.update_idletasks()
		pipeline = rs.pipeline()
		config = rs.config()
		try:
			config.enable_stream(rs.stream.depth, w, h, rs.format.z16, fps)
			config.enable_stream(rs.stream.color, w, h, rs.format.rgb8, fps)
			profile = pipeline.start(config)
		except Exception as e:
			self.ankle_rs_status.set("(接続失敗)")
			messagebox.showerror("接続確認 失敗",
				f"D405を起動できませんでした:\n{e}\n\n"
				"別のアプリがカメラを掴んでいる/USBが弱い/解像度非対応 の可能性があります。")
			return
		info_lines = []
		try:
			dev = profile.get_device()
			try:
				name = dev.get_info(rs.camera_info.name)
				info_lines.append(f"デバイス: {name}")
			except Exception:
				pass
			try:
				serial = dev.get_info(rs.camera_info.serial_number)
				info_lines.append(f"シリアル: {serial}")
			except Exception:
				pass
			try:
				fw = dev.get_info(rs.camera_info.firmware_version)
				info_lines.append(f"ファームウェア: {fw}")
			except Exception:
				pass
			depth_sensor = dev.first_depth_sensor()
			ds = float(depth_sensor.get_depth_scale())
			info_lines.append(f"Depth Scale: {ds:.6f} m/unit (= {ds*1000:.4f} mm/unit)")
			color_profile = profile.get_stream(rs.stream.color).as_video_stream_profile()
			intr = color_profile.get_intrinsics()
			info_lines.append(f"RGB内部: fx={intr.fx:.1f}, fy={intr.fy:.1f}, "
			                  f"cx={intr.ppx:.1f}, cy={intr.ppy:.1f}, {intr.width}x{intr.height}")
			info_lines.append(f"歪み係数: {[round(c,5) for c in intr.coeffs]}")

			# --- ステレオ構成の確認 ---
			# 左右カメラで ArUco を検出して三角測量する手法が有効かを判断するには
			# 基線長 B が要る。奥行き精度は sigma_Z = Z^2 * sigma_disp / (f * B) で、
			# B が短いと単眼PnPの横方向精度に全く歯が立たない。
			try:
				ir1 = profile.get_stream(rs.stream.infrared, 1)
				ir2 = profile.get_stream(rs.stream.infrared, 2)
				extr = ir1.get_extrinsics_to(ir2)
				baseline_mm = float(np.linalg.norm(np.asarray(extr.translation))) * 1000.0
				info_lines.append("")
				info_lines.append(f"■ ステレオ構成")
				info_lines.append(f"基線長 (左右カメラ間): {baseline_mm:.2f} mm")
				Z = 220.0
				sig_px = 0.054      # 実測のコーナー検出精度
				marker_mm = 20.0
				if baseline_mm > 0.1:
					sig_z = Z * Z * sig_px / (intr.fx * baseline_mm)
					tilt_stereo = np.degrees(np.arctan2(np.sqrt(2) * sig_z, marker_mm))
					lat_mm = sig_px / intr.fx * Z
					tilt_mono = 2.6 * np.degrees(np.arctan2(
						lat_mm, marker_mm / 2 * np.sqrt(2)))
					info_lines.append(
						f"距離 {Z:.0f}mm / 角検出 {sig_px} px のとき:")
					info_lines.append(
						f"  三角測量の奥行き精度: {sig_z:.3f} mm "
						f"→ 傾き {tilt_stereo:.2f}°")
					info_lines.append(
						f"  単眼PnPの横方向精度:   {lat_mm:.4f} mm "
						f"→ 傾き {tilt_mono:.2f}°")
					if tilt_stereo > tilt_mono:
						info_lines.append(
							f"  ⇒ 三角測量した3D点だけで姿勢を組むと "
							f"{tilt_stereo / max(tilt_mono, 1e-9):.1f}倍 悪化します。")
						info_lines.append(
							f"    左右の再投影誤差を同時最小化する方式 (バンドル調整) なら "
							f"観測が2倍になり 1.4〜1.8倍の改善が見込めます。")
					else:
						info_lines.append("  ⇒ 三角測量が有利な構成です。")
			except Exception as e:
				info_lines.append(f"(ステレオ情報の取得に失敗: {e})")

			# 録画時に左右ストリームを保存できるか
			try:
				dev2 = profile.get_device()
				sensors = dev2.query_sensors()
				sn = []
				for s in sensors:
					for sp in s.get_stream_profiles()[:200]:
						try:
							if sp.stream_type() == rs.stream.infrared:
								sn.append(sp.stream_index())
						except Exception:
							pass
				if sn:
					info_lines.append(
						f"赤外(左右)ストリーム: index {sorted(set(sn))} が利用可能")
					info_lines.append(
						"※ ステレオ方式を使うには、録画時にこれらを .db3 に含める必要があります"
						" (既存の録画には入っていません)")
			except Exception:
				pass
		finally:
			try:
				pipeline.stop()
			except Exception:
				pass
		text = "\n".join(info_lines) if info_lines else "(情報取得なし)"
		self.ankle_rs_status.set("接続OK: " + " / ".join(info_lines[:2]) if info_lines else "接続OK")
		messagebox.showinfo("D405 接続確認 OK", text)

	def _ankle_rs_run_preview_and_record(self, bag_path: str) -> tuple:
		"""2フェーズ撮影: (1)プレビュー→開始で(2)録画→停止。

		操作: **tk Toplevelコントロールパネル** の [録画開始/停止] [キャンセル] ボタン、
		     または SPACE / ESC (パネルにフォーカスがあるとき)、
		     または cv2 の動画ウィンドウで SPACE / ESC (cv2にフォーカスがあるとき)、
		     または cv2 ウィンドウの × 。

		bind_all を使わないのは、それが「元のボタン」のSPACE=invoke を防げないため。
		Toplevelを最前面+focus_forceで表示し、Toplevel自身に bind することで
		「元ボタンにSPACEが誤って行く」問題を根本回避する。

		Returns: (success: bool, frame_count: int, error_msg: str or None)
		"""
		if not self._ankle_check_rs():
			return False, 0, "pyrealsense2 が利用不可"
		try:
			import cv2
			import pyrealsense2 as rs
			import numpy as np
		except Exception as e:
			return False, 0, f"依存インポート失敗: {e}"

		w, h, fps = self._ankle_rs_parse_resolution()

		# --- Toplevel コントロールパネル ---
		self._ankle_rs_action = None   # None / "space" / "escape"

		panel = tk.Toplevel(self)
		panel.title("D405 撮影コントロール")
		panel.transient(self)
		try:
			panel.attributes("-topmost", True)
		except Exception:
			pass
		# 位置: メインウィンドウの右上寄り
		try:
			px = self.winfo_rootx() + max(self.winfo_width() - 320, 40)
			py = self.winfo_rooty() + 80
			panel.geometry(f"300x220+{px}+{py}")
		except Exception:
			panel.geometry("300x220")

		state_var = tk.StringVar(value="準備中…")
		action_btn_text = tk.StringVar(value="録画開始 (SPACE)")

		tk.Label(panel, text="D405 撮影コントロール",
		         font=(self.ui_font_family, 10, "bold")).pack(pady=(10, 4))
		tk.Label(panel, textvariable=state_var, fg="#005580",
		         font=(self.ui_font_family, 9)).pack(pady=(0, 8))
		action_btn = tk.Button(panel, textvariable=action_btn_text, width=22, height=2,
		                        bg="#e0f0ff", relief="raised",
		                        command=lambda: (setattr(self, "_ankle_rs_action", "space")))
		action_btn.pack(pady=4)
		cancel_btn = tk.Button(panel, text="キャンセル / 停止 (ESC)", width=22, height=1,
		                        command=lambda: (setattr(self, "_ankle_rs_action", "escape")))
		cancel_btn.pack(pady=(2, 6))
		tk.Label(panel,
		         text="cv2の×またはSPACE/ESCキーでも操作可",
		         fg="gray", font=(self.ui_font_family, 8)).pack(pady=(2, 4))

		# キーバインド (Toplevel限定 — bind_allしないので元ボタンに漏れない)
		panel.bind("<space>", lambda e: (setattr(self, "_ankle_rs_action", "space")))
		panel.bind("<Escape>", lambda e: (setattr(self, "_ankle_rs_action", "escape")))
		# パネルの×ボタン
		panel.protocol("WM_DELETE_WINDOW",
		                lambda: setattr(self, "_ankle_rs_action", "escape"))

		# 表示 + フォーカス強奪 (元ボタンからフォーカスを奪う)
		panel.update_idletasks()
		try:
			panel.focus_force()
		except Exception:
			pass

		def _consume_action() -> str:
			"""Toplevel/ボタン/cv2 キーの3経路から action を取り出す (取れなければ '')。"""
			a = self._ankle_rs_action
			if a:
				self._ankle_rs_action = None
				return a
			try:
				k = cv2.waitKey(1) & 0xFF
			except Exception:
				k = 255
			if k == 27:
				return "escape"
			if k == ord(" "):
				return "space"
			return ""

		def _window_closed(name: str) -> bool:
			try:
				return cv2.getWindowProperty(name, cv2.WND_PROP_VISIBLE) < 1.0
			except Exception:
				return False

		WIN_COLOR = "D405 color"
		WIN_DEPTH = "D405 depth"
		WIN_REC = "D405 REC"
		WIN_ARUCO = "D405 ArUco tracking (pose axes)"

		def _cleanup_panel():
			try: panel.destroy()
			except Exception: pass

		# =============== 単一パイプライン方式 ===============
		# 前回まで Phase 1→ Phase 2 でpipeline を stop→ start していたが、
		# pyrealsense2 の内部リソース解放が間に合わず Python がクラッシュする問題があった。
		# 対策: 最初から enable_record_to_file(bag_path) で pipeline を1本だけ起動し、
		#      recorder.pause()/resume() で「書き込むかどうか」だけ切り替える。
		#      これで pipeline は途中で stop/start されず、クラッシュしない。
		self.ankle_rs_status.set(f"起動中… ({w}x{h}@{fps})")
		state_var.set(f"起動中… ({w}x{h}@{fps})")
		action_btn_text.set("録画開始 (SPACE)")
		try:
			action_btn.configure(bg="#e0f0ff")
		except Exception:
			pass
		self.update_idletasks()

		pipeline = rs.pipeline()
		config = rs.config()
		try:
			config.enable_stream(rs.stream.depth, w, h, rs.format.z16, fps)
			config.enable_stream(rs.stream.color, w, h, rs.format.rgb8, fps)
			config.enable_record_to_file(bag_path)
			profile = pipeline.start(config)
		except Exception as e:
			_cleanup_panel()
			return False, 0, f"パイプライン起動失敗: {e}"

		# recorder を取得して直ちに pause (書き込ませない=preview 状態)
		recorder = None
		try:
			recorder = profile.get_device().as_recorder()
			recorder.pause()
		except Exception as e:
			print(f"[ankle rs] recorder取得/pause失敗: {e}")

		# ウィンドウ
		try:
			cv2.namedWindow(WIN_COLOR, cv2.WINDOW_NORMAL)
			cv2.namedWindow(WIN_DEPTH, cv2.WINDOW_NORMAL)
		except Exception:
			pass

		colorizer = rs.colorizer()
		align = rs.align(rs.stream.color)
		start_record = False
		cancelled = False
		frame_count = 0
		recording = False   # True になったら書き込み中

		# ---- 録画中に表示映像 (オーバーレイ込み) を MP4 でも自動保存 ----
		# .db3 と同じフォルダ・同じベース名に _color.mp4 / _depth.mp4 / _aruco.mp4 を出す
		bag_p = Path(bag_path)
		mp4_color_path = str(bag_p.with_name(bag_p.stem + "_color.mp4"))
		mp4_depth_path = str(bag_p.with_name(bag_p.stem + "_depth.mp4"))
		mp4_aruco_path = str(bag_p.with_name(bag_p.stem + "_aruco.mp4"))
		video_writer_color = None
		video_writer_depth = None
		video_writer_aruco = None
		_MP4_FOURCC = cv2.VideoWriter_fourcc(*"mp4v")

		def _ensure_video_writers(w_px: int, h_px: int, fps_hint: float):
			"""録画開始直後の1フレーム目でwriterを初期化する (フレームサイズが確定してから)。"""
			nonlocal video_writer_color, video_writer_depth
			if video_writer_color is None:
				try:
					video_writer_color = cv2.VideoWriter(
						mp4_color_path, _MP4_FOURCC, float(fps_hint), (int(w_px), int(h_px)))
					if not video_writer_color.isOpened():
						print(f"[ankle rs] color mp4 writer失敗 → 無効化")
						video_writer_color = None
				except Exception as e:
					print(f"[ankle rs] color mp4 writer例外: {e}")
					video_writer_color = None
			if video_writer_depth is None:
				try:
					video_writer_depth = cv2.VideoWriter(
						mp4_depth_path, _MP4_FOURCC, float(fps_hint), (int(w_px), int(h_px)))
					if not video_writer_depth.isOpened():
						print(f"[ankle rs] depth mp4 writer失敗 → 無効化")
						video_writer_depth = None
				except Exception as e:
					print(f"[ankle rs] depth mp4 writer例外: {e}")
					video_writer_depth = None

		def _ensure_aruco_video_writer(w_px: int, h_px: int, fps_hint: float):
			nonlocal video_writer_aruco
			if video_writer_aruco is None:
				try:
					video_writer_aruco = cv2.VideoWriter(
						mp4_aruco_path, _MP4_FOURCC, float(fps_hint), (int(w_px), int(h_px)))
					if not video_writer_aruco.isOpened():
						print(f"[ankle rs] aruco mp4 writer失敗 → 無効化")
						video_writer_aruco = None
				except Exception as e:
					print(f"[ankle rs] aruco mp4 writer例外: {e}")
					video_writer_aruco = None

		# ---- ArUco 追跡プレビュー 初期化 (プレビュー中も録画中も第3ウィンドウで表示) ----
		# ②の辞書・実寸、D405のRGB内部パラメータを使ってリアルタイム検出+PnP+軸描画。
		# 失敗しても本体機能 (録画自体) には影響しないよう、エラー時は静かにスキップする。
		aruco_ok = False
		aruco_state = {
			"detector": None, "dictionary": None, "params": None, "use_new_api": False,
			"obj_pts": None, "K": None, "dist": None, "marker_size": 20.0,
			"dict_name": "",
		}
		try:
			_dict_name = str(self.ankle_aruco_dict_var.get())
			_marker_size = float(self.ankle_marker_size_mm.get())
			if _marker_size <= 0:
				raise ValueError("marker size must be > 0")
			_obj_pts = np.asarray(self._ankle_marker_obj_points(_marker_size), dtype=np.float64)
			_det, _dic, _prm, _new = self._ankle_make_detector(_dict_name)
			_cprof = profile.get_stream(rs.stream.color).as_video_stream_profile()
			_intr = _cprof.get_intrinsics()
			_K = np.array([[_intr.fx, 0, _intr.ppx],
			               [0, _intr.fy, _intr.ppy],
			               [0, 0, 1]], dtype=np.float64)
			_dist_c = np.array(_intr.coeffs, dtype=np.float64)
			aruco_state.update({
				"detector": _det, "dictionary": _dic, "params": _prm, "use_new_api": _new,
				"obj_pts": _obj_pts, "K": _K, "dist": _dist_c,
				"marker_size": _marker_size, "dict_name": _dict_name,
			})
			cv2.namedWindow(WIN_ARUCO, cv2.WINDOW_NORMAL)
			aruco_ok = True
			print(f"[ankle rs] ArUco追跡プレビュー: 辞書={_dict_name}, 実寸={_marker_size:.1f}mm")
		except Exception as e:
			print(f"[ankle rs] ArUco追跡プレビュー初期化失敗 (この機能はスキップ): {e}")

		def _draw_aruco_overlay(bgr_clean, gray):
			"""1フレームで検出→軸描画。 (overlay_img, n_detected) を返す。aruco_ok=False時はNone。"""
			nonlocal aruco_ok
			if not aruco_ok:
				return None, 0
			try:
				if aruco_state["use_new_api"] and aruco_state["detector"] is not None:
					corners, ids, _ = aruco_state["detector"].detectMarkers(gray)
				else:
					corners, ids, _ = cv2.aruco.detectMarkers(
						gray, aruco_state["dictionary"], parameters=aruco_state["params"])
			except Exception as e:
				print(f"[ankle rs] aruco detect失敗: {e}")
				return None, 0
			out = bgr_clean.copy()
			n_det = 0 if ids is None else len(ids)
			# ヘッダ情報
			cv2.putText(out, f"ArUco tracking ({aruco_state['dict_name']}, {aruco_state['marker_size']:.1f}mm)",
			            (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 255, 255), 1, cv2.LINE_AA)
			if n_det == 0:
				cv2.putText(out, "No ArUco detected", (10, 55),
				            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2, cv2.LINE_AA)
				return out, 0
			try:
				cv2.aruco.drawDetectedMarkers(out, corners, ids)
			except Exception:
				pass
			# 各マーカーで PnP → drawFrameAxes + ID テキスト
			K = aruco_state["K"]; dist_c = aruco_state["dist"]
			obj_pts_ = aruco_state["obj_pts"]
			ax_len = float(aruco_state["marker_size"]) * 0.7
			for corner, mid_arr in zip(corners, ids):
				try:
					mid = int(mid_arr[0] if hasattr(mid_arr, '__len__') else mid_arr)
				except Exception:
					continue
				img_pts = corner.reshape(-1, 2).astype(np.float64)
				try:
					ok, rvec, tvec = cv2.solvePnP(
						obj_pts_, img_pts, K, dist_c, flags=cv2.SOLVEPNP_IPPE)
				except Exception:
					ok = False
				if not ok:
					continue
				try:
					cv2.drawFrameAxes(out, K, dist_c, rvec, tvec, ax_len, 3)
				except Exception:
					pass
				try:
					cm = img_pts.mean(axis=0)
					cx_m, cy_m = int(cm[0]), int(cm[1])
					cv2.putText(out, f"id={mid}", (cx_m + 10, cy_m - 10),
					            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 100, 0), 2, cv2.LINE_AA)
				except Exception:
					pass
			cv2.putText(out, f"Detected: {n_det}", (10, out.shape[0] - 15),
			            cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 200, 0), 2, cv2.LINE_AA)
			return out, n_det

		# 露光設定 (深度 + カラー)
		try:
			depth_sensor = profile.get_device().first_depth_sensor()
			self._ankle_rs_configure_depth_exposure(depth_sensor)
		except Exception as e:
			print(f"[ankle rs] 深度露光設定失敗: {e}")
		try:
			self._ankle_rs_configure_color_exposure(profile)
		except Exception as e:
			print(f"[ankle rs] カラー露光設定失敗: {e}")

		# 破棄フレーム数
		try:
			n_discard = max(0, int(self.ankle_rs_discard_frames.get()))
		except Exception:
			n_discard = 45

		# パネルにフォーカス
		try: panel.focus_force()
		except Exception: pass

		self.ankle_rs_status.set(f"プレビュー中… ({w}x{h}@{fps})")
		state_var.set(f"プレビュー中\n({w}x{h}@{fps})")

		try:
			while True:
				# --- 停止/開始条件 ---
				if _window_closed(WIN_COLOR) or _window_closed(WIN_DEPTH):
					if recording:
						break   # 録画中は停止として扱う
					cancelled = True; break
				a = _consume_action()
				if a == "escape":
					if recording:
						break   # 録画中は停止
					cancelled = True; break
				if a == "space":
					if not recording:
						# --- プレビュー → 録画へ切替 ---
						# 破棄フレーム (露光/AE安定待ち) を pause状態で消費
						state_var.set(f"露光安定待ち…\n(0/{n_discard}フレーム)")
						self.update_idletasks()
						for i in range(n_discard):
							try:
								pipeline.wait_for_frames(timeout_ms=2000)
							except Exception:
								break
							if i % max(fps // 2, 1) == 0:
								state_var.set(f"露光安定待ち…\n({i+1}/{n_discard}フレーム)")
								self.update_idletasks()
						# 録画開始
						try:
							if recorder is not None:
								recorder.resume()
						except Exception as e:
							print(f"[ankle rs] recorder.resume失敗: {e}")
						recording = True
						start_record = True
						frame_count = 0
						action_btn_text.set("録画停止 (SPACE)")
						try:
							action_btn.configure(bg="#ffe0e0")
						except Exception:
							pass
						self.ankle_rs_status.set("録画中…")
						state_var.set("録画中 (0フレーム)")
						self.update_idletasks()
						try: panel.focus_force()
						except Exception: pass
						continue
					else:
						# 録画中 → 停止
						break

				# --- 通常フレーム処理 ---
				try:
					success, frames = pipeline.try_wait_for_frames(timeout_ms=100)
				except Exception as e:
					print(f"[ankle rs] frame取得エラー: {e}")
					if recording:
						break
					cancelled = True; break
				if not success:
					self.update()
					continue

				# ---- 表示 (色+深度両方、プレビュー/録画で内容切替) ----
				aligned = align.process(frames)
				color = aligned.get_color_frame()
				depth = aligned.get_depth_frame()
				if not color or not depth:
					self.update(); continue
				if recording:
					frame_count += 1
				rgb = np.asanyarray(color.get_data())
				bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
				# ArUcoウィンドウ用のクリーンな色画像 (十字/RECオーバーレイなし)
				bgr_clean_for_aruco = bgr.copy() if aruco_ok else None
				gray_for_aruco = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY) if aruco_ok else None
				colored_depth = colorizer.colorize(depth)
				depth_bgr = cv2.cvtColor(np.asanyarray(colored_depth.get_data()), cv2.COLOR_RGB2BGR)
				ch, cw, _ = bgr.shape
				cx, cy = cw // 2, ch // 2
				try:
					dist = depth.get_distance(cx, cy)
				except Exception:
					dist = 0.0
				# 中心距離十字は色/深度両方に描画
				for img in (bgr, depth_bgr):
					cv2.line(img, (cx - 10, cy), (cx + 10, cy), (255, 255, 255), 1)
					cv2.line(img, (cx, cy - 10), (cx, cy + 10), (255, 255, 255), 1)
					cv2.putText(img, f"{dist:.3f} m", (cx + 15, cy - 15),
					            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1, cv2.LINE_AA)
				# オーバーレイ (色ウィンドウのみ)
				if recording:
					cv2.putText(bgr, f"REC  {frame_count} frames  (t={frame_count/max(fps,1):.1f}s)",
					            (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2, cv2.LINE_AA)
					cv2.putText(bgr, "Use panel or SPACE/ESC to stop", (10, 60),
					            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2, cv2.LINE_AA)
					cv2.circle(bgr, (25, 90), 8, (0, 0, 255), -1)
				else:
					cv2.putText(bgr, "PREVIEW  Use panel or SPACE/ESC", (10, 30),
					            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2, cv2.LINE_AA)
				# ウィンドウ閉じ判定 (imshow前に)
				if _window_closed(WIN_COLOR) or _window_closed(WIN_DEPTH):
					if recording:
						break
					cancelled = True; break
				# MP4 に書き出し (録画中のみ、オーバーレイ込み映像)
				if recording:
					_ensure_video_writers(bgr.shape[1], bgr.shape[0], fps)
					if video_writer_color is not None:
						try: video_writer_color.write(bgr)
						except Exception as e: print(f"[ankle rs] color mp4 write失敗: {e}")
					if video_writer_depth is not None:
						try: video_writer_depth.write(depth_bgr)
						except Exception as e: print(f"[ankle rs] depth mp4 write失敗: {e}")
				cv2.imshow(WIN_COLOR, bgr)
				cv2.imshow(WIN_DEPTH, depth_bgr)
				# ---- ArUco追跡ウィンドウ (プレビュー+録画中の両方で表示) ----
				if aruco_ok:
					if _window_closed(WIN_ARUCO):
						# ユーザーがArUcoウィンドウの×を閉じた → 以降スキップ (録画は継続)
						aruco_ok = False
					else:
						overlay, n_det = _draw_aruco_overlay(bgr_clean_for_aruco, gray_for_aruco)
						if overlay is not None:
							cv2.imshow(WIN_ARUCO, overlay)
							# 録画中はMP4にも保存
							if recording:
								_ensure_aruco_video_writer(overlay.shape[1], overlay.shape[0], fps)
								if video_writer_aruco is not None:
									try: video_writer_aruco.write(overlay)
									except Exception as e: print(f"[ankle rs] aruco mp4 write失敗: {e}")
				# 進捗表示更新 (fpsごと)
				if recording and frame_count % max(fps, 1) == 0:
					sec = frame_count / max(fps, 1)
					self.ankle_rs_status.set(f"録画中… {frame_count}f (t≈{sec:.1f}s)")
					state_var.set(f"録画中 ({frame_count}フレーム / {sec:.1f}s)")
					self.update_idletasks()
				self.update()
		finally:
			try:
				if recorder is not None and recording:
					recorder.pause()
			except Exception:
				pass
			try: pipeline.stop()
			except Exception: pass
			# MP4 writer をきちんとリリース (これをやらないとファイルが破損)
			try:
				if video_writer_color is not None:
					video_writer_color.release()
			except Exception: pass
			try:
				if video_writer_depth is not None:
					video_writer_depth.release()
			except Exception: pass
			try:
				if video_writer_aruco is not None:
					video_writer_aruco.release()
			except Exception: pass
			try: cv2.destroyAllWindows()
			except Exception: pass
			try: cv2.waitKey(1)
			except Exception: pass
			_cleanup_panel()

		# キャンセル時は生成された(ほぼ空の) .db3 ファイルと MP4 も削除
		if cancelled and not start_record:
			for _p in (bag_path, mp4_color_path, mp4_depth_path, mp4_aruco_path):
				try:
					_pp = Path(_p)
					if _pp.exists():
						_pp.unlink()
				except Exception:
					pass
			self.ankle_rs_status.set("(録画キャンセル)")
			return False, 0, None
		if not start_record:
			# 想定外
			self.ankle_rs_status.set("(録画されませんでした)")
			return False, 0, None
		# 保存された MP4 のパスを self に貯めておく (on_ankle_rs_capture が拾える)
		self._ankle_rs_last_mp4 = {
			"color": mp4_color_path if Path(mp4_color_path).exists() else None,
			"depth": mp4_depth_path if Path(mp4_depth_path).exists() else None,
			"aruco": mp4_aruco_path if Path(mp4_aruco_path).exists() else None,
		}
		return True, frame_count, None

	def on_ankle_rs_capture(self) -> None:
		"""D405プレビュー+録画→.db3保存 → ankle_depth_pathに自動セット。"""
		if not self._ankle_check_rs():
			return
		# クリックされたボタンにフォーカスが残ると SPACE でボタンが再invokeされるので、
		# ここで先にフォーカスを root へ逃がす (保険)
		try:
			self.focus_set()
		except Exception:
			pass
		# 録画先 (pyrealsense2 2.58+ は .db3 拡張子を要求)
		default_dir = Path(__file__).parent / "cache"
		try:
			default_dir.mkdir(parents=True, exist_ok=True)
		except Exception:
			pass
		import datetime as _dt
		default_name = f"ankle_d405_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.db3"
		bag_path = filedialog.asksaveasfilename(
			title="録画先 .db3 ファイル",
			initialdir=str(default_dir),
			initialfile=default_name,
			defaultextension=".db3",
			filetypes=[("RealSense recording", "*.db3"), ("すべてのファイル", "*.*")])
		if not bag_path:
			return
		messagebox.showinfo(
			"D405 録画",
			"プレビューウィンドウが開きます。\n\n"
			"  ・SPACE = 破棄フレーム待ち → 録画開始 / 録画中は停止\n"
			"  ・ESC   = 中止\n\n"
			"D405 は近距離用 (7-50cm) です。マーカーが視野内でくっきり見える距離で録画してください。")
		ok, frame_count, err = self._ankle_rs_run_preview_and_record(bag_path)
		if err:
			self.ankle_rs_status.set(f"(エラー: {err})")
			messagebox.showerror("D405 録画 失敗", f"録画に失敗しました:\n{err}")
			return
		if not ok:
			self.ankle_rs_status.set("(録画されませんでした)")
			return
		# 保存ファイル一覧を表示 (.db3 + _color.mp4 + _depth.mp4)
		def _size_mb(p: str) -> float:
			try:
				return Path(p).stat().st_size / (1024 * 1024)
			except Exception:
				return -1
		db3_sz = _size_mb(bag_path)
		mp4s = getattr(self, "_ankle_rs_last_mp4", None) or {}
		mp4_color = mp4s.get("color")
		mp4_depth = mp4s.get("depth")
		mp4_aruco = mp4s.get("aruco")
		self.ankle_rs_status.set(f"録画完了: {frame_count}f  ({db3_sz:.1f} MB db3)  → {Path(bag_path).name}")
		self.ankle_depth_path.set(bag_path)
		lines = [f"{frame_count} フレームを保存しました。",
		         "",
		         "▼ 保存ファイル (同一フォルダ):"]
		lines.append(f"  {Path(bag_path).name}  ({db3_sz:.1f} MB)  ← 解析用 (深度含む)")
		if mp4_color:
			lines.append(f"  {Path(mp4_color).name}  ({_size_mb(mp4_color):.1f} MB)  ← 動画プレイヤーで確認可")
		if mp4_depth:
			lines.append(f"  {Path(mp4_depth).name}  ({_size_mb(mp4_depth):.1f} MB)  ← 深度サーモグラフ動画")
		if mp4_aruco:
			lines.append(f"  {Path(mp4_aruco).name}  ({_size_mb(mp4_aruco):.1f} MB)  ← ArUco追跡映像 (軸+ID描画)")
		lines += ["",
		          "「② 深度データ」に .db3 を自動セットしました。",
		          "続けて ④「ArUco検出+PnP実行」で解析できます。",
		          "",
		          "MP4は Windows Media Player 等で直接再生できます。"]
		messagebox.showinfo("D405 録画完了", "\n".join(lines))

	# endregion ankle simulator

	def _create_org_tab(self) -> None:
		"""ORGタブのUIを構築"""
		# スクロール可能なメインフレーム
		canvas = tk.Canvas(self.org_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.org_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)

		scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
		)

		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)

		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")

		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="ORG - RobotScan データ可視化",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# ファイル選択フレーム
		file_frame = ttk.LabelFrame(container, text="ファイル選択", style="Bold.TLabelframe")
		file_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
		for i in range(3):
			file_frame.columnconfigure(i, weight=[0, 1, 0][i])

		# RobotScan STL/OBJ
		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="RobotScan モデル (STL/OBJ)",
			textvariable=self.org_model_path,
			command=self.choose_org_model,
		)

		# RobotScan PP
		self._add_file_row(
			parent=file_frame,
			row=1,
			label_text="RobotScan 座標ポイント (PP)",
			textvariable=self.org_pp_path,
			command=self.choose_org_pp,
		)

		# ORG座標入力フレーム
		coord_frame = ttk.LabelFrame(container, text="ORG座標入力（C_o座標系）", style="Bold.TLabelframe")
		coord_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		coord_frame.columnconfigure(1, weight=1)
		coord_frame.columnconfigure(3, weight=1)
		coord_frame.columnconfigure(5, weight=1)

		ttk.Label(coord_frame, text="X (mm):").grid(row=0, column=0, sticky="e", padx=(12, 5), pady=8)
		self.org_x_entry = ttk.Entry(coord_frame, textvariable=self.org_coord_x, width=15)
		self.org_x_entry.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=8)

		ttk.Label(coord_frame, text="Y (mm):").grid(row=0, column=2, sticky="e", padx=(12, 5), pady=8)
		self.org_y_entry = ttk.Entry(coord_frame, textvariable=self.org_coord_y, width=15)
		self.org_y_entry.grid(row=0, column=3, sticky="w", padx=(0, 12), pady=8)

		ttk.Label(coord_frame, text="Z (mm):").grid(row=0, column=4, sticky="e", padx=(12, 5), pady=8)
		self.org_z_entry = ttk.Entry(coord_frame, textvariable=self.org_coord_z, width=15)
		self.org_z_entry.grid(row=0, column=5, sticky="w", padx=(0, 12), pady=8)

		# 可視化ボタン
		action_frame = ttk.Frame(container)
		action_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.org_visualize_button = ttk.Button(
			action_frame,
			text="モデル + 座標ポイントを可視化",
			command=self.on_visualize_org,
			state="disabled"
		)
		self.org_visualize_button.grid(row=0, column=0, sticky="ew")

		# Variable trace
		self.org_model_path.trace_add("write", lambda *_: self.update_org_button_states())
		self.org_pp_path.trace_add("write", lambda *_: self.update_org_button_states())

		# 初期状態を更新
		self.update_org_button_states()

	def _create_a_b_tab(self) -> None:
		"""A-BタブのUIを構築"""
		container = self.a_b_tab
		container.columnconfigure(0, weight=1)

		title_label = ttk.Label(
			container,
			text="A-B モデル表示",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		file_frame = ttk.LabelFrame(container, text="モデル選択", style="Bold.TLabelframe")
		file_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
		for i in range(3):
			file_frame.columnconfigure(i, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="モデルA (STL)",
			textvariable=self.a_b_model_a_path,
			command=self.choose_a_b_model_a,
		)

		self._add_file_row(
			parent=file_frame,
			row=1,
			label_text="モデルB (STL)",
			textvariable=self.a_b_model_b_path,
			command=self.choose_a_b_model_b,
		)

		action_frame = ttk.Frame(container)
		action_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.a_b_visualize_button = ttk.Button(
			action_frame,
			text="AとBを同時表示",
			command=self.on_visualize_a_b,
			state="disabled",
		)
		self.a_b_visualize_button.grid(row=0, column=0, sticky="ew")

		self.a_b_diff_button = ttk.Button(
			action_frame,
			text="A-Bを実行",
			command=self.on_execute_a_b_heatmap,
			state="disabled",
		)
		self.a_b_diff_button.grid(row=1, column=0, sticky="ew", pady=(8, 0))

		# Variable trace
		self.a_b_model_a_path.trace_add("write", lambda *_: self.update_a_b_button_states())
		self.a_b_model_b_path.trace_add("write", lambda *_: self.update_a_b_button_states())

		# 初期状態を更新
		self.update_a_b_button_states()

	def _create_heatmap_tab(self) -> None:
		"""HeatMapタブのUIを構築"""
		container = self.heatmap_tab
		container.columnconfigure(0, weight=1)

		title_label = ttk.Label(
			container,
			text="HeatMap (B→A 距離)",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		info_label = ttk.Label(
			container,
			text="モデルBの各点からモデルA表面までの距離をヒートマップ表示します。\n可視化画面で点をクリックすると距離(mm)が表示されます。",
			justify=tk.LEFT
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 8))

		file_frame = ttk.LabelFrame(container, text="モデル選択", style="Bold.TLabelframe")
		file_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		for i in range(3):
			file_frame.columnconfigure(i, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="モデルA (参照: 表面) (STL/OBJ)",
			textvariable=self.heatmap_model_a_path,
			command=self.choose_heatmap_model_a,
		)

		self._add_file_row(
			parent=file_frame,
			row=1,
			label_text="モデルB (距離計算対象) (STL/OBJ)",
			textvariable=self.heatmap_model_b_path,
			command=self.choose_heatmap_model_b,
		)

		# 表示モデル切り替え
		display_frame = ttk.LabelFrame(container, text="表示モデル切り替え", style="Bold.TLabelframe")
		display_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		display_frame.columnconfigure(0, weight=1)
		display_frame.columnconfigure(1, weight=1)

		self.heatmap_show_model_a_check = ttk.Checkbutton(
			display_frame,
			text="モデルAを表示",
			variable=self.heatmap_show_model_a,
		)
		self.heatmap_show_model_a_check.grid(row=0, column=0, sticky="w", padx=12, pady=8)

		self.heatmap_show_model_b_check = ttk.Checkbutton(
			display_frame,
			text="モデルBを表示",
			variable=self.heatmap_show_model_b,
		)
		self.heatmap_show_model_b_check.grid(row=0, column=1, sticky="w", padx=12, pady=8)

		action_frame = ttk.Frame(container)
		action_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.heatmap_visualize_button = ttk.Button(
			action_frame,
			text="AとBを同時表示",
			command=self.on_visualize_heatmap_models,
			state="disabled",
		)
		self.heatmap_visualize_button.grid(row=0, column=0, sticky="ew")

		self.heatmap_execute_button = ttk.Button(
			action_frame,
			text="B→A 距離ヒートマップ",
			command=self.on_execute_heatmap_b_to_a,
			state="disabled",
		)
		self.heatmap_execute_button.grid(row=1, column=0, sticky="ew", pady=(8, 0))

		# Variable trace
		self.heatmap_model_a_path.trace_add("write", lambda *_: self.update_heatmap_button_states())
		self.heatmap_model_b_path.trace_add("write", lambda *_: self.update_heatmap_button_states())

		# 初期状態を更新
		self.update_heatmap_button_states()

	def _create_c_fix_tab(self) -> None:
		"""C fixタブのコンテンツを作成"""
		container = self.c_fix_tab
		container.columnconfigure(0, weight=1)
		
		# タイトル
		title_label = ttk.Label(
			container,
			text="Coordinate Fix ツール",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))
		
		# 説明
		info_label = ttk.Label(
			container,
			text="Simulatorタブで選択されたモデルと特徴点を使用して座標系を確認・修正します。",
			justify=tk.LEFT,
			wraplength=600
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))
		
		# 修正量入力フレーム
		correction_frame = ttk.LabelFrame(container, text="修正量", style="Bold.TLabelframe")
		correction_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		correction_frame.columnconfigure(1, weight=1)
		correction_frame.columnconfigure(3, weight=1)
		correction_frame.columnconfigure(5, weight=1)
		
		# 修正量の変数
		self.c_fix_correction_x = tk.DoubleVar(value=0.0)
		self.c_fix_correction_y = tk.DoubleVar(value=0.0)
		self.c_fix_correction_z = tk.DoubleVar(value=0.0)
		
		ttk.Label(correction_frame, text="X (mm):").grid(row=0, column=0, sticky="e", padx=(12, 5), pady=8)
		self.c_fix_x_entry = ttk.Entry(correction_frame, textvariable=self.c_fix_correction_x, width=15)
		self.c_fix_x_entry.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=8)
		
		ttk.Label(correction_frame, text="Y (mm):").grid(row=0, column=2, sticky="e", padx=(12, 5), pady=8)
		self.c_fix_y_entry = ttk.Entry(correction_frame, textvariable=self.c_fix_correction_y, width=15)
		self.c_fix_y_entry.grid(row=0, column=3, sticky="w", padx=(0, 12), pady=8)
		
		ttk.Label(correction_frame, text="Z (mm):").grid(row=0, column=4, sticky="e", padx=(12, 5), pady=8)
		self.c_fix_z_entry = ttk.Entry(correction_frame, textvariable=self.c_fix_correction_z, width=15)
		self.c_fix_z_entry.grid(row=0, column=5, sticky="w", padx=(0, 12), pady=8)
		
		# 可視化ボタン
		viz_frame = ttk.LabelFrame(container, text="可視化", style="Bold.TLabelframe")
		viz_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		viz_frame.columnconfigure(0, weight=1)
		
		self.c_fix_visualize_button = ttk.Button(
			viz_frame,
			text="モデル + 特徴点を可視化",
			command=self.on_c_fix_visualize,
			state="disabled"
		)
		self.c_fix_visualize_button.grid(row=0, column=0, sticky="ew", padx=12, pady=12)
		
		# 変換ボタン
		transform_frame = ttk.LabelFrame(container, text="変換", style="Bold.TLabelframe")
		transform_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 12))
		transform_frame.columnconfigure(0, weight=1)
		transform_frame.columnconfigure(1, weight=1)
		
		self.c_fix_transform_prox_button = ttk.Button(
			transform_frame,
			text="近位ABCD修正後を保存",
			command=self.on_c_fix_transform_prox,
			state="disabled"
		)
		self.c_fix_transform_prox_button.grid(row=0, column=0, sticky="ew", padx=12, pady=12)
		
		self.c_fix_transform_dist_button = ttk.Button(
			transform_frame,
			text="遠位ABC修正後を保存",
			command=self.on_c_fix_transform_dist,
			state="disabled"
		)
		self.c_fix_transform_dist_button.grid(row=0, column=1, sticky="ew", padx=12, pady=12)
		
		# Variable trace for button state updates
		self.prox_model_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.prox_pp_abcd_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.prox_pp_olmn_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.dist_model_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.dist_pp_abc_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		self.dist_pp_olmn_path.trace_add("write", lambda *_: self.update_c_fix_button_states())
		
		# 初期状態を更新
		self.update_c_fix_button_states()

	def _create_fitting_tab(self) -> None:
		"""Fittingタブのコンテンツを作成"""
		# スクロール可能なキャンバスを作成
		canvas = tk.Canvas(self.fitting_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.fitting_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)
		
		scrollable_frame.bind(
			"<Configure>",
			lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
		)
		
		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)
		
		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")
		
		# マウスホイールでスクロール（Mac/Windows/Linux 両対応）
		def _on_mousewheel(event):
			step = _mousewheel_units(event)
			if step:
				canvas.yview_scroll(step, "units")
		canvas.bind_all("<MouseWheel>", _on_mousewheel)
		canvas.bind_all("<Button-4>", _on_mousewheel)
		canvas.bind_all("<Button-5>", _on_mousewheel)
		
		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="3Dモデル フィッティングツール",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# 説明（コンパクト版）
		description_text = (
			"親任意領域に対して子任意領域をフィッティングします (RANSAC/ICP)\n"
			"その後、親モデルファイルに対して同じ動き方で子モデルファイルを移動させます。"
		)
		description = ttk.Label(
			container,
			text=description_text,
			justify=tk.LEFT
		)
		description.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 4))
		
		# パラメータ説明ボタン
		help_button = ttk.Button(container, text="パラメータの意味を表示", command=self.show_fitting_help)
		help_button.grid(row=2, column=0, sticky="w", padx=12, pady=(0, 8))

		# Section: ファイル選択（親）
		parent_frame = ttk.LabelFrame(container, text="親ファイル（固定側）", style="Bold.TLabelframe")
		parent_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 8))
		for i in range(3):
			parent_frame.columnconfigure(i, weight=1)

		# 親モデルファイル
		self._add_file_row(
			parent=parent_frame,
			row=0,
			label_text="親モデルファイル (STL/OBJ)",
			textvariable=self.fitting_parent_model_path,
			command=self.choose_fitting_parent_model,
		)

		# 親任意領域ファイル
		self._add_file_row(
			parent=parent_frame,
			row=1,
			label_text="親任意領域ファイル (STL/OBJ)",
			textvariable=self.fitting_parent_region_path,
			command=self.choose_fitting_parent_region,
		)

		# Section: ファイル選択（子）
		child_frame = ttk.LabelFrame(container, text="子ファイル（移動側）", style="Bold.TLabelframe")
		child_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 8))
		for i in range(3):
			child_frame.columnconfigure(i, weight=1)

		# 子モデルファイル
		self._add_file_row(
			parent=child_frame,
			row=0,
			label_text="子モデルファイル (STL/OBJ)",
			textvariable=self.fitting_child_model_path,
			command=self.choose_fitting_child_model,
		)

		# 子任意領域ファイル
		self._add_file_row(
			parent=child_frame,
			row=1,
			label_text="子任意領域ファイル (STL/OBJ)",
			textvariable=self.fitting_child_region_path,
			command=self.choose_fitting_child_region,
		)

		# 子特徴点ファイル
		self._add_file_row(
			parent=child_frame,
			row=2,
			label_text="子特徴点ファイル (PP)",
			textvariable=self.fitting_child_pp_path,
			command=self.choose_fitting_child_pp,
		)

		# 子ヒートマップ表示用モデル（任意）
		self._add_file_row(
			parent=child_frame,
			row=3,
			label_text="子ヒートマップ表示用モデル (STL/OBJ)",
			textvariable=self.fitting_child_heatmap_model_path,
			command=self.choose_fitting_child_heatmap_model,
		)


		# 操作ボタン段（可視化）
		viz_frame = ttk.LabelFrame(container, text="モデル確認", style="Bold.TLabelframe")
		viz_frame.grid(row=5, column=0, sticky="nsew", padx=12, pady=(0, 8))
		viz_frame.columnconfigure(0, weight=1)
		viz_frame.columnconfigure(1, weight=1)

		self.fitting_viz_parent_button = ttk.Button(
			viz_frame, 
			text="親を可視化（モデル + 任意領域）",
			command=self.on_visualize_fitting_parent,
			width=30
		)
		self.fitting_viz_parent_button.grid(row=0, column=0, sticky="ew", padx=8, pady=6)

		self.fitting_viz_child_button = ttk.Button(
			viz_frame, 
			text="子を可視化（モデル + 任意領域）",
			command=self.on_visualize_fitting_child,
			width=30
		)
		self.fitting_viz_child_button.grid(row=0, column=1, sticky="ew", padx=8, pady=6)

		# パラメータ設定フレーム
		param_frame = ttk.LabelFrame(container, text="フィッティングパラメータ", style="Bold.TLabelframe")
		param_frame.grid(row=6, column=0, sticky="nsew", padx=12, pady=(0, 8))
		param_frame.columnconfigure(1, weight=1)
		param_frame.columnconfigure(3, weight=1)
		param_frame.columnconfigure(5, weight=1)
		
		# RANSACパラメータ
		ransac_label = ttk.Label(param_frame, text="RANSACパラメータ", font=(self.ui_font_family, 9, "bold"))
		ransac_label.grid(row=0, column=0, columnspan=6, sticky="w", padx=8, pady=(6, 3))
		
		ttk.Label(param_frame, text="距離閾値 (mm):").grid(row=1, column=0, sticky="e", padx=(8, 5), pady=4)
		self.ransac_distance_entry = ttk.Entry(param_frame, textvariable=self.ransac_distance_threshold, width=10)
		self.ransac_distance_entry.grid(row=1, column=1, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="最大反復回数:").grid(row=1, column=2, sticky="e", padx=(8, 5), pady=4)
		self.ransac_max_iterations_entry = ttk.Entry(param_frame, textvariable=self.ransac_max_iterations, width=10)
		self.ransac_max_iterations_entry.grid(row=1, column=3, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="信頼度:").grid(row=1, column=4, sticky="e", padx=(8, 5), pady=4)
		self.ransac_confidence_entry = ttk.Entry(param_frame, textvariable=self.ransac_confidence, width=10)
		self.ransac_confidence_entry.grid(row=1, column=5, sticky="w", padx=(0, 8), pady=4)
		
		# ICPパラメータ
		icp_label = ttk.Label(param_frame, text="ICPパラメータ", font=(self.ui_font_family, 9, "bold"))
		icp_label.grid(row=2, column=0, columnspan=6, sticky="w", padx=8, pady=(8, 3))
		
		ttk.Label(param_frame, text="閾値 (mm):").grid(row=3, column=0, sticky="e", padx=(8, 5), pady=4)
		self.icp_threshold_entry = ttk.Entry(param_frame, textvariable=self.icp_threshold, width=10)
		self.icp_threshold_entry.grid(row=3, column=1, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="最大反復回数:").grid(row=3, column=2, sticky="e", padx=(8, 5), pady=4)
		self.icp_max_iterations_entry = ttk.Entry(param_frame, textvariable=self.icp_max_iterations, width=10)
		self.icp_max_iterations_entry.grid(row=3, column=3, sticky="w", padx=(0, 8), pady=4)
		
		ttk.Label(param_frame, text="サンプル点数:").grid(row=3, column=4, sticky="e", padx=(8, 5), pady=4)
		self.sample_points_entry = ttk.Entry(param_frame, textvariable=self.sample_points, width=10)
		self.sample_points_entry.grid(row=3, column=5, sticky="w", padx=(0, 8), pady=4)
		
		# デフォルトに戻すボタン
		btn_reset_params = ttk.Button(param_frame, text="デフォルトに戻す", command=self.reset_fitting_parameters)
		btn_reset_params.grid(row=4, column=0, columnspan=6, pady=(6, 6))

		# 操作ボタン段（実行）
		action_frame = ttk.Frame(container)
		action_frame.grid(row=7, column=0, sticky="nsew", padx=12, pady=(0, 12))
		action_frame.columnconfigure(0, weight=1)

		self.fitting_execute_button = ttk.Button(
			action_frame, 
			text="フィッティング実行",
			command=self.on_fitting_execute,
			state="disabled"
		)
		self.fitting_execute_button.grid(row=0, column=0, sticky="ew", pady=4)

		# Variable trace for dynamic state updates
		self.fitting_parent_model_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_parent_region_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_child_model_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_child_region_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		self.fitting_child_heatmap_model_path.trace_add("write", lambda *_: self.update_fitting_button_states())
		
		# 初期状態を更新
		self.update_fitting_button_states()

	def _create_cartilage_separation_tab(self) -> None:
		"""軟骨分離タブ: 近位（骨盤）・遠位（大腿骨）それぞれのサブタブを持つ"""
		# スクロール可能なキャンバス
		canvas = tk.Canvas(self.cartilage_sep_tab, highlightthickness=0)
		scrollbar = ttk.Scrollbar(self.cartilage_sep_tab, orient="vertical", command=canvas.yview)
		scrollable_frame = ttk.Frame(canvas)
		scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
		canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
		canvas.configure(yscrollcommand=scrollbar.set)
		canvas.pack(side="left", fill="both", expand=True)
		scrollbar.pack(side="right", fill="y")
		def _on_mousewheel_cs(event):
			step = _mousewheel_units(event)
			if step:
				canvas.yview_scroll(step, "units")
		canvas.bind_all("<MouseWheel>", _on_mousewheel_cs)
		canvas.bind_all("<Button-4>", _on_mousewheel_cs)
		canvas.bind_all("<Button-5>", _on_mousewheel_cs)
		
		container = scrollable_frame
		container.columnconfigure(0, weight=1)

		# タイトル
		ttk.Label(container, text="軟骨分離ツール (Cartilage Separation)",
			font=(self.ui_font_family, 12, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))

		ttk.Label(container, justify=tk.LEFT, wraplength=700, text=(
			"骨+軟骨モデル（体積メッシュ）と骨のみモデル（体積メッシュ）を位置合わせし、\n"
			"ブーリアン差分（モデル1 − モデル2）で軟骨領域を分離します。\n"
			"近位（骨盤）・遠位（大腿骨）それぞれで実行してください。\n"
			"分離後のモデルは Simulator タブの軟骨モデル入力欄から読み込めます。")
		).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 8))

		# === 近位 / 遠位 サブノートブック ===
		cs_notebook = ttk.Notebook(container)
		cs_notebook.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 8))

		# --- 近位サブタブ ---
		prox_tab = ttk.Frame(cs_notebook, padding=8)
		cs_notebook.add(prox_tab, text=self._joint_label("cs_prox_tab"))
		self._joint_widgets["cs_notebook"] = cs_notebook
		self._joint_widgets["cs_prox_tab"] = prox_tab
		self._build_cs_side_ui(prox_tab, side="prox")

		# --- 遠位サブタブ ---
		dist_tab = ttk.Frame(cs_notebook, padding=8)
		cs_notebook.add(dist_tab, text=self._joint_label("cs_dist_tab"))
		self._joint_widgets["cs_dist_tab"] = dist_tab
		self._build_cs_side_ui(dist_tab, side="dist")

		# === 共有: RANSAC/ICPパラメータ ===
		param_frame = ttk.LabelFrame(container, text="フィッティングパラメータ（共有）", style="Bold.TLabelframe")
		param_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 8))
		param_frame.columnconfigure(1, weight=1)
		param_frame.columnconfigure(3, weight=1)
		param_frame.columnconfigure(5, weight=1)

		ttk.Label(param_frame, text="RANSACパラメータ", font=(self.ui_font_family, 9, "bold")).grid(row=0, column=0, columnspan=6, sticky="w", padx=8, pady=(6, 3))
		ttk.Label(param_frame, text="距離閾値 (mm):").grid(row=1, column=0, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_ransac_distance_threshold, width=10).grid(row=1, column=1, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="最大反復回数:").grid(row=1, column=2, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_ransac_max_iterations, width=10).grid(row=1, column=3, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="信頼度:").grid(row=1, column=4, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_ransac_confidence, width=10).grid(row=1, column=5, sticky="w", padx=(0, 8), pady=4)

		ttk.Label(param_frame, text="ICPパラメータ", font=(self.ui_font_family, 9, "bold")).grid(row=2, column=0, columnspan=6, sticky="w", padx=8, pady=(8, 3))
		ttk.Label(param_frame, text="閾値 (mm):").grid(row=3, column=0, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_icp_threshold, width=10).grid(row=3, column=1, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="最大反復回数:").grid(row=3, column=2, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_icp_max_iterations, width=10).grid(row=3, column=3, sticky="w", padx=(0, 8), pady=4)
		ttk.Label(param_frame, text="サンプル点数:").grid(row=3, column=4, sticky="e", padx=(8, 5), pady=4)
		ttk.Entry(param_frame, textvariable=self.cs_sample_points, width=10).grid(row=3, column=5, sticky="w", padx=(0, 8), pady=4)
		btn_row = ttk.Frame(param_frame)
		btn_row.grid(row=4, column=0, columnspan=6, pady=(6, 6))
		ttk.Button(btn_row, text="デフォルトに戻す", command=self.reset_cs_fitting_parameters).pack(side="left", padx=4)
		ttk.Button(btn_row, text="現在の値をデフォルトとして保存", command=self._save_cs_fitting_defaults).pack(side="left", padx=4)

		# === 共有: スケール補正 ===
		scale_frame = ttk.LabelFrame(container, text="スケール補正", style="Bold.TLabelframe")
		scale_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 8))

		ttk.Checkbutton(scale_frame, text="スケール補正を有効化",
			variable=self.cs_enable_scaling).grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		ttk.Label(scale_frame, text="3Dスキャナ（骨+軟骨）と CT（骨のみ）のように撮影装置が異なる場合、\n"
			"モデル間の寸法差（スケール差）を ICP で自動補正します。\n"
			"同一装置で撮影した場合はOFFのままで構いません。",
			foreground="gray", justify="left").grid(row=1, column=0, sticky="w", padx=24, pady=(0, 8))

		# === 共有: 軟骨抽出パラメータ ===
		extract_frame = ttk.LabelFrame(container, text="軟骨抽出パラメータ", style="Bold.TLabelframe")
		extract_frame.grid(row=5, column=0, sticky="nsew", padx=12, pady=(0, 8))

		ttk.Label(extract_frame, text="距離閾値 [mm]:").grid(row=0, column=0, sticky="e", padx=(8, 5), pady=(6, 2))
		ttk.Entry(extract_frame, textvariable=self.cs_distance_threshold, width=8).grid(row=0, column=1, sticky="w", padx=(0, 8), pady=(6, 2))
		ttk.Label(extract_frame, text="骨表面からこの距離以上離れた部分を軟骨として抽出します。\n"
			"値を小さくすると薄い軟骨も拾えます（推奨: 0.1〜0.5mm）。",
			foreground="gray", justify="left").grid(row=1, column=0, columnspan=2, sticky="w", padx=24, pady=(0, 8))

		# === 共有: 可視化オプション ===
		vis_frame = ttk.LabelFrame(container, text="可視化オプション（共有）", style="Bold.TLabelframe")
		vis_frame.grid(row=6, column=0, sticky="nsew", padx=12, pady=(0, 8))
		vis_frame.columnconfigure(1, weight=1); vis_frame.columnconfigure(3, weight=1)
		ttk.Label(vis_frame, text="骨の色:").grid(row=0, column=0, sticky="e", padx=(8, 5), pady=6)
		ttk.Button(vis_frame, text="■ 変更", command=self._choose_cs_bone_color).grid(row=0, column=1, sticky="w", padx=(0, 8), pady=6)
		ttk.Label(vis_frame, text="骨の透明度:").grid(row=0, column=2, sticky="e", padx=(8, 5), pady=6)
		ttk.Entry(vis_frame, textvariable=self.cs_bone_opacity, width=8).grid(row=0, column=3, sticky="w", padx=(0, 8), pady=6)
		ttk.Label(vis_frame, text="軟骨の色:").grid(row=1, column=0, sticky="e", padx=(8, 5), pady=6)
		ttk.Button(vis_frame, text="■ 変更", command=self._choose_cs_cartilage_color).grid(row=1, column=1, sticky="w", padx=(0, 8), pady=6)
		ttk.Label(vis_frame, text="軟骨の透明度:").grid(row=1, column=2, sticky="e", padx=(8, 5), pady=6)
		ttk.Entry(vis_frame, textvariable=self.cs_cartilage_opacity, width=8).grid(row=1, column=3, sticky="w", padx=(0, 8), pady=6)

	def _build_cs_side_ui(self, parent, side: str) -> None:
		"""近位/遠位共通のファイル入力・ボタンUIを構築
		Args:
			side: "prox" or "dist"
		"""
		parent.columnconfigure(0, weight=1)
		side_label = self._joint_label("cs_prox_tab" if side == "prox" else "cs_dist_tab")

		# 変数を取得
		if side == "prox":
			v_m1w = self.cs_prox_model1_whole_path
			v_m1r = self.cs_prox_model1_region_path
			v_m2w = self.cs_prox_model2_whole_path
			v_m2r = self.cs_prox_model2_region_path
		else:
			v_m1w = self.cs_dist_model1_whole_path
			v_m1r = self.cs_dist_model1_region_path
			v_m2w = self.cs_dist_model2_whole_path
			v_m2r = self.cs_dist_model2_region_path

		# モデル1（骨+軟骨）
		m1f = ttk.LabelFrame(parent, text=f"{side_label}: モデル1（骨＋軟骨）", style="Bold.TLabelframe")
		m1f.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3): m1f.columnconfigure(i, weight=1)
		self._add_file_row(parent=m1f, row=0, label_text="全体モデル (STL/OBJ)", textvariable=v_m1w,
			command=lambda s=side: self._choose_cs_file(s, "m1w"))
		self._add_file_row(parent=m1f, row=1, label_text="任意領域 (STL/OBJ) ※共通骨領域", textvariable=v_m1r,
			command=lambda s=side: self._choose_cs_file(s, "m1r"))

		# モデル2（骨のみ）
		m2f = ttk.LabelFrame(parent, text=f"{side_label}: モデル2（骨のみ）", style="Bold.TLabelframe")
		m2f.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
		for i in range(3): m2f.columnconfigure(i, weight=1)
		self._add_file_row(parent=m2f, row=0, label_text="全体モデル (STL/OBJ)", textvariable=v_m2w,
			command=lambda s=side: self._choose_cs_file(s, "m2w"))
		self._add_file_row(parent=m2f, row=1, label_text="任意領域 (STL/OBJ) ※共通骨領域", textvariable=v_m2r,
			command=lambda s=side: self._choose_cs_file(s, "m2r"))

		# 確認 + 実行ボタン
		action_frame = ttk.Frame(parent)
		action_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 8))
		action_frame.columnconfigure(0, weight=1)
		action_frame.columnconfigure(1, weight=1)

		viz_btn = ttk.Button(action_frame, text="モデルを確認",
			command=lambda s=side: self._on_cs_visualize_side(s))
		viz_btn.grid(row=0, column=0, sticky="ew", padx=(0, 4), pady=4)

		exec_btn = ttk.Button(action_frame, text=f"{side_label} 軟骨分離を実行",
			command=lambda s=side: self._on_cs_execute_side(s))
		exec_btn.grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=4)

		# ボタン参照を保持
		if side == "prox":
			self.cs_prox_viz_btn = viz_btn
			self.cs_prox_exec_btn = exec_btn
		else:
			self.cs_dist_viz_btn = viz_btn
			self.cs_dist_exec_btn = exec_btn

		# 変数のtrace
		for v in [v_m1w, v_m1r, v_m2w, v_m2r]:
			v.trace_add("write", lambda *_, s=side: self._update_cs_side_buttons(s))

	# region FEM解析タブ
	def _create_fem_tab(self) -> None:
		"""FEM解析タブのコンテンツを作成"""
		container = self.fem_tab
		container.columnconfigure(0, weight=1)

		if not _HAS_FEM:
			ttk.Label(container, text="FEM解析モジュール (fem_contact_solver_2.py) が見つかりません。\n同じディレクトリに配置してください。",
				foreground="red").grid(row=0, column=0, pady=20)
			return

		row_idx = 0

		# タイトル
		ttk.Label(container, text="FEM 接触解析（複合モデル: 骨+軟骨）",
			font=(self.ui_font_family, 11, "bold")).grid(row=row_idx, column=0, sticky="w", pady=(0, 4))
		row_idx += 1
		ttk.Label(container, text="骨と軟骨を分離したモデルに対し、軟骨面同士の接触応力を有限要素法で解析します。",
			foreground="gray").grid(row=row_idx, column=0, sticky="w", pady=(0, 8))
		row_idx += 1

		# === 入力メッシュ ===
		mesh_frame = ttk.LabelFrame(container, text="入力メッシュ（複合モデル）", style="Bold.TLabelframe")
		mesh_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		mesh_frame.columnconfigure(1, weight=1)
		row_idx += 1

		# 近位
		fem_prox_header = ttk.Label(mesh_frame, text=f"【{self._joint_label('cs_prox_tab')}】", font=(self.ui_font_family, 9, "bold"))
		fem_prox_header.grid(row=0, column=0, columnspan=3, sticky="w", padx=8, pady=(6, 2))
		self._joint_widgets["fem_prox_header"] = fem_prox_header
		ttk.Label(mesh_frame, text="骨モデル:").grid(row=1, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_prox_bone_entry = ttk.Entry(mesh_frame, textvariable=self.fem_prox_bone_path, width=60)
		fem_prox_bone_entry.grid(row=1, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_prox_bone_entry, self.fem_prox_bone_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("prox_bone")).grid(row=1, column=2, padx=(4, 8), pady=2)
		ttk.Label(mesh_frame, text="軟骨モデル:").grid(row=2, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_prox_cart_entry = ttk.Entry(mesh_frame, textvariable=self.fem_prox_cartilage_path, width=60)
		fem_prox_cart_entry.grid(row=2, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_prox_cart_entry, self.fem_prox_cartilage_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("prox_cartilage")).grid(row=2, column=2, padx=(4, 8), pady=2)

		# 遠位
		fem_dist_header = ttk.Label(mesh_frame, text=f"【{self._joint_label('cs_dist_tab')}】", font=(self.ui_font_family, 9, "bold"))
		fem_dist_header.grid(row=3, column=0, columnspan=3, sticky="w", padx=8, pady=(8, 2))
		self._joint_widgets["fem_dist_header"] = fem_dist_header
		ttk.Label(mesh_frame, text="骨モデル:").grid(row=4, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_dist_bone_entry = ttk.Entry(mesh_frame, textvariable=self.fem_dist_bone_path, width=60)
		fem_dist_bone_entry.grid(row=4, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_dist_bone_entry, self.fem_dist_bone_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("dist_bone")).grid(row=4, column=2, padx=(4, 8), pady=2)
		ttk.Label(mesh_frame, text="軟骨モデル:").grid(row=5, column=0, sticky="w", padx=(20, 4), pady=2)
		fem_dist_cart_entry = ttk.Entry(mesh_frame, textvariable=self.fem_dist_cartilage_path, width=60)
		fem_dist_cart_entry.grid(row=5, column=1, sticky="ew", padx=4, pady=2)
		self._register_dnd(fem_dist_cart_entry, self.fem_dist_cartilage_path)
		ttk.Button(mesh_frame, text="参照...", width=8,
			command=lambda: self._choose_fem_file("dist_cartilage")).grid(row=5, column=2, padx=(4, 8), pady=(2, 8))

		# === 材料パラメータ ===
		mat_frame = ttk.LabelFrame(container, text="材料パラメータ", style="Bold.TLabelframe")
		mat_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		# 軟骨
		ttk.Label(mat_frame, text="《軟骨》", font=(self.ui_font_family, 9, "bold")).grid(
			row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		ttk.Label(mat_frame, text="ヤング率 E [MPa]:").grid(row=0, column=1, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_cart_E, width=10).grid(row=0, column=2, padx=4, pady=2)
		ttk.Label(mat_frame, text="ポアソン比 ν:").grid(row=0, column=3, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_cart_nu, width=10).grid(row=0, column=4, padx=4, pady=2)
		ttk.Label(mat_frame, text="厚さ [mm]:").grid(row=0, column=5, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_cart_thickness, width=10).grid(row=0, column=6, padx=(4, 8), pady=2)

		# 骨（参考表示）
		ttk.Label(mat_frame, text="《骨》", font=(self.ui_font_family, 9, "bold")).grid(
			row=1, column=0, sticky="w", padx=8, pady=(4, 2))
		ttk.Label(mat_frame, text="ヤング率 E [MPa]:").grid(row=1, column=1, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_bone_E, width=10).grid(row=1, column=2, padx=4, pady=2)
		ttk.Label(mat_frame, text="ポアソン比 ν:").grid(row=1, column=3, sticky="w", padx=4, pady=2)
		ttk.Entry(mat_frame, textvariable=self.fem_bone_nu, width=10).grid(row=1, column=4, padx=4, pady=2)
		ttk.Label(mat_frame, text="（剛体BC扱い）", foreground="gray").grid(
			row=1, column=5, columnspan=2, sticky="w", padx=4, pady=(4, 8))

		# === 接触パラメータ ===
		contact_frame = ttk.LabelFrame(container, text="接触パラメータ", style="Bold.TLabelframe")
		contact_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		ttk.Label(contact_frame, text="ペナルティ剛性 [MPa/mm]:").grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		ttk.Entry(contact_frame, textvariable=self.fem_penalty_stiffness, width=10).grid(row=0, column=1, padx=4, pady=(6, 2))
		ttk.Label(contact_frame, text="接触判定閾値 [mm]:").grid(row=0, column=2, sticky="w", padx=(16, 4), pady=(6, 2))
		ttk.Entry(contact_frame, textvariable=self.fem_contact_tolerance, width=10).grid(row=0, column=3, padx=(4, 8), pady=(6, 8))

		# === 解析オプション ===
		opt_frame = ttk.LabelFrame(container, text="解析オプション", style="Bold.TLabelframe")
		opt_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		ttk.Label(opt_frame, text="境界条件:").grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		bc_combo = ttk.Combobox(opt_frame, textvariable=self.fem_boundary_mode, values=["auto", "rim"],
			state="readonly", width=8)
		bc_combo.grid(row=0, column=1, padx=4, pady=(6, 2))
		ttk.Label(opt_frame, text="最大節点数:").grid(row=0, column=2, sticky="w", padx=(16, 4), pady=(6, 2))
		ttk.Entry(opt_frame, textvariable=self.fem_max_nodes, width=10).grid(row=0, column=3, padx=4, pady=(6, 2))
		ttk.Button(opt_frame, text="デフォルトに戻す", command=self._reset_fem_parameters).grid(
			row=0, column=4, padx=(16, 8), pady=(6, 8))

		# === 実行ボタン ===
		btn_frame = ttk.Frame(container)
		btn_frame.grid(row=row_idx, column=0, sticky="w", pady=(0, 8))
		row_idx += 1

		self.fem_execute_btn = ttk.Button(btn_frame, text="解析を実行", command=self._on_fem_execute)
		self.fem_execute_btn.pack(side="left", padx=(0, 8))
		self.fem_visualize_btn = ttk.Button(btn_frame, text="結果を可視化", command=self._on_fem_visualize, state="disabled")
		self.fem_visualize_btn.pack(side="left", padx=(0, 8))
		self.fem_report_btn = ttk.Button(btn_frame, text="レポート生成", command=self._on_fem_generate_report, state="disabled")
		self.fem_report_btn.pack(side="left", padx=(0, 8))

		# === 結果サマリー ===
		result_frame = ttk.LabelFrame(container, text="結果サマリー", style="Bold.TLabelframe")
		result_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		result_frame.columnconfigure(0, weight=1)
		row_idx += 1

		self.fem_result_text = tk.Text(result_frame, height=10, state="disabled", wrap="word",
			font=(self.ui_font_family, 9))
		fem_result_scroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.fem_result_text.yview)
		self.fem_result_text.configure(yscrollcommand=fem_result_scroll.set)
		self.fem_result_text.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
		fem_result_scroll.grid(row=0, column=1, sticky="ns", padx=(0, 8), pady=8)

		# === 可視化オプション ===
		vis_frame = ttk.LabelFrame(container, text="可視化オプション", style="Bold.TLabelframe")
		vis_frame.grid(row=row_idx, column=0, sticky="nsew", pady=(0, 8))
		row_idx += 1

		ttk.Label(vis_frame, text="表示スカラー:").grid(row=0, column=0, sticky="w", padx=8, pady=(6, 2))
		scalar_combo = ttk.Combobox(vis_frame, textvariable=self.fem_scalar_name,
			values=["contact_pressure", "von_mises_stress", "max_principal_stress",
				"displacement_magnitude", "penetration_depth"],
			state="readonly", width=24)
		scalar_combo.grid(row=0, column=1, padx=4, pady=(6, 2))

		ttk.Label(vis_frame, text="カラーマップ:").grid(row=0, column=2, sticky="w", padx=(16, 4), pady=(6, 2))
		cmap_combo = ttk.Combobox(vis_frame, textvariable=self.fem_cmap,
			values=["jet", "plasma", "coolwarm", "viridis", "hot"],
			state="readonly", width=12)
		cmap_combo.grid(row=0, column=3, padx=4, pady=(6, 2))

		ttk.Checkbutton(vis_frame, text="相手軟骨を表示", variable=self.fem_show_dist_mesh).grid(
			row=1, column=0, columnspan=2, sticky="w", padx=8, pady=(2, 8))
		ttk.Checkbutton(vis_frame, text="骨を半透明表示", variable=self.fem_show_bone).grid(
			row=1, column=2, columnspan=2, sticky="w", padx=8, pady=(2, 8))

		# ボタン状態の初期設定
		self._update_fem_buttons()

		# パス変更時にボタン状態を更新
		for v in [self.fem_prox_cartilage_path, self.fem_dist_cartilage_path]:
			v.trace_add("write", lambda *_: self._update_fem_buttons())
	# --- FEM解析 コールバックメソッド群 ---

	def _choose_fem_file(self, which: str) -> None:
		"""FEM解析タブ用ファイル選択ダイアログ"""
		titles = {
			"prox_bone": "近位 骨モデル",
			"prox_cartilage": "近位 軟骨モデル",
			"dist_bone": "遠位 骨モデル",
			"dist_cartilage": "遠位 軟骨モデル",
		}
		var_map = {
			"prox_bone": self.fem_prox_bone_path,
			"prox_cartilage": self.fem_prox_cartilage_path,
			"dist_bone": self.fem_dist_bone_path,
			"dist_cartilage": self.fem_dist_cartilage_path,
		}
		path = filedialog.askopenfilename(
			title=f"FEM解析: {titles.get(which, which)}を選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			var_map[which].set(path)

	def _update_fem_buttons(self) -> None:
		"""FEM解析タブのボタン有効/無効制御"""
		has_cartilage = (bool(self.fem_prox_cartilage_path.get().strip()) and
						 bool(self.fem_dist_cartilage_path.get().strip()))
		try:
			self.fem_execute_btn.config(state="normal" if (has_cartilage and not self.fem_is_running) else "disabled")
		except Exception:
			pass
		try:
			self.fem_visualize_btn.config(state="normal" if self.fem_last_results is not None else "disabled")
		except Exception:
			pass
		try:
			self.fem_report_btn.config(state="normal" if self.fem_last_results is not None else "disabled")
		except Exception:
			pass

	def _reset_fem_parameters(self) -> None:
		"""FEM解析パラメータをデフォルト値に戻す"""
		self.fem_cart_E.set(10.0)
		self.fem_cart_nu.set(0.45)
		self.fem_cart_thickness.set(2.0)
		self.fem_bone_E.set(17000.0)
		self.fem_bone_nu.set(0.3)
		self.fem_penalty_stiffness.set(500.0)
		self.fem_contact_tolerance.set(2.0)
		self.fem_boundary_mode.set("auto")
		self.fem_max_nodes.set(30000)

	def _on_fem_execute(self) -> None:
		"""FEM接触解析を実行（バックグラウンドスレッド）"""
		if not _HAS_FEM:
			messagebox.showerror("エラー", "FEMモジュールが利用できません。")
			return

		prox_cart_path = self.fem_prox_cartilage_path.get().strip()
		dist_cart_path = self.fem_dist_cartilage_path.get().strip()

		if not prox_cart_path or not dist_cart_path:
			messagebox.showerror("エラー", "近位・遠位の軟骨モデルを選択してください。")
			return

		for p, name in [(prox_cart_path, "近位軟骨"), (dist_cart_path, "遠位軟骨")]:
			if not Path(p).exists():
				messagebox.showerror("エラー", f"{name}ファイルが見つかりません:\n{p}")
				return

		# ボタン無効化
		self.fem_is_running = True
		self._update_fem_buttons()

		# 結果テキストをクリア
		self.fem_result_text.config(state="normal")
		self.fem_result_text.delete("1.0", "end")
		self.fem_result_text.insert("1.0", "解析を実行中...")
		self.fem_result_text.config(state="disabled")

		# 骨モデルパス（可視化用、解析には不要）
		prox_bone_path = self.fem_prox_bone_path.get().strip()
		dist_bone_path = self.fem_dist_bone_path.get().strip()

		def run_analysis():
			try:
				prox_cart = pv.read(prox_cart_path)
				dist_cart = pv.read(dist_cart_path)

				# 骨メッシュ（可視化用に読み込み）
				prox_bone = pv.read(prox_bone_path) if prox_bone_path and Path(prox_bone_path).exists() else None
				dist_bone = pv.read(dist_bone_path) if dist_bone_path and Path(dist_bone_path).exists() else None

				material = MaterialProperties(
					E=self.fem_cart_E.get(),
					nu=self.fem_cart_nu.get(),
					thickness=self.fem_cart_thickness.get(),
				)
				contact = ContactParameters(
					penalty_stiffness=self.fem_penalty_stiffness.get(),
					contact_tolerance=self.fem_contact_tolerance.get(),
				)
				solver = FEMContactSolver(material=material, contact=contact, verbose=True, shared_cache=self._shared_fem_cache)
				results = solver.analyze(
					prox_cart, dist_cart,
					boundary_mode=self.fem_boundary_mode.get(),
					max_nodes=self.fem_max_nodes.get(),
				)
				self._fem_thread_result = results
				self._fem_thread_prox_cart = prox_cart
				self._fem_thread_dist_cart = dist_cart
				self._fem_thread_prox_bone = prox_bone
				self._fem_thread_dist_bone = dist_bone
				self._fem_thread_error = None
			except Exception as e:
				self._fem_thread_result = None
				self._fem_thread_error = e
				import traceback
				traceback.print_exc()

		self._fem_thread = threading.Thread(target=run_analysis, daemon=True)
		self._fem_thread.start()
		self._poll_fem_thread()

	def _poll_fem_thread(self) -> None:
		"""FEM解析スレッドの完了チェック"""
		if self._fem_thread.is_alive():
			self.after(200, self._poll_fem_thread)
			return

		self.fem_is_running = False

		if hasattr(self, '_fem_thread_error') and self._fem_thread_error is not None:
			self.fem_result_text.config(state="normal")
			self.fem_result_text.delete("1.0", "end")
			self.fem_result_text.insert("1.0", f"解析エラー:\n{self._fem_thread_error}")
			self.fem_result_text.config(state="disabled")
			messagebox.showerror("FEM解析エラー", str(self._fem_thread_error))
		elif hasattr(self, '_fem_thread_result') and self._fem_thread_result is not None:
			self.fem_last_results = self._fem_thread_result
			self.fem_last_prox_cart_mesh = self._fem_thread_prox_cart
			self.fem_last_dist_cart_mesh = self._fem_thread_dist_cart
			self.fem_last_prox_bone_mesh = getattr(self, '_fem_thread_prox_bone', None)
			self.fem_last_dist_bone_mesh = getattr(self, '_fem_thread_dist_bone', None)

			# サマリー表示
			self.fem_result_text.config(state="normal")
			self.fem_result_text.delete("1.0", "end")
			self.fem_result_text.insert("1.0", self.fem_last_results.summary())
			self.fem_result_text.config(state="disabled")
			messagebox.showinfo("完了", "FEM接触解析が完了しました。")

		self._update_fem_buttons()

	def _on_fem_visualize(self) -> None:
		"""FEM解析結果をPyVistaで可視化"""
		if self.fem_last_results is None or self.fem_last_prox_cart_mesh is None:
			messagebox.showwarning("警告", "先に解析を実行してください。")
			return

		results = self.fem_last_results
		scalar_name = self.fem_scalar_name.get()
		cmap = self.fem_cmap.get()

		# 近位軟骨メッシュにFEM結果を適用
		mesh_vis = self.fem_last_prox_cart_mesh.copy()
		if results.n_nodes == mesh_vis.n_points:
			apply_fem_results_to_mesh(mesh_vis, results)
		else:
			messagebox.showwarning("警告",
				f"FEM結果の節点数({results.n_nodes})がメッシュ({mesh_vis.n_points})と不一致です。\n"
				"デシメーションが適用された可能性があります。")
			return

		# スカラーラベル定義
		scalar_labels = {
			'contact_pressure': '接触圧 [MPa]',
			'von_mises_stress': 'von Mises応力 [MPa]',
			'max_principal_stress': '最大主応力 [MPa]',
			'displacement_magnitude': '変位量 [mm]',
			'penetration_depth': '侵入量 [mm]',
		}
		label = scalar_labels.get(scalar_name, scalar_name)

		# プロッター構築
		plotter = pv.Plotter(window_size=(1200, 800))
		plotter.set_background('white')

		# 近位軟骨（FEM結果）
		plotter.add_mesh(
			mesh_vis, scalars=scalar_name, cmap=cmap, show_edges=False,
			scalar_bar_args={'title': label, 'color': 'black'},
			opacity=1.0,
		)

		# 遠位軟骨（半透明）
		if self.fem_show_dist_mesh.get() and self.fem_last_dist_cart_mesh is not None:
			plotter.add_mesh(
				self.fem_last_dist_cart_mesh, color='#FFB6C1', opacity=0.3, show_edges=False,
			)

		# 骨（半透明）
		if self.fem_show_bone.get():
			if self.fem_last_prox_bone_mesh is not None:
				plotter.add_mesh(
					self.fem_last_prox_bone_mesh, color='#F5DEB3', opacity=0.2, show_edges=False,
				)
			if self.fem_last_dist_bone_mesh is not None:
				plotter.add_mesh(
					self.fem_last_dist_bone_mesh, color='#DEB887', opacity=0.2, show_edges=False,
				)

		plotter.add_axes()

		# サマリーテキスト
		text = (
			f"接触面積: {results.contact_area:.1f} mm²\n"
			f"最大接触圧: {results.peak_contact_pressure:.3f} MPa\n"
			f"接触節点: {results.n_contact_nodes:,}\n"
			f"解析時間: {results.solve_time_sec:.1f} 秒"
		)
		plotter.add_text(text, position='upper_right', font_size=10, color='black')

		plotter.show()

	def _on_fem_generate_report(self) -> None:
		"""FEM解析レポートをExcel+Word形式で自動生成（FEMタブから実行）"""
		if not _HAS_REPORT:
			messagebox.showerror("エラー",
				"レポート生成モジュールが利用できません。\n"
				"fem_report_generator_v2.py が同じフォルダにあることを確認してください。\n"
				"また以下がインストールされていることを確認してください:\n"
				"  pip install python-docx matplotlib openpyxl")
			return

		if self.fem_last_results is None:
			messagebox.showwarning("警告", "先にFEM解析を実行してください。")
			return

		# 保存先フォルダを選択
		from tkinter import filedialog
		output_dir = filedialog.askdirectory(title="レポート保存先フォルダを選択")
		if not output_dir:
			return

		# プログレス表示
		self.fem_report_btn.config(state="disabled", text="生成中...")
		self.root.update_idletasks()

		try:
			from fem_report_generator_v2 import generate_report_from_simulator
			generate_report_from_simulator(self, output_dir=output_dir)
			messagebox.showinfo("完了", f"レポートを保存しました:\n{output_dir}")
		except Exception as e:
			messagebox.showerror("エラー", f"レポート生成中にエラーが発生しました:\n{e}")
		finally:
			self.fem_report_btn.config(state="normal", text="レポート生成")

	def _run_fem_on_current_frame(self, prox_cartilage_mesh, dist_cartilage_mesh,
								   prox_bone_mesh=None, dist_bone_mesh=None) -> None:
		"""Simulator連携: 現在フレームの姿勢でFEM解析を実行

		on_animate() 内のFEMボタンから呼ばれる。
		メッシュは既にアニメーション姿勢に変換済み。
		"""
		if not _HAS_FEM:
			messagebox.showerror("エラー", "FEMモジュールが利用できません。")
			return

		if prox_cartilage_mesh is None or dist_cartilage_mesh is None:
			messagebox.showwarning("警告", "軟骨モデルが読み込まれていません。\nSimulatorタブで軟骨モデルを指定してください。")
			return

		try:
			material = MaterialProperties(
				E=self.fem_cart_E.get(),
				nu=self.fem_cart_nu.get(),
				thickness=self.fem_cart_thickness.get(),
			)
			contact = ContactParameters(
				penalty_stiffness=self.fem_penalty_stiffness.get(),
				contact_tolerance=self.fem_contact_tolerance.get(),
			)
			solver = FEMContactSolver(material=material, contact=contact, verbose=True, shared_cache=self._shared_fem_cache)

			# 現在のメッシュ位置のコピーで解析
			prox_cart_copy = prox_cartilage_mesh.copy()
			dist_cart_copy = dist_cartilage_mesh.copy()

			# キャッシュ存在確認 → 状態ウィンドウ表示
			boundary_mode = self.fem_boundary_mode.get()
			max_nodes = self.fem_max_nodes.get()
			from_cache = solver.has_cache(prox_cart_copy, dist_cart_copy, boundary_mode, max_nodes)
			status_win = tk.Toplevel(self)
			status_win.title("FEM解析")
			status_win.resizable(False, False)
			status_win.transient(self)
			msg = "キャッシュからFEM結果を読み込んでいます..." if from_cache else "FEM接触解析を実行中..."
			ttk.Label(status_win, text=msg, padding=20).pack()
			status_win.update_idletasks()
			x = self.winfo_rootx() + (self.winfo_width() - status_win.winfo_width()) // 2
			y = self.winfo_rooty() + (self.winfo_height() - status_win.winfo_height()) // 2
			status_win.geometry(f"+{x}+{y}")
			status_win.update()

			results = solver.analyze(
				prox_cart_copy, dist_cart_copy,
				boundary_mode=boundary_mode,
				max_nodes=max_nodes,
			)
			status_win.destroy()

			# 結果を可視化
			scalar_name = self.fem_scalar_name.get()
			cmap = self.fem_cmap.get()

			mesh_vis = prox_cart_copy.copy()
			if results.n_nodes == mesh_vis.n_points:
				apply_fem_results_to_mesh(mesh_vis, results)
			else:
				# デシメーション時は解析メッシュを使用
				mesh_vis = pv.PolyData(solver._nodes, prox_cart_copy.faces)
				apply_fem_results_to_mesh(mesh_vis, results)

			scalar_labels = {
				'contact_pressure': '接触圧 [MPa]',
				'von_mises_stress': 'von Mises応力 [MPa]',
				'max_principal_stress': '最大主応力 [MPa]',
				'displacement_magnitude': '変位量 [mm]',
				'penetration_depth': '侵入量 [mm]',
			}
			label = scalar_labels.get(scalar_name, scalar_name)

			plotter = pv.Plotter(window_size=(1200, 800))
			plotter.set_background('white')

			plotter.add_mesh(
				mesh_vis, scalars=scalar_name, cmap=cmap, show_edges=False,
				scalar_bar_args={'title': label, 'color': 'black'},
				opacity=1.0,
			)

			# 遠位軟骨
			plotter.add_mesh(dist_cart_copy, color='#FFB6C1', opacity=0.3, show_edges=False)

			# 骨（半透明）
			if prox_bone_mesh is not None:
				plotter.add_mesh(prox_bone_mesh.copy(), color='#F5DEB3', opacity=0.2, show_edges=False)
			if dist_bone_mesh is not None:
				plotter.add_mesh(dist_bone_mesh.copy(), color='#DEB887', opacity=0.2, show_edges=False)

			plotter.add_axes()
			text = (
				f"接触面積: {results.contact_area:.1f} mm²\n"
				f"最大接触圧: {results.peak_contact_pressure:.3f} MPa\n"
				f"接触節点: {results.n_contact_nodes:,}\n"
				f"解析時間: {results.solve_time_sec:.1f} 秒"
			)
			plotter.add_text(text, position='upper_right', font_size=10, color='black')
			plotter.show()

		except Exception as e:
			try:
				status_win.destroy()
			except Exception:
				pass
			messagebox.showerror("FEM解析エラー", f"解析中にエラーが発生しました:\n{e}")
			import traceback
			traceback.print_exc()

	# endregion FEM解析タブ

	def _create_stl2asc_tab(self) -> None:
		"""STL2ASCタブのコンテンツを作成"""
		container = self.stl2asc_tab
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="STL to ASC Converter",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# 説明
		info_label = ttk.Label(
			container,
			text="STLファイルから点群データ(ASC形式)を抽出・変換します。",
			justify=tk.LEFT,
			wraplength=600
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))

		# ファイル選択変数
		self.stl2asc_input_path = tk.StringVar(value="")

		# ファイル選択UI
		file_frame = ttk.LabelFrame(container, text="STLファイル選択", style="Bold.TLabelframe")
		file_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		file_frame.columnconfigure(1, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="STLファイル (.stl)",
			textvariable=self.stl2asc_input_path,
			command=self.choose_stl_file
		)

		# 変換ボタン
		action_frame = ttk.LabelFrame(container, text="操作", style="Bold.TLabelframe")
		action_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		
		self.stl2asc_convert_button = ttk.Button(
			action_frame,
			text="ASCとして保存",
			command=self.on_stl2asc_convert
		)
		self.stl2asc_convert_button.pack(fill="x", padx=12, pady=12)

	def choose_stl_file(self) -> None:
		"""STLファイルを選択"""
		path = filedialog.askopenfilename(
			title="STLファイルを選択",
			filetypes=[("STL files", "*.stl"), ("All files", "*.*")]
		)
		if path:
			self.stl2asc_input_path.set(path)

	def on_stl2asc_convert(self) -> None:
		"""STLをASCに変換して保存"""
		stl_path = self.stl2asc_input_path.get().strip()
		if not stl_path:
			messagebox.showwarning("警告", "STLファイルを選択してください。")
			return
		
		if not os.path.exists(stl_path):
			messagebox.showerror("エラー", f"ファイルが見つかりません:\n{stl_path}")
			return
			
		# 保存先を選択
		asc_path = filedialog.asksaveasfilename(
			title="ASCファイルとして保存",
			defaultextension=".asc",
			filetypes=[("ASC files", "*.asc"), ("All files", "*.*")],
			initialfile=Path(stl_path).stem + ".asc"
		)
		
		if not asc_path:
			return

		try:
			# PyVistaを使って読み込み
			mesh = pv.read(stl_path)
			points = mesh.points
			
			if points is None or len(points) == 0:
				messagebox.showerror("エラー", "点群データを読み込めませんでした。")
				return

			# ASCとして保存 (X Y Z format)
			np.savetxt(asc_path, points, fmt="%.8f", delimiter=" ")
			
			messagebox.showinfo("完了", f"保存しました:\n{asc_path}\n点数: {len(points)}")
			
		except Exception as e:
			messagebox.showerror("エラー", f"変換に失敗しました:\n{e}\n{traceback.format_exc()}")

	def _create_randomizer_tab(self) -> None:
		"""Randomizerタブのコンテンツを作成"""
		container = self.randomizer_tab
		container.columnconfigure(0, weight=1)

		# タイトル
		title_label = ttk.Label(
			container,
			text="Random Transform (ワールド座標ランダム化)",
			font=(self.ui_font_family, 12, "bold"),
		)
		title_label.grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))

		# 説明
		info_label = ttk.Label(
			container,
			text="選択したSTL/OBJモデルに対し、指定範囲内でランダムな平行移動・回転を適用して保存します。",
			justify=tk.LEFT,
			wraplength=600
		)
		info_label.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 12))

		# ファイル選択変数
		# (Note: self.randomizer_input_path, self.random_trans_range, self.random_rot_range are initialized in __init__)

		# ファイル選択
		file_frame = ttk.LabelFrame(container, text="対象モデル選択", style="Bold.TLabelframe")
		file_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 12))
		file_frame.columnconfigure(1, weight=1)

		self._add_file_row(
			parent=file_frame,
			row=0,
			label_text="モデルファイル (STL/OBJ)",
			textvariable=self.randomizer_input_path,
			command=self.choose_randomizer_input
		)

		# パラメータ設定
		param_frame = ttk.LabelFrame(container, text="ランダム化パラメータ", style="Bold.TLabelframe")
		param_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 12))
		param_frame.columnconfigure(1, weight=1)

		# 平行移動範囲
		ttk.Label(param_frame, text="平行移動範囲 (±mm):").grid(row=0, column=0, sticky="e", padx=(12, 6), pady=6)
		ttk.Entry(param_frame, textvariable=self.random_trans_range).grid(row=0, column=1, sticky="w", padx=(0, 12), pady=6)
		
		# 回転範囲
		ttk.Label(param_frame, text="回転範囲 (±deg):").grid(row=1, column=0, sticky="e", padx=(12, 6), pady=6)
		ttk.Entry(param_frame, textvariable=self.random_rot_range).grid(row=1, column=1, sticky="w", padx=(0, 12), pady=6)

		# 実行ボタン
		action_frame = ttk.Frame(container)
		action_frame.grid(row=4, column=0, sticky="nsew", padx=12, pady=(0, 12))

		self.randomizer_execute_button = ttk.Button(
			action_frame,
			text="ランダム変換を実行して保存...",
			command=self.on_randomizer_execute
		)
		self.randomizer_execute_button.pack(fill="x", pady=12)
		
	def _create_pos_checker_tab(self) -> None:
		"""PosCheckerタブのコンテンツを作成"""
		container = self.pos_checker_tab
		container.columnconfigure(0, weight=1)

		# Variables
		self.pos_check_model_path = tk.StringVar(value="")
		self.pos_check_pp_pv_path = tk.StringVar(value="")
		self.pos_check_pp_fm_path = tk.StringVar(value="")

		# Model Input
		model_frame = ttk.LabelFrame(container, text="1. モデル選択 (STL/OBJ)", style="Bold.TLabelframe")
		model_frame.pack(fill="x", padx=10, pady=5)
		model_frame.columnconfigure(1, weight=1)

		ttk.Label(model_frame, text="モデルファイル:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
		pc_model_entry = ttk.Entry(model_frame, textvariable=self.pos_check_model_path)
		pc_model_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_model_entry, self.pos_check_model_path)
		ttk.Button(model_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_model_path, "model")).grid(row=0, column=2, padx=5, pady=5)

		# PP Inputs
		pp_frame = ttk.LabelFrame(container, text="2. PPファイル選択", style="Bold.TLabelframe")
		pp_frame.pack(fill="x", padx=10, pady=5)
		pp_frame.columnconfigure(1, weight=1)

		# PV (Proximal/Pelvis)
		ttk.Label(pp_frame, text="PP (PV/Proximal):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
		pc_pv_entry = ttk.Entry(pp_frame, textvariable=self.pos_check_pp_pv_path)
		pc_pv_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_pv_entry, self.pos_check_pp_pv_path)
		ttk.Button(pp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_pp_pv_path, "pp")).grid(row=0, column=2, padx=5, pady=5)
		ttk.Label(pp_frame, text="※ Simulator基本機能(A,B,C,D,O)で構築").grid(row=1, column=1, sticky="w", padx=5)

		# FM (Distal/Femur)
		ttk.Label(pp_frame, text="PP (FM/Distal):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
		pc_fm_entry = ttk.Entry(pp_frame, textvariable=self.pos_check_pp_fm_path)
		pc_fm_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_fm_entry, self.pos_check_pp_fm_path)
		ttk.Button(pp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_pp_fm_path, "pp")).grid(row=2, column=2, padx=5, pady=5)
		ttk.Label(pp_frame, text="※ Simulator遠位機能(A,B,O)で構築").grid(row=3, column=1, sticky="w", padx=5)

		# Execute 1
		btn_frame = ttk.Frame(container)
		btn_frame.pack(fill="x", padx=10, pady=10)
		ttk.Button(btn_frame, text="機能1: 可視化 & 変換行列(Cp->Cf)算出", command=self.on_pos_checker_execute).pack(fill="x", ipady=5)
		
		# Separator
		ttk.Separator(container, orient="horizontal").pack(fill="x", padx=10, pady=10)
		
		# Function 2: Matrix Comparison
		# Variables for Function 2
		self.pos_check_mat1_path = tk.StringVar(value="")
		self.pos_check_mat2_path = tk.StringVar(value="")

		comp_frame = ttk.LabelFrame(container, text="機能2: 変換行列の比較 (Input 1 ➙ Input 2)", style="Bold.TLabelframe")
		comp_frame.pack(fill="x", padx=10, pady=5)
		comp_frame.columnconfigure(1, weight=1)

		# Input 1
		ttk.Label(comp_frame, text="Input 1 (行列ファイル):").grid(row=0, column=0, padx=5, pady=5, sticky="e")
		pc_mat1_entry = ttk.Entry(comp_frame, textvariable=self.pos_check_mat1_path)
		pc_mat1_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_mat1_entry, self.pos_check_mat1_path)
		ttk.Button(comp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_mat1_path, "txt")).grid(row=0, column=2, padx=5, pady=5)

		# Input 2
		ttk.Label(comp_frame, text="Input 2 (行列ファイル):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
		pc_mat2_entry = ttk.Entry(comp_frame, textvariable=self.pos_check_mat2_path)
		pc_mat2_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
		self._register_dnd(pc_mat2_entry, self.pos_check_mat2_path)
		ttk.Button(comp_frame, text="参照...", command=lambda: self._choose_pos_check_file(self.pos_check_mat2_path, "txt")).grid(row=1, column=2, padx=5, pady=5)

		# Execute 2
		btn_frame2 = ttk.Frame(container)
		btn_frame2.pack(fill="x", padx=10, pady=10)
		ttk.Button(btn_frame2, text="機能2: 比較行列算出 & 保存", command=self.on_pos_checker_compare).pack(fill="x", ipady=5)

	def _choose_pos_check_file(self, target_var: tk.StringVar, ftype: str) -> None:
		if ftype == "model":
			filetypes = [("3D Model", "*.stl *.obj"), ("All files", "*.*")]
			title = "モデルファイルを選択"
		elif ftype == "pp":
			filetypes = [("PickedPoints", "*.pp"), ("All files", "*.*")]
			title = "PPファイルを選択"
		else: # txt
			filetypes = [("Text files", "*.txt"), ("All files", "*.*")]
			title = "変換行列ファイルを選択"
			
		path = filedialog.askopenfilename(title=title, filetypes=filetypes)
		if path:
			target_var.set(path)

	def on_pos_checker_execute(self) -> None:
		"""PosChecker実行: 可視化と行列計算"""
		model_path = self.pos_check_model_path.get().strip()
		pp_pv_path = self.pos_check_pp_pv_path.get().strip()
		pp_fm_path = self.pos_check_pp_fm_path.get().strip()

		if not all([model_path, pp_pv_path, pp_fm_path]):
			messagebox.showwarning("入力不足", "全てのファイルを選択してください。")
			return

		try:
			import pyvista as pv
			
			# 1. Load Data
			mesh = pv.read(model_path)
			pts_pv, lbl_pv = self._parse_pp_file(pp_pv_path)
			pts_fm, lbl_fm = self._parse_pp_file(pp_fm_path)
			
			# 2. Build Coordinate Systems
			# PV (Proximal): A, B, C, D, O
			try:
				res_pv = self._build_coordinate_system(pts_pv, lbl_pv)
				origin_pv, x_pv, y_pv, z_pv = res_pv
			except Exception as e:
				messagebox.showerror("座標系構築エラー (PV)", f"PV(Proximal)側の座標系構築に失敗しました:\n{e}")
				return

			# FM (Distal): A, B, O
			try:
				res_fm = self._build_coordinate_system_dist(pts_fm, lbl_fm)
				origin_fm, x_fm, y_fm, z_fm = res_fm
			except Exception as e:
				messagebox.showerror("座標系構築エラー (FM)", f"FM(Distal)側の座標系構築に失敗しました:\n{e}")
				return

			# 3. Calculate Transform Cp -> Cf
			# T_pv_world: PV to World
			R_pv = np.column_stack([x_pv, y_pv, z_pv])
			T_pv_world = np.eye(4)
			T_pv_world[:3, :3] = R_pv
			T_pv_world[:3, 3] = origin_pv

			# T_fm_world: FM to World
			R_fm = np.column_stack([x_fm, y_fm, z_fm])
			T_fm_world = np.eye(4)
			T_fm_world[:3, :3] = R_fm
			T_fm_world[:3, 3] = origin_fm
			
			# T_pv_fm: FM relative to PV (Coordinate of FM origin observed from PV, etc?)
			# "CpからCfまでの同時変換行列" implies taking a point in Cf and expressing it in Cp?
			# Or expressing the frame Cf in Cp.
			# T_relative = T_pv_world^(-1) * T_fm_world
			T_world_pv = np.linalg.inv(T_pv_world)
			T_rel = T_world_pv @ T_fm_world
			
			# 4. Display Result
			self._show_pos_checker_result(T_rel, mesh, pts_pv, lbl_pv, pts_fm, lbl_fm, 
										  (origin_pv, x_pv, y_pv, z_pv), 
										  (origin_fm, x_fm, y_fm, z_fm))

		except Exception as e:
			messagebox.showerror("エラー", f"実行中にエラーが発生しました:\n{e}")
			traceback.print_exc()

	def _show_pos_checker_result(self, matrix, mesh, pts_pv, lbl_pv, pts_fm, lbl_fm, cs_pv, cs_fm):
		# Result Window
		res_win = tk.Toplevel(self)
		res_win.title("PosChecker 結果")
		res_win.geometry("600x500")
		
		# Text Area
		text_frame = ttk.Frame(res_win)
		text_frame.pack(fill="both", expand=True, padx=10, pady=10)
		
		result_text = "【同時変換行列 (Cp -> Cf)】\n"
		result_text += "※ Cp(PV)座標系から見たCf(FM)座標系の位置姿勢\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{matrix[0][0]:12.6f} {matrix[0][1]:12.6f} {matrix[0][2]:12.6f} {matrix[0][3]:12.6f}\n"
		result_text += f"{matrix[1][0]:12.6f} {matrix[1][1]:12.6f} {matrix[1][2]:12.6f} {matrix[1][3]:12.6f}\n"
		result_text += f"{matrix[2][0]:12.6f} {matrix[2][1]:12.6f} {matrix[2][2]:12.6f} {matrix[2][3]:12.6f}\n"
		result_text += f"{matrix[3][0]:12.6f} {matrix[3][1]:12.6f} {matrix[3][2]:12.6f} {matrix[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"
		
		# Translation / Rotation analysis
		trans = matrix[:3, 3]
		rot_mat = matrix[:3, :3]
		# Simple Euler (XYZ)
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0

		result_text += "【解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"

		st = tk.Text(text_frame, wrap="none", font=(self.mono_font_family, 10))
		st.pack(side="left", fill="both", expand=True)
		sb = ttk.Scrollbar(text_frame, orient="vertical", command=st.yview)
		sb.pack(side="right", fill="y")
		st.config(yscrollcommand=sb.set)
		st.insert("1.0", result_text)
		st.config(state="disabled") # Read-only
		
		# Buttons
		btn_frame = ttk.Frame(res_win)
		btn_frame.pack(fill="x", padx=10, pady=10)
		
		def save_txt():
			path = filedialog.asksaveasfilename(
				title="結果を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if path:
				try:
					with open(path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"保存しました:\n{path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(btn_frame, text="テキストとして保存...", command=save_txt).pack(side="left", padx=5)
		ttk.Button(btn_frame, text="閉じる", command=res_win.destroy).pack(side="right", padx=5)

		# Visualize (PyVista)
		# numpy array の真偽値判定エラーを回避するため、Noneチェックを明示的に行う
		if mesh is not None and pts_pv is not None and pts_fm is not None: 
			# Create Plotter
			plotter = pv.Plotter(title="PosChecker Visualization")
			plotter.set_background("white")
			
			# Add Mesh
			plotter.add_mesh(mesh, color="lightgray", opacity=0.5, show_edges=False)
			
			# Add PV Points (Blue-ish)
			plotter.add_points(pts_pv, color="blue", point_size=10, render_points_as_spheres=True)
			plotter.add_point_labels(pts_pv, lbl_pv, font_size=12, text_color="blue", show_points=False)
			
			# Add FM Points (Red-ish)
			plotter.add_points(pts_fm, color="red", point_size=10, render_points_as_spheres=True)
			plotter.add_point_labels(pts_fm, lbl_fm, font_size=12, text_color="red", show_points=False)
			
			# Visualize Coordinate Systems
			origin_pv, x_pv, y_pv, z_pv = cs_pv
			origin_fm, x_fm, y_fm, z_fm = cs_fm
			
			axis_len = 50.0
			
			# PV System (Solid lines)
			plotter.add_arrows(origin_pv, x_pv, mag=axis_len, color='red', label='Cp X')
			plotter.add_arrows(origin_pv, y_pv, mag=axis_len, color='green', label='Cp Y')
			plotter.add_arrows(origin_pv, z_pv, mag=axis_len, color='blue', label='Cp Z')
			plotter.add_mesh(pv.Sphere(radius=2, center=origin_pv), color="black")
			plotter.add_point_labels(np.array([origin_pv]), ["Cp"], point_size=0, text_color="black", font_size=16, show_points=False)

			# FM System (Dotted lines? PyVista arrows usually solid. Use lighter colors or tubes)
			# Use slightly different colors (Magenta, Yellow, Cyan) to distinguish
			plotter.add_arrows(origin_fm, x_fm, mag=axis_len, color='magenta', label='Cf X')
			plotter.add_arrows(origin_fm, y_fm, mag=axis_len, color='orange', label='Cf Y') # Changed yellow to orange for visibility
			plotter.add_arrows(origin_fm, z_fm, mag=axis_len, color='cyan', label='Cf Z')
			plotter.add_mesh(pv.Sphere(radius=2, center=origin_fm), color="gray")
			plotter.add_point_labels(np.array([origin_fm]), ["Cf"], point_size=0, text_color="gray", font_size=16, show_points=False)

			plotter.add_legend()
			plotter.show()

	def on_pos_checker_compare(self) -> None:
		"""PosChecker 機能2: 変換行列の比較 (Input 1 -> Input 2)"""
		mat1_path = self.pos_check_mat1_path.get().strip()
		mat2_path = self.pos_check_mat2_path.get().strip()

		if not all([mat1_path, mat2_path]):
			messagebox.showwarning("入力不足", "Input 1 と Input 2 の両方の行列ファイルを選択してください。")
			return

		try:
			# 1. Parse Matrices
			T1 = self._parse_matrix_file(mat1_path)
			T2 = self._parse_matrix_file(mat2_path)

			if T1 is None:
				messagebox.showerror("形式エラー", f"Input 1 のファイルを解析できませんでした:\n{mat1_path}")
				return
			if T2 is None:
				messagebox.showerror("形式エラー", f"Input 2 のファイルを解析できませんでした:\n{mat2_path}")
				return

			# 2. Calculate Relative Transform Input 1 -> Input 2
			# Assuming T1 and T2 are poses relative to a common base (Cp)
			# P_Cp = T1 * P_Input1
			# P_Cp = T2 * P_Input2
			# T1 * P_Input1 = T2 * P_Input2
			# P_Input1 = T1^-1 * T2 * P_Input2
			# So Transform (1 -> 2) is T1^-1 * T2. This matrix maps points in Frame 2 to Frame 1.
			# This represents the pose of Frame 2 relative to Frame 1.
			T_12 = np.linalg.inv(T1) @ T2
			
			# 3. Show Result
			# We reuse _show_pos_checker_result but pass None for visualization args
			# And we need to adjust the title in the result text, but _show_pos_checker_result has hardcoded title.
			# So we might want to make _show_pos_checker_result more flexible or create a new method.
			# I'll create a dedicated display method to match the request exactly.
			
			self._show_pos_checker_compare_result(T_12, mat1_path, mat2_path)

		except Exception as e:
			messagebox.showerror("エラー", f"実行中にエラーが発生しました:\n{e}")
			traceback.print_exc()

	def _parse_matrix_file(self, path: str) -> Optional[np.ndarray]:
		"""テキストファイルから4x4行列を読み取る (PosChecker出力形式対応)"""
		try:
			matrix_lines = []
			with open(path, "r", encoding="utf-8") as f:
				lines = f.readlines()
			
			# シンプルなヒューリスティック: 4つの数値が並んでいる行を4行連続で探す
			for i in range(len(lines) - 3):
				candidates = lines[i:i+4]
				temp_matrix = []
				failed = False
				for line in candidates:
					parts = line.replace(",", " ").split() # カンマかスペース区切り
					# filter non-numeric parts if necessary, but assuming clean lines inside dashes
					# Try parsing 4 floats
					floats = []
					for p in parts:
						try:
							floats.append(float(p))
						except ValueError:
							pass
					
					if len(floats) == 4:
						temp_matrix.append(floats)
					else:
						failed = True
						break
				
				if not failed and len(temp_matrix) == 4:
					return np.array(temp_matrix)
			
			return None
		except Exception:
			return None

	def _show_pos_checker_compare_result(self, matrix, path1, path2):
		res_win = tk.Toplevel(self)
		res_win.title("PosChecker 比較結果")
		res_win.geometry("600x500")

		text_frame = ttk.Frame(res_win)
		text_frame.pack(fill="both", expand=True, padx=10, pady=10)

		name1 = os.path.basename(path1)
		name2 = os.path.basename(path2)

		result_text = "【変換行列比較 (Input 1 ➙ Input 2)】\n"
		result_text += f"Input 1: {path1}\n"
		result_text += f"Input 2: {path2}\n"
		result_text += "※ Input 1 のCf座標系から見た Input 2 のCf座標系の位置姿勢 (T_1_inv * T_2)\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{matrix[0][0]:12.6f} {matrix[0][1]:12.6f} {matrix[0][2]:12.6f} {matrix[0][3]:12.6f}\n"
		result_text += f"{matrix[1][0]:12.6f} {matrix[1][1]:12.6f} {matrix[1][2]:12.6f} {matrix[1][3]:12.6f}\n"
		result_text += f"{matrix[2][0]:12.6f} {matrix[2][1]:12.6f} {matrix[2][2]:12.6f} {matrix[2][3]:12.6f}\n"
		result_text += f"{matrix[3][0]:12.6f} {matrix[3][1]:12.6f} {matrix[3][2]:12.6f} {matrix[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"

		# Analysis
		trans = matrix[:3, 3]
		rot_mat = matrix[:3, :3]
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0

		dist_mm = np.linalg.norm(trans)

		result_text += "【相対変位解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Straight Dist(mm): {dist_mm:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"

		st = tk.Text(text_frame, wrap="none", font=(self.mono_font_family, 10))
		st.pack(side="left", fill="both", expand=True)
		sb = ttk.Scrollbar(text_frame, orient="vertical", command=st.yview)
		sb.pack(side="right", fill="y")
		st.config(yscrollcommand=sb.set)
		st.insert("1.0", result_text)
		st.config(state="disabled")

		btn_frame = ttk.Frame(res_win)
		btn_frame.pack(fill="x", padx=10, pady=10)

		def save_txt():
			path = filedialog.asksaveasfilename(
				title="結果を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if path:
				try:
					with open(path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"保存しました:\n{path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(btn_frame, text="テキストとして保存...", command=save_txt).pack(side="left", padx=5)
		ttk.Button(btn_frame, text="閉じる", command=res_win.destroy).pack(side="right", padx=5)


	def _create_coordinate_creater_tab(self) -> None:
		"""Coordinate Createrタブのコンテンツを作成"""
		container = self.coordinate_creater_tab
		container.columnconfigure(0, weight=1)
		
		# Variables
		self.cc_fe = tk.DoubleVar(value=0.0)
		self.cc_vv = tk.DoubleVar(value=0.0)
		self.cc_ie = tk.DoubleVar(value=0.0)
		self.cc_ml = tk.DoubleVar(value=0.0)
		self.cc_ap = tk.DoubleVar(value=0.0)
		self.cc_pd = tk.DoubleVar(value=0.0)
		
		# Container Frame
		input_frame = ttk.LabelFrame(container, text="パラメータ入力", style="Bold.TLabelframe")
		input_frame.pack(fill="x", padx=10, pady=10)
		input_frame.columnconfigure(1, weight=1)
		input_frame.columnconfigure(4, weight=1)
		
		# Rotations (FE, VV, IE)
		ttk.Label(input_frame, text="Rotation Parameters (deg)", font=(self.ui_font_family, 10, "bold")).grid(row=0, column=0, columnspan=2, pady=(5, 5), sticky="w", padx=5)
		
		ttk.Label(input_frame, text="FE (Flexion/Extension):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_fe).grid(row=1, column=1, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="VV (Varus/Valgus):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_vv).grid(row=2, column=1, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="IE (Internal/External):").grid(row=3, column=0, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_ie).grid(row=3, column=1, padx=5, pady=5, sticky="ew")
		
		ttk.Separator(input_frame, orient="vertical").grid(row=1, column=2, rowspan=3, sticky="ns", padx=10, pady=5)
		
		# Translations (ML, AP, PD)
		ttk.Label(input_frame, text="Translation Parameters (mm)", font=(self.ui_font_family, 10, "bold")).grid(row=0, column=3, columnspan=2, pady=(5, 5), sticky="w", padx=5)
		
		ttk.Label(input_frame, text="ML (Medial/Lateral):").grid(row=1, column=3, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_ml).grid(row=1, column=4, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="AP (Anterior/Posterior):").grid(row=2, column=3, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_ap).grid(row=2, column=4, padx=5, pady=5, sticky="ew")

		ttk.Label(input_frame, text="PD (Proximal/Distal):").grid(row=3, column=3, padx=5, pady=5, sticky="e")
		ttk.Entry(input_frame, textvariable=self.cc_pd).grid(row=3, column=4, padx=5, pady=5, sticky="ew")

		# Execute
		btn_frame = ttk.Frame(container)
		btn_frame.pack(fill="x", padx=10, pady=15)
		ttk.Button(btn_frame, text="変換行列を生成 & 表示", command=self.on_coordinate_creater_execute).pack(fill="x", ipady=5)

	def on_coordinate_creater_execute(self) -> None:
		"""Coordinate Creater実行"""
		try:
			fe = self.cc_fe.get()
			vv = self.cc_vv.get()
			ie = self.cc_ie.get()
			ml = self.cc_ml.get()
			ap = self.cc_ap.get()
			pd = self.cc_pd.get()
			
			# Build Matrix (Order should match _build_transform_matrix: rz(FE), rx(VV), ry(IE), ml, ap, pd)
			matrix = self._build_transform_matrix(fe, vv, ie, ml, ap, pd)
			
			self._show_coordinate_creater_result(matrix, fe, vv, ie, ml, ap, pd)
			
		except Exception as e:
			messagebox.showerror("エラー", f"値の取得または計算に失敗しました:\n{e}")

	def _show_coordinate_creater_result(self, matrix, fe, vv, ie, ml, ap, pd):
		res_win = tk.Toplevel(self)
		res_win.title("Coordinate Creater 結果")
		res_win.geometry("600x500")

		text_frame = ttk.Frame(res_win)
		text_frame.pack(fill="both", expand=True, padx=10, pady=10)

		# PosChecker形式に合わせる
		result_text = "【変換行列生成結果】\n"
		result_text += f"Input: FE={fe}, VV={vv}, IE={ie}, ML={ml}, AP={ap}, PD={pd}\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{matrix[0][0]:12.6f} {matrix[0][1]:12.6f} {matrix[0][2]:12.6f} {matrix[0][3]:12.6f}\n"
		result_text += f"{matrix[1][0]:12.6f} {matrix[1][1]:12.6f} {matrix[1][2]:12.6f} {matrix[1][3]:12.6f}\n"
		result_text += f"{matrix[2][0]:12.6f} {matrix[2][1]:12.6f} {matrix[2][2]:12.6f} {matrix[2][3]:12.6f}\n"
		result_text += f"{matrix[3][0]:12.6f} {matrix[3][1]:12.6f} {matrix[3][2]:12.6f} {matrix[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"
		
		# Analysis
		trans = matrix[:3, 3]
		rot_mat = matrix[:3, :3]
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0

		result_text += "【解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"

		st = tk.Text(text_frame, wrap="none", font=(self.mono_font_family, 10))
		st.pack(side="left", fill="both", expand=True)
		sb = ttk.Scrollbar(text_frame, orient="vertical", command=st.yview)
		sb.pack(side="right", fill="y")
		st.config(yscrollcommand=sb.set)
		st.insert("1.0", result_text)
		st.config(state="disabled")

		btn_frame = ttk.Frame(res_win)
		btn_frame.pack(fill="x", padx=10, pady=10)

		def save_txt():
			path = filedialog.asksaveasfilename(
				title="結果を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if path:
				try:
					with open(path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"保存しました:\n{path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(btn_frame, text="テキストとして保存...", command=save_txt).pack(side="left", padx=5)
		ttk.Button(btn_frame, text="閉じる", command=res_win.destroy).pack(side="right", padx=5)


	def _create_pp_merge_tab(self) -> None:
		"""pp mergeタブのコンテンツを作成"""
		container = self.pp_merge_tab
		container.columnconfigure(0, weight=1)

		# PPファイル入力エリア (最大6つ)
		list_frame = ttk.LabelFrame(container, text="統合するPPファイル (最大6つ)", style="Bold.TLabelframe")
		list_frame.pack(fill="both", expand=True, padx=5, pady=5)
		list_frame.columnconfigure(1, weight=1)

		self.pp_merge_vars = []
		
		# 6つの入力行を作成
		for i in range(6):
			var = tk.StringVar(value="")
			self.pp_merge_vars.append(var)
			
			row = i
			ttk.Label(list_frame, text=f"File {i+1}:").grid(row=row, column=0, sticky="w", padx=5, pady=5)
			
			entry = ttk.Entry(list_frame, textvariable=var)
			entry.grid(row=row, column=1, sticky="ew", padx=5, pady=5)
			
			# 参照ボタン
			ttk.Button(list_frame, text="参照...", width=8, 
					   command=lambda v=var: self._choose_pp_merge_file(v)).grid(row=row, column=2, sticky="e", padx=2, pady=5)
			
			# 削除（クリア）ボタン
			ttk.Button(list_frame, text="クリア", width=6,
					   command=lambda v=var: v.set("")).grid(row=row, column=3, sticky="e", padx=2, pady=5)

		# 実行ボタン
		ttk.Button(container, text="PPファイルを統合して保存...", command=self.on_pp_merge_execute).pack(fill="x", padx=5, pady=10)

	def _choose_pp_merge_file(self, target_var: tk.StringVar) -> None:
		path = filedialog.askopenfilename(
			title="PPファイルを選択",
			filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")]
		)
		if path:
			target_var.set(path)
			
	def on_pp_merge_execute(self) -> None:
		"""複数のPPファイルを統合して保存"""
		# 空でないパスを収集
		paths = [v.get().strip() for v in self.pp_merge_vars if v.get().strip()]
		
		if len(paths) < 2:
			messagebox.showwarning("入力不足", "少なくとも2つのPPファイルを選択してください。")
			return
			
		try:
			all_pts = []
			all_lbls = []
			
			print(f"[統合] {len(paths)}個のファイルを処理します...")
			
			for i, path in enumerate(paths):
				print(f"[統合] 読み込み中 ({i+1}/{len(paths)}): {path}")
				pts, lbls = self._parse_pp_file(path)
				all_pts.append(pts)
				all_lbls.extend(lbls)
				print(f" -> {len(pts)}点")
			
			# 統合
			if not all_pts:
				messagebox.showwarning("データなし", "有効な点データが見つかりませんでした。")
				return

			merged_pts = np.vstack(all_pts)
			merged_lbls = all_lbls
			print(f"[統合] 合計: {len(merged_pts)}点")
			
			# 保存先選択
			default_name = Path(paths[0]).stem + "_merged.pp"
			save_path = filedialog.asksaveasfilename(
				title="統合されたPPファイルを保存",
				defaultextension=".pp",
				filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")],
				initialfile=default_name
			)
			
			if not save_path:
				return
				
			# 保存
			save_child_pp_file(save_path, merged_pts, merged_lbls)
			messagebox.showinfo("完了", f"{len(paths)}個のPPファイルを統合しました:\n合計点数: {len(merged_pts)}\n\n保存先:\n{save_path}")
			
		except Exception as e:
			messagebox.showerror("エラー", f"PPファイルの統合に失敗しました:\n{e}")
			traceback.print_exc()

	# ================================================================
	# Cache Settings タブ
	# ================================================================

	def _create_cache_tab(self) -> None:
		"""Cache Settingsタブのコンテンツを作成"""
		container = self.cache_tab
		container.columnconfigure(0, weight=1)

		# --- NASパス設定 ---
		nas_frame = ttk.LabelFrame(container, text="NAS共有キャッシュ設定")
		nas_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
		nas_frame.columnconfigure(1, weight=1)

		ttk.Label(nas_frame, text="NASパス:").grid(row=0, column=0, sticky="w", padx=4, pady=4)
		ttk.Entry(nas_frame, textvariable=self.cache_nas_path).grid(row=0, column=1, sticky="ew", padx=4, pady=4)
		ttk.Button(nas_frame, text="参照...", command=self._choose_cache_nas_path).grid(row=0, column=2, padx=4, pady=4)

		ttk.Label(nas_frame, text="例: Z:/FRS_cache  や  //NAS_NAME/share/FRS_cache",
				  foreground="gray").grid(row=1, column=0, columnspan=3, sticky="w", padx=8)

		ttk.Checkbutton(nas_frame, text="キャッシュ有効", variable=self.cache_enabled).grid(
			row=2, column=0, columnspan=2, sticky="w", padx=4, pady=4)

		# --- 操作ボタン ---
		btn_frame = ttk.Frame(container)
		btn_frame.grid(row=1, column=0, sticky="ew", pady=(0, 8))

		ttk.Button(btn_frame, text="NASパスを適用（再初期化）",
				   command=self._on_cache_reinit).pack(side="left", padx=4)
		ttk.Button(btn_frame, text="統計情報を表示",
				   command=self._on_cache_show_stats).pack(side="left", padx=4)

		# --- 統計情報表示エリア ---
		self._cache_stats_text = tk.Text(container, height=12, width=60, state="disabled",
										 font=("Consolas", 9))
		self._cache_stats_text.grid(row=2, column=0, sticky="nsew", pady=(0, 8))
		container.rowconfigure(2, weight=1)

		# --- キャッシュ一覧 ---
		list_frame = ttk.LabelFrame(container, text="計算済みデータ一覧")
		list_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 8))
		container.rowconfigure(3, weight=1)

		list_btn_frame = ttk.Frame(list_frame)
		list_btn_frame.pack(fill=tk.X, padx=4, pady=4)
		ttk.Button(list_btn_frame, text="一覧を更新",
				   command=self._on_cache_list_refresh).pack(side="left", padx=4)
		ttk.Button(list_btn_frame, text="選択行を削除",
				   command=self._on_cache_list_delete_selected).pack(side="left", padx=4)

		cols = ("type", "xlsx", "prox_stl", "dist_stl", "frames", "date")
		col_labels = {"type": "種類", "xlsx": "運動データ(xlsx)", "prox_stl": "近位STL",
					  "dist_stl": "遠位STL", "frames": "フレーム数", "date": "計算日時"}
		self._cache_list_tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=8)
		for c in cols:
			self._cache_list_tree.heading(c, text=col_labels[c])
		self._cache_list_tree.column("type",     width=70,  anchor="center")
		self._cache_list_tree.column("xlsx",     width=220, anchor="w")
		self._cache_list_tree.column("prox_stl", width=140, anchor="w")
		self._cache_list_tree.column("dist_stl", width=140, anchor="w")
		self._cache_list_tree.column("frames",   width=80,  anchor="center")
		self._cache_list_tree.column("date",     width=130, anchor="center")

		sb = ttk.Scrollbar(list_frame, orient="vertical", command=self._cache_list_tree.yview)
		self._cache_list_tree.configure(yscrollcommand=sb.set)
		self._cache_list_tree.pack(side="left", fill="both", expand=True, padx=(4, 0), pady=4)
		sb.pack(side="left", fill="y", pady=4)

		# --- 危険ゾーン ---
		danger_frame = ttk.LabelFrame(container, text="キャッシュ削除")
		danger_frame.grid(row=4, column=0, sticky="ew", pady=(0, 8))

		ttk.Button(danger_frame, text="NASキャッシュを全削除",
				   command=self._on_cache_clear_all).pack(side="left", padx=4, pady=4)

	def _choose_cache_nas_path(self) -> None:
		d = filedialog.askdirectory(title="NASキャッシュディレクトリを選択")
		if d:
			self.cache_nas_path.set(d)

	def _on_cache_reinit(self) -> None:
		self._reinit_shared_cache()
		self._on_cache_show_stats()
		messagebox.showinfo("キャッシュ", "共有キャッシュを再初期化しました。")

	def _on_cache_list_refresh(self) -> None:
		"""キャッシュ一覧を更新する"""
		tree = self._cache_list_tree
		tree.delete(*tree.get_children())

		# NASベースパス配下の全サブフォルダをスキャン（フォルダ名の変遷に対応）
		nas_base = self.cache_nas_path.get().strip()
		all_meta = []

		if nas_base:
			from pathlib import Path
			base = Path(nas_base)
			print(f"[CacheList] NASベースパス: {base}")
			# 候補フォルダを全て試す（ドットあり・なし両方）
			candidate_dirs = [
				(base / "fem_cache",     "fem"),
				(base / ".fem_cache",    "fem"),
				(base / "overlap_cache", "overlap"),
				(base / ".overlap_cache","overlap"),
			]
			visited_keys = set()
			for scan_dir, cache_type in candidate_dirs:
				if not scan_dir.exists():
					print(f"[CacheList]   スキップ（存在しない）: {scan_dir}")
					continue
				print(f"[CacheList]   スキャン中: {scan_dir}")
				# メタデータJSONを読み込み
				import json as _json
				for meta_file in scan_dir.glob("_meta_*.json"):
					try:
						with open(meta_file, 'r', encoding='utf-8') as f:
							meta = _json.load(f)
						mk = meta_file.stem[6:]
						if mk in visited_keys:
							continue
						meta['_meta_key'] = mk
						meta['_scan_dir'] = str(scan_dir)
						all_meta.append(meta)
						visited_keys.add(mk)
					except Exception:
						pass
				# メタデータのないPKLファイルをスキャン
				for pkl_file in scan_dir.glob("*.pkl"):
					if pkl_file.name.startswith('_'):
						continue
					stem = pkl_file.stem
					if stem.startswith('fem_'):
						key = stem[4:]
					elif stem.startswith('overlap_'):
						key = stem[8:]
					else:
						key = stem
					if key in visited_keys:
						continue
					try:
						stat = pkl_file.stat()
						mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
						size_mb = round(stat.st_size / (1024**2), 1)
						all_meta.append({
							'type': cache_type,
							'xlsx': '（情報なし）',
							'prox_stl': '（情報なし）',
							'dist_stl': '（情報なし）',
							'frames': '-',
							'date': mtime.strftime('%Y-%m-%d %H:%M'),
							'size_mb': size_mb,
							'_meta_key': key,
							'_no_meta': True,
							'_scan_dir': str(scan_dir),
						})
						visited_keys.add(key)
					except Exception:
						pass
			print(f"[CacheList] 合計 {len(all_meta)} 件")
		else:
			print("[CacheList] NASパス未設定")

		# 日時でソート（新しい順）
		all_meta.sort(key=lambda m: m.get('date', ''), reverse=True)
		for meta in all_meta:
			cache_type = meta.get('type', 'unknown')
			type_label = {"fem": "FEM", "overlap": "Overlap"}.get(cache_type, cache_type)
			if meta.get('_no_meta'):
				type_label += "※"  # メタデータなし（旧キャッシュ）
			if cache_type == 'fem':
				frames_str = f"{meta.get('frames_cached','?')}/{meta.get('frames','?')}"
			else:
				frames_str = str(meta.get('frames', '-'))
			tree.insert("", "end",
				values=(
					type_label,
					meta.get('xlsx', ''),
					meta.get('prox_stl', ''),
					meta.get('dist_stl', ''),
					frames_str,
					meta.get('date', ''),
				),
				tags=(cache_type, meta.get('_meta_key', ''))
			)
		if not all_meta:
			tree.insert("", "end", values=("", "キャッシュなし（NAS未接続またはデータなし）", "", "", "", ""))

	def _on_cache_list_delete_selected(self) -> None:
		"""選択行のキャッシュを削除する"""
		tree = self._cache_list_tree
		selected = tree.selection()
		if not selected:
			messagebox.showwarning("選択なし", "削除する行を選択してください。")
			return
		if not messagebox.askyesno("確認", f"{len(selected)}件のキャッシュを削除しますか？"):
			return
		for item in selected:
			tags = tree.item(item, "tags")
			if len(tags) >= 2:
				cache_type, meta_key = tags[0], tags[1]
				cache = self._shared_fem_cache if cache_type == 'fem' else self._shared_overlap_cache
				if cache and meta_key:
					cache.delete_dataset(meta_key)
		self._on_cache_list_refresh()
		messagebox.showinfo("完了", "選択したキャッシュを削除しました。")

	def _on_cache_show_stats(self) -> None:
		lines = []
		for name, cache in [("FEM", self._shared_fem_cache), ("Overlap", self._shared_overlap_cache)]:
			if cache is None:
				lines.append(f"[{name}] 未初期化")
				continue
			s = cache.stats()
			lines.append(f"[{name}]")
			lines.append(f"  NAS接続:  {'OK' if s['nas_available'] else 'なし'}")
			if s['nas_available']:
				lines.append(f"  NAS:      {s['nas_count']}件 ({s['nas_size_mb']:.1f} MB)")
			lines.append(f"  メモリ:   {s['memory_count']}件（セッション内）")
			lines.append("")
		self._cache_stats_text.config(state="normal")
		self._cache_stats_text.delete("1.0", "end")
		self._cache_stats_text.insert("1.0", "\n".join(lines))
		self._cache_stats_text.config(state="disabled")

	def _on_cache_clear_all(self) -> None:
		if not messagebox.askyesno("確認", "NASキャッシュを全削除しますか？\nこの操作は取り消せません。"):
			return
		if self._shared_fem_cache:
			self._shared_fem_cache.clear_all()
		if self._shared_overlap_cache:
			self._shared_overlap_cache.clear_all()
		self._on_cache_show_stats()
		messagebox.showinfo("完了", "NASキャッシュを削除しました。")

	def choose_randomizer_input(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルファイルを選択",
			filetypes=[("3D Model", "*.stl *.obj"), ("All files", "*.*")]
		)
		if path:
			self.randomizer_input_path.set(path)

	def on_randomizer_execute(self) -> None:
		"""ランダム変換を実行して保存"""
		input_path = self.randomizer_input_path.get().strip()
		if not input_path:
			messagebox.showwarning("警告", "モデルファイルを選択してください。")
			return
		
		if not os.path.exists(input_path):
			messagebox.showerror("エラー", f"ファイルが見つかりません:\n{input_path}")
			return

		# 保存先を選択
		save_path = filedialog.asksaveasfilename(
			title="変換後のモデルを保存",
			defaultextension=".stl",
			filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj"), ("All files", "*.*")],
			initialfile=Path(input_path).stem + "_random" + Path(input_path).suffix
		)
		
		if not save_path:
			return

		try:
			# パラメータ取得
			try:
				trans_range = self.random_trans_range.get()
				rot_range = self.random_rot_range.get()
			except tk.TclError:
				messagebox.showerror("エラー", "数値パラメータが不正です。")
				return

			# モデル読み込み
			mesh = pv.read(input_path)
			
			# ランダム変換生成
			# Translation
			t = np.random.uniform(-trans_range, trans_range, 3)
			
			# Rotation (Euler angles -> Rotation Matrix)
			# deg -> rad
			rx = np.deg2rad(np.random.uniform(-rot_range, rot_range))
			ry = np.deg2rad(np.random.uniform(-rot_range, rot_range))
			rz = np.deg2rad(np.random.uniform(-rot_range, rot_range))
			
			Rx = np.array([
				[1, 0, 0, 0],
				[0, np.cos(rx), -np.sin(rx), 0],
				[0, np.sin(rx), np.cos(rx), 0],
				[0, 0, 0, 1]
			])
			Ry = np.array([
				[np.cos(ry), 0, np.sin(ry), 0],
				[0, 1, 0, 0],
				[-np.sin(ry), 0, np.cos(ry), 0],
				[0, 0, 0, 1]
			])
			Rz = np.array([
				[np.cos(rz), -np.sin(rz), 0, 0],
				[np.sin(rz), np.cos(rz), 0, 0],
				[0, 0, 1, 0],
				[0, 0, 0, 1]
			])
			
			# Combined Rotation
			R = Rz @ Ry @ Rx
			
			# Translation Matrix
			T = np.eye(4)
			T[:3, 3] = t
			
			# Final Transform
			M = T @ R
			
			# Apply transform
			mesh.transform(M)
			
			# Save
			mesh.save(save_path)
			
			msg = (
				f"完了しました。\n\n"
				f"Translation (mm):\n  X={t[0]:.2f}, Y={t[1]:.2f}, Z={t[2]:.2f}\n"
				f"Rotation (deg):\n  X={np.rad2deg(rx):.1f}, Y={np.rad2deg(ry):.1f}, Z={np.rad2deg(rz):.1f}\n\n"
				f"Saved to:\n{save_path}"
			)
			messagebox.showinfo("成功", msg)
			
		except Exception as e:
			messagebox.showerror("エラー", f"処理に失敗しました:\n{e}\n{traceback.format_exc()}")

	# ----- UI Helpers -----
	def _setup_fonts(self) -> None:
		"""環境に合わせたUI/等幅フォントを選択し、Tkのデフォルトに適用する"""
		try:
			families = set(tkfont.families())
		except Exception:
			families = set()

		def pick(candidates, fallback=None):
			for fam in candidates:
				if fam in families:
					return fam
			return fallback or tkfont.nametofont("TkDefaultFont").cget("family")

		# 日本語表示を考慮した候補（プラットフォーム別に優先順位を設定）
		if IS_MACOS:
			ui_candidates = [
				"Hiragino Sans", "Hiragino Kaku Gothic Pro", "Hiragino Kaku Gothic ProN",
				".AppleSystemUIFont", "Helvetica Neue", "Lucida Grande", "Arial",
			]
			mono_candidates = [
				"Menlo", "Monaco", "SF Mono", "Osaka-Mono", "Courier New",
			]
		elif IS_WINDOWS:
			ui_candidates = [
				"Yu Gothic UI", "Meiryo UI", "Segoe UI", "メイリオ",
				"Yu Gothic", "MS UI Gothic", "MS Gothic", "Arial",
			]
			mono_candidates = [
				"Consolas", "MS Gothic", "MeiryoKe_Console", "Courier New",
			]
		else:  # Linux
			ui_candidates = [
				"Noto Sans CJK JP", "Noto Sans JP", "Takao Gothic",
				"IPAGothic", "VL Gothic", "DejaVu Sans", "Arial",
			]
			mono_candidates = [
				"Noto Sans Mono CJK JP", "DejaVu Sans Mono", "Monospace", "Courier New",
			]

		self.ui_font_family = pick(ui_candidates)
		self.mono_font_family = pick(mono_candidates, fallback=self.ui_font_family)

		# Tkのデフォルトフォント更新
		try:
			default_font = tkfont.nametofont("TkDefaultFont")
			default_font.configure(family=self.ui_font_family, size=11)
			for name in ("TkTextFont", "TkHeadingFont", "TkMenuFont"):
				try:
					f = tkfont.nametofont(name)
					f.configure(family=self.ui_font_family, size=11)
				except Exception:
					pass
			try:
				fixed = tkfont.nametofont("TkFixedFont")
				fixed.configure(family=self.mono_font_family, size=11)
			except Exception:
				pass
		except Exception:
			pass

	# ----- ドラッグ&ドロップ ヘルパー -----
	@staticmethod
	def _parse_dnd_path(data: str) -> str:
		"""ドロップされたデータからファイルパスを抽出する。
		macOS/Windows/Linuxのいずれの形式にも対応。
		"""
		raw = data.strip()
		# 複数ファイルの場合は最初の1つだけ使用
		# macOSでは中括弧で囲まれることがある: {/path/to/file}
		if raw.startswith("{") and "}" in raw:
			raw = raw[1:raw.index("}")]
		# 改行区切りの場合
		if "\n" in raw:
			raw = raw.split("\n")[0].strip()
		# スペースを含むパスが複数トークンに分割される場合への対処
		# file:// URI形式の場合
		if raw.startswith("file://"):
			raw = urllib.parse.unquote(raw[7:])
			# macOSでは file:///path の形式
			if raw.startswith("//"):
				raw = raw[2:]
		return raw

	def _register_dnd(self, widget, textvariable: tk.StringVar) -> None:
		"""ウィジェットにドラッグ&ドロップを登録する。

		macOS/Windows/Linux 差異への対応:
		  - <<Drop>> と <<Drop:DND_Files>> の両方をバインド（Aqua では後者しか発火しない事例がある）
		  - 初回登録時のみ失敗をコンソール表示（以降は静かにスキップ）
		"""
		if not _HAS_DND:
			return
		try:
			widget.drop_target_register(DND_FILES)
			def on_drop(event):
				path = self._parse_dnd_path(event.data)
				if path and os.path.exists(path):
					textvariable.set(path)
			widget.dnd_bind("<<Drop>>", on_drop)
			widget.dnd_bind("<<Drop:DND_Files>>", on_drop)
		except Exception as e:
			if not getattr(self, "_dnd_warn_shown", False):
				print(f"[DnD] 登録に失敗しました（以降のDnDは無効化されます）: {e}")
				self._dnd_warn_shown = True

	# ----- PyVista プロッタ 協調表示ヘルパー -----
	def _show_plotter_coop(self, plotter) -> None:
		"""pv.Plotter を Tk mainloop と協調しながら非同期表示する。

		macOS では Tk と VTK(Cocoa) のイベントループが競合し、
		Tk コールバック内から plotter.show() を呼ぶとハングする。
		本ヘルパーは既存 on_visualize_all / on_animate と同一パターン
		（interactive_update=True + Tk after() による update() ポンプ）を
		用いてハングを回避する。Windows/Linux でも副作用は無い
		（UIの応答性が上がる副次効果のみ）。
		"""
		# 非同期表示モードに切替（古い pyvista では auto_close のみ）
		try:
			plotter.show(auto_close=False, interactive_update=True)
		except TypeError:
			plotter.show(auto_close=False)

		# プロッタ本体を GC から守る（closureが持つが念のため保険）
		if not hasattr(self, "_active_plotters"):
			self._active_plotters = []
		self._active_plotters.append(plotter)

		alive = [True]

		def _on_exit(*_a):
			alive[0] = False

		try:
			plotter.iren.add_observer('ExitEvent', _on_exit)
		except Exception:
			pass

		def _tick():
			if not alive[0]:
				try:
					self._active_plotters.remove(plotter)
				except (ValueError, AttributeError):
					pass
				return
			try:
				plotter.update()
			except Exception:
				alive[0] = False
				try:
					self._active_plotters.remove(plotter)
				except (ValueError, AttributeError):
					pass
				return
			self.after(30, _tick)

		self.after(30, _tick)

	def _add_file_row(self, parent: ttk.Frame, row: int, label_text: str, textvariable: tk.StringVar, command):
		"""Add a file selection row with label, readonly entry (showing filename only), and browse button.
		ドラッグ&ドロップにも対応（tkinterdnd2が利用可能な場合）。
		ラベルウィジェットを返す（後から text を差し替えられるように）。
		"""
		row_label = ttk.Label(parent, text=label_text)
		row_label.grid(row=row, column=0, sticky="w", padx=12, pady=6)

		# Create a frame to hold the entry with custom display
		entry_frame = ttk.Frame(parent)
		entry_frame.grid(row=row, column=1, sticky="ew", padx=12, pady=6)

		# Create a StringVar that displays only the filename
		display_var = tk.StringVar(value="")

		# Update display when the actual variable changes
		def update_display(*args):
			full_path = textvariable.get()
			if full_path:
				# Show only filename
				display_var.set(Path(full_path).name)
			else:
				display_var.set("")

		textvariable.trace_add("write", update_display)
		update_display()  # Initialize

		entry = ttk.Entry(entry_frame, textvariable=display_var, state="readonly")
		entry.pack(fill="x", expand=True)

		# ドラッグ&ドロップ対応
		# macOS の readonly Entry はドロップ受け取りが不安定なため、
		# 周辺の Label / Frame でも受け付けて取りこぼしを防ぐ。
		self._register_dnd(entry, textvariable)
		self._register_dnd(entry_frame, textvariable)
		self._register_dnd(row_label, textvariable)

		# Add tooltip showing full path on hover
		def show_tooltip(event):
			full_path = textvariable.get()
			if full_path:
				# Create tooltip
				tooltip = tk.Toplevel()
				tooltip.wm_overrideredirect(True)
				tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
				label = ttk.Label(tooltip, text=full_path, background="lightyellow", relief="solid", borderwidth=1, padding=4)
				label.pack()
				entry.tooltip = tooltip

		def hide_tooltip(event):
			if hasattr(entry, 'tooltip'):
				entry.tooltip.destroy()
				del entry.tooltip

		entry.bind("<Enter>", show_tooltip)
		entry.bind("<Leave>", hide_tooltip)

		ttk.Button(parent, text="参照...", command=command).grid(row=row, column=2, sticky="e", padx=12, pady=6)
		return row_label

	# ----- File pickers -----
	def choose_prox_model(self) -> None:
		path = filedialog.askopenfilename(
			title="近位モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.prox_model_path.set(path)

	def choose_prox_pp_abcd(self) -> None:
		path = filedialog.askopenfilename(
			title="近位特徴点 ABCD (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.prox_pp_abcd_path.set(path)

	def choose_prox_pp_olmn(self) -> None:
		path = filedialog.askopenfilename(
			title="近位特徴点 OLMN (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.prox_pp_olmn_path.set(path)

	def choose_dist_model(self) -> None:
		path = filedialog.askopenfilename(
			title="遠位モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.dist_model_path.set(path)

	def choose_dist_pp_abc(self) -> None:
		path = filedialog.askopenfilename(
			title="遠位特徴点 ABC (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.dist_pp_abc_path.set(path)

	def choose_dist_pp_olmn(self) -> None:
		path = filedialog.askopenfilename(
			title="遠位特徴点 OLMN (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.dist_pp_olmn_path.set(path)

	def choose_prox_cartilage_model(self) -> None:
		"""近位軟骨モデルを選択"""
		path = filedialog.askopenfilename(
			title="近位軟骨モデルを選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			self.prox_cartilage_model_path.set(path)

	def choose_dist_cartilage_model(self) -> None:
		"""遠位軟骨モデルを選択"""
		path = filedialog.askopenfilename(
			title="遠位軟骨モデルを選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			self.dist_cartilage_model_path.set(path)

	# ----- File pickers (Fitting) -----
	def choose_fitting_parent_model(self) -> None:
		path = filedialog.askopenfilename(
			title="親モデルファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_parent_model_path.set(path)

	def choose_fitting_parent_region(self) -> None:
		path = filedialog.askopenfilename(
			title="親任意領域ファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_parent_region_path.set(path)

	def choose_fitting_child_model(self) -> None:
		path = filedialog.askopenfilename(
			title="子モデルファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_model_path.set(path)

	def choose_fitting_child_region(self) -> None:
		path = filedialog.askopenfilename(
			title="子任意領域ファイルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_region_path.set(path)

	def choose_fitting_child_pp(self) -> None:
		path = filedialog.askopenfilename(
			title="子特徴点 (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_pp_path.set(path)

	def choose_fitting_child_heatmap_model(self) -> None:
		path = filedialog.askopenfilename(
			title="子ヒートマップ表示用モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.fitting_child_heatmap_model_path.set(path)

	def choose_transform_group(self) -> None:
		path = filedialog.askopenfilename(
			title="同次変換行列群ファイルを選択",
			filetypes=[
				("すべてのファイル", "*.*"),
				("Excelファイル", "*.xlsx"),
				("KKRファイル", "*.KKR *.kkr"),
			],
		)
		if path:
			# 拡張子を確認
			ext = Path(path).suffix.lower()
			if ext not in ['.xlsx', '.kkr']:
				messagebox.showwarning("ファイル形式エラー", "選択されたファイル形式はサポートされていません。\n.xlsx または .kkr ファイルを選択してください。")
				return
			
			self.transform_group_path.set(path)

	# ----- File pickers (ORG) -----
	def choose_org_model(self) -> None:
		path = filedialog.askopenfilename(
			title="RobotScan モデルを選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.org_model_path.set(path)

	def choose_org_pp(self) -> None:
		path = filedialog.askopenfilename(
			title="RobotScan 座標ポイント (PP) を選択",
			filetypes=[
				("特徴点ファイル", "*.pp"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.org_pp_path.set(path)

	# ----- File pickers (A-B) -----
	def choose_a_b_model_a(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルA (STL) を選択",
			filetypes=[
				("STL", "*.stl"),
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.a_b_model_a_path.set(path)

	def choose_a_b_model_b(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルB (STL) を選択",
			filetypes=[
				("STL", "*.stl"),
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.a_b_model_b_path.set(path)

	# ----- File pickers (HeatMap) -----
	def choose_heatmap_model_a(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルA (参照) を選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.heatmap_model_a_path.set(path)

	def choose_heatmap_model_b(self) -> None:
		path = filedialog.askopenfilename(
			title="モデルB (距離計算対象) を選択",
			filetypes=[
				("3Dモデル", "*.obj *.stl"),
				("OBJ", "*.obj"),
				("STL", "*.stl"),
				("すべてのファイル", "*.*"),
			],
		)
		if path:
			self.heatmap_model_b_path.set(path)

	# ----- Cartilage Separation: Generic File Picker -----
	def _choose_cs_file(self, side: str, which: str) -> None:
		"""軟骨分離用ファイル選択（side='prox'|'dist', which='m1w'|'m1r'|'m2w'|'m2r'）"""
		titles = {
			"m1w": "骨＋軟骨 全体モデル", "m1r": "骨＋軟骨 任意領域",
			"m2w": "骨のみ 全体モデル", "m2r": "骨のみ 任意領域",
		}
		var_map = {
			("prox", "m1w"): self.cs_prox_model1_whole_path, ("prox", "m1r"): self.cs_prox_model1_region_path,
			("prox", "m2w"): self.cs_prox_model2_whole_path, ("prox", "m2r"): self.cs_prox_model2_region_path,
			("dist", "m1w"): self.cs_dist_model1_whole_path, ("dist", "m1r"): self.cs_dist_model1_region_path,
			("dist", "m2w"): self.cs_dist_model2_whole_path, ("dist", "m2r"): self.cs_dist_model2_region_path,
		}
		side_label = "近位" if side == "prox" else "遠位"
		path = filedialog.askopenfilename(
			title=f"{side_label}: {titles[which]}を選択",
			filetypes=[("3Dモデル", "*.obj *.stl"), ("OBJ", "*.obj"), ("STL", "*.stl"), ("すべてのファイル", "*.*")],
		)
		if path:
			var_map[(side, which)].set(path)

	def _update_cs_side_buttons(self, side: str) -> None:
		"""近位/遠位ボタンの活性制御"""
		if side == "prox":
			m1w = bool(self.cs_prox_model1_whole_path.get().strip())
			m1r = bool(self.cs_prox_model1_region_path.get().strip())
			m2w = bool(self.cs_prox_model2_whole_path.get().strip())
			m2r = bool(self.cs_prox_model2_region_path.get().strip())
			try: self.cs_prox_viz_btn.config(state="normal" if m1w else "disabled")
			except: pass
			try: self.cs_prox_exec_btn.config(state="normal" if (m1w and m1r and m2w and m2r) else "disabled")
			except: pass
		else:
			m1w = bool(self.cs_dist_model1_whole_path.get().strip())
			m1r = bool(self.cs_dist_model1_region_path.get().strip())
			m2w = bool(self.cs_dist_model2_whole_path.get().strip())
			m2r = bool(self.cs_dist_model2_region_path.get().strip())
			try: self.cs_dist_viz_btn.config(state="normal" if m1w else "disabled")
			except: pass
			try: self.cs_dist_exec_btn.config(state="normal" if (m1w and m1r and m2w and m2r) else "disabled")
			except: pass

	def _on_cs_visualize_side(self, side: str) -> None:
		"""近位/遠位モデルの確認可視化"""
		if side == "prox":
			w_path = self.cs_prox_model1_whole_path.get().strip()
			r_path = self.cs_prox_model1_region_path.get().strip()
			w2_path = self.cs_prox_model2_whole_path.get().strip()
		else:
			w_path = self.cs_dist_model1_whole_path.get().strip()
			r_path = self.cs_dist_model1_region_path.get().strip()
			w2_path = self.cs_dist_model2_whole_path.get().strip()
		side_label = self._joint_label("cs_prox_tab" if side == "prox" else "cs_dist_tab")
		if not w_path:
			messagebox.showwarning("入力不足", "モデル1の全体モデルを選択してください。")
			return
		try:
			plotter = pv.Plotter(title=f"{side_label} モデル確認")
			plotter.set_background('white')
			plotter.add_mesh(pv.read(w_path), color='lightblue', opacity=0.5, show_edges=False, label='骨+軟骨 全体')
			if r_path:
				plotter.add_mesh(pv.read(r_path), color='red', opacity=0.8, show_edges=False, label='任意領域')
			if w2_path:
				plotter.add_mesh(pv.read(w2_path), color='wheat', opacity=0.3, show_edges=False, label='骨のみ 全体')
			plotter.add_axes(); plotter.add_legend(); plotter.show()
			self._restore_window_geometry()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{str(e)}")

	def _on_cs_execute_side(self, side: str) -> None:
		"""近位/遠位の軟骨分離を実行（ブーリアン差分）"""
		if side == "prox":
			m1w_path = self.cs_prox_model1_whole_path.get().strip()
			m1r_path = self.cs_prox_model1_region_path.get().strip()
			m2w_path = self.cs_prox_model2_whole_path.get().strip()
			m2r_path = self.cs_prox_model2_region_path.get().strip()
		else:
			m1w_path = self.cs_dist_model1_whole_path.get().strip()
			m1r_path = self.cs_dist_model1_region_path.get().strip()
			m2w_path = self.cs_dist_model2_whole_path.get().strip()
			m2r_path = self.cs_dist_model2_region_path.get().strip()

		side_label = self._joint_label("cs_prox_tab" if side == "prox" else "cs_dist_tab")
		# 既存のon_cs_executeを呼び出す（パスを一時的にprox変数にセット）
		# on_cs_executeはcs_prox_*を参照するので、dist実行時はswap
		if side == "dist":
			# 一時的にprox変数にdist値をセット
			save = [self.cs_prox_model1_whole_path.get(), self.cs_prox_model1_region_path.get(),
					self.cs_prox_model2_whole_path.get(), self.cs_prox_model2_region_path.get()]
			self.cs_prox_model1_whole_path.set(m1w_path)
			self.cs_prox_model1_region_path.set(m1r_path)
			self.cs_prox_model2_whole_path.set(m2w_path)
			self.cs_prox_model2_region_path.set(m2r_path)

		# 実行中の側を記録（on_cs_execute内でSimulator反映に使用）
		self._cs_executing_side = side
		self.on_cs_execute()

		if side == "dist":
			# 元に戻す
			self.cs_prox_model1_whole_path.set(save[0])
			self.cs_prox_model1_region_path.set(save[1])
			self.cs_prox_model2_whole_path.set(save[2])
			self.cs_prox_model2_region_path.set(save[3])

	def _choose_cs_bone_color(self) -> None:
		"""骨の色を選択"""
		color = colorchooser.askcolor(initialcolor=self.cs_bone_color, title="骨の表示色を選択")
		if color[1]:
			self.cs_bone_color = color[1]

	def _choose_cs_cartilage_color(self) -> None:
		"""軟骨の色を選択"""
		color = colorchooser.askcolor(initialcolor=self.cs_cartilage_color, title="軟骨の表示色を選択")
		if color[1]:
			self.cs_cartilage_color = color[1]

	# --- フィッティングパラメータ デフォルト保存/復元 ---

	_CS_FITTING_DEFAULTS_BUILTIN = {
		"ransac_distance_threshold": 1.0,
		"ransac_max_iterations": 1000,
		"ransac_confidence": 0.99,
		"icp_threshold": 5.0,
		"icp_max_iterations": 2000,
		"sample_points": 10000,
		"distance_threshold": 0.5,
	}

	def _cs_fitting_defaults_path(self) -> Path:
		"""ユーザーデフォルト保存先"""
		return Path(__file__).parent / ".frs_fitting_defaults.json"

	def _load_cs_fitting_defaults(self) -> dict:
		"""ユーザーデフォルトを読み込み。なければビルトインを返す"""
		p = self._cs_fitting_defaults_path()
		if p.exists():
			try:
				with p.open("r", encoding="utf-8") as f:
					data = json.load(f)
				# ビルトインのキーが全て揃っているか補完
				merged = dict(self._CS_FITTING_DEFAULTS_BUILTIN)
				merged.update(data)
				return merged
			except Exception:
				pass
		return dict(self._CS_FITTING_DEFAULTS_BUILTIN)

	def _save_cs_fitting_defaults(self) -> None:
		"""現在のフィッティングパラメータをユーザーデフォルトとして保存"""
		data = {
			"ransac_distance_threshold": self.cs_ransac_distance_threshold.get(),
			"ransac_max_iterations": self.cs_ransac_max_iterations.get(),
			"ransac_confidence": self.cs_ransac_confidence.get(),
			"icp_threshold": self.cs_icp_threshold.get(),
			"icp_max_iterations": self.cs_icp_max_iterations.get(),
			"sample_points": self.cs_sample_points.get(),
			"distance_threshold": self.cs_distance_threshold.get(),
		}
		try:
			p = self._cs_fitting_defaults_path()
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
			messagebox.showinfo("保存完了",
				"現在のフィッティングパラメータをデフォルトとして保存しました。\n\n"
				f"RANSAC 距離閾値: {data['ransac_distance_threshold']}\n"
				f"RANSAC 最大反復: {data['ransac_max_iterations']}\n"
				f"RANSAC 信頼度: {data['ransac_confidence']}\n"
				f"ICP 閾値: {data['icp_threshold']}\n"
				f"ICP 最大反復: {data['icp_max_iterations']}\n"
				f"サンプル点数: {data['sample_points']}\n"
				f"距離閾値: {data['distance_threshold']}")
		except Exception as e:
			messagebox.showerror("保存エラー", f"デフォルト保存に失敗しました:\n{e}")

	def reset_cs_fitting_parameters(self) -> None:
		"""フィッティングパラメータをユーザーデフォルト（なければビルトイン）に戻す"""
		d = self._load_cs_fitting_defaults()
		self.cs_ransac_distance_threshold.set(d["ransac_distance_threshold"])
		self.cs_ransac_max_iterations.set(int(d["ransac_max_iterations"]))
		self.cs_ransac_confidence.set(d["ransac_confidence"])
		self.cs_icp_threshold.set(d["icp_threshold"])
		self.cs_icp_max_iterations.set(int(d["icp_max_iterations"]))
		self.cs_sample_points.set(int(d["sample_points"]))
		self.cs_distance_threshold.set(d["distance_threshold"])

	def on_cs_execute(self) -> None:
		"""
		【軟骨分離】メイン実行メソッド
		
		■ 処理フロー:
		  1. 任意領域同士を RANSAC + ICP でフィッティング → 変換行列T取得
		  2. T をモデル2（骨のみ）全体に適用 → モデル1と同じ座標系に
		  3. ブーリアン差分: モデル1（骨+軟骨） − モデル2（骨のみ） = 軟骨（＋関節唇）
		  4. 結果を可視化＋STL保存オプション
		
		■ 前提条件:
		  すべてのモデルは体積メッシュ（内部にもメッシュがある）であること。
		  FEMに使用するため、表面だけでなく内部も埋まっている必要がある。
		"""
		# 入力チェック
		if not all([
			self.cs_prox_model1_whole_path.get().strip(),
			self.cs_prox_model1_region_path.get().strip(),
			self.cs_prox_model2_whole_path.get().strip(),
			self.cs_prox_model2_region_path.get().strip(),
		]):
			messagebox.showerror("エラー", "4つのモデルファイルをすべて選択してください。")
			return
		
		try:
			import copy
			import trimesh  # type: ignore
			
			def load_mesh_o3d(file_path):
				"""メッシュをOpen3Dで読み込む"""
				try:
					print(f"ファイル読み込み中: {Path(file_path).name}")
					tm = trimesh.load(file_path, force='mesh')
					vertices = np.array(tm.vertices)
					triangles = np.array(tm.faces)
					mesh = o3d.geometry.TriangleMesh()
					mesh.vertices = o3d.utility.Vector3dVector(vertices)
					mesh.triangles = o3d.utility.Vector3iVector(triangles)
					mesh.compute_vertex_normals()
					print(f"  ✓ 成功: 頂点数={len(vertices)}, 三角形数={len(triangles)}")
					return mesh
				except Exception as e:
					print(f"  ✗ エラー: {e}")
					messagebox.showerror("エラー", f"ファイル読み込み失敗:\n{Path(file_path).name}\n\n{str(e)}")
					return None
			
			print("\n" + "="*60)
			print("=== 軟骨分離処理開始 ===")
			print("="*60)
			print(f"モデル1全体: {Path(self.cs_prox_model1_whole_path.get()).name}")
			print(f"モデル1任意領域: {Path(self.cs_prox_model1_region_path.get()).name}")
			print(f"モデル2全体: {Path(self.cs_prox_model2_whole_path.get()).name}")
			print(f"モデル2任意領域: {Path(self.cs_prox_model2_region_path.get()).name}")
			
			# --- Phase 1: 任意領域の RANSAC + ICP フィッティング ---
			print("\n--- Phase 1: 任意領域のフィッティング ---")
			
			region1_mesh = load_mesh_o3d(self.cs_prox_model1_region_path.get())
			region2_mesh = load_mesh_o3d(self.cs_prox_model2_region_path.get())
			
			if not all([region1_mesh, region2_mesh]):
				return
			
			# 色設定して初期状態を表示
			region1_mesh.paint_uniform_color([1.0, 0.0, 0.0])  # 赤 = モデル1任意領域
			region2_mesh.paint_uniform_color([0.0, 0.0, 1.0])  # 青 = モデル2任意領域
			
			print("\n[初期状態] 任意領域を表示中...")
			o3d.visualization.draw_geometries(
				[copy.deepcopy(region1_mesh), copy.deepcopy(region2_mesh)],
				window_name="初期状態: モデル1任意領域(赤) + モデル2任意領域(青)")
			self._restore_window_geometry()
			
			# 主軸による事前アライメント
			print("\n=== 主軸による事前アライメント ===")
			
			def get_principal_axis(mesh):
				pcd = mesh.sample_points_uniformly(number_of_points=20000)
				pts = np.asarray(pcd.points)
				center = np.mean(pts, axis=0)
				cov = np.cov((pts - center).T)
				eigvals, eigvecs = np.linalg.eigh(cov)
				return center, eigvecs[:, np.argmax(eigvals)]
			
			def align_axis(mesh, src_axis, tgt_axis):
				v = np.cross(src_axis, tgt_axis)
				c = np.dot(src_axis, tgt_axis)
				if np.linalg.norm(v) < 1e-8:
					R = np.eye(3)
				else:
					vx = np.array([[0, -v[2], v[1]], [v[2], 0, -v[0]], [-v[1], v[0], 0]])
					R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
				mesh_rot = copy.deepcopy(mesh)
				mesh_rot.rotate(R, center=(0, 0, 0))
				return mesh_rot, R
			
			c_r1, axis_r1 = get_principal_axis(region1_mesh)
			c_r2, axis_r2 = get_principal_axis(region2_mesh)
			
			if np.dot(axis_r1, axis_r2) < 0:
				axis_r2 = -axis_r2
				print("主軸の向きを反転しました")
			
			region2_aligned, R_align = align_axis(region2_mesh, axis_r2, axis_r1)
			c_r2_aligned, _ = get_principal_axis(region2_aligned)
			translation_vector = c_r1 - c_r2_aligned
			region2_aligned.translate(translation_vector, relative=True)
			
			# 主軸アライメントの変換行列
			axis_transform = np.eye(4)
			axis_transform[:3, :3] = R_align
			axis_transform[:3, 3] = c_r1 - R_align @ c_r2
			
			print(f"主軸アライメント変換行列:\n{axis_transform}")
			
			# 主軸アライメント後を表示
			print("\n[主軸アライメント後] 表示中...")
			r1_disp = copy.deepcopy(region1_mesh)
			r2_disp = copy.deepcopy(region2_aligned)
			r1_disp.paint_uniform_color([1.0, 0.0, 0.0])
			r2_disp.paint_uniform_color([0.0, 1.0, 0.0])
			o3d.visualization.draw_geometries(
				[r1_disp, r2_disp],
				window_name="主軸・重心一致後: モデル1(赤) + モデル2(緑)")
			self._restore_window_geometry()
			
			# RANSAC
			print("\n=== RANSAC位置合わせ ===")
			parent_pcd = region1_mesh.sample_points_uniformly(number_of_points=self.cs_sample_points.get())
			child_pcd = region2_aligned.sample_points_uniformly(number_of_points=self.cs_sample_points.get())
			
			# voxel_size自動推定
			try:
				num_points = len(parent_pcd.points)
				sample_n = min(500, max(10, num_points))
				sample_indices = np.random.choice(num_points, sample_n, replace=False)
				kdt = o3d.geometry.KDTreeFlann(parent_pcd)
				nn_dists = []
				for idx in sample_indices:
					[_, idxs, dists] = kdt.search_knn_vector_3d(parent_pcd.points[idx], 2)
					if len(dists) >= 2:
						nn_dists.append(float(np.sqrt(dists[1])))
				voxel_size = max(float(np.mean(nn_dists)) * 1.5, 1e-6) if nn_dists else max(self.cs_ransac_distance_threshold.get(), 1e-6)
			except Exception:
				voxel_size = max(self.cs_ransac_distance_threshold.get(), 1e-6)
			
			print(f"voxel_size={voxel_size:.6f}")
			
			parent_down = parent_pcd.voxel_down_sample(voxel_size)
			child_down = child_pcd.voxel_down_sample(voxel_size)
			
			radius_normal = voxel_size * 2
			parent_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			child_down.estimate_normals(o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			
			radius_feature = voxel_size * 5
			parent_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				parent_down, o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			child_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				child_down, o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			
			dist_threshold_ransac = self.cs_ransac_distance_threshold.get()
			result_ransac = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
				child_down, parent_down, child_fpfh, parent_fpfh,
				mutual_filter=True,
				max_correspondence_distance=dist_threshold_ransac,
				estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPoint(self.cs_enable_scaling.get()),
				ransac_n=3,
				checkers=[
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(0.9),
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(dist_threshold_ransac)
				],
				criteria=o3d.pipelines.registration.RANSACConvergenceCriteria(
					self.cs_ransac_max_iterations.get(), self.cs_ransac_confidence.get())
			)
			
			ransac_transform = result_ransac.transformation
			print(f"RANSAC fitness={result_ransac.fitness:.6f}, RMSE={result_ransac.inlier_rmse:.6f}")
			
			# RANSAC後を表示
			region2_ransac = copy.deepcopy(region2_aligned)
			region2_ransac.transform(ransac_transform)
			
			print("\n[RANSAC後] 表示中...")
			r1_disp = copy.deepcopy(region1_mesh)
			r2_disp = copy.deepcopy(region2_ransac)
			r1_disp.paint_uniform_color([1.0, 0.0, 0.0])
			r2_disp.paint_uniform_color([1.0, 1.0, 0.0])
			o3d.visualization.draw_geometries(
				[r1_disp, r2_disp],
				window_name="RANSAC後: モデル1(赤) + モデル2(黄)")
			self._restore_window_geometry()
			
			# ICP
			print("\n=== ICP精密位置合わせ ===")
			result_icp = o3d.pipelines.registration.registration_icp(
				child_down, parent_down,
				self.cs_icp_threshold.get(),
				ransac_transform,
				o3d.pipelines.registration.TransformationEstimationPointToPoint(self.cs_enable_scaling.get()),
				o3d.pipelines.registration.ICPConvergenceCriteria(max_iteration=self.cs_icp_max_iterations.get())
			)
			
			icp_transform = result_icp.transformation
			print(f"ICP fitness={result_icp.fitness:.6f}, RMSE={result_icp.inlier_rmse:.6f}")

			# スケール補正が有効な場合、スケール係数をログ出力
			if self.cs_enable_scaling.get():
				scale_factor = np.cbrt(np.linalg.det(np.array(icp_transform[:3, :3])))
				print(f"  スケール補正係数: {scale_factor:.6f} ({(scale_factor-1)*100:.2f}%)")
			
			# 最終変換行列 = ICP @ axis_transform （元の座標系からの合成変換）
			final_transform = icp_transform @ axis_transform
			print(f"\n最終変換行列:\n{final_transform}")
			
			# ICP後を表示
			region2_final = copy.deepcopy(region2_mesh)
			region2_final.transform(final_transform)
			
			print("\n[ICP後] 表示中...")
			r1_disp = copy.deepcopy(region1_mesh)
			r2_disp = copy.deepcopy(region2_final)
			r1_disp.paint_uniform_color([1.0, 0.0, 0.0])
			r2_disp.paint_uniform_color([0.0, 1.0, 0.0])
			o3d.visualization.draw_geometries(
				[r1_disp, r2_disp],
				window_name="ICP後（最終）: モデル1(赤) + モデル2(緑)")
			self._restore_window_geometry()

			# --- フィッティング確認 → キャンセル可能 ---
			scale_info = ""
			if self.cs_enable_scaling.get():
				sf = np.cbrt(np.linalg.det(np.array(icp_transform[:3, :3])))
				scale_info = f"\n  スケール補正: {sf:.6f} ({(sf-1)*100:.2f}%)\n"
			proceed = messagebox.askyesno("フィッティング確認",
				f"位置合わせの結果を確認してください。\n\n"
				f"  ICP fitness: {result_icp.fitness:.6f}\n"
				f"  ICP RMSE:    {result_icp.inlier_rmse:.6f}\n"
				f"{scale_info}\n"
				f"このまま軟骨分離を実行しますか？\n"
				f"（フィッティングが不十分な場合は「いいえ」でキャンセル）")
			if not proceed:
				print("ユーザーがキャンセルしました。")
				messagebox.showinfo("キャンセル", "軟骨分離をキャンセルしました。\nパラメータを調整して再実行してください。")
				return

			# --- Phase 2: モデル2全体を変換 ---
			print("\n--- Phase 2: モデル2全体に変換行列を適用 ---")
			
			model1_whole = pv.read(self.cs_prox_model1_whole_path.get())
			model2_whole = pv.read(self.cs_prox_model2_whole_path.get())
			
			# 変換を適用
			model2_transformed = model2_whole.copy()
			model2_transformed.transform(final_transform, inplace=True)
			
			print(f"モデル1: {model1_whole.n_points}点, {model1_whole.n_cells}面")
			print(f"モデル2（変換後）: {model2_transformed.n_points}点, {model2_transformed.n_cells}面")
			
			# --- Phase 3: 距離ベース軟骨抽出 ---
			print("\n--- Phase 3: 軟骨領域の抽出（距離ベース）---")
			print("  モデル1(骨+軟骨)の表面から、モデル2(骨)の外側にある領域を軟骨として抽出します。")

			cartilage_mesh = None

			try:
				# 両モデルの表面メッシュを取得
				m1_surface = model1_whole.extract_surface()
				m2_surface = model2_transformed.extract_surface()

				print(f"  モデル1 表面: {m1_surface.n_points}点, {m1_surface.n_cells}面")
				print(f"  モデル2 表面: {m2_surface.n_points}点, {m2_surface.n_cells}面")

				# --- VTK implicit distance による符号付き距離計算 ---
				# vtkImplicitPolyDataDistance: 幾何学的に正確な符号付き距離
				# 正 = Model2の外側（軟骨）、負 = Model2の内側（骨）
				print("  VTK implicit distance で符号付き距離を計算中...")
				m1_with_dist = m1_surface.compute_implicit_distance(m2_surface, inplace=False)
				signed_distances = m1_with_dist['implicit_distance']

				# 統計出力
				n_outside = np.sum(signed_distances > 0)
				n_inside = np.sum(signed_distances <= 0)
				print(f"  外側（軟骨候補）: {n_outside}点 ({100*n_outside/max(1,m1_surface.n_points):.1f}%)")
				print(f"  内側（骨）:       {n_inside}点 ({100*n_inside/max(1,m1_surface.n_points):.1f}%)")
				if n_outside > 0:
					outside_dists = signed_distances[signed_distances > 0]
					print(f"  外側距離: 平均 {outside_dists.mean():.3f}mm, 最大 {outside_dists.max():.3f}mm")
				if n_inside > 0:
					inside_dists = signed_distances[signed_distances <= 0]
					print(f"  内側距離: 平均 {inside_dists.mean():.3f}mm, 最小 {inside_dists.min():.3f}mm")

				# 閾値適用: signed_distance > threshold の頂点を軟骨として抽出
				threshold = self.cs_distance_threshold.get()
				mask = signed_distances > threshold
				n_selected = np.sum(mask)
				print(f"\n  距離 > {threshold:.2f}mm の頂点: {n_selected}/{m1_surface.n_points} ({100*n_selected/max(1,m1_surface.n_points):.1f}%)")

				if n_selected > 0:
					# 頂点マスクからセル（面）を抽出
					cartilage_mesh = m1_surface.extract_points(mask, adjacent_cells=True)
					cartilage_mesh = cartilage_mesh.extract_surface()

					if cartilage_mesh.n_cells > 0:
						# クリーンアップ: 小さい断片を除去
						try:
							connected = cartilage_mesh.connectivity(largest=False)
							if 'RegionId' in connected.array_names:
								region_ids = connected['RegionId']
								unique_ids, counts = np.unique(region_ids, return_counts=True)
								min_cells = max(10, cartilage_mesh.n_cells * 0.01)
								large_regions = unique_ids[counts >= min_cells]
								if len(large_regions) > 0 and len(large_regions) < len(unique_ids):
									keep_mask = np.isin(region_ids, large_regions)
									cartilage_mesh = connected.extract_points(keep_mask, adjacent_cells=True).extract_surface()
									removed = len(unique_ids) - len(large_regions)
									print(f"  小断片 {removed}個を除去")
						except Exception:
							pass

						print(f"  ✓ 軟骨抽出 成功: {cartilage_mesh.n_points}点, {cartilage_mesh.n_cells}面")
					else:
						cartilage_mesh = None
						print("  ✗ セル抽出結果が空")
				else:
					print("  ✗ 閾値を超える頂点なし（閾値を小さくしてみてください）")

			except Exception as e_dist:
				print(f"  ✗ 距離ベース抽出 失敗: {e_dist}")
				import traceback; traceback.print_exc()

			if cartilage_mesh is None or cartilage_mesh.n_cells == 0:
				messagebox.showerror("エラー",
					"軟骨領域の抽出に失敗しました。\n\n"
					"考えられる対処:\n"
					"・「距離閾値」を小さくする（現在: {:.2f}mm → 0.1mm等）\n"
					"・位置合わせを改善（RANSAC/ICPパラメータを調整）\n"
					"・「スケール補正」を有効にする".format(self.cs_distance_threshold.get()))
				return

			print(f"  ✓ 軟骨＋関節唇: {cartilage_mesh.n_points}点, {cartilage_mesh.n_cells}面")
			
			# --- Phase 4: 結果の可視化 ---
			print("\n--- Phase 4: 結果可視化 ---")
			
			plotter = pv.Plotter()
			plotter.set_background('white')
			
			cart_opacity = self.cs_cartilage_opacity.get()
			bone_opacity = self.cs_bone_opacity.get()
			
			# 軟骨（差分結果）
			plotter.add_mesh(cartilage_mesh, color=self.cs_cartilage_color, opacity=cart_opacity,
				show_edges=False, label=f'軟骨+関節唇 ({cartilage_mesh.n_cells}面)')
			
			# モデル2（座標一致済み・骨のみ）半透明表示
			plotter.add_mesh(model2_transformed, color=self.cs_bone_color, opacity=bone_opacity,
				show_edges=False, label=f'骨のみモデル ({model2_transformed.n_cells}面)')
			
			# モデル1全体を薄く重ねて確認用
			plotter.add_mesh(model1_whole, color='gray', opacity=0.08,
				show_edges=False, label='モデル1全体（参照）')
			
			plotter.add_axes()
			plotter.add_legend()
			plotter.add_text("ブーリアン差分: モデル1 − モデル2", position='upper_right', font_size=10)
			plotter.show()
			# PyVistaリソースを明示解放（ウィンドウ閉じた後のフリーズを軽減）
			plotter.close()
			del plotter
			import gc; gc.collect()
			self._restore_window_geometry()
			self.update_idletasks()  # tkinterイベントを処理してUI応答性を回復

			# --- Phase 5: 保存オプション ---
			print("\n--- Phase 5: 保存 ---")
			
			want_save = messagebox.askyesno("保存確認",
				f"分離結果を保存しますか？\n\n"
				f"・軟骨（＋関節唇）: {cartilage_mesh.n_cells}面\n"
				f"・モデル2（座標系一致済み）: {model2_transformed.n_cells}面")
			
			if want_save:
				# 軟骨メッシュを保存
				original = Path(self.cs_prox_model1_whole_path.get())
				default_name = original.stem + "_cartilage.stl"
				save_path = filedialog.asksaveasfilename(
					title="軟骨モデルを保存",
					defaultextension=".stl",
					initialfile=default_name,
					filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj"), ("All files", "*.*")]
				)
				if save_path:
					cartilage_mesh.save(save_path)
					print(f"軟骨モデルを保存: {Path(save_path).name}")
					# Simulatorタブの軟骨モデルパスに自動反映
					side = getattr(self, '_cs_executing_side', 'prox')
					if side == "dist":
						self.dist_cartilage_model_path.set(save_path)
						print(f"  → Simulatorタブ「遠位軟骨モデル」に反映しました")
					else:
						self.prox_cartilage_model_path.set(save_path)
						print(f"  → Simulatorタブ「近位軟骨モデル」に反映しました")

				# 座標系一致済みのモデル2を保存
				want_m2 = messagebox.askyesno("座標系一致済みモデル保存",
					"モデル2（骨のみ）の座標系一致済みモデルも保存しますか？\n"
					"（FEMで骨と軟骨に異なる物性値を設定する際に使用）")
				if want_m2:
					original2 = Path(self.cs_prox_model2_whole_path.get())
					default_name2 = original2.stem + "_aligned.stl"
					save_path2 = filedialog.asksaveasfilename(
						title="座標系一致済みモデル2を保存",
						defaultextension=".stl",
						initialfile=default_name2,
						filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj"), ("All files", "*.*")]
					)
					if save_path2:
						model2_transformed.save(save_path2)
						print(f"座標系一致済みモデル2を保存: {Path(save_path2).name}")
			
			# --- カラー統合モデルの保存 ---
			want_combined = messagebox.askyesno(
				"カラー統合モデル保存",
				"骨モデルと軟骨モデルを色分けした状態で\n"
				"1つの3Dモデル（PLY形式）として保存しますか？\n\n"
				f"  骨モデル : {self.cs_bone_color}  ({model2_transformed.n_cells}面)\n"
				f"  軟骨モデル: {self.cs_cartilage_color}  ({cartilage_mesh.n_cells}面)\n\n"
				"MeshLabやBlenderで色付きのまま開けます。")
			if want_combined:
				def _hex_to_rgb(hex_color):
					h = hex_color.lstrip('#')
					return [int(h[i:i+2], 16) for i in (0, 2, 4)]

				def _write_colored_ply(save_path, meshes_with_colors):
					"""
					頂点カラー付きPLY（バイナリ）を直接書き出す。
					PyVistaのsave()はRGB属性を無視するためカスタム実装。
					meshes_with_colors: [(pv_mesh, [r,g,b]), ...]
					"""
					import struct
					all_verts  = []
					all_colors = []
					all_faces  = []
					vert_offset = 0
					for mesh, color in meshes_with_colors:
						tri = mesh.triangulate()
						pts = tri.points.astype(np.float32)
						all_verts.append(pts)
						all_colors.append(np.full((len(pts), 3), color, dtype=np.uint8))
						all_faces.append(tri.faces.reshape(-1, 4)[:, 1:].astype(np.int32) + vert_offset)
						vert_offset += len(pts)
					verts  = np.concatenate(all_verts,  axis=0)
					colors = np.concatenate(all_colors, axis=0)
					faces  = np.concatenate(all_faces,  axis=0)
					header = (
						"ply\n"
						"format binary_little_endian 1.0\n"
						f"element vertex {len(verts)}\n"
						"property float x\n"
						"property float y\n"
						"property float z\n"
						"property uchar red\n"
						"property uchar green\n"
						"property uchar blue\n"
						f"element face {len(faces)}\n"
						"property list uchar int vertex_indices\n"
						"end_header\n"
					).encode('ascii')
					with open(save_path, 'wb') as fp:
						fp.write(header)
						for (x, y, z), (r, g, b) in zip(verts, colors):
							fp.write(struct.pack('<fffBBB', x, y, z, r, g, b))
						for (v0, v1, v2) in faces:
							fp.write(struct.pack('<Biii', 3, v0, v1, v2))

				bone_rgb = _hex_to_rgb(self.cs_bone_color)
				cart_rgb = _hex_to_rgb(self.cs_cartilage_color)

				original_path = Path(self.cs_prox_model1_whole_path.get())
				default_name = original_path.stem + "_カラー統合.ply"
				save_path_combined = filedialog.asksaveasfilename(
					title="カラー統合モデルを保存（骨＋軟骨・色分け済み）",
					defaultextension=".ply",
					initialdir=str(original_path.parent),
					initialfile=default_name,
					filetypes=[
						("PLY files（色情報あり）", "*.ply"),
						("All files", "*.*")
					]
				)
				if save_path_combined:
					_write_colored_ply(
						save_path_combined,
						[
							(model2_transformed, bone_rgb),
							(cartilage_mesh,     cart_rgb),
						]
					)
					print(f"カラー統合モデルを保存: {save_path_combined}")
					messagebox.showinfo(
						"保存完了",
						f"カラー統合モデルを保存しました。\n\n"
						f"保存先:\n{save_path_combined}\n\n"
						f"MeshLab / Blender / Windows 3Dビューアで色付きのまま開けます。")

			print("\n=== 軟骨分離処理完了 ===")

		except Exception as e:
			messagebox.showerror("エラー", f"軟骨分離中にエラーが発生しました:\n{str(e)}")
			print(f"エラー詳細: {e}")
			traceback.print_exc()

	# ----- Actions (Dummy) -----
	def on_visualize_prox(self) -> None:
		# 近位モデルと近位PPを表示（PyVista）
		model_path = self.prox_model_path.get().strip()
		pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		# 膝関節 (ISB) では OLMN ファイルは不要
		knee_mode = (self.joint_var.get() == 2)
		if not model_path or not pp_abcd_path or (not knee_mode and not pp_olmn_path):
			if knee_mode:
				messagebox.showwarning("入力不足", "近位モデルと近位特徴点(ABCD)を選択してください。")
			else:
				messagebox.showwarning("入力不足", "近位モデルと近位特徴点(ABCD, OLMN)を選択してください。")
			return

		# ライブラリの確認（遅延インポート）
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return

		# PP読み込み（PickedPoints XML対応）- 膝関節モードでは OLMN なしも可
		try:
			points_abcd, labels_abcd = self._parse_pp_file(pp_abcd_path)
			if pp_olmn_path:
				points_olmn, labels_olmn = self._parse_pp_file(pp_olmn_path)
				points = np.vstack([points_abcd, points_olmn])
				labels = labels_abcd + labels_olmn
			else:
				points = points_abcd
				labels = labels_abcd
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 座標系構築（A, B, C, D, O点から）
		try:
			origin, x_axis, y_axis, z_axis = self._build_coordinate_system(points, labels)
			# 自動オフセット設定（O'点がある場合）
			self._try_auto_set_offset(points, labels, origin, x_axis, y_axis, z_axis,
			                          self.prox_offset_x, self.prox_offset_y, self.prox_offset_z)
		except Exception as e:
			messagebox.showwarning("座標系構築", f"座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			origin, x_axis, y_axis, z_axis = None, None, None, None

		# メッシュ読み込み（PyVista）
		try:
			mesh_full = pv.read(model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"近位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# 球体で寛骨臼領域を抽出（座標系が構築できた場合）
		mesh_extracted = None
		origin_prime = None
		if origin is not None:
			try:
				# O'原点位置を計算（ローカル座標系に沿ったオフセット）
				origin_prime = origin + (
					self.prox_offset_x.get() * x_axis +
					self.prox_offset_y.get() * y_axis +
					self.prox_offset_z.get() * z_axis
				)
				print(f"[可視化] 近位O: {origin}, 近位O': {origin_prime}, オフセット: [{self.prox_offset_x.get()}, {self.prox_offset_y.get()}, {self.prox_offset_z.get()}]")
				
				radius = self.prox_radius.get()
				mesh_extracted = self._extract_region_by_sphere(mesh_full, origin_prime, radius)
				if mesh_extracted.n_points == 0:
					messagebox.showwarning("抽出失敗", f"半径 {radius}mm の球体内にメッシュが見つかりませんでした。")
					mesh_extracted = None
			except Exception as e:
				messagebox.showwarning("球体抽出", f"球体抽出に失敗しました:\n{e}\n\n元のメッシュのみ表示します。")
				mesh_extracted = None

		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		# ウィンドウサイズをディスプレイの90%に設定
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)

		# プロッタ作成（格子/目盛りなし、モデル+点のみ）
		plotter = pv.Plotter(title="近位モデル/特徴点の可視化", window_size=(window_width, window_height))
		plotter.set_background("white")
		
		# 元のメッシュ全体（半透明・グレー）
		mesh_full_actor = plotter.add_mesh(mesh_full, color="lightgray", smooth_shading=True, show_edges=False, opacity=0.3)
		
		# 抽出された領域（設定したカラー・不透明）
		mesh_extracted_actor = None
		if mesh_extracted is not None and mesh_extracted.n_points > 0:
			mesh_extracted_actor = plotter.add_mesh(mesh_extracted, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=1.0)
		
		# 点群（設定したカラーを使用）
		points_actor = plotter.add_points(points, color=self.prox_color, point_size=12, render_points_as_spheres=True)
		
		# 点のラベルを表示
		label_actors = []
		for i, (pt, label) in enumerate(zip(points, labels)):
			label_actor = plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.prox_color, bold=True, shadow=True, show_points=False)
			label_actors.append(label_actor)
		
		# 座標系の表示（構築できた場合）
		axis_actors = []
		axis_prime_actors = []
		if origin is not None and x_axis is not None and y_axis is not None and z_axis is not None:
			axis_length = 50.0  # 軸の長さ
			# O座標系（実線）
			# X軸（赤）
			x_line = pv.Line(origin, origin + x_axis * axis_length)
			x_actor = plotter.add_mesh(x_line, color="red", line_width=3)
			axis_actors.append(x_actor)
			# Y軸（緑）
			y_line = pv.Line(origin, origin + y_axis * axis_length)
			y_actor = plotter.add_mesh(y_line, color="green", line_width=3)
			axis_actors.append(y_actor)
			# Z軸（青）
			z_line = pv.Line(origin, origin + z_axis * axis_length)
			z_actor = plotter.add_mesh(z_line, color="blue", line_width=3)
			axis_actors.append(z_actor)
			
			# O'座標系（破線風・細め）
			if origin_prime is not None:
				axis_prime_length = 40.0  # O'の軸は少し短く
				dash_length = 3.0  # 破線の長さ
				gap_length = 2.0   # 破線の間隔
				
				# X'軸（赤・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = origin_prime + x_axis * current_pos
					end = origin_prime + x_axis * min(current_pos + dash_length, axis_prime_length)
					x_prime_segment = pv.Line(start, end)
					x_prime_actor = plotter.add_mesh(x_prime_segment, color="darkred", line_width=2)
					axis_prime_actors.append(x_prime_actor)
					current_pos += dash_length + gap_length
				
				# Y'軸（緑・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = origin_prime + y_axis * current_pos
					end = origin_prime + y_axis * min(current_pos + dash_length, axis_prime_length)
					y_prime_segment = pv.Line(start, end)
					y_prime_actor = plotter.add_mesh(y_prime_segment, color="darkgreen", line_width=2)
					axis_prime_actors.append(y_prime_actor)
					current_pos += dash_length + gap_length
				
				# Z'軸（青・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = origin_prime + z_axis * current_pos
					end = origin_prime + z_axis * min(current_pos + dash_length, axis_prime_length)
					z_prime_segment = pv.Line(start, end)
					z_prime_actor = plotter.add_mesh(z_prime_segment, color="darkblue", line_width=2)
					axis_prime_actors.append(z_prime_actor)
					current_pos += dash_length + gap_length
				
				# O'の原点を球で表示
				o_prime_sphere = pv.Sphere(radius=2.0, center=origin_prime)
				o_prime_actor = plotter.add_mesh(o_prime_sphere, color="orange", opacity=0.8)
				axis_prime_actors.append(o_prime_actor)
		
		# 左上にチェックボックスUI
		def toggle_full_model(state):
			"""元のメッシュ全体の表示/非表示"""
			mesh_full_actor.SetVisibility(state)
		
		def toggle_extracted_model(state):
			"""抽出された領域の表示/非表示"""
			if mesh_extracted_actor is not None:
				mesh_extracted_actor.SetVisibility(state)
		
		def toggle_points(state):
			points_actor.SetVisibility(state)
			for label_actor in label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_axes(state):
			"""O座標系の表示/非表示を切り替え"""
			for actor in axis_actors:
				actor.SetVisibility(state)
		
		def toggle_axes_prime(state):
			"""O'座標系の表示/非表示を切り替え"""
			for actor in axis_prime_actors:
				actor.SetVisibility(state)
		
		def update_full_opacity(value):
			"""元のメッシュの透明度を変更"""
			mesh_full_actor.GetProperty().SetOpacity(value)
		
		def update_extracted_opacity(value):
			"""抽出された領域の透明度を変更"""
			if mesh_extracted_actor is not None:
				mesh_extracted_actor.GetProperty().SetOpacity(value)
		
		# チェックボックスとラベル（設定したカラーを使用）
		plotter.add_text("Full", position=(10, 10), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_full_model, value=True, position=(100, 12), size=20, border_size=1, color_on="lightgray", color_off="white")
		
		if mesh_extracted_actor is not None:
			plotter.add_text("Ext", position=(10, 40), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_extracted_model, value=True, position=(100, 42), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			checkbox_offset = 70
		else:
			checkbox_offset = 40
		
		plotter.add_text("PP", position=(10, checkbox_offset), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_points, value=True, position=(100, checkbox_offset+2), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		# 座標系の表示チェックボックス（構築できた場合のみ）
		if axis_actors:
			plotter.add_text("O", position=(10, checkbox_offset+30), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_axes, value=True, position=(100, checkbox_offset+32), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			checkbox_offset += 30
		
		# O'座標系の表示チェックボックス（O'が存在する場合のみ）
		if axis_prime_actors:
			plotter.add_text("O'", position=(10, checkbox_offset+30), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_axes_prime, value=True, position=(100, checkbox_offset+32), size=20, border_size=1, color_on="orange", color_off="white")
		
		# 透明度スライダー（元のメッシュ用）
		plotter.add_slider_widget(
			update_full_opacity,
			rng=[0.0, 1.0],
			value=0.3,
			title="Full ",
			pointa=(0.05, 0.02),
			pointb=(0.15, 0.02),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color="lightgray",
		)
		
		# 透明度スライダー（抽出領域用）
		if mesh_extracted_actor is not None:
			plotter.add_slider_widget(
				update_extracted_opacity,
				rng=[0.0, 1.0],
				value=1.0,
				title=" ",
				pointa=(0.05, 0.06),
				pointb=(0.15, 0.06),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.prox_color,
			)

		# 不要な装飾は追加しない（grid/axes/bounds未追加）
		# 自動でカメラをフィット
		self._show_plotter_coop(plotter)  # ユーザーはマウスで視点操作可能

	def on_visualize_dist(self) -> None:
		# 遠位モデルと遠位PPを表示（PyVista）
		dist_model_path = self.dist_model_path.get().strip()
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		# 膝関節 (ISB) では OLMN ファイルは不要
		knee_mode = (self.joint_var.get() == 2)
		if not dist_model_path or not dist_pp_abc_path or (not knee_mode and not dist_pp_olmn_path):
			if knee_mode:
				messagebox.showwarning("入力不足", "遠位モデルと遠位特徴点(ABCD)を選択してください。")
			else:
				messagebox.showwarning("入力不足", "遠位モデルと遠位特徴点(ABC, OLMN)を選択してください。")
			return

		# ライブラリの確認（遅延インポート）
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return

		# PP読み込み（PickedPoints XML対応）- 膝関節モードでは OLMN なしも可
		try:
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			if dist_pp_olmn_path:
				dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
				dist_points = np.vstack([dist_points_abc, dist_points_olmn])
				dist_labels = dist_labels_abc + dist_labels_olmn
			else:
				dist_points = dist_points_abc
				dist_labels = dist_labels_abc
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 座標系構築（A, B, O点から - 遠位用）
		try:
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = self._build_coordinate_system_dist(dist_points, dist_labels)
			# 自動オフセット設定（遠位O'点がある場合）
			self._try_auto_set_offset(dist_points, dist_labels, dist_origin, dist_x_axis, dist_y_axis, dist_z_axis,
			                          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z)
		except Exception as e:
			messagebox.showwarning("座標系構築", f"座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = None, None, None, None

		# メッシュ読み込み（PyVista）
		try:
			dist_mesh_full = pv.read(dist_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"遠位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# 球体で大腿骨頭領域を抽出（座標系が構築できた場合）
		dist_mesh_extracted = None
		dist_origin_prime = None
		if dist_origin is not None:
			try:
				# O'原点位置を計算（ローカル座標系に沿ったオフセット）
				# 遠位座標系は dist_origin, dist_x_axis, ... で定義されている
				dist_origin_prime = dist_origin + (
					self.dist_offset_x.get() * dist_x_axis +
					self.dist_offset_y.get() * dist_y_axis +
					self.dist_offset_z.get() * dist_z_axis
				)
				print(f"[可視化] 遠位O: {dist_origin}, 遠位O': {dist_origin_prime}, オフセット: [{self.dist_offset_x.get()}, {self.dist_offset_y.get()}, {self.dist_offset_z.get()}]")
				
				radius = self.dist_radius.get()
				dist_mesh_extracted = self._extract_region_by_sphere(dist_mesh_full, dist_origin_prime, radius)
				if dist_mesh_extracted.n_points == 0:
					messagebox.showwarning("抽出失敗", f"半径 {radius}mm の球体内にメッシュが見つかりませんでした。")
					dist_mesh_extracted = None
			except Exception as e:
				messagebox.showwarning("球体抽出", f"球体抽出に失敗しました:\n{e}\n\n元のメッシュのみ表示します。")
				dist_mesh_extracted = None

		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		# ウィンドウサイズをディスプレイの90%に設定
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)

		# プロッタ作成（格子/目盛りなし、モデル+点のみ）
		dist_plotter = pv.Plotter(title="遠位モデル/特徴点の可視化", window_size=(window_width, window_height))
		dist_plotter.set_background("white")
		
		# 元のメッシュ全体（半透明・グレー）
		dist_mesh_full_actor = dist_plotter.add_mesh(dist_mesh_full, color="lightgray", smooth_shading=True, show_edges=False, opacity=0.3)
		
		# 抽出された領域（設定したカラー・不透明）
		dist_mesh_extracted_actor = None
		if dist_mesh_extracted is not None and dist_mesh_extracted.n_points > 0:
			dist_mesh_extracted_actor = dist_plotter.add_mesh(dist_mesh_extracted, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=1.0)
		
		# 点群（設定したカラーを使用）
		dist_points_actor = dist_plotter.add_points(dist_points, color=self.dist_color, point_size=12, render_points_as_spheres=True)
		
		# 点のラベルを表示
		dist_label_actors = []
		for i, (pt, label) in enumerate(zip(dist_points, dist_labels)):
			dist_label_actor = dist_plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.dist_color, bold=True, shadow=True, show_points=False)
			dist_label_actors.append(dist_label_actor)
		
		# 座標系の表示（構築できた場合）
		dist_axis_actors = []
		dist_axis_prime_actors = []  # O'座標系用
		if dist_origin is not None and dist_x_axis is not None and dist_y_axis is not None and dist_z_axis is not None:
			axis_length = 50.0  # 軸の長さ
			# O座標系（実線）
			# X軸（赤）
			dist_x_line = pv.Line(dist_origin, dist_origin + dist_x_axis * axis_length)
			dist_x_actor = dist_plotter.add_mesh(dist_x_line, color="red", line_width=3)
			dist_axis_actors.append(dist_x_actor)
			# Y軸（緑）
			dist_y_line = pv.Line(dist_origin, dist_origin + dist_y_axis * axis_length)
			dist_y_actor = dist_plotter.add_mesh(dist_y_line, color="green", line_width=3)
			dist_axis_actors.append(dist_y_actor)
			# Z軸（青）
			dist_z_line = pv.Line(dist_origin, dist_origin + dist_z_axis * axis_length)
			dist_z_actor = dist_plotter.add_mesh(dist_z_line, color="blue", line_width=3)
			dist_axis_actors.append(dist_z_actor)

			# O'座標系（破線風・細め）
			if dist_origin_prime is not None:
				axis_prime_length = 40.0  # O'の軸は少し短く
				dash_length = 3.0  # 破線の長さ
				gap_length = 2.0   # 破線の間隔
				
				# X'軸（赤・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = dist_origin_prime + dist_x_axis * current_pos
					end = dist_origin_prime + dist_x_axis * min(current_pos + dash_length, axis_prime_length)
					x_prime_segment = pv.Line(start, end)
					x_prime_actor = dist_plotter.add_mesh(x_prime_segment, color="darkred", line_width=2)
					dist_axis_prime_actors.append(x_prime_actor)
					current_pos += dash_length + gap_length
				
				# Y'軸（緑・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = dist_origin_prime + dist_y_axis * current_pos
					end = dist_origin_prime + dist_y_axis * min(current_pos + dash_length, axis_prime_length)
					y_prime_segment = pv.Line(start, end)
					y_prime_actor = dist_plotter.add_mesh(y_prime_segment, color="darkgreen", line_width=2)
					dist_axis_prime_actors.append(y_prime_actor)
					current_pos += dash_length + gap_length
				
				# Z'軸（青・破線風）
				current_pos = 0.0
				while current_pos < axis_prime_length:
					start = dist_origin_prime + dist_z_axis * current_pos
					end = dist_origin_prime + dist_z_axis * min(current_pos + dash_length, axis_prime_length)
					z_prime_segment = pv.Line(start, end)
					z_prime_actor = dist_plotter.add_mesh(z_prime_segment, color="darkblue", line_width=2)
					dist_axis_prime_actors.append(z_prime_actor)
					current_pos += dash_length + gap_length
				
				# O'の原点を球で表示
				o_prime_sphere = pv.Sphere(radius=2.0, center=dist_origin_prime)
				o_prime_actor = dist_plotter.add_mesh(o_prime_sphere, color="orange", opacity=0.8)
				dist_axis_prime_actors.append(o_prime_actor)
		
		# 左上にチェックボックスUI
		def toggle_dist_full_model(state):
			"""元のメッシュ全体の表示/非表示"""
			dist_mesh_full_actor.SetVisibility(state)
		
		def toggle_dist_extracted_model(state):
			"""抽出された領域の表示/非表示"""
			if dist_mesh_extracted_actor is not None:
				dist_mesh_extracted_actor.SetVisibility(state)
		
		def toggle_dist_points(state):
			dist_points_actor.SetVisibility(state)
			for dist_label_actor in dist_label_actors:
				dist_label_actor.SetVisibility(state)
		
		def toggle_dist_axes(state):
			"""座標系の表示/非表示を切り替え"""
			for actor in dist_axis_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_axes_prime(state):
			"""O'座標系の表示/非表示を切り替え"""
			for actor in dist_axis_prime_actors:
				actor.SetVisibility(state)
		
		def update_dist_full_opacity(value):
			"""元のメッシュの透明度を変更"""
			dist_mesh_full_actor.GetProperty().SetOpacity(value)
		
		def update_dist_extracted_opacity(value):
			"""抽出された領域の透明度を変更"""
			if dist_mesh_extracted_actor is not None:
				dist_mesh_extracted_actor.GetProperty().SetOpacity(value)
		
		# チェックボックスとラベル（設定したカラーを使用）
		dist_plotter.add_text("Full", position=(10, 10), font_size=10, color="black")
		dist_plotter.add_checkbox_button_widget(toggle_dist_full_model, value=True, position=(100, 12), size=20, border_size=1, color_on="lightgray", color_off="white")
		
		if dist_mesh_extracted_actor is not None:
			dist_plotter.add_text("Ext", position=(10, 40), font_size=10, color="black")
			dist_plotter.add_checkbox_button_widget(toggle_dist_extracted_model, value=True, position=(100, 42), size=20, border_size=1, color_on=self.dist_color, color_off="white")
			dist_checkbox_offset = 70
		else:
			dist_checkbox_offset = 40
		
		dist_plotter.add_text("PP", position=(10, dist_checkbox_offset), font_size=10, color="black")
		dist_plotter.add_checkbox_button_widget(toggle_dist_points, value=True, position=(100, dist_checkbox_offset+2), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		# 座標系の表示チェックボックス（構築できた場合のみ）
		if dist_axis_actors:
			dist_plotter.add_text("O", position=(10, dist_checkbox_offset+30), font_size=10, color="black")
			dist_plotter.add_checkbox_button_widget(toggle_dist_axes, value=True, position=(100, dist_checkbox_offset+32), size=20, border_size=1, color_on=self.dist_color, color_off="white")
			dist_checkbox_offset += 30
			
		# O'座標系の表示チェックボックス（O'が存在する場合のみ）
		if dist_axis_prime_actors:
			dist_plotter.add_text("O'", position=(10, dist_checkbox_offset+30), font_size=10, color="black")
			dist_plotter.add_checkbox_button_widget(toggle_dist_axes_prime, value=True, position=(100, dist_checkbox_offset+32), size=20, border_size=1, color_on="orange", color_off="white")
		
		# 透明度スライダー（元のメッシュ用）
		dist_plotter.add_slider_widget(
			update_dist_full_opacity,
			rng=[0.0, 1.0],
			value=0.3,
			title="Full ",
			pointa=(0.05, 0.02),
			pointb=(0.15, 0.02),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color="lightgray",
		)
		
		# 透明度スライダー（抽出領域用）
		if dist_mesh_extracted_actor is not None:
			dist_plotter.add_slider_widget(
				update_dist_extracted_opacity,
				rng=[0.0, 1.0],
				value=1.0,
				title="",
				pointa=(0.05, 0.06),
				pointb=(0.15, 0.06),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.dist_color,
			)

		# 不要な装飾は追加しない（grid/axes/bounds未追加）
		# 自動でカメラをフィット
		self._show_plotter_coop(dist_plotter)  # ユーザーはマウスで視点操作可能

	def on_visualize_all(self) -> None:
		# 近位・遠位モデルと特徴点を同時表示（PyVista）
		prox_model_path = self.prox_model_path.get().strip()
		prox_pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		prox_pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		dist_model_path = self.dist_model_path.get().strip()
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		
		# 膝関節 (ISB仕様) では OLMN ファイルは不要（原点は ABCD相当の中で決まる）
		knee_mode = (self.joint_var.get() == 2)
		required_paths = [prox_model_path, prox_pp_abcd_path, dist_model_path, dist_pp_abc_path]
		if not knee_mode:
			required_paths.extend([prox_pp_olmn_path, dist_pp_olmn_path])
		if not all(required_paths):
			if knee_mode:
				messagebox.showwarning("入力不足", "近位・遠位のモデルと特徴点(ABCD相当)を選択してください。")
			else:
				messagebox.showwarning("入力不足", "近位・遠位のモデルと特徴点(ABCD, OLMN, ABC, OLMN)をすべて選択してください。")
			return

		# ライブラリの確認（遅延インポート）
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return

		# 近位PP読み込み（膝関節モードでは OLMN ファイルなしでも可）
		try:
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(prox_pp_abcd_path)
			if prox_pp_olmn_path:
				prox_points_olmn, prox_labels_olmn = self._parse_pp_file(prox_pp_olmn_path)
				prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
				prox_labels = prox_labels_abcd + prox_labels_olmn
			else:
				prox_points = prox_points_abcd
				prox_labels = prox_labels_abcd
		except Exception as e:
			messagebox.showerror("近位PP読み込み失敗", f"近位PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 遠位PP読み込み（膝関節モードでは OLMN ファイルなしでも可）
		try:
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			if dist_pp_olmn_path:
				dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
				dist_points = np.vstack([dist_points_abc, dist_points_olmn])
				dist_labels = dist_labels_abc + dist_labels_olmn
			else:
				dist_points = dist_points_abc
				dist_labels = dist_labels_abc
		except Exception as e:
			messagebox.showerror("遠位PP読み込み失敗", f"遠位PPファイルの読み込みに失敗しました:\n{e}")
			return

		# 近位座標系構築
		try:
			prox_origin, prox_x_axis, prox_y_axis, prox_z_axis = self._build_coordinate_system(prox_points, prox_labels)
			# 自動オフセット設定
			self._try_auto_set_offset(prox_points, prox_labels, prox_origin, prox_x_axis, prox_y_axis, prox_z_axis,
			                          self.prox_offset_x, self.prox_offset_y, self.prox_offset_z)
		except Exception as e:
			messagebox.showwarning("近位座標系構築", f"近位座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			prox_origin, prox_x_axis, prox_y_axis, prox_z_axis = None, None, None, None

		# 遠位座標系構築
		try:
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = self._build_coordinate_system_dist(dist_points, dist_labels)
			# 自動オフセット設定
			self._try_auto_set_offset(dist_points, dist_labels, dist_origin, dist_x_axis, dist_y_axis, dist_z_axis,
			                          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z)
		except Exception as e:
			messagebox.showwarning("遠位座標系構築", f"遠位座標系の構築に失敗しました:\n{e}\n\n可視化は続行します。")
			dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = None, None, None, None

		# 近位メッシュ読み込み
		try:
			prox_mesh = pv.read(prox_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"近位モデルの読み込みに失敗しました:\n{e}")
			return

		# 遠位メッシュ読み込み
		try:
			dist_mesh = pv.read(dist_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"遠位モデルの読み込みに失敗しました:\n{e}")
			return

		# 遠位メッシュの元のコピーを保存（変換前）
		dist_mesh_original = dist_mesh.copy()
		
		# 遠位座標系の元の値を保存（変換前）
		dist_origin_original = dist_origin.copy() if dist_origin is not None else None
		dist_x_axis_original = dist_x_axis.copy() if dist_x_axis is not None else None
		dist_y_axis_original = dist_y_axis.copy() if dist_y_axis is not None else None
		dist_z_axis_original = dist_z_axis.copy() if dist_z_axis is not None else None

		# ========== ステップ1: 近位座標系をワールド座標系に揃える ==========
		prox_to_world = None
		if prox_origin is not None and prox_x_axis is not None and prox_y_axis is not None and prox_z_axis is not None:
			# 近位座標系 → ワールド座標系への変換行列を構築
			# 近位座標系の軸ベクトル [x, y, z] を列に持つ行列の逆行列
			prox_rotation_matrix = np.column_stack([prox_x_axis, prox_y_axis, prox_z_axis])
			prox_to_world_rotation = prox_rotation_matrix.T  # 逆行列（直交行列なので転置）
			
			# 4x4同次変換行列
			prox_to_world = np.eye(4)
			prox_to_world[:3, :3] = prox_to_world_rotation
			prox_to_world[:3, 3] = -prox_to_world_rotation @ prox_origin
			
			# 近位メッシュを変換
			prox_mesh_homo = np.hstack([prox_mesh.points, np.ones((prox_mesh.points.shape[0], 1))])
			prox_mesh_transformed = (prox_to_world @ prox_mesh_homo.T).T
			prox_mesh.points = prox_mesh_transformed[:, :3]
			
			# 近位特徴点を変換
			prox_points_homo = np.hstack([prox_points, np.ones((prox_points.shape[0], 1))])
			prox_points_transformed = (prox_to_world @ prox_points_homo.T).T
			prox_points = prox_points_transformed[:, :3]
			
			# 近位座標系を更新（ワールド座標系と一致）
			prox_origin = np.array([0.0, 0.0, 0.0])
			prox_x_axis = np.array([1.0, 0.0, 0.0])
			prox_y_axis = np.array([0.0, 1.0, 0.0])
			prox_z_axis = np.array([0.0, 0.0, 1.0])
		
		# ========== ステップ2: 遠位を近位と同じ変換でワールド座標系に移動 ==========
		# 遠位メッシュと遠位座標系をワールド座標系に変換（近位と同じ変換）
		# この段階では遠位メッシュと遠位座標系の関係は保たれる
		dist_origin_in_world = None
		dist_x_axis_in_world = None
		dist_y_axis_in_world = None
		dist_z_axis_in_world = None
		
		if dist_origin_original is not None and dist_x_axis_original is not None and dist_y_axis_original is not None and dist_z_axis_original is not None:
			if prox_to_world is not None:
				# 遠位メッシュをワールド座標系に変換
				dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
				dist_mesh_transformed = (prox_to_world @ dist_mesh_homo.T).T
				dist_mesh.points = dist_mesh_transformed[:, :3]
				
				# 遠位特徴点をワールド座標系に変換
				dist_points_homo = np.hstack([dist_points, np.ones((dist_points.shape[0], 1))])
				dist_points_transformed = (prox_to_world @ dist_points_homo.T).T
				dist_points = dist_points_transformed[:, :3]
				
				# 遠位座標系をワールド座標系に変換
				dist_origin_homo = np.append(dist_origin_original, 1)
				dist_origin_in_world = (prox_to_world @ dist_origin_homo)[:3]
				
				prox_to_world_rotation = prox_to_world[:3, :3]
				dist_x_axis_in_world = prox_to_world_rotation @ dist_x_axis_original
				dist_y_axis_in_world = prox_to_world_rotation @ dist_y_axis_original
				dist_z_axis_in_world = prox_to_world_rotation @ dist_z_axis_original
			else:
				# 近位変換がない場合はそのまま
				dist_origin_in_world = dist_origin_original
				dist_x_axis_in_world = dist_x_axis_original
				dist_y_axis_in_world = dist_y_axis_original
				dist_z_axis_in_world = dist_z_axis_original
			
			# この時点で、遠位メッシュと遠位座標系はワールド座標系に配置されているが
			# 遠位メッシュと遠位座標系の相対関係は元のまま保たれている
			
			# ========== ステップ3: 遠位座標系を近位座標系（=ワールド座標系）に揃える ==========
			# 遠位座標系 → ワールド座標系への変換行列を構築
			dist_rotation_matrix = np.column_stack([dist_x_axis_in_world, dist_y_axis_in_world, dist_z_axis_in_world])
			dist_to_world_rotation = dist_rotation_matrix.T
			
			dist_to_world = np.eye(4)
			dist_to_world[:3, :3] = dist_to_world_rotation
			dist_to_world[:3, 3] = -dist_to_world_rotation @ dist_origin_in_world
			
			# 遠位メッシュを遠位座標系基準に変換（原点中心、軸を単位ベクトルに）
			dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
			dist_mesh_transformed = (dist_to_world @ dist_mesh_homo.T).T
			dist_mesh.points = dist_mesh_transformed[:, :3]
			
			# 遠位特徴点を遠位座標系基準に変換
			dist_points_homo = np.hstack([dist_points, np.ones((dist_points.shape[0], 1))])
			dist_points_transformed = (dist_to_world @ dist_points_homo.T).T
			dist_points = dist_points_transformed[:, :3]
			
			# 遠位座標系を更新（近位座標系=ワールド座標系と一致）
			dist_origin = np.array([0.0, 0.0, 0.0])
			dist_x_axis = np.array([1.0, 0.0, 0.0])
			dist_y_axis = np.array([0.0, 1.0, 0.0])
			dist_z_axis = np.array([0.0, 0.0, 1.0])

		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)

		# プロッタ作成（関節種別でタイトル切替）
		plotter_title = "膝関節全体モデル/特徴点の可視化 (ISB)" if knee_mode else "股関節全体モデル/特徴点の可視化"
		all_plotter = pv.Plotter(title=plotter_title, window_size=(window_width, window_height))
		all_plotter.set_background("white")
		
		# 近位モデル（設定したカラーを使用）
		prox_mesh_actor = all_plotter.add_mesh(prox_mesh, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=1.0)
		# 遠位モデル（設定したカラーを使用）
		dist_mesh_actor = all_plotter.add_mesh(dist_mesh, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=1.0)
		
		# 軟骨モデルの読み込みと表示（任意）
		prox_cart_actor = None
		dist_cart_actor = None
		prox_cart_path = self.prox_cartilage_model_path.get().strip()
		dist_cart_path = self.dist_cartilage_model_path.get().strip()
		
		if prox_cart_path:
			try:
				prox_cart_mesh = pv.read(prox_cart_path)
				# 骨と同じ変換を適用
				if prox_to_world is not None:
					prox_cart_homo = np.hstack([prox_cart_mesh.points, np.ones((prox_cart_mesh.points.shape[0], 1))])
					prox_cart_mesh.points = (prox_to_world @ prox_cart_homo.T).T[:, :3]
				prox_cart_actor = all_plotter.add_mesh(prox_cart_mesh, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='近位軟骨')
				print(f"近位軟骨モデル読み込み成功: {Path(prox_cart_path).name}")
			except Exception as e:
				print(f"近位軟骨モデル読み込み失敗: {e}")
		
		if dist_cart_path:
			try:
				dist_cart_mesh_viz = pv.read(dist_cart_path)
				# 骨と同じ変換を適用（prox_to_world → dist_to_world）
				if prox_to_world is not None:
					dist_cart_homo = np.hstack([dist_cart_mesh_viz.points, np.ones((dist_cart_mesh_viz.points.shape[0], 1))])
					dist_cart_mesh_viz.points = (prox_to_world @ dist_cart_homo.T).T[:, :3]
				if dist_origin_in_world is not None:
					dist_cart_homo2 = np.hstack([dist_cart_mesh_viz.points, np.ones((dist_cart_mesh_viz.points.shape[0], 1))])
					dist_cart_mesh_viz.points = (dist_to_world @ dist_cart_homo2.T).T[:, :3]
				dist_cart_actor = all_plotter.add_mesh(dist_cart_mesh_viz, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='遠位軟骨')
				print(f"遠位軟骨モデル読み込み成功: {Path(dist_cart_path).name}")
			except Exception as e:
				print(f"遠位軟骨モデル読み込み失敗: {e}")
		
		# 近位特徴点の表示
		prox_points_actor = all_plotter.add_points(prox_points, color=self.prox_color, point_size=12, render_points_as_spheres=True)
		
		# 近位点のラベルを表示
		prox_label_actors = []
		for i, (pt, label) in enumerate(zip(prox_points, prox_labels)):
			prox_label_actor = all_plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.prox_color, bold=True, shadow=True, show_points=False)
			prox_label_actors.append(prox_label_actor)
		
		# 遠位特徴点の表示
		dist_points_actor = all_plotter.add_points(dist_points, color=self.dist_color, point_size=12, render_points_as_spheres=True)
		
		# 遠位点のラベルを表示
		dist_label_actors = []
		for i, (pt, label) in enumerate(zip(dist_points, dist_labels)):
			dist_label_actor = all_plotter.add_point_labels([pt], [label], point_size=0, font_size=12, text_color=self.dist_color, bold=True, shadow=True, show_points=False)
			dist_label_actors.append(dist_label_actor)
		
		# 近位座標系の表示
		prox_axis_actors = []
		if prox_origin is not None and prox_x_axis is not None and prox_y_axis is not None and prox_z_axis is not None:
			axis_length = 50.0
			# X軸（赤）
			x_line = pv.Line(prox_origin, prox_origin + prox_x_axis * axis_length)
			x_actor = all_plotter.add_mesh(x_line, color="red", line_width=3)
			prox_axis_actors.append(x_actor)
			# Y軸（緑）
			y_line = pv.Line(prox_origin, prox_origin + prox_y_axis * axis_length)
			y_actor = all_plotter.add_mesh(y_line, color="green", line_width=3)
			prox_axis_actors.append(y_actor)
			# Z軸（青）
			z_line = pv.Line(prox_origin, prox_origin + prox_z_axis * axis_length)
			z_actor = all_plotter.add_mesh(z_line, color="blue", line_width=3)
			prox_axis_actors.append(z_actor)

		# 【追加】近位A-O線とO-遠位A線の角度表示
		# 動的に更新するために変数を初期化
		dist_a_initial = None
		line_dist_actor = None
		line_dist_source = None # PyVistaのラインソース
		angle_text_actor = None
		prox_a_static = None # 固定
		prox_o_static = np.array([0., 0., 0.])
		
		# オプションが有効な場合のみ実行
		if self.show_ao_angle.get():
			try:
				# 点Aの検索（大文字小文字区別なし）
				prox_a_idx = -1
				for i, lbl in enumerate(prox_labels):
					if lbl.strip().upper() == 'A':
						prox_a_idx = i
						break
				
				dist_a_idx = -1
				for i, lbl in enumerate(dist_labels):
					if lbl.strip().upper() == 'A':
						dist_a_idx = i
						break
				
				if prox_a_idx != -1 and dist_a_idx != -1:
					pt_prox_a = prox_points[prox_a_idx]
					pt_dist_a = dist_points[dist_a_idx]
					origin_pt = prox_o_static
					
					# 状態保存（スライダー更新用）
					prox_a_static = pt_prox_a
					dist_a_initial = pt_dist_a.copy() # 初期位置（ワールド変換後、変形前）
					
					# ベクトル計算: A_prox -> O
					vec_prox_ao = origin_pt - pt_prox_a
					norm_prox = np.linalg.norm(vec_prox_ao)
					
					# ベクトル計算: O -> A_dist
					vec_dist_oa = pt_dist_a - origin_pt
					norm_dist = np.linalg.norm(vec_dist_oa)
					
					if norm_prox > 1e-6 and norm_dist > 1e-6:
						# 角度計算 (cos theta)
						cos_theta = np.dot(vec_prox_ao, vec_dist_oa) / (norm_prox * norm_dist)
						cos_theta = np.clip(cos_theta, -1.0, 1.0)
						angle_deg = np.degrees(np.arccos(cos_theta))
						
						# 線分の可視化
						# Prox A -> O (マゼンタ)
						line_prox = pv.Line(pt_prox_a, origin_pt)
						all_plotter.add_mesh(line_prox, color="magenta", line_width=4, label="Prox A-O")
						
						# O -> Dist A (シアン)
						# ここでLineオブジェクトを保持しておき、後でpointsを更新する
						line_dist_source = pv.Line(origin_pt, pt_dist_a)
						line_dist_actor = all_plotter.add_mesh(line_dist_source, color="cyan", line_width=4, label="O-Dist A")
						
						# テキスト表示
						info_text = (
							f"Angle: {angle_deg:.1f} deg\n"
							f"(Prox A-O vs O-Dist A)"
						)
						
						angle_text_actor = all_plotter.add_text(
							info_text,
							position=(10, window_height - 150),
							font_size=12,
							color="black",
							shadow=True
						)
						print(f"[Angle Analysis] {info_text.replace(chr(10), ', ')}")

			except Exception as e:
				print(f"角度計算エラー: {e}")

		
		# 遠位座標系の表示

		
		# 遠位座標系の表示
		dist_axis_actors = []
		# ※ここでは追加しない（後述のスライダー用コードで追加し、重複を避ける）

		# 同次変換行列用のスライダー（遠位モデルの変換）
		# 遠位メッシュの変換後のコピーを保持（スライダーの基準）
		dist_mesh_for_slider = dist_mesh.copy()
		
		# 遠位座標系の初期値を保存（変換後の値）
		dist_origin_initial = dist_origin.copy() if dist_origin is not None else None
		dist_x_axis_initial = dist_x_axis.copy() if dist_x_axis is not None else None
		dist_y_axis_initial = dist_y_axis.copy() if dist_y_axis is not None else None
		dist_z_axis_initial = dist_z_axis.copy() if dist_z_axis is not None else None
		
		# 遠位座標系の軸ラインを保持（動的更新用）
		# 近位軸に合わせて長さと色を統一
		DIST_AXIS_LENGTH = 50.0
		dist_x_line = None
		dist_y_line = None
		dist_z_line = None
		if dist_origin is not None and dist_x_axis is not None and dist_y_axis is not None and dist_z_axis is not None:
			dist_x_line = pv.Line(dist_origin, dist_origin + dist_x_axis * DIST_AXIS_LENGTH)
			dist_y_line = pv.Line(dist_origin, dist_origin + dist_y_axis * DIST_AXIS_LENGTH)
			dist_z_line = pv.Line(dist_origin, dist_origin + dist_z_axis * DIST_AXIS_LENGTH)

			# 遠位座標系の軸をプロッターに追加（スライダーで動的に更新される）
			# 近位と同じ配色・太さに統一
			dx_actor = all_plotter.add_mesh(dist_x_line, color='red',   line_width=3, label='Dist X')
			dy_actor = all_plotter.add_mesh(dist_y_line, color='green', line_width=3, label='Dist Y')
			dz_actor = all_plotter.add_mesh(dist_z_line, color='blue',  line_width=3, label='Dist Z')
			
			# トグル制御用にアクターをリストに追加
			dist_axis_actors.append(dx_actor)
			dist_axis_actors.append(dy_actor)
			dist_axis_actors.append(dz_actor)
		
		# チェックボックスUI（BooleanVar に保存して再起動後も復元）
		# 重要: VTKコールバック内で SetVisibility を直接呼ぶと、VTK内部の状態と
		# 競合して Python が segfault することがある（try/except では捕えられない）。
		# self.after(0, ...) で Tk のメインループに退避してから実行する。
		def _safe_set_visibility(actors, state):
			for a in actors:
				if a is None:
					continue
				try:
					a.SetVisibility(bool(state))
				except Exception as e:
					print(f"[viz toggle] SetVisibility failed: {e}")

		# VTKコールバックからTk操作を避けるため、可視化変更はキューに積んで
		# Tkのrender_loop内でまとめて適用する。
		pending_visibility = {
			"prox_model": None,
			"prox_pp": None,
			"prox_axes": None,
			"dist_model": None,
			"dist_pp": None,
			"dist_axes": None,
		}

		def _queue_visibility(key, state):
			try:
				pending_visibility[key] = bool(state)
			except Exception as e:
				print(f"[viz toggle queue] {key}: {e}")

		def toggle_prox_model(state):
			_queue_visibility("prox_model", state)

		def toggle_prox_points(state):
			_queue_visibility("prox_pp", state)

		def toggle_dist_model(state):
			_queue_visibility("dist_model", state)

		def toggle_dist_points(state):
			_queue_visibility("dist_pp", state)

		def toggle_prox_axes(state):
			_queue_visibility("prox_axes", state)

		def toggle_dist_axes(state):
			_queue_visibility("dist_axes", state)

		# 起動時の保存済み表示状態の反映は show() 完了後に遅延適用する。
		# （init 中の SetVisibility が VTK 内部の初期化と競合してクラッシュする
		#  可能性があるため、show() で完全に立ち上がってから適用）
		print(f"[viz init] prox_model={self.viz_show_prox_model.get()} "
		      f"prox_pp={self.viz_show_prox_pp.get()} prox_axes={self.viz_show_prox_axes.get()} "
		      f"dist_model={self.viz_show_dist_model.get()} dist_pp={self.viz_show_dist_pp.get()} "
		      f"dist_axes={self.viz_show_dist_axes.get()}")
		print(f"[viz init] actors: prox_mesh={prox_mesh_actor is not None} "
		      f"prox_pts={prox_points_actor is not None} prox_axes_n={len(prox_axis_actors)} "
		      f"dist_mesh={dist_mesh_actor is not None} dist_pts={dist_points_actor is not None} "
		      f"dist_axes_n={len(dist_axis_actors)}")

		def _apply_saved_viz_state():
			"""保存済みの表示/非表示状態を実 actor に適用（show() 後に呼ぶ）。"""
			_safe_set_visibility([prox_mesh_actor], self.viz_show_prox_model.get())
			_safe_set_visibility([prox_points_actor] + list(prox_label_actors), self.viz_show_prox_pp.get())
			_safe_set_visibility(prox_axis_actors, self.viz_show_prox_axes.get())
			_safe_set_visibility([dist_mesh_actor], self.viz_show_dist_model.get())
			_safe_set_visibility([dist_points_actor] + list(dist_label_actors), self.viz_show_dist_pp.get())
			_safe_set_visibility(dist_axis_actors, self.viz_show_dist_axes.get())
		
		def update_prox_opacity(value):
			prox_mesh_actor.GetProperty().SetOpacity(value)
		
		def update_dist_opacity(value):
			dist_mesh_actor.GetProperty().SetOpacity(value)
		
		# チェックボックスとラベル - 近位（設定したカラーを使用）
		all_plotter.add_text("Prox Model", position=(10, 10), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_prox_model, value=self.viz_show_prox_model.get(), position=(130, 12), size=20, border_size=1, color_on=self.prox_color, color_off="white")

		all_plotter.add_text("Prox PP", position=(10, 40), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_prox_points, value=self.viz_show_prox_pp.get(), position=(130, 42), size=20, border_size=1, color_on=self.prox_color, color_off="white")

		prox_y_offset = 70
		if prox_axis_actors:
			all_plotter.add_text("Prox Axes", position=(10, prox_y_offset), font_size=10, color="black")
			all_plotter.add_checkbox_button_widget(toggle_prox_axes, value=self.viz_show_prox_axes.get(), position=(130, prox_y_offset+2), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			prox_y_offset += 30

		# チェックボックスとラベル - 遠位（設定したカラーを使用）
		all_plotter.add_text("Dist Model", position=(10, prox_y_offset), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_dist_model, value=self.viz_show_dist_model.get(), position=(130, prox_y_offset+2), size=20, border_size=1, color_on=self.dist_color, color_off="white")

		all_plotter.add_text("Dist PP", position=(10, prox_y_offset+30), font_size=10, color="black")
		all_plotter.add_checkbox_button_widget(toggle_dist_points, value=self.viz_show_dist_pp.get(), position=(130, prox_y_offset+32), size=20, border_size=1, color_on=self.dist_color, color_off="white")

		if dist_axis_actors:
			all_plotter.add_text("Dist Axes", position=(10, prox_y_offset+60), font_size=10, color="black")
			all_plotter.add_checkbox_button_widget(toggle_dist_axes, value=self.viz_show_dist_axes.get(), position=(130, prox_y_offset+62), size=20, border_size=1, color_on=self.dist_color, color_off="white")

		# 透明度スライダー - 近位（設定したカラーを使用）
		all_plotter.add_slider_widget(
			update_prox_opacity,
			rng=[0.0, 1.0],
			value=1.0,
			title=" ",
			pointa=(0.08, 0.025),
			pointb=(0.18, 0.025),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color=self.prox_color,
		)
		
		# 透明度スライダー - 遠位（設定したカラーを使用）
		all_plotter.add_slider_widget(
			update_dist_opacity,
			rng=[0.0, 1.0],
			value=1.0,
			title=" ",
			pointa=(0.08, 0.085),
			pointb=(0.18, 0.085),
			style='modern',
			tube_width=0.01,
			slider_width=0.02,
			color=self.dist_color,
		)
		
		
		# 変換パラメータの初期値
		transform_params = {
			'FE': 0.0,  # Flexion/Extension (Z軸回転)
			'VV': 0.0,  # Varus/Valgus (X軸回転)
			'IE': 0.0,  # Internal/External rotation (Y軸回転)
			'ML': 0.0,  # Medial/Lateral translation
			'AP': 0.0,  # Anterior/Posterior translation
			'PD': 0.0,  # Proximal/Distal translation
		}

		# 靭帯リスト（apply_transformで終点を追従させるためここで宣言）
		# 各要素: {'name', 'prox_point' (ワールド固定), 'dist_point_local' (脛骨ローカル),
		#         'thickness', 'color', 'line_source', 'actor', 'label_actor'}
		ligaments = []
		
		def apply_transform():
			"""現在のパラメータで遠位モデルと座標系を変換"""
			# 同次変換行列を構築（ワールド座標系基準）
			# 膝関節 (ISB): FE=X, VV=Y, IE=Z 回転、ML/AP/PD は X/Y/Z 並進（連成なし）
			# 股関節 (既存): FE=Z, VV=X, IE=Y 回転、並進は連成式
			if knee_mode:
				matrix = self._build_transform_matrix_knee_isb(
					fe=transform_params['FE'],
					vv=transform_params['VV'],
					ie=transform_params['IE'],
					ml=transform_params['ML'],
					ap=transform_params['AP'],
					pd=transform_params['PD'],
				)
			else:
				matrix = self._build_transform_matrix(
					rz=transform_params['FE'],
					rx=transform_params['VV'],
					ry=transform_params['IE'],
					ml=transform_params['ML'],
					ap=transform_params['AP'],
					pd=transform_params['PD']
				)
			
			# 遠位メッシュを変換
			# dist_mesh_for_sliderは遠位座標系の初期位置からの相対位置を保持
			# 1. 遠位座標系の原点を中心に配置
			if dist_origin_initial is not None:
				centered_points = dist_mesh_for_slider.points.copy() - dist_origin_initial
			else:
				centered_points = dist_mesh_for_slider.points.copy()
			
			# 2. 同次座標に変換
			ones = np.ones((centered_points.shape[0], 1))
			points_homogeneous = np.hstack([centered_points, ones])
			
			# 3. 同次変換行列を適用
			transformed_homogeneous = (matrix @ points_homogeneous.T).T
			transformed_points = transformed_homogeneous[:, :3]
			
			# 4. 原点に戻す
			if dist_origin_initial is not None:
				transformed_points = transformed_points + dist_origin_initial
			
			# 5. メッシュの頂点を更新
			dist_mesh.points = transformed_points
			
			# 【追加】角度線の動的更新
			if (dist_a_initial is not None and line_dist_source is not None and 
				angle_text_actor is not None and prox_a_static is not None):
				
				# Dist A点にも同じ変換を適用
				# dist_a_initialはワールド同次変換適用後の初期位置。
				# ここではapply_transformと同じロジックで変換する
				
				# 1. 中心化（dist_origin_initialは[0,0,0]なのでそのまま）
				pt_centered = dist_a_initial
				
				# 2. 同次座標 & 変換
				pt_homo = np.append(pt_centered, 1.0)
				pt_transformed = (matrix @ pt_homo)[:3]
				
				# 3. 原点戻し（+0 なのでそのまま）
				new_dist_a = pt_transformed
				
				# ライン更新: O(0,0,0) -> New Dist A
				line_dist_source.points = np.array([prox_o_static, new_dist_a])
				
				# 角度再計算
				vec_prox_ao = prox_o_static - prox_a_static
				vec_dist_oa = new_dist_a - prox_o_static
				
				n_p = np.linalg.norm(vec_prox_ao)
				n_d = np.linalg.norm(vec_dist_oa)
				
				if n_p > 1e-6 and n_d > 1e-6:
					cos_t = np.dot(vec_prox_ao, vec_dist_oa) / (n_p * n_d)
					cos_t = np.clip(cos_t, -1.0, 1.0)
					deg = np.degrees(np.arccos(cos_t))
					
					# テキスト更新
					new_text = (
						f"Angle: {deg:.1f} deg\n"
						f"(Prox A-O vs O-Dist A)"
					)
					angle_text_actor.SetInput(new_text)

			# 遠位座標系の軸を更新
			if (dist_origin_initial is not None and dist_x_axis_initial is not None and 
				dist_y_axis_initial is not None and dist_z_axis_initial is not None and
				dist_x_line is not None and dist_y_line is not None and dist_z_line is not None):
				
				# 遠位座標系の原点を変換
				origin_centered = np.array([0, 0, 0, 1])
				transformed_origin = (matrix @ origin_centered)[:3]
				transformed_origin = transformed_origin + dist_origin_initial
				
				# 遠位座標系の軸を変換
				rotation_matrix = matrix[:3, :3]
				transformed_x_axis = rotation_matrix @ dist_x_axis_initial
				transformed_y_axis = rotation_matrix @ dist_y_axis_initial
				transformed_z_axis = rotation_matrix @ dist_z_axis_initial
				
				axis_length = 50.0  # 近位と同じ長さに統一
				# X軸の更新
				dist_x_line.points = np.array([transformed_origin, transformed_origin + transformed_x_axis * axis_length])
				# Y軸の更新
				dist_y_line.points = np.array([transformed_origin, transformed_origin + transformed_y_axis * axis_length])
				# Z軸の更新
				dist_z_line.points = np.array([transformed_origin, transformed_origin + transformed_z_axis * axis_length])

			# 靭帯チューブの遠位端を追従させる（遠位側はローカル→現在表示位置に変換）
			# Tube は pv.Line + tube() の polydata を毎フレーム再生成して mapper 差し替え。
			if ligaments and dist_origin_initial is not None:
				for lig in ligaments:
					try:
						local_pt = lig['dist_point_local']
						centered = local_pt - dist_origin_initial
						centered_homo = np.append(centered, 1.0)
						transformed_homo = matrix @ centered_homo
						dist_pt_world = transformed_homo[:3] + dist_origin_initial
						tube_pd = _build_tube_polydata(lig['prox_point'], dist_pt_world, lig['thickness'])
						if lig.get('actor') is not None:
							lig['actor'].GetMapper().SetInputData(tube_pd)
						if lig.get('label_actor') is not None:
							mid = (lig['prox_point'] + dist_pt_world) / 2.0
							try:
								lig['label_actor'].SetPosition(*mid)
							except Exception:
								pass
					except Exception:
						pass

		# スライダー関数（角度用）
		def update_FE(value):
			transform_params['FE'] = value
			apply_transform()
		
		def update_VV(value):
			transform_params['VV'] = value
			apply_transform()
		
		def update_IE(value):
			transform_params['IE'] = value
			apply_transform()
		
		# スライダー関数（変位用）
		def update_ML(value):
			transform_params['ML'] = value
			apply_transform()
		
		def update_AP(value):
			transform_params['AP'] = value
			apply_transform()
		
		def update_PD(value):
			transform_params['PD'] = value
			apply_transform()
		
		# スライダーを追加（右側に配置）
		slider_x_start = 0.75
		slider_x_end = 0.95
		
		# スライダーウィジェットの参照を保持
		slider_widgets = {}
		
		# event_type='always' でドラッグ中もコールバックが連続発火 → 3Dモデルがリアルタイム追従
		# 古いPyVistaでは event_type 引数が無いため、未対応バージョン用のフォールバックを用意
		def _make_slider(callback, rng, title, posy, color):
			kwargs = dict(
				rng=rng, value=0.0, title=title,
				pointa=(slider_x_start, posy), pointb=(slider_x_end, posy),
				style='modern', tube_width=0.005, slider_width=0.015, color=color,
			)
			try:
				return all_plotter.add_slider_widget(callback, event_type='always', **kwargs)
			except TypeError:
				return all_plotter.add_slider_widget(callback, **kwargs)

		# 角度スライダー（度）
		slider_widgets['FE'] = _make_slider(update_FE, [-180.0, 180.0], "FE", 0.85, "blue")
		slider_widgets['VV'] = _make_slider(update_VV, [-180.0, 180.0], "VV", 0.75, "red")
		slider_widgets['IE'] = _make_slider(update_IE, [-180.0, 180.0], "IE", 0.65, "green")

		# 変位スライダー（mm）
		slider_widgets['ML'] = _make_slider(update_ML, [-50.0, 50.0], "ML", 0.50, "blue")
		slider_widgets['AP'] = _make_slider(update_AP, [-50.0, 50.0], "AP", 0.40, "red")
		slider_widgets['PD'] = _make_slider(update_PD, [-50.0, 50.0], "PD", 0.30, "green")
		
		# Resetボタンの追加
		def reset_all():
			"""全てのパラメータを0にリセット"""
			# パラメータを0に設定
			transform_params['FE'] = 0.0
			transform_params['VV'] = 0.0
			transform_params['IE'] = 0.0
			transform_params['ML'] = 0.0
			transform_params['AP'] = 0.0
			transform_params['PD'] = 0.0
			
			# スライダーの値を0に更新
			for key, widget in slider_widgets.items():
				# PyVistaのスライダーウィジェットの値を更新
				widget.GetRepresentation().SetValue(0.0)
			
			# 変換を適用
			apply_transform()
		
		# Resetボタンを追加
		all_plotter.add_text(
			"Reset",
			position=(slider_x_start * all_plotter.window_size[0], 0.15 * all_plotter.window_size[1]),
			font_size=12,
			color="white",
			name="reset_button"
		)
		
		# ボタンクリックのためのキーボードショートカット（'r'キー）
		all_plotter.add_key_event('r', reset_all)
		all_plotter.add_key_event('R', reset_all)

		# ===== 自動再生コントロールウィンドウ（Tkinter） =====
		# 6DOF それぞれに: 再生/停止トグル、方向 (+/-)、速度 (任意入力) を提供。
		# 範囲端 (FE/VV/IE は ±180°、ML/AP/PD は ±50mm) に達したら反転（往復運動）。
		# 値は20msごとに増分し、PyVistaウィジェットと apply_transform に同期反映する。

		DOF_CONFIG = {
			'FE': {'rng': (-180.0, 180.0), 'default_speed': 30.0, 'unit': '°/s', 'kind': 'rot'},
			'VV': {'rng': (-180.0, 180.0), 'default_speed': 30.0, 'unit': '°/s', 'kind': 'rot'},
			'IE': {'rng': (-180.0, 180.0), 'default_speed': 30.0, 'unit': '°/s', 'kind': 'rot'},
			'ML': {'rng': (-50.0, 50.0),   'default_speed': 10.0, 'unit': 'mm/s', 'kind': 'trans'},
			'AP': {'rng': (-50.0, 50.0),   'default_speed': 10.0, 'unit': 'mm/s', 'kind': 'trans'},
			'PD': {'rng': (-50.0, 50.0),   'default_speed': 10.0, 'unit': 'mm/s', 'kind': 'trans'},
		}
		INTERVAL_MS = 20  # ~50Hz 更新

		auto_state = {}
		for dof in DOF_CONFIG:
			auto_state[dof] = {'playing': False, 'direction': 1, 'after_id': None}

		ctrl_window = tk.Toplevel(self)
		ctrl_window.title("自動再生コントロール" + (" [膝関節 ISB]" if knee_mode else " [股関節]"))
		ctrl_window.geometry("680x780")
		ctrl_window.resizable(True, True)
		ctrl_window.attributes('-topmost', True)

		# スクロール可能なコンテナ
		ctrl_canvas = tk.Canvas(ctrl_window, highlightthickness=0)
		ctrl_scrollbar = ttk.Scrollbar(ctrl_window, orient="vertical", command=ctrl_canvas.yview)
		ctrl_scroll_frame = ttk.Frame(ctrl_canvas)
		ctrl_scroll_frame.bind("<Configure>", lambda e: ctrl_canvas.configure(scrollregion=ctrl_canvas.bbox("all")))
		ctrl_canvas.create_window((0, 0), window=ctrl_scroll_frame, anchor="nw")
		ctrl_canvas.configure(yscrollcommand=ctrl_scrollbar.set)
		ctrl_canvas.pack(side="left", fill="both", expand=True)
		ctrl_scrollbar.pack(side="right", fill="y")
		def _on_ctrl_mousewheel(event):
			try:
				step = _mousewheel_units(event)
				if step:
					ctrl_canvas.yview_scroll(step, "units")
			except Exception:
				pass
		ctrl_canvas.bind_all("<MouseWheel>", _on_ctrl_mousewheel)
		ctrl_canvas.bind_all("<Button-4>", _on_ctrl_mousewheel)
		ctrl_canvas.bind_all("<Button-5>", _on_ctrl_mousewheel)

		# 個別自動再生セクション (LabelFrame + 単一grid で揃える)
		auto_frame = ttk.LabelFrame(ctrl_scroll_frame, text="個別自動再生")
		auto_frame.pack(fill="x", padx=8, pady=(8, 4))
		# 列幅は uniform="auto_col" + 固定widthで揃える
		hdr_font = (self.ui_font_family, 9, "bold")
		ttk.Label(auto_frame, text="DOF",   width=6,  font=hdr_font).grid(row=0, column=0, padx=4, pady=(4, 2), sticky="w")
		ttk.Label(auto_frame, text="現在値", width=10, font=hdr_font).grid(row=0, column=1, padx=4, pady=(4, 2), sticky="w")
		ttk.Label(auto_frame, text="方向",   width=6,  font=hdr_font).grid(row=0, column=2, padx=4, pady=(4, 2), sticky="w")
		ttk.Label(auto_frame, text="速度",   width=10, font=hdr_font).grid(row=0, column=3, padx=4, pady=(4, 2), sticky="w")
		ttk.Label(auto_frame, text="単位",   width=8,  font=hdr_font).grid(row=0, column=4, padx=4, pady=(4, 2), sticky="w")
		ttk.Label(auto_frame, text="再生",   width=8,  font=hdr_font).grid(row=0, column=5, padx=4, pady=(4, 2), sticky="w")

		ctrl_widgets = {}  # 各DOFのwidget参照

		def update_value_display(dof):
			try:
				ctrl_widgets[dof]['value_label'].config(text=f"{transform_params[dof]:+.2f}")
			except Exception:
				pass

		def make_animate_step(dof):
			lo, hi = DOF_CONFIG[dof]['rng']
			def animate_step():
				st = auto_state[dof]
				if not st['playing']:
					return
				# 速度を取得（不正値の場合は前回値維持）
				try:
					speed = float(ctrl_widgets[dof]['speed_var'].get())
				except (ValueError, tk.TclError):
					speed = DOF_CONFIG[dof]['default_speed']
				delta = st['direction'] * speed * (INTERVAL_MS / 1000.0)
				new_val = transform_params[dof] + delta
				# 範囲端で反転（往復運動）
				if new_val > hi:
					new_val = hi - (new_val - hi)
					st['direction'] = -1
					ctrl_widgets[dof]['dir_btn'].config(text="-")
				elif new_val < lo:
					new_val = lo + (lo - new_val)
					st['direction'] = 1
					ctrl_widgets[dof]['dir_btn'].config(text="+")
				new_val = max(lo, min(hi, new_val))
				transform_params[dof] = new_val
				# PyVistaスライダーウィジェットの値も同期
				try:
					slider_widgets[dof].GetRepresentation().SetValue(new_val)
				except Exception:
					pass
				# 3Dモデルに反映
				try:
					apply_transform()
				except Exception:
					pass
				update_value_display(dof)
				# 次フレームを予約
				st['after_id'] = self.after(INTERVAL_MS, animate_step)
			return animate_step

		def toggle_play(dof):
			st = auto_state[dof]
			st['playing'] = not st['playing']
			if st['playing']:
				ctrl_widgets[dof]['play_btn'].config(text="停止")
				make_animate_step(dof)()
			else:
				ctrl_widgets[dof]['play_btn'].config(text="再生")
				if st['after_id'] is not None:
					try:
						self.after_cancel(st['after_id'])
					except Exception:
						pass
					st['after_id'] = None

		def toggle_direction(dof):
			st = auto_state[dof]
			st['direction'] = -st['direction']
			ctrl_widgets[dof]['dir_btn'].config(text="+" if st['direction'] > 0 else "-")

		for row_idx, dof in enumerate(['FE', 'VV', 'IE', 'ML', 'AP', 'PD'], start=1):
			cfg = DOF_CONFIG[dof]
			ttk.Label(auto_frame, text=dof, width=6).grid(row=row_idx, column=0, padx=4, pady=2, sticky="w")
			value_label = ttk.Label(auto_frame, text="+0.00", width=10)
			value_label.grid(row=row_idx, column=1, padx=4, pady=2, sticky="w")
			dir_btn = ttk.Button(auto_frame, text="+", width=4,
				command=lambda d=dof: toggle_direction(d))
			dir_btn.grid(row=row_idx, column=2, padx=4, pady=2)
			speed_var = tk.DoubleVar(value=cfg['default_speed'])
			speed_entry = ttk.Entry(auto_frame, textvariable=speed_var, width=10)
			speed_entry.grid(row=row_idx, column=3, padx=4, pady=2)
			ttk.Label(auto_frame, text=cfg['unit'], width=8).grid(row=row_idx, column=4, padx=4, pady=2, sticky="w")
			play_btn = ttk.Button(auto_frame, text="再生", width=8,
				command=lambda d=dof: toggle_play(d))
			play_btn.grid(row=row_idx, column=5, padx=4, pady=2)
			ctrl_widgets[dof] = {
				'value_label': value_label,
				'dir_btn': dir_btn,
				'speed_var': speed_var,
				'play_btn': play_btn,
			}

		# 全停止 + リセットボタン
		btm = ttk.Frame(ctrl_scroll_frame)
		btm.pack(fill="x", padx=8, pady=(4, 8))
		def stop_all_autoplay():
			for d in DOF_CONFIG:
				if auto_state[d]['playing']:
					toggle_play(d)
		def reset_from_ctrl():
			stop_all_autoplay()
			reset_all()
			for d in DOF_CONFIG:
				update_value_display(d)
		ttk.Button(btm, text="全停止", command=stop_all_autoplay).pack(side="left", padx=4)
		ttk.Button(btm, text="全リセット (0に戻す)", command=reset_from_ctrl).pack(side="left", padx=4)

		# ===== シーケンス再生 (パターン1〜5) =====
		# 各パターンに DOF・方向・速度・変位を設定。チェックされたものを上から順番に実行する。
		# 各パターンは「現在値から指定変位だけ動かす」累積動作。
		seq_state = {
			'running': False,
			'step_idx': 0,
			'queue': [],         # 実行待ちパターンのリスト
			'accumulated': 0.0,  # 現在ステップで動いた量
			'after_id': None,
		}
		PATTERN_COUNT = 5
		DOF_OPTIONS = ['FE', 'VV', 'IE', 'ML', 'AP', 'PD']

		seq_frame = ttk.LabelFrame(ctrl_scroll_frame, text="シーケンス再生 (パターン1〜5)")
		seq_frame.pack(fill="x", padx=8, pady=(4, 4))

		# ヘッダ行 + 各行を一つの grid に入れる（カラム揃え）
		ttk.Label(seq_frame, text="有効",   width=5,  font=hdr_font).grid(row=0, column=0, padx=4, pady=(4, 2))
		ttk.Label(seq_frame, text="パターン", width=8,  font=hdr_font).grid(row=0, column=1, padx=4, pady=(4, 2))
		ttk.Label(seq_frame, text="DOF",   width=6,  font=hdr_font).grid(row=0, column=2, padx=4, pady=(4, 2))
		ttk.Label(seq_frame, text="方向",   width=5,  font=hdr_font).grid(row=0, column=3, padx=4, pady=(4, 2))
		ttk.Label(seq_frame, text="速度",   width=10, font=hdr_font).grid(row=0, column=4, padx=4, pady=(4, 2))
		ttk.Label(seq_frame, text="変位",   width=10, font=hdr_font).grid(row=0, column=5, padx=4, pady=(4, 2))
		ttk.Label(seq_frame, text="単位 (速度/変位)", width=18, font=hdr_font).grid(row=0, column=6, padx=4, pady=(4, 2))

		pattern_widgets = []  # 各行のtk変数群
		for p_idx in range(1, PATTERN_COUNT + 1):
			enabled_v = tk.BooleanVar(value=False)
			dof_v = tk.StringVar(value='FE')
			dir_v = tk.StringVar(value='+')
			speed_v = tk.DoubleVar(value=10.0)
			disp_v = tk.DoubleVar(value=20.0)
			unit_label = ttk.Label(seq_frame, text="°/s, °", width=18)

			def make_unit_updater(dv, lbl):
				def upd(*_):
					kind = DOF_CONFIG.get(dv.get(), {}).get('kind', 'rot')
					lbl.config(text="°/s, °" if kind == 'rot' else "mm/s, mm")
				return upd

			ttk.Checkbutton(seq_frame, variable=enabled_v).grid(row=p_idx, column=0, padx=4, pady=1)
			ttk.Label(seq_frame, text=f"パターン{p_idx}", width=8).grid(row=p_idx, column=1, padx=4, pady=1, sticky="w")
			dof_cb = ttk.Combobox(seq_frame, textvariable=dof_v, values=DOF_OPTIONS, width=5, state='readonly')
			dof_cb.grid(row=p_idx, column=2, padx=4, pady=1)
			dof_cb.bind('<<ComboboxSelected>>', make_unit_updater(dof_v, unit_label))
			ttk.Combobox(seq_frame, textvariable=dir_v, values=['+', '-'], width=4, state='readonly').grid(row=p_idx, column=3, padx=4, pady=1)
			ttk.Entry(seq_frame, textvariable=speed_v, width=10).grid(row=p_idx, column=4, padx=4, pady=1)
			ttk.Entry(seq_frame, textvariable=disp_v, width=10).grid(row=p_idx, column=5, padx=4, pady=1)
			unit_label.grid(row=p_idx, column=6, padx=4, pady=1, sticky="w")

			pattern_widgets.append({
				'enabled': enabled_v, 'dof': dof_v, 'dir': dir_v,
				'speed': speed_v, 'disp': disp_v, 'unit_label': unit_label,
			})

		# ステータス表示 (シーケンス枠内の最後の行を使う)
		seq_status_label = ttk.Label(seq_frame, text="状態: 停止中", foreground="gray")
		seq_status_label.grid(row=PATTERN_COUNT + 1, column=0, columnspan=7, padx=4, pady=(4, 2), sticky="w")

		# 制御ボタン
		seq_btn_frame = ttk.Frame(seq_frame)
		seq_btn_frame.grid(row=PATTERN_COUNT + 2, column=0, columnspan=7, padx=4, pady=(0, 6), sticky="w")

		def stop_sequence():
			seq_state['running'] = False
			if seq_state['after_id'] is not None:
				try:
					self.after_cancel(seq_state['after_id'])
				except Exception:
					pass
				seq_state['after_id'] = None
			try:
				seq_start_btn.config(text="シーケンス再生開始")
				seq_status_label.config(text="状態: 停止中", foreground="gray")
			except Exception:
				pass

		def sequence_step():
			if not seq_state['running']:
				return
			# 全パターン消化?
			if seq_state['step_idx'] >= len(seq_state['queue']):
				try:
					seq_status_label.config(text="状態: 完了", foreground="green")
				except Exception:
					pass
				stop_sequence()
				return

			current = seq_state['queue'][seq_state['step_idx']]
			dof = current['dof']
			cfg = DOF_CONFIG[dof]
			lo, hi = cfg['rng']

			# 1ステップ分の増分（方向×速度×dt）
			delta_per_step = current['dir'] * current['speed'] * (INTERVAL_MS / 1000.0)
			remaining_disp = current['target_disp'] - seq_state['accumulated']

			if abs(delta_per_step) >= remaining_disp:
				# 最終ステップ: ぴったり目標分だけ動かす
				delta = current['dir'] * remaining_disp
			else:
				delta = delta_per_step

			new_val = transform_params[dof] + delta
			# 範囲外ならクランプ（往復はしない: シーケンスは指定変位を達成して次へ）
			new_val_clamped = max(lo, min(hi, new_val))
			actual_delta = new_val_clamped - transform_params[dof]
			transform_params[dof] = new_val_clamped

			seq_state['accumulated'] += abs(actual_delta)

			# PyVistaスライダーと表示を同期
			try:
				slider_widgets[dof].GetRepresentation().SetValue(new_val_clamped)
			except Exception:
				pass
			try:
				apply_transform()
			except Exception:
				pass
			update_value_display(dof)

			# ステータス更新
			try:
				seq_status_label.config(
					text=f"状態: パターン{current['idx']} 実行中 ({dof} {seq_state['accumulated']:.2f}/{current['target_disp']:.2f})",
					foreground="blue")
			except Exception:
				pass

			# 範囲端で動けなくなった場合、強制的に次へ（無限ループ回避）
			step_done = (seq_state['accumulated'] >= current['target_disp'] - 1e-6) or (abs(actual_delta) < 1e-9)

			if step_done:
				seq_state['step_idx'] += 1
				seq_state['accumulated'] = 0.0

			seq_state['after_id'] = self.after(INTERVAL_MS, sequence_step)

		def start_sequence():
			# 個別自動再生は止めてから開始（競合回避）
			stop_all_autoplay()
			# 有効パターンを集める
			queue = []
			for i, w in enumerate(pattern_widgets, start=1):
				if not w['enabled'].get():
					continue
				try:
					sp = float(w['speed'].get())
					dp = float(w['disp'].get())
				except (ValueError, tk.TclError):
					continue
				# 速度と変位は絶対値で扱う（符号は「方向」コンボボックスが担当）
				sp_abs = abs(sp)
				dp_abs = abs(dp)
				if sp_abs <= 0 or dp_abs <= 0:
					continue
				queue.append({
					'idx': i,
					'dof': w['dof'].get(),
					'dir': 1 if w['dir'].get() == '+' else -1,
					'speed': sp_abs,
					'target_disp': dp_abs,
				})
			if not queue:
				try:
					seq_status_label.config(text="状態: チェック済みパターンが無いか入力値が不正", foreground="red")
				except Exception:
					pass
				return
			seq_state['running'] = True
			seq_state['step_idx'] = 0
			seq_state['queue'] = queue
			seq_state['accumulated'] = 0.0
			try:
				seq_start_btn.config(text="シーケンス停止")
			except Exception:
				pass
			sequence_step()

		def toggle_sequence():
			if seq_state['running']:
				stop_sequence()
			else:
				start_sequence()

		seq_start_btn = ttk.Button(seq_btn_frame, text="シーケンス再生開始", command=toggle_sequence)
		seq_start_btn.pack(side="left", padx=4)

		# ===== プリセット永続化（姿勢 + シーケンス） =====
		presets_data = self._load_presets()
		# 形式: {"poses": [{"name": str, "params": {FE,VV,IE,ML,AP,PD}}, ...],
		#       "sequences": [{"name": str, "patterns": [{...}x5]}, ...]}

		def persist_presets():
			"""現在のpresets_dataをファイルに保存。"""
			self._save_presets(presets_data)

		# ===== 姿勢の保存・呼び出し =====
		pose_frame = ttk.LabelFrame(ctrl_scroll_frame, text="姿勢の保存・呼び出し")
		pose_frame.pack(fill="x", padx=8, pady=(8, 4))

		pose_top = ttk.Frame(pose_frame)
		pose_top.pack(fill="x", padx=4, pady=(4, 4))
		ttk.Label(pose_top, text="現在のFE/VV/IE/ML/AP/PD値を名前付きで保存します。").pack(side="left")

		# 保存済み姿勢のリストUI（行を動的に追加・削除）
		pose_list_frame = ttk.Frame(pose_frame)
		pose_list_frame.pack(fill="x", padx=4, pady=(0, 4))

		pose_row_widgets = []  # 各行の {'frame': ..., 'name_var': ..., 'pose': dict}

		def refresh_pose_list():
			# 既存の行を破棄
			for w in pose_row_widgets:
				try: w['frame'].destroy()
				except Exception: pass
			pose_row_widgets.clear()
			# 再構築
			for idx, pose in enumerate(presets_data.get("poses", [])):
				row = ttk.Frame(pose_list_frame)
				row.pack(fill="x", pady=1)
				name_var = tk.StringVar(value=pose.get("name", f"姿勢{idx+1}"))
				ttk.Entry(row, textvariable=name_var, width=20).pack(side="left", padx=2)
				# 値プレビュー
				p = pose.get("params", {})
				preview = f"FE:{p.get('FE',0):+.1f} VV:{p.get('VV',0):+.1f} IE:{p.get('IE',0):+.1f} | ML:{p.get('ML',0):+.1f} AP:{p.get('AP',0):+.1f} PD:{p.get('PD',0):+.1f}"
				ttk.Label(row, text=preview, foreground="gray", font=(self.ui_font_family, 8)).pack(side="left", padx=8)
				def apply_pose(i=idx):
					stop_all_autoplay(); stop_sequence()
					p2 = presets_data["poses"][i].get("params", {})
					for d in DOF_OPTIONS:
						if d in p2:
							transform_params[d] = float(p2[d])
							try: slider_widgets[d].GetRepresentation().SetValue(transform_params[d])
							except Exception: pass
							update_value_display(d)
					try: apply_transform()
					except Exception: pass
				def delete_pose(i=idx):
					try: presets_data["poses"].pop(i)
					except Exception: pass
					persist_presets()
					refresh_pose_list()
				def rename_pose(*_, i=idx, nv=name_var):
					try: presets_data["poses"][i]["name"] = nv.get()
					except Exception: pass
					persist_presets()
				name_var.trace_add("write", rename_pose)
				ttk.Button(row, text="適用", width=6, command=apply_pose).pack(side="right", padx=2)
				ttk.Button(row, text="削除", width=6, command=delete_pose).pack(side="right", padx=2)
				pose_row_widgets.append({'frame': row, 'name_var': name_var})

		def save_current_pose():
			# 名前のデフォルト
			default_name = f"姿勢{len(presets_data.get('poses', [])) + 1}"
			new_pose = {
				"name": default_name,
				"params": {d: float(transform_params[d]) for d in DOF_OPTIONS},
			}
			presets_data.setdefault("poses", []).append(new_pose)
			persist_presets()
			refresh_pose_list()

		ttk.Button(pose_top, text="現在の姿勢を保存", command=save_current_pose).pack(side="right", padx=2)
		refresh_pose_list()

		# ===== シーケンスプリセット =====
		preset_frame = ttk.LabelFrame(ctrl_scroll_frame, text="シーケンスプリセット")
		preset_frame.pack(fill="x", padx=8, pady=(8, 8))

		preset_top = ttk.Frame(preset_frame)
		preset_top.pack(fill="x", padx=4, pady=(4, 4))
		ttk.Label(preset_top, text="現在のシーケンス（パターン1〜5の内容）を名前付きで保存します。").pack(side="left")

		preset_list_frame = ttk.Frame(preset_frame)
		preset_list_frame.pack(fill="x", padx=4, pady=(0, 4))

		preset_row_widgets = []

		def refresh_preset_list():
			for w in preset_row_widgets:
				try: w['frame'].destroy()
				except Exception: pass
			preset_row_widgets.clear()
			for idx, preset in enumerate(presets_data.get("sequences", [])):
				row = ttk.Frame(preset_list_frame)
				row.pack(fill="x", pady=1)
				name_var = tk.StringVar(value=preset.get("name", f"シーケンス{idx+1}"))
				ttk.Entry(row, textvariable=name_var, width=20).pack(side="left", padx=2)
				patterns = preset.get("patterns", [])
				enabled_count = sum(1 for p in patterns if p.get("enabled"))
				ttk.Label(row, text=f"有効パターン数: {enabled_count}", foreground="gray", font=(self.ui_font_family, 8)).pack(side="left", padx=8)
				def load_preset(i=idx):
					stop_sequence()
					patterns2 = presets_data["sequences"][i].get("patterns", [])
					for j, w_row in enumerate(pattern_widgets):
						if j < len(patterns2):
							p = patterns2[j]
							try:
								w_row['enabled'].set(bool(p.get('enabled', False)))
								w_row['dof'].set(str(p.get('dof', 'FE')))
								w_row['dir'].set(str(p.get('dir', '+')))
								w_row['speed'].set(float(p.get('speed', 10.0)))
								w_row['disp'].set(float(p.get('disp', 20.0)))
								# 単位ラベル更新
								kind = DOF_CONFIG.get(w_row['dof'].get(), {}).get('kind', 'rot')
								w_row['unit_label'].config(text="°/s, °" if kind == 'rot' else "mm/s, mm")
							except Exception:
								pass
						else:
							w_row['enabled'].set(False)
				def delete_preset(i=idx):
					try: presets_data["sequences"].pop(i)
					except Exception: pass
					persist_presets()
					refresh_preset_list()
				def rename_preset(*_, i=idx, nv=name_var):
					try: presets_data["sequences"][i]["name"] = nv.get()
					except Exception: pass
					persist_presets()
				name_var.trace_add("write", rename_preset)
				ttk.Button(row, text="読込", width=6, command=load_preset).pack(side="right", padx=2)
				ttk.Button(row, text="削除", width=6, command=delete_preset).pack(side="right", padx=2)
				preset_row_widgets.append({'frame': row, 'name_var': name_var})

		def save_current_sequence():
			default_name = f"シーケンス{len(presets_data.get('sequences', [])) + 1}"
			patterns_snapshot = []
			for w in pattern_widgets:
				try:
					patterns_snapshot.append({
						'enabled': bool(w['enabled'].get()),
						'dof': str(w['dof'].get()),
						'dir': str(w['dir'].get()),
						'speed': float(w['speed'].get()),
						'disp': float(w['disp'].get()),
					})
				except Exception:
					patterns_snapshot.append({'enabled': False, 'dof': 'FE', 'dir': '+', 'speed': 10.0, 'disp': 20.0})
			presets_data.setdefault("sequences", []).append({
				"name": default_name,
				"patterns": patterns_snapshot,
			})
			persist_presets()
			refresh_preset_list()

		ttk.Button(preset_top, text="現在のシーケンスを保存", command=save_current_sequence).pack(side="right", padx=2)
		refresh_preset_list()

		# ===== 靭帯作成 =====
		# 「追加」→ 近位モデル上で点をクリック → 遠位モデル上で点をクリック → 2点をつなぐラインを描画。
		# 名前・太さ・色は後から変更可能。遠位点は脛骨ローカル座標で保持し、スライダー操作で追従する。
		# 靭帯は presets ファイルに永続化される。
		from tkinter import colorchooser as _colorchooser

		# 関節種別ごとに別キーで保存（モデルが違うと点位置の意味が変わるため）
		lig_save_key = "ligaments_knee" if knee_mode else "ligaments_hip"
		loading_ligaments = [False]  # 復元中フラグ（loopで persist しないように）

		def persist_ligaments():
			if loading_ligaments[0]:
				return
			try:
				presets_data[lig_save_key] = [
					{
						'name': lig['name'],
						'prox_point': lig['prox_point'].tolist(),
						'dist_point_local': lig['dist_point_local'].tolist(),
						'thickness': float(lig['thickness']),
						'color': lig['color'],
						'visible': bool(lig.get('visible', True)),
					}
					for lig in ligaments
				]
				persist_presets()
			except Exception as e:
				print(f"[靭帯永続化] 失敗: {e}")

		lig_frame = ttk.LabelFrame(ctrl_scroll_frame, text="靭帯作成")
		lig_frame.pack(fill="x", padx=8, pady=(8, 8))

		lig_top = ttk.Frame(lig_frame)
		lig_top.pack(fill="x", padx=4, pady=(4, 2))
		lig_status_label = ttk.Label(lig_top, text="状態: 待機中", foreground="gray")
		lig_status_label.pack(side="left", padx=2)

		pick_state = {'mode': 'idle', 'prox_point': None}  # 'idle' / 'waiting_prox' / 'waiting_dist' / 'processing'

		# 右クリックでサーフェスピッキングするVTKオブザーバ
		# 左クリックドラッグはカメラ回転のまま使えるよう、VTKの RightButtonPressEvent を直接フック
		import vtk as _vtk
		lig_cell_picker = _vtk.vtkCellPicker()
		lig_cell_picker.SetTolerance(0.005)
		# 骨・軟骨メッシュのみをピック対象に限定（座標軸・PP点・ラベル等は除外）
		try:
			lig_cell_picker.SetPickFromList(True)
			lig_cell_picker.InitializePickList()
			if prox_mesh_actor is not None:
				lig_cell_picker.AddPickList(prox_mesh_actor)
			if dist_mesh_actor is not None:
				lig_cell_picker.AddPickList(dist_mesh_actor)
			if prox_cart_actor is not None:
				lig_cell_picker.AddPickList(prox_cart_actor)
			if dist_cart_actor is not None:
				lig_cell_picker.AddPickList(dist_cart_actor)
		except Exception as e:
			print(f"[靭帯ピック] PickList 設定失敗: {e}")

		def set_lig_status(text, color="gray"):
			try:
				lig_status_label.config(text=text, foreground=color)
			except Exception:
				pass

		def cancel_pick():
			pick_state['mode'] = 'idle'
			pick_state['prox_point'] = None
			set_lig_status("状態: 待機中", "gray")

		def build_current_matrix():
			"""現在のスライダー値から変換行列を組み立てる。"""
			if knee_mode:
				return self._build_transform_matrix_knee_isb(
					fe=transform_params['FE'], vv=transform_params['VV'], ie=transform_params['IE'],
					ml=transform_params['ML'], ap=transform_params['AP'], pd=transform_params['PD'])
			else:
				return self._build_transform_matrix(
					rz=transform_params['FE'], rx=transform_params['VV'], ry=transform_params['IE'],
					ml=transform_params['ML'], ap=transform_params['AP'], pd=transform_params['PD'])

		def to_dist_local(world_pt):
			"""現在表示中の遠位点(world)を、脛骨ローカル(slider=0時)座標に逆変換。"""
			if dist_origin_initial is None:
				return np.array(world_pt, dtype=float).copy()
			matrix = build_current_matrix()
			try:
				inv = np.linalg.inv(matrix)
			except Exception:
				inv = np.eye(4)
			centered = np.array(world_pt, dtype=float) - dist_origin_initial
			centered_homo = np.append(centered, 1.0)
			local_centered = (inv @ centered_homo)[:3]
			return local_centered + dist_origin_initial

		def _compute_dist_world_for_lig(lig):
			"""靭帯の遠位点 (ローカル) を、現在の変換で表示位置 (ワールド) に。"""
			if dist_origin_initial is None:
				return np.array(lig['dist_point_local'], dtype=float)
			m = build_current_matrix()
			centered = lig['dist_point_local'] - dist_origin_initial
			homo = np.append(centered, 1.0)
			return (m @ homo)[:3] + dist_origin_initial

		def _build_tube_polydata(prox_pt, dist_pt, radius):
			"""2点間のチューブ polydata を生成。半径はワールド単位(mm)。"""
			line = pv.Line(prox_pt, dist_pt, resolution=2)
			# n_sides を大きくすると滑らか。負・ゼロ半径は VTK が嫌うので最小値を保証。
			r = max(0.05, float(radius))
			return line.tube(radius=r, n_sides=14)

		def regenerate_ligament_tube(lig):
			"""指定靭帯のチューブ形状を再生成して actor の入力を差し替える。"""
			dist_world = _compute_dist_world_for_lig(lig)
			tube_pd = _build_tube_polydata(lig['prox_point'], dist_world, lig['thickness'])
			if lig.get('actor') is not None:
				try:
					lig['actor'].GetMapper().SetInputData(tube_pd)
				except Exception:
					pass
			# ラベル位置も中点に追従
			if lig.get('label_actor') is not None:
				mid = (np.array(lig['prox_point']) + dist_world) / 2.0
				try:
					lig['label_actor'].SetPosition(*mid)
				except Exception:
					pass

		def add_ligament(name, prox_pt, dist_local_pt, thickness=1.5, color="#FF8800", visible=True):
			"""靭帯データを構築してプロッタにチューブを追加。

			thickness は ワールド単位 (mm) のチューブ半径。ズームしてもボーンとの
			相対太さが変わらない（pv.Tube は実3D形状）。
			visible: 初期表示状態（False で隠す）。
			"""
			prox_arr = np.array(prox_pt, dtype=float)
			dist_local_arr = np.array(dist_local_pt, dtype=float)
			# 現在の表示位置に遠位点を変換
			if dist_origin_initial is not None:
				matrix = build_current_matrix()
				centered = dist_local_arr - dist_origin_initial
				centered_homo = np.append(centered, 1.0)
				transformed_homo = matrix @ centered_homo
				dist_world = transformed_homo[:3] + dist_origin_initial
			else:
				dist_world = dist_local_arr
			tube_pd = _build_tube_polydata(prox_arr, dist_world, thickness)
			actor = all_plotter.add_mesh(tube_pd, color=color, smooth_shading=True)
			# 中点に名前ラベル
			mid = (prox_arr + dist_world) / 2.0
			label_actor = None
			try:
				label_actor = all_plotter.add_point_labels(
					[mid], [name], point_size=0, font_size=10, text_color=color,
					bold=True, shadow=True, show_points=False)
			except Exception:
				pass
			# 初期可視状態を反映
			try:
				if actor is not None:
					actor.SetVisibility(bool(visible))
				if label_actor is not None:
					label_actor.SetVisibility(bool(visible))
			except Exception:
				pass
			lig = {
				'name': name,
				'prox_point': prox_arr,
				'dist_point_local': dist_local_arr,
				'thickness': float(thickness),
				'color': color,
				'visible': bool(visible),
				'actor': actor,
				'label_actor': label_actor,
			}
			ligaments.append(lig)
			refresh_ligament_list()
			persist_ligaments()

		def toggle_ligament_visibility(idx, visible):
			"""指定靭帯の表示/非表示を切替。"""
			try:
				lig = ligaments[idx]
				lig['visible'] = bool(visible)
				if lig.get('actor') is not None:
					try: lig['actor'].SetVisibility(bool(visible))
					except Exception: pass
				if lig.get('label_actor') is not None:
					try: lig['label_actor'].SetVisibility(bool(visible))
					except Exception: pass
				persist_ligaments()
			except Exception as e:
				print(f"[靭帯表示切替] 失敗: {e}")

		def on_pick_callback(point):
			"""ピックされた3D点を処理する。pick_state['mode']で分岐。

			VTKイベントコールバック内では add_mesh 等が落ちるので、
			レンダラ操作は self.after(...) で次のイベントループに遅延させる。
			"""
			p = np.array(point, dtype=float)
			if pick_state['mode'] == 'waiting_prox':
				pick_state['prox_point'] = p
				pick_state['mode'] = 'waiting_dist'
				self.after(0, lambda: set_lig_status("状態: 遠位上を右クリックで点を追加…", "blue"))
			elif pick_state['mode'] == 'waiting_dist':
				dist_picked = p.copy()
				prox = pick_state['prox_point']
				pick_state['prox_point'] = None
				pick_state['mode'] = 'processing'

				def deferred_create():
					try:
						dist_local = to_dist_local(dist_picked)
						default_name = f"靭帯{len(ligaments) + 1}"
						add_ligament(default_name, prox, dist_local)
						set_lig_status(f"状態: 「{default_name}」を作成しました", "green")
					except Exception as e:
						set_lig_status(f"靭帯作成エラー: {e}", "red")
					finally:
						pick_state['mode'] = 'idle'

				self.after(50, deferred_create)

		def on_right_button_press(obj, event):
			"""右クリック時のVTKオブザーバ。pickingモード中だけ動作する。

			Pick() を VTK の RightButtonPressEvent 内で直接呼ぶと VTK 内部状態の
			再入で落ちることがあるので、クリック座標だけ捕まえて self.after で遅延実行する。
			"""
			if pick_state['mode'] not in ('waiting_prox', 'waiting_dist'):
				return  # ピッキング待機中でなければ無視（カメラ操作はそのまま）
			try:
				x, y = obj.GetEventPosition()
			except Exception:
				return

			def do_pick():
				try:
					if pick_state['mode'] not in ('waiting_prox', 'waiting_dist'):
						return
					renderer = all_plotter.renderer
					lig_cell_picker.Pick(x, y, 0, renderer)
					# PickList で骨・軟骨に限定済み。それ以外はヒットしない。
					if lig_cell_picker.GetCellId() < 0:
						return
					# どのアクターに当たったかチェック（PickList 設定でも念のため確認）
					picked_actor = lig_cell_picker.GetActor()
					allowed = {prox_mesh_actor, dist_mesh_actor, prox_cart_actor, dist_cart_actor}
					if picked_actor not in allowed:
						return  # 想定外アクターは無視
					point = lig_cell_picker.GetPickPosition()
					on_pick_callback(np.array(point, dtype=float))
				except Exception as e:
					print(f"[靭帯ピック] 失敗: {e}")

			# 30ms 遅延でVTKの右クリックイベント処理が完全に戻ったあとに実行
			self.after(30, do_pick)

		# 右クリックオブザーバをセッション中ずっと登録しておく（モード判定で実行制御）
		try:
			all_plotter.iren.add_observer('RightButtonPressEvent', on_right_button_press)
		except Exception as e:
			print(f"[靭帯] 右クリック observer 登録失敗: {e}")

		def begin_pick():
			if pick_state['mode'] != 'idle':
				cancel_pick()
				return
			pick_state['mode'] = 'waiting_prox'
			pick_state['prox_point'] = None
			set_lig_status("状態: 近位上を右クリックで点を追加…", "blue")

		ttk.Button(lig_top, text="追加", command=begin_pick).pack(side="right", padx=2)
		ttk.Button(lig_top, text="キャンセル", command=cancel_pick).pack(side="right", padx=2)

		lig_list_frame = ttk.Frame(lig_frame)
		lig_list_frame.pack(fill="x", padx=4, pady=(2, 6))
		lig_row_widgets = []

		def remove_ligament(idx):
			try:
				lig = ligaments[idx]
				# PyVista actor削除
				if lig.get('actor') is not None:
					try: all_plotter.remove_actor(lig['actor'])
					except Exception: pass
				if lig.get('label_actor') is not None:
					try: all_plotter.remove_actor(lig['label_actor'])
					except Exception: pass
				ligaments.pop(idx)
			except Exception:
				pass
			refresh_ligament_list()
			persist_ligaments()

		def update_ligament_thickness(idx, value):
			try:
				lig = ligaments[idx]
				w = max(0.05, float(value))
				lig['thickness'] = w
				# チューブを再生成 (半径=mm単位、ワールド空間)
				regenerate_ligament_tube(lig)
				persist_ligaments()
			except (ValueError, tk.TclError):
				pass

		def update_ligament_color(idx):
			try:
				lig = ligaments[idx]
				result = _colorchooser.askcolor(color=lig['color'], parent=ctrl_window, title=f"「{lig['name']}」の色を選択")
				if result and result[1]:
					hex_color = result[1]
					lig['color'] = hex_color
					persist_ligaments()
					if lig.get('actor') is not None:
						try:
							# hex → RGB(0-1)
							r = int(hex_color[1:3], 16) / 255.0
							g = int(hex_color[3:5], 16) / 255.0
							b = int(hex_color[5:7], 16) / 255.0
							lig['actor'].GetProperty().SetColor(r, g, b)
						except Exception:
							pass
					# ラベルの色も更新（再生成）
					if lig.get('label_actor') is not None:
						try:
							all_plotter.remove_actor(lig['label_actor'])
						except Exception:
							pass
						try:
							dist_world = _compute_dist_world_for_lig(lig)
							mid_pt = (lig['prox_point'] + dist_world) / 2.0
							lig['label_actor'] = all_plotter.add_point_labels(
								[mid_pt], [lig['name']], point_size=0, font_size=10,
								text_color=hex_color, bold=True, shadow=True, show_points=False)
						except Exception:
							pass
					refresh_ligament_list()
			except Exception as e:
				print(f"[靭帯色] 失敗: {e}")

		def update_ligament_name(idx, name_var):
			try:
				lig = ligaments[idx]
				new_name = name_var.get()
				lig['name'] = new_name
				persist_ligaments()
				# ラベルを再生成
				if lig.get('label_actor') is not None:
					try:
						all_plotter.remove_actor(lig['label_actor'])
					except Exception:
						pass
					try:
						dist_world = _compute_dist_world_for_lig(lig)
						mid_pt = (lig['prox_point'] + dist_world) / 2.0
						lig['label_actor'] = all_plotter.add_point_labels(
							[mid_pt], [new_name], point_size=0, font_size=10,
							text_color=lig['color'], bold=True, shadow=True, show_points=False)
					except Exception:
						pass
			except Exception:
				pass

		def refresh_ligament_list():
			# 既存行を破棄
			for w in lig_row_widgets:
				try: w['frame'].destroy()
				except Exception: pass
			lig_row_widgets.clear()
			# ヘッダ行（先頭に1回だけ）
			if ligaments:
				hdr_row = ttk.Frame(lig_list_frame)
				hdr_row.pack(fill="x", pady=(2, 1))
				ttk.Label(hdr_row, text="表示", width=5, font=hdr_font).pack(side="left", padx=2)
				ttk.Label(hdr_row, text="名前", width=14, font=hdr_font).pack(side="left", padx=2)
				ttk.Label(hdr_row, text="太さ", width=6, font=hdr_font).pack(side="left", padx=2)
				ttk.Label(hdr_row, text="色",   width=6, font=hdr_font).pack(side="left", padx=2)
				lig_row_widgets.append({'frame': hdr_row})
			# 各靭帯行
			for idx, lig in enumerate(ligaments):
				row = ttk.Frame(lig_list_frame)
				row.pack(fill="x", pady=1)

				# 表示/非表示チェックボックス
				vis_var = tk.BooleanVar(value=bool(lig.get('visible', True)))
				vis_chk = ttk.Checkbutton(row, variable=vis_var,
					command=lambda i=idx, vv=vis_var: toggle_ligament_visibility(i, vv.get()))
				vis_chk.pack(side="left", padx=(8, 4))

				name_var = tk.StringVar(value=lig['name'])
				name_entry = ttk.Entry(row, textvariable=name_var, width=14)
				name_entry.pack(side="left", padx=2)
				name_var.trace_add("write", lambda *_, i=idx, nv=name_var: update_ligament_name(i, nv))

				thick_var = tk.DoubleVar(value=lig['thickness'])
				thick_entry = ttk.Entry(row, textvariable=thick_var, width=6)
				thick_entry.pack(side="left", padx=2)
				thick_var.trace_add("write", lambda *_, i=idx, tv=thick_var: update_ligament_thickness(i, tv.get()))

				# カラーボタン（現色を背景に表示）
				color_btn = tk.Button(row, text=" ", bg=lig['color'], width=4,
					command=lambda i=idx: update_ligament_color(i))
				color_btn.pack(side="left", padx=2)

				ttk.Button(row, text="削除", width=6, command=lambda i=idx: remove_ligament(i)).pack(side="right", padx=2)
				lig_row_widgets.append({'frame': row, 'name_var': name_var, 'vis_var': vis_var})

		refresh_ligament_list()

		# 保存済み靭帯の復元（関節種別ごと）
		saved_ligs = presets_data.get(lig_save_key, []) or []
		if saved_ligs:
			loading_ligaments[0] = True
			try:
				for sl in saved_ligs:
					try:
						add_ligament(
							name=str(sl.get('name', f"靭帯{len(ligaments)+1}")),
							prox_pt=np.array(sl.get('prox_point', [0,0,0]), dtype=float),
							dist_local_pt=np.array(sl.get('dist_point_local', [0,0,0]), dtype=float),
							thickness=float(sl.get('thickness', 1.5)),
							color=str(sl.get('color', '#FF8800')),
							visible=bool(sl.get('visible', True)),
						)
					except Exception as e:
						print(f"[靭帯復元] スキップ: {e}")
			finally:
				loading_ligaments[0] = False

		# ===== PyVistaウィンドウをノンブロッキング表示 + レンダループ =====
		plotter_alive = [True]

		def on_plotter_exit(*args):
			plotter_alive[0] = False

		try:
			all_plotter.iren.add_observer('ExitEvent', on_plotter_exit)
		except Exception:
			pass

		def on_ctrl_close():
			stop_all_autoplay()
			stop_sequence()
			# 表示状態などをここで永続化
			try:
				self._save_state()
			except Exception as e:
				print(f"[on_ctrl_close save] {e}")
			try:
				ctrl_window.destroy()
			except Exception:
				pass
		ctrl_window.protocol("WM_DELETE_WINDOW", on_ctrl_close)

		try:
			all_plotter.show(auto_close=False, interactive_update=True)
		except TypeError:
			all_plotter.show(auto_close=False)

		# show() 直後に保存済みの表示状態を適用（チェックボックスとactor の状態を一致させる）
		try:
			self.after(300, _apply_saved_viz_state)
		except Exception as e:
			print(f"[viz preset schedule] {e}")

		# レンダループ: PyVista の VTK イベントを Tk の after で駆動
		render_after_id = [None]
		def _apply_pending_visibility():
			"""チェックボックスからの可視化変更を安全に適用。"""
			try:
				if pending_visibility["prox_model"] is not None:
					state = pending_visibility["prox_model"]
					_safe_set_visibility([prox_mesh_actor], state)
					self.viz_show_prox_model.set(state)
					pending_visibility["prox_model"] = None
				if pending_visibility["prox_pp"] is not None:
					state = pending_visibility["prox_pp"]
					_safe_set_visibility([prox_points_actor] + list(prox_label_actors), state)
					self.viz_show_prox_pp.set(state)
					pending_visibility["prox_pp"] = None
				if pending_visibility["prox_axes"] is not None:
					state = pending_visibility["prox_axes"]
					_safe_set_visibility(prox_axis_actors, state)
					self.viz_show_prox_axes.set(state)
					pending_visibility["prox_axes"] = None
				if pending_visibility["dist_model"] is not None:
					state = pending_visibility["dist_model"]
					_safe_set_visibility([dist_mesh_actor], state)
					self.viz_show_dist_model.set(state)
					pending_visibility["dist_model"] = None
				if pending_visibility["dist_pp"] is not None:
					state = pending_visibility["dist_pp"]
					_safe_set_visibility([dist_points_actor] + list(dist_label_actors), state)
					self.viz_show_dist_pp.set(state)
					pending_visibility["dist_pp"] = None
				if pending_visibility["dist_axes"] is not None:
					state = pending_visibility["dist_axes"]
					_safe_set_visibility(dist_axis_actors, state)
					self.viz_show_dist_axes.set(state)
					pending_visibility["dist_axes"] = None
			except Exception as e:
				print(f"[viz toggle apply] {e}")

		def render_loop():
			if not plotter_alive[0]:
				stop_all_autoplay()
				stop_sequence()
				try:
					ctrl_window.destroy()
				except Exception:
					pass
				return
			try:
				_apply_pending_visibility()
				all_plotter.update()
			except Exception:
				plotter_alive[0] = False
				stop_all_autoplay()
				stop_sequence()
				try:
					ctrl_window.destroy()
				except Exception:
					pass
				return
			render_after_id[0] = self.after(30, render_loop)
		render_after_id[0] = self.after(30, render_loop)

	def on_visualize_options(self) -> None:
		"""可視化オプションウィンドウを開く"""
		# 既に開いている場合はフォーカス
		if hasattr(self, '_options_window') and self._options_window is not None and self._options_window.winfo_exists():
			self._options_window.lift()
			self._options_window.focus_force()
			return

		from tkinter import colorchooser

		# 新しいウィンドウを作成
		self._options_window = tk.Toplevel(self)
		options_window = self._options_window
		options_window.title("可視化オプション")
		options_window.geometry("650x750")
		options_window.resizable(True, True)
		options_window.transient(self)  # メインウィンドウの前面に固定

		# タイトル
		title_label = ttk.Label(options_window, text="可視化オプション", font=(self.ui_font_family, 12, "bold"))
		title_label.pack(pady=(10, 20))
		
		# ========== 色設定セクション ==========
		color_section_label = ttk.Label(options_window, text="● 色設定", font=(self.ui_font_family, 10, "bold"))
		color_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# 近位カラー設定
		prox_frame = ttk.Frame(options_window)
		prox_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(prox_frame, text="近位カラー:").pack(side="left")
		prox_color_label = tk.Label(prox_frame, text="　　　", bg=self.prox_color, relief="solid", borderwidth=1, width=10)
		prox_color_label.pack(side="left", padx=10)
		
		def choose_prox_color():
			color = colorchooser.askcolor(initialcolor=self.prox_color, title="近位カラーを選択", parent=options_window)
			options_window.lift()
			options_window.focus_force()
			if color[1]:  # color[1]は16進数カラーコード
				self.prox_color = color[1]
				prox_color_label.config(bg=self.prox_color)
				self._save_state()
		
		ttk.Button(prox_frame, text="変更...", command=choose_prox_color).pack(side="left")
		
		# 遠位カラー設定
		dist_frame = ttk.Frame(options_window)
		dist_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(dist_frame, text="遠位カラー:").pack(side="left")
		dist_color_label = tk.Label(dist_frame, text="　　　", bg=self.dist_color, relief="solid", borderwidth=1, width=10)
		dist_color_label.pack(side="left", padx=10)
		
		def choose_dist_color():
			color = colorchooser.askcolor(initialcolor=self.dist_color, title="遠位カラーを選択", parent=options_window)
			options_window.lift()
			options_window.focus_force()
			if color[1]:  # color[1]は16進数カラーコード
				self.dist_color = color[1]
				dist_color_label.config(bg=self.dist_color)
				self._save_state()
		
		ttk.Button(dist_frame, text="変更...", command=choose_dist_color).pack(side="left")

		
		# オーバーラップ体積カラー設定
		overlap_frame = ttk.Frame(options_window)
		overlap_frame.pack(fill="x", padx=40, pady=5)

		ttk.Label(overlap_frame, text="オーバーラップ体積カラー:").pack(side="left")
		overlap_color_label = tk.Label(overlap_frame, text="　　　", bg=self.overlap_color, relief="solid", borderwidth=1, width=10)
		overlap_color_label.pack(side="left", padx=10)

		def choose_overlap_color():
			color = colorchooser.askcolor(initialcolor=self.overlap_color, title="オーバーラップ体積カラーを選択", parent=options_window)
			options_window.lift()
			options_window.focus_force()
			if color[1]:
				self.overlap_color = color[1]
				overlap_color_label.config(bg=self.overlap_color)
				self._save_state()

		ttk.Button(overlap_frame, text="変更...", command=choose_overlap_color).pack(side="left")
		
		# ========== 関節範囲の半径設定セクション ==========
		ttk.Separator(options_window, orient="horizontal").pack(fill="x", padx=20, pady=15)
		
		radius_section_label = ttk.Label(options_window, text="● 関節範囲の半径（球体抽出用）", font=(self.ui_font_family, 10, "bold"))
		radius_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# 近位側半径設定
		prox_radius_frame = ttk.Frame(options_window)
		prox_radius_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(prox_radius_frame, text="近位側半径 (mm):").pack(side="left")
		prox_radius_entry = ttk.Entry(prox_radius_frame, textvariable=self.prox_radius, width=10)
		prox_radius_entry.pack(side="left", padx=10)
		ttk.Label(prox_radius_frame, text="※寛骨臼領域を抽出", foreground="gray").pack(side="left")
		
		# 遠位側半径設定
		dist_radius_frame = ttk.Frame(options_window)
		dist_radius_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(dist_radius_frame, text="遠位側半径 (mm):").pack(side="left")
		dist_radius_entry = ttk.Entry(dist_radius_frame, textvariable=self.dist_radius, width=10)
		dist_radius_entry.pack(side="left", padx=10)
		ttk.Label(dist_radius_frame, text="※大腿骨頭領域を抽出", foreground="gray").pack(side="left")
		
		# セパレータ
		ttk.Separator(options_window, orient="horizontal").pack(fill="x", padx=20, pady=15)
		
		# O'原点位置オフセット設定
		offset_section_label = ttk.Label(options_window, text="● O'原点位置オフセット（ローカル座標系に沿った移動）", font=(self.ui_font_family, 10, "bold"))
		offset_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# 近位O'オフセット設定
		prox_offset_label = ttk.Label(options_window, text="近位O'オフセット (mm):", font=(self.ui_font_family, 9))
		prox_offset_label.pack(anchor="w", padx=40, pady=(5, 2))
		
		prox_offset_frame = ttk.Frame(options_window)
		prox_offset_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(prox_offset_frame, text="X:").pack(side="left")
		prox_offset_x_entry = ttk.Entry(prox_offset_frame, textvariable=self.prox_offset_x, width=8)
		prox_offset_x_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(prox_offset_frame, text="Y:").pack(side="left")
		prox_offset_y_entry = ttk.Entry(prox_offset_frame, textvariable=self.prox_offset_y, width=8)
		prox_offset_y_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(prox_offset_frame, text="Z:").pack(side="left")
		prox_offset_z_entry = ttk.Entry(prox_offset_frame, textvariable=self.prox_offset_z, width=8)
		prox_offset_z_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(prox_offset_frame, text="※抽出中心 = O + X*x軸 + Y*y軸 + Z*z軸", foreground="gray").pack(side="left")
		
		# 遠位O'オフセット設定
		dist_offset_label = ttk.Label(options_window, text="遠位O'オフセット (mm):", font=(self.ui_font_family, 9))
		dist_offset_label.pack(anchor="w", padx=40, pady=(10, 2))
		
		dist_offset_frame = ttk.Frame(options_window)
		dist_offset_frame.pack(fill="x", padx=40, pady=5)
		
		ttk.Label(dist_offset_frame, text="X:").pack(side="left")
		dist_offset_x_entry = ttk.Entry(dist_offset_frame, textvariable=self.dist_offset_x, width=8)
		dist_offset_x_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(dist_offset_frame, text="Y:").pack(side="left")
		dist_offset_y_entry = ttk.Entry(dist_offset_frame, textvariable=self.dist_offset_y, width=8)
		dist_offset_y_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(dist_offset_frame, text="Z:").pack(side="left")
		dist_offset_z_entry = ttk.Entry(dist_offset_frame, textvariable=self.dist_offset_z, width=8)
		dist_offset_z_entry.pack(side="left", padx=(5, 15))
		
		ttk.Label(dist_offset_frame, text="※抽出中心 = O + X*x軸 + Y*y軸 + Z*z軸", foreground="gray").pack(side="left")

		# セパレータ
		ttk.Separator(options_window, orient="horizontal").pack(fill="x", padx=20, pady=15)
		
		# その他設定
		misc_section_label = ttk.Label(options_window, text="● その他", font=(self.ui_font_family, 10, "bold"))
		misc_section_label.pack(anchor="w", padx=20, pady=(0, 10))
		
		# A-O線表示
		ao_frame = ttk.Frame(options_window)
		ao_frame.pack(fill="x", padx=40, pady=5)
		ttk.Checkbutton(ao_frame, text=self._joint_label("ao_checkbox"), variable=self.show_ao_angle).pack(side="left")
		
		# 保存と閉じるボタン
		button_frame = ttk.Frame(options_window)
		button_frame.pack(pady=20)
		
		def save_and_close():
			self._save_state()
			options_window.destroy()
		
		ttk.Button(button_frame, text="保存して閉じる", command=save_and_close).pack(side="left", padx=5)
		ttk.Button(button_frame, text="閉じる", command=options_window.destroy).pack(side="left", padx=5)

	def on_check_transform(self) -> None:
		"""同次変換行列群の確認: xlsxファイルから変換行列を読み込んで表示"""
		transform_path = self.transform_group_path.get()
		if not transform_path:
			messagebox.showwarning("同次変換行列群の確認", "同次変換行列群ファイルを選択してください。")
			return
		
		try:
			# xlsxファイルを読み込み
			data = self._load_transform_matrices(transform_path)
			if not data:
				messagebox.showwarning("同次変換行列群の確認", "変換行列データを読み込めませんでした。")
				return
			
			# 確認ウィンドウを表示
			self._show_transform_matrices(data)
			
		except Exception as e:
			messagebox.showerror("エラー", f"同次変換行列の読み込みに失敗しました:\n{str(e)}")
	
	def _load_transform_matrices(self, filepath: str):
		"""xlsxまたはKKRファイルから同次変換行列データを読み込む
		
		Args:
			filepath: xlsxまたはKKRファイルのパス
			
		Returns:
			list of dict: [{"time": float, "matrix": np.ndarray(4,4)}, ...]
		"""
		# ファイル拡張子でフォーマットを判定
		file_ext = Path(filepath).suffix.lower()
		if file_ext in ['.kkr']:
			return self._load_transform_matrices_from_kkr(filepath)
		else:
			return self._load_transform_matrices_from_xlsx(filepath)
	
	def _load_transform_matrices_from_xlsx(self, filepath: str):
		"""xlsxファイルから同次変換行列データを読み込む
		
		Args:
			filepath: xlsxファイルのパス
			
		Returns:
			list of dict: [{"time": float, "matrix": np.ndarray(4,4)}, ...]
		"""
		try:
			# data_only=Falseで数式も読み取れるようにする
			wb = openpyxl.load_workbook(filepath, data_only=False)
			sheet = wb.active
			
			# 列のマッピング (AB=FE, AC=ML, AD=VV, AE=AP, AF=IE, AG=PD)
			# AB=28, AC=29, AD=30, AE=31, AF=32, AG=33
			col_indices = {
				"FE": 28,  # Z軸回転
				"ML": 29,  # Z軸変位
				"VV": 30,  # X軸回転
				"AP": 31,  # X軸変位
				"IE": 32,  # Y軸回転
				"PD": 33,  # Y軸変位
			}
			
			data = []
			# 16行目以降を読み取り
			for row_idx in range(16, sheet.max_row + 1):
				# A列からTime値を取得（列番号1）
				time_cell = sheet.cell(row=row_idx, column=1)
				time_val = time_cell.value
				
				if time_val is None:
					continue
				
				# Time値を数値に変換
				try:
					time_val = float(time_val)
				except (TypeError, ValueError):
					continue
				
				# 各列の値を取得
				try:
					fe = float(sheet.cell(row=row_idx, column=col_indices["FE"]).value or 0)  # Z軸回転
					ml = float(sheet.cell(row=row_idx, column=col_indices["ML"]).value or 0)  # Z軸変位
					vv = float(sheet.cell(row=row_idx, column=col_indices["VV"]).value or 0)  # X軸回転
					ap = float(sheet.cell(row=row_idx, column=col_indices["AP"]).value or 0)  # X軸変位
					ie = float(sheet.cell(row=row_idx, column=col_indices["IE"]).value or 0)  # Y軸回転
					pd = float(sheet.cell(row=row_idx, column=col_indices["PD"]).value or 0)  # Y軸変位
				except (TypeError, ValueError):
					continue
				
				# 同次変換行列を構築 (Z-X-Y オイラー角 + 変位)
				matrix = self._build_transform_matrix(fe, vv, ie, ml, ap, pd)
				
				data.append({
					"time": time_val,
					"matrix": matrix,
					"angles": (fe, vv, ie),
					"translations": (ml, ap, pd)
				})
			
			wb.close()
			
			# データを時間順にソート（昇順）
			data.sort(key=lambda x: x["time"])
			
			print(f"[データ読み込み] {len(data)}フレーム読み込み完了")
			if data:
				print(f"[データ読み込み] 時間範囲: {data[0]['time']:.3f}s ~ {data[-1]['time']:.3f}s")
			
			return data
			
		except Exception as e:
			raise Exception(f"xlsx読み込みエラー: {str(e)}")
	
	def _load_transform_matrices_from_kkr(self, filepath: str):
		"""KKRファイルから同次変換行列データを読み込む
		
		Args:
			filepath: KKRファイルのパス
			
		Returns:
			list of dict: [{"time": float, "matrix": np.ndarray(4,4)}, ...]
		"""
		try:
			import csv
			
			if not Path(filepath).exists():
				raise FileNotFoundError(f"ファイルが見つかりません: {filepath}")
				
			data = []
			
			with open(filepath, 'r', encoding='utf-8') as f:
				# メタデータ行をスキップして、実際のデータヘッダーを見つける
				lines = f.readlines()
				header_line_idx = None
				
				for i, line in enumerate(lines):
					if 'Time' in line and 'FE' in line and 'ML' in line:
						header_line_idx = i
						break
				
				if header_line_idx is None:
					raise Exception("KKRファイルにデータヘッダーが見つかりません")
				
				# ヘッダー行からデータ行までを再構築
				data_lines = lines[header_line_idx:]
				reader = csv.DictReader(data_lines)
				
				for row in reader:
					try:
						# Time列を取得
						time_val = float(row['Time'])
						
						# FE, ML, VV, AP, IE, PD列を取得
						fe = float(row['FE'])  # Z軸回転
						ml = float(row['ML'])  # Z軸変位（Medial/Lateral）
						vv = float(row['VV'])  # X軸回転
						ap = float(row['AP'])  # X軸変位（Anterior/Posterior）
						ie = float(row['IE'])  # Y軸回転
						pd = float(row['PD'])  # Y軸変位（Proximal/Distal）
					except (KeyError, TypeError, ValueError) as e:
						# 列が存在しない、または変換できない場合はスキップ
						continue
					
					# 同次変換行列を構築 (Z-X-Y オイラー角 + 変位)
					matrix = self._build_transform_matrix(fe, vv, ie, ml, ap, pd)
					
					data.append({
						"time": time_val,
						"matrix": matrix,
						"angles": (fe, vv, ie),
						"translations": (ml, ap, pd)
					})
			
			# データを時間順にソート（昇順）
			data.sort(key=lambda x: x["time"])
			
			print(f"[データ読み込み] {len(data)}フレーム読み込み完了")
			if data:
				print(f"[データ読み込み] 時間範囲: {data[0]['time']:.3f}s ~ {data[-1]['time']:.3f}s")
			
			return data
			
		except Exception as e:
			raise Exception(f"KKR読み込みエラー: {str(e)}")
	
	def _build_transform_matrix_knee_isb(self, fe: float, vv: float, ie: float,
	                                     ml: float, ap: float, pd: float) -> np.ndarray:
		"""膝関節 ISB / Grood-Suntay 簡略化版の4x4同次変換行列。

		ワールド軸 = 大腿骨座標系 (X=外側, Y=前方, Z=近位) に整列済みのメッシュに対して使う。
		- FE (Flexion/Extension): 大腿骨X軸まわりの回転 [deg]
		- VV (Varus/Valgus):      Y軸まわりの回転 [deg]
		- IE (Internal/External): Z軸まわりの回転 [deg]
		- ML / AP / PD: X / Y / Z 方向の独立並進 [mm]（股関節版のような連成は無い）

		合成順序: R = Rx(FE) @ Ry(VV) @ Rz(IE)。小角度では順序非依存。
		"""
		fe_rad = np.deg2rad(fe)
		vv_rad = np.deg2rad(vv)
		ie_rad = np.deg2rad(ie)

		Rx = np.array([
			[1, 0, 0],
			[0, np.cos(fe_rad), -np.sin(fe_rad)],
			[0, np.sin(fe_rad),  np.cos(fe_rad)],
		])
		Ry = np.array([
			[ np.cos(vv_rad), 0, np.sin(vv_rad)],
			[ 0, 1, 0],
			[-np.sin(vv_rad), 0, np.cos(vv_rad)],
		])
		Rz = np.array([
			[np.cos(ie_rad), -np.sin(ie_rad), 0],
			[np.sin(ie_rad),  np.cos(ie_rad), 0],
			[0, 0, 1],
		])
		R = Rx @ Ry @ Rz

		# 並進は「回転後の脛骨ローカルフレーム」で適用する。
		# こうしないと FE=-90° 等で脛骨が回転した後、ML/AP/PD のスライダーが
		# 関節座標系ではなく世界軸に沿って動いてしまい、AP と PD が入れ替わって
		# 見える等の不一致が生じる。R @ t_local で「脛骨と一緒に回る軸」に並進する。
		t_local = np.array([ml, ap, pd], dtype=float)
		t_world = R @ t_local

		T = np.eye(4, dtype=float)
		T[:3, :3] = R
		T[:3, 3] = t_world
		return T

	def _build_transform_matrix(self, rz: float, rx: float, ry: float,
							  ml: float, ap: float, pd: float) -> np.ndarray:
		"""Z-X-Y オイラー角と変位から4x4同次変換行列を構築
		
		Args:
			rz, rx, ry: Z, X, Y軸周りの回転角度 (度) [FE, VV, IE]
			ml: Medial/Lateral変位 (mm)
			ap: Anterior/Posterior変位 (mm)
			pd: Proximal/Distal変位 (mm)
			
		Returns:
			4x4 同次変換行列
		"""
		# 度をラジアンに変換
		rz_rad = np.deg2rad(rz)
		rx_rad = np.deg2rad(rx)
		ry_rad = np.deg2rad(ry)
		
		# Z軸回転行列
		Rz = np.array([
			[np.cos(rz_rad), -np.sin(rz_rad), 0],
			[np.sin(rz_rad),  np.cos(rz_rad), 0],
			[0,               0,              1]
		])
		
		# X軸回転行列
		Rx = np.array([
			[1, 0,               0],
			[0, np.cos(rx_rad), -np.sin(rx_rad)],
			[0, np.sin(rx_rad),  np.cos(rx_rad)]
		])
		
		# Y軸回転行列
		Ry = np.array([
			[ np.cos(ry_rad), 0, np.sin(ry_rad)],
			[ 0,              1, 0],
			[-np.sin(ry_rad), 0, np.cos(ry_rad)]
		])
		
		# 合成回転行列 (Intrinsic: Y -> X -> Z)
		# R = Rz @ Rx @ Ry
		R = Rz @ Rx @ Ry

		# --- 指定数式による並進ベクトルの計算 ---
		# x = AP * cos(FE) - PD * sin(FE) * cos(VV)
		# y = AP * sin(FE) + PD * cos(FE) * cos(VV)
		# z = ML + PD * sin(VV)
		cf = np.cos(rz_rad)
		sf = np.sin(rz_rad)
		cv = np.cos(rx_rad)
		sv = np.sin(rx_rad)
		x = (ap * cf) - (pd * sf * cv)
		y = (ap * sf) + (pd * cf * cv)
		z = ml + (pd * sv)
		t_vec = np.array([x, y, z], dtype=float)

		# 4x4同次変換行列へセット
		T = np.eye(4, dtype=float)
		T[:3, :3] = R
		T[:3, 3] = t_vec

		return T
	
	def _show_transform_matrices(self, data):
		"""変位・姿勢変化データを表示するウィンドウ"""
		window = tk.Toplevel(self)
		window.title("変位・姿勢変化データの確認")
		window.geometry("600x500")
		
		# スクロール可能なテキストウィジェット
		frame = ttk.Frame(window)
		frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
		
		scrollbar = ttk.Scrollbar(frame)
		scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
		
		text_widget = tk.Text(frame, wrap=tk.NONE, yscrollcommand=scrollbar.set, font=(self.mono_font_family, 10))
		text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		scrollbar.config(command=text_widget.yview)
		
		# ヘッダー行を表示（各時刻に対して 4x4 行列を出力）
		text_widget.insert(tk.END, f"{'Time':>8}    {'Transform Matrix (4x4)'}\n")
		text_widget.insert(tk.END, "" + "=" * 70 + "\n")
		
		# データを整形して表示（各時刻ごとに行列を見やすく表示）
		for item in data:
			time = item["time"]
			matrix = item.get("matrix")
			if matrix is None:
				# 予備: 行列が未計算なら angles/translations から構築
				angles = item.get("angles", (0.0, 0.0, 0.0))
				trans = item.get("translations", (0.0, 0.0, 0.0))
				matrix = self._build_transform_matrix(angles[0], angles[1], angles[2], trans[0], trans[1], trans[2])
			
			# 時刻行
			text_widget.insert(tk.END, f"{time:8.3f}\n")
			# 行列を 3 桁整形で 4 行に分けて表示
			for r in range(4):
				row_vals = [matrix[r, c] for c in range(4)]
				row_str = '  '.join([f"{v:12.6f}" for v in row_vals])
				text_widget.insert(tk.END, f"    {row_str}\n")
			text_widget.insert(tk.END, "\n")
		
		text_widget.config(state=tk.DISABLED)
		
		# 閉じるボタン
		close_btn = ttk.Button(window, text="閉じる", command=window.destroy)
		close_btn.pack(pady=10)

	def _get_cache_hash(self, prox_mesh, dist_mesh, transform_data, simplify_mesh=False):
		"""キャッシュファイル用のハッシュを生成（コンテンツベース）

		ファイルパスではなくデータの中身に基づくため、異なるPC・OSで
		同じデータなら同じハッシュが生成される。

		Args:
			prox_mesh: 近位メッシュ
			dist_mesh: 遠位メッシュ
			transform_data: 変換データ
			simplify_mesh: メッシュ簡略化フラグ

		Returns:
			str: ハッシュ文字列
		"""
		hasher = hashlib.sha256()
		hasher.update(b"v6_content_based")

		# メッシュの頂点座標（ファイルパスではなく実データ）
		hasher.update(np.round(np.array(prox_mesh.points), 2).tobytes())
		hasher.update(np.round(np.array(dist_mesh.points), 2).tobytes())

		# 変換データ（行列の中身）
		hasher.update(str(len(transform_data)).encode())
		for tf in transform_data:
			m = tf['matrix']
			hasher.update(np.round(np.array(m), 6).tobytes())

		# パラメータ（抽出領域・ヒートマップに影響）
		params = (
			f"pr:{self.prox_radius.get()}|dr:{self.dist_radius.get()}|"
			f"po:{self.prox_offset_x.get()}/{self.prox_offset_y.get()}/{self.prox_offset_z.get()}|"
			f"do:{self.dist_offset_x.get()}/{self.dist_offset_y.get()}/{self.dist_offset_z.get()}|"
			f"simplify:{simplify_mesh}"
		)
		hasher.update(params.encode('utf-8'))

		return hasher.hexdigest()[:20]

	def _get_cache_filepath(self, cache_hash):
		"""キャッシュファイルのパスを取得（SharedCacheManager対応）

		SharedCacheManagerが設定されている場合はNAS→ローカルの順に検索。
		設定されていない場合は従来のローカルキャッシュ。

		Args:
			cache_hash: ハッシュ文字列

		Returns:
			Path: キャッシュファイルのパス
		"""
		# SharedCacheManagerが設定されている場合
		if hasattr(self, '_shared_overlap_cache') and self._shared_overlap_cache is not None:
			scm = self._shared_overlap_cache
			if scm.is_nas_available() and scm._nas_dir is not None:
				nas_file = scm._nas_dir / f"overlap_{cache_hash}.pkl"
				return nas_file  # 存在有無にかかわらずNASパスを返す

		# NAS未設定時は git 管理下の共有キャッシュ（cache/overlap/）を使う。
		# 内容ベースのハッシュキーなので、サーバーで計算→git push→他PCでpull すれば
		# 同じデータに対して同じファイルが見つかり、再計算なしで表示できる。
		cache_dir = Path(__file__).parent / "cache" / "overlap"
		cache_dir.mkdir(parents=True, exist_ok=True)
		new_path = cache_dir / f"overlap_{cache_hash}.pkl"
		# 後方互換: 旧 .overlap_cache/ に既存キャッシュがあれば読めるようにする（保存は新フォルダ）
		if not new_path.exists():
			old_path = Path(__file__).parent / ".overlap_cache" / f"overlap_{cache_hash}.pkl"
			if old_path.exists():
				return old_path
		return new_path

	def _save_overlap_cache(self, cache_filepath, overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes=None):
		"""オーバーラップデータとヒートマップをキャッシュファイルに保存
		
		Args:
			cache_filepath: キャッシュファイルのパス
			overlap_meshes: オーバーラップメッシュのリスト
			overlap_areas: オーバーラップ面積のリスト
			overlap_depths: 接触深度のリスト
			heatmap_meshes: ヒートマップメッシュのリスト（オプション）
		"""
		print(f"[CACHE DEBUG] Saving to: {cache_filepath}")
		print(f"[CACHE DEBUG] Overlap entries: {len(overlap_meshes)}")
		if heatmap_meshes:
			print(f"[CACHE DEBUG] Heatmap entries: {len(heatmap_meshes)}")

		try:
			# メッシュデータをシリアライズ可能な形式に変換
			cache_data = {
				'overlap': [],
				'heatmap': [],
				'heatmap_base': None # 最適化用共通ジオメトリ
			}
			
			# オーバーラップデータ
			for mesh, area, depth in zip(overlap_meshes, overlap_areas, overlap_depths):
				if mesh is not None and mesh.n_points > 0:
					cache_data['overlap'].append({
						'points': mesh.points,
						'faces': mesh.faces,
						'area': area,
						'depth': depth
					})
				else:
					cache_data['overlap'].append(None)
			
			# ヒートマップデータ（最適化：共通ジオメトリを抽出）
			if isinstance(heatmap_meshes, _LazyHeatmapList):
				# 遅延リストは base + 距離配列を直接保存（メッシュ再生成なし＝高速・省メモリ）
				base = heatmap_meshes._base
				if base is not None and base.n_points > 0:
					cache_data['heatmap_base'] = {
						'n_points': base.n_points, 'n_faces': base.n_cells,
						'points': base.points, 'faces': base.faces,
					}
					print(f"[キャッシュ] ヒートマップ共通ジオメトリを抽出: {base.n_points}点, {base.n_cells}面")
				for d in heatmap_meshes.distances:
					if d is None:
						cache_data['heatmap'].append(None)
					else:
						cache_data['heatmap'].append({'type': 'ref', 'distance': np.asarray(d, dtype=np.float32)})
			elif heatmap_meshes is not None:
				# ベースメッシュを探す（最初の非空メッシュ）
				base_mesh_info = None
				for mesh in heatmap_meshes:
					if mesh is not None and mesh.n_points > 0:
						base_mesh_info = {
							'n_points': mesh.n_points,
							'n_faces': mesh.n_cells,
							'points': mesh.points,
							'faces': mesh.faces
						}
						cache_data['heatmap_base'] = base_mesh_info
						print(f"[キャッシュ] ヒートマップ共通ジオメトリを抽出: {mesh.n_points}点, {mesh.n_cells}面")
						break
				
				for mesh in heatmap_meshes:
					if mesh is not None and mesh.n_points > 0:
						# ベースメッシュと一致するか確認
						is_same_topology = (base_mesh_info is not None and 
										  mesh.n_points == base_mesh_info['n_points'] and 
										  mesh.n_cells == base_mesh_info['n_faces'])
						
						heatmap_dict = {}
						if is_same_topology:
							# ジオメトリが同じならスカラーのみ保存（タイプ: ref）
							heatmap_dict['type'] = 'ref'
						else:
							# 異なるなら全保存（タイプ: full）
							heatmap_dict['type'] = 'full'
							heatmap_dict['points'] = mesh.points
							heatmap_dict['faces'] = mesh.faces if mesh.faces is not None and len(mesh.faces) > 0 else None
						
						# スカラーデータ（distance）を保存（float32でファイル容量・時間を半減）
						if 'distance' in mesh.array_names:
							heatmap_dict['distance'] = np.asarray(mesh['distance'], dtype=np.float32)

						cache_data['heatmap'].append(heatmap_dict)
					else:
						cache_data['heatmap'].append(None)
			
			# ファイルに保存（アトミック: 一時ファイルへ書いてから置換。中断時の破損を防ぐ）
			tmp_path = Path(str(cache_filepath) + ".tmp")
			with open(tmp_path, 'wb') as f:
				pickle.dump(cache_data, f, protocol=pickle.HIGHEST_PROTOCOL)
			os.replace(str(tmp_path), str(cache_filepath))

			print(f"[キャッシュ] 保存完了: {cache_filepath}")
			if heatmap_meshes is not None:
				print(f"[キャッシュ] ヒートマップデータも保存: {len(heatmap_meshes)}フレーム")

			# SharedCacheManagerが設定されている場合、NASにも同期＋メタデータ保存
			if hasattr(self, '_shared_overlap_cache') and self._shared_overlap_cache is not None:
				scm = self._shared_overlap_cache
				if scm.is_nas_available() and scm._nas_dir is not None:
					try:
						import shutil
						nas_file = scm._nas_dir / cache_filepath.name
						if not nas_file.exists():
							scm._nas_dir.mkdir(parents=True, exist_ok=True)
							shutil.copy2(str(cache_filepath), str(nas_file))
							print(f"[キャッシュ] NASに同期完了: {nas_file}")
					except Exception as e2:
						print(f"[キャッシュ] NAS同期エラー（無視）: {e2}")
					# メタデータ保存
					try:
						cache_hash = cache_filepath.stem.replace('overlap_', '')
						size_mb = round(cache_filepath.stat().st_size / (1024**2), 1) if cache_filepath.exists() else 0
						meta = {
							'type': 'overlap',
							'xlsx': os.path.basename(self.transform_group_path.get()),
							'prox_stl': os.path.basename(self.prox_model_path.get()),
							'dist_stl': os.path.basename(self.dist_model_path.get()),
							'frames': len(overlap_meshes),
							'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
							'size_mb': size_mb,
						}
						scm.save_metadata(cache_hash, meta)
					except Exception as e3:
						print(f"[キャッシュ] メタデータ保存エラー（無視）: {e3}")
		except Exception as e:
			print(f"[キャッシュ] 保存失敗 (詳細): {e}")
			import traceback
			traceback.print_exc()
			messagebox.showerror("キャッシュ保存エラー", f"キャッシュの保存に失敗しました。\n\n詳細: {e}")

	def _load_overlap_cache(self, cache_filepath):
		"""キャッシュファイルからオーバーラップデータとヒートマップを読み込み
		
		Args:
			cache_filepath: キャッシュファイルのパス
			
		Returns:
			tuple: (オーバーラップメッシュのリスト, オーバーラップ面積のリスト, 接触深度のリスト, ヒートマップメッシュのリスト)、またはNone
		"""
		try:
			if not cache_filepath.exists():
				return None
			
			# 進捗ダイアログを作成
			progress_window = tk.Toplevel(self)
			progress_window.title("キャッシュ読み込み中")
			progress_window.geometry("450x150")
			progress_window.resizable(False, False)
			progress_window.attributes('-topmost', True)
			
			ttk.Label(progress_window, text="キャッシュファイルを読み込んでいます...", font=(self.ui_font_family, 11)).pack(pady=10)
			
			progress_label = ttk.Label(progress_window, text="0 / ? フレーム", font=(self.ui_font_family, 10))
			progress_label.pack(pady=5)
			
			progress_bar = ttk.Progressbar(progress_window, mode='determinate', length=400)
			progress_bar.pack(pady=10)
			
			cancel_var = tk.BooleanVar(value=False)
			ttk.Button(progress_window, text="キャンセル", command=lambda: cancel_var.set(True)).pack(pady=5)
			
			progress_window.update()
			
			# ファイルから読み込み
			print(f"[キャッシュ] pickle読み込み開始: {cache_filepath}")
			with open(cache_filepath, 'rb') as f:
				cache_data = pickle.load(f)
			print(f"[キャッシュ] pickle読み込み完了")
			
			# 新形式（辞書型）と旧形式（リスト型）の両方に対応
			if isinstance(cache_data, dict):
				# 新形式: オーバーラップとヒートマップを含む
				overlap_meshes = []
				overlap_areas = []
				overlap_depths = []
				heatmap_meshes = []
				
				# オーバーラップデータを再構築（高速化：最小限の処理）
				overlap_data = cache_data.get('overlap', [])
				for data in overlap_data:
					if data is None:
						overlap_meshes.append(pv.PolyData())
						overlap_areas.append(0.0)
						overlap_depths.append(0.0)
					else:
						# メッシュ作成をスキップ、データだけ保持（必要に応じて後で作成）
						overlap_meshes.append(None)  # プレースホルダー
						overlap_areas.append(data.get('area', 0.0))
						overlap_depths.append(data.get('depth', 0.0))
				
				# ヒートマップデータを再構築（進捗表示付き）
				heatmap_data = cache_data.get('heatmap', [])
				total_frames = len(heatmap_data)
				progress_bar['maximum'] = total_frames
				
				# 共通ジオメトリを事前に復元（最適化対応）
				base_mesh_info = cache_data.get('heatmap_base')
				base_mesh_cache = None
				if base_mesh_info:
					try:
						base_mesh_cache = pv.PolyData(base_mesh_info['points'], base_mesh_info['faces'])
						print(f"[キャッシュ] 共通ジオメトリを復元: {base_mesh_cache.n_points}点")
					except Exception as e:
						print(f"[キャッシュ] 共通ジオメトリ復元失敗: {e}")

				print(f"[キャッシュ] ヒートマップを遅延復元（メッシュは再生時に必要分だけ生成）: {total_frames}フレーム")

				# 全フレーム同一形状(base)で距離のみ異なる典型ケースは、遅延リストで即時復元する
				# （19000超のPolyDataを先に全部作らないので、読込後の待ちが大幅に短縮）。
				all_ref = (base_mesh_cache is not None) and all(
					(d is None) or (d.get('type') == 'ref') for d in heatmap_data)

				if all_ref:
					distances = [None if d is None else d.get('distance') for d in heatmap_data]
					heatmap_meshes = _LazyHeatmapList(base_mesh_cache, distances)
					progress_bar['value'] = total_frames
					progress_label.config(text=f"{total_frames} / {total_frames} フレーム (100%)")
					progress_window.update()
				else:
					# 形状がフレームごとに異なる場合は従来どおり個別復元
					for i, data in enumerate(heatmap_data):
						if cancel_var.get():
							progress_window.destroy()
							print("[キャッシュ] 読み込みがキャンセルされました")
							return None
						if data is None:
							heatmap_meshes.append(pv.PolyData())
						else:
							mesh = None
							if 'type' in data and data['type'] == 'ref' and base_mesh_cache is not None:
								mesh = base_mesh_cache.copy(deep=False)
							else:
								if 'faces' in data and data['faces'] is not None:
									mesh = pv.PolyData(data['points'], data['faces'])
								elif 'points' in data:
									mesh = pv.PolyData(data['points'])
							if mesh:
								if 'distance' in data:
									mesh['distance'] = data['distance']
								heatmap_meshes.append(mesh)
							else:
								heatmap_meshes.append(pv.PolyData())
						if i % 100 == 0 or i == total_frames - 1:
							progress_bar['value'] = i + 1
							progress_label.config(text=f"{i + 1} / {total_frames} フレーム ({(i+1)/total_frames*100:.1f}%)")
							progress_window.update()

				progress_window.destroy()
				
				print(f"[キャッシュ] 読み込み完了: {cache_filepath}")
				if heatmap_meshes:
					print(f"[キャッシュ] ヒートマップデータも読み込み: {len(heatmap_meshes)}フレーム")
				return overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes
			
			else:
				# 旧形式: オーバーラップのみ（後方互換性）
				overlap_meshes = []
				overlap_areas = []
				overlap_depths = []
				
				total_frames = len(cache_data)
				progress_bar['maximum'] = total_frames
				
				for i, data in enumerate(cache_data):
					# キャンセルチェック
					if cancel_var.get():
						progress_window.destroy()
						print("[キャッシュ] 読み込みがキャンセルされました")
						return None
					
					if data is None:
						overlap_meshes.append(None)
						overlap_areas.append(0.0)
						overlap_depths.append(0.0)
					else:
						overlap_meshes.append(None)  # プレースホルダー
						overlap_areas.append(data.get('area', 0.0))
						overlap_depths.append(data.get('depth', 0.0))
					
					# 100フレームごとに進捗更新
					if i % 100 == 0 or i == total_frames - 1:
						progress_bar['value'] = i + 1
						progress_label.config(text=f"{i + 1} / {total_frames} フレーム ({(i+1)/total_frames*100:.1f}%)")
						progress_window.update()
				
				progress_window.destroy()
				
				print(f"[キャッシュ] 読み込み完了（旧形式）: {cache_filepath}")
				return overlap_meshes, overlap_areas, overlap_depths, []
		
		except Exception as e:
			print(f"[キャッシュ] 読み込み失敗: {e}")
			# 進捗ウィンドウを閉じる
			try:
				progress_window.destroy()
			except:
				pass
			return None

	def _precompute_overlap_volumes(self, prox_mesh, dist_mesh, transform_data, progress_callback=None, use_cache=True, use_parallel=True, use_bbox_check=True, simplify_mesh=False):
		"""オーバーラップ体積機能は無効化されています。ヒートマップのみ後段で計算します。

		Returns:
		    tuple: 空の (overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes)
		"""
		print("[事前計算] オーバーラップ体積機能は無効化されています。ヒートマップのみ計算されます。")
		
		# キャッシュが有効な場合、キャッシュファイルの存在を確認
		if use_cache:
			cache_hash = self._get_cache_hash(prox_mesh, dist_mesh, transform_data, simplify_mesh)
			cache_filepath = self._get_cache_filepath(cache_hash)
			
			if cache_filepath.exists():
				# キャッシュファイルのサイズと更新日時を取得
				file_size_mb = cache_filepath.stat().st_size / (1024 * 1024)
				file_mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(cache_filepath.stat().st_mtime))
				
				# ユーザーに確認
				response = messagebox.askyesnocancel(
					"キャッシュファイルが見つかりました",
					f"条件に合うキャッシュファイルが見つかりました。\n\n"
					f"ファイル: {cache_filepath.name}\n"
					f"サイズ: {file_size_mb:.2f} MB\n"
					f"作成日時: {file_mtime}\n"
					f"フレーム数: {len(transform_data)}\n\n"
					f"このキャッシュを使用しますか？\n\n"
					f"「はい」: キャッシュを使用（高速）\n"
					f"「いいえ」: キャッシュを削除して再計算\n"
					f"「キャンセル」: 処理を中断"
				)
				
				if response is None:  # キャンセル
					print("[キャッシュ] 処理がキャンセルされました")
					return [], [], [], []
				elif response:  # はい（キャッシュを使用）
					print(f"[キャッシュ] キャッシュファイルを使用します: {cache_filepath}")
					try:
						result = self._load_overlap_cache(cache_filepath)
						if result:
							overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes = result
							print(f"[キャッシュ] 読み込み完了: {len(overlap_meshes)}フレーム")
							return overlap_meshes, overlap_areas, overlap_depths, heatmap_meshes
						else:
							print("[キャッシュ] 読み込み失敗、再計算します")
					except Exception as e:
						print(f"[キャッシュ] 読み込みエラー: {e}")
						messagebox.showwarning("キャッシュエラー", f"キャッシュの読み込みに失敗しました。\n再計算します。\n\nエラー: {e}")
				else:  # いいえ（キャッシュを削除）
					try:
						cache_filepath.unlink()
						print(f"[キャッシュ] キャッシュファイルを削除しました: {cache_filepath}")
						messagebox.showinfo("キャッシュ削除", "キャッシュファイルを削除しました。\n新しく計算を開始します。")
					except Exception as e:
						print(f"[キャッシュ] 削除エラー: {e}")
						messagebox.showerror("削除エラー", f"キャッシュファイルの削除に失敗しました。\n\nエラー: {e}")
		
		return [], [], [], []

	def _compute_contact_depth(self, prox_mesh, dist_mesh, overlap_mesh=None):
		"""接触深度を計算
		
		接触深度は、2つのメッシュが重なっている領域において、どれだけ侵入しているかを表します。
		複数の方法で計算を試み、最も適切な結果を返します。
		
		方法1: オーバーラップメッシュを使用（最も正確）
		方法2: 遠位メッシュの頂点と近位メッシュの距離
		方法3: 両メッシュの最小距離
		
		Args:
			prox_mesh: 近位メッシュ（PyVistaメッシュ）
			dist_mesh: 遠位メッシュ（PyVistaメッシュ）
			overlap_mesh: オーバーラップメッシュ（オプション）
			
		Returns:
			float: 接触深度（mm）、接触がない場合は0.0
		"""
		try:
			import pyvista as pv
			
			# 方法1: オーバーラップメッシュが提供されている場合
			if overlap_mesh is not None and overlap_mesh.n_points > 0:
				# オーバーラップ領域の中心から両メッシュまでの平均距離を計算
				overlap_points = overlap_mesh.points
				
				# サンプリング（計算時間短縮のため、最大1000点）
				if len(overlap_points) > 1000:
					indices = np.random.choice(len(overlap_points), 1000, replace=False)
					sample_points = overlap_points[indices]
				else:
					sample_points = overlap_points
				
				# 各点から近位メッシュと遠位メッシュへの距離を計算
				prox_distances = []
				dist_distances = []
				
				for point in sample_points:
					# 近位メッシュへの距離
					try:
						idx = prox_mesh.find_closest_point(point)
						if isinstance(idx, (int, np.integer)) and idx >= 0:
							closest_pt = prox_mesh.points[idx]
							prox_dist = np.linalg.norm(point - closest_pt)
							prox_distances.append(prox_dist)
					except Exception:
						pass
					
					# 遠位メッシュへの距離
					try:
						idx = dist_mesh.find_closest_point(point)
						if isinstance(idx, (int, np.integer)) and idx >= 0:
							closest_pt = dist_mesh.points[idx]
							dist_dist = np.linalg.norm(point - closest_pt)
							dist_distances.append(dist_dist)
					except Exception:
						pass
				
				if len(prox_distances) > 0 and len(dist_distances) > 0:
					# 両方の平均距離の和を深度とする
					mean_depth = np.mean(prox_distances) + np.mean(dist_distances)
					print(f"[接触深度] 方法1（オーバーラップメッシュ）: {mean_depth:.4f} mm")
					return mean_depth
			
			# 方法2: 遠位メッシュの頂点から近位メッシュへの最小距離を計算
			# （オーバーラップがあれば、一部の点は非常に近いはず）
			dist_points = dist_mesh.points
			
			# サンプリング（計算時間短縮）
			if len(dist_points) > 2000:
				indices = np.random.choice(len(dist_points), 2000, replace=False)
				sample_dist_points = dist_points[indices]
			else:
				sample_dist_points = dist_points
			
			# 各点から近位メッシュへの距離を計算
			min_distances = []
			for point in sample_dist_points:
				try:
					idx = prox_mesh.find_closest_point(point)
					if isinstance(idx, (int, np.integer)) and idx >= 0:
						closest_pt = prox_mesh.points[idx]
						distance = np.linalg.norm(point - closest_pt)
						min_distances.append(distance)
				except Exception:
					pass
			
			if len(min_distances) > 0:
				# 最小距離の下位10%の平均を深度とする（最も近い点 = 最も侵入している点）
				sorted_distances = np.sort(min_distances)
				n_close = max(1, len(sorted_distances) // 10)
				close_distances = sorted_distances[:n_close]
				mean_depth = np.mean(close_distances)
				print(f"[接触深度] 方法2（最小距離法）: {mean_depth:.4f} mm")
				return mean_depth
			
			# 方法3: 両メッシュ間の最小距離を直接計算
			# これは計算コストが高いため、最後の手段
			try:
				# 近位メッシュの中心付近の点を取得
				prox_center = prox_mesh.center
				prox_points = prox_mesh.points
				
				# 中心に近い点を選択（上位20%）
				distances_to_center = np.linalg.norm(prox_points - prox_center, axis=1)
				n_select = max(100, len(prox_points) // 5)
				close_indices = np.argpartition(distances_to_center, n_select)[:n_select]
				selected_prox_points = prox_points[close_indices]
				
				# 選択された近位点から遠位メッシュへの最小距離
				min_dist_list = []
				for point in selected_prox_points:
					try:
						idx = dist_mesh.find_closest_point(point)
						if isinstance(idx, (int, np.integer)) and idx >= 0:
							closest_pt = dist_mesh.points[idx]
							distance = np.linalg.norm(point - closest_pt)
							min_dist_list.append(distance)
					except Exception:
						pass
				
				if len(min_dist_list) > 0:
					# 最小距離の下位10%の平均
					sorted_distances = np.sort(min_dist_list)
					n_close = max(1, len(sorted_distances) // 10)
					close_distances = sorted_distances[:n_close]
					mean_depth = np.mean(close_distances)
					print(f"[接触深度] 方法3（双方向最小距離）: {mean_depth:.4f} mm")
					return mean_depth
			except Exception as e:
				print(f"[接触深度] 方法3でエラー: {e}")
			
			return 0.0
		
		except Exception as e:
			print(f"[接触深度計算] エラー: {e}")
			import traceback
			traceback.print_exc()
			return 0.0

	def _show_precompute_dialog(self, total_frames, has_cartilage=False):
		"""事前計算の設定と進捗を表示するダイアログ

		Args:
			total_frames: 総フレーム数
			has_cartilage: FEM解析対象（関節領域または軟骨モデル）が利用可能かどうか

		Returns:
			tuple: (dialog_window, progress_var, status_var, cancel_var, options_dict, start_var)
		"""
		# プログレスダイアログを作成
		progress_window = tk.Toplevel(self)
		progress_window.title("事前計算")
		progress_window.geometry("550x750")
		progress_window.resizable(False, False)
		progress_window.grab_set()  # モーダルダイアログにする

		# プログレスバー
		progress_var = tk.DoubleVar()
		status_var = tk.StringVar(value="オプションを設定の上、「計算開始」をクリックしてください")
		cancel_var = tk.BooleanVar(value=False)
		start_var = tk.BooleanVar(value=False)  # 計算開始フラグ
		skip_var = tk.BooleanVar(value=False)  # スキップフラグ

		# 最適化オプション
		enable_precompute_var = tk.BooleanVar(value=True)  # 事前計算を行うかどうか
		use_parallel_var = tk.BooleanVar(value=True)
		use_cache_var = tk.BooleanVar(value=True)
		use_bbox_var = tk.BooleanVar(value=True)
		use_simplify_var = tk.BooleanVar(value=True)
		enable_fem_precompute_var = tk.BooleanVar(value=False)  # FEM事前計算（デフォルトOFF。必要時のみチェック）
		
		# タイトル
		title_label = ttk.Label(
			progress_window,
			text="事前計算",
			font=(self.ui_font_family, 12, "bold")
		)
		title_label.pack(pady=10)
		
		# 説明文
		info_label = ttk.Label(
			progress_window,
			text=f"フレーム数: {total_frames}\n"
			     f"事前計算することで、アニメーション中の処理が高速化されます。",
			justify=tk.LEFT
		)
		info_label.pack(pady=5)
		
		# 事前計算有効化チェックボックス
		enable_frame = ttk.Frame(progress_window)
		enable_frame.pack(pady=5, padx=10, fill=tk.X)
		
		enable_check = ttk.Checkbutton(
			enable_frame,
			text="ヒートマップを事前計算する（オフの場合、ヒートマップ表示は簡易/無効になります）",
			variable=enable_precompute_var,
			command=lambda: update_option_states()
		)
		enable_check.pack(anchor=tk.W)
		
		# 最適化オプションフレーム
		options_frame = ttk.LabelFrame(progress_window, text="高速化オプション", padding=10, style="Bold.TLabelframe")
		options_frame.pack(pady=5, padx=10, fill=tk.X)
		
		parallel_check = ttk.Checkbutton(
			options_frame,
			text=f"並列処理を使用（{max(1, cpu_count()-1)}コア）",
			variable=use_parallel_var
		)
		parallel_check.pack(anchor=tk.W, pady=2)
		
		cache_check = ttk.Checkbutton(
			options_frame,
			text="キャッシュファイルを使用（次回以降高速化）",
			variable=use_cache_var
		)
		cache_check.pack(anchor=tk.W, pady=2)
		
		bbox_check = ttk.Checkbutton(
			options_frame,
			text="バウンディングボックスチェック（交差なしをスキップ）",
			variable=use_bbox_var
		)
		bbox_check.pack(anchor=tk.W, pady=2)
		
		simplify_check = ttk.Checkbutton(
			options_frame,
			text="メッシュ簡略化（精度低下、高速化）",
			variable=use_simplify_var
		)
		simplify_check.pack(anchor=tk.W, pady=2)

		# FEM事前計算フレーム
		fem_frame = ttk.LabelFrame(progress_window, text="FEM接触解析事前計算", padding=10, style="Bold.TLabelframe")
		fem_frame.pack(pady=5, padx=10, fill=tk.X)

		fem_check = ttk.Checkbutton(
			fem_frame,
			text="FEM接触解析を事前計算する（関節領域で計算・時間がかかります）",
			variable=enable_fem_precompute_var,
			state="normal" if (_HAS_FEM and has_cartilage) else "disabled"
		)
		fem_check.pack(anchor=tk.W, pady=2)

		if not _HAS_FEM:
			fem_check_note = ttk.Label(fem_frame, text="FEMモジュール (fem_contact_solver_2.py) が見つかりません")
			fem_check_note.pack(anchor=tk.W, padx=20, pady=2)
		elif not has_cartilage:
			fem_check_note = ttk.Label(fem_frame, text="関節領域が抽出できていません（半径設定を確認してください）")
			fem_check_note.pack(anchor=tk.W, padx=20, pady=2)

		# オプションの有効/無効を切り替える関数
		def update_option_states():
			state = "normal" if enable_precompute_var.get() else "disabled"
			parallel_check.config(state=state)
			cache_check.config(state=state)
			bbox_check.config(state=state)
			simplify_check.config(state=state)
		
		# 状態表示
		status_label = ttk.Label(progress_window, textvariable=status_var)
		status_label.pack(pady=5)
		
		# プログレスバー
		progress_bar = ttk.Progressbar(
			progress_window,
			variable=progress_var,
			maximum=100,
			length=500,
			mode='determinate'
		)
		progress_bar.pack(pady=10)
		
		# パーセンテージ表示
		percent_label = ttk.Label(progress_window, text="0%")
		percent_label.pack(pady=5)
		
		# ボタンフレーム
		button_frame = ttk.Frame(progress_window)
		button_frame.pack(pady=10)
		
		# 計算開始ボタン
		def start_computation():
			if not enable_precompute_var.get():
				# 事前計算をスキップする場合
				skip_var.set(True)
				progress_window.destroy()
			else:
				# 計算開始
				start_var.set(True)
				start_button.config(state="disabled")
				skip_button.config(state="disabled")
				# オプションを無効化
				enable_check.config(state="disabled")
				parallel_check.config(state="disabled")
				cache_check.config(state="disabled")
				bbox_check.config(state="disabled")
				simplify_check.config(state="disabled")
		
		start_button = ttk.Button(
			button_frame,
			text="計算開始",
			command=start_computation
		)
		start_button.pack(side=tk.LEFT, padx=5)
		
		# スキップボタン（計算せずに進む）
		def skip_computation():
			skip_var.set(True)
			progress_window.destroy()
		
		skip_button = ttk.Button(
			button_frame,
			text="計算せずに進む",
			command=skip_computation
		)
		skip_button.pack(side=tk.LEFT, padx=5)

		# ウィンドウの×で閉じられた場合は「計算せずに進む」と同じ扱い（計算を始めない）
		progress_window.protocol("WM_DELETE_WINDOW", skip_computation)
		
		# キャンセルボタン
		def cancel_computation():
			cancel_var.set(True)
			progress_window.destroy()
		
		cancel_button = ttk.Button(
			button_frame,
			text="キャンセル",
			command=cancel_computation
		)
		cancel_button.pack(side=tk.LEFT, padx=5)
		
		# 進捗更新関数
		def update_progress(current, total, message):
			if cancel_var.get():
				return False  # キャンセルされた
			
			percentage = (current / total) * 100
			progress_var.set(percentage)
			status_var.set(message)
			percent_label.config(text=f"{percentage:.1f}%")
			progress_window.update()
			return True  # 継続
		
		# ウィンドウを中央に配置
		progress_window.transient(self)
		progress_window.update_idletasks()
		x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
		y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
		progress_window.geometry(f"+{x}+{y}")

		# ※ 自動スタートはしない。ユーザーが FEM/ヒートマップ等のオプションを設定してから
		#   「計算開始」または「計算せずに進む」を押すまで待機する。
		#   （以前はここで after(200, start_computation) により即時開始していたため、
		#     FEM事前計算などをオフにできなかった。）

		options_dict = {
			'use_parallel': use_parallel_var,
			'use_cache': use_cache_var,
			'use_bbox': use_bbox_var,
			'use_simplify': use_simplify_var,
			'enable_precompute': enable_precompute_var,
			'enable_fem_precompute': enable_fem_precompute_var
		}

		return progress_window, update_progress, cancel_var, options_dict, start_var, skip_var

	def show_fitting_help(self) -> None:
		"""フィッティングパラメータの説明を表示"""
		help_text = (
			"【パラメータの意味】\n\n"
			"■ RANSACパラメータ（粗い位置合わせ）\n"
			"  ・距離閾値: 対応点とみなす最大距離 (mm)\n"
			"  ・最大反復回数: ランダムサンプリングの試行回数\n"
			"  ・信頼度: 良い変換を見つける確率 (0.0～1.0)\n\n"
			"■ ICPパラメータ（精密な位置合わせ）\n"
			"  ・閾値: 対応点とみなす最大距離 (mm)\n"
			"  ・最大反復回数: 位置合わせの反復回数\n"
			"  ・サンプル点数: 計算に使う点の数（多いほど精密だが遅い）"
		)
		messagebox.showinfo("パラメータの説明", help_text)
	
	def reset_fitting_parameters(self) -> None:
		"""フィッティングパラメータをデフォルト値に戻す"""
		
		# RANSACパラメータ
		self.ransac_distance_threshold.set(1.0)
		self.ransac_max_iterations.set(1000)
		self.ransac_confidence.set(0.99)
		
		# ICPパラメータ
		self.icp_threshold.set(5.0)
		self.icp_max_iterations.set(2000)
		self.sample_points.set(10000)
		
		messagebox.showinfo("完了", "パラメータをデフォルト値に戻しました")

	def on_fitting_execute(self) -> None:
		"""
		【重要メソッド】フィッティング実行（Open3Dベース）
		
		RANSACとICPを使用して、親任意領域に子任意領域を位置合わせします。
		
		■ 処理フロー:
		  1. 親任意領域・子任意領域メッシュ読み込み
		  2. 主軸による事前アライメント（重心と主軸を一致）
		  3. RANSACによる粗い位置合わせ
		     - FPFHフィーチャー抽出
		     - ランダムサンプリングで最適変換を探索
		  4. ICPによる精密な位置合わせ
		     - 点対点ICP（Point-to-Point）
		     - 反復的に最適変換を収束
		  5. 結果の可視化と保存オプション表示
		  
		■ パラメータ:
		  - RANSACパラメータ（UI設定）:
		    * 距離閾値: 対応点とみなす最大距離 (mm)
		    * 最大反復回数: ランダムサンプリングの試行回数
		    * 信頼度: 良い変換を見つける確率 (0.0～1.0)
		  - ICPパラメータ（UI設定）:
		    * 閾値: 対応点とみなす最大距離 (mm)
		    * 最大反復回数: 位置合わせの反復回数
		    * サンプル点数: 計算に使う点の数
		  
		■ 変換行列の合成:
		  最終的な変換行列 = ICP変換 × RANSAC変換 × 主軸アライメント変換
		  この変換を子モデル全体に適用することで、フィッティング結果を得ます。
		  
		■ 改修ポイント:
		  - 行3180付近: RANSAC実装（FPFHフィーチャー計算）
		  - 行3250付近: ICP実装（反復最適化）
		  - 行3360付近: 結果メニュー表示
		  
		■ トラブルシューティング:
		  - RANSAC失敗: 距離閾値を大きく、または事前アライメントを確認
		  - ICP収束不良: 閾値を大きく、または反復回数を増やす
		  - メッシュ読み込み失敗: trimeshライブラリのインストールを確認
		"""
		# 必要なファイルがすべて選択されているか確認
		if not all([
			self.fitting_parent_region_path.get(),
			self.fitting_child_region_path.get()
		]):
			messagebox.showerror("エラー", "親任意領域と子任意領域を選択してください。")
			return
		
		try:
			import copy
			import trimesh # type: ignore
			
			# メッシュファイルを読み込む（Open3D形式）
			def load_mesh_o3d(file_path):
				"""メッシュファイルをOpen3Dで読み込む"""
				try:
					print(f"ファイル読み込み中: {Path(file_path).name}")
					# trimeshで読み込み
					trimesh_mesh = trimesh.load(file_path, force='mesh')
					vertices = np.array(trimesh_mesh.vertices)
					triangles = np.array(trimesh_mesh.faces)
					
					# Open3Dメッシュに変換
					mesh = o3d.geometry.TriangleMesh()
					mesh.vertices = o3d.utility.Vector3dVector(vertices)
					mesh.triangles = o3d.utility.Vector3iVector(triangles)
					mesh.compute_vertex_normals()
					
					print(f"  ✓ 成功: 頂点数={len(vertices)}, 三角形数={len(triangles)}")
					return mesh
				except Exception as e:
					print(f"  ✗ エラー: {e}")
					messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました:\n{Path(file_path).name}\n\n{str(e)}")
					return None
			
			# 2つの任意領域モデルを読み込み
			print("\n=== フィッティング処理開始 ===")
			print(f"親任意領域: {Path(self.fitting_parent_region_path.get()).name}")
			print(f"子任意領域: {Path(self.fitting_child_region_path.get()).name}")
			
			parent_region_mesh = load_mesh_o3d(self.fitting_parent_region_path.get())
			child_region_mesh = load_mesh_o3d(self.fitting_child_region_path.get())
			
			if not all([parent_region_mesh, child_region_mesh]):
				return
			
			# 色を設定
			parent_region_mesh.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_region_mesh.paint_uniform_color([0.0, 0.0, 1.0])  # 青
			
			# 初期状態を表示
			print("\n[初期状態] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_mesh)
			o3d.visualization.draw_geometries([parent_display, child_display],
										  window_name="初期状態: 親(赤) + 子(青)")
			# ウィンドウサイズを復元
			self._restore_window_geometry()
			
			# 主軸による事前アライメント
			print("\n=== 主軸による事前アライメント ===")
			
			# 主軸計算関数
			def get_principal_axis(mesh):
				"""メッシュの主軸（最大固有ベクトル）と重心を計算"""
				pcd = mesh.sample_points_uniformly(number_of_points=20000)
				pts = np.asarray(pcd.points)
				center = np.mean(pts, axis=0)
				cov = np.cov((pts - center).T)
				eigvals, eigvecs = np.linalg.eigh(cov)
				principal_axis = eigvecs[:, np.argmax(eigvals)]
				return center, principal_axis
			
			# 主軸回転関数
			def align_axis(mesh, src_axis, tgt_axis):
				"""主軸の向きを合わせる"""
				v = np.cross(src_axis, tgt_axis)
				c = np.dot(src_axis, tgt_axis)
				if np.linalg.norm(v) < 1e-8:
					R = np.eye(3)
				else:
					vx = np.array([[0, -v[2], v[1]],
								   [v[2], 0, -v[0]],
								   [-v[1], v[0], 0]])
					R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
				mesh_rot = copy.deepcopy(mesh)
				mesh_rot.rotate(R, center=(0, 0, 0))
				return mesh_rot, R
			
			# 主軸を計算
			c_parent, axis_parent = get_principal_axis(parent_region_mesh)
			c_child, axis_child = get_principal_axis(child_region_mesh)
			
			print(f"親の重心: {c_parent}")
			print(f"親の主軸: {axis_parent}")
			print(f"子の重心: {c_child}")
			print(f"子の主軸: {axis_child}")
			
			# 主軸の向きを一致させる
			if np.dot(axis_parent, axis_child) < 0:
				axis_child = -axis_child
				print("主軸の向きを反転しました")
			
			# 子メッシュの主軸を親メッシュに合わせる
			child_region_aligned, R_align = align_axis(child_region_mesh, axis_child, axis_parent)
			
			# 重心を一致させる
			c_child_aligned, _ = get_principal_axis(child_region_aligned)
			translation_vector = c_parent - c_child_aligned
			child_region_aligned.translate(translation_vector, relative=True)
			
			# 主軸アライメントの変換行列を構築（子の元の位置から変換後の位置への変換）
			# 1. 元の子の重心を原点に移動
			# 2. 回転を適用
			# 3. 新しい位置（親の重心）に移動
			axis_transform = np.eye(4)
			axis_transform[:3, :3] = R_align
			axis_transform[:3, 3] = c_parent - R_align @ c_child
			
			print(f"主軸アライメント変換行列:")
			print(axis_transform)
			
			# 主軸・重心一致後を表示
			print("\n[主軸アライメント後] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_aligned)
			parent_display.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_display.paint_uniform_color([0.0, 1.0, 0.0])  # 緑
			
			# 主軸を線で表示
			line_set = o3d.geometry.LineSet()
			points = [c_parent, c_parent + axis_parent*50, c_parent, c_parent + axis_child*50]
			lines = [[0, 1], [2, 3]]
			colors = [[1, 0, 0], [0, 0, 1]]
			line_set.points = o3d.utility.Vector3dVector(points)
			line_set.lines = o3d.utility.Vector2iVector(lines)
			line_set.colors = o3d.utility.Vector3dVector(colors)
			
			o3d.visualization.draw_geometries([parent_display, child_display, line_set],
										  window_name="主軸・重心一致後: 親(赤) + 子(緑)")
			# RANSACによる粗い位置合わせ
			print("\n=== RANSAC位置合わせ開始 ===")
			
			# 点群に変換
			parent_pcd = parent_region_mesh.sample_points_uniformly(number_of_points=10000)
			child_pcd = child_region_aligned.sample_points_uniformly(number_of_points=10000)
			
			print(f"点群サンプル数 - 親: {len(parent_pcd.points)}, 子: {len(child_pcd.points)}")
			
			# RANSACパラメータを取得
			# 信頼度は0.0-1.0のfloatとして扱う（UI値のスケーリングバグ修正）
			max_iterations = int(self.ransac_max_iterations.get())
			confidence = float(self.ransac_confidence.get())
			
			# voxel_sizeを点群から自動推定（平均最近傍距離に基づく）
			try:
				num_points = len(parent_pcd.points)
				sample_n = min(500, max(10, num_points))
				sample_indices = np.random.choice(num_points, sample_n, replace=False)
				kdt = o3d.geometry.KDTreeFlann(parent_pcd)
				nn_dists = []
				for idx in sample_indices:
					[_, idxs, dists] = kdt.search_knn_vector_3d(parent_pcd.points[idx], 2)
					if len(dists) >= 2:
						nn_dists.append(float(np.sqrt(dists[1])))
				if len(nn_dists) > 0:
					mean_nn = float(np.mean(nn_dists))
					voxel_size = max(mean_nn * 1.5, 1e-6)
				else:
					voxel_size = max(self.ransac_distance_threshold.get(), 1e-6)
			except Exception as e:
				voxel_size = max(self.ransac_distance_threshold.get(), 1e-6)
				print(f"[警告] voxel_size自動推定に失敗しました: {e}. UI値を使用します: {voxel_size}")
			
			print(f"RANSACパラメータ: voxel_size={voxel_size:.6f}, max_iterations={max_iterations}, confidence={confidence}")
			# デバッグ: バウンディングボックス情報
			minb = parent_region_mesh.get_min_bound()
			maxb = parent_region_mesh.get_max_bound()
			bbox_diag = np.linalg.norm(maxb - minb)
			print(f"親メッシュバウンディングボックス: min={minb}, max={maxb}, diag={bbox_diag:.3f}")
			
			# ボクセルダウンサンプリング
			parent_down = parent_pcd.voxel_down_sample(voxel_size)
			child_down = child_pcd.voxel_down_sample(voxel_size)
			
			print(f"ダウンサンプリング後 - 親: {len(parent_down.points)}, 子: {len(child_down.points)}")
			
			# 法線を計算
			radius_normal = voxel_size * 2
			parent_down.estimate_normals(
				search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			child_down.estimate_normals(
				search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=radius_normal, max_nn=30))
			
			# FPFHフィーチャーを計算
			radius_feature = voxel_size * 5
			print(f"FPFH計算中... (radius: {radius_feature})")
			
			parent_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				parent_down,
				o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			child_fpfh = o3d.pipelines.registration.compute_fpfh_feature(
				child_down,
				o3d.geometry.KDTreeSearchParamHybrid(radius=radius_feature, max_nn=100))
			
			# RANSACによる位置合わせ
			distance_threshold = voxel_size * 1.5
			print(f"RANSAC実行中... (distance_threshold: {distance_threshold})")
			
			result_ransac = o3d.pipelines.registration.registration_ransac_based_on_feature_matching(
				child_down, parent_down, child_fpfh, parent_fpfh,
				mutual_filter=True,
				max_correspondence_distance=distance_threshold,
				estimation_method=o3d.pipelines.registration.TransformationEstimationPointToPoint(False),
				ransac_n=4,
				checkers=[
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnEdgeLength(0.9),
					o3d.pipelines.registration.CorrespondenceCheckerBasedOnDistance(distance_threshold)
				],
				criteria=o3d.pipelines.registration.RANSACConvergenceCriteria(max_iterations, confidence))
			
			ransac_transform = result_ransac.transformation
			print(f"RANSAC完了 - Fitness: {result_ransac.fitness:.4f}, RMSE: {result_ransac.inlier_rmse:.4f}")
			print(f"変換行列:\n{ransac_transform}")
			
			# RANSAC結果をメッシュに適用
			child_region_ransac = copy.deepcopy(child_region_aligned)
			child_region_ransac.transform(ransac_transform)
			
			# RANSAC結果を可視化
			print("\n[RANSAC結果] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_ransac)
			parent_display.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_display.paint_uniform_color([0.0, 1.0, 1.0])  # シアン
			o3d.visualization.draw_geometries([parent_display, child_display],
										  window_name="RANSAC後: 親(赤) + 子(シアン)")
			# ウィンドウサイズを復元
			self._restore_window_geometry()
			
			# ICPによる精密な位置合わせ
			print("\n=== ICP位置合わせ開始 ===")
			threshold = self.icp_threshold.get()
			if threshold <= 0.0:
				print(f"[警告] ICP閾値が0以下({threshold})です。デフォルトの閾値(1e-6)を使用します。")
				threshold = 1e-6
			max_iterations = self.icp_max_iterations.get()
			
			print(f"ICPパラメータ: threshold={threshold}, max_iterations={max_iterations}")
			
			# 点群に変換
			parent_icp_pcd = parent_region_mesh.sample_points_uniformly(number_of_points=20000)
			child_icp_pcd = child_region_ransac.sample_points_uniformly(number_of_points=20000)
			
			print(f"ICP実行中... (点数: 親={len(parent_icp_pcd.points)}, 子={len(child_icp_pcd.points)})")
			
			# ICPによる位置合わせ
			result_icp = o3d.pipelines.registration.registration_icp(
				child_icp_pcd, parent_icp_pcd, threshold, np.eye(4),
				o3d.pipelines.registration.TransformationEstimationPointToPoint(),
				o3d.pipelines.registration.ICPConvergenceCriteria(max_iteration=max_iterations))
			
			icp_transform = result_icp.transformation
			print(f"ICP完了 - Fitness: {result_icp.fitness:.4f}, RMSE: {result_icp.inlier_rmse:.4f}")
			print(f"変換行列:\n{icp_transform}")
			
			# ICP結果をメッシュに適用
			child_region_final = copy.deepcopy(child_region_ransac)
			child_region_final.transform(icp_transform)
			
			# ICP結果を可視化
			print("\n[ICP結果(最終)] 可視化中...")
			parent_display = copy.deepcopy(parent_region_mesh)
			child_display = copy.deepcopy(child_region_final)
			parent_display.paint_uniform_color([1.0, 0.0, 0.0])  # 赤
			child_display.paint_uniform_color([0.0, 0.0, 1.0])  # 青
			o3d.visualization.draw_geometries([parent_display, child_display],
										  window_name="ICP後(最終結果): 親(赤) + 子(青)")
			# ウィンドウサイズを復元
			self._restore_window_geometry()
			
			print("\n=== フィッティング処理完了 ===")
			
			# ユーザーに結果の確認を求める
			result = messagebox.askyesno(
				"フィッティング結果の確認",
				"フィッティング結果に満足しましたか？\n\n"
				"「はい」を選択すると、変換後の操作メニューを表示します。"
			)
			
			if not result:
				print("フィッティング処理をキャンセルしました")
				messagebox.showinfo("キャンセル", "フィッティング処理をキャンセルしました")
				return
			
			# 子モデルに変換を適用して新しい子モデルを作成
			print("\n=== 子モデルへの変換適用 ===")
			
			# 子モデル（フルモデル）を読み込み
			child_model_path = self.fitting_child_model_path.get()
			if not child_model_path:
				messagebox.showerror("エラー", "子モデルが選択されていません")
				return
			
			print(f"子モデル読み込み中: {Path(child_model_path).name}")
			child_model_mesh = load_mesh_o3d(child_model_path)
			print(f"  ✓ 頂点数={len(child_model_mesh.vertices)}, 三角形数={len(child_model_mesh.triangles)}")
			
			# 合成変換行列を計算（主軸アライメント + RANSAC + ICP）
			# フィッティングで求めた変換：子の任意領域 → 親の任意領域
			print("\n=== 変換行列の合成 ===")
			print("1. 主軸アライメント変換（既に計算済み）")
			print("2. RANSAC変換:")
			print(ransac_transform)
			print("3. ICP変換:")
			print(icp_transform)
			
			# 全変換を合成（子→親の変換）
			child_to_parent_transform = icp_transform @ ransac_transform @ axis_transform
			
			print("子→親の変換行列:")
			print(child_to_parent_transform)
			
			# 子モデル全体に変換を適用
			child_model_transformed = copy.deepcopy(child_model_mesh)
			child_model_transformed.transform(child_to_parent_transform)
			print("  ✓ 子モデルに変換を適用しました")
			
			# 統合GUIを表示
			self._show_fitting_result_menu(
				child_model_transformed,
				child_to_parent_transform,
				child_model_path,
				load_mesh_o3d
			)
			
		except Exception as e:
			messagebox.showerror("エラー", f"フィッティング処理中にエラーが発生しました:\n{str(e)}")
			print(f"エラー詳細: {e}")
			import traceback
			traceback.print_exc()
		finally:
			# メインウィンドウのサイズを復元
			self._restore_window_geometry()
	
	def _show_fitting_result_menu(self, child_model_transformed, child_to_parent_transform, child_model_path, load_mesh_o3d):
		"""フィッティング結果の操作メニューGUIを表示
		
		Args:
			child_model_transformed: 変換後の子モデル（Open3D）
			child_to_parent_transform: 変換行列
			child_model_path: 元の子モデルパス
			load_mesh_o3d: メッシュ読み込み関数
		"""
		# 新しいウィンドウを作成
		menu_window = tk.Toplevel(self)
		menu_window.title("フィッティング完了 - 操作メニュー")
		menu_window.geometry("800x850") # サイズ変更
		menu_window.resizable(True, True) # リサイズ可能に
		
		# ウィンドウを中央に配置
		menu_window.update_idletasks()
		x = (menu_window.winfo_screenwidth() // 2) - (800 // 2)
		y = (menu_window.winfo_screenheight() // 2) - (850 // 2)
		menu_window.geometry(f"+{x}+{y}")
		
		# メインフレーム
		main_frame = ttk.Frame(menu_window, padding=20)
		main_frame.pack(fill=tk.BOTH, expand=True)
		
		# タイトル
		title_label = ttk.Label(
			main_frame,
			text="フィッティング完了！",
			font=(self.ui_font_family, 14, "bold")
		)
		title_label.pack(pady=(0, 10))

		# 同時変換行列の表示（テキストエリア）
		matrix_frame = ttk.LabelFrame(main_frame, text="計算された同時変換行列 (Child ➙ Parent)", padding=10)
		matrix_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

		matrix_text = tk.Text(matrix_frame, height=14, width=65, font=(self.mono_font_family, 10))
		matrix_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		
		matrix_scrollbar = ttk.Scrollbar(matrix_frame, orient="vertical", command=matrix_text.yview)
		matrix_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
		matrix_text.config(yscrollcommand=matrix_scrollbar.set)
		
		# 行列文字列作成
		m = child_to_parent_transform
		
		# PosCheckerと同じフォーマットで解析情報を付与
		result_text = "【フィッティング結果変換行列 (Child ➙ Parent)】\n"
		result_text += "※ Child(移動側)からParent(基準側)への座標変換\n"
		result_text += "--------------------------------------------------\n"
		result_text += f"{m[0][0]:12.6f} {m[0][1]:12.6f} {m[0][2]:12.6f} {m[0][3]:12.6f}\n"
		result_text += f"{m[1][0]:12.6f} {m[1][1]:12.6f} {m[1][2]:12.6f} {m[1][3]:12.6f}\n"
		result_text += f"{m[2][0]:12.6f} {m[2][1]:12.6f} {m[2][2]:12.6f} {m[2][3]:12.6f}\n"
		result_text += f"{m[3][0]:12.6f} {m[3][1]:12.6f} {m[3][2]:12.6f} {m[3][3]:12.6f}\n"
		result_text += "--------------------------------------------------\n\n"
		
		# 解析 (Translation / Rotation)
		trans = m[:3, 3]
		norm_val = np.linalg.norm(trans)
		rot_mat = m[:3, :3]
		import math
		sy = math.sqrt(rot_mat[0,0] * rot_mat[0,0] +  rot_mat[1,0] * rot_mat[1,0])
		if not sy < 1e-6:
			x = math.atan2(rot_mat[2,1] , rot_mat[2,2])
			y = math.atan2(-rot_mat[2,0], sy)
			z = math.atan2(rot_mat[1,0], rot_mat[0,0])
		else:
			x = math.atan2(-rot_mat[1,2], rot_mat[1,1])
			y = math.atan2(-rot_mat[2,0], sy)
			z = 0
			
		result_text += "【解析】\n"
		result_text += f"Translation (mm): X={trans[0]:.4f}, Y={trans[1]:.4f}, Z={trans[2]:.4f}\n"
		result_text += f"Norm (mm)       : {norm_val:.4f}\n"
		result_text += f"Rotation (deg)  : X={math.degrees(x):.4f}, Y={math.degrees(y):.4f}, Z={math.degrees(z):.4f}\n"
		
		matrix_text.insert("1.0", result_text)
		matrix_text.config(state="disabled") # 編集不可

		# 行列保存ボタン
		def save_matrix_txt():
			matrix_path = filedialog.asksaveasfilename(
				title="同時変換行列を保存",
				defaultextension=".txt",
				filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
			)
			if matrix_path:
				try:
					with open(matrix_path, "w", encoding="utf-8") as f:
						f.write(result_text)
					messagebox.showinfo("保存完了", f"行列を保存しました:\n{matrix_path}")
				except Exception as e:
					messagebox.showerror("エラー", f"保存に失敗しました:\n{e}")

		ttk.Button(matrix_frame, text="行列をテキストとして保存", command=save_matrix_txt).pack(anchor="e", pady=(5, 0))

		# 説明
		info_label = ttk.Label(
			main_frame,
			text="変換が完了しました。以下のボタンから必要な操作を選択してください。",
			wraplength=750
		)
		info_label.pack(pady=(0, 20))
		
		# ボタンフレーム
		button_frame = ttk.Frame(main_frame)
		button_frame.pack(fill=tk.BOTH, expand=True)
		
		# 保存済みフラグ
		model_saved = [False]  # リストで包んで参照渡し
		pp_saved = [False]
		saved_model_path = [None]
		
		# ヒートマップデータキャッシュ
		heatmap_data_cache = {
			'distances': None,
			'points': None,
			'info': ""
		}
		
		# 1. 変換後のモデル保存
		def save_model():
			original_child_path = Path(child_model_path)
			init_dir = str(original_child_path.parent)
			init_name = original_child_path.stem + ".Fitted.stl"
			output_path = filedialog.asksaveasfilename(
				title="変換後の子モデルを保存",
				defaultextension=".stl",
				filetypes=[
					("STL files", "*.stl"),
					("OBJ files", "*.obj"),
					("PLY files", "*.ply"),
					("All files", "*.*")
				],
				initialfile=init_name,
				initialdir=init_dir
			)
			
			if output_path:
				print(f"\n変換後の子モデルを保存中: {Path(output_path).name}")
				o3d.io.write_triangle_mesh(output_path, child_model_transformed)
				print(f"  ✓ 保存完了: {output_path}")
				model_saved[0] = True
				saved_model_path[0] = output_path
				btn_save_model.config(text="✓ モデル保存完了", state="disabled")
				messagebox.showinfo("保存完了", f"モデルを保存しました:\n{output_path}")
		
		btn_save_model = ttk.Button(
			button_frame,
			text="変換後のモデルを保存",
			command=save_model
		)
		btn_save_model.pack(fill=tk.X, pady=5)
		
		# 2. 変換後のPP保存
		def save_pp():
			try:
				child_pp_path = self.fitting_child_pp_path.get()
				if not child_pp_path:
					child_pp_path = filedialog.askopenfilename(
						title="子PPファイルを選択",
						filetypes=[("PickedPoints/XML", "*.pp;*.xml;*.txt"), ("All files", "*.*")]
					)
					if not child_pp_path:
						return
					self.fitting_child_pp_path.set(child_pp_path)
				
				# PP読み込み
				points, labels = self._parse_pp_file(child_pp_path)
				filtered_points = np.array(points)
				filtered_labels = list(labels)
				
				# 変換適用
				if len(filtered_points) > 0:
					homo = np.hstack([filtered_points, np.ones((len(filtered_points), 1))])
					transformed = (child_to_parent_transform @ homo.T).T[:, :3]
				else:
					transformed = np.empty((0, 3))
				
				# 保存ダイアログ
				pp_dir = str(Path(child_pp_path).parent)
				base_name = Path(child_pp_path).stem
				out_name = base_name + ".Fitted_points.pp"
				save_pp_path = filedialog.asksaveasfilename(
					title="変換後の子PPを保存",
					defaultextension=".pp",
					filetypes=[("PickedPoints/XML", "*.pp;*.xml"), ("All files", "*.*")],
					initialdir=pp_dir,
					initialfile=out_name
				)
				
				if save_pp_path:
					save_child_pp_file(save_pp_path, transformed, filtered_labels)
					print(f"  ✓ 子PP保存完了: {save_pp_path}")
					pp_saved[0] = True
					btn_save_pp.config(text="✓ PP保存完了", state="disabled")
					messagebox.showinfo("保存完了", f"子PPを保存しました:\n{save_pp_path}")
			except Exception as e:
				print(f"[PP保存] エラー: {e}")
				messagebox.showerror("エラー", f"子PP保存でエラーが発生しました:\n{e}")
		
		btn_save_pp = ttk.Button(
			button_frame,
			text="変換後のPPを保存",
			command=save_pp
		)
		btn_save_pp.pack(fill=tk.X, pady=5)
		
		# 区切り線
		separator = ttk.Separator(button_frame, orient='horizontal')
		separator.pack(fill=tk.X, pady=(15, 5))

		# --- 可視化設定 (色設定) ---
		settings_frame = ttk.LabelFrame(button_frame, text="可視化の色設定", padding=10)
		settings_frame.pack(fill=tk.X, pady=(0, 15))
		
		# デフォルト色設定
		viz_colors = {
			'parent': '#FF0000',     # 赤
			'child': '#00FF00',      # 緑
			'cmap': 'turbo'          # ヒートマップ
		}
		
		# 色選択コールバック
		def pick_parent_color():
			c = colorchooser.askcolor(color=viz_colors['parent'], title="親モデルの色を選択")[1]
			if c:
				viz_colors['parent'] = c
				lbl_parent_color.config(bg=c)

		def pick_child_color():
			c = colorchooser.askcolor(color=viz_colors['child'], title="子モデルの色を選択")[1]
			if c:
				viz_colors['child'] = c
				lbl_child_color.config(bg=c)
				
		# 親モデル色
		f_colors = ttk.Frame(settings_frame)
		f_colors.pack(fill=tk.X)
		
		f_p = ttk.Frame(f_colors)
		f_p.pack(side=tk.LEFT, padx=5)
		ttk.Button(f_p, text="親モデル色", command=pick_parent_color, width=12).pack(side=tk.LEFT)
		lbl_parent_color = tk.Label(f_p, bg=viz_colors['parent'], width=4, relief="solid")
		lbl_parent_color.pack(side=tk.LEFT, padx=2)
		
		# 子モデル色
		f_c = ttk.Frame(f_colors)
		f_c.pack(side=tk.LEFT, padx=5)
		ttk.Button(f_c, text="子モデル色", command=pick_child_color, width=12).pack(side=tk.LEFT)
		lbl_child_color = tk.Label(f_c, bg=viz_colors['child'], width=4, relief="solid")
		lbl_child_color.pack(side=tk.LEFT, padx=2)

		# 3. 普通に可視化
		def visualize_simple():
			print("\n[可視化] Open3Dで表示中...")
			
			# hex to rgb helper
			def h2r(h):
				h = h.lstrip('#')
				return [int(h[i:i+2], 16)/255.0 for i in (0, 2, 4)]
			
			p_col = h2r(viz_colors['parent'])
			c_col = h2r(viz_colors['child'])
			
			geoms = []
			
			# 親モデル
			parent_model_path = self.fitting_parent_model_path.get()
			if parent_model_path:
				parent_model_mesh = load_mesh_o3d(parent_model_path)
				parent_model_mesh.paint_uniform_color(p_col)
				geoms.append(parent_model_mesh)
			
			# 変換後の子モデル
			child_display = copy.deepcopy(child_model_transformed)
			child_display.paint_uniform_color(c_col)
			geoms.append(child_display)
			
			o3d.visualization.draw_geometries(
				geoms,
				window_name="フィッティング結果: シンプル表示"
			)
			# ウィンドウサイズを復元
			self._restore_window_geometry()
		
		btn_visualize_simple = ttk.Button(
			button_frame,
			text="普通に可視化する",
			command=visualize_simple
		)
		btn_visualize_simple.pack(fill=tk.X, pady=5)
		
		# 4. ヒートマップありで可視化
		def visualize_heatmap():
			print("\n[ヒートマップ可視化] 準備中...")
			
			# 優先順位: 指定されたヒートマップ用モデル > (ダイアログ選択: 任意領域 vs 全体)
			heatmap_model_path_val = self.fitting_child_heatmap_model_path.get().strip()
			use_region_only = False
			
			# ヒートマップ表示用モデルが指定されていない場合のみ、ダイアログを表示
			if not heatmap_model_path_val:
				# 子任意領域のみを対象にするか確認（カスタムダイアログ）
				selection_var = tk.IntVar(value=-1) # 1: Region, 0: Full
				
				d = tk.Toplevel(menu_window)
				d.title("ヒートマップ対象選択")
				d.geometry("450x230")
				d.resizable(False, False)
				
				# 中央配置
				d.update_idletasks()
				dx = (d.winfo_screenwidth() // 2) - (450 // 2)
				dy = (d.winfo_screenheight() // 2) - (230 // 2)
				d.geometry(f"+{dx}+{dy}")
				
				ttk.Label(d, text="距離計算の基準メッシュを選択してください", font=(self.ui_font_family, 11, "bold")).pack(pady=(20, 10))
				ttk.Label(d, text="どちらのモデルに対して距離ヒートマップを表示しますか？", font=(self.ui_font_family, 10)).pack(pady=(0, 20))
				
				def select_region():
					selection_var.set(1)
					d.destroy()
					
				def select_full():
					selection_var.set(0)
					d.destroy()
					
				btn_region = ttk.Button(d, text="子任意領域のみ (高速)", command=select_region, width=30)
				btn_region.pack(pady=2)
				ttk.Label(d, text="※ フィッティングに使用した部分のみ", font=(self.ui_font_family, 9), foreground="gray").pack(pady=(0, 12))
				
				btn_full = ttk.Button(d, text="子モデル全体 (低速)", command=select_full, width=30)
				btn_full.pack(pady=2)
				ttk.Label(d, text="※ 全体の干渉を確認できます（計算に時間がかかります）", font=(self.ui_font_family, 9), foreground="gray").pack(pady=(0, 10))
				
				d.transient(menu_window)
				d.grab_set()
				menu_window.wait_window(d)
				
				if selection_var.get() == -1:
					print("ヒートマップ表示キャンセル")
					return

				use_region_only = (selection_var.get() == 1)
			
			# 親モデル読み込み
			parent_model_path = self.fitting_parent_model_path.get()
			if not parent_model_path:
				messagebox.showerror("エラー", "親モデルが選択されていません")
				return
			
			parent_model_mesh_pv = pv.read(parent_model_path)
			
			heatmap_mesh_to_display = None

			# ヒートマップ表示用モデルがある場合（子モデル側にヒートマップを表示）
			if heatmap_model_path_val:
				print(f"ヒートマップ表示用モデルを使用: {Path(heatmap_model_path_val).name}")
				# 指定モデルを読み込み、フィッティング変換を適用
				hm_o3d = load_mesh_o3d(heatmap_model_path_val)
				if hm_o3d is None:
					messagebox.showerror("エラー", "ヒートマップ表示用モデルの読み込みに失敗しました")
					return
				hm_transformed = copy.deepcopy(hm_o3d)
				hm_transformed.transform(child_to_parent_transform)
				verts = np.asarray(hm_transformed.vertices)
				tris = np.asarray(hm_transformed.triangles)
				faces_pv = np.hstack([np.full((len(tris), 1), 3), tris]) if len(tris) > 0 else np.empty((0, 4), dtype=int)
				
				# 計算対象（Subject）: ヒートマップ表示用モデル（子）
				subject_vertices = verts
				# メッシュが巨大な場合は簡略化（ユーザー要望: 軽量化）
				if len(verts) > 30000:
					print(f"  メッシュ簡略化中 ({len(verts)} points)...")
					tmp_mesh = pv.PolyData(verts, faces_pv)
					subject_mesh = tmp_mesh.decimate(target_reduction=0.5)
					subject_vertices = np.array(subject_mesh.points)
					print(f"  -> 簡略化後: {len(subject_vertices)} points")
				else:
					subject_mesh = pv.PolyData(verts, faces_pv)
				
				# 距離基準（Target）: 親モデル全体
				# ユーザー提案: バウンディングボックス（長方形）で親モデルを切り抜く（ROI抽出）
				print(f"  距離計算用ターゲット: 親モデル全体（ROI抽出）")
				
				# バウンディングボックス計算
				bounds = subject_mesh.bounds # (xmin, xmax, ymin, ymax, zmin, zmax)
				margin = 15.0 # mm (接触判定用なので余裕を持たせる。これ以上離れた点は精密計算しても意味が薄い)
				
				clip_box_bounds = [
					bounds[0] - margin, bounds[1] + margin,
					bounds[2] - margin, bounds[3] + margin,
					bounds[4] - margin, bounds[5] + margin
				]
				
				try:
					# PyVistaのclip_boxでROI抽出
					# invert=False: 箱の中身を残す
					print(f"  親モデルをクリッピング中 (ROI Margin: {margin}mm)...")
					target_mesh = parent_model_mesh_pv.clip_box(clip_box_bounds, invert=False)
					
					# クリップ結果が空（全く重なっていない）場合の対策
					if target_mesh.n_points == 0:
						print("  警告: ROI内に親モデルが存在しません。親モデル全体を使用します。")
						target_mesh = parent_model_mesh_pv
					else:
						print(f"  -> 親モデル要素数: {parent_model_mesh_pv.n_points} -> {target_mesh.n_points} points")
				
				except Exception as e:
					print(f"  クリッピング失敗: {e}")
					target_mesh = parent_model_mesh_pv
				
				calc_label = "距離ヒートマップを計算中（子ヒートマップモデル(簡略) → 親モデル(ROI)）..."
				
			# 指定がない場合（親モデル側にヒートマップを表示して、子との距離を見る）
			else:
				if use_region_only:
					print("子任意領域に変換を適用中...")
					child_region_path = self.fitting_child_region_path.get()
					child_region_mesh_o3d = load_mesh_o3d(child_region_path)
					child_region_transformed = copy.deepcopy(child_region_mesh_o3d)
					child_region_transformed.transform(child_to_parent_transform)
					region_vertices = np.asarray(child_region_transformed.vertices)
					region_triangles = np.asarray(child_region_transformed.triangles)
					faces_pv = np.hstack([np.full((len(region_triangles), 1), 3), region_triangles])
					reference_mesh = pv.PolyData(region_vertices, faces_pv)
				else:
					child_vertices = np.asarray(child_model_transformed.vertices)
					child_triangles = np.asarray(child_model_transformed.triangles)
					faces_pv = np.hstack([np.full((len(child_triangles), 1), 3), child_triangles])
					reference_mesh = pv.PolyData(child_vertices, faces_pv)
				
				# 計算対象（Subject）: 親モデル
				subject_vertices = np.array(parent_model_mesh_pv.points)
				subject_mesh = parent_model_mesh_pv.copy()
				
				# 距離基準（Target）: 子モデル（領域 or 全体）
				target_mesh = reference_mesh
				
				calc_label = "距離ヒートマップを計算中（親モデル → 子モデル）..."

			print(f"  計算対象頂点数: {len(subject_vertices)} 点")
			print(calc_label)
			
			# プログレスウィンドウを作成
			progress_win = tk.Toplevel(menu_window)
			progress_win.title("計算中...")
			progress_win.geometry("400x120")
			progress_win.resizable(False, False)
			progress_win.attributes('-topmost', True)
			
			# 中央配置
			progress_win.update_idletasks()
			px = (progress_win.winfo_screenwidth() // 2) - (400 // 2)
			py = (progress_win.winfo_screenheight() // 2) - (120 // 2)
			progress_win.geometry(f"+{px}+{py}")
			
			ttk.Label(progress_win, text=calc_label, font=(self.ui_font_family, 10)).pack(pady=(15, 5))
			
			progress_var = tk.DoubleVar(value=0)
			progress_bar = ttk.Progressbar(progress_win, variable=progress_var, maximum=100)
			progress_bar.pack(fill=tk.X, padx=20, pady=5)
			
			status_label = ttk.Label(progress_win, text="0%", font=(self.ui_font_family, 9), foreground="gray")
			status_label.pack(pady=2)
			
			progress_win.update()
			
			def update_progress(val):
				if val > 100:
					status_label.config(text="補間計算中...")
				else:
					progress_var.set(val)
					status_label.config(text=f"{val:.0f}%")
				progress_win.update()
			
			try:
				# 高速化のため、Open3DのPoint-to-Point距離計算を使用
				# 元々は「近傍点のみPyVistaの精密計算(Point-to-Surface)」をループで行っていたが、
				# 点数が多いと非常に時間がかかるため(Pythonループのオーバーヘッド)、
				# ターゲットメッシュの「頂点」に加えて「面の中心点」もサンプリングして密度を上げ、
				# Open3DのKDTree(C++)で一括計算する方式に変更。
				
				status_label.config(text="距離分布を計算中...")
				progress_win.update()
				
				# 計算対象（Subject）
				pcd_s = o3d.geometry.PointCloud()
				pcd_s.points = o3d.utility.Vector3dVector(subject_vertices)
				
				# 距離基準（Target）
				# 頂点だけでなく面の中心点も追加することで、メッシュ表面への距離をより正確に近似する
				tgt_pts = np.array(target_mesh.points)
				try:
					if target_mesh.n_cells > 0:
						status_label.config(text="メッシュ表面点をサンプリング中...")
						progress_win.update()
						centers = target_mesh.cell_centers().points
						tgt_pts = np.vstack([tgt_pts, centers])
				except Exception as e:
					print(f"面中心計算スキップ: {e}")
				
				pcd_t = o3d.geometry.PointCloud()
				pcd_t.points = o3d.utility.Vector3dVector(tgt_pts)
				
				status_label.config(text="最近傍探索を実行中(KDTree)...")
				progress_win.update()
				
				# 全点に対して高速計算
				distances = np.asarray(pcd_s.compute_point_cloud_distance(pcd_t))
				
				progress_var.set(100)
				status_label.config(text="計算完了")
				progress_win.update()
				
				# 3. 遠すぎる点の除外（分布を見やすくするため、閾値以上は丸めるか除外）
				# ここでは生の値を保持するが、可視化時にclimで調整される
				
			finally:
				progress_win.destroy()
			
			# 計算対象メッシュにヒートマップを適用
			heatmap_mesh_to_display = subject_mesh # 既にPolyData
			heatmap_mesh_to_display['distance'] = distances
			
			# キャッシュに保存
			heatmap_data_cache['distances'] = distances
			heatmap_data_cache['points'] = np.array(subject_mesh.points)
			heatmap_data_cache['info'] = calc_label
			
			# エクセル出力ボタンを有効化
			try:
				btn_export_heatmap.config(state="normal", text="ヒートマップ分布をExcelに出力 (データあり)")
			except Exception as e:
				print(f"ボタン有効化エラー: {e}")
			
			max_dist = np.max(distances)
			mean_dist = np.mean(distances)
			min_dist = np.min(distances)
			
			print(f"\n距離統計:")
			print(f"  最小距離: {min_dist:.4f} mm")
			print(f"  平均距離: {mean_dist:.4f} mm")
			print(f"  最大距離: {max_dist:.4f} mm")
			
			# 子モデル全体も準備
			child_vertices_full = np.asarray(child_model_transformed.vertices)
			child_triangles_full = np.asarray(child_model_transformed.triangles)
			faces_pv_full = np.hstack([np.full((len(child_triangles_full), 1), 3), child_triangles_full])
			child_model_pv = pv.PolyData(child_vertices_full, faces_pv_full)
			
			# PyVistaで可視化
			print("\n可視化ウィンドウを起動中...")
			self._show_fitting_result_with_heatmap(
				parent_model_mesh_pv,
				child_model_pv,
				heatmap_mesh_to_display,
				None,  # parent_mesh_o3d (不要)
				child_model_transformed,
				child_to_parent_transform,
				parent_color=viz_colors['parent'],
				child_color=viz_colors['child'],
				cmap='turbo'
			)
		
		btn_visualize_heatmap = ttk.Button(
			button_frame,
			text="ヒートマップありで可視化する（時間かかるよ）",
			command=visualize_heatmap
		)
		btn_visualize_heatmap.pack(fill=tk.X, pady=5)

		# 4.5 ヒートマップデータ出力
		def export_heatmap_data():
			try:
				import pandas as pd
				import numpy as np
				
				distances = heatmap_data_cache['distances']
				points = heatmap_data_cache['points']
				
				if distances is None:
					return
				
				# NaN除去
				valid_mask = ~np.isnan(distances)
				n_total = len(distances)
				distances = distances[valid_mask]
				points = points[valid_mask]
				n_valid = len(distances)
				
				if n_valid < n_total:
					print(f"  Note: NaNを含む {n_total - n_valid} 点を除外しました")

				# 保存ダイアログ
				default_name = "Heatmap_Distribution.xlsx"
				if self.fitting_child_model_path.get():
					default_name = Path(self.fitting_child_model_path.get()).stem + "_Heatmap.xlsx"

				out_path = filedialog.asksaveasfilename(
					title="ヒートマップ分布データの保存",
					defaultextension=".xlsx",
					filetypes=[("Excel files", "*.xlsx")],
					initialfile=default_name
				)
				
				if not out_path:
					return
				
				print(f"Excel保存開始: {out_path}")
				
				# 1. 基本統計量
				stats_df = pd.DataFrame({
					"Item": ["Min (mm)", "Max (mm)", "Mean (mm)", "Std (mm)", "Count"],
					"Value": [
						np.min(distances),
						np.max(distances),
						np.mean(distances),
						np.std(distances),
						len(distances)
					]
				})
				
				# 2. 分布（ヒストグラム） 0.1mm刻み
				# バケット作成
				max_dist = np.max(distances)
				max_val = np.ceil(max_dist * 10) / 10.0 # 0.1単位で切り上げ
				if max_val < 0.1: max_val = 0.1
				
				bins = np.arange(0, max_val + 0.1001, 0.1) # 0, 0.1, 0.2 ...
				
				# pd.cutでビン詰め
				cuts = pd.cut(distances, bins=bins, right=True, include_lowest=True)
				counts = cuts.value_counts().sort_index()
				
				dist_df = pd.DataFrame({
					"Range": counts.index.astype(str),
					"Count": counts.values,
					"Percentage": (counts.values / len(distances) * 100)
				})
				
				# 3. 生データ (座標 + 距離)
				# 100万行制限チェック
				if len(distances) < 1000000:
					raw_df = pd.DataFrame({
						"X": points[:, 0],
						"Y": points[:, 1],
						"Z": points[:, 2],
						"Distance_mm": distances
					})
					write_raw = True
				else:
					print("  データ数が多すぎるためRawDataシートはスキップします")
					write_raw = False
				
				# ExcelWriterで複数シートに書き出し
				with pd.ExcelWriter(out_path) as writer:
					stats_df.to_excel(writer, sheet_name="Summary", index=False)
					dist_df.to_excel(writer, sheet_name="Distribution", index=False)
					if write_raw:
						raw_df.to_excel(writer, sheet_name="RawData", index=False)
					
				messagebox.showinfo("保存完了", f"ヒートマップデータを保存しました:\n{out_path}")
				print("  ✓ 保存完了")
				
			except ImportError:
				messagebox.showerror("エラー", "pandasライブラリが見つかりません。\n'pip install pandas openpyxl'を実行してください。")
			except Exception as e:
				print(f"Excel保存エラー: {e}")
				print(traceback.format_exc())
				messagebox.showerror("エラー", f"Excel保存に失敗しました:\n{e}")

		btn_export_heatmap = ttk.Button(
			button_frame,
			text="ヒートマップ分布をExcelに出力",
			command=export_heatmap_data,
			state="disabled"
		)
		btn_export_heatmap.pack(fill=tk.X, pady=5)
		
		# 区切り線
		separator2 = ttk.Separator(button_frame, orient='horizontal')
		separator2.pack(fill=tk.X, pady=15)
		
		# 5. 閉じるボタン
		def close_menu():
			menu_window.destroy()
			if model_saved[0] or pp_saved[0]:
				messagebox.showinfo("完了", "フィッティング処理が完了しました。")
		
		btn_close = ttk.Button(
			button_frame,
			text="閉じる",
			command=close_menu
		)
		btn_close.pack(fill=tk.X, pady=5)
		
		# ウィンドウを最前面に
		menu_window.transient(self)
		menu_window.grab_set()
		self.wait_window(menu_window)
	
	def _show_fitting_result_with_heatmap(self, parent_mesh_pv, child_mesh_pv, child_heatmap_mesh, parent_mesh_o3d, child_mesh_o3d, transform, parent_color='red', child_color='green', cmap='turbo'):
		"""フィッティング結果を距離ヒートマップ付きで可視化
		
		Args:
			parent_mesh_pv: 親メッシュ（PyVista）
			child_mesh_pv: 子メッシュ全体（PyVista、通常表示用）
			child_heatmap_mesh: ヒートマップ用メッシュ（PyVista、距離データ付き）
			parent_mesh_o3d: 親メッシュ（Open3D）
			child_mesh_o3d: 子メッシュ（Open3D）
			transform: 変換行列
			parent_color: 親モデルの色 (default: 'red')
			child_color: 子モデルの色 (default: 'green')
			cmap: ヒートマップのカラーマップ (default: 'turbo')
		"""
		plotter = pv.Plotter()
		plotter.set_background('white')
		
		# アクター辞書（表示切り替え用）
		actors = {}
		
		# 親モデル（赤）
		if parent_mesh_pv is not None:
			actor_parent = plotter.add_mesh(
				parent_mesh_pv,
				color=parent_color,
				opacity=0.5,
				show_edges=False,
				label='親モデル'
			)
			actors['parent'] = actor_parent
		
		# 子モデル（緑、通常表示用）
		if child_mesh_pv is not None:
			# 通常表示用（緑）
			actor_child = plotter.add_mesh(
				child_mesh_pv,
				color=child_color,
				opacity=0.5,
				show_edges=False,
				label='子モデル'
			)
			actors['child'] = actor_child
		
		# ヒートマップ表示用（別メッシュ）
		if child_heatmap_mesh is not None and 'distance' in child_heatmap_mesh.array_names:
			actor_heatmap = plotter.add_mesh(
				child_heatmap_mesh,
				scalars='distance',
				cmap=cmap,  # カラーマップ
				clim=[0.0, 5.0],  # カラースケール: 0～5mm
				show_edges=False,
				opacity=1.0,
				label='距離ヒートマップ',
				scalar_bar_args={
					'title': '距離 (mm)',
					'vertical': True,
				'title_font_size': 16,
				'label_font_size': 14
			}
		)
		actors['heatmap'] = actor_heatmap
		# 初期状態は非表示に設定
		actor_heatmap.SetVisibility(False)
		def update_parent_opacity(value):
			if 'parent' in actors:
				actors['parent'].GetProperty().SetOpacity(value)
		
		plotter.add_slider_widget(
			update_parent_opacity,
			[0.0, 1.0],
			value=0.5,
			title='親モデル透明度',
			pointa=(0.65, 0.92),
			pointb=(0.95, 0.92),
			style='modern'
		)
		
		# 透明度スライダー（子モデル）
		def update_child_opacity(value):
			if 'child' in actors and actors['child'].GetVisibility():
				actors['child'].GetProperty().SetOpacity(value)
		
		plotter.add_slider_widget(
			update_child_opacity,
			[0.0, 1.0],
			value=0.5,
			title='子モデル透明度',
			pointa=(0.65, 0.87),
			pointb=(0.95, 0.87),
			style='modern'
		)
		
		# チェックボックス（親モデル表示切り替え）
		def toggle_parent(state):
			if 'parent' in actors:
				actors['parent'].SetVisibility(state)
		
		plotter.add_checkbox_button_widget(
			toggle_parent,
			value=True,
			position=(10, 150),
			size=30,
			border_size=2,
			color_on='red',
			color_off='grey',
			background_color='white'
		)
		
		# ラベル（親モデル）
		plotter.add_text('親モデル', position=(45, 155), font_size=10, color='black')
		
		# チェックボックス（子モデル表示切り替え）
		def toggle_child(state):
			if 'child' in actors:
				actors['child'].SetVisibility(state)
		
		plotter.add_checkbox_button_widget(
			toggle_child,
			value=True,
			position=(10, 110),
			size=30,
			border_size=2,
			color_on='green',
			color_off='grey',
			background_color='white'
		)
		
		# ラベル（子モデル）
		plotter.add_text('子モデル', position=(45, 115), font_size=10, color='black')
		
		# チェックボックス（ヒートマップ表示切り替え）
		def toggle_heatmap(state):
			if 'heatmap' in actors:
				actors['heatmap'].SetVisibility(state)
		
		plotter.add_checkbox_button_widget(
			toggle_heatmap,
			value=False,
			position=(10, 70),
			size=30,
			border_size=2,
			color_on='orange',
			color_off='grey',
			background_color='white'
		)
		
		# ラベル（ヒートマップ）
		plotter.add_text('距離ヒートマップ', position=(45, 75), font_size=10, color='black')
		
		# ヘルプテキスト
		help_text = (
			'フィッティング結果\n'
			'チェックボックス: 表示切り替え\n'
			'スライダー: 透明度調整\n'
			'距離ヒートマップ: 青(近い)→赤(遠い)'
		)
		plotter.add_text(
			help_text,
			position=(10, 10),
			font_size=9,
			color='black',
			font='arial'
		)
		
		print("\n[最終結果] 可視化ウィンドウを表示中...")
		print("  - 親モデル(赤)と子モデル(緑)の透明度を調整できます")
		print("  - チェックボックスで表示/非表示を切り替えられます")
		print("  - 距離ヒートマップで位置合わせ精度を確認できます")
		
		plotter.show()
	
	def _compute_distance_heatmap_fast(self, vertices: np.ndarray, parent_mesh: pv.PolyData, max_sample_points: int = 10000, progress_callback=None) -> np.ndarray:
		"""高速な距離ヒートマップ計算（ダウンサンプリング + 補間）
		
		Args:
			vertices: 子メッシュの頂点座標 (N, 3)
			parent_mesh: 親メッシュ（PyVista）
			max_sample_points: 最大サンプル点数
			progress_callback: 進捗更新用コールバック関数(0.0~100.0)
		
		Returns:
			距離配列 (N,)
		"""
		n_vertices = len(vertices)
		print(f"  元の頂点数: {n_vertices}")

		# --- GPU一括計算パス（GPUがある場合のみ。厳密な点→三角形距離） ---
		# GPUがあれば全頂点を一括計算（サンプリング/補間不要）。CPUのみの環境では
		# 全対比較(O(N×M))は不利なので、従来のVTK逐次パスにフォールバックする。
		try:
			import frs_gpu
			if frs_gpu.has_torch() and frs_gpu.get_device() in ("cuda", "mps"):
				parent_tri = parent_mesh
				if not isinstance(parent_tri, pv.PolyData):
					parent_tri = parent_tri.extract_surface()
				parent_tri = parent_tri.triangulate()
				pverts = np.asarray(parent_tri.points, dtype=np.float64)
				pfaces_raw = np.asarray(parent_tri.faces)
				if pfaces_raw.size >= 4:
					pfaces = pfaces_raw.reshape(-1, 4)[:, 1:4].astype(np.int64)
					print(f"  [GPU] {frs_gpu.device_info()} で全{n_vertices}点を一括距離計算（三角形{len(pfaces)}）")
					d_gpu = frs_gpu.point_to_mesh_distance(np.asarray(vertices, dtype=np.float64), pverts, pfaces, batch=1024)
					if d_gpu is not None and len(d_gpu) == n_vertices:
						if progress_callback:
							progress_callback(100.0)
						print("  ✓ GPU距離計算完了")
						return d_gpu
		except Exception as e:
			print(f"  [GPU] 失敗のためCPUにフォールバック: {e}")

		# サンプリング比率を計算
		if n_vertices > max_sample_points:
			sample_rate = max_sample_points / n_vertices
			sample_indices = np.random.choice(n_vertices, max_sample_points, replace=False)
			sample_vertices = vertices[sample_indices]
			print(f"  サンプリング: {max_sample_points} 点 ({sample_rate*100:.1f}%)")
		else:
			sample_indices = np.arange(n_vertices)
			sample_vertices = vertices
			print(f"  サンプリング不要")
		
		# 距離計算（逐次処理）
		print(f"  距離計算中...")
		sample_distances = np.zeros(len(sample_vertices))
		
		progress_step = max(1, len(sample_vertices) // 20)  # 5%ごとに進捗表示
		for i, vertex in enumerate(sample_vertices):
			_, closest_point = parent_mesh.find_closest_cell(vertex, return_closest_point=True)
			sample_distances[i] = np.linalg.norm(vertex - closest_point)
			
			if (i + 1) % progress_step == 0:
				progress = (i + 1) / len(sample_vertices) * 100
				print(f"    進行状況: {progress:.0f}%")
				if progress_callback:
					progress_callback(progress)
		
		# 100%通知
		if progress_callback:
			progress_callback(100.0)
		
		# 全頂点への補間（サンプリングした場合）
		if n_vertices > max_sample_points:
			print(f"  補間中...")
			if progress_callback:
				progress_callback(101.0) # 100%超えで補間中表現
				
			try:
				from scipy.interpolate import NearestNDInterpolator
				interpolator = NearestNDInterpolator(sample_vertices, sample_distances)
				distances = interpolator(vertices)
			except ImportError:
				print("  警告: scipyが見つかりません。サンプル点のみ使用します")
				# scipyがない場合は、最近傍を単純に探す
				distances = np.zeros(n_vertices)
				for i, vertex in enumerate(vertices):
					dists = np.linalg.norm(sample_vertices - vertex, axis=1)
					nearest_idx = np.argmin(dists)
					distances[i] = sample_distances[nearest_idx]
		else:
			distances = sample_distances
		
		print(f"  ✓ 距離計算完了")
		return distances
	
	def _voxel_downsample(self, points: np.ndarray, voxel_size: float) -> np.ndarray:
		"""ボクセルグリッドによるダウンサンプリング
		
		Args:
			points: 点群 (N, 3)
			voxel_size: ボクセルサイズ (mm)
			
		Returns:
			ダウンサンプリングされた点群 (M, 3)
		"""
		# 各点をボクセルグリッドにマッピング
		voxel_indices = np.floor(points / voxel_size).astype(np.int32)
		
		# 重複を除去（各ボクセルに1点のみ残す）
		unique_voxels, inverse_indices = np.unique(voxel_indices, axis=0, return_inverse=True)
		
		# 各ボクセルの代表点を計算（平均）
		downsampled = np.zeros((len(unique_voxels), 3))
		for i in range(len(unique_voxels)):
			mask = inverse_indices == i
			downsampled[i] = np.mean(points[mask], axis=0)
		
		return downsampled
	
	def _ransac_registration(self, source: np.ndarray, target: np.ndarray) -> np.ndarray:
		"""RANSACによる位置合わせ（対応点ベース）
		
		Args:
			source: 移動させる点群 (N, 3)
			target: ターゲット点群 (M, 3)
			
		Returns:
			4x4 変換行列、失敗時はNone
		"""
		from scipy.spatial import cKDTree
		
		distance_threshold = self.ransac_distance_threshold.get()
		max_iterations = self.ransac_max_iterations.get()
		confidence = self.ransac_confidence.get()
		
		print(f"RANSACパラメータ: 距離閾値={distance_threshold}mm, 最大反復={max_iterations}, 信頼度={confidence}")
		
		# ステップ1: 最近傍対応点を事前計算（KDTreeを使用）
		print("  [1/3] 最近傍対応点を計算中...")
		target_tree = cKDTree(target)
		distances, correspondences = target_tree.query(source)
		
		# 対応点ペアを作成（距離が近い順にソート）
		valid_pairs = []
		for i, (dist, target_idx) in enumerate(zip(distances, correspondences)):
			if dist < distance_threshold * 5:  # 初期対応は緩い閾値
				valid_pairs.append((i, target_idx, dist))
		
		valid_pairs.sort(key=lambda x: x[2])  # 距離でソート
		
		if len(valid_pairs) < 100:
			print(f"  警告: 初期対応点が少なすぎます（{len(valid_pairs)}組）")
			print(f"  ヒント: ボクセルサイズを小さくするか、2つのモデルを手動で近づけてください")
			return None
		
		print(f"  対応点候補: {len(valid_pairs)}組")
		
		# ステップ2: RANSACで最良の変換を探索
		print(f"  [2/3] RANSAC実行中...")
		best_transform = None
		best_inliers = 0
		estimated_inlier_ratio = 0.1
		
		for iteration in range(max_iterations):
			# 対応点ペアからランダムに3組選択
			if len(valid_pairs) < 3:
				return None
			
			sample_pairs = np.random.choice(len(valid_pairs), min(3, len(valid_pairs)), replace=False)
			
			source_indices = [valid_pairs[idx][0] for idx in sample_pairs]
			target_indices = [valid_pairs[idx][1] for idx in sample_pairs]
			
			source_sample = source[source_indices]
			target_sample = target[target_indices]
			
			# 3点対応から変換行列を推定
			transform = self._estimate_transform_from_points(source_sample, target_sample)
			
			if transform is None:
				continue
			
			# すべての点に変換を適用
			source_transformed = self._apply_transform_to_points(source, transform)
			
			# インライア数をカウント（厳密な閾値で）
			distances, _ = target_tree.query(source_transformed)
			inliers = np.sum(distances < distance_threshold)
			
			if inliers > best_inliers:
				best_inliers = inliers
				best_transform = transform
				current_ratio = inliers / len(source)
				
				if iteration % 100 == 0 or iteration < 10:
					print(f"    反復 {iteration}: インライア数={inliers}/{len(source)} ({100*current_ratio:.1f}%)")
				
				# 適応的な早期終了
				if current_ratio > estimated_inlier_ratio:
					estimated_inlier_ratio = current_ratio
					
					if estimated_inlier_ratio > 0.01:
						w_cubed = estimated_inlier_ratio ** 3
						if w_cubed < 0.9999:
							required_iterations = int(np.log(1 - confidence) / np.log(1 - w_cubed))
							
							if iteration >= required_iterations:
								print(f"    信頼度 {confidence} を満たす反復回数を達成")
								break
		
		# ステップ3: 結果の評価
		print(f"  [3/3] 結果評価...")
		if best_transform is not None:
			final_ratio = best_inliers / len(source)
			print(f"RANSAC完了: インライア数={best_inliers}/{len(source)} ({100*final_ratio:.1f}%)")
			
			if final_ratio < 0.3:
				print(f"  警告: インライア率が低いです（{100*final_ratio:.1f}%）")
				print(f"  ヒント: パラメータ調整が必要かもしれません")
				# それでも結果を返す（ICPで改善できる可能性がある）
		else:
			print("RANSACに失敗しました")
		
		return best_transform
	
	def _estimate_transform_from_points(self, source: np.ndarray, target: np.ndarray) -> np.ndarray:
		"""3点対応から変換行列を推定（SVDベース）
		
		Args:
			source: ソース点群 (3, 3)
			target: ターゲット点群 (3, 3)
			
		Returns:
			4x4 変換行列、失敗時はNone
		"""
		try:
			# 重心を計算
			source_center = np.mean(source, axis=0)
			target_center = np.mean(target, axis=0)
			
			# 中心化
			source_centered = source - source_center
			target_centered = target - target_center
			
			# 共分散行列
			H = source_centered.T @ target_centered
			
			# SVD
			U, S, Vt = np.linalg.svd(H)
			R = Vt.T @ U.T
			
			# 反射を防ぐ
			if np.linalg.det(R) < 0:
				Vt[-1, :] *= -1
				R = Vt.T @ U.T
			
			# 平行移動ベクトル
			t = target_center - R @ source_center
			
			# 4x4変換行列
			transform = np.eye(4)
			transform[:3, :3] = R
			transform[:3, 3] = t
			
			return transform
			
		except Exception:
			return None
	
	def _apply_transform_to_points(self, points: np.ndarray, transform: np.ndarray) -> np.ndarray:
		"""点群に変換行列を適用
		
		Args:
			points: 点群 (N, 3)
			transform: 4x4 変換行列
			
		Returns:
			変換後の点群 (N, 3)
		"""
		ones = np.ones((len(points), 1))
		points_homogeneous = np.hstack([points, ones])
		points_transformed = (transform @ points_homogeneous.T).T[:, :3]
		return points_transformed

	def on_visualize_fitting_parent(self) -> None:
		"""親モデルと親任意領域を同時に可視化"""
		parent_model_path = self.fitting_parent_model_path.get()
		parent_region_path = self.fitting_parent_region_path.get()
		
		if not parent_model_path and not parent_region_path:
			messagebox.showerror("エラー", "親モデルまたは親任意領域ファイルを選択してください。")
			return
		
		try:
			# ディスプレイサイズの取得
			screen_width = self.winfo_screenwidth()
			screen_height = self.winfo_screenheight()
			window_width = int(screen_width * 0.9)
			window_height = int(screen_height * 0.9)
			
			plotter = pv.Plotter(title="親モデルの可視化", window_size=(window_width, window_height))
			plotter.set_background("white")
			
			model_actor = None
			region_actor = None
			
			# 親モデル（若干透明）
			if parent_model_path:
				parent_model = pv.read(parent_model_path)
				model_actor = plotter.add_mesh(
					parent_model, 
					color='lightcoral', 
					opacity=0.4,
					smooth_shading=True,
					show_edges=False, 
					label='親モデル'
				)
			
			# 親任意領域（不透明）
			if parent_region_path:
				parent_region = pv.read(parent_region_path)
				region_actor = plotter.add_mesh(
					parent_region, 
					color='red', 
					opacity=1.0,
					smooth_shading=True,
					show_edges=False, 
					label='親任意領域'
				)
			
			# チェックボックスとスライダー
			def toggle_model(state):
				if model_actor:
					model_actor.SetVisibility(state)
			
			def toggle_region(state):
				if region_actor:
					region_actor.SetVisibility(state)
			
			def update_model_opacity(value):
				if model_actor:
					model_actor.GetProperty().SetOpacity(value)
			
			def update_region_opacity(value):
				if region_actor:
					region_actor.GetProperty().SetOpacity(value)
			
			# UI配置
			checkbox_y = 10
			if model_actor:
				plotter.add_text("親モデル", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_model, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="lightcoral", color_off="white")
				checkbox_y += 30
			
			if region_actor:
				plotter.add_text("親任意領域", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_region, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="red", color_off="white")
			
			# 透明度スライダー
			if model_actor:
				plotter.add_slider_widget(
					update_model_opacity,
					rng=[0.0, 1.0],
					value=0.4,
					title="親モデル ",
					pointa=(0.05, 0.92),
					pointb=(0.15, 0.92),
					style='modern',
					color="lightcoral"
				)
			
			if region_actor:
				plotter.add_slider_widget(
					update_region_opacity,
					rng=[0.0, 1.0],
					value=1.0,
					title="親任意領域 ",
					pointa=(0.05, 0.85),
					pointb=(0.15, 0.85),
					style='modern',
					color="red"
				)
			
			plotter.add_legend()
			plotter.show()
			
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_visualize_fitting_child(self) -> None:
		"""子モデル、子任意領域、子特徴点を同時に可視化"""
		child_model_path = self.fitting_child_model_path.get()
		child_region_path = self.fitting_child_region_path.get()
		child_pp_path = self.fitting_child_pp_path.get()
		
		if not child_model_path and not child_region_path and not child_pp_path:
			messagebox.showerror("エラー", "子モデル、子任意領域、または子特徴点ファイルを選択してください。")
			return
		
		try:
			# ディスプレイサイズの取得
			screen_width = self.winfo_screenwidth()
			screen_height = self.winfo_screenheight()
			window_width = int(screen_width * 0.9)
			window_height = int(screen_height * 0.9)
			
			plotter = pv.Plotter(title="子モデルの可視化", window_size=(window_width, window_height))
			plotter.set_background("white")
			
			model_actor = None
			region_actor = None
			points_actor = None
			
			# 子モデル（若干透明）
			if child_model_path:
				child_model = pv.read(child_model_path)
				model_actor = plotter.add_mesh(
					child_model, 
					color='lightblue', 
					opacity=0.4,
					smooth_shading=True,
					show_edges=False, 
					label='子モデル'
				)
			
			# 子任意領域（不透明）
			if child_region_path:
				child_region = pv.read(child_region_path)
				region_actor = plotter.add_mesh(
					child_region, 
					color='blue', 
					opacity=1.0,
					smooth_shading=True,
					show_edges=False, 
					label='子任意領域'
				)
			
			# 子特徴点（PP）- すべての点を表示（LMNも含む）
			if child_pp_path:
				points, labels = self._parse_pp_file(child_pp_path)
				
				if len(points) > 0:
					# 特徴点を表示
					points_actor = plotter.add_points(
						points,
						color='green',
						point_size=15,
						render_points_as_spheres=True,
						label='子特徴点'
					)
					
					# ラベル表示
					for point, label in zip(points, labels):
						plotter.add_point_labels(
							[point],
							[label],
							font_size=12,
							text_color='green',
							point_size=0,
							always_visible=True,
							shape_opacity=0.5,
							bold=True
						)
			
			# チェックボックスとスライダー
			def toggle_model(state):
				if model_actor:
					model_actor.SetVisibility(state)
			
			def toggle_region(state):
				if region_actor:
					region_actor.SetVisibility(state)
			
			def update_model_opacity(value):
				if model_actor:
					model_actor.GetProperty().SetOpacity(value)
			
			def update_region_opacity(value):
				if region_actor:
					region_actor.GetProperty().SetOpacity(value)
			
			# UI配置
			checkbox_y = 10
			if model_actor:
				plotter.add_text("子モデル", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_model, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="lightblue", color_off="white")
				checkbox_y += 30
			
			if region_actor:
				plotter.add_text("子任意領域", position=(10, checkbox_y), font_size=10, color="black")
				plotter.add_checkbox_button_widget(toggle_region, value=True, position=(100, checkbox_y+2), 
													size=20, border_size=1, color_on="blue", color_off="white")
			
			# 透明度スライダー
			if model_actor:
				plotter.add_slider_widget(
					update_model_opacity,
					rng=[0.0, 1.0],
					value=0.4,
					title="子モデル ",
					pointa=(0.05, 0.92),
					pointb=(0.15, 0.92),
					style='modern',
					color="lightblue"
				)
			
			if region_actor:
				plotter.add_slider_widget(
					update_region_opacity,
					rng=[0.0, 1.0],
					value=1.0,
					title="子任意領域 ",
					pointa=(0.05, 0.85),
					pointb=(0.15, 0.85),
					style='modern',
					color="blue"
				)
			
			plotter.add_legend()
			plotter.show()
			
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_visualize_org(self) -> None:
		"""ORG: RobotScan モデルと座標ポイントを可視化"""
		model_path = self.org_model_path.get().strip()
		pp_path = self.org_pp_path.get().strip()
		
		if not model_path or not pp_path:
			messagebox.showerror("エラー", "モデルファイルと座標ポイントファイルを選択してください。")
			return
		
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista"
			)
			return
		
		# PP読み込み
		try:
			points, labels = self._parse_pp_file(pp_path)
		except Exception as e:
			messagebox.showerror("エラー", f"PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# CL点と座標系の計算
		# PPファイルの順序: L, M, N
		C_cl_origin = None
		C_cl_x_axis = None
		C_cl_y_axis = None
		C_cl_z_axis = None
		C_o_origin = None
		C_o_x_axis = None
		C_o_y_axis = None
		C_o_z_axis = None
		L_prime_point = None
		
		if len(points) >= 3:
			L_point = points[0]  # L点
			M_point = points[1]  # M点
			N_point = points[2]  # N点
			
			# MとNの中点を計算（CL点）
			CL_point = np.array([
				(M_point[0] + N_point[0]) / 2.0,
				(M_point[1] + N_point[1]) / 2.0,
				(M_point[2] + N_point[2]) / 2.0
			])
			
			# LMベクトルを作成
			LM_vector = M_point - L_point
			LM_unit = LM_vector / np.linalg.norm(LM_vector)
			
			# NからLM直線への垂線の足L'を計算
			# L'はLM上の点で、NL'がLMに垂直
			LN_vector = N_point - L_point
			projection_length = np.dot(LN_vector, LM_unit)
			L_prime_point = L_point + projection_length * LM_unit
			
			# C_cl座標系を構築
			# 原点: CL
			C_cl_origin = CL_point
			
			# X軸: M → L'
			x_vec = L_prime_point - M_point
			C_cl_x_axis = x_vec / np.linalg.norm(x_vec)
			
			# Z軸: N → L'
			z_vec = L_prime_point - N_point
			C_cl_z_axis = z_vec / np.linalg.norm(z_vec)
			
			# Y軸: Z × X
			C_cl_y_axis = np.cross(C_cl_z_axis, C_cl_x_axis)
			C_cl_y_axis = C_cl_y_axis / np.linalg.norm(C_cl_y_axis)
			
			# C_o座標系: C_clをY軸方向に250mm平行移動
			C_o_origin = C_cl_origin + -250.0 * C_cl_y_axis
			C_o_x_axis = C_cl_x_axis
			C_o_y_axis = C_cl_y_axis
			C_o_z_axis = C_cl_z_axis
			
			# CL点とL'点を追加
			points = np.vstack([points, CL_point, L_prime_point])
			labels = labels + ['CL', "L'"]
		
		# メッシュ読み込み
		try:
			mesh = pv.read(model_path)
		except Exception as e:
			messagebox.showerror("エラー", f"モデルの読み込みに失敗しました:\n{e}")
			return
		
		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)
		
		# プロッタ作成
		plotter = pv.Plotter(title="ORG - RobotScan データ可視化", window_size=(window_width, window_height))
		plotter.set_background("white")
		
		# メッシュ表示
		mesh_actor = plotter.add_mesh(
			mesh,
			color='lightgray',
			smooth_shading=True,
			show_edges=False,
			opacity=0.8,
			label='RobotScan モデル'
		)
		
		# 座標ポイント表示
		points_actor = plotter.add_points(
			points,
			color='red',
			point_size=15,
			render_points_as_spheres=True,
			label='座標ポイント'
		)
		
		# ラベル表示（各点の名前）
		for i, (point, label) in enumerate(zip(points, labels)):
			plotter.add_point_labels(
				[point],
				[label],
				font_size=12,
				text_color='black',
				point_size=0,
				always_visible=True,
				shape_opacity=0.5
			)
		
		# C_cl座標系の表示
		c_cl_actors = []
		if C_cl_origin is not None:
			axis_length = 50.0  # 座標軸の長さ
			
			# X軸（赤）
			x_line = pv.Line(C_cl_origin, C_cl_origin + axis_length * C_cl_x_axis)
			c_cl_actors.append(plotter.add_mesh(x_line, color='red', line_width=5, label='C_cl X'))
			
			# Y軸（緑）
			y_line = pv.Line(C_cl_origin, C_cl_origin + axis_length * C_cl_y_axis)
			c_cl_actors.append(plotter.add_mesh(y_line, color='green', line_width=5, label='C_cl Y'))
			
			# Z軸（青）
			z_line = pv.Line(C_cl_origin, C_cl_origin + axis_length * C_cl_z_axis)
			c_cl_actors.append(plotter.add_mesh(z_line, color='blue', line_width=5, label='C_cl Z'))
			
			# 原点（CL）に小さな球
			c_cl_actors.append(plotter.add_mesh(
				pv.Sphere(radius=3.0, center=C_cl_origin),
				color='yellow',
				label='C_cl原点'
			))
		
		# C_o座標系の表示
		c_o_actors = []
		if C_o_origin is not None:
			axis_length = 50.0
			
			# X軸（濃い赤）
			x_line = pv.Line(C_o_origin, C_o_origin + axis_length * C_o_x_axis)
			c_o_actors.append(plotter.add_mesh(x_line, color='darkred', line_width=5, label='C_o X'))
			
			# Y軸（濃い緑）
			y_line = pv.Line(C_o_origin, C_o_origin + axis_length * C_o_y_axis)
			c_o_actors.append(plotter.add_mesh(y_line, color='darkgreen', line_width=5, label='C_o Y'))
			
			# Z軸（濃い青）
			z_line = pv.Line(C_o_origin, C_o_origin + axis_length * C_o_z_axis)
			c_o_actors.append(plotter.add_mesh(z_line, color='darkblue', line_width=5, label='C_o Z'))
			
			# 原点に小さな球
			c_o_actors.append(plotter.add_mesh(
				pv.Sphere(radius=3.0, center=C_o_origin),
				color='orange',
				label='C_o原点'
			))
		
		# ORG座標をC_o座標系で表示
		org_actor = None
		if C_o_origin is not None and C_o_x_axis is not None:
			try:
				# ORG座標を取得（C_o座標系での座標）
				org_x = self.org_coord_x.get()
				org_y = self.org_coord_y.get()
				org_z = self.org_coord_z.get()
				
				# C_o座標系からワールド座標系に変換
				# ORG_world = C_o_origin + org_x * C_o_x_axis + org_y * C_o_y_axis + org_z * C_o_z_axis
				org_world = C_o_origin + org_x * C_o_x_axis + org_y * C_o_y_axis + org_z * C_o_z_axis
				
				# ORG点を表示（紫の大きな球）
				org_actor = plotter.add_mesh(
					pv.Sphere(radius=5.0, center=org_world),
					color='purple',
					label='ORG'
				)
				
				# ORGラベル
				plotter.add_point_labels(
					[org_world],
					['ORG'],
					font_size=14,
					text_color='purple',
					point_size=0,
					always_visible=True,
					shape_opacity=0.7,
					bold=True
				)
			except (ValueError, TypeError):
				pass  # 座標入力が不正な場合は無視
		
		# チェックボックス
		def toggle_mesh(state):
			mesh_actor.SetVisibility(state)
		
		def toggle_points(state):
			points_actor.SetVisibility(state)
		
		def toggle_c_cl(state):
			for actor in c_cl_actors:
				actor.SetVisibility(state)
		
		def toggle_c_o(state):
			for actor in c_o_actors:
				actor.SetVisibility(state)
		
		def update_mesh_opacity(value):
			mesh_actor.GetProperty().SetOpacity(value)
		
		plotter.add_text("Model", position=(10, 10), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_mesh, value=True, position=(100, 12), size=20, border_size=1, color_on="lightgray", color_off="white")
		
		plotter.add_text("Points", position=(10, 40), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_points, value=True, position=(100, 42), size=20, border_size=1, color_on="red", color_off="white")
		
		if c_cl_actors:
			plotter.add_text("C_cl", position=(10, 70), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_c_cl, value=True, position=(100, 72), size=20, border_size=1, color_on="yellow", color_off="white")
		
		if c_o_actors:
			plotter.add_text("C_o", position=(10, 100), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_c_o, value=True, position=(100, 102), size=20, border_size=1, color_on="orange", color_off="white")
		
		if org_actor:
			def toggle_org(state):
				org_actor.SetVisibility(state)
			
			plotter.add_text("ORG", position=(10, 130), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_org, value=True, position=(100, 132), size=20, border_size=1, color_on="purple", color_off="white")
		
		# 透明度スライダー
		plotter.add_slider_widget(
			update_mesh_opacity,
			rng=[0.0, 1.0],
			value=0.8,
			title="Model ",
			pointa=(0.05, 0.02),
			pointb=(0.15, 0.02),
			style='modern',
			color="lightgray"
		)
		
		plotter.add_legend()
		plotter.show()
		
		# ウィンドウを閉じた後、ORG座標の保存を確認
		if C_o_origin is not None and org_actor is not None:
			try:
				# ORG座標を取得
				org_x = self.org_coord_x.get()
				org_y = self.org_coord_y.get()
				org_z = self.org_coord_z.get()
				
				# C_o座標系からワールド座標系に変換
				org_world = C_o_origin + org_x * C_o_x_axis + org_y * C_o_y_axis + org_z * C_o_z_axis
				
				# 保存確認ダイアログ
				result = messagebox.askyesno(
					"ORG座標の保存",
					f"ORG座標をPPファイルに保存しますか？\n\n"
					f"C_o座標系: ({org_x:.2f}, {org_y:.2f}, {org_z:.2f})\n"
					f"ワールド座標: ({org_world[0]:.2f}, {org_world[1]:.2f}, {org_world[2]:.2f})"
				)
				
				if result:
					self._save_org_to_pp(pp_path, org_world)
			except (ValueError, TypeError):
				pass  # 座標が不正な場合は何もしない

	def _save_org_to_pp(self, pp_path: str, org_world: np.ndarray) -> None:
		"""ORG座標をPPファイルに追加保存（混在形式対応：XML + テキスト行を保持）"""
		try:
			import xml.etree.ElementTree as ET
			from datetime import datetime
			
			# ファイル全体を読み込み、XML部分とテキスト行部分を分離
			with open(pp_path, 'r', encoding='utf-8') as f:
				content = f.read()
			
			# XML終了タグを探す
			xml_end = content.find('</PickedPoints>')
			if xml_end == -1:
				raise ValueError("PickedPoints終了タグが見つかりません")
			
			xml_end += len('</PickedPoints>')
			xml_part = content[:xml_end]
			text_part = content[xml_end:].strip()
			
			# XML部分をパース
			root = ET.fromstring(xml_part)
			
			# O点が既に存在するか確認
			existing_o = None
			for point in root.findall('point'):
				if point.get('name') == 'O':
					existing_o = point
					break
			
			if existing_o is not None:
				# 既存のO点を更新
				existing_o.set('x', f"{org_world[0]:.5f}")
				existing_o.set('y', f"{org_world[1]:.5f}")
				existing_o.set('z', f"{org_world[2]:.5f}")
			else:
				# 新しいO点を追加
				new_point = ET.Element('point')
				new_point.set('x', f"{org_world[0]:.5f}")
				new_point.set('y', f"{org_world[1]:.5f}")
				new_point.set('z', f"{org_world[2]:.5f}")
				new_point.set('active', '1')
				new_point.set('name', 'O')
				root.append(new_point)
			
			# DocumentDataのDateTime要素を更新
			doc_data = root.find('DocumentData')
			if doc_data is not None:
				date_time = doc_data.find('DateTime')
				if date_time is not None:
					now = datetime.now()
					date_time.set('date', now.strftime('%Y-%m-%d'))
					date_time.set('time', now.strftime('%H:%M:%S'))
			
			# XML部分を文字列に変換
			tree = ET.ElementTree(root)
			xml_str = io.BytesIO()
			tree.write(xml_str, encoding='utf-8', xml_declaration=True)
			xml_output = xml_str.getvalue().decode('utf-8')
			
			# テキスト行部分を結合して書き戻し
			final_content = xml_output
			if text_part:
				final_content += '\n' + text_part + '\n'
			
			with open(pp_path, 'w', encoding='utf-8') as f:
				f.write(final_content)
			
			messagebox.showinfo(
				"保存完了",
				f"ORG座標をO点として保存しました:\n{pp_path}"
			)
		except Exception as e:
			messagebox.showerror("保存エラー", f"PPファイルへの保存に失敗しました:\n{e}")

	def on_visualize_a_b(self) -> None:
		"""A-B: 2つのモデルを同時に可視化"""
		model_a_path = self.a_b_model_a_path.get().strip()
		model_b_path = self.a_b_model_b_path.get().strip()

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path)
			mesh_b = pv.read(model_b_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="A-B Visualization")
			plotter.set_background("white")
			plotter.add_mesh(mesh_a, color="lightgray", opacity=0.6, show_edges=False, label="Model A")
			plotter.add_mesh(mesh_b, color="tomato", opacity=0.6, show_edges=False, label="Model B")
			plotter.add_legend()
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_execute_a_b_heatmap(self) -> None:
		"""A-B: AからBメッシュまでの距離をヒートマップ表示し、点群をExcel出力"""
		model_a_path = self.a_b_model_a_path.get().strip()
		model_b_path = self.a_b_model_b_path.get().strip()

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path).triangulate().clean()
			mesh_b = pv.read(model_b_path).triangulate().clean()
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		if mesh_a.n_points == 0 or mesh_b.n_points == 0:
			messagebox.showerror("エラー", "メッシュの頂点数が0です。ファイルを確認してください。")
			return

		try:
			max_points = int(self.sample_points.get()) if hasattr(self, "sample_points") else 10000
			distances = self._compute_distance_heatmap_fast(mesh_a.points, mesh_b, max_sample_points=max_points)
		except Exception as e:
			messagebox.showerror("計算失敗", f"距離計算に失敗しました:\n{e}")
			return

		try:
			heatmap_mesh = mesh_a.copy(deep=False)
			heatmap_mesh['distance'] = np.array(distances, dtype=float)
		except Exception as e:
			messagebox.showerror("エラー", f"ヒートマップ生成に失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="A-B Distance Heatmap")
			plotter.set_background("white")
			min_d = float(np.min(distances))
			max_d = float(np.max(distances))
			plotter.add_mesh(
				heatmap_mesh,
				scalars="distance",
				cmap="turbo",
				clim=[min_d, max_d],
				show_edges=False,
				scalar_bar_args={
					"title": "Distance (mm)",
					"n_labels": 5,
					"fmt": "%.2f",
				},
			)
			plotter.add_mesh(mesh_b, color="lightgray", opacity=0.35, show_edges=False, label="Model B")
			plotter.add_legend()
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")
			return

		# Excel出力
		if not messagebox.askyesno("Excel出力", "点群情報をExcelで保存しますか？"):
			return

		save_path = filedialog.asksaveasfilename(
			title="点群情報を保存",
			defaultextension=".xlsx",
			filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
		)
		if not save_path:
			return

		try:
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "A_to_B_Distance"
			ws.append(["index", "x", "y", "z", "distance_mm"])
			for i, (pt, d) in enumerate(zip(mesh_a.points, distances)):
				ws.append([i, float(pt[0]), float(pt[1]), float(pt[2]), float(d)])
			wb.save(save_path)
			messagebox.showinfo("保存完了", f"保存しました:\n{save_path}")
		except Exception as e:
			messagebox.showerror("保存失敗", f"Excel保存に失敗しました:\n{e}")

	def on_visualize_heatmap_models(self) -> None:
		"""HeatMap: 2つのモデルを同時に可視化"""
		model_a_path = self.heatmap_model_a_path.get().strip()
		model_b_path = self.heatmap_model_b_path.get().strip()
		show_model_a = bool(self.heatmap_show_model_a.get())
		show_model_b = bool(self.heatmap_show_model_b.get())

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		if not (show_model_a or show_model_b):
			messagebox.showwarning("表示設定", "表示するモデルを1つ以上選択してください。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path)
			mesh_b = pv.read(model_b_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="HeatMap Models Visualization")
			plotter.set_background("white")
			if show_model_a:
				plotter.add_mesh(mesh_a, color="lightgray", opacity=0.6, show_edges=False, label="Model A (reference)")
			if show_model_b:
				plotter.add_mesh(mesh_b, color="tomato", opacity=0.6, show_edges=False, label="Model B (target)")
			plotter.add_legend()
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")

	def on_execute_heatmap_b_to_a(self) -> None:
		"""HeatMap: BからA表面までの距離をヒートマップ表示（クリックで距離表示）"""
		model_a_path = self.heatmap_model_a_path.get().strip()
		model_b_path = self.heatmap_model_b_path.get().strip()
		show_model_a = bool(self.heatmap_show_model_a.get())
		show_model_b = bool(self.heatmap_show_model_b.get())

		if not model_a_path or not model_b_path:
			messagebox.showwarning("入力不足", "モデルAとモデルBを選択してください。")
			return

		if not show_model_b:
			messagebox.showwarning("表示設定", "ヒートマップ表示にはモデルBの表示が必要です。")
			return

		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"エラー",
				"PyVistaがインストールされていません。\n"
				"以下のコマンドでインストールしてください:\n"
				"pip install pyvista",
			)
			return

		try:
			mesh_a = pv.read(model_a_path).triangulate().clean()
			mesh_b = pv.read(model_b_path).triangulate().clean()
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"モデルの読み込みに失敗しました:\n{e}")
			return

		if mesh_a.n_points == 0 or mesh_b.n_points == 0:
			messagebox.showerror("エラー", "メッシュの頂点数が0です。ファイルを確認してください。")
			return

		try:
			max_points = int(self.sample_points.get()) if hasattr(self, "sample_points") else 10000
			distances = self._compute_distance_heatmap_fast(mesh_b.points, mesh_a, max_sample_points=max_points)
		except Exception as e:
			messagebox.showerror("計算失敗", f"距離計算に失敗しました:\n{e}")
			return

		try:
			heatmap_mesh = mesh_b.copy(deep=False)
			heatmap_mesh["distance"] = np.array(distances, dtype=float)
		except Exception as e:
			messagebox.showerror("エラー", f"ヒートマップ生成に失敗しました:\n{e}")
			return

		try:
			plotter = pv.Plotter(title="HeatMap (B→A Distance)")
			plotter.set_background("white")
			min_d = float(np.min(distances))
			max_d = float(np.max(distances))
			plotter.add_mesh(
				heatmap_mesh,
				scalars="distance",
				cmap="turbo",
				clim=[min_d, max_d],
				show_edges=False,
				label="Model B (heatmap)",
				scalar_bar_args={
					"title": "Distance B→A (mm)",
					"n_labels": 5,
					"fmt": "%.2f",
				},
			)
			if show_model_a:
				plotter.add_mesh(mesh_a, color="lightgray", opacity=0.35, show_edges=False, label="Model A (reference)", pickable=False)
			plotter.add_legend()
			plotter.add_text("Click a point to see distance (mm)", position=(10, 10), font_size=9, color="black", name="pick_hint")
			pick_text_actor = plotter.add_text("Picked distance: --", position=(10, 28), font_size=10, color="black", name="pick_value")

			def _on_pick(*args):
				point = None
				if len(args) >= 1:
					candidate = args[0]
					if hasattr(candidate, "GetPickPosition"):
						try:
							pos = candidate.GetPickPosition()
							point = np.array([pos[0], pos[1], pos[2]], dtype=float)
						except Exception:
							point = None
					elif isinstance(candidate, (list, tuple, np.ndarray)) and len(candidate) >= 3:
						point = np.array(candidate[:3], dtype=float)
					elif isinstance(candidate, pv.PolyData):
						if candidate.n_points > 0:
							point = candidate.points[0]
				if point is None:
					return
				try:
					idx = heatmap_mesh.find_closest_point(point)
					d = float(heatmap_mesh["distance"][idx])
					pick_text_actor.SetInput(f"Picked distance: {d:.3f} mm")
					plotter.render()
				except Exception:
					pass

			plotter.enable_point_picking(callback=_on_pick, show_point=True, use_picker=True, show_message=False)
			plotter.show()
		except Exception as e:
			messagebox.showerror("エラー", f"可視化に失敗しました:\n{e}")
			return

	def on_c_fix_visualize(self) -> None:
		"""C. Fixタブでの可視化：近位・遠位モデルと全特徴点を表示"""
		# ファイルパスを取得
		prox_model_path = self.prox_model_path.get().strip()
		prox_pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		prox_pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		dist_model_path = self.dist_model_path.get().strip()
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		
		if not all([prox_model_path, prox_pp_abcd_path, prox_pp_olmn_path,
					dist_model_path, dist_pp_abc_path, dist_pp_olmn_path]):
			messagebox.showwarning("入力不足", "Simulatorタブですべてのファイルを選択してください。")
			return
		
		# ライブラリの確認
		try:
			import pyvista as pv
		except Exception:
			messagebox.showerror(
				"ライブラリ未導入",
				"PyVista のインストールが必要です。以下を順に実行してください:\n\n"
				"pip install vtk\n"
				"pip install pyvista",
			)
			return
		
		# PP読み込み（近位）
		try:
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(prox_pp_abcd_path)
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(prox_pp_olmn_path)
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"近位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# PP読み込み（遠位）
		try:
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"遠位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# メッシュ読み込み（近位）
		try:
			prox_mesh = pv.read(prox_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"近位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# メッシュ読み込み（遠位）
		try:
			dist_mesh = pv.read(dist_model_path)
		except Exception as e:
			messagebox.showerror("読み込み失敗", f"遠位モデルの読み込みに失敗しました:\n{e}")
			return
		
		# 近位CL座標系の構築
		prox_C_cl_origin = None
		prox_C_cl_x_axis = None
		prox_C_cl_y_axis = None
		prox_C_cl_z_axis = None
		prox_CL_point = None
		prox_L_prime_point = None
		
		# 近位: L, M, N点を探す
		prox_L_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'L'), None)
		prox_M_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'M'), None)
		prox_N_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'N'), None)
		
		if prox_L_idx is not None and prox_M_idx is not None and prox_N_idx is not None:
			prox_L_point = prox_points[prox_L_idx]
			prox_M_point = prox_points[prox_M_idx]
			prox_N_point = prox_points[prox_N_idx]
			
			# MとNの中点を計算（CL点）
			prox_CL_point = np.array([
				(prox_M_point[0] + prox_N_point[0]) / 2.0,
				(prox_M_point[1] + prox_N_point[1]) / 2.0,
				(prox_M_point[2] + prox_N_point[2]) / 2.0
			])
			
			# LMベクトルを作成
			prox_LM_vector = prox_M_point - prox_L_point
			prox_LM_unit = prox_LM_vector / np.linalg.norm(prox_LM_vector)
			
			# NからLM直線への垂線の足L'を計算
			prox_LN_vector = prox_N_point - prox_L_point
			prox_projection_length = np.dot(prox_LN_vector, prox_LM_unit)
			prox_L_prime_point = prox_L_point + prox_projection_length * prox_LM_unit
			
			# C_cl座標系を構築
			prox_C_cl_origin = prox_CL_point
			
			# X軸: M → L'
			prox_x_vec = prox_L_prime_point - prox_M_point
			prox_C_cl_x_axis = prox_x_vec / np.linalg.norm(prox_x_vec)
			
			# Z軸: N → L'
			prox_z_vec = prox_L_prime_point - prox_N_point
			prox_C_cl_z_axis = prox_z_vec / np.linalg.norm(prox_z_vec)
			
			# Y軸: Z × X
			prox_C_cl_y_axis = np.cross(prox_C_cl_z_axis, prox_C_cl_x_axis)
			prox_C_cl_y_axis = prox_C_cl_y_axis / np.linalg.norm(prox_C_cl_y_axis)
			
			# CL点とL'点を点群に追加
			prox_points = np.vstack([prox_points, prox_CL_point, prox_L_prime_point])
			prox_labels = prox_labels + ['CL', "L'"]
		
		# 修正量を取得（近位用）
		correction_x = self.c_fix_correction_x.get()
		correction_y = self.c_fix_correction_y.get()
		correction_z = self.c_fix_correction_z.get()
		
		# 近位のA'B'C'D'点を計算
		prox_corrected_points = None
		prox_corrected_labels = []
		if prox_C_cl_origin is not None:
			# A, B, C, D点を探す
			prox_ABCD_indices = []
			prox_ABCD_labels_found = []
			for target_label in ['A', 'B', 'C', 'D']:
				idx = next((i for i, lbl in enumerate(prox_labels) if lbl == target_label), None)
				if idx is not None:
					prox_ABCD_indices.append(idx)
					prox_ABCD_labels_found.append(target_label)
			
			if prox_ABCD_indices:
				# CL座標系に沿って修正量を適用
				correction_vector = (correction_x * prox_C_cl_x_axis + 
									 correction_y * prox_C_cl_y_axis + 
									 correction_z * prox_C_cl_z_axis)
				
				prox_corrected_points = []
				for idx, label in zip(prox_ABCD_indices, prox_ABCD_labels_found):
					original_point = prox_points[idx]
					corrected_point = original_point + correction_vector
					prox_corrected_points.append(corrected_point)
					prox_corrected_labels.append(f"{label}'")
				
				prox_corrected_points = np.array(prox_corrected_points)
		
		# 遠位CL座標系の構築
		dist_C_cl_origin = None
		dist_C_cl_x_axis = None
		dist_C_cl_y_axis = None
		dist_C_cl_z_axis = None
		dist_CL_point = None
		dist_L_prime_point = None
		
		# 遠位: L, M, N点を探す
		dist_L_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'L'), None)
		dist_M_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'M'), None)
		dist_N_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'N'), None)
		
		if dist_L_idx is not None and dist_M_idx is not None and dist_N_idx is not None:
			dist_L_point = dist_points[dist_L_idx]
			dist_M_point = dist_points[dist_M_idx]
			dist_N_point = dist_points[dist_N_idx]
			
			# MとNの中点を計算（CL点）
			dist_CL_point = np.array([
				(dist_M_point[0] + dist_N_point[0]) / 2.0,
				(dist_M_point[1] + dist_N_point[1]) / 2.0,
				(dist_M_point[2] + dist_N_point[2]) / 2.0
			])
			
			# LMベクトルを作成
			dist_LM_vector = dist_M_point - dist_L_point
			dist_LM_unit = dist_LM_vector / np.linalg.norm(dist_LM_vector)
			
			# NからLM直線への垂線の足L'を計算
			dist_LN_vector = dist_N_point - dist_L_point
			dist_projection_length = np.dot(dist_LN_vector, dist_LM_unit)
			dist_L_prime_point = dist_L_point + dist_projection_length * dist_LM_unit
			
			# C_cl座標系を構築
			dist_C_cl_origin = dist_CL_point
			
			# X軸: M → L'
			dist_x_vec = dist_L_prime_point - dist_M_point
			dist_C_cl_x_axis = dist_x_vec / np.linalg.norm(dist_x_vec)
			
			# Z軸: N → L'
			dist_z_vec = dist_L_prime_point - dist_N_point
			dist_C_cl_z_axis = dist_z_vec / np.linalg.norm(dist_z_vec)
			
			# Y軸: Z × X
			dist_C_cl_y_axis = np.cross(dist_C_cl_z_axis, dist_C_cl_x_axis)
			dist_C_cl_y_axis = dist_C_cl_y_axis / np.linalg.norm(dist_C_cl_y_axis)
			
			# CL点とL'点を点群に追加
			dist_points = np.vstack([dist_points, dist_CL_point, dist_L_prime_point])
			dist_labels = dist_labels + ['CL', "L'"]
		
		# 遠位のA'B'C'点を計算（遠位はA, B, Cのみ）
		dist_corrected_points = None
		dist_corrected_labels = []
		if dist_C_cl_origin is not None:
			# A, B, C点を探す
			dist_ABC_indices = []
			dist_ABC_labels_found = []
			for target_label in ['A', 'B', 'C']:
				idx = next((i for i, lbl in enumerate(dist_labels) if lbl == target_label), None)
				if idx is not None:
					dist_ABC_indices.append(idx)
					dist_ABC_labels_found.append(target_label)
			
			if dist_ABC_indices:
				# CL座標系に沿って修正量を適用
				correction_vector = (correction_x * dist_C_cl_x_axis + 
									 correction_y * dist_C_cl_y_axis + 
									 correction_z * dist_C_cl_z_axis)
				
				dist_corrected_points = []
				for idx, label in zip(dist_ABC_indices, dist_ABC_labels_found):
					original_point = dist_points[idx]
					corrected_point = original_point + correction_vector
					dist_corrected_points.append(corrected_point)
					dist_corrected_labels.append(f"{label}'")
				
				dist_corrected_points = np.array(dist_corrected_points)
		
		# ディスプレイサイズの取得
		screen_width = self.winfo_screenwidth()
		screen_height = self.winfo_screenheight()
		window_width = int(screen_width * 0.9)
		window_height = int(screen_height * 0.9)
		
		# プロッタ作成
		plotter = pv.Plotter(title="C. Fix - 座標系確認", window_size=(window_width, window_height))
		plotter.set_background("white")
		
		# 近位メッシュ（水色）
		prox_mesh_actor = plotter.add_mesh(prox_mesh, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=0.5)
		
		# 遠位メッシュ（ピンク）
		dist_mesh_actor = plotter.add_mesh(dist_mesh, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=0.5)
		
		# 近位点群（水色）
		prox_points_actor = plotter.add_points(prox_points, color=self.prox_color, point_size=15, render_points_as_spheres=True)
		
		# 近位点ラベル
		prox_label_actors = []
		for pt, label in zip(prox_points, prox_labels):
			label_actor = plotter.add_point_labels([pt], [f"Prox_{label}"], point_size=0, font_size=14, 
											   text_color=self.prox_color, bold=True, shadow=True, show_points=False)
		
		# 遠位点ラベル
		dist_label_actors = []
		for pt, label in zip(dist_points, dist_labels):
			label_actor = plotter.add_point_labels([pt], [f"Dist_{label}"], point_size=0, font_size=14, 
											   text_color=self.dist_color, bold=True, shadow=True, show_points=False)
		
		prox_corrected_actors = []
		prox_corrected_label_actors = []
		if prox_corrected_points is not None and len(prox_corrected_points) > 0:
			prox_corrected_points_actor = plotter.add_points(
				prox_corrected_points, 
				color='darkturquoise', 
				point_size=18, 
				render_points_as_spheres=True
			)
			prox_corrected_actors.append(prox_corrected_points_actor)
			
			# 修正点のラベル
			for pt, label in zip(prox_corrected_points, prox_corrected_labels):
				label_actor = plotter.add_point_labels(
					[pt], [f"Prox_{label}"], 
					point_size=0, font_size=16, 
					text_color='darkturquoise', 
					bold=True, shadow=True
				)
				prox_corrected_label_actors.append(label_actor)
		
		# 遠位修正点群（A'B'C'）- 濃いピンク
		dist_corrected_actors = []
		dist_corrected_label_actors = []
		if dist_corrected_points is not None and len(dist_corrected_points) > 0:
			dist_corrected_points_actor = plotter.add_points(
				dist_corrected_points, 
				color='deeppink', 
				point_size=18, 
				render_points_as_spheres=True
			)
			dist_corrected_actors.append(dist_corrected_points_actor)
			
			# 修正点のラベル
			for pt, label in zip(dist_corrected_points, dist_corrected_labels):
				label_actor = plotter.add_point_labels(
					[pt], [f"Dist_{label}"], 
					point_size=0, font_size=16, 
					text_color='deeppink', 
					bold=True, shadow=True
				)
				dist_corrected_label_actors.append(label_actor)
		
		# 近位CL座標系の可視化
		prox_cl_actors = []
		if prox_C_cl_origin is not None:
			axis_length = 50.0
			# X軸（赤）
			prox_cl_x_line = pv.Line(prox_C_cl_origin, prox_C_cl_origin + prox_C_cl_x_axis * axis_length)
			prox_cl_x_actor = plotter.add_mesh(prox_cl_x_line, color="red", line_width=4)
			prox_cl_actors.append(prox_cl_x_actor)
			# Y軸（緑）
			prox_cl_y_line = pv.Line(prox_C_cl_origin, prox_C_cl_origin + prox_C_cl_y_axis * axis_length)
			prox_cl_y_actor = plotter.add_mesh(prox_cl_y_line, color="green", line_width=4)
			prox_cl_actors.append(prox_cl_y_actor)
			# Z軸（青）
			prox_cl_z_line = pv.Line(prox_C_cl_origin, prox_C_cl_origin + prox_C_cl_z_axis * axis_length)
			prox_cl_z_actor = plotter.add_mesh(prox_cl_z_line, color="blue", line_width=4)
			prox_cl_actors.append(prox_cl_z_actor)
			# 原点（CL）に小さな球
			prox_cl_sphere = plotter.add_mesh(
				pv.Sphere(radius=3.0, center=prox_C_cl_origin),
				color='yellow',
				opacity=0.8
			)
			prox_cl_actors.append(prox_cl_sphere)
		
		# 遠位CL座標系の可視化
		dist_cl_actors = []
		if dist_C_cl_origin is not None:
			axis_length = 50.0
			# X軸（暗赤）
			dist_cl_x_line = pv.Line(dist_C_cl_origin, dist_C_cl_origin + dist_C_cl_x_axis * axis_length)
			dist_cl_x_actor = plotter.add_mesh(dist_cl_x_line, color="darkred", line_width=4)
			dist_cl_actors.append(dist_cl_x_actor)
			# Y軸（暗緑）
			dist_cl_y_line = pv.Line(dist_C_cl_origin, dist_C_cl_origin + dist_C_cl_y_axis * axis_length)
			dist_cl_y_actor = plotter.add_mesh(dist_cl_y_line, color="darkgreen", line_width=4)
			dist_cl_actors.append(dist_cl_y_actor)
			# Z軸（暗青）
			dist_cl_z_line = pv.Line(dist_C_cl_origin, dist_C_cl_origin + dist_C_cl_z_axis * axis_length)
			dist_cl_z_actor = plotter.add_mesh(dist_cl_z_line, color="darkblue", line_width=4)
			dist_cl_actors.append(dist_cl_z_actor)
			# 原点（CL）に小さな球
			dist_cl_sphere = plotter.add_mesh(
				pv.Sphere(radius=3.0, center=dist_C_cl_origin),
				color='orange',
				opacity=0.8
			)
			dist_cl_actors.append(dist_cl_sphere)
		
		# チェックボックスUI
		def toggle_prox_mesh(state):
			prox_mesh_actor.SetVisibility(state)
		
		def toggle_dist_mesh(state):
			dist_mesh_actor.SetVisibility(state)
		
		def toggle_prox_points(state):
			prox_points_actor.SetVisibility(state)
			for label_actor in prox_label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_dist_points(state):
			dist_points_actor.SetVisibility(state)
			for label_actor in dist_label_actors:
				label_actor.SetVisibility(state)
		
		def toggle_prox_cl(state):
			for actor in prox_cl_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_cl(state):
			for actor in dist_cl_actors:
				actor.SetVisibility(state)
		
		def toggle_prox_corrected(state):
			for actor in prox_corrected_actors:
				actor.SetVisibility(state)
			for actor in prox_corrected_label_actors:
				actor.SetVisibility(state)
		
		def toggle_dist_corrected(state):
			for actor in dist_corrected_actors:
				actor.SetVisibility(state)
			for actor in dist_corrected_label_actors:
				actor.SetVisibility(state)
		
		# チェックボックス配置
		plotter.add_text("Prox Mesh", position=(10, 10), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_prox_mesh, value=True, position=(120, 12), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		plotter.add_text("Dist Mesh", position=(10, 40), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_dist_mesh, value=True, position=(120, 42), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		plotter.add_text("Prox PP", position=(10, 70), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_prox_points, value=True, position=(120, 72), size=20, border_size=1, color_on=self.prox_color, color_off="white")
		
		plotter.add_text("Dist PP", position=(10, 100), font_size=10, color="black")
		plotter.add_checkbox_button_widget(toggle_dist_points, value=True, position=(120, 102), size=20, border_size=1, color_on=self.dist_color, color_off="white")
		
		# CL座標系のチェックボックス
		checkbox_y = 130
		if prox_cl_actors:
			plotter.add_text("Prox C_cl", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_prox_cl, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="yellow", color_off="white")
			checkbox_y += 30
		
		if dist_cl_actors:
			plotter.add_text("Dist C_cl", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_dist_cl, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="orange", color_off="white")
			checkbox_y += 30
		
		# 修正点のチェックボックス
		if prox_corrected_actors:
			plotter.add_text("Prox A'B'C'D'", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_prox_corrected, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="darkturquoise", color_off="white")
			checkbox_y += 30
		
		if dist_corrected_actors:
			plotter.add_text("Dist A'B'C'", position=(10, checkbox_y), font_size=10, color="black")
			plotter.add_checkbox_button_widget(toggle_dist_corrected, value=True, position=(120, checkbox_y+2), size=20, border_size=1, color_on="deeppink", color_off="white")
		
		# カメラを全体が見えるように自動調整
		plotter.reset_camera()
		plotter.show()

	def on_c_fix_transform_prox(self) -> None:
		"""C. Fix: 近位ABCD修正後の座標をPPファイルとして保存"""
		prox_pp_abcd_path = self.prox_pp_abcd_path.get().strip()
		prox_pp_olmn_path = self.prox_pp_olmn_path.get().strip()
		
		if not all([prox_pp_abcd_path, prox_pp_olmn_path]):
			messagebox.showwarning("入力不足", "近位のPPファイルを選択してください。")
			return
		
		try:
			# PP読み込み（近位）
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(prox_pp_abcd_path)
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(prox_pp_olmn_path)
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"近位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# 近位CL座標系の構築
		prox_L_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'L'), None)
		prox_M_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'M'), None)
		prox_N_idx = next((i for i, lbl in enumerate(prox_labels) if lbl == 'N'), None)
		
		if prox_L_idx is None or prox_M_idx is None or prox_N_idx is None:
			messagebox.showerror("エラー", "近位のL, M, N点が見つかりません。")
			return
		
		prox_L_point = prox_points[prox_L_idx]
		prox_M_point = prox_points[prox_M_idx]
		prox_N_point = prox_points[prox_N_idx]
		
		prox_CL_point = (prox_M_point + prox_N_point) / 2.0
		prox_LM_vector = prox_M_point - prox_L_point
		prox_LM_unit = prox_LM_vector / np.linalg.norm(prox_LM_vector)
		prox_LN_vector = prox_N_point - prox_L_point
		prox_projection_length = np.dot(prox_LN_vector, prox_LM_unit)
		prox_L_prime_point = prox_L_point + prox_projection_length * prox_LM_unit
		
		prox_C_cl_origin = prox_CL_point
		prox_x_vec = prox_L_prime_point - prox_M_point
		prox_C_cl_x_axis = prox_x_vec / np.linalg.norm(prox_x_vec)
		prox_z_vec = prox_L_prime_point - prox_N_point
		prox_C_cl_z_axis = prox_z_vec / np.linalg.norm(prox_z_vec)
		prox_C_cl_y_axis = np.cross(prox_C_cl_z_axis, prox_C_cl_x_axis)
		prox_C_cl_y_axis = prox_C_cl_y_axis / np.linalg.norm(prox_C_cl_y_axis)
		
		# 修正量を取得
		correction_x = self.c_fix_correction_x.get()
		correction_y = self.c_fix_correction_y.get()
		correction_z = self.c_fix_correction_z.get()
		
		# 近位ABCD点を修正
		prox_ABCD_corrected = {}
		for target_label in ['A', 'B', 'C', 'D']:
			idx = next((i for i, lbl in enumerate(prox_labels) if lbl == target_label), None)
			if idx is not None:
				original_point = prox_points[idx]
				correction_vector = (correction_x * prox_C_cl_x_axis + 
									 correction_y * prox_C_cl_y_axis + 
									 correction_z * prox_C_cl_z_axis)
				corrected_point = original_point + correction_vector
				prox_ABCD_corrected[target_label] = corrected_point
		
		# 近位ABCDファイルの保存
		prox_abcd_save_path = filedialog.asksaveasfilename(
			title="近位ABCD修正後のPPファイルを保存",
			defaultextension=".pp",
			filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")],
			initialdir=str(Path(prox_pp_abcd_path).parent),
			initialfile=Path(prox_pp_abcd_path).stem + "_corrected.pp"
		)
		
		if not prox_abcd_save_path:
			messagebox.showinfo("キャンセル", "保存がキャンセルされました。")
			return
		
		try:
			# ABCDの順で保存
			prox_abcd_points_list = []
			prox_abcd_labels_list = []
			for label in ['A', 'B', 'C', 'D']:
				if label in prox_ABCD_corrected:
					prox_abcd_points_list.append(prox_ABCD_corrected[label])
					prox_abcd_labels_list.append(label)
			
			if prox_abcd_points_list:
				save_child_pp_file(prox_abcd_save_path, np.array(prox_abcd_points_list), prox_abcd_labels_list)
				messagebox.showinfo("完了", f"近位ABCD修正後のPPファイルを保存しました:\n{prox_abcd_save_path}")
		except Exception as e:
			messagebox.showerror("保存エラー", f"ファイルの保存に失敗しました:\n{e}")

	def on_c_fix_transform_dist(self) -> None:
		"""C. Fix: 遠位ABC修正後の座標をPPファイルとして保存"""
		dist_pp_abc_path = self.dist_pp_abc_path.get().strip()
		dist_pp_olmn_path = self.dist_pp_olmn_path.get().strip()
		
		if not all([dist_pp_abc_path, dist_pp_olmn_path]):
			messagebox.showwarning("入力不足", "遠位のPPファイルを選択してください。")
			return
		
		try:
			# PP読み込み（遠位）
			dist_points_abc, dist_labels_abc = self._parse_pp_file(dist_pp_abc_path)
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(dist_pp_olmn_path)
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
		except Exception as e:
			messagebox.showerror("PP読み込み失敗", f"遠位PPファイルの読み込みに失敗しました:\n{e}")
			return
		
		# 遠位CL座標系の構築
		dist_L_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'L'), None)
		dist_M_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'M'), None)
		dist_N_idx = next((i for i, lbl in enumerate(dist_labels) if lbl == 'N'), None)
		
		if dist_L_idx is None or dist_M_idx is None or dist_N_idx is None:
			messagebox.showerror("エラー", "遠位のL, M, N点が見つかりません。")
			return
		
		dist_L_point = dist_points[dist_L_idx]
		dist_M_point = dist_points[dist_M_idx]
		dist_N_point = dist_points[dist_N_idx]
		
		dist_CL_point = (dist_M_point + dist_N_point) / 2.0
		dist_LM_vector = dist_M_point - dist_L_point
		dist_LM_unit = dist_LM_vector / np.linalg.norm(dist_LM_vector)
		dist_LN_vector = dist_N_point - dist_L_point
		dist_projection_length = np.dot(dist_LN_vector, dist_LM_unit)
		dist_L_prime_point = dist_L_point + dist_projection_length * dist_LM_unit
		
		dist_C_cl_origin = dist_CL_point
		dist_x_vec = dist_L_prime_point - dist_M_point
		dist_C_cl_x_axis = dist_x_vec / np.linalg.norm(dist_x_vec)
		dist_z_vec = dist_L_prime_point - dist_N_point
		dist_C_cl_z_axis = dist_z_vec / np.linalg.norm(dist_z_vec)
		dist_C_cl_y_axis = np.cross(dist_C_cl_z_axis, dist_C_cl_x_axis)
		dist_C_cl_y_axis = dist_C_cl_y_axis / np.linalg.norm(dist_C_cl_y_axis)
		
		# 修正量を取得
		correction_x = self.c_fix_correction_x.get()
		correction_y = self.c_fix_correction_y.get()
		correction_z = self.c_fix_correction_z.get()
		
		# 遠位ABC点を修正
		dist_ABC_corrected = {}
		for target_label in ['A', 'B', 'C']:
			idx = next((i for i, lbl in enumerate(dist_labels) if lbl == target_label), None)
			if idx is not None:
				original_point = dist_points[idx]
				correction_vector = (correction_x * dist_C_cl_x_axis + 
									 correction_y * dist_C_cl_y_axis + 
									 correction_z * dist_C_cl_z_axis)
				corrected_point = original_point + correction_vector
				dist_ABC_corrected[target_label] = corrected_point
		
		# 遠位ABCファイルの保存
		dist_abc_save_path = filedialog.asksaveasfilename(
			title="遠位ABC修正後のPPファイルを保存",
			defaultextension=".pp",
			filetypes=[("PickedPoints", "*.pp"), ("All files", "*.*")],
			initialdir=str(Path(dist_pp_abc_path).parent),
			initialfile=Path(dist_pp_abc_path).stem + "_corrected.pp"
		)
		
		if not dist_abc_save_path:
			messagebox.showinfo("キャンセル", "保存がキャンセルされました。")
			return
		
		try:
			# ABCの順で保存
			dist_abc_points_list = []
			dist_abc_labels_list = []
			for label in ['A', 'B', 'C']:
				if label in dist_ABC_corrected:
					dist_abc_points_list.append(dist_ABC_corrected[label])
					dist_abc_labels_list.append(label)
			
			if dist_abc_points_list:
				save_child_pp_file(dist_abc_save_path, np.array(dist_abc_points_list), dist_abc_labels_list)
				messagebox.showinfo("完了", f"遠位ABC修正後のPPファイルを保存しました:\n{dist_abc_save_path}")
		except Exception as e:
			messagebox.showerror("保存エラー", f"ファイルの保存に失敗しました:\n{e}")

	def on_animate(self) -> None:
		"""
		【最重要メソッド】アニメーション実行: 近位固定、遠位を同次変換行列に沿って動かす
		
		このメソッドは約1400行のコードで構成されており、以下の処理を行います:
		
		■ 処理フロー:
		  1. ファイル読み込み（近位・遠位モデル、PP、変換行列）
		  2. 座標系構築（C-RobotScan座標系とC-CL座標系）
		  3. ワールド座標系への統一変換
		  4. ヒートマップ事前計算（オプション、並列処理対応）
		  5. PyVistaプロッターUI構築
		  6. アニメーションループ（リアルタイム座標変換と可視化）
		  
		■ 座標系の種類:
		  - C-RobotScan: ロボット座標系（元データ）
		  - C-CL: 解剖学的座標系（L, M, N点から構築）
		  - ワールド座標系: 可視化用の統一座標系
		  
		■ 主要な内部関数:
		  - show_frame(frame_idx): 指定フレームを描画
		  - animation_loop(): アニメーション再生ループ
		  - on_playback_scale_change(frame_idx): スライダー操作でのフレーム移動
		  - export_to_csv(): 接触深度データをCSVエクスポート
		  
		■ 改修ポイント:
		  - 行5520付近: ヒートマップ計算処理（並列化可能）
		  - 行6230-6540: show_frame関数（座標変換ロジック）
		  - 行6541-6650: animation_loop関数（再生速度制御）
		  
		■ パフォーマンス最適化:
		  - キャッシュ機能: 2回目以降はヒートマップを再利用
		  - 並列処理: マルチコアCPUでヒートマップ計算を高速化
		  - ダウンサンプリング: 点群数を削減して計算を軽量化
		  
		■ トラブルシューティング:
		  - メッシュ読み込み失敗: ファイルパスとフォーマットを確認
		  - 座標系構築失敗: L, M, N点の存在を確認
		  - アニメーション動作不良: transform_dataの行列が正しいか確認
		"""
		# 必要なファイルがすべて選択されているか確認
		if not all([
			self.prox_model_path.get(),
			self.prox_pp_abcd_path.get(),
			self.prox_pp_olmn_path.get(),
			self.dist_model_path.get(),
			self.dist_pp_abc_path.get(),
			self.dist_pp_olmn_path.get(),
			self.transform_group_path.get()
		]):
			messagebox.showerror("エラー", "全てのファイルを選択してください。")
			return
		
		try:
			# 変位・姿勢変化データを読み込み
			transform_data = self._load_transform_matrices(self.transform_group_path.get())
			if not transform_data:
				messagebox.showerror("エラー", "変位・姿勢変化データが空です。")
				return
			
			# 近位モデルとPPを読み込み - 2つのPPファイルを統合
			prox_mesh = pv.read(self.prox_model_path.get())
			prox_points_abcd, prox_labels_abcd = self._parse_pp_file(self.prox_pp_abcd_path.get())
			prox_points_olmn, prox_labels_olmn = self._parse_pp_file(self.prox_pp_olmn_path.get())
			prox_points = np.vstack([prox_points_abcd, prox_points_olmn])
			prox_labels = prox_labels_abcd + prox_labels_olmn
			
			# 遠位モデルとPPを読み込み - 2つのPPファイルを統合
			dist_mesh = pv.read(self.dist_model_path.get())
			dist_points_abc, dist_labels_abc = self._parse_pp_file(self.dist_pp_abc_path.get())
			dist_points_olmn, dist_labels_olmn = self._parse_pp_file(self.dist_pp_olmn_path.get())
			dist_points = np.vstack([dist_points_abc, dist_points_olmn])
			dist_labels = dist_labels_abc + dist_labels_olmn
			
			# 近位座標系を構築（_build_coordinate_systemを使用）
			try:
				prox_origin, prox_x_axis, prox_y_axis, prox_z_axis = self._build_coordinate_system(prox_points, prox_labels)
				# 自動オフセット設定
				self._try_auto_set_offset(prox_points, prox_labels, prox_origin, prox_x_axis, prox_y_axis, prox_z_axis,
				                          self.prox_offset_x, self.prox_offset_y, self.prox_offset_z)
			except Exception as e:
				messagebox.showerror("近位座標系構築", f"近位座標系の構築に失敗しました:\n{e}")
				return
			
			# 遠位座標系を構築
			try:
				dist_origin, dist_x_axis, dist_y_axis, dist_z_axis = self._build_coordinate_system_dist(dist_points, dist_labels)
				# 自動オフセット設定
				self._try_auto_set_offset(dist_points, dist_labels, dist_origin, dist_x_axis, dist_y_axis, dist_z_axis,
				                          self.dist_offset_x, self.dist_offset_y, self.dist_offset_z)
			except Exception as e:
				messagebox.showerror("遠位座標系構築", f"遠位座標系の構築に失敗しました:\n{e}")
				return
			
			# --- Step 1: 近位座標系をワールド座標系に揃える変換 ---
			prox_rotation_matrix = np.column_stack([prox_x_axis, prox_y_axis, prox_z_axis])
			prox_to_world = np.eye(4)
			prox_to_world[:3, :3] = prox_rotation_matrix.T
			prox_to_world[:3, 3] = -prox_rotation_matrix.T @ prox_origin
			
			# 近位メッシュを変換
			prox_mesh_homo = np.hstack([prox_mesh.points, np.ones((prox_mesh.points.shape[0], 1))])
			prox_mesh.points = (prox_to_world @ prox_mesh_homo.T).T[:, :3]
			
			# 近位座標系をワールド座標系に変換
			prox_origin = np.array([0.0, 0.0, 0.0])
			prox_x_axis = np.array([1.0, 0.0, 0.0])
			prox_y_axis = np.array([0.0, 1.0, 0.0])
			prox_z_axis = np.array([0.0, 0.0, 1.0])
			
			# --- Step 2: 遠位メッシュと座標系に同じ変換を適用 ---
			dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
			dist_mesh.points = (prox_to_world @ dist_mesh_homo.T).T[:, :3]
			
			# 遠位座標系の原点と軸を変換
			dist_origin_homo = np.append(dist_origin, 1)
			dist_origin_in_world = (prox_to_world @ dist_origin_homo)[:3]
			
			dist_rotation_matrix = np.column_stack([dist_x_axis, dist_y_axis, dist_z_axis])
			dist_rotation_in_world = prox_rotation_matrix.T @ dist_rotation_matrix
			dist_x_axis_in_world = dist_rotation_in_world[:, 0]
			dist_y_axis_in_world = dist_rotation_in_world[:, 1]
			dist_z_axis_in_world = dist_rotation_in_world[:, 2]
			
			# --- Step 3: 遠位座標系をワールド座標系に揃える変換 ---
			dist_to_world = np.eye(4)
			dist_to_world[:3, :3] = dist_rotation_in_world.T
			dist_to_world[:3, 3] = -dist_rotation_in_world.T @ dist_origin_in_world
			
			# 遠位メッシュを変換
			dist_mesh_homo = np.hstack([dist_mesh.points, np.ones((dist_mesh.points.shape[0], 1))])
			dist_mesh.points = (dist_to_world @ dist_mesh_homo.T).T[:, :3]
			
			# 遠位座標系をワールド座標系に変換
			dist_origin = np.array([0.0, 0.0, 0.0])
			dist_x_axis = np.array([1.0, 0.0, 0.0])
			dist_y_axis = np.array([0.0, 1.0, 0.0])
			dist_z_axis = np.array([0.0, 0.0, 1.0])
			
			# --- 軟骨メッシュの読み込みと変換（任意） ---
			prox_cartilage_mesh = None
			dist_cartilage_mesh = None
			dist_cartilage_for_anim = None
			
			prox_cart_path = self.prox_cartilage_model_path.get().strip()
			dist_cart_path = self.dist_cartilage_model_path.get().strip()
			
			if prox_cart_path:
				try:
					prox_cartilage_mesh = pv.read(prox_cart_path)
					# 骨と同じ変換を適用: prox_to_world
					pc_homo = np.hstack([prox_cartilage_mesh.points, np.ones((prox_cartilage_mesh.points.shape[0], 1))])
					prox_cartilage_mesh.points = (prox_to_world @ pc_homo.T).T[:, :3]
					print(f"近位軟骨モデル読み込み: {Path(prox_cart_path).name} ({prox_cartilage_mesh.n_points}点)")
				except Exception as e:
					print(f"近位軟骨モデル読み込み失敗: {e}")
					prox_cartilage_mesh = None
			
			if dist_cart_path:
				try:
					dist_cartilage_mesh = pv.read(dist_cart_path)
					# 骨と同じ変換を適用: prox_to_world → dist_to_world
					dc_homo = np.hstack([dist_cartilage_mesh.points, np.ones((dist_cartilage_mesh.points.shape[0], 1))])
					dist_cartilage_mesh.points = (prox_to_world @ dc_homo.T).T[:, :3]
					dc_homo2 = np.hstack([dist_cartilage_mesh.points, np.ones((dist_cartilage_mesh.points.shape[0], 1))])
					dist_cartilage_mesh.points = (dist_to_world @ dc_homo2.T).T[:, :3]
					# アニメーション用に初期位置を保存
					dist_cartilage_for_anim = dist_cartilage_mesh.copy()
					print(f"遠位軟骨モデル読み込み: {Path(dist_cart_path).name} ({dist_cartilage_mesh.n_points}点)")
				except Exception as e:
					print(f"遠位軟骨モデル読み込み失敗: {e}")
					dist_cartilage_mesh = None

			# --- 関節領域の抽出 ---
			# O'原点位置を計算（ローカル座標系に沿ったオフセット）
			# Step 1, 3でメッシュと座標系はワールド原点(0,0,0)と軸(1,0,0等)に整列済み
			prox_origin_prime = np.array([
				self.prox_offset_x.get(),
				self.prox_offset_y.get(),
				self.prox_offset_z.get()
			])
			
			dist_origin_prime = np.array([
				self.dist_offset_x.get(),
				self.dist_offset_y.get(),
				self.dist_offset_z.get()
			])
			
			print(f"[抽出] 近位O: {prox_origin}, 近位O': {prox_origin_prime}, オフセット設定: [{self.prox_offset_x.get()}, {self.prox_offset_y.get()}, {self.prox_offset_z.get()}]")
			print(f"[抽出] 遠位O: {dist_origin}, 遠位O': {dist_origin_prime}, オフセット設定: [{self.dist_offset_x.get()}, {self.dist_offset_y.get()}, {self.dist_offset_z.get()}]")
			
			# 近位関節領域（寛骨臼）を球体で抽出（O'を中心に）
			prox_joint_region = self._extract_region_by_sphere(prox_mesh, prox_origin_prime, self.prox_radius.get())
			
			# 遠位関節領域（大腿骨頭）を球体で抽出（O'を中心に）
			dist_joint_region = self._extract_region_by_sphere(dist_mesh, dist_origin_prime, self.dist_radius.get())
			
			# 関節領域が空の場合は警告
			if prox_joint_region.n_points == 0:
				messagebox.showwarning("警告", "近位関節領域（寛骨臼）が抽出できませんでした。半径を確認してください。")
			if dist_joint_region.n_points == 0:
				messagebox.showwarning("警告", "遠位関節領域（大腿骨頭）が抽出できませんでした。半径を確認してください。")
			
			# --- オーバーラップ体積の事前計算 ---
			# 事前計算ダイアログを表示
			has_cartilage = (prox_cartilage_mesh is not None and dist_cartilage_for_anim is not None)
			has_joint_regions = (prox_joint_region is not None and prox_joint_region.n_points > 0
			                     and dist_joint_region is not None and dist_joint_region.n_points > 0)
			has_fem_target = has_cartilage or has_joint_regions
			progress_window, update_progress, cancel_var, options_dict, start_var, skip_var = self._show_precompute_dialog(len(transform_data), has_cartilage=has_fem_target)
			
			# ユーザーがボタンを押すまで待機（モーダルダイアログ）
			# ボタンが押されるとdialog_closedがTrueになる
			while progress_window.winfo_exists():
				if start_var.get() or skip_var.get() or cancel_var.get():
					break
				progress_window.update()
				time.sleep(0.01)
			
			# キャンセルされた場合は終了
			if cancel_var.get():
				try:
					progress_window.destroy()
				except:
					pass
				return
			
			# オーバーラップ計算を実行
			overlap_precomputed = []
			overlap_areas_precomputed = []  # 面積リストを追加
			overlap_depths_precomputed = []  # 深度リストを追加
			heatmap_precomputed = []  # 距離ヒートマップリストを追加
			fem_pressure_precomputed = []  # FEM接触圧リスト
			prox_fem_surface = None  # FEM近位表面メッシュ

			if skip_var.get() or not start_var.get():
				# 事前計算をスキップ（「計算せずに進む」/×閉じ/未開始のいずれも計算しない）
				try:
					progress_window.destroy()
				except:
					pass
				messagebox.showinfo("スキップ", "事前計算をスキップしました。\nオーバーラップ表示は無効になります。")
			else:
				# 事前計算を実行する
				try:
					# 最適化オプションを取得
					use_parallel = options_dict['use_parallel'].get()
					use_cache = options_dict['use_cache'].get()
					use_bbox = options_dict['use_bbox'].get()
					use_simplify = options_dict['use_simplify'].get()
					
					# 事前計算を実行
					def progress_wrapper(current, total, msg):
						if cancel_var.get():
							return False
						return update_progress(current, total, msg)
					
					# オーバーラップ体積の計算
					result = self._precompute_overlap_volumes(
						prox_mesh, dist_mesh, transform_data, 
						progress_callback=progress_wrapper,
						use_cache=use_cache,
						use_parallel=use_parallel,
						use_bbox_check=use_bbox,
						simplify_mesh=use_simplify
					)
					
					# 戻り値を展開（ヒートマップを含む場合と含まない場合の両方に対応）
					if len(result) == 4:
						overlap_precomputed, overlap_areas_precomputed, overlap_depths_precomputed, heatmap_precomputed = result
						if heatmap_precomputed:
							print(f"[キャッシュ] ヒートマップがキャッシュから読み込まれました: {len(heatmap_precomputed)}フレーム")
							# 最初のフレームの情報を表示
							if len(heatmap_precomputed) > 0:
								first_heatmap = heatmap_precomputed[0]
								print(f"[キャッシュ] 最初のヒートマップ: {first_heatmap.n_points}点")
								print(f"[キャッシュ] スカラー配列: {first_heatmap.array_names}")
								if 'distance' in first_heatmap.array_names:
									distances = first_heatmap['distance']
									print(f"[キャッシュ] 距離範囲: min={distances.min():.2f}, max={distances.max():.2f}")
						else:
							print(f"[キャッシュ] ヒートマップは空でした")
					else:
						overlap_precomputed, overlap_areas_precomputed, overlap_depths_precomputed = result
						heatmap_precomputed = []  # ヒートマップがない場合は空リスト
						print(f"[キャッシュ] 旧形式のキャッシュ（ヒートマップなし）")
					
					if cancel_var.get():
						messagebox.showinfo("キャンセル", "事前計算がキャンセルされました。\nオーバーラップ表示は無効になります。")
						overlap_precomputed = []
						overlap_areas_precomputed = []
						overlap_depths_precomputed = []
						heatmap_precomputed = []
					else:
						# 新規計算フラグ（キャッシュ保存判定用）
						heatmap_newly_computed = False
						
						# 距離ヒートマップの事前計算（キャッシュから読み込まれていない場合のみ）
						if not heatmap_precomputed:
							print(f"[事前計算] 距離ヒートマップを{len(transform_data)}フレーム分計算します...")
							print(f"[事前計算] 近位関節領域: {prox_joint_region.n_points}点")
							print(f"[事前計算] 遠位関節領域: {dist_joint_region.n_points}点")
							
							if prox_joint_region.n_points == 0 or dist_joint_region.n_points == 0:
								print(f"[警告] 関節領域が空のため、ヒートマップ計算をスキップします")
								heatmap_precomputed = []
							else:
								# 事前最適化: 表面抽出を1回だけ実行（ループ外で）
								prox_surface = prox_joint_region.extract_surface() if hasattr(prox_joint_region, 'extract_surface') else prox_joint_region
								dist_surface = dist_joint_region.extract_surface() if hasattr(dist_joint_region, 'extract_surface') else dist_joint_region
								print(f"[最適化] 表面抽出完了: prox={prox_surface.n_points}点, dist={dist_surface.n_points}点")
								
								# Phase 3最適化: 並列処理は一旦無効化（安定性優先）
								use_parallel_heatmap = False  # use_parallel and len(transform_data) > 1
								if use_parallel_heatmap:
									print(f"[Phase 3] 並列処理を使用してヒートマップ計算を高速化します")
								else:
									print(f"[最適化] 逐次処理でヒートマップを計算します（安定性優先）")
								
								# Phase 2最適化: prox_points配列を事前抽出（不変なので1回だけ）
								prox_points = np.asarray(prox_joint_region.points)
								print(f"[最適化] prox_points配列を事前抽出: {len(prox_points)}点")
								
								# 事前最適化: VTKエラー出力を1回だけ設定（ループ外で）
								import vtk
								import os
								vtk_error_output = vtk.vtkFileOutputWindow()
								vtk_error_output.SetFileName("nul" if os.name == 'nt' else "/dev/null")
								vtk_original_output = vtk.vtkOutputWindow.GetInstance()
								vtk.vtkOutputWindow.SetInstance(vtk_error_output)

								# --- 符号付き距離を Open3D BVH で一括高速計算（VTK implicit と一致・大幅高速化） ---
								# BVHにより O(N log M)。GPU総当たりやVTK per-frame より桁違いに速く、追加ライブラリ不要。
								_use_fast_heatmap = False
								try:
									print(f"[高速] Open3D BVH でヒートマップ{len(transform_data)}フレームを一括計算します")
									heatmap_precomputed = self._precompute_heatmaps_o3d(
										prox_joint_region, dist_surface, prox_points, transform_data,
										update_progress, cancel_var)
									_use_fast_heatmap = True
									print(f"[高速] ヒートマップ {len(heatmap_precomputed)}フレーム計算完了")
								except Exception as e:
									print(f"[高速] Open3D計算に失敗、CPU逐次にフォールバック: {e}")
									_use_fast_heatmap = False
									heatmap_precomputed = []

								try:
									if _use_fast_heatmap:
										pass  # 既に高速パスで計算済み
									elif use_parallel_heatmap:
										# Phase 3: スレッド並列処理でヒートマップを計算（GIL制約あるが、I/O待機が多いため有効）
										from concurrent.futures import ThreadPoolExecutor, as_completed
										import multiprocessing
										
										max_workers = min(multiprocessing.cpu_count(), len(transform_data))
										print(f"[Phase 3] {max_workers}個のワーカーでスレッド並列処理を実行")
										
										# 並列処理用のヘルパー関数
										def compute_single_heatmap(args):
											i, tf_info, prox_region, dist_region, prox_surf, dist_surf, prox_pts = args
											try:
												# 遠位関節領域を変換
												dist_joint_transformed = self._apply_transform(dist_region.copy(), tf_info['matrix'])
												dist_surface_transformed = self._apply_transform(dist_surf.copy(), tf_info['matrix'])
												
												# ヒートマップを計算
												heatmap = self._compute_distance_heatmap(prox_region, dist_joint_transformed, prox_surf, dist_surface_transformed, prox_pts)
												return (i, heatmap, None)
											except Exception as e:
												return (i, None, str(e))
										
										# 並列タスクを準備
										tasks = [
											(i, tf_info, prox_joint_region, dist_joint_region, prox_surface, dist_surface, prox_points)
											for i, tf_info in enumerate(transform_data)
										]
										
										# 結果を格納する配列（インデックス順に保持）
										heatmap_results = [None] * len(transform_data)
										
										with ThreadPoolExecutor(max_workers=max_workers) as executor:
											futures = {executor.submit(compute_single_heatmap, task): task[0] for task in tasks}
											
											for future in as_completed(futures):
												if cancel_var.get():
													break
												
												i, heatmap, error = future.result()
												
												if error:
													print(f"[並列処理] フレーム {i}: 計算失敗 - {error}")
													# ダミーヒートマップを作成
													empty_heatmap = prox_joint_region.copy()
													if empty_heatmap.n_points > 0:
														empty_heatmap['distance'] = np.zeros(empty_heatmap.n_points)
													heatmap_results[i] = empty_heatmap
												else:
													heatmap_results[i] = heatmap
												
												# 進捗更新
												completed = sum(1 for h in heatmap_results if h is not None)
												if not update_progress(completed, len(transform_data), f"並列ヒートマップ計算中: {completed}/{len(transform_data)}"):
													break
										
										# 結果をリストに追加
										for heatmap in heatmap_results:
											if heatmap is not None:
												heatmap_precomputed.append(heatmap)
										
										# デバッグ: 最初のフレームの情報を表示
										if len(heatmap_precomputed) > 0:
											first_heatmap = heatmap_precomputed[0]
											print(f"[並列処理] フレーム0のヒートマップ: {first_heatmap.n_points}点, スカラー: {first_heatmap.array_names}")
											if 'distance' in first_heatmap.array_names:
												distances = first_heatmap['distance']
												print(f"[並列処理] 距離の範囲: min={distances.min():.2f}mm, max={distances.max():.2f}mm, mean={distances.mean():.2f}mm")
										
									else:
										# 従来の逐次処理
										for i, tf_info in enumerate(transform_data):
											if cancel_var.get():
												break
											
											try:
												# 遠位関節領域を変換（最適化: copyを最小限に）
												dist_joint_transformed = self._apply_transform(dist_joint_region.copy(), tf_info['matrix'])
												dist_surface_transformed = self._apply_transform(dist_surface.copy(), tf_info['matrix'])
												
												# 距離ヒートマップを計算（prox_joint_regionは変更されないため.copy()不要、表面は事前抽出済み）
												heatmap = self._compute_distance_heatmap(prox_joint_region, dist_joint_transformed, prox_surface, dist_surface_transformed, prox_points)
												heatmap_precomputed.append(heatmap)
												
												# デバッグログは100フレーム毎のみ（I/O負荷削減）
												if i % 100 == 0:
													print(f"[進捗] フレーム{i}/{len(transform_data)}完了")
												
												# デバッグ: 最初のフレームの情報を表示
												if i == 0:
													print(f"[事前計算] フレーム0のヒートマップ: {heatmap.n_points}点, スカラー: {heatmap.array_names}")
													if 'distance' in heatmap.array_names:
														distances = heatmap['distance']
														print(f"[事前計算] 距離の範囲: min={distances.min():.2f}mm, max={distances.max():.2f}mm, mean={distances.mean():.2f}mm")
												
												# 進捗更新を10フレーム毎にバッチ化（GUI更新負荷削減）
												if (i + 1) % 10 == 0 or (i + 1) == len(transform_data):
													if not update_progress(i + 1, len(transform_data), f"ヒートマップ計算中: {i+1}/{len(transform_data)}"):
														break
											
											except Exception as e:
												# import traceback # 既にグローバルでインポート済み
												print(f"[エラー] フレーム {i}: ヒートマップ計算失敗")
												print(f"[エラー詳細] {e}")
												print(f"[スタックトレース]\n{traceback.format_exc()}")
												# 空のメッシュにダミーデータを追加
												try:
													empty_heatmap = prox_joint_region.copy()
													if empty_heatmap.n_points > 0:
														empty_heatmap['distance'] = np.zeros(empty_heatmap.n_points)
													heatmap_precomputed.append(empty_heatmap)
												except Exception as e2:
													print(f"[エラー] ダミーヒートマップ作成も失敗: {e2}")
								finally:
									# VTKエラー出力を元に戻す
									vtk.vtkOutputWindow.SetInstance(vtk_original_output)
								
								# 新規計算完了フラグを立てる
								heatmap_newly_computed = True
						else:
							print(f"[キャッシュ] ヒートマップはキャッシュから読み込み済みです")
						
						if cancel_var.get():
							messagebox.showinfo("キャンセル", "事前計算がキャンセルされました。\nオーバーラップ表示は無効になります。")
							overlap_precomputed = []
							overlap_areas_precomputed = []
							overlap_depths_precomputed = []
							heatmap_precomputed = []
						else:
							# ヒートマップも含めてキャッシュに保存（新規計算した場合のみ）。
							# 大量フレームだと保存に時間がかかるため、バックグラウンドで書き出し、
							# アニメーションは待たずに先に表示する（保存は継続・完了時にログ表示）。
							if use_cache and heatmap_newly_computed and heatmap_precomputed:
								cache_hash = self._get_cache_hash(prox_mesh, dist_mesh, transform_data, use_simplify)
								cache_filepath = self._get_cache_filepath(cache_hash)
								print(f"[キャッシュ] ヒートマップをバックグラウンドで保存します（アニメは先に表示）...")

								def _bg_save_cache(cf=cache_filepath, ov=overlap_precomputed,
								                   oa=overlap_areas_precomputed, od=overlap_depths_precomputed,
								                   hm=heatmap_precomputed):
									try:
										self._save_overlap_cache(cf, ov, oa, od, hm)
										print(f"[キャッシュ] ★バックグラウンド保存 完了: {cf}")
									except Exception as _e:
										print(f"[キャッシュ] バックグラウンド保存エラー: {_e}")

								threading.Thread(target=_bg_save_cache, daemon=True).start()
							
							print(f"[事前計算] 完了: {len(overlap_precomputed)}フレーム分のオーバーラップ, {len(heatmap_precomputed)}フレーム分のヒートマップ")

					# --- FEM接触解析の事前計算（関節領域で実行） ---
					fem_pressure_precomputed = []  # List of (pressure_array, results_summary_dict) per frame
					prox_fem_surface = None  # For mapping pressure results back to heatmap points
					if options_dict.get('enable_fem_precompute') and options_dict['enable_fem_precompute'].get():
						if prox_joint_region.n_points > 0 and dist_joint_region.n_points > 0 and _HAS_FEM:
							print(f"[FEM事前計算] 関節領域で{len(transform_data)}フレーム分のFEM解析を実行します...")
							print(f"[FEM事前計算] 近位関節領域: {prox_joint_region.n_points}点")
							print(f"[FEM事前計算] 遠位関節領域: {dist_joint_region.n_points}点")
							try:
								material = MaterialProperties(
									E=self.fem_cart_E.get(),
									nu=self.fem_cart_nu.get(),
									thickness=self.fem_cart_thickness.get(),
								)
								contact = ContactParameters(
									penalty_stiffness=self.fem_penalty_stiffness.get(),
									contact_tolerance=self.fem_contact_tolerance.get(),
								)
								# 近位関節領域の表面を事前抽出（FEM解析用）
								prox_fem_surface = prox_joint_region.extract_surface() if hasattr(prox_joint_region, 'extract_surface') else prox_joint_region
								print(f"[FEM事前計算] 近位FEM表面: {prox_fem_surface.n_points}点")
								for i, tf_info in enumerate(transform_data):
									if cancel_var.get():
										break
									try:
										# 遠位関節領域をこのフレームの姿勢に変換
										matrix = tf_info['matrix']
										dist_joint_transformed = self._apply_transform(dist_joint_region.copy(), matrix)
										dist_fem_surface = dist_joint_transformed.extract_surface() if hasattr(dist_joint_transformed, 'extract_surface') else dist_joint_transformed
										# FEM解析を実行
										solver = FEMContactSolver(material=material, contact=contact, verbose=False, shared_cache=self._shared_fem_cache)
										results = solver.analyze(
											prox_fem_surface.copy(), dist_fem_surface.copy(),
											boundary_mode=self.fem_boundary_mode.get(),
											max_nodes=self.fem_max_nodes.get(),
										)
										# contact_pressure配列を保存（prox_fem_surface.n_points長）
										# ヒートマップのprox_joint_regionに対応させるため、必要に応じてマッピング
										if results.n_nodes == prox_fem_surface.n_points:
											pressure = results.contact_pressure.copy()
										else:
											# デシメーションが起きた場合、元の近位メッシュにマッピング
											from scipy.spatial import cKDTree
											tree = cKDTree(solver._nodes)
											_, idx = tree.query(prox_fem_surface.points, k=1)
											pressure = results.contact_pressure[idx]
										summary = {
											'contact_area': results.contact_area,
											'peak_contact_pressure': results.peak_contact_pressure,
											'n_contact_nodes': results.n_contact_nodes,
											'solve_time_sec': results.solve_time_sec,
										}
										fem_pressure_precomputed.append((pressure, summary))
									except Exception as e:
										print(f"[FEM事前計算] フレーム{i}でエラー: {e}")
										fem_pressure_precomputed.append(None)
									# 進捗更新
									if (i + 1) % 5 == 0 or (i + 1) == len(transform_data):
										update_progress(i + 1, len(transform_data), f"FEM解析: フレーム {i+1}/{len(transform_data)}")
								n_success = sum(1 for x in fem_pressure_precomputed if x is not None)
								print(f"[FEM事前計算] 完了: {n_success}/{len(transform_data)}フレーム成功")
								# FEMメタデータ保存
								if self._shared_fem_cache and n_success > 0:
									try:
										xlsx_name = os.path.basename(self.transform_group_path.get())
										prox_name = os.path.basename(self.prox_model_path.get())
										dist_name = os.path.basename(self.dist_model_path.get())
										meta_key = compute_content_hash(xlsx_name, prox_name, dist_name)
										meta = {
											'type': 'fem',
											'xlsx': xlsx_name,
											'prox_stl': prox_name,
											'dist_stl': dist_name,
											'frames': len(transform_data),
											'frames_cached': n_success,
											'date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
										}
										self._shared_fem_cache.save_metadata(meta_key, meta)
									except Exception as _e:
										print(f"[FEM] メタデータ保存エラー（無視）: {_e}")
							except Exception as e:
								print(f"[FEM事前計算] 初期化エラー: {e}")
								traceback.print_exc()
								fem_pressure_precomputed = []
						else:
							if not _HAS_FEM:
								print("[FEM事前計算] FEMモジュールが利用できません")
							else:
								print("[FEM事前計算] 関節領域が空のためスキップ")
					
				except Exception as e:
					messagebox.showerror("エラー", f"事前計算中にエラーが発生しました:\n{e}")
					overlap_precomputed = []
					overlap_areas_precomputed = []
					overlap_depths_precomputed = []
					heatmap_precomputed = []
				
				finally:
					# プログレスダイアログを閉じる
					try:
						progress_window.destroy()
					except:
						pass

			# --- プロッターの作成 ---
			# 空のメッシュをプロット可能にする設定（プロッター作成前に設定）
			pv.global_theme.allow_empty_mesh = True
			
			# ディスプレイサイズの取得
			screen_width = self.winfo_screenwidth()
			screen_height = self.winfo_screenheight()
			window_width = int(screen_width * 0.9)
			window_height = int(screen_height * 0.9)
			
			anim_plotter = pv.Plotter(title="アニメーション実行", window_size=(window_width, window_height))
			anim_plotter.set_background("white")
			
			# 近位モデルを追加（透明度0.5でフィッティングと合わせる）
			prox_actor = anim_plotter.add_mesh(prox_mesh, color=self.prox_color, smooth_shading=True, show_edges=False, opacity=0.5)
			
			# 近位座標系の軸を表示
			axis_length = 50.0
			prox_x_line = pv.Line(prox_origin, prox_origin + prox_x_axis * axis_length)
			prox_y_line = pv.Line(prox_origin, prox_origin + prox_y_axis * axis_length)
			prox_z_line = pv.Line(prox_origin, prox_origin + prox_z_axis * axis_length)
			prox_x_actor = anim_plotter.add_mesh(prox_x_line, color='red', line_width=3, label='Prox X')
			prox_y_actor = anim_plotter.add_mesh(prox_y_line, color='green', line_width=3, label='Prox Y')
			prox_z_actor = anim_plotter.add_mesh(prox_z_line, color='blue', line_width=3, label='Prox Z')
			
			# 遠位モデルを追加（透明度0.5）
			dist_actor = anim_plotter.add_mesh(dist_mesh, color=self.dist_color, smooth_shading=True, show_edges=False, opacity=0.3)
			
			# 軟骨モデルを追加（任意）
			prox_cart_actor = None
			dist_cart_actor = None
			if prox_cartilage_mesh is not None:
				prox_cart_actor = anim_plotter.add_mesh(prox_cartilage_mesh, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='近位軟骨')
			if dist_cartilage_mesh is not None:
				dist_cart_actor = anim_plotter.add_mesh(dist_cartilage_mesh, color=self.cs_cartilage_color, smooth_shading=True, show_edges=False, opacity=0.8, label='遠位軟骨')
			
			# 遠位関節領域（球体抽出部分）を強調表示（不透明、色違い）
			# これによりユーザーはオフセットの設定が正しいか視覚的に確認できる
			dist_region_actor_mesh = dist_joint_region.copy()
			dist_region_actor = anim_plotter.add_mesh(dist_region_actor_mesh, color="#00FFFF", smooth_shading=True, show_edges=False, opacity=1.0, label='Extracted Distal Region')
			dist_region_actor.SetVisibility(False)  # デフォルトでは非表示にする（ユーザーの要望）

			# 遠位座標系の軸を表示
			dist_x_line = pv.Line(dist_origin, dist_origin + dist_x_axis * axis_length)
			dist_y_line = pv.Line(dist_origin, dist_origin + dist_y_axis * axis_length)
			dist_z_line = pv.Line(dist_origin, dist_origin + dist_z_axis * axis_length)
			dist_x_actor = anim_plotter.add_mesh(dist_x_line, color='red', line_width=3, label='Dist X')
			dist_y_actor = anim_plotter.add_mesh(dist_y_line, color='green', line_width=3, label='Dist Y')
			dist_z_actor = anim_plotter.add_mesh(dist_z_line, color='blue', line_width=3, label='Dist Z')
			
			# オーバーラップ体積メッシュは廃止

			# --- 抽出領域（関節領域）の表示 ---
			# 近位関節領域と遠位関節領域は削除（ヒートマップで代替）
			
			# 距離ヒートマップメッシュを初期化
			# ヒートマップデータがある場合は最初のフレームで初期化、ない場合はダミー
			print(f"[デバッグ] ヒートマップ初期化: len={len(heatmap_precomputed)}")
			if len(heatmap_precomputed) > 0:
				print(f"[デバッグ] heatmap_precomputed[0]: {heatmap_precomputed[0].n_points}点")
				if heatmap_precomputed[0].n_points > 0:
					heatmap_mesh = heatmap_precomputed[0].copy()
					# スカラー 'distance' が無ければダミーで埋める（キャッシュ/計算の不整合対策）
					if 'distance' not in heatmap_mesh.array_names:
						print(f"[警告] 最初のヒートマップに 'distance' スカラーがありません。ダミー値で埋めます。")
						try:
							n = heatmap_mesh.n_points
							heatmap_mesh['distance'] = np.zeros(n)
						except Exception:
							# それでも失敗する場合は空のダミーメッシュに置き換え
							heatmap_mesh = pv.PolyData()
							heatmap_mesh.points = np.array([[0, 0, 0]])
							heatmap_mesh['distance'] = np.array([0.0])
					print(f"[デバッグ] ヒートマップメッシュを最初のフレームで初期化: {heatmap_mesh.n_points}点")
					print(f"[デバッグ] スカラー: {heatmap_mesh.array_names}")
				else:
					heatmap_mesh = pv.PolyData()
					heatmap_mesh.points = np.array([[0, 0, 0]])
					heatmap_mesh['distance'] = np.array([0.0])
					print(f"[デバッグ] ヒートマップメッシュをダミーで初期化（0点のため）")
			else:
				heatmap_mesh = pv.PolyData()
				# 空のメッシュにダミーのスカラーを追加（エラー回避）
				heatmap_mesh.points = np.array([[0, 0, 0]])
				heatmap_mesh['distance'] = np.array([0.0])
				print(f"[デバッグ] ヒートマップメッシュをダミーで初期化（データなし）")
			
			# カラースケール: 範囲 [-10, 0] mm 固定。めり込み(負)=赤(深いほど濃い) → 接触(0)=緑。
			# 離間(正=範囲外)は透明にして素のモデルを見せる（above-range透明）。
			CLIM_LO, CLIM_HI = -10.0, 0.0
			heatmap_cmap = 'RdYlGn'
			try:
				from matplotlib.colors import LinearSegmentedColormap
				heatmap_cmap = LinearSegmentedColormap.from_list(
					'contact_pen',
					[(0.0, (0.75, 0.0, 0.0)),   # -10mm: 濃い赤
					 (0.5, (1.0, 0.6, 0.0)),    # -5mm:  オレンジ
					 (1.0, (0.1, 0.75, 0.1))])  # 0mm(接触): 緑
			except Exception:
				heatmap_cmap = 'RdYlGn'

			heatmap_actor = anim_plotter.add_mesh(
				heatmap_mesh,
				scalars='distance',
				cmap=heatmap_cmap,     # -10mm=赤 → 0mm=緑
				clim=[CLIM_LO, CLIM_HI],
				opacity=1.0,           # 全体不透明度（Heatmapスライダーで調整）
				show_edges=False,
				label='Heatmap',
			)
			# 範囲外(離間 >0mm)は透明 → 素のモデルを見せる
			try:
				_hm_lut = heatmap_actor.GetMapper().GetLookupTable()
				_hm_lut.SetUseAboveRangeColor(True)
				_hm_lut.SetAboveRangeColor(0.0, 0.0, 0.0, 0.0)
			except Exception:
				pass
			
			# Z-fighting対策：骨の透明度=1.0でもヒートマップが隠れないよう強めのオフセット + Translucent強制
			# （bones を opaque パスで描いた後、heatmap を translucent パスで手前寄せして描く）
			heatmap_mapper = heatmap_actor.GetMapper()
			heatmap_mapper.SetResolveCoincidentTopologyToPolygonOffset()
			heatmap_mapper.SetRelativeCoincidentTopologyPolygonOffsetParameters(-2.0, -100.0)
			try:
				heatmap_actor.ForceTranslucentOn()  # opaque bones の後に描画されるので確実に前面
			except Exception:
				pass

			# 境界線（0mm等高線）を近位抽出領域メッシュ表面に張り付けて表示
			print(f"[初期化] 境界線（等高線）の初期化を開始...")
			boundary_actor = None
			# 近位抽出領域メッシュに距離スカラーをマッピングして等高線抽出
			if len(heatmap_precomputed) > 0 and 'distance' in heatmap_precomputed[0].array_names and prox_joint_region.n_points > 0:
				try:
					print(f"[初期化] 近位メッシュをコピー中...")
					# 近位抽出領域メッシュに距離値をマッピング
					prox_with_distance = prox_joint_region.copy()
					
					print(f"[初期化] ヒートマップデータを取得中...")
					# ヒートマップの点と距離値を取得
					heatmap_points = heatmap_precomputed[0].points
					heatmap_distances = heatmap_precomputed[0]['distance']
					
					print(f"[初期化] KDTreeを構築中（{len(heatmap_points)}点）...")
					# 近位メッシュの各頂点に最近傍の距離値を割り当て
					from scipy.spatial import cKDTree
					tree = cKDTree(heatmap_points)
					
					print(f"[初期化] 最近傍探索を実行中（{prox_with_distance.n_points}点）...")
					_, indices = tree.query(prox_with_distance.points, k=1)
					prox_with_distance['distance'] = heatmap_distances[indices]
					print(f"[初期化] 距離マッピング完了")
					
					# メッシュ情報を表示
					print(f"[デバッグ] 近位メッシュ: {prox_with_distance.n_points}点, {prox_with_distance.n_cells}面")
					print(f"[デバッグ] 距離範囲: min={prox_with_distance['distance'].min():.3f}, max={prox_with_distance['distance'].max():.3f}")
					
					print(f"[初期化] 等高線（0mm）を抽出中...")
					# PyVistaのcontour機能で距離=0.0mmの等高線を面メッシュから抽出
					contour_mesh = prox_with_distance.contour(isosurfaces=[0.0], scalars='distance')
					print(f"[初期化] 等高線抽出完了")
					
					if contour_mesh.n_points > 0:
						print(f"[デバッグ] 等高線抽出成功: {contour_mesh.n_points}点, {contour_mesh.n_lines}線")
						# 境界線メッシュを作成（後で更新用）
						boundary_mesh = contour_mesh.copy()
					else:
						print(f"[警告] 等高線が抽出されませんでした（0mm付近の点がメッシュ上にない可能性）")
						boundary_mesh = pv.PolyData()
				except Exception as e:
					print(f"[警告] 等高線抽出失敗: {e}")
					# import traceback # 既にグローバルでインポート済み
					traceback.print_exc()
					boundary_mesh = pv.PolyData()
			else:
				boundary_mesh = pv.PolyData()
			
			print(f"[初期化] 境界線actorを作成中...")
			# 境界線actorを一度だけ作成（以降はboundary_meshのポイントを更新するだけで点滅しない）
			boundary_actor = anim_plotter.add_mesh(
				boundary_mesh,
				color='white',
				line_width=5,
				render_lines_as_tubes=True,
				lighting=False
			)
			# 骨の透明度=1.0でも境界線が隠れないよう手前寄せ + Translucent強制
			try:
				_b_mapper = boundary_actor.GetMapper()
				_b_mapper.SetResolveCoincidentTopologyToPolygonOffset()
				_b_mapper.SetRelativeCoincidentTopologyLineOffsetParameters(-3.0, -200.0)
				boundary_actor.ForceTranslucentOn()
			except Exception:
				pass

			print(f"[初期化] 境界線actor作成完了: type={type(boundary_actor)}")
			
			print(f"[初期化] 表示設定を構成中...")
			# ヒートマップがある場合は初期表示ON、ない場合は非表示
			if len(heatmap_precomputed) > 0:
				print(f"[デバッグ] ヒートマップデータあり: {len(heatmap_precomputed)}フレーム")
				print(f"[デバッグ] 最初のフレームのヒートマップ: {heatmap_precomputed[0].n_points}点")
				if 'distance' in heatmap_precomputed[0].array_names:
					print(f"[デバッグ] distanceスカラーあり")
				heatmap_actor.SetVisibility(True)
				boundary_actor.SetVisibility(True)
				print(f"[デバッグ] ヒートマップ表示をONに設定")
				print(f"[デバッグ] 境界線表示をONに設定（visibility={boundary_actor.GetVisibility()}）")
			else:
				print(f"[デバッグ] ヒートマップデータなし")
				heatmap_actor.SetVisibility(False)
				boundary_actor.SetVisibility(False)

			# --- FEM接触圧メッシュの初期化（ヒートマップと同じ形状、異なるスカラー） ---
			print(f"[初期化] FEM接触圧メッシュの初期化を開始...")
			pressure_mesh = None
			pressure_actor = None
			heatmap_display_mode = [0]  # 0=distance, 1=contact_pressure

			if fem_pressure_precomputed and any(x is not None for x in fem_pressure_precomputed):
				try:
					# 最初の有効なフレームを探す
					first_pressure = None
					first_summary = None
					for fp in fem_pressure_precomputed:
						if fp is not None:
							first_pressure, first_summary = fp
							break

					if first_pressure is not None:
						# ヒートマップと同じジオメトリを持つメッシュを作成
						pressure_mesh = heatmap_mesh.copy()

						# 圧力データをマッピング
						if prox_fem_surface is not None and prox_fem_surface.n_points > 0:
							if prox_fem_surface.n_points == pressure_mesh.n_points:
								pressure_mesh['contact_pressure'] = first_pressure.copy()
							else:
								from scipy.spatial import cKDTree
								tree = cKDTree(prox_fem_surface.points)
								_, idx = tree.query(pressure_mesh.points, k=1)
								pressure_mesh['contact_pressure'] = first_pressure[idx]
						else:
							pressure_mesh['contact_pressure'] = first_pressure.copy()

						# 全フレームの圧力範囲を走査してスケールを決定
						p_max = 0.0
						for fp in fem_pressure_precomputed:
							if fp is not None:
								p_max = max(p_max, float(np.nanmax(fp[0])))
						if p_max < 1e-6:
							p_max = 1.0

						print(f"[初期化] 圧力範囲: 0 - {p_max:.2f} MPa")

						pressure_actor = anim_plotter.add_mesh(
							pressure_mesh,
							scalars='contact_pressure',
							cmap='hot_r',
							clim=[0, p_max],
							show_edges=False,
							opacity=1.0,
							scalar_bar_args={'title': '接触圧 [MPa]', 'color': 'black'},
							label='FEM接触圧'
						)

						# Z-fighting対策（骨の透明度=1.0でも接触圧が隠れないよう強めに手前寄せ）
						p_mapper = pressure_actor.GetMapper()
						p_mapper.SetResolveCoincidentTopologyToPolygonOffset()
						p_mapper.SetRelativeCoincidentTopologyPolygonOffsetParameters(-2.0, -100.0)
						try:
							pressure_actor.ForceTranslucentOn()
						except Exception:
							pass

						pressure_actor.SetVisibility(False)
						print(f"[初期化] FEM接触圧メッシュを初期化: {pressure_mesh.n_points}点, p_max={p_max:.2f}MPa")
				except Exception as e:
					print(f"[警告] FEM接触圧メッシュの初期化失敗: {e}")
					traceback.print_exc()
					pressure_mesh = None
					pressure_actor = None
			else:
				print(f"[初期化] FEM事前計算データがないため、接触圧メッシュは作成されません")

			print(f"[初期化] ウィジェット（表示制御）を構成中...")
			# オーバーラップメッシュの枠線表示状態
			overlap_edges_visible = [True]  # デフォルトで枠線表示
			
			# --- ウィジェット: 表示/非表示と透明度 ---
			# 近位モデルの表示/非表示
			def toggle_prox_model(state):
				prox_actor.SetVisibility(state)
			
			# 近位座標系の表示/非表示
			def toggle_prox_axes(state):
				prox_x_actor.SetVisibility(state)
				prox_y_actor.SetVisibility(state)
				prox_z_actor.SetVisibility(state)
			
			# 遠位モデルの表示/非表示
			def toggle_dist_model(state):
				dist_actor.SetVisibility(state)
			
			# 遠位座標系の表示/非表示
			def toggle_dist_axes(state):
				dist_x_actor.SetVisibility(state)
				dist_y_actor.SetVisibility(state)
				dist_z_actor.SetVisibility(state)

			# ヒートマップの表示/非表示
			def toggle_heatmap(state):
				if heatmap_precomputed:  # 事前計算済みの場合のみ有効
					heatmap_actor.SetVisibility(state)
				else:
					heatmap_actor.SetVisibility(False)
			
			# 境界線（0mm等高線）の表示/非表示
			def toggle_boundary(state):
				if boundary_actor is not None:
					boundary_actor.SetVisibility(state)

			# FEM接触圧の表示切り替え
			def toggle_pressure_display(state):
				if state:
					# 圧力表示ON: distanceをOFF、pressureをON
					heatmap_actor.SetVisibility(False)
					if pressure_actor is not None:
						pressure_actor.SetVisibility(True)
					heatmap_display_mode[0] = 1
				else:
					# 圧力表示OFF: distanceをON（ヒートマップが事前計算されている場合）
					if heatmap_precomputed:
						heatmap_actor.SetVisibility(True)
					if pressure_actor is not None:
						pressure_actor.SetVisibility(False)
					heatmap_display_mode[0] = 0

			# 近位モデルの透明度
			def update_prox_opacity(value):
				prox_actor.GetProperty().SetOpacity(value)
			
			# 遠位モデルの透明度
			def update_dist_opacity(value):
				dist_actor.GetProperty().SetOpacity(value)
			
			# オーバーラップ体積の透明度制御は削除
			
			# チェックボックスとラベル - 近位
			anim_plotter.add_text("Prox Model", position=(10, 10), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_prox_model, value=True, position=(130, 12), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			
			anim_plotter.add_text("Prox Axes", position=(10, 40), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_prox_axes, value=True, position=(130, 42), size=20, border_size=1, color_on=self.prox_color, color_off="white")
			
			# チェックボックスとラベル - 遠位
			anim_plotter.add_text("Dist Model", position=(10, 70), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_dist_model, value=True, position=(130, 72), size=20, border_size=1, color_on=self.dist_color, color_off="white")
			
			anim_plotter.add_text("Dist Axes", position=(10, 100), font_size=10, color="black")
			anim_plotter.add_checkbox_button_widget(toggle_dist_axes, value=True, position=(130, 102), size=20, border_size=1, color_on=self.dist_color, color_off="white")

			# チェックボックスとラベル - ヒートマップ（事前計算されている場合のみ表示）
			heatmap_enabled = len(heatmap_precomputed) > 0
			if heatmap_enabled:
				anim_plotter.add_text("Heatmap", position=(10, 130), font_size=10, color="black")
				anim_plotter.add_checkbox_button_widget(
					toggle_heatmap, 
					value=True,  # 初期表示ON
					position=(130, 132), 
					size=20, 
					border_size=1, 
					color_on="orange", 
					color_off="white"
				)
				
				# チェックボックスとラベル - 境界線（0mm等高線）
				anim_plotter.add_text("Boundary", position=(10, 160), font_size=10, color="black")
				anim_plotter.add_checkbox_button_widget(
					toggle_boundary,
					value=True,  # 初期表示ON
					position=(130, 162),
					size=20,
					border_size=1,
					color_on="white",
					color_off="gray"
				)

			# チェックボックスとラベル - FEM接触圧（事前計算されている場合のみ表示）
			if fem_pressure_precomputed and any(x is not None for x in fem_pressure_precomputed):
				anim_plotter.add_text("FEM圧", position=(10, 190), font_size=10, color="black")
				anim_plotter.add_checkbox_button_widget(
					toggle_pressure_display,
					value=False,  # デフォルトで非表示
					position=(130, 192),
					size=20,
					border_size=1,
					color_on="purple",
					color_off="white"
				)
				next_y = 220
			else:
				next_y = 190
			
			# Overlap UIは削除済み
			
			# 透明度スライダー - 近位
			anim_plotter.add_slider_widget(
				update_prox_opacity,
				rng=[0.0, 1.0],
				value=0.5,
				title=" ",
				pointa=(0.10, 0.025),
				pointb=(0.20, 0.025),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.prox_color,
			)
			
			# 透明度スライダー - 遠位
			anim_plotter.add_slider_widget(
				update_dist_opacity,
				rng=[0.0, 1.0],
				value=0.5,
				title=" ",
				pointa=(0.10, 0.085),
				pointb=(0.20, 0.085),
				style='modern',
				tube_width=0.01,
				slider_width=0.02,
				color=self.dist_color,
			)
			
			# 重複体積の透明度スライダーは削除
			
			# --- アニメーション設定 ---
			# 遠位メッシュの初期状態を保存（スライダーと同じロジック用）
			dist_mesh_for_anim = dist_mesh.copy()
			dist_origin_initial = np.array([0.0, 0.0, 0.0])  # 変換後の遠位座標系原点
			dist_x_axis_initial = np.array([1.0, 0.0, 0.0])
			dist_y_axis_initial = np.array([0.0, 1.0, 0.0])
			dist_z_axis_initial = np.array([0.0, 0.0, 1.0])

			# 近位メッシュ・軟骨・座標系の初期状態を保存（「脛骨を固定」モードで inv(M) を適用するため）
			prox_mesh_for_anim = prox_mesh.copy()
			prox_cartilage_for_anim = prox_cartilage_mesh.copy() if prox_cartilage_mesh is not None else None
			prox_origin_initial = np.array([0.0, 0.0, 0.0])
			prox_x_axis_initial = np.array([1.0, 0.0, 0.0])
			prox_y_axis_initial = np.array([0.0, 1.0, 0.0])
			prox_z_axis_initial = np.array([0.0, 0.0, 1.0])

			# 脛骨固定モード（True: 脛骨=dist を固定・大腿骨=prox が動く / False: 従来通り prox 固定・dist 動く）
			fix_tibia_mode = [False]

			# 遠位関節領域の初期状態を保存（メッシュと初期点群）
			dist_region_for_anim = dist_joint_region.copy()  # dist_region_meshではなくdist_joint_regionから
			
			# 現在のフレーム番号
			current_frame = [0]  # リストで包んで参照を保持（見かけのフレーム: Excelデータのインデックス）
			actual_frame_counter = [0]  # 実際のフレームカウンタ（5ms間隔で増加）
			animation_start_time = [None]  # アニメーション開始時刻（実時間）
			playback_speed = [1.0]  # 再生速度（1.0 = 等倍速）
			last_scale_update_time = [0.0]  # 最後に再生バーを更新した時刻
			scale_update_interval = [0.2]  # 再生バー更新間隔（秒）- YouTubeスタイル
			print(f"[初期化] アニメーション制御変数を初期化中...")
			after_id = [None]  # Tkinter afterのID（アニメーション用）
			check_id = [None]  # Tkinter afterのID（ウィンドウチェック用）
			is_animation_active = [True]  # アニメーションが有効かどうか
			is_paused = [False]  # 一時停止フラグ
			is_seeking = [False]  # シーク中フラグ（再生バー操作中）
			is_programmatic_update = [False]  # プログラムから再生バーを更新中かどうか
			user_is_dragging = [False]  # ユーザーがドラッグ中かどうか
			
			print(f"[初期化] 再生コントロールウィンドウを作成中...")
			# データ総数をログ出力
			print(f"[アニメーション] データ総数: {len(transform_data)} フレーム")
			
			# CSVの時間データを取得（各フレームのTime値）
			frame_times = [frame_data["time"] for frame_data in transform_data]
			print(f"[アニメーション] Time範囲: {frame_times[0]:.3f}s ~ {frame_times[-1]:.3f}s")
			
			# データが昇順に並んでいるか確認
			is_sorted = all(frame_times[i] <= frame_times[i+1] for i in range(len(frame_times)-1))
			print(f"[アニメーション] データは時間順: {is_sorted}")
			if not is_sorted:
				print("[警告] データが時間順に並んでいません！")
			
			# 最大時間を取得（ループ用）
			max_time = frame_times[-1]
			
			# 再生コントロール用のTkinterウィンドウを作成
			control_window = tk.Toplevel(self)
			control_window.title("再生コントロール")
			control_window.geometry("850x360")
			control_window.resizable(True, True)  # リサイズ可能（Mac対応）
			control_window.minsize(400, 200)       # 最小サイズを設定
			control_window.attributes('-topmost', True)  # 常に最前面に表示
			
			# 情報表示フレーム
			info_frame = ttk.Frame(control_window)
			info_frame.pack(pady=5, padx=10, fill=tk.X)
			
			# フレーム情報表示（横並び）
			visible_frame_label = ttk.Label(
				info_frame,
				text=f"Frame: 0/{len(transform_data)-1} | Time: {frame_times[0]:.3f}s",
				font=(self.ui_font_family, 10)
			)
			visible_frame_label.pack(side=tk.LEFT, padx=5)
			
			actual_time_label = ttk.Label(
				info_frame,
				text=f"Actual: 0.000s",
				font=(self.ui_font_family, 10),
				foreground="blue"
			)
			actual_time_label.pack(side=tk.LEFT, padx=5)
			
			# 再生速度表示
			speed_label = ttk.Label(
				info_frame,
				text=f"Speed: 1.0x",
				font=(self.ui_font_family, 10),
				foreground="darkgreen"
			)
			speed_label.pack(side=tk.LEFT, padx=5)
			
			# オーバーラップ面積表示（新規追加）
			overlap_area_label = ttk.Label(
				info_frame,
				text=f"Area: 0.00 mm²",
				font=(self.ui_font_family, 10),
				foreground="red"
			)
			if len(overlap_areas_precomputed) > 0:
				overlap_area_label.pack(side=tk.LEFT, padx=5)
			
			# 接触深度表示（新規追加）
			overlap_depth_label = ttk.Label(
				info_frame,
				text=f"Depth: 0.00 mm",
				font=(self.ui_font_family, 10),
				foreground="purple"
			)
			if len(overlap_depths_precomputed) > 0:
				overlap_depth_label.pack(side=tk.LEFT, padx=5)

			# 最大めり込み量表示（新規追加）
			max_penetration_label = ttk.Label(
				info_frame,
				text=f"Max Pent: 0.00 mm",
				font=(self.ui_font_family, 10),
				foreground="red"
			)
			max_penetration_label.pack(side=tk.LEFT, padx=5)
			
			# 再生バー（スライダー）
			# YouTubeスタイル: マウスイベントを追跡
			def on_scale_button_press(event):
				"""マウスボタンが押された（ドラッグ開始）"""
				user_is_dragging[0] = True
			
			def on_scale_button_release(event):
				"""マウスボタンが離された（ドラッグ終了）"""
				user_is_dragging[0] = False
			
			playback_scale = tk.Scale(
				control_window,
				from_=0,
				to=len(transform_data)-1,
				orient=tk.HORIZONTAL,
				label="Frame Position",
				command=lambda val: on_playback_scale_change(int(val))
			)
			playback_scale.pack(pady=5, padx=10, fill=tk.X)
			
			# マウスイベントをバインド
			playback_scale.bind("<ButtonPress-1>", on_scale_button_press)
			playback_scale.bind("<ButtonRelease-1>", on_scale_button_release)
			
			# 再生速度スライダー
			def on_speed_change(val):
				"""再生速度が変更されたときの処理"""
				speed = float(val)
				playback_speed[0] = speed
				speed_label.config(text=f"Speed: {speed:.2f}x")
				# 再生中の場合、開始時刻を調整
				if animation_start_time[0] is not None and not is_paused[0]:
					actual_time = actual_frame_counter[0] * 0.005
					animation_start_time[0] = time.time() - (actual_time / speed)
			
			speed_scale = tk.Scale(
				control_window,
				from_=0.25,
				to=10.0,
				resolution=0.25,
				orient=tk.HORIZONTAL,
				label="Playback Speed (0.25x - 10x)",
				command=on_speed_change
			)
			speed_scale.set(1.0)  # 初期値: 等倍速
			speed_scale.pack(pady=5, padx=10, fill=tk.X)
			
			# 一時停止/再生ボタン
			button_frame = ttk.Frame(control_window)
			button_frame.pack(pady=5)
			
			pause_button = ttk.Button(button_frame, text="一時停止", width=15)
			pause_button.pack(side=tk.LEFT, padx=5)
			
			# CSVエクスポートボタン（新規追加）
			def export_to_csv():
				"""オーバーラップ面積と接触深度データをCSVにエクスポート"""
				if len(overlap_areas_precomputed) == 0:
					messagebox.showwarning("エクスポート", "オーバーラップデータがありません。\n事前計算を実行してください。")
					return
				
				# ファイル保存ダイアログ
				filepath = filedialog.asksaveasfilename(
					title="CSVファイルを保存",
					defaultextension=".csv",
					filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
				)
				
				if not filepath:
					return  # キャンセルされた
				
				try:
					import csv
					with open(filepath, 'w', newline='', encoding='utf-8') as f:
						writer = csv.writer(f)
						# ヘッダー行
						writer.writerow(["IE (deg)", "Overlap Area (mm^2)", "Contact Depth (mm)"])
						
						# データ行
						for i, frame_data in enumerate(transform_data):
							ie_angle = frame_data["angles"][2]  # IE角度（Y軸回転）
							area = overlap_areas_precomputed[i] if i < len(overlap_areas_precomputed) else 0.0
							depth = overlap_depths_precomputed[i] if i < len(overlap_depths_precomputed) else 0.0
							writer.writerow([f"{ie_angle:.6f}", f"{area:.6f}", f"{depth:.6f}"])
					
					messagebox.showinfo("エクスポート完了", f"CSVファイルを保存しました:\n{filepath}")
				
				except Exception as e:
					messagebox.showerror("エクスポート失敗", f"CSVファイルの保存に失敗しました:\n{e}")
			
			export_button = ttk.Button(button_frame, text="CSV出力", width=15, command=export_to_csv)
			if len(overlap_areas_precomputed) > 0:
				export_button.pack(side=tk.LEFT, padx=5)

			# --- このモデルを出力 ボタン (新規機能) ---
			def export_current_frame_model():
				"""現在のフレームの近位・遠位（変換後）モデルを1つのファイルに出力"""
				# 一時停止してから保存処理に入る
				current_was_paused = is_paused[0]
				is_paused[0] = True
				pause_button.config(text="再生")
				
				try:
					# 現在のフレームインデックス
					idx = current_frame[0]
					
					# 保存先を選択
					save_path = filedialog.asksaveasfilename(
						title=f"現在のモデルを出力 (Frame {idx})",
						defaultextension=".stl",
						filetypes=[("STL files", "*.stl"), ("OBJ files", "*.obj")],
						initialfile=f"model_frame_{idx}.stl"
					)
					
					if not save_path:
						# キャンセルされたら再生状態を戻す
						if not current_was_paused:
							toggle_pause()
						return
					
					# 1. 近位メッシュのコピー（変更不要）
					prox_copy = prox_mesh.copy()
					
					# 2. 遠位メッシュは既に変換済み（anim_plotter内で変換されている）
					# しかし、PyVista上のMeshは `dist_mesh.points` が `show_frame` で更新されているので
					# 現在の `dist_mesh` をそのまま使えばよい
					dist_copy = dist_mesh.copy()
					
					# 3. 2つのメッシュを結合
					merged_mesh = prox_copy + dist_copy
					
					# 4. 保存
					merged_mesh.save(save_path)
					
					messagebox.showinfo("出力完了", f"モデルを出力しました:\nFrame: {idx}\nPath: {save_path}")
					
				except Exception as e:
					messagebox.showerror("エラー", f"モデル出力中にエラーが発生しました:\n{e}\n{traceback.format_exc()}")
				
				# 再生状態を戻す（ただし、明示的に止めたい場合もあるので、ユーザーが再開ボタンを押す運用でもよいが、ここでは元の状態に戻す）
				if not current_was_paused:
					toggle_pause()

			export_model_button = ttk.Button(button_frame, text="このモデルを出力", width=18, command=export_current_frame_model)
			export_model_button.pack(side=tk.LEFT, padx=5)

			# --- 「脛骨を固定」トグル（大腿骨=prox / 脛骨=dist の視覚的な固定側を入替） ---
			fix_tibia_button = ttk.Button(button_frame, text="脛骨を固定", width=18)
			def toggle_fix_tibia():
				fix_tibia_mode[0] = not fix_tibia_mode[0]
				if fix_tibia_mode[0]:
					fix_tibia_button.config(text="脛骨固定中 (元に戻す)")
				else:
					fix_tibia_button.config(text="脛骨を固定")
				# 現在のフレームで即座に再描画してモード切替を反映
				try:
					show_frame(current_frame[0], force_render=True)
				except Exception as e:
					print(f"[脛骨固定トグル] 再描画失敗: {e}")
			fix_tibia_button.config(command=toggle_fix_tibia)
			fix_tibia_button.pack(side=tk.LEFT, padx=5)

			# --- FEM接触解析ボタン（関節領域がある場合に表示）---
			if _HAS_FEM and prox_joint_region.n_points > 0 and dist_joint_region.n_points > 0:
				def run_fem_current_frame():
					"""現在フレームの姿勢でFEM接触解析を実行（関節領域で計算）"""
					was_paused = is_paused[0]
					is_paused[0] = True
					pause_button.config(text="再生")
					# 現在フレームの遠位関節領域を変換
					frame_idx = current_frame[0] if hasattr(current_frame, '__getitem__') else 0
					matrix = transform_data[frame_idx]['matrix']
					dist_joint_current = self._apply_transform(dist_joint_region.copy(), matrix)
					prox_fem = prox_joint_region.extract_surface() if hasattr(prox_joint_region, 'extract_surface') else prox_joint_region
					dist_fem = dist_joint_current.extract_surface() if hasattr(dist_joint_current, 'extract_surface') else dist_joint_current
					self._run_fem_on_current_frame(
						prox_fem, dist_fem,
						prox_bone_mesh=prox_mesh, dist_bone_mesh=dist_mesh,
					)
					if not was_paused:
						toggle_pause()

				fem_anim_btn = ttk.Button(button_frame, text="FEM解析（関節領域）", width=18, command=run_fem_current_frame)
				fem_anim_btn.pack(side=tk.LEFT, padx=5)

			# --- レポート生成ボタン ---
			if _HAS_REPORT and _HAS_FEM and prox_joint_region.n_points > 0 and dist_joint_region.n_points > 0:
				def generate_animation_report():
					"""アニメーションデータからExcel+Wordレポートを生成"""
					was_paused = is_paused[0]
					is_paused[0] = True
					pause_button.config(text="再生")

					# 保存先フォルダを選択
					from tkinter import filedialog as fd_report
					output_dir = fd_report.askdirectory(title="レポート保存先フォルダを選択")
					if not output_dir:
						if not was_paused:
							toggle_pause()
						return

					# 元のExcelファイルパス
					source_xlsx = self.transform_group_path.get()

					# プログレス表示用
					report_progress_window = tk.Toplevel(control_window)
					report_progress_window.title("レポート生成中")
					report_progress_window.geometry("400x120")
					report_progress_window.transient(control_window)
					report_progress_label = ttk.Label(report_progress_window, text="レポートを生成しています...", font=("Arial", 11))
					report_progress_label.pack(pady=(15, 5))
					report_progress_bar = ttk.Progressbar(report_progress_window, mode='determinate', length=350)
					report_progress_bar.pack(pady=5)
					report_progress_window.update()

					def progress_cb(message, percent):
						try:
							report_progress_label.config(text=message)
							report_progress_bar['value'] = percent
							report_progress_window.update()
						except Exception:
							pass

					try:
						result = generate_report_from_animation(
							app=self,
							transform_data=transform_data,
							source_xlsx_path=source_xlsx,
							output_dir=output_dir,
							prox_joint_region=prox_joint_region,
							dist_joint_region=dist_joint_region,
							prox_bone_mesh=prox_mesh,
							dist_bone_mesh=dist_mesh,
							dist_origin_initial=dist_origin_initial,
							progress_callback=progress_cb,
						)
						report_progress_window.destroy()

						# 結果表示
						msg_parts = ["レポート生成が完了しました。\n"]
						if result.get('excel_path'):
							msg_parts.append(f"Excel: {os.path.basename(result['excel_path'])}")
						if result.get('word_path'):
							msg_parts.append(f"Word: {os.path.basename(result['word_path'])}")
						msg_parts.append(f"\n保存先: {output_dir}")
						messagebox.showinfo("レポート生成完了", "\n".join(msg_parts))

					except Exception as e:
						try:
							report_progress_window.destroy()
						except Exception:
							pass
						messagebox.showerror("エラー", f"レポート生成中にエラーが発生しました:\n{e}")

					if not was_paused:
						toggle_pause()

				report_anim_btn = ttk.Button(button_frame, text="レポート生成", width=14, command=generate_animation_report)
				report_anim_btn.pack(side=tk.LEFT, padx=5)
			# --------------------------------------------
			
			def on_playback_scale_change(frame_idx):
				"""再生バースライダーが操作されたときの処理（YouTubeスタイル）"""
				# プログラムからの更新の場合は何もしない（ユーザー操作のみ処理）
				if is_programmatic_update[0]:
					return
				
				# ユーザーがドラッグ中の場合のみシークする
				if user_is_dragging[0]:
					is_seeking[0] = True
					current_frame[0] = frame_idx
					show_frame(frame_idx, force_render=True)
					# 見かけのフレーム表示を更新
					time_val = frame_times[frame_idx]
					visible_frame_label.config(text=f"Frame: {frame_idx}/{len(transform_data)-1} | Time: {time_val:.3f}s")
					# シーク時に実際の時間もリセット
					actual_frame_counter[0] = int(time_val / 0.005)
					actual_time = actual_frame_counter[0] * 0.005
					actual_time_label.config(text=f"Actual: {actual_time:.3f}s")
					# アニメーション開始時刻もリセット（再生中の場合、再生速度を考慮）
					if animation_start_time[0] is not None:
						animation_start_time[0] = time.time() - (actual_time / playback_speed[0])
					is_seeking[0] = False
			
			def toggle_pause():
				"""一時停止/再生を切り替え"""
				is_paused[0] = not is_paused[0]
				if is_paused[0]:
					pause_button.config(text="再生")
				else:
					pause_button.config(text="一時停止")
					# 再生再開時に実時間をリセット（再生速度を考慮）
					actual_time = actual_frame_counter[0] * 0.005
					animation_start_time[0] = time.time() - (actual_time / playback_speed[0])
			
			pause_button.config(command=toggle_pause)
			
			# ウィンドウが閉じられたときの処理
			def on_control_window_close():
				is_animation_active[0] = False
				control_window.destroy()
			
			control_window.protocol("WM_DELETE_WINDOW", on_control_window_close)
			
			print(f"[初期化] アニメーション関数を定義中...")
			# 再生バー用の変数（PyVistaスライダーは削除）
			frame_text_actor = [None]  # 見かけのフレーム（CSVのTime値）
			actual_frame_text_actor = [None]  # 実際のフレーム（0.005秒ごと）
			six_axis_text_actor = [None]  # 6軸データ表示用
			
			def show_frame(frame_idx, force_render=False):
				"""指定されたフレームを表示"""
				nonlocal boundary_actor, boundary_mesh  # 外側のboundary_actorとboundary_meshを参照
				
				# アニメーションが無効化されているかチェック（force_render時は無視）
				if not force_render and not is_animation_active[0]:
					return
				
				# ウィンドウが閉じられていないかチェック
				if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
					is_animation_active[0] = False
					return
				
				# render windowが有効かチェック
				if hasattr(anim_plotter, 'render_window'):
					try:
						if not anim_plotter.render_window:
							is_animation_active[0] = False
							return
					except:
						is_animation_active[0] = False
						return
				
				try:
					if frame_idx >= len(transform_data):
						frame_idx = 0  # ループ再生
					
					# 現在のフレームのデータを取得
					try:
						frame_data = transform_data[frame_idx]
						time_val = frame_data["time"]
						angles = frame_data["angles"]  # [FE, VV, IE]
						trans = frame_data["translations"]  # [ML, AP, PD]
					except Exception as data_err:
						print(f"[エラー] フレーム{frame_idx}のデータ取得失敗: {data_err}")
						return
					
					# 6軸データをPyVistaウィンドウに表示（右上）
					six_axis_text = (
						f"FE: {angles[0]:+7.2f}°\n"
						f"VV: {angles[1]:+7.2f}°\n"
						f"IE: {angles[2]:+7.2f}°\n"
						f"ML: {trans[0]:+7.2f} mm\n"
						f"AP: {trans[1]:+7.2f} mm\n"
						f"PD: {trans[2]:+7.2f} mm"
					)
					
					# 初回のみactor作成、以降はテキスト更新のみ（点滅防止）
					if six_axis_text_actor[0] is None:
						try:
							six_axis_text_actor[0] = anim_plotter.add_text(
								six_axis_text,
								position='upper_right',
								font_size=12,
								color='black',
								font='courier'
							)
						except:
							pass
					else:
						# actorが既に存在する場合、テキストだけ更新
						try:
							six_axis_text_actor[0].SetText(3, six_axis_text)
						except:
							pass
					
					# 同次変換行列を構築（スライダーと同じロジック）
					matrix = self._build_transform_matrix(
						rz=angles[0],  # FE
						rx=angles[1],  # VV
						ry=angles[2],  # IE
						ml=trans[0],   # ML
						ap=trans[1],   # AP
						pd=trans[2]    # PD
					)

					# 「脛骨を固定」モードの分岐:
					#   通常 (False): 大腿骨=prox 固定・脛骨=dist を matrix で駆動（従来通り）
					#   固定 (True):  脛骨=dist 固定・大腿骨=prox を inv(matrix) で駆動（相対姿勢は等価）
					if fix_tibia_mode[0]:
						try:
							inv_matrix = np.linalg.inv(matrix)
						except Exception:
							inv_matrix = np.eye(4)
						dist_M = None       # dist は初期位置に戻す
						prox_M = inv_matrix # prox に逆変換を適用
					else:
						dist_M = matrix
						prox_M = None       # prox は初期位置のまま

					# ピボット付き 4x4 変換を点群に適用するローカルヘルパー
					def _apply_pivot(pts_arr, M4, pivot):
						centered = pts_arr - pivot
						homo = np.hstack([centered, np.ones((centered.shape[0], 1))])
						return (M4 @ homo.T).T[:, :3] + pivot

					axis_length = 50.0

					# --- 遠位メッシュの更新（dist_M があれば変換、無ければ初期位置に戻す） ---
					if dist_M is not None:
						dist_mesh.points = _apply_pivot(dist_mesh_for_anim.points.copy(), dist_M, dist_origin_initial)
					else:
						dist_mesh.points = dist_mesh_for_anim.points.copy()

					# 遠位軟骨メッシュ
					if dist_cartilage_for_anim is not None and dist_cartilage_mesh is not None:
						if dist_M is not None:
							dist_cartilage_mesh.points = _apply_pivot(dist_cartilage_for_anim.points.copy(), dist_M, dist_origin_initial)
						else:
							dist_cartilage_mesh.points = dist_cartilage_for_anim.points.copy()

					# --- 遠位関節領域（抽出球）---
					if dist_region_for_anim.n_points > 0:
						if dist_M is not None:
							dist_region_actor_mesh.points = _apply_pivot(dist_region_for_anim.points.copy(), dist_M, dist_origin_initial)
						else:
							dist_region_actor_mesh.points = dist_region_for_anim.points.copy()

					# --- 遠位座標系の軸 ---
					if dist_M is not None:
						td_origin = (dist_M @ np.array([0, 0, 0, 1]))[:3] + dist_origin_initial
						td_rot = dist_M[:3, :3]
						tdx = td_rot @ dist_x_axis_initial
						tdy = td_rot @ dist_y_axis_initial
						tdz = td_rot @ dist_z_axis_initial
						dist_x_line.points = np.array([td_origin, td_origin + tdx * axis_length])
						dist_y_line.points = np.array([td_origin, td_origin + tdy * axis_length])
						dist_z_line.points = np.array([td_origin, td_origin + tdz * axis_length])
					else:
						dist_x_line.points = np.array([dist_origin_initial, dist_origin_initial + dist_x_axis_initial * axis_length])
						dist_y_line.points = np.array([dist_origin_initial, dist_origin_initial + dist_y_axis_initial * axis_length])
						dist_z_line.points = np.array([dist_origin_initial, dist_origin_initial + dist_z_axis_initial * axis_length])

					# --- 近位側（脛骨固定モードでのみ動く） ---
					if prox_M is not None:
						prox_mesh.points = _apply_pivot(prox_mesh_for_anim.points.copy(), prox_M, dist_origin_initial)
						if prox_cartilage_for_anim is not None and prox_cartilage_mesh is not None:
							prox_cartilage_mesh.points = _apply_pivot(prox_cartilage_for_anim.points.copy(), prox_M, dist_origin_initial)
						tp_origin = (prox_M @ np.array([0, 0, 0, 1]))[:3] + dist_origin_initial
						tp_rot = prox_M[:3, :3]
						tpx = tp_rot @ prox_x_axis_initial
						tpy = tp_rot @ prox_y_axis_initial
						tpz = tp_rot @ prox_z_axis_initial
						prox_x_line.points = np.array([tp_origin, tp_origin + tpx * axis_length])
						prox_y_line.points = np.array([tp_origin, tp_origin + tpy * axis_length])
						prox_z_line.points = np.array([tp_origin, tp_origin + tpz * axis_length])
					else:
						# 通常モード: 近位は初期位置のまま（再アサインで確実に戻す）
						prox_mesh.points = prox_mesh_for_anim.points.copy()
						if prox_cartilage_for_anim is not None and prox_cartilage_mesh is not None:
							prox_cartilage_mesh.points = prox_cartilage_for_anim.points.copy()
						prox_x_line.points = np.array([prox_origin_initial, prox_origin_initial + prox_x_axis_initial * axis_length])
						prox_y_line.points = np.array([prox_origin_initial, prox_origin_initial + prox_y_axis_initial * axis_length])
						prox_z_line.points = np.array([prox_origin_initial, prox_origin_initial + prox_z_axis_initial * axis_length])
					
					# 重複体積の更新ロジックは削除
					
					# ヒートマップの更新（事前計算済みの場合、可視性に関係なく常に更新）
					if heatmap_precomputed and frame_idx < len(heatmap_precomputed):
						try:
							heatmap_data = heatmap_precomputed[frame_idx]
						except Exception as e:
							print(f"[エラー] フレーム{frame_idx}: ヒートマップデータ取得失敗: {e}")
							return
						
						if heatmap_data is not None and heatmap_data.n_points > 0:
							# distanceスカラーの存在確認
							if 'distance' in heatmap_data.array_names:
								# デバッグ: 最初のフレームのみログ出力
								if frame_idx == 0:
									print(f"[デバッグ] フレーム0のヒートマップ更新: {heatmap_data.n_points}点")
									distances = heatmap_data['distance']
									print(f"[デバッグ] 距離範囲: min={distances.min():.2f}, max={distances.max():.2f}")
								
								# メッシュを安全に更新
								try:
									# 【重要】heatmap_data は _LazyHeatmapList 由来で _base メッシュの
									# vtkPoints を shallow 共有している。copy_structure すると
									# heatmap_mesh も同じ vtkPoints を指してしまい、
									# 「heatmap_mesh.points = ...」の代入で _base 側まで書換わり、
									# 次フレーム以降に変換が累積してメッシュが飛んでいく（葉のように舞う）。
									# → deep_copy で独立した vtkPoints を持たせ、安全に変換を適用する。
									if prox_M is not None and hasattr(heatmap_mesh, 'deep_copy'):
										# 脛骨固定モード: 独立コピー → 逆変換を適用
										heatmap_mesh.deep_copy(heatmap_data)
										heatmap_mesh.points = _apply_pivot(
											np.asarray(heatmap_mesh.points), prox_M, dist_origin_initial
										)
									elif hasattr(heatmap_mesh, 'copy_structure'):
										# 通常モード: 変換不要なので shallow でOK（軽量）
										heatmap_mesh.copy_structure(heatmap_data)
										heatmap_mesh.point_data['distance'] = heatmap_data['distance']
									else:
										# copy_structureが使えない場合は直接代入
										heatmap_mesh.points = heatmap_data.points.copy()
										heatmap_mesh['distance'] = heatmap_data['distance'].copy()
										if hasattr(heatmap_data, 'faces') and heatmap_data.faces is not None:
											if hasattr(heatmap_mesh, 'faces'):
												heatmap_mesh.faces = heatmap_data.faces.copy()
										if prox_M is not None:
											heatmap_mesh.points = _apply_pivot(
												np.asarray(heatmap_mesh.points), prox_M, dist_origin_initial
											)
								except Exception as hm_err:
									if frame_idx == 0:
										print(f"[エラー] フレーム{frame_idx}: メッシュ更新失敗: {hm_err}")
									# 更新失敗時はダミーデータ
									try:
										heatmap_mesh.points = np.array([[0, 0, 0]])
										heatmap_mesh['distance'] = np.array([0.0])
									except:
										pass
								
								# 境界線（0mm等高線）を毎フレーム更新（actorの可視性制御で点滅防止）
								try:
									# 近位抽出領域メッシュに距離値をマッピング
									prox_with_distance = prox_joint_region.copy()
									
									# ヒートマップの点と距離値を取得
									heatmap_points = heatmap_data.points
									heatmap_distances = heatmap_data['distance']
									
									# 近位メッシュの各頂点に最近傍の距離値を割り当て
									from scipy.spatial import cKDTree
									tree = cKDTree(heatmap_points)
									_, indices = tree.query(prox_with_distance.points, k=1)
									prox_with_distance['distance'] = heatmap_distances[indices]
									
									# PyVistaのcontour機能で距離=0.0mmの等高線を抽出
									contour_mesh = prox_with_distance.contour(isosurfaces=[0.0], scalars='distance')
									
									if contour_mesh.n_points > 0:
										# boundary_meshを更新（actorはそのまま、メッシュだけ更新で点滅防止）
										try:
											# shallow_copyまたはcopy_structureで安全に更新
											if hasattr(boundary_mesh, 'copy_structure'):
												boundary_mesh.copy_structure(contour_mesh)
											else:
												boundary_mesh.points = contour_mesh.points.copy()
												if hasattr(contour_mesh, 'lines') and contour_mesh.lines is not None:
													boundary_mesh.lines = contour_mesh.lines.copy()
											# 脛骨固定モードでは、境界線も prox と一緒に inv(matrix) で移動
											if prox_M is not None:
												boundary_mesh.points = _apply_pivot(
													np.asarray(boundary_mesh.points), prox_M, dist_origin_initial
												)
											# actorを表示
											if boundary_actor is not None:
												try:
													boundary_actor.SetVisibility(True)
												except:
													pass
										except Exception as update_err:
											if frame_idx == 0:
												print(f"[エラー] フレーム{frame_idx}: boundary_mesh更新失敗: {update_err}")
											# 更新失敗時はactorを非表示に
											if boundary_actor is not None:
												try:
													boundary_actor.SetVisibility(False)
												except:
													pass
									else:
										# 等高線がない場合はactorを非表示に（メッシュは触らない）
										if boundary_actor is not None:
											try:
												boundary_actor.SetVisibility(False)
											except:
												pass
								except Exception as e:
									if frame_idx == 0:
										print(f"[エラー] フレーム{frame_idx}: 境界線抽出失敗: {e}")
									# 等高線抽出失敗時はactorを非表示に
									if boundary_actor is not None:
										try:
											boundary_actor.SetVisibility(False)
										except:
											pass
							else:
								# distanceスカラーがない場合
								if frame_idx == 0:
									print(f"[警告] フレーム0のヒートマップにdistanceスカラーがありません: {heatmap_data.array_names}")
								# 空のメッシュの場合はダミーデータで初期化
								try:
									heatmap_mesh.points = np.array([[0, 0, 0]])
									heatmap_mesh['distance'] = np.array([0.0])
								except:
									pass
								# 境界線を非表示に
								if boundary_actor is not None:
									try:
										boundary_actor.SetVisibility(False)
									except:
										pass
						else:
							# 空のメッシュの場合はダミーデータで初期化
							try:
								heatmap_mesh.points = np.array([[0, 0, 0]])
								heatmap_mesh['distance'] = np.array([0.0])
							except:
								pass
							# 境界線を非表示に
							if boundary_actor is not None:
								try:
									boundary_actor.SetVisibility(False)
								except:
									pass
					else:
						# ヒートマップデータがない場合
						if frame_idx == 0:
							print(f"[デバッグ] フレーム0: heatmap_precomputedなし（len={len(heatmap_precomputed) if heatmap_precomputed else 0}）")
						# 境界線を非表示に
						if boundary_actor is not None:
							try:
								boundary_actor.SetVisibility(False)
							except:
								pass
					
					# FEM接触圧メッシュの更新（毎フレーム更新）
					if fem_pressure_precomputed and frame_idx < len(fem_pressure_precomputed) and pressure_mesh is not None:
						try:
							fem_data = fem_pressure_precomputed[frame_idx]
							if fem_data is not None:
								pressure_arr, summary = fem_data
								# pressure_meshはheatmap_meshと同じ点を持つので、点を更新
								pressure_mesh.points = heatmap_mesh.points.copy()
								# 圧力値をマッピング（KDTree は prox 初期フレームで作られているので、
								# 脛骨固定モードでもクエリ点は「変換前 heatmap 位置」を用いる）
								if prox_fem_surface is not None:
									if len(pressure_arr) == pressure_mesh.n_points:
										pressure_mesh['contact_pressure'] = pressure_arr.copy()
									else:
										from scipy.spatial import cKDTree
										tree = cKDTree(prox_fem_surface.points)
										query_pts = (
											heatmap_data.points
											if 'heatmap_data' in locals() and heatmap_data is not None
											else pressure_mesh.points
										)
										_, idx = tree.query(query_pts, k=1)
										pressure_mesh['contact_pressure'] = pressure_arr[idx]
								else:
									pressure_mesh['contact_pressure'] = pressure_arr.copy()
						except Exception as e:
							if frame_idx == 0:
								print(f"[エラー] フレーム{frame_idx}: FEM接触圧更新失敗: {e}")
				
					# 面積表示を更新（新規追加）
					if len(overlap_areas_precomputed) > 0 and frame_idx < len(overlap_areas_precomputed):
						current_area = overlap_areas_precomputed[frame_idx]
						try:
							overlap_area_label.config(text=f"Area: {current_area:.2f} mm²")
						except:
							pass  # ウィンドウが閉じられている場合
					
					# 深度表示を更新（新規追加）
					if len(overlap_depths_precomputed) > 0 and frame_idx < len(overlap_depths_precomputed):
						current_depth = overlap_depths_precomputed[frame_idx]
						try:
							overlap_depth_label.config(text=f"Depth: {current_depth:.2f} mm")
						except:
							pass  # ウィンドウが閉じられている場合

					# 最大めり込み量表示を更新（ヒートマップから計算）
					if heatmap_precomputed and frame_idx < len(heatmap_precomputed):
						try:
							hm = heatmap_precomputed[frame_idx]
							if hm is not None and hm.n_points > 0 and 'distance' in hm.array_names:
								d = hm['distance']
								# 符号付き距離: 負の値がめり込み
								# 最小値（負の最大値）を取得し、負の場合のみ絶対値を表示
								min_dist = np.nanmin(d)
								max_penetration = abs(min_dist) if min_dist < 0 else 0.0
								max_penetration_label.config(text=f"Max Pent: {max_penetration:.2f} mm")
							else:
								max_penetration_label.config(text=f"Max Pent: -- mm")
						except Exception as e:
							print(f"[警告] めり込み量計算エラー: {e}")
							try:
								max_penetration_label.config(text=f"Max Pent: -- mm")
							except:
								pass
					

					# PyVistaの描画を更新（強制レンダリングまたはアクティブ時）
					if force_render or is_animation_active[0]:
						try:
							if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
								is_animation_active[0] = False
								return
							if hasattr(anim_plotter, 'render_window') and anim_plotter.render_window:
								try:
									anim_plotter.render()
								except Exception as render_err:
									# OpenGLエラーを無視（ログのみ出力）
									if frame_idx == 0:
										print(f"[エラー] フレーム{frame_idx}: レンダリング失敗: {render_err}")
									if not force_render:
										is_animation_active[0] = False
										if after_id[0] is not None:
											try:
												self.after_cancel(after_id[0])
											except:
												pass
											after_id[0] = None
							else:
								# render_windowが無効になった
								is_animation_active[0] = False
								return
						except Exception as check_err:
							# チェック自体が失敗
							is_animation_active[0] = False
							return
				except Exception as e:
					# エラー内容を詳細に表示
					print(f"[致命的エラー] フレーム{frame_idx}でshow_frame()がクラッシュ: {e}")
					traceback.print_exc()
					if not force_render:
						is_animation_active[0] = False
						if after_id[0] is not None:
							try:
								self.after_cancel(after_id[0])
							except:
								pass
							after_id[0] = None
			
			def animation_loop():
				"""アニメーションループ（実時間に同期して5ms間隔で進める）"""
				# デバッグ: ループが呼ばれていることを確認
				# print("[DEBUG] animation_loop() called")
				
				# アニメーションが無効化されているかチェック
				if not is_animation_active[0]:
					print("[アニメーション] is_animation_active is False, stopping")
					if after_id[0] is not None:
						self.after_cancel(after_id[0])
						after_id[0] = None
					return
				
				# ウィンドウが閉じられたかチェック
				if hasattr(anim_plotter, 'closed') and anim_plotter.closed:
					is_animation_active[0] = False
					if after_id[0] is not None:
						self.after_cancel(after_id[0])
						after_id[0] = None
					print("[アニメーション] ウィンドウが閉じられました")
					return
				
				# 一時停止中またはシーク中は更新しない
				if not is_paused[0] and not is_seeking[0]:
					# アニメーション開始時刻を記録（初回のみ）
					if animation_start_time[0] is None:
						animation_start_time[0] = time.time()
					
					# 実際の経過時間を計算し、再生速度を適用
					elapsed_real_time = time.time() - animation_start_time[0]
					elapsed_animation_time = elapsed_real_time * playback_speed[0]  # 再生速度を適用
					
					# 経過時間から実際のフレームカウンタを計算（5ms単位）
					actual_frame_counter[0] = int(elapsed_animation_time / 0.005)
					actual_time = actual_frame_counter[0] * 0.005
					
					# デバッグ: 時間の詳細を表示（最初の10秒間のみ）
					# if elapsed_real_time < 10.0 and actual_frame_counter[0] % 20 == 0:
					#     print(f"[TIME] real={elapsed_real_time:.3f}s, anim={elapsed_animation_time:.3f}s, actual={actual_time:.3f}s, speed={playback_speed[0]:.2f}x")
					
					# 最大時間を超えたらループ
					if actual_time > max_time:
						actual_frame_counter[0] = 0
						actual_time = 0.0
						animation_start_time[0] = time.time()  # リセット
						print("[アニメーション] ループ再生")
					
					# 実際の時間表示を更新
					try:
						actual_time_label.config(text=f"Actual: {actual_time:.3f}s")
					except:
						pass
					
					# 実際の時間に最も近いExcelデータのフレームを見つける（改善版）
					# 逆走を防ぐため、現在のフレームより前には戻らないようにする
					target_frame = current_frame[0]  # デフォルトは現在のフレーム
					
					# 現在のフレームから先を検索（前進のみ）
					for i in range(current_frame[0], len(frame_times)):
						if frame_times[i] <= actual_time:
							target_frame = i
						else:
							# 次のフレームの時間を超えたら停止
							break
					
					# もし適切なフレームが見つからなかった場合（actual_timeが最初のフレームより前）
					# その場合のみ、最初から検索
					if target_frame == current_frame[0] and current_frame[0] > 0:
						if actual_time < frame_times[current_frame[0]]:
							# actual_timeが現在のフレームより前の場合のみ、0から検索
							target_frame = 0
							for i in range(len(frame_times)):
								if frame_times[i] <= actual_time:
									target_frame = i
								else:
									break
					
					# デバッグ: 逆走が発生した場合のみログ出力
					if target_frame < current_frame[0]:
						print(f"[WARNING] 逆走検出: Frame {current_frame[0]} -> {target_frame}")
						print(f"  actual_time={actual_time:.3f}s, current_frame_time={frame_times[current_frame[0]]:.3f}s, target_frame_time={frame_times[target_frame]:.3f}s")
						print(f"  elapsed_real_time={elapsed_real_time:.3f}s, speed={playback_speed[0]:.2f}x")
					
					# フレームが変わった場合のみ更新
					if target_frame != current_frame[0]:
						current_frame[0] = target_frame
						show_frame(current_frame[0])
						# 見かけのフレーム表示を更新
						try:
							time_val = frame_times[current_frame[0]]
							visible_frame_label.config(text=f"Frame: {current_frame[0]}/{len(transform_data)-1} | Time: {time_val:.3f}s")
						except:
							pass  # ウィンドウが閉じられている場合
					
					# 再生バーは200msごとにのみ更新（YouTubeスタイル: ドラッグ中は更新しない）
					current_time = time.time()
					if not user_is_dragging[0] and (current_time - last_scale_update_time[0] >= scale_update_interval[0]):
						try:
							is_programmatic_update[0] = True
							playback_scale.set(current_frame[0])
							is_programmatic_update[0] = False
							last_scale_update_time[0] = current_time
						except:
							pass
				
				# 次のチェックを予約（5msごと）
				after_id[0] = self.after(5, animation_loop)
			
			def stop_animation():
				"""アニメーションを停止"""
				print("[アニメーション] 停止します")
				is_animation_active[0] = False
				
				# タイマーをキャンセル
				if after_id[0] is not None:
					try:
						self.after_cancel(after_id[0])
					except:
						pass
					after_id[0] = None
				
				# ウィンドウを閉じる（安全に）
				try:
					if hasattr(anim_plotter, 'close'):
						# 既に閉じられていないかチェック
						if not (hasattr(anim_plotter, 'closed') and anim_plotter.closed):
							# render_windowが有効かチェック
							if hasattr(anim_plotter, 'render_window') and anim_plotter.render_window:
								anim_plotter.close()
				except Exception as e:
					print(f"[アニメーション] クローズエラー（無視）: {e}")
			
			# ウィンドウのクローズコールバックを設定
			def on_window_close():
				"""ウィンドウが閉じられたときの処理"""
				print("[アニメーション] ウィンドウクローズイベント")
				is_animation_active[0] = False
				if after_id[0] is not None:
					try:
						self.after_cancel(after_id[0])
					except:
						pass
					after_id[0] = None
				if check_id[0] is not None:
					try:
						self.after_cancel(check_id[0])
					except:
						pass
					check_id[0] = None
			
			# カメラ位置の設定
			anim_plotter.camera_position = 'iso'
			anim_plotter.reset_camera()
			
			# 最初のフレームを表示
			show_frame(0)
			
			# 一時停止/再生ボタン（Pキー）
			anim_plotter.add_key_event('p', toggle_pause)
			anim_plotter.add_key_event('P', toggle_pause)
			
			# アニメーション開始
			print("[アニメーション] 開始します（実時間に同期、5ms間隔で更新）")
			print("[アニメーション] ウィンドウの×ボタンまたはQキーで終了してください")
			print("[アニメーション] Pキーで一時停止/再生")
			
			# ウィンドウクローズ時のコールバックを設定
			def on_window_close():
				"""ウィンドウが閉じられた時に呼ばれるコールバック"""
				print("[アニメーション] ウィンドウが閉じられました。タイマーを停止します。")
				is_animation_active[0] = False
				if after_id[0] is not None:
					try:
						self.after_cancel(after_id[0])
					except:
						pass
					after_id[0] = None
				if check_id[0] is not None:
					try:
						self.after_cancel(check_id[0])
					except:
						pass
					check_id[0] = None
				# 再生コントロールウィンドウも閉じる
				try:
					control_window.destroy()
				except:
					pass
			
			# ウィンドウが閉じられた時のコールバックを登録
			try:
				anim_plotter.iren.add_observer('ExitEvent', lambda obj, event: on_window_close())
			except:
				pass  # コールバック登録に失敗しても続行
			
			print(f"[初期化] PyVistaウィンドウを表示中...")
			# ウィンドウを表示（ノンブロッキング）してからアニメーション開始
			try:
				anim_plotter.show(auto_close=False, interactive_update=True)
			except TypeError:
				anim_plotter.show(auto_close=False)
			
			print(f"[初期化完了] アニメーションループを開始します")
			# ウィンドウが表示されたのでアニメーションループを開始
			after_id[0] = self.after(5, animation_loop)
			
		except FileNotFoundError as e:
			messagebox.showerror("ファイルエラー", f"ファイルが見つかりません:\n{e}")
		except Exception as e:
			messagebox.showerror("エラー", f"アニメーション実行中にエラーが発生しました:\n{e}\n\n{traceback.format_exc()}")

	# ----- Joint labels (関節種別ごとのUIラベル) -----
	def _joint_label(self, key: str) -> str:
		"""関節種別 (joint_var) に応じたUIラベル文字列を返す。

		key により対象UI要素を識別する。膝関節では ISB / Grood-Suntay 1983 を採用するため、
		近位=大腿骨、遠位=脛骨 となる。
		"""
		knee = (self.joint_var.get() == 2)
		table = {
			"prox_file_frame":  ("近位ファイル選択（大腿骨側）"  if knee else "近位ファイル選択（寛骨臼側）"),
			"dist_file_frame":  ("遠位ファイル選択（脛骨側）"    if knee else "遠位ファイル選択（大腿骨側）"),
			"cs_prox_tab":      ("近位（大腿骨）"                if knee else "近位（骨盤）"),
			"cs_dist_tab":      ("遠位（脛骨）"                  if knee else "遠位（大腿骨）"),
			"ao_checkbox":      ("大腿骨A-O線と脛骨A-O線を表示する" if knee else "骨盤A-O線と大腿骨A-Oを表示する"),
			# PPファイル選択行のラベル
			"prox_pp_abcd_label": ("近位特徴点 ABCD (PP)" if knee else "近位特徴点 ABCD (PP)"),
			"prox_pp_olmn_label": ("近位特徴点 OLMN (PP) ※膝関節では任意" if knee else "近位特徴点 OLMN (PP)"),
			"dist_pp_abc_label":  ("遠位特徴点 ABCD (PP)" if knee else "遠位特徴点 ABC (PP)"),
			"dist_pp_olmn_label": ("遠位特徴点 OLMN (PP) ※膝関節では任意" if knee else "遠位特徴点 OLMN (PP)"),
			# 特徴点の意味を示す凡例
			"prox_legend": (
				"特徴点: A=大腿骨頭中心, B=顆間中点(原点), C=内側後顆, D=外側後顆 (ISB / Grood-Suntay 1983)"
				if knee else
				"特徴点: A,B,C,D=寛骨臼の参照点, O=原点(寛骨臼中心) — OLMN ファイルに O 等を入れて結合"
			),
			"dist_legend": (
				"特徴点: A=顆間隆起中点(原点), B=足関節中心, C=内側プラトー, D=外側プラトー (ISB / Grood-Suntay 1983)"
				if knee else
				"特徴点: A,B=大腿骨頭軸の参照点, O=原点(大腿骨頭中心) — OLMN ファイルに O 等を入れて結合"
			),
		}
		return table.get(key, key)

	def _apply_joint_labels(self) -> None:
		"""joint_var に応じて、登録済みウィジェットのラベルを更新する。"""
		w = self._joint_widgets
		try:
			if "prox_file_frame" in w:
				w["prox_file_frame"].configure(text=self._joint_label("prox_file_frame"))
			if "dist_file_frame" in w:
				w["dist_file_frame"].configure(text=self._joint_label("dist_file_frame"))
			if "cs_notebook" in w and "cs_prox_tab" in w and "cs_dist_tab" in w:
				w["cs_notebook"].tab(w["cs_prox_tab"], text=self._joint_label("cs_prox_tab"))
				w["cs_notebook"].tab(w["cs_dist_tab"], text=self._joint_label("cs_dist_tab"))
			if "fem_prox_header" in w:
				w["fem_prox_header"].configure(text=f"【{self._joint_label('cs_prox_tab')}】")
			if "fem_dist_header" in w:
				w["fem_dist_header"].configure(text=f"【{self._joint_label('cs_dist_tab')}】")
			# PPファイル行のラベル
			for key in ("prox_pp_abcd_label", "prox_pp_olmn_label",
			            "dist_pp_abc_label", "dist_pp_olmn_label",
			            "prox_legend", "dist_legend"):
				if key in w:
					w[key].configure(text=self._joint_label(key))
		except Exception as e:
			print(f"[ラベル更新] 失敗: {e}")

	def _on_joint_changed(self) -> None:
		"""関節種別ラジオボタンが切り替わった時のハンドラ。

		現在の状態を保存してから、新しい関節の状態ファイルがあれば読み込む。
		ラベルとボタン状態も更新する。
		"""
		# 直前の関節の状態を保存（_prev_joint が無ければ初回切替なのでスキップ）
		prev = getattr(self, "_prev_joint", None)
		new = self.joint_var.get()
		if prev is not None and prev != new and prev in (1, 2):
			try:
				self._save_state_for_joint(prev)
			except Exception as e:
				print(f"[関節切替] 直前状態の保存に失敗: {e}")
		# 新しい関節の状態ファイルがあれば読み込む（パス変数等を上書き）
		if new in (1, 2) and prev != new:
			try:
				self._load_paths_for_joint(new)
			except Exception as e:
				print(f"[関節切替] 新状態の読込に失敗: {e}")
		self._prev_joint = new
		self._apply_joint_labels()
		self.update_button_states()

	# ----- State control -----
	def update_button_states(self) -> None:
		# 可視化条件
		# 膝関節 (ISB仕様) では OLMN ファイルは不要なのでオプション扱い
		knee_mode = (self.joint_var.get() == 2)
		olmn_prox_ok = knee_mode or bool(self.prox_pp_olmn_path.get())
		olmn_dist_ok = knee_mode or bool(self.dist_pp_olmn_path.get())
		all_files_selected = (
			bool(self.prox_model_path.get())
			and bool(self.prox_pp_abcd_path.get())
			and olmn_prox_ok
			and bool(self.dist_model_path.get())
			and bool(self.dist_pp_abc_path.get())
			and olmn_dist_ok
		)
		joint_selected = self.joint_var.get() in (1, 2)
		# 近位のみ
		prox_ready = joint_selected and bool(self.prox_model_path.get()) and bool(self.prox_pp_abcd_path.get()) and olmn_prox_ok
		self.visualize_prox_button.state(["!disabled"] if prox_ready else ["disabled"])
		# 遠位のみ
		dist_ready = joint_selected and bool(self.dist_model_path.get()) and bool(self.dist_pp_abc_path.get()) and olmn_dist_ok
		self.visualize_dist_button.state(["!disabled"] if dist_ready else ["disabled"])
		# 全体
		all_ready = joint_selected and all_files_selected
		self.visualize_all_button.state(["!disabled"] if all_ready else ["disabled"])

		# 同次変換行列群の確認: 同ファイル選択済み
		check_enabled = bool(self.transform_group_path.get())
		self.check_tf_button.state(["!disabled"] if check_enabled else ["disabled"])

		# アニメーション実行: 4ファイル + 同次変換行列群 選択済み
		animate_enabled = all_files_selected and check_enabled
		self.animate_button.state(["!disabled"] if animate_enabled else ["disabled"])

	def update_fitting_button_states(self) -> None:
		"""Fittingタブのボタン状態を更新"""
		# 各ファイルの選択状態
		parent_model_selected = bool(self.fitting_parent_model_path.get())
		parent_region_selected = bool(self.fitting_parent_region_path.get())
		child_model_selected = bool(self.fitting_child_model_path.get())
		child_region_selected = bool(self.fitting_child_region_path.get())
		
		# 統合可視化ボタン（どちらかのファイルが選択されていればOK）
		parent_viz_enabled = parent_model_selected or parent_region_selected
		self.fitting_viz_parent_button.state(["!disabled"] if parent_viz_enabled else ["disabled"])
		
		child_viz_enabled = child_model_selected or child_region_selected
		self.fitting_viz_child_button.state(["!disabled"] if child_viz_enabled else ["disabled"])

		# フィッティング実行: 4つのファイルすべて選択済み
		all_fitting_files_selected = (
			parent_model_selected
			and parent_region_selected
			and child_model_selected
			and child_region_selected
		)
		self.fitting_execute_button.state(["!disabled"] if all_fitting_files_selected else ["disabled"])

	def update_org_button_states(self) -> None:
		"""ORGタブのボタン状態を更新"""
		model_selected = bool(self.org_model_path.get())
		pp_selected = bool(self.org_pp_path.get())
		
		# 両方のファイルが選択されている場合のみ可視化ボタンを有効化
		viz_enabled = model_selected and pp_selected
		self.org_visualize_button.state(["!disabled"] if viz_enabled else ["disabled"])

	def update_a_b_button_states(self) -> None:
		"""A-Bタブのボタン状態を更新"""
		model_a_selected = bool(self.a_b_model_a_path.get())
		model_b_selected = bool(self.a_b_model_b_path.get())
		viz_enabled = model_a_selected and model_b_selected
		self.a_b_visualize_button.state(["!disabled"] if viz_enabled else ["disabled"])
		self.a_b_diff_button.state(["!disabled"] if viz_enabled else ["disabled"])

	def update_heatmap_button_states(self) -> None:
		"""HeatMapタブのボタン状態を更新"""
		model_a_selected = bool(self.heatmap_model_a_path.get())
		model_b_selected = bool(self.heatmap_model_b_path.get())
		viz_enabled = model_a_selected and model_b_selected
		self.heatmap_visualize_button.state(["!disabled"] if viz_enabled else ["disabled"])
		self.heatmap_execute_button.state(["!disabled"] if viz_enabled else ["disabled"])

	def update_c_fix_button_states(self) -> None:
		"""C fixタブのボタン状態を更新"""
		# Simulatorタブのファイルがすべて選択されているか確認
		all_files_selected = (
			bool(self.prox_model_path.get())
			and bool(self.prox_pp_abcd_path.get())
			and bool(self.prox_pp_olmn_path.get())
			and bool(self.dist_model_path.get())
			and bool(self.dist_pp_abc_path.get())
			and bool(self.dist_pp_olmn_path.get())
		)
		self.c_fix_visualize_button.state(["!disabled"] if all_files_selected else ["disabled"])
		
		# 近位ファイルが選択されているか確認
		prox_files_selected = (
			bool(self.prox_pp_abcd_path.get())
			and bool(self.prox_pp_olmn_path.get())
		)
		self.c_fix_transform_prox_button.state(["!disabled"] if prox_files_selected else ["disabled"])
		
		# 遠位ファイルが選択されているか確認
		dist_files_selected = (
			bool(self.dist_pp_abc_path.get())
			and bool(self.dist_pp_olmn_path.get())
		)
		self.c_fix_transform_dist_button.state(["!disabled"] if dist_files_selected else ["disabled"])

	# ----- Persistence -----
	def _presets_file_path(self) -> Path:
		"""姿勢・シーケンスプリセットの保存先ファイルパス。"""
		import platform
		filename = "frs2015_presets.json"
		if platform.system() == "Darwin":
			state_dir = Path.home() / ".frs_simulator"
			state_dir.mkdir(parents=True, exist_ok=True)
			return state_dir / filename
		else:
			return Path(__file__).with_name(filename)

	def _load_presets(self) -> dict:
		"""プリセットファイルを読み込む。存在しない場合は空の構造を返す。"""
		p = self._presets_file_path()
		default = {"poses": [], "sequences": []}
		if not p.exists():
			return default
		try:
			data = json.load(p.open("r", encoding="utf-8"))
			if not isinstance(data, dict):
				return default
			data.setdefault("poses", [])
			data.setdefault("sequences", [])
			return data
		except Exception as e:
			print(f"[プリセット読込] 失敗: {e}")
			return default

	def _save_presets(self, data: dict) -> None:
		"""プリセットをファイルに保存。"""
		p = self._presets_file_path()
		try:
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
		except Exception as e:
			print(f"[プリセット保存] 失敗: {e}")

	def _state_file_path(self, joint: int = None) -> Path:
		"""関節種別ごとの状態ファイルパスを返す。

		joint=None の場合は現在の joint_var を使用。
		- joint==2 (膝関節): frs2015_gui_state_knee.json
		- それ以外: frs2015_gui_state.json （股関節・既存ファイル）
		"""
		import platform
		if joint is None:
			joint = self.joint_var.get()
		filename = "frs2015_gui_state_knee.json" if joint == 2 else "frs2015_gui_state.json"
		if platform.system() == "Darwin":
			# macOS: ホームディレクトリ配下の隠しフォルダに保存（スクリプトディレクトリが書き込み不可の場合に対応）
			state_dir = Path.home() / ".frs_simulator"
			state_dir.mkdir(parents=True, exist_ok=True)
			return state_dir / filename
		else:
			return Path(__file__).with_name(filename)

	# ----- knee simulator 専用の状態保存（hipの関節ラジオと独立） -----
	def _knee_state_file_path(self) -> Path:
		"""knee simulator タブ専用の状態ファイルパス。"""
		import platform
		filename = "frs2015_gui_state_knee_sim.json"
		if platform.system() == "Darwin":
			state_dir = Path.home() / ".frs_simulator"
			state_dir.mkdir(parents=True, exist_ok=True)
			return state_dir / filename
		return Path(__file__).with_name(filename)

	def _knee_state_vars(self) -> dict:
		"""保存/復元する knee 変数を {キー: (var, 型)} で返す。"""
		d = {
			"knee_initial_scan": (self.knee_initial_scan_path, str),
			"knee_initial_pp": (self.knee_initial_pp_path, str),
			"knee_femur_model": (self.knee_femur_model_path, str),
			"knee_tibia_model": (self.knee_tibia_model_path, str),
			"knee_transform": (self.knee_transform_path, str),
			"knee_reg_femur_src": (self.knee_reg_femur_src_path, str),
			"knee_reg_femur_tgt": (self.knee_reg_femur_tgt_path, str),
			"knee_reg_tibia_src": (self.knee_reg_tibia_src_path, str),
			"knee_reg_tibia_tgt": (self.knee_reg_tibia_tgt_path, str),
			"knee_side": (self.knee_side_var, int),
			"knee_w_scan_deg": (self.knee_w_scan_deg, float),
			"knee_femur_color": (self.knee_femur_color_var, str),
			"knee_tibia_color": (self.knee_tibia_color_var, str),
		}
		# 骨ごとの位置合わせパラメータ・方式・プリセット選択
		for bone in ("femur", "tibia"):
			for key, kind, default, _lbl in self._knee_param_spec():
				d[f"knee_{bone}_{key}"] = (self.knee_reg_pvars[bone][key], int if kind == "int" else float)
			d[f"knee_{bone}_method"] = (self.knee_reg_method_var[bone], str)
			d[f"knee_{bone}_scaling"] = (self.knee_reg_scaling_var[bone], bool)
			d[f"knee_{bone}_preview"] = (self.knee_reg_preview_var[bone], bool)
			d[f"knee_{bone}_preset_sel"] = (self.knee_preset_sel[bone], str)
		return d

	def _save_knee_state(self) -> None:
		"""knee simulator の全「試験タブ」を専用ファイルへ保存する。"""
		# アクティブタブを最新状態に更新
		try:
			if getattr(self, "_knee_tabs", None):
				self._knee_tabs[self._knee_active_tab]['snapshot'] = self._knee_snapshot_current()
		except Exception:
			pass
		data = {
			"tabs": getattr(self, "_knee_tabs", []),
			"active": getattr(self, "_knee_active_tab", 0),
		}
		try:
			p = self._knee_state_file_path()
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
			print(f"[knee状態保存] {p}（{len(data['tabs'])}タブ）")
		except Exception as e:
			print(f"[knee状態保存] 失敗: {e}")

	def _load_knee_state(self) -> None:
		"""knee simulator の状態を専用ファイルから復元する（複数タブ対応・旧形式互換）。"""
		try:
			p = self._knee_state_file_path()
			if not p.exists():
				return
			data = json.load(p.open("r", encoding="utf-8"))
		except Exception as e:
			print(f"[knee状態復元] 失敗: {e}")
			return

		# 新形式: 複数タブ
		if isinstance(data, dict) and isinstance(data.get("tabs"), list) and data["tabs"]:
			self._knee_tabs = data["tabs"]
			try:
				self._knee_active_tab = int(data.get("active", 0))
			except Exception:
				self._knee_active_tab = 0
			if self._knee_active_tab < 0 or self._knee_active_tab >= len(self._knee_tabs):
				self._knee_active_tab = 0
			# アクティブタブの状態をライブ変数へ反映（UI構築は後段だが変数は既に存在）
			self._knee_restore_snapshot(self._knee_tabs[self._knee_active_tab].get("snapshot", {}))
			return

		# 旧形式(フラット): ライブ変数へ復元（後段の _knee_init_tabs が「試験1」化する）
		for key, (var, typ) in self._knee_state_vars().items():
			if key not in data:
				continue
			try:
				val = data[key]
				if typ is bool:
					var.set(bool(val))
				elif typ is int:
					var.set(int(val))
				elif typ is float:
					var.set(float(val))
				else:
					var.set(str(val))
			except Exception:
				pass
		for key, attr in (("_femur_reg_T", "_knee_femur_reg_T"), ("_tibia_reg_T", "_knee_tibia_reg_T")):
			if data.get(key) is not None:
				try:
					setattr(self, attr, np.array(data[key], dtype=float))
				except Exception:
					pass

	# パス系（関節依存）の状態キー一覧
	_JOINT_PATH_KEYS = (
		"prox_model", "prox_pp_abcd", "prox_pp_olmn",
		"dist_model", "dist_pp_abc", "dist_pp_olmn",
		"transform_group",
		"prox_cartilage_model", "dist_cartilage_model",
		"fem_prox_bone", "fem_prox_cartilage",
		"fem_dist_bone", "fem_dist_cartilage",
		"cs_prox_model1_whole", "cs_prox_model1_region",
		"cs_prox_model2_whole", "cs_prox_model2_region",
		"cs_dist_model1_whole", "cs_dist_model1_region",
		"cs_dist_model2_whole", "cs_dist_model2_region",
	)

	def _load_state(self) -> None:
		# まず股関節ファイル（既存メイン）を読み、joint_var を取得
		main_path = self._state_file_path(joint=1)
		data = None
		try:
			if main_path.exists():
				data = json.load(main_path.open("r", encoding="utf-8"))
		except Exception:
			data = None

		# joint_var の暫定値を判定
		j_initial = 0
		if data is not None:
			try:
				j_initial = int(data.get("joint", 0))
				if j_initial not in (0, 1, 2):
					j_initial = 0
			except Exception:
				j_initial = 0

		# 膝関節モードなら膝関節ファイルを読み直し（存在すれば）
		if j_initial == 2:
			knee_path = self._state_file_path(joint=2)
			try:
				if knee_path.exists():
					data = json.load(knee_path.open("r", encoding="utf-8"))
			except Exception:
				pass

		if data is None:
			return

		# 反映（妥当性チェックは軽め）
		try:
			j = int(data.get("joint", 0))
			self.joint_var.set(j if j in (1, 2) else 0)
		except Exception:
			self.joint_var.set(0)

		self.prox_model_path.set(str(data.get("prox_model", "")))
		self.prox_pp_abcd_path.set(str(data.get("prox_pp_abcd", "")))
		self.prox_pp_olmn_path.set(str(data.get("prox_pp_olmn", "")))
		self.dist_model_path.set(str(data.get("dist_model", "")))
		self.dist_pp_abc_path.set(str(data.get("dist_pp_abc", "")))
		self.dist_pp_olmn_path.set(str(data.get("dist_pp_olmn", "")))
		self.transform_group_path.set(str(data.get("transform_group", "")))
		
		# Fittingファイルの復元
		self.fitting_parent_model_path.set(str(data.get("fitting_parent_model", "")))
		self.fitting_parent_region_path.set(str(data.get("fitting_parent_region", "")))
		self.fitting_child_model_path.set(str(data.get("fitting_child_model", "")))
		self.fitting_child_region_path.set(str(data.get("fitting_child_region", "")))
		self.fitting_child_pp_path.set(str(data.get("fitting_child_pp", "")))
		
		# ORGファイルの復元
		self.org_model_path.set(str(data.get("org_model", "")))
		self.org_pp_path.set(str(data.get("org_pp", "")))

		# A-Bファイルの復元
		self.a_b_model_a_path.set(str(data.get("a_b_model_a", "")))
		self.a_b_model_b_path.set(str(data.get("a_b_model_b", "")))

		# HeatMapファイルの復元
		self.heatmap_model_a_path.set(str(data.get("heatmap_model_a", "")))
		self.heatmap_model_b_path.set(str(data.get("heatmap_model_b", "")))
		try:
			self.heatmap_show_model_a.set(bool(data.get("heatmap_show_model_a", True)))
		except Exception:
			self.heatmap_show_model_a.set(True)
		try:
			self.heatmap_show_model_b.set(bool(data.get("heatmap_show_model_b", True)))
		except Exception:
			self.heatmap_show_model_b.set(True)
		
		# ORG座標の復元
		try:
			self.org_coord_x.set(float(data.get("org_coord_x", 0.0)))
		except (ValueError, TypeError):
			self.org_coord_x.set(0.0)
		
		try:
			self.org_coord_y.set(float(data.get("org_coord_y", 0.0)))
		except (ValueError, TypeError):
			self.org_coord_y.set(0.0)
		
		try:
			self.org_coord_z.set(float(data.get("org_coord_z", 0.0)))
		except (ValueError, TypeError):
			self.org_coord_z.set(0.0)
		
		# Fittingパラメータの復元
		# RANSACパラメータ
		try:
			self.ransac_distance_threshold.set(float(data.get("ransac_distance_threshold", 1.0)))
		except (ValueError, TypeError):
			self.ransac_distance_threshold.set(1.0)
		
		try:
			self.ransac_max_iterations.set(int(data.get("ransac_max_iterations", 1000)))
		except (ValueError, TypeError):
			self.ransac_max_iterations.set(1000)
		
		try:
			self.ransac_confidence.set(float(data.get("ransac_confidence", 0.99)))
		except (ValueError, TypeError):
			self.ransac_confidence.set(0.99)
		
		# ICPパラメータ
		try:
			self.icp_threshold.set(float(data.get("icp_threshold", 5.0)))
		except (ValueError, TypeError):
			self.icp_threshold.set(5.0)
		
		try:
			self.icp_max_iterations.set(int(data.get("icp_max_iterations", 2000)))
		except (ValueError, TypeError):
			self.icp_max_iterations.set(2000)
		
		try:
			self.sample_points.set(int(data.get("sample_points", 10000)))
		except (ValueError, TypeError):
			self.sample_points.set(10000)
		
		# カラー設定の復元
		self.prox_color = data.get("prox_color", "#B0C4DE")
		self.dist_color = data.get("dist_color", "#FFB6C1")
		self.overlap_color = data.get("overlap_color", "#BC0051")
		
		# 関節範囲の半径の復元
		try:
			self.prox_radius.set(float(data.get("prox_radius", 30.0)))
		except (ValueError, TypeError):
			self.prox_radius.set(30.0)
		
		try:
			self.dist_radius.set(float(data.get("dist_radius", 30.0)))
		except (ValueError, TypeError):
			self.dist_radius.set(30.0)
		
		# O'原点位置オフセットの復元
		try:
			self.prox_offset_x.set(float(data.get("prox_offset_x", 0.0)))
		except (ValueError, TypeError):
			self.prox_offset_x.set(0.0)
		
		try:
			self.prox_offset_y.set(float(data.get("prox_offset_y", 0.0)))
		except (ValueError, TypeError):
			self.prox_offset_y.set(0.0)
		
		try:
			self.prox_offset_z.set(float(data.get("prox_offset_z", 0.0)))
		except (ValueError, TypeError):
			self.prox_offset_z.set(0.0)
		
		try:
			self.dist_offset_x.set(float(data.get("dist_offset_x", 0.0)))
		except (ValueError, TypeError):
			self.dist_offset_x.set(0.0)
		
		try:
			self.dist_offset_y.set(float(data.get("dist_offset_y", 0.0)))
		except (ValueError, TypeError):
			self.dist_offset_y.set(0.0)
		
		try:
			self.dist_offset_z.set(float(data.get("dist_offset_z", 0.0)))
		except (ValueError, TypeError):
			self.dist_offset_z.set(0.0)

		# 軟骨分離ファイルの復元（近位）
		self.cs_prox_model1_whole_path.set(str(data.get("cs_model1_whole", data.get("cs_prox_model1_whole", ""))))
		self.cs_prox_model1_region_path.set(str(data.get("cs_model1_region", data.get("cs_prox_model1_region", ""))))
		self.cs_prox_model2_whole_path.set(str(data.get("cs_model2_whole", data.get("cs_prox_model2_whole", ""))))
		self.cs_prox_model2_region_path.set(str(data.get("cs_model2_region", data.get("cs_prox_model2_region", ""))))

		# 軟骨分離ファイルの復元（遠位）
		self.cs_dist_model1_whole_path.set(str(data.get("cs_dist_model1_whole", "")))
		self.cs_dist_model1_region_path.set(str(data.get("cs_dist_model1_region", "")))
		self.cs_dist_model2_whole_path.set(str(data.get("cs_dist_model2_whole", "")))
		self.cs_dist_model2_region_path.set(str(data.get("cs_dist_model2_region", "")))

		# Simulator 軟骨モデルの復元
		self.prox_cartilage_model_path.set(str(data.get("prox_cartilage_model", "")))
		self.dist_cartilage_model_path.set(str(data.get("dist_cartilage_model", "")))

		# 軟骨分離パラメータの復元
		try:
			self.cs_ransac_distance_threshold.set(float(data.get("cs_ransac_distance_threshold", 1.0)))
		except (ValueError, TypeError):
			self.cs_ransac_distance_threshold.set(1.0)
		try:
			self.cs_ransac_max_iterations.set(int(data.get("cs_ransac_max_iterations", 1000)))
		except (ValueError, TypeError):
			self.cs_ransac_max_iterations.set(1000)
		try:
			self.cs_ransac_confidence.set(float(data.get("cs_ransac_confidence", 0.99)))
		except (ValueError, TypeError):
			self.cs_ransac_confidence.set(0.99)
		try:
			self.cs_icp_threshold.set(float(data.get("cs_icp_threshold", 5.0)))
		except (ValueError, TypeError):
			self.cs_icp_threshold.set(5.0)
		try:
			self.cs_icp_max_iterations.set(int(data.get("cs_icp_max_iterations", 2000)))
		except (ValueError, TypeError):
			self.cs_icp_max_iterations.set(2000)
		try:
			self.cs_sample_points.set(int(data.get("cs_sample_points", 10000)))
		except (ValueError, TypeError):
			self.cs_sample_points.set(10000)
		try:
			self.cs_distance_threshold.set(float(data.get("cs_distance_threshold", 0.5)))
		except (ValueError, TypeError):
			self.cs_distance_threshold.set(0.5)
		try:
			self.cs_enable_scaling.set(bool(data.get("cs_enable_scaling", False)))
		except (ValueError, TypeError):
			self.cs_enable_scaling.set(False)

		# 軟骨分離カラー・透明度の復元
		self.cs_bone_color = data.get("cs_bone_color", "#F5DEB3")
		self.cs_cartilage_color = data.get("cs_cartilage_color", "#00CED1")
		try:
			self.cs_bone_opacity.set(float(data.get("cs_bone_opacity", 0.4)))
		except (ValueError, TypeError):
			self.cs_bone_opacity.set(0.4)
		try:
			self.cs_cartilage_opacity.set(float(data.get("cs_cartilage_opacity", 1.0)))
		except (ValueError, TypeError):
			self.cs_cartilage_opacity.set(1.0)

		# FEM解析パラメータの復元
		self.fem_prox_bone_path.set(str(data.get("fem_prox_bone", "")))
		self.fem_prox_cartilage_path.set(str(data.get("fem_prox_cartilage", "")))
		self.fem_dist_bone_path.set(str(data.get("fem_dist_bone", "")))
		self.fem_dist_cartilage_path.set(str(data.get("fem_dist_cartilage", "")))
		try: self.fem_cart_E.set(float(data.get("fem_cart_E", 10.0)))
		except (ValueError, TypeError): self.fem_cart_E.set(10.0)
		try: self.fem_cart_nu.set(float(data.get("fem_cart_nu", 0.45)))
		except (ValueError, TypeError): self.fem_cart_nu.set(0.45)
		try: self.fem_cart_thickness.set(float(data.get("fem_cart_thickness", 2.0)))
		except (ValueError, TypeError): self.fem_cart_thickness.set(2.0)
		try: self.fem_bone_E.set(float(data.get("fem_bone_E", 17000.0)))
		except (ValueError, TypeError): self.fem_bone_E.set(17000.0)
		try: self.fem_bone_nu.set(float(data.get("fem_bone_nu", 0.3)))
		except (ValueError, TypeError): self.fem_bone_nu.set(0.3)
		try: self.fem_penalty_stiffness.set(float(data.get("fem_penalty_stiffness", 500.0)))
		except (ValueError, TypeError): self.fem_penalty_stiffness.set(500.0)
		try: self.fem_contact_tolerance.set(float(data.get("fem_contact_tolerance", 2.0)))
		except (ValueError, TypeError): self.fem_contact_tolerance.set(2.0)
		self.fem_boundary_mode.set(str(data.get("fem_boundary_mode", "auto")))
		try: self.fem_max_nodes.set(int(data.get("fem_max_nodes", 30000)))
		except (ValueError, TypeError): self.fem_max_nodes.set(30000)
		self.fem_scalar_name.set(str(data.get("fem_scalar_name", "contact_pressure")))
		self.fem_cmap.set(str(data.get("fem_cmap", "jet")))
		try: self.fem_show_dist_mesh.set(bool(data.get("fem_show_dist_mesh", True)))
		except (ValueError, TypeError): self.fem_show_dist_mesh.set(True)
		try: self.fem_show_bone.set(bool(data.get("fem_show_bone", True)))
		except (ValueError, TypeError): self.fem_show_bone.set(True)
		try: self.show_fem_analysis.set(bool(data.get("show_fem_analysis", False)))
		except (ValueError, TypeError): self.show_fem_analysis.set(False)

		# 共有キャッシュ設定の復元（キャッシュ初期化は _load_state() 呼び出し元が行う）
		self.cache_nas_path.set(str(data.get("cache_nas_path", "")))
		try: self.cache_enabled.set(bool(data.get("cache_enabled", True)))
		except (ValueError, TypeError): self.cache_enabled.set(True)

		# 可視化ウィンドウの表示/非表示
		for key, var in (
			("viz_show_prox_model", self.viz_show_prox_model),
			("viz_show_prox_pp",    self.viz_show_prox_pp),
			("viz_show_prox_axes",  self.viz_show_prox_axes),
			("viz_show_dist_model", self.viz_show_dist_model),
			("viz_show_dist_pp",    self.viz_show_dist_pp),
			("viz_show_dist_axes",  self.viz_show_dist_axes),
		):
			try: var.set(bool(data.get(key, True)))
			except (ValueError, TypeError): var.set(True)

	def _save_state(self) -> None:
		data = {
			"joint": self.joint_var.get(),
			"prox_model": self.prox_model_path.get(),
			"prox_pp_abcd": self.prox_pp_abcd_path.get(),
			"prox_pp_olmn": self.prox_pp_olmn_path.get(),
			"dist_model": self.dist_model_path.get(),
			"dist_pp_abc": self.dist_pp_abc_path.get(),
			"dist_pp_olmn": self.dist_pp_olmn_path.get(),
			"transform_group": self.transform_group_path.get(),
			"fitting_parent_model": self.fitting_parent_model_path.get(),
			"fitting_parent_region": self.fitting_parent_region_path.get(),
			"fitting_child_model": self.fitting_child_model_path.get(),
			"fitting_child_region": self.fitting_child_region_path.get(),
			"fitting_child_pp": self.fitting_child_pp_path.get(),
			"org_model": self.org_model_path.get(),
			"org_pp": self.org_pp_path.get(),
			"a_b_model_a": self.a_b_model_a_path.get(),
			"a_b_model_b": self.a_b_model_b_path.get(),
			"heatmap_model_a": self.heatmap_model_a_path.get(),
			"heatmap_model_b": self.heatmap_model_b_path.get(),
			"heatmap_show_model_a": self.heatmap_show_model_a.get(),
			"heatmap_show_model_b": self.heatmap_show_model_b.get(),
			"org_coord_x": self.org_coord_x.get(),
			"org_coord_y": self.org_coord_y.get(),
			"org_coord_z": self.org_coord_z.get(),
			"ransac_distance_threshold": self.ransac_distance_threshold.get(),
			"ransac_max_iterations": self.ransac_max_iterations.get(),
			"ransac_confidence": self.ransac_confidence.get(),
			"icp_threshold": self.icp_threshold.get(),
			"icp_max_iterations": self.icp_max_iterations.get(),
			"sample_points": self.sample_points.get(),
			"prox_color": self.prox_color,
			"dist_color": self.dist_color,
			"overlap_color": self.overlap_color,
			"prox_radius": self.prox_radius.get(),
			"dist_radius": self.dist_radius.get(),
			"prox_offset_x": self.prox_offset_x.get(),
			"prox_offset_y": self.prox_offset_y.get(),
			"prox_offset_z": self.prox_offset_z.get(),
			"dist_offset_x": self.dist_offset_x.get(),
			"dist_offset_y": self.dist_offset_y.get(),
			"dist_offset_z": self.dist_offset_z.get(),
			# 軟骨分離（近位）
			"cs_prox_model1_whole": self.cs_prox_model1_whole_path.get(),
			"cs_prox_model1_region": self.cs_prox_model1_region_path.get(),
			"cs_prox_model2_whole": self.cs_prox_model2_whole_path.get(),
			"cs_prox_model2_region": self.cs_prox_model2_region_path.get(),
			# 軟骨分離（遠位）
			"cs_dist_model1_whole": self.cs_dist_model1_whole_path.get(),
			"cs_dist_model1_region": self.cs_dist_model1_region_path.get(),
			"cs_dist_model2_whole": self.cs_dist_model2_whole_path.get(),
			"cs_dist_model2_region": self.cs_dist_model2_region_path.get(),
			# 軟骨分離パラメータ
			"cs_ransac_distance_threshold": self.cs_ransac_distance_threshold.get(),
			"cs_ransac_max_iterations": self.cs_ransac_max_iterations.get(),
			"cs_ransac_confidence": self.cs_ransac_confidence.get(),
			"cs_icp_threshold": self.cs_icp_threshold.get(),
			"cs_icp_max_iterations": self.cs_icp_max_iterations.get(),
			"cs_sample_points": self.cs_sample_points.get(),
			"cs_distance_threshold": self.cs_distance_threshold.get(),
			"cs_enable_scaling": self.cs_enable_scaling.get(),
			"cs_bone_color": self.cs_bone_color,
			"cs_cartilage_color": self.cs_cartilage_color,
			"cs_bone_opacity": self.cs_bone_opacity.get(),
			"cs_cartilage_opacity": self.cs_cartilage_opacity.get(),
			# Simulator 軟骨モデル
			"prox_cartilage_model": self.prox_cartilage_model_path.get(),
			"dist_cartilage_model": self.dist_cartilage_model_path.get(),
			# FEM解析
			"fem_prox_bone": self.fem_prox_bone_path.get(),
			"fem_prox_cartilage": self.fem_prox_cartilage_path.get(),
			"fem_dist_bone": self.fem_dist_bone_path.get(),
			"fem_dist_cartilage": self.fem_dist_cartilage_path.get(),
			"fem_cart_E": self.fem_cart_E.get(),
			"fem_cart_nu": self.fem_cart_nu.get(),
			"fem_cart_thickness": self.fem_cart_thickness.get(),
			"fem_bone_E": self.fem_bone_E.get(),
			"fem_bone_nu": self.fem_bone_nu.get(),
			"fem_penalty_stiffness": self.fem_penalty_stiffness.get(),
			"fem_contact_tolerance": self.fem_contact_tolerance.get(),
			"fem_boundary_mode": self.fem_boundary_mode.get(),
			"fem_max_nodes": self.fem_max_nodes.get(),
			"fem_scalar_name": self.fem_scalar_name.get(),
			"fem_cmap": self.fem_cmap.get(),
			"fem_show_dist_mesh": self.fem_show_dist_mesh.get(),
			"fem_show_bone": self.fem_show_bone.get(),
			"show_fem_analysis": self.show_fem_analysis.get(),
			# 共有キャッシュ設定
			"cache_nas_path": self.cache_nas_path.get(),
			"cache_enabled": self.cache_enabled.get(),
			# 可視化ウィンドウの表示/非表示
			"viz_show_prox_model": self.viz_show_prox_model.get(),
			"viz_show_prox_pp":    self.viz_show_prox_pp.get(),
			"viz_show_prox_axes":  self.viz_show_prox_axes.get(),
			"viz_show_dist_model": self.viz_show_dist_model.get(),
			"viz_show_dist_pp":    self.viz_show_dist_pp.get(),
			"viz_show_dist_axes":  self.viz_show_dist_axes.get(),
		}
		joint = self.joint_var.get()
		try:
			p = self._state_file_path()
			with p.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
			print(f"[状態保存] {p}")
		except Exception as e:
			print(f"[状態保存] 失敗: {e}")

		# 膝関節モードで保存した場合、メイン（股関節）ファイルの joint フィールドだけ
		# 更新しておく（次回起動時に「最後は膝関節だった」と判定するため）
		if joint == 2:
			try:
				main_path = self._state_file_path(joint=1)
				if main_path.exists():
					main_data = json.load(main_path.open("r", encoding="utf-8"))
				else:
					main_data = {}
				main_data["joint"] = 2
				with main_path.open("w", encoding="utf-8") as f:
					json.dump(main_data, f, ensure_ascii=False, indent=2)
			except Exception as e:
				print(f"[メイン joint更新] 失敗: {e}")

	def _save_state_for_joint(self, joint: int) -> None:
		"""指定された関節用のファイルに、現在のパス変数群を保存する。
		関節切替時に「切替前の状態」を保存する用途。joint_var とは独立して動く。
		"""
		# 一時的に joint_var を切替前の値に見立てて _save_state を呼ぶのは副作用が大きい。
		# 代わりに、ターゲットファイル単位で書き込む。
		# パス系のキーのみを書き込む（共通設定は現状の _save_state が担当）
		target = self._state_file_path(joint=joint)
		try:
			# 既存ファイルがあれば一度読み込んで、その上にパスを上書き
			if target.exists():
				try:
					data = json.load(target.open("r", encoding="utf-8"))
				except Exception:
					data = {}
			else:
				data = {}
			data["joint"] = joint
			data["prox_model"] = self.prox_model_path.get()
			data["prox_pp_abcd"] = self.prox_pp_abcd_path.get()
			data["prox_pp_olmn"] = self.prox_pp_olmn_path.get()
			data["dist_model"] = self.dist_model_path.get()
			data["dist_pp_abc"] = self.dist_pp_abc_path.get()
			data["dist_pp_olmn"] = self.dist_pp_olmn_path.get()
			data["transform_group"] = self.transform_group_path.get()
			data["prox_cartilage_model"] = self.prox_cartilage_model_path.get()
			data["dist_cartilage_model"] = self.dist_cartilage_model_path.get()
			data["fem_prox_bone"] = self.fem_prox_bone_path.get()
			data["fem_prox_cartilage"] = self.fem_prox_cartilage_path.get()
			data["fem_dist_bone"] = self.fem_dist_bone_path.get()
			data["fem_dist_cartilage"] = self.fem_dist_cartilage_path.get()
			data["cs_prox_model1_whole"] = self.cs_prox_model1_whole_path.get()
			data["cs_prox_model1_region"] = self.cs_prox_model1_region_path.get()
			data["cs_prox_model2_whole"] = self.cs_prox_model2_whole_path.get()
			data["cs_prox_model2_region"] = self.cs_prox_model2_region_path.get()
			data["cs_dist_model1_whole"] = self.cs_dist_model1_whole_path.get()
			data["cs_dist_model1_region"] = self.cs_dist_model1_region_path.get()
			data["cs_dist_model2_whole"] = self.cs_dist_model2_whole_path.get()
			data["cs_dist_model2_region"] = self.cs_dist_model2_region_path.get()
			with target.open("w", encoding="utf-8") as f:
				json.dump(data, f, ensure_ascii=False, indent=2)
			print(f"[関節別状態保存] joint={joint} → {target}")
		except Exception as e:
			print(f"[関節別状態保存] 失敗: {e}")

	def _load_paths_for_joint(self, joint: int) -> None:
		"""指定された関節用のファイルからパス変数を読み込む。
		存在しない場合はパス系変数を空文字にリセットする。
		"""
		target = self._state_file_path(joint=joint)
		if not target.exists():
			# 関節別ファイルが無い場合、パス変数をリセット（前の関節の値が残らないように）
			for v in (self.prox_model_path, self.prox_pp_abcd_path, self.prox_pp_olmn_path,
					  self.dist_model_path, self.dist_pp_abc_path, self.dist_pp_olmn_path,
					  self.transform_group_path,
					  self.prox_cartilage_model_path, self.dist_cartilage_model_path,
					  self.fem_prox_bone_path, self.fem_prox_cartilage_path,
					  self.fem_dist_bone_path, self.fem_dist_cartilage_path,
					  self.cs_prox_model1_whole_path, self.cs_prox_model1_region_path,
					  self.cs_prox_model2_whole_path, self.cs_prox_model2_region_path,
					  self.cs_dist_model1_whole_path, self.cs_dist_model1_region_path,
					  self.cs_dist_model2_whole_path, self.cs_dist_model2_region_path):
				v.set("")
			return
		try:
			data = json.load(target.open("r", encoding="utf-8"))
		except Exception as e:
			print(f"[関節別パス読込] 失敗: {e}")
			return
		self.prox_model_path.set(str(data.get("prox_model", "")))
		self.prox_pp_abcd_path.set(str(data.get("prox_pp_abcd", "")))
		self.prox_pp_olmn_path.set(str(data.get("prox_pp_olmn", "")))
		self.dist_model_path.set(str(data.get("dist_model", "")))
		self.dist_pp_abc_path.set(str(data.get("dist_pp_abc", "")))
		self.dist_pp_olmn_path.set(str(data.get("dist_pp_olmn", "")))
		self.transform_group_path.set(str(data.get("transform_group", "")))
		self.prox_cartilage_model_path.set(str(data.get("prox_cartilage_model", "")))
		self.dist_cartilage_model_path.set(str(data.get("dist_cartilage_model", "")))
		self.fem_prox_bone_path.set(str(data.get("fem_prox_bone", "")))
		self.fem_prox_cartilage_path.set(str(data.get("fem_prox_cartilage", "")))
		self.fem_dist_bone_path.set(str(data.get("fem_dist_bone", "")))
		self.fem_dist_cartilage_path.set(str(data.get("fem_dist_cartilage", "")))
		self.cs_prox_model1_whole_path.set(str(data.get("cs_prox_model1_whole", "")))
		self.cs_prox_model1_region_path.set(str(data.get("cs_prox_model1_region", "")))
		self.cs_prox_model2_whole_path.set(str(data.get("cs_prox_model2_whole", "")))
		self.cs_prox_model2_region_path.set(str(data.get("cs_prox_model2_region", "")))
		self.cs_dist_model1_whole_path.set(str(data.get("cs_dist_model1_whole", "")))
		self.cs_dist_model1_region_path.set(str(data.get("cs_dist_model1_region", "")))
		self.cs_dist_model2_whole_path.set(str(data.get("cs_dist_model2_whole", "")))
		self.cs_dist_model2_region_path.set(str(data.get("cs_dist_model2_region", "")))

	def _on_close(self) -> None:
		self._save_state()
		self._save_knee_state()
		self._save_ankle_state()
		self.destroy()
	
	def _save_initial_geometry(self) -> None:
		"""初期ウィンドウサイズを保存"""
		try:
			self.update_idletasks()
			self._initial_window_geometry = self.geometry()
		except Exception:
			pass
	
	def _restore_window_geometry(self) -> None:
		"""ウィンドウサイズを復元"""
		try:
			if self._initial_window_geometry:
				self.geometry(self._initial_window_geometry)
				self.update_idletasks()
		except Exception:
			pass

	# ----- Utils -----
	def _apply_transform(self, mesh: pv.PolyData, transform_matrix: np.ndarray) -> pv.PolyData:
		"""メッシュに同次変換行列を適用
		
		Args:
			mesh: 変換するメッシュ
			transform_matrix: 4x4同次変換行列
			
		Returns:
			pv.PolyData: 変換されたメッシュ
		"""
		if mesh.n_points == 0:
			return mesh
		
		# 同次座標に変換
		points_homo = np.hstack([mesh.points, np.ones((mesh.points.shape[0], 1))])
		
		# 変換を適用
		transformed_points = (transform_matrix @ points_homo.T).T[:, :3]
		
		# 新しいメッシュを作成
		transformed_mesh = mesh.copy()
		transformed_mesh.points = transformed_points
		
		return transformed_mesh
	
	def _compute_distance_heatmap(self, prox_joint_region: pv.PolyData, dist_joint_region: pv.PolyData, prox_surface: pv.PolyData = None, dist_surface: pv.PolyData = None, prox_points: np.ndarray = None) -> pv.PolyData:
		"""
		近位関節領域の各頂点に対して、遠位関節領域表面までの符号付き距離を計算してヒートマップを作成
		
		このメソッドは接触状態の可視化に使用されます。距離が小さいほど接触に近い状態を示します。
		
		Args:
			prox_joint_region: 近位関節領域（寛骨臼）のメッシュ
			dist_joint_region: 遠位関節領域（大腿骨頭）のメッシュ
			prox_surface: 事前抽出された近位表面（オプション、高速化用）
			dist_surface: 事前抽出された遠位表面（オプション、高速化用）
			prox_points: 事前抽出された近位点群（オプション、高速化用）
			
		Returns:
			pv.PolyData: "distance" スカラーに符号付き距離を持つ点群
			            - 正の値: 遠位が近位の外側にある（非接触）
			            - 0付近: 接触状態
			            - 負の値: 侵入状態（オーバーラップ）
			
		■ 処理フロー:
		  1. バウンディングボックスチェック（10mm以上離れている場合は簡易計算）
		  2. 近位表面抽出（prox_surfaceが指定されていない場合）
		  3. 遠位表面抽出（dist_surfaceが指定されていない場合）
		  4. 近位点群の各点について遠位表面への最近接点を検索
		  5. 距離計算と符号決定（法線ベクトルの内積で判定）
		  
		■ 最適化手法:
		  - バウンディングボックスによる早期リターン
		  - 表面抽出の再利用（prox_surface, dist_surface指定時）
		  - 点群のダウンサンプリング（prox_points指定時）
		  - KD-Tree高速近傍探索（PyVistaのfind_closest_cell使用）
		  
		■ 改修ポイント:
		  - 行7260付近: 表面抽出ロジック
		  - 行7300付近: 距離計算と符号決定
		  
		■ パフォーマンス注意:
		  - 大規模メッシュ（10万点以上）では処理時間が長くなる可能性あり
		  - 並列処理での呼び出しを推奨（_precompute_overlap_volumes経由）
		"""
		if prox_joint_region.n_points == 0 or dist_joint_region.n_points == 0:
			return prox_joint_region.copy() if prox_joint_region.n_points > 0 else pv.PolyData()
		
		# バウンディングボックスの事前チェック（交差判定の前に実施）
		try:
			prox_bounds = prox_joint_region.bounds
			dist_bounds = dist_joint_region.bounds
			
			# バウンディングボックスが遠く離れている場合（閾値: 10mm以上）
			# X, Y, Z軸それぞれで距離をチェック
			x_gap = max(0, dist_bounds[0] - prox_bounds[1], prox_bounds[0] - dist_bounds[1])
			y_gap = max(0, dist_bounds[2] - prox_bounds[3], prox_bounds[2] - dist_bounds[3])
			z_gap = max(0, dist_bounds[4] - prox_bounds[5], prox_bounds[4] - dist_bounds[5])
			
			# いずれかの軸で10mm以上離れている場合は、簡易計算のみ実行
			if x_gap > 10 or y_gap > 10 or z_gap > 10:
				# 中心間距離を全点の距離として設定（概算）
				prox_center = prox_joint_region.center
				dist_center = dist_joint_region.center
				approx_distance = np.linalg.norm(prox_center - dist_center)
				# メッシュ構造を保持して結果を返す
				result = prox_joint_region.copy(deep=False)
				result['distance'] = np.full(prox_joint_region.n_points, approx_distance)
				return result
		except Exception as e:
			# バウンディングボックスチェックに失敗した場合は通常計算を続行
			pass
		
		# 正確な符号付き距離計算：boolean_intersection で実際の重複を判定
		try:
			# Phase 2最適化: 事前抽出された配列があれば使用
			pts = prox_points if prox_points is not None else np.asarray(prox_joint_region.points)
			if pts.size == 0:
				return pv.PolyData(pts)

			# まず最近傍距離を計算
			from scipy.spatial import cKDTree
			dist_points = np.asarray(dist_joint_region.points)
			if dist_points.shape[0] == 0:
				if prox_points is None or len(prox_points) == prox_joint_region.n_points:
					result = prox_joint_region.copy(deep=False)
				else:
					result = pv.PolyData(pts)
				result['distance'] = np.full(len(pts), 100.0)  # 遠位メッシュがない場合は大きな値
				return result
			
			# KDTreeで各近位点から遠位メッシュへの最小距離を計算（最近傍インデックスも保存）
			tree = cKDTree(dist_points)
			d_vals, closest_indices = tree.query(pts, k=1)
			
			# 最適化: 最近傍距離で早期スキップ判定（結果は変わらない）
			min_distance = np.min(d_vals)
			if min_distance > 0.5:
				# 全ての点が0.5m以上離れている場合、boolean_intersectionは不要
				# （0.5m未満の点のみを負にする処理なので、結果は同じ）
				print(f"[最適化] 最小距離{min_distance:.2f}mm > 0.5m のためboolean_intersectionをスキップ")
				
				# 元のメッシュ構造を保持できるか確認
				if prox_points is None or len(prox_points) == prox_joint_region.n_points:
					result = prox_joint_region.copy(deep=False)
				else:
					result = pv.PolyData(pts)
					
				result['distance'] = np.array(d_vals, dtype=float)
				return result
			
			# 【超高速化】Signed Distance で埋め込み検出（10-50倍高速）
			try:
				# 事前抽出された表面を使用（高速化）、なければその場で抽出
				if dist_surface is not None:
					dist_mesh_for_implicit = dist_surface
				else:
					dist_mesh_for_implicit = dist_joint_region
					if hasattr(dist_joint_region, 'extract_surface'):
						dist_mesh_for_implicit = dist_joint_region.extract_surface()
				
				# PyVistaのimplicit distance（signed distance）を計算
				# 遠位メッシュを暗黙的関数として、近位点からの符号付き距離を計算
				# 負 = 遠位メッシュの内部、正 = 外部
				prox_point_cloud = pv.PolyData(pts)
				signed_distances = prox_point_cloud.compute_implicit_distance(dist_mesh_for_implicit, inplace=False)
				
				if 'implicit_distance' in signed_distances.point_data:
					implicit_dist = signed_distances.point_data['implicit_distance']
					
					# 埋め込み判定: implicit_distance が負 = 遠位メッシュの内部にある
					embedded_mask = implicit_dist < 0
					
					# デバッグ: 統計表示
					n_embedded = np.sum(embedded_mask)
					if n_embedded > 0:
						print(f"[Signed Distance] 埋め込み検出: {n_embedded}個 / {len(pts)}点")
						embedded_implicit = implicit_dist[embedded_mask]
						print(f"[Signed Distance] 埋め込み深さ: min={embedded_implicit.min():.3f}mm, max={embedded_implicit.max():.3f}mm, mean={embedded_implicit.mean():.3f}mm")
					
					# 埋め込まれている点: implicit_distanceを使用（既に符号付き）
					# KDTree距離（d_vals）と組み合わせる: 埋め込みなら負、外なら正
					# d_valsの絶対値を保持しつつ、埋め込みの場合のみ負にする
					d_vals[embedded_mask] = -np.abs(d_vals[embedded_mask])
					
				else:
					print(f"[警告] implicit_distance計算失敗 - フォールバック判定")
					# フォールバック: 距離が0.2mm未満なら埋め込みとみなす
					very_close_mask = d_vals < 0.2
					if np.any(very_close_mask):
						d_vals[very_close_mask] = -d_vals[very_close_mask]
				
			except Exception as e:
				# フォールバック: 距離が極めて小さい点を負にする（0.2mm未満）
				inside_mask = d_vals < 0.2
				if np.any(inside_mask):
					d_vals[inside_mask] = -d_vals[inside_mask]
			
			try:
				# 結果メッシュを作成（元のメッシュ構造を保持）
				# prox_pointsが指定されている場合はそれがprox_joint_regionの頂点に対応すると仮定
				# もしprox_pointsがダウンサンプリングされている場合は、元のメッシュ構造を使えないためPoint Cloudになる
				if prox_points is None or len(prox_points) == prox_joint_region.n_points:
					result = prox_joint_region.copy(deep=False)
				else:
					# ダウンサンプリングされている場合はPoint Cloudとして作成せざるを得ない
					result = pv.PolyData(pts)
				
				# 距離配列を設定
				result['distance'] = np.array(d_vals, dtype=float)
				return result
				
			except Exception as e2:
				print(f"[警告] 結果メッシュ生成エラー: {e2}")
				result = pv.PolyData(pts)
				result['distance'] = np.array(d_vals, dtype=float)
				return result

		except Exception as e:
			print(f"[警告] 距離ヒートマップ計算エラー: {e}")
			# フォールバックは元のメッシュ構造を維持
			fallback = prox_joint_region.copy(deep=False) if prox_joint_region.n_points > 0 else pv.PolyData()
			if fallback.n_points > 0:
				fallback['distance'] = np.full(fallback.n_points, 100.0)  # エラー時は全点を遠くに
			else:
				# 点がない場合はダミー
				fallback = pv.PolyData(np.array([[0.0, 0.0, 0.0]]))
				fallback['distance'] = np.array([100.0])
			return fallback

	def _precompute_heatmaps_o3d(self, prox_joint_region, dist_surface, prox_points, transform_data,
	                             update_progress, cancel_var):
		"""全フレームの符号付き距離ヒートマップを Open3D の BVH(RaycastingScene) で高速計算する。

		BVHにより O(N log M) で、GPU総当たり(O(N×M))やVTK per-frameより大幅に高速。
		剛体不変性を利用し、遠位メッシュ(固定)にシーンを一度だけ構築し、近位点を各フレームの
		inv(T_f) で逆変換してまとめて問い合わせる。値は VTK compute_implicit_distance と一致。

		Returns: フレームごとの pv.PolyData(prox領域 + 'distance') のリスト。
		"""
		import open3d as o3d
		# 遠位表面の三角形（固定・元姿勢）でシーンを1回だけ構築
		dm = dist_surface
		if not isinstance(dm, pv.PolyData):
			dm = dm.extract_surface()
		dm = dm.triangulate()
		dverts = np.asarray(dm.points, dtype=np.float32)
		dfaces_raw = np.asarray(dm.faces)
		if dfaces_raw.size < 4:
			raise ValueError("遠位メッシュに三角形がありません")
		dfaces = dfaces_raw.reshape(-1, 4)[:, 1:4].astype(np.int32)
		scene = o3d.t.geometry.RaycastingScene()
		scene.add_triangles(o3d.t.geometry.TriangleMesh(
			o3d.core.Tensor(dverts, o3d.core.float32),
			o3d.core.Tensor(dfaces, o3d.core.int32)))

		pts = np.asarray(prox_points) if prox_points is not None else np.asarray(prox_joint_region.points)
		n_pts = len(pts)
		if n_pts == 0:
			return []
		homog = np.hstack([pts, np.ones((n_pts, 1))]).astype(np.float64)  # (P,4)
		same_struct = (prox_points is None or n_pts == prox_joint_region.n_points)

		n_frames = len(transform_data)
		# 距離配列だけを貯める（メッシュは作らない）。全フレーム同一形状なので遅延リストで返す。
		results_d = [None] * n_frames
		# 近位点を逆変換して複数フレームをまとめて問い合わせ（目安: 約800万点/チャンク）
		frames_per_chunk = max(1, min(256, max(1, 8_000_000 // max(n_pts, 1))))
		for start in range(0, n_frames, frames_per_chunk):
			if cancel_var.get():
				break
			chunk = transform_data[start:start + frames_per_chunk]
			big = np.empty((len(chunk) * n_pts, 3), dtype=np.float32)
			for k, tf in enumerate(chunk):
				Tinv = np.linalg.inv(np.asarray(tf['matrix'], dtype=float))
				big[k * n_pts:(k + 1) * n_pts] = (Tinv @ homog.T).T[:, :3].astype(np.float32)
			sd = scene.compute_signed_distance(o3d.core.Tensor(big, o3d.core.float32)).numpy()
			for k in range(len(chunk)):
				# float32で保持（可視化には十分。容量・保存/読込時間を半減）。viewでなくcopyで独立化
				results_d[start + k] = np.array(sd[k * n_pts:(k + 1) * n_pts], dtype=np.float32)
			done = min(start + len(chunk), n_frames)
			if not update_progress(done, n_frames, f"[高速] ヒートマップ計算中: {done}/{n_frames}"):
				break
		# 共通ジオメトリ（近位領域メッシュ）＋各フレーム距離配列 の遅延リストを返す
		base_mesh = prox_joint_region if same_struct else pv.PolyData(pts)
		distances = [d for d in results_d if d is not None]
		return _LazyHeatmapList(base_mesh, distances)

	def _precompute_heatmaps_gpu(self, prox_joint_region, dist_surface, prox_points, transform_data,
	                             update_progress, cancel_var):
		"""全フレームの符号付き距離ヒートマップをGPUで一括計算する（frs_gpu使用）。

		剛体変換の不変性を利用: 遠位メッシュ(固定)に対し、近位点を各フレームの inv(T_f) で
		逆変換して符号付き距離を求める。これにより遠位三角形をGPUに一度載せるだけで、
		全フレームをまとめてGPU計算できる。値はVTK compute_implicit_distanceと一致。

		Returns: フレームごとの pv.PolyData(prox領域 + 'distance') のリスト。
		"""
		import frs_gpu
		# 遠位表面の三角形（固定・元姿勢）
		dm = dist_surface
		if not isinstance(dm, pv.PolyData):
			dm = dm.extract_surface()
		dm = dm.triangulate()
		dverts = np.asarray(dm.points, dtype=np.float64)
		dfaces_raw = np.asarray(dm.faces)
		if dfaces_raw.size < 4:
			raise ValueError("遠位メッシュに三角形がありません")
		dfaces = dfaces_raw.reshape(-1, 4)[:, 1:4].astype(np.int64)

		pts = np.asarray(prox_points) if prox_points is not None else np.asarray(prox_joint_region.points)
		n_pts = len(pts)
		if n_pts == 0:
			return []
		homog = np.hstack([pts, np.ones((n_pts, 1))])  # (P,4)
		same_struct = (prox_points is None or n_pts == prox_joint_region.n_points)

		n_frames = len(transform_data)
		results = [None] * n_frames
		# 近位点を逆変換して、複数フレームをまとめてGPUへ（目安: 約600万点/チャンク）
		frames_per_chunk = max(1, min(128, max(1, 6_000_000 // max(n_pts, 1))))
		for start in range(0, n_frames, frames_per_chunk):
			if cancel_var.get():
				break
			chunk = transform_data[start:start + frames_per_chunk]
			big = np.empty((len(chunk) * n_pts, 3), dtype=np.float64)
			for k, tf in enumerate(chunk):
				Tinv = np.linalg.inv(np.asarray(tf['matrix'], dtype=float))
				big[k * n_pts:(k + 1) * n_pts] = (Tinv @ homog.T).T[:, :3]
			signed = frs_gpu.signed_point_to_mesh_distance(big, dverts, dfaces, batch=256)
			if signed is None:
				raise RuntimeError("GPU符号付き距離が利用できません")
			for k in range(len(chunk)):
				d = signed[k * n_pts:(k + 1) * n_pts]
				hm = prox_joint_region.copy(deep=False) if same_struct else pv.PolyData(pts)
				hm['distance'] = np.asarray(d, dtype=float)
				results[start + k] = hm
			done = min(start + len(chunk), n_frames)
			if not update_progress(done, n_frames, f"[GPU] ヒートマップ計算中: {done}/{n_frames}"):
				break
		return [h for h in results if h is not None]

		
		# 変換を適用
		transformed_points = (transform_matrix @ points_homo.T).T[:, :3]
		
		# 新しいメッシュを作成
		transformed_mesh = mesh.copy()
		transformed_mesh.points = transformed_points
		
		return transformed_mesh
	
	# ----- Utils: Sphere Extraction -----
	def _extract_region_by_sphere(self, mesh: pv.PolyData, origin: np.ndarray, radius: float) -> pv.PolyData:
		"""球体の内部にあるメッシュ部分を抽出
		
		Args:
			mesh: 元のメッシュ
			origin: 座標系の原点（球の中心）
			radius: 球の半径（mm）
			
		Returns:
			pv.PolyData: 球内に含まれるメッシュ部分（表面メッシュとして有効なもの）
		"""
		if mesh is None or mesh.n_points == 0:
			return pv.PolyData()
		
		try:
			# メッシュの全頂点を取得
			points = mesh.points
			
			# 各点から原点までの距離を計算
			distances = np.linalg.norm(points - origin, axis=1)
			
			# 半径以内の点のマスクを作成
			mask = distances <= radius
			
			if not np.any(mask):
				return pv.PolyData()
				
			# extract_pointsで点を抽出（adjacent_cells=Trueでセルも含める）
			# PyVistaのバージョンによってはUnstructuredGridが返る
			extracted = mesh.extract_points(mask, adjacent_cells=True)
			
			# UnstructuredGridの場合はPolyData（表面）に変換して、面情報を正しく保持する
			# これにより点群ではなくメッシュとして描画されるようになる
			if hasattr(extracted, 'extract_surface'):
				surface = extracted.extract_surface()
				return surface
			elif isinstance(extracted, pv.PolyData):
				return extracted.clean()
			else:
				# 変換できない場合はクリーニングだけ試みる
				if hasattr(extracted, 'clean'):
					return extracted.clean()
				return extracted
				
		except Exception as e:
			print(f"領域抽出エラー: {e}")
			return pv.PolyData()

	# ----- Utils: Coordinate System Builder -----
	def _build_coordinate_system(self, points, labels):
		"""近位側の座標系を構築。

		joint_var に応じて股関節/膝関節（ISB / Grood-Suntay 1983）を切替える。
		- 股関節: A, B, C, D, O の5点から構築（既存仕様）
		- 膝関節 (大腿骨): A, B, C, D の4点から ISB 仕様で構築
		  - A = 大腿骨頭中心（近位機械軸端）
		  - B = 大腿骨遠位機械軸端（顆間中点）= 原点
		  - C = 内側後顆点
		  - D = 外側後顆点
		"""
		if self.joint_var.get() == 2:
			return self._build_coordinate_system_knee_femur(points, labels)
		import numpy as np
		
		# ラベルから点を取得
		point_dict = {}
		for i, label in enumerate(labels):
			label_upper = label.strip().upper()
			if label_upper in ['A', 'B', 'C', 'D', 'O']:
				point_dict[label_upper] = points[i]
		
		# 必要な点が揃っているか確認
		required = ['A', 'B', 'C', 'D', 'O']
		missing = [p for p in required if p not in point_dict]
		if missing:
			raise ValueError(f"必要な点が見つかりません: {', '.join(missing)}")
		
		A = np.array(point_dict['A'])
		B = np.array(point_dict['B'])
		C = np.array(point_dict['C'])
		D = np.array(point_dict['D'])
		O = np.array(point_dict['O'])
		
		# 原点
		origin = O
		
		# Z軸: B→Aの方向
		z_axis = A - B
		z_axis = z_axis / np.linalg.norm(z_axis)  # 正規化
		
		# CとDの中点
		mid_CD = (C + D) / 2.0
		
		# AB線上の点でmid_CDに最も近い点を求める
		# AB線: P = B + t*(A-B)
		# mid_CDからPへのベクトルがAB線に垂直になるtを求める
		AB = A - B
		t = np.dot(mid_CD - B, AB) / np.dot(AB, AB)
		closest_point = B + t * AB
		
		# mid_CDからAB線への垂線ベクトル
		perp = closest_point - mid_CD
		if np.linalg.norm(perp) < 1e-10:
			raise ValueError("CとDの中点がAB線上にあるため、X軸を定義できません")
		
		# X軸: 垂線方向でZ軸に直交
		x_axis = perp - np.dot(perp, z_axis) * z_axis  # Z軸成分を除去
		x_axis = x_axis / np.linalg.norm(x_axis)  # 正規化
		
		# Y軸: Z×X軸
		y_axis = np.cross(z_axis, x_axis)
		y_axis = y_axis / np.linalg.norm(y_axis)  # 正規化
		
		return origin, x_axis, y_axis, z_axis

	def _build_coordinate_system_dist(self, points, labels):
		"""遠位側の座標系を構築。

		joint_var に応じて股関節/膝関節（ISB / Grood-Suntay 1983）を切替える。
		- 股関節: A, B, O の3点から構築（既存仕様）
		- 膝関節 (脛骨): A, B, C, D の4点から ISB 仕様で構築
		  - A = 顆間隆起中点（近位機械軸端）= 原点
		  - B = 足関節中心（遠位機械軸端）
		  - C = 内側脛骨プラトー中心
		  - D = 外側脛骨プラトー中心
		"""
		if self.joint_var.get() == 2:
			return self._build_coordinate_system_knee_tibia(points, labels)
		import numpy as np
		
		# ラベルから点を取得
		point_dict = {}
		for i, label in enumerate(labels):
			label_upper = label.strip().upper()
			if label_upper in ['A', 'B', 'O']:
				point_dict[label_upper] = points[i]
		
		# 必要な点が揃っているか確認
		required = ['A', 'B', 'O']
		missing = [p for p in required if p not in point_dict]
		if missing:
			raise ValueError(f"必要な点が見つかりません: {', '.join(missing)}")
		
		A = np.array(point_dict['A'])
		B = np.array(point_dict['B'])
		O = np.array(point_dict['O'])
		
		# 原点
		origin = O
		
		# Z軸: B→Aの方向
		z_axis = A - B
		z_axis = z_axis / np.linalg.norm(z_axis)  # 正規化
		
		# BとAの中点
		mid_AB = (A + B) / 2.0
		
		# 中点からOに向かうベクトル
		vec_to_O = O - mid_AB
		if np.linalg.norm(vec_to_O) < 1e-10:
			raise ValueError("BとAの中点とOが一致しているため、Y軸を定義できません")
		
		# Y軸: ベクトルをZ軸に直交させる
		y_axis = vec_to_O - np.dot(vec_to_O, z_axis) * z_axis  # Z軸成分を除去
		if np.linalg.norm(y_axis) < 1e-10:
			raise ValueError("OがAB線上にあるため、Y軸を定義できません")
		y_axis = y_axis / np.linalg.norm(y_axis)  # 正規化
		
		# X軸: Y×Z軸
		x_axis = np.cross(y_axis, z_axis)
		x_axis = x_axis / np.linalg.norm(x_axis)  # 正規化

		return origin, x_axis, y_axis, z_axis

	# ----- Utils: Knee Coordinate System (ISB / Grood-Suntay 1983) -----
	def _build_coordinate_system_knee_femur(self, points, labels):
		"""膝関節 大腿骨側の ISB 座標系を構築。

		Grood & Suntay (1983) に準拠:
		- 原点: B (大腿骨遠位機械軸端 = 顆間中点)
		- Z軸 (機械軸): B→A方向（近位正）
		- Y軸 (前方): 前額面の法線。CD後顆連結線とZ軸の外積から導出
		- X軸 (屈曲軸): Y×Z軸（右膝で外側正）

		Args:
			points: 各点の座標 [(x,y,z), ...]
			labels: 各点のラベル ["A", "B", "C", "D", ...]
				A = 大腿骨頭中心
				B = 顆間中点（最遠位）
				C = 内側後顆点
				D = 外側後顆点
		"""
		import numpy as np
		point_dict = {}
		for i, label in enumerate(labels):
			lab = label.strip().upper()
			if lab in ['A', 'B', 'C', 'D']:
				point_dict[lab] = points[i]
		required = ['A', 'B', 'C', 'D']
		missing = [p for p in required if p not in point_dict]
		if missing:
			raise ValueError(f"膝関節(大腿骨)に必要な点が見つかりません: {', '.join(missing)}")

		A = np.array(point_dict['A'], dtype=float)
		B = np.array(point_dict['B'], dtype=float)
		C = np.array(point_dict['C'], dtype=float)
		D = np.array(point_dict['D'], dtype=float)

		# 原点: B (顆間中点 = 大腿骨遠位機械軸端)
		origin = B

		# Z軸: 機械軸（近位正方向 = B→A）
		z_vec = A - B
		if np.linalg.norm(z_vec) < 1e-10:
			raise ValueError("A (大腿骨頭中心) と B (顆間中点) が一致しているため、Z軸を定義できません")
		z_axis = z_vec / np.linalg.norm(z_vec)

		# 後顆連結線 D-C (内側→外側)
		cd_vec = D - C
		if np.linalg.norm(cd_vec) < 1e-10:
			raise ValueError("C (内側後顆) と D (外側後顆) が一致しているため、座標系を定義できません")

		# Y軸 (前方): cross(Z, D-C). 右膝で前方を向く
		y_vec = np.cross(z_axis, cd_vec)
		if np.linalg.norm(y_vec) < 1e-10:
			raise ValueError("後顆連結線が機械軸と平行のため、Y軸を定義できません")
		y_axis = y_vec / np.linalg.norm(y_vec)

		# X軸 (屈曲軸): Y×Z (右手系完成)
		x_axis = np.cross(y_axis, z_axis)
		x_axis = x_axis / np.linalg.norm(x_axis)

		return origin, x_axis, y_axis, z_axis

	def _build_coordinate_system_knee_tibia(self, points, labels):
		"""膝関節 脛骨側の ISB 座標系を構築。

		Grood & Suntay (1983) に準拠:
		- 原点: A (顆間隆起中点 = 脛骨近位機械軸端)
		- z軸 (機械軸): B→A方向（近位正）
		- y軸 (前方): プラトー連結線CDとz軸の外積から導出
		- x軸: y×z（右膝で外側正）

		Args:
			points: 各点の座標
			labels: 各点のラベル
				A = 顆間隆起中点（最近位）
				B = 足関節中心（最遠位）
				C = 内側プラトー中心
				D = 外側プラトー中心
		"""
		import numpy as np
		point_dict = {}
		for i, label in enumerate(labels):
			lab = label.strip().upper()
			if lab in ['A', 'B', 'C', 'D']:
				point_dict[lab] = points[i]
		required = ['A', 'B', 'C', 'D']
		missing = [p for p in required if p not in point_dict]
		if missing:
			raise ValueError(f"膝関節(脛骨)に必要な点が見つかりません: {', '.join(missing)}")

		A = np.array(point_dict['A'], dtype=float)
		B = np.array(point_dict['B'], dtype=float)
		C = np.array(point_dict['C'], dtype=float)
		D = np.array(point_dict['D'], dtype=float)

		# 原点: A (顆間隆起中点)
		origin = A

		# z軸: 機械軸（近位正方向 = B→A）
		z_vec = A - B
		if np.linalg.norm(z_vec) < 1e-10:
			raise ValueError("A (顆間隆起中点) と B (足関節中心) が一致しているため、z軸を定義できません")
		z_axis = z_vec / np.linalg.norm(z_vec)

		# プラトー連結線 D-C (内側→外側)
		cd_vec = D - C
		if np.linalg.norm(cd_vec) < 1e-10:
			raise ValueError("C (内側プラトー) と D (外側プラトー) が一致しているため、座標系を定義できません")

		# y軸 (前方): cross(z, D-C)
		y_vec = np.cross(z_axis, cd_vec)
		if np.linalg.norm(y_vec) < 1e-10:
			raise ValueError("プラトー連結線が機械軸と平行のため、y軸を定義できません")
		y_axis = y_vec / np.linalg.norm(y_vec)

		# x軸: y×z
		x_axis = np.cross(y_axis, z_axis)
		x_axis = x_axis / np.linalg.norm(x_axis)

		return origin, x_axis, y_axis, z_axis

	def _try_auto_set_offset(self, points, labels, origin, x_axis, y_axis, z_axis, off_x_var, off_y_var, off_z_var):
		"""O'点が存在し、かつオフセットが未設定(0)の場合、自動的にオフセットを設定する"""
		import numpy as np
		# 現在のオフセット値を確認
		try:
			curr_x = off_x_var.get()
			curr_y = off_y_var.get()
			curr_z = off_z_var.get()
		except:
			return

		# 既に設定されている場合は何もしない（0.001mm以上の絶対値がある場合）
		if abs(curr_x) > 1e-3 or abs(curr_y) > 1e-3 or abs(curr_z) > 1e-3:
			return

		# O'を探す
		op_idx = -1
		target_labels = ["O'", "O_PRIME", "OPRIME", "O_DASH", "ODASH"]
		
		for i, label in enumerate(labels):
			lab = label.strip().upper()
			if lab in target_labels:
				op_idx = i
				break
		
		if op_idx != -1:
			O_prime = points[op_idx]
			
			# Global Offset Vector
			diff = O_prime - origin
			
			# Project to Local Coordinate System (x, y, z axes are normalized)
			# Local Offset = Dot(Diff, Axis)
			local_off_x = np.dot(diff, x_axis)
			local_off_y = np.dot(diff, y_axis)
			local_off_z = np.dot(diff, z_axis)
			
			# Update Variables
			off_x_var.set(float(local_off_x))
			off_y_var.set(float(local_off_y))
			off_z_var.set(float(local_off_z))
			
			print(f"[自動設定] O'点を検出しました。オフセットを自動設定します: [{local_off_x:.3f}, {local_off_y:.3f}, {local_off_z:.3f}]")

	# ----- Utils: PP loader -----
	def _parse_pp_file(self, path: str):
		"""PPファイルから点群を抽出。
		対応形式:
		- GOM Inspect 等の PickedPoints XML 形式 (<PickedPoints><point x="" y="" z="" name="" active="1"/> ...)
		- テキスト行形式: "x y z" / "name x y z" / CSV(3〜4列)
		- 混在形式: XMLブロックの後にテキスト行が続く場合にも対応
		- 行頭が#の行はコメント扱い
		"""
		p = Path(path)
		if not p.exists():
			raise FileNotFoundError(f"PPファイルが存在しません: {path}")

		# ファイル全文を読み込む
		full_text = p.read_text(encoding="utf-8", errors="ignore")
		pts = []
		labels = []

		# まずXMLブロック（先頭〜</PickedPoints>まで）を安全に抽出してパース
		xml_end_idx = full_text.lower().find('</pickedpoints>')
		if full_text.lstrip().startswith('<') and xml_end_idx != -1:
			xml_block = full_text[:xml_end_idx + len('</PickedPoints>')]
			try:
				import xml.etree.ElementTree as ET
				root = ET.fromstring(xml_block)
				if root.tag.lower() == 'pickedpoints':
					for pt in root.findall('.//point'):
						active = pt.get('active')
						if active is not None and str(active) not in ('1', 'true', 'True'):
							continue
						name = pt.get('name') or f"P{len(pts)+1}"
						x = float(pt.get('x'))
						y = float(pt.get('y'))
						z = float(pt.get('z'))
						pts.append([x, y, z])
						labels.append(name)
			except Exception:
				# XML抽出/パース失敗時は後続のテキスト処理へ
				pass

		# XMLの後に続くテキスト部分を個別にパース（混在形式対応）
		text_tail = full_text[xml_end_idx + len('</PickedPoints>'):] if xml_end_idx != -1 else full_text
		with io.StringIO(text_tail) as f:
			for line in f:
				s = line.strip()
				if not s or s.startswith('#'):
					continue
				# カンマ or 空白区切り
				parts = [t for t in s.replace(',', ' ').split() if t]
				if len(parts) < 3:
					continue
				# 後ろ3つをx,y,zとして解釈
				try:
					x, y, z = float(parts[-3]), float(parts[-2]), float(parts[-1])
					pts.append([x, y, z])
					name = ' '.join(parts[:-3]) if len(parts) > 3 else f"P{len(pts)}"
					labels.append(name)
				except Exception:
					continue
		import numpy as np
		if not pts:
			raise ValueError("PPファイルから座標を抽出できませんでした。")
		return np.asarray(pts, dtype=float), labels

	def _uniform_sample(self, points: np.ndarray, n_samples: int) -> np.ndarray:
		"""点群から均一にサンプリング
		
		Args:
			points: 点群 (N, 3)
			n_samples: サンプリング点数
		
		Returns:
			サンプリングされた点群 (n_samples, 3)
		"""
		if len(points) <= n_samples:
			return points
		
		# ランダムにサンプリング
		indices = np.random.choice(len(points), n_samples, replace=False)
		return points[indices]


# グローバル関数（multiprocessing用）
def _compute_overlap_for_frame(args):
	"""単一フレームのオーバーラップ体積を計算（並列処理用）
	
	Args:
		args: (prox_mesh_points, prox_mesh_faces, dist_mesh_points_initial, 
		       dist_mesh_faces, dist_origin_initial, matrix, frame_idx, use_bbox_check)
	
	Returns:
		tuple: (frame_idx, overlap_mesh_points, overlap_mesh_faces, overlap_area, contact_depth)
	"""
	try:
		import pyvista as pv
		import numpy as np
		
		(prox_mesh_points, prox_mesh_faces, dist_mesh_points_initial, 
		 dist_mesh_faces, dist_origin_initial, matrix, frame_idx, use_bbox_check) = args
		
		# 近位メッシュを再構築
		prox_mesh = pv.PolyData(prox_mesh_points, prox_mesh_faces)
		
		# 遠位メッシュを変換
		centered_points = dist_mesh_points_initial - dist_origin_initial
		ones = np.ones((centered_points.shape[0], 1))
		points_homogeneous = np.hstack([centered_points, ones])
		transformed_homogeneous = (matrix @ points_homogeneous.T).T
		transformed_points = transformed_homogeneous[:, :3] + dist_origin_initial
		
		# 変換された遠位メッシュを作成
		dist_mesh = pv.PolyData(transformed_points, dist_mesh_faces)
		
		# バウンディングボックスチェック（オプション）
		if use_bbox_check:
			prox_bounds = prox_mesh.bounds
			dist_bounds = dist_mesh.bounds
			
			# バウンディングボックスが交差していない場合はスキップ
			if (prox_bounds[1] < dist_bounds[0] or prox_bounds[0] > dist_bounds[1] or
				prox_bounds[3] < dist_bounds[2] or prox_bounds[2] > dist_bounds[3] or
				prox_bounds[5] < dist_bounds[4] or prox_bounds[4] > dist_bounds[5]):
				return (frame_idx, None, None, 0.0, 0.0)
		
		# 距離ベースの事前チェック：最小距離が閾値以上なら交差なし
		# 近位メッシュからサンプル点を取得（高速化のため全点は使わない）
		sample_size = min(1000, prox_mesh.n_points)
		if prox_mesh.n_points > sample_size:
			sample_indices = np.random.choice(prox_mesh.n_points, sample_size, replace=False)
			sample_points = prox_mesh.points[sample_indices]
		else:
			sample_points = prox_mesh.points
		
		# 各サンプル点から遠位メッシュへの最小距離を計算
		min_distances = []
		for point in sample_points:
			_, closest_point = dist_mesh.find_closest_cell(point, return_closest_point=True)
			distance = np.linalg.norm(point - closest_point)
			min_distances.append(distance)
		
		# 最小距離が0.1mm以上なら接触なしと判定
		min_dist = min(min_distances) if min_distances else float('inf')
		if min_dist > 0.1:  # 閾値: 0.1mm
			return (frame_idx, None, None, 0.0, 0.0)
		
		# オーバーラップ体積を計算
		# VTKの警告・エラーログを一時的に抑制
		import vtk
		import os
		error_output = vtk.vtkFileOutputWindow()
		error_output.SetFileName("nul" if os.name == 'nt' else "/dev/null")
		vtk.vtkOutputWindow.SetInstance(error_output)
		
		overlap_result = prox_mesh.boolean_intersection(dist_mesh)
		
		# ログ出力を元に戻す
		vtk.vtkOutputWindow.SetInstance(None)
		
		if overlap_result.n_points > 0:
			# メッシュデータと面積を返す
			area = overlap_result.area
			# 面積が極端に小さい場合は数値誤差とみなす
			if area < 0.01:  # 0.01 mm² 以下は無視
				return (frame_idx, None, None, 0.0, 0.0)
			
			# 接触深度を計算（オーバーラップメッシュを渡す）
			depth = _compute_contact_depth_global(prox_mesh, dist_mesh, overlap_result)
			return (frame_idx, overlap_result.points, overlap_result.faces, area, depth)
		else:
			return (frame_idx, None, None, 0.0, 0.0)
	
	except Exception as e:
		print(f"[並列計算] フレーム {frame_idx}: エラー - {e}")
		return (frame_idx, None, None, 0.0, 0.0)


def _compute_contact_depth_global(prox_mesh, dist_mesh, overlap_mesh=None):
	"""接触深度を計算（グローバル関数・並列処理用）
	
	Args:
		prox_mesh: 近位メッシュ（PyVistaメッシュ）
		dist_mesh: 遠位メッシュ（PyVistaメッシュ）
		overlap_mesh: オーバーラップメッシュ（オプション）
		
	Returns:
		float: 接触深度（mm）、接触がない場合は0.0
	"""
	try:
		import pyvista as pv
		import numpy as np
		
		# 方法1: オーバーラップメッシュが提供されている場合
		# オーバーラップメッシュの厚み（バウンディングボックスの最小辺）を深度とする
		if overlap_mesh is not None and overlap_mesh.n_points > 0:
			try:
				# オーバーラップ領域のバウンディングボックスを取得
				bounds = overlap_mesh.bounds
				# X, Y, Z方向の幅を計算
				x_width = bounds[1] - bounds[0]
				y_width = bounds[3] - bounds[2]
				z_width = bounds[5] - bounds[4]
				# 最小幅を深度とする（貫入深さの近似）
				min_width = min(x_width, y_width, z_width)
				return max(0.0, min_width)  # 負の値は返さない
			except Exception:
				pass
		
		# 方法2: 遠位メッシュの頂点から近位メッシュへの最小距離を計算
		dist_points = dist_mesh.points
		
		# サンプリング（計算時間短縮）
		if len(dist_points) > 2000:
			indices = np.random.choice(len(dist_points), 2000, replace=False)
			sample_dist_points = dist_points[indices]
		else:
			sample_dist_points = dist_points
		
		# 各点から近位メッシュへの距離を計算
		min_distances = []
		for point in sample_dist_points:
			try:
				idx = prox_mesh.find_closest_point(point)
				if isinstance(idx, (int, np.integer)) and idx >= 0:
					closest_pt = prox_mesh.points[idx]
					distance = np.linalg.norm(point - closest_pt)
					min_distances.append(distance)
			except Exception:
				pass
		
		if len(min_distances) > 0:
			# 最小距離の下位10%の平均を深度とする
			sorted_distances = np.sort(min_distances)
			n_close = max(1, len(sorted_distances) // 10)
			close_distances = sorted_distances[:n_close]
			mean_depth = np.mean(close_distances)
			return mean_depth
		
		return 0.0
	
	except Exception as e:
		print(f"[接触深度計算] エラー: {e}")
		return 0.0

	def _uniform_sample(self, points: np.ndarray, n_samples: int) -> np.ndarray:
		"""点群から均一にサンプリング
		
		Args:
			points: 点群 (N, 3)
			n_samples: サンプリング点数
		
		Returns:
			サンプリングされた点群 (n_samples, 3)
		"""
		if len(points) <= n_samples:
			return points
		
		# ランダムにサンプリング
		indices = np.random.choice(len(points), n_samples, replace=False)
		return points[indices]
	
	def _icp_registration(self, source_points: np.ndarray, target_points: np.ndarray) -> Optional[np.ndarray]:
		"""ICP（Iterative Closest Point）による位置合わせ
		
		Args:
			source_points: 子の点群 (N, 3)
			target_points: 親の点群 (M, 3)
		
		Returns:
			変換行列 (4, 4) または None（失敗時）
		"""
		max_iterations = self.icp_max_iterations.get()
		threshold = self.icp_threshold.get()
		
		print(f"ICP設定: 最大反復回数={max_iterations}, 距離閾値={threshold}mm")
		
		# KDTreeを構築（親点群）
		tree = cKDTree(target_points)
		
		# 現在の変換（単位行列から開始）
		current_transform = np.eye(4)
		prev_error = float('inf')
		
		transformed = source_points.copy()
		
		for iteration in range(max_iterations):
			# 最近傍点を探索
			distances, indices = tree.query(transformed)
			
			# 距離閾値以下の点のみ使用
			valid_mask = distances < threshold
			valid_source = transformed[valid_mask]
			valid_target = target_points[indices[valid_mask]]
			
			if len(valid_source) < 3:
				print(f"反復{iteration+1}: 対応点が不足（{len(valid_source)}点）")
				break
			
			# 平均誤差を計算
			mean_error = np.mean(distances[valid_mask])
			
			if iteration % 50 == 0 or iteration == max_iterations - 1:
				print(f"反復{iteration+1}: 平均誤差={mean_error:.4f}mm, 対応点数={len(valid_source)}")
			
			# 収束判定
			if abs(prev_error - mean_error) < 0.0001:
				print(f"収束しました（反復{iteration+1}）")
				break
			
			prev_error = mean_error
			
			# 変換行列を推定
			transform = self._estimate_transform_svd(valid_source, valid_target)
			
			# 変換を適用
			transformed = self._apply_transform_to_points(transformed, transform)
			
			# 累積変換を更新
			current_transform = transform @ current_transform
		
		# 最終的な対応点数と誤差を計算
		distances, _ = tree.query(transformed)
		valid_mask = distances < threshold
		final_inliers = np.sum(valid_mask)
		final_error = np.mean(distances[valid_mask]) if final_inliers > 0 else float('inf')
		
		print(f"\nICP完了:")
		print(f"  最終平均誤差: {final_error:.4f}mm")
		print(f"  対応点数: {final_inliers} / {len(source_points)}")
		
		return current_transform
	
	def _estimate_transform_svd(self, source: np.ndarray, target: np.ndarray) -> np.ndarray:
		"""SVDを使用して剛体変換を推定（重心ベース）
		
		Args:
			source: 子の点群 (N, 3)
			target: 親の点群 (N, 3)
		
		Returns:
			変換行列 (4, 4)
		"""
		# 重心を計算
		centroid_source = np.mean(source, axis=0)
		centroid_target = np.mean(target, axis=0)
		
		# 重心を原点に移動
		source_centered = source - centroid_source
		target_centered = target - centroid_target
		
		# 共分散行列を計算
		H = source_centered.T @ target_centered
		
		# SVD分解
		U, S, Vt = np.linalg.svd(H)
		
		# 回転行列を計算
		R = Vt.T @ U.T
		
		# 特殊な反射を修正
		if np.linalg.det(R) < 0:
			Vt[-1, :] *= -1
			R = Vt.T @ U.T
		
		# 平行移動を計算
		t = centroid_target - R @ centroid_source
		
		# 4x4変換行列を構築
		transform = np.eye(4)
		transform[:3, :3] = R
		transform[:3, 3] = t
		
		return transform
	
	def _save_transformed_model(self, transform: np.ndarray) -> None:
		"""変換済みの子モデル全体を保存
		
		Args:
			transform: 変換行列 (4, 4)
		"""
		try:
			# 子モデル全体を読み込み
			child_model_mesh = pv.read(self.fitting_child_model_path.get())
			
			print(f"\n子モデル全体を変換中...")
			print(f"  頂点数: {child_model_mesh.n_points}")
			print(f"  三角形数: {child_model_mesh.n_cells}")
			
			# 変換を適用
			child_model_transformed = child_model_mesh.copy()
			child_model_transformed.transform(transform, inplace=True)
			
			# 保存先を選択
			original_path = Path(self.fitting_child_model_path.get())
			default_filename = original_path.stem + "_fitted" + original_path.suffix
			
			save_path = filedialog.asksaveasfilename(
				title="変換済みモデルを保存",
				defaultextension=original_path.suffix,
				initialfile=default_filename,
				filetypes=[
					("STL files", "*.stl"),
					("OBJ files", "*.obj"),
					("All files", "*.*")
				]
			)
			
			if save_path:
				child_model_transformed.save(save_path)
				print(f"\n変換済みモデルを保存しました: {Path(save_path).name}")

				# 子PPも同じ変換で保存（存在する場合）
				# ユーザーにPP保存の意志確認
				want_save_pp = messagebox.askyesno(
					"確認",
					"子特徴点 (PP) も同じ変換で保存しますか？"
				)
				child_pp_path = self.fitting_child_pp_path.get().strip()
				if want_save_pp and not child_pp_path:
					# 未選択ならここで選択を促す
					messagebox.showinfo("情報", "子PPファイルが未選択です。選択してください。")
					self.choose_fitting_child_pp()
					child_pp_path = self.fitting_child_pp_path.get().strip()
				if want_save_pp and child_pp_path:
					try:
						pp_points, pp_labels = self._parse_pp_file(child_pp_path)
						# L/M/Nは保存対象から除外する（仕様に応じて変更可）
						filtered_points = []
						filtered_labels = []
						for p, lb in zip(pp_points, pp_labels):
							if lb not in ['L', 'M', 'N']:
								filtered_points.append(p)
								filtered_labels.append(lb)
						if not filtered_points:
							filtered_points = pp_points
							filtered_labels = pp_labels
						filtered_points = np.asarray(filtered_points, dtype=float)
						# 変換を適用
						ones = np.ones((filtered_points.shape[0], 1))
						points_h = np.hstack([filtered_points, ones])
						points_t = (transform @ points_h.T).T[:, :3]
						# 保存先提案（元モデルの隣に .pp）
						pp_default = Path(save_path).with_suffix('').with_name(Path(save_path).stem + "_points.pp")
						pp_save_path = filedialog.asksaveasfilename(
							title="変換済み子特徴点 (PP) を保存",
							defaultextension=".pp",
							initialfile=pp_default.name,
							filetypes=[("PP files", "*.pp"), ("All files", "*.*")]
						)
						if pp_save_path:
							self.save_child_pp_file(pp_save_path, points_t, filtered_labels)
							print(f"変換済み子PPを保存しました: {Path(pp_save_path).name}")
						else:
							print("子PPの保存はキャンセルされました。")
					except Exception as e:
						print(f"子PPの保存に失敗: {e}")
				
				# 確認のため可視化
				response = messagebox.askyesno("確認", 
					f"変換済みモデルを保存しました。\n"
					f"保存先: {Path(save_path).name}\n\n"
					f"親モデルと変換済み子モデルを表示しますか？")
				
				if response:
					# 親モデルを読み込み
					parent_model_mesh = pv.read(self.fitting_parent_model_path.get())
					
					# 最終結果を可視化
					plotter = pv.Plotter()
					plotter.set_background('white')
					
					plotter.add_mesh(parent_model_mesh, color='red', opacity=0.5, show_edges=False, label='親モデル')
					plotter.add_mesh(child_model_transformed, color='cyan', opacity=0.5, show_edges=False, label='子モデル（フィッティング後）')
					
					plotter.add_axes()
					plotter.add_legend()
					
					print("\n[最終結果] 親モデル + 変換済み子モデルを表示中...")
					plotter.show()
			
		except Exception as e:
			messagebox.showerror("エラー", f"モデル保存中にエラーが発生しました:\n{str(e)}")
			print(f"エラー詳細: {e}")
			import traceback
			traceback.print_exc()
	
	def _get_principal_axis(self, mesh: pv.PolyData) -> tuple:
		"""メッシュの主軸（最大固有ベクトル）と重心を計算
		
		Args:
			mesh: PyVistaメッシュ
		
		Returns:
			tuple: (重心, 主軸の単位ベクトル)
		"""
		points = mesh.points
		
		# 重心を計算
		center = np.mean(points, axis=0)
		
		# 共分散行列を計算
		centered = points - center
		cov = np.cov(centered.T)
		
		# 固有値・固有ベクトルを計算
		eigvals, eigvecs = np.linalg.eigh(cov)
		
		# 最大固有値に対応する固有ベクトル（主軸）
		principal_axis = eigvecs[:, np.argmax(eigvals)]
		
		return center, principal_axis
	
	def _align_axis(self, mesh: pv.PolyData, src_axis: np.ndarray, tgt_axis: np.ndarray) -> tuple:
		"""主軸をターゲット軸に合わせる回転を適用
		
		Args:
			mesh: PyVistaメッシュ
			src_axis: ソース軸（単位ベクトル）
			tgt_axis: ターゲット軸（単位ベクトル）
		
		Returns:
			tuple: (回転後のメッシュ, 回転行列)
		"""
		# 外積で回転軸を計算
		v = np.cross(src_axis, tgt_axis)
		c = np.dot(src_axis, tgt_axis)
		
		# すでに平行な場合
		if np.linalg.norm(v) < 1e-8:
			return mesh.copy(), np.eye(3)
		
		# Rodriguesの回転公式で回転行列を計算
		vx = np.array([
			[0, -v[2], v[1]],
			[v[2], 0, -v[0]],
			[-v[1], v[0], 0]
		])
		R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
		# メッシュに回転適用
		rotated = mesh.copy()
		rotated.points = (R @ rotated.points.T).T
		return rotated, R

	def save_child_model_cache(self, cache_path: str, child_model: object) -> None:
		"""子モデルのキャッシュ保存（例: pickle形式）"""
		import pickle
		with open(cache_path, 'wb') as f:
			pickle.dump(child_model, f)

	def choose_save_child_pp_directory(self) -> str:
		"""子新PPファイル保存用ディレクトリ選択ダイアログ"""
		dir_path = filedialog.askdirectory(title="子新PPファイルの保存先ディレクトリを選択")
		return dir_path

	def save_child_pp_file(self, save_path: str, points: np.ndarray, labels: list) -> None:
		"""子特徴点（PP）をXML形式で保存する
		<PickedPoints><point x="..." y="..." z="..." active="1" name="..."/></PickedPoints>
		"""
		import xml.etree.ElementTree as ET
		root = ET.Element('PickedPoints')
		for pt, label in zip(points, labels):
			point_elem = ET.SubElement(root, 'point')
			point_elem.set('x', f"{pt[0]:.8f}")
			point_elem.set('y', f"{pt[1]:.8f}")
			point_elem.set('z', f"{pt[2]:.8f}")
			point_elem.set('active', "1")
			point_elem.set('name', label)
		tree = ET.ElementTree(root)
		tree.write(save_path, encoding="utf-8", xml_declaration=True)
		
		R = np.eye(3) + vx + vx @ vx * ((1 - c) / (np.linalg.norm(v) ** 2))
		
		# 4x4変換行列に変換
		transform = np.eye(4)
		transform[:3, :3] = R
		
		# メッシュに適用
		mesh_rotated = mesh.copy()
		mesh_rotated.transform(transform, inplace=True)
		
		return mesh_rotated, R
# endregion MainMenuGUIクラス


# region メインエントリーポイント
def main():
	"""メインエントリーポイント"""
	# --- プラットフォーム固有の初期化 ---
	if IS_WINDOWS:
		# Windows: Open3DによるDPI設定の強制変更（ウィンドウ縮小現象）への対策
		# アプリケーション起動時から高DPI対応(System DPI Aware)としてマークすることで、
		# 途中で設定が切り替わるのを防ぎ、サイズ変化を抑制する。
		try:
			import ctypes
			ctypes.windll.shcore.SetProcessDpiAwareness(1)
		except Exception:
			pass
	elif IS_MACOS:
		# macOS: Retinaディスプレイ対応とアプリケーション名の設定
		try:
			# macOSでTkinterアプリをDockで正しく表示させる
			import subprocess
			# PyObjCが利用可能な場合、追加の最適化
		except Exception:
			pass

	app = MainMenuGUI()

	# macOS固有のウィンドウ設定
	if IS_MACOS:
		try:
			# macOSでウィンドウを前面に持ってくる
			app.lift()
			app.attributes('-topmost', True)
			app.after(100, lambda: app.attributes('-topmost', False))
		except Exception:
			pass

	app.mainloop()


if __name__ == "__main__":
	# macOS/Windows両対応: spawnモードでの並列処理に必要
	from multiprocessing import freeze_support
	freeze_support()
	main()
# endregion メインエントリーポイント

# ============================================================================
# ファイル終端
# ============================================================================
# FRS.Simulator.py 1.0
# 
# このファイルは約8100行のPythonコードで構成されています。
# 
# ■ ファイル構成:
#   - 目次: 行1-70
#   - インポート: 行71-97
#   - グローバル関数: 行98-123
#   - MainMenuGUIクラス: 行124-8098
#   - main()関数: 行8099-8112
#
# ■ 今後の改善案:
#   1. on_animate()の分割: 1400行→複数の小関数に分割
#   2. キャッシュシステムの強化: SQLiteベースのDBキャッシュ
#   3. ユニットテスト追加: 座標変換・幾何計算の検証
#   4. 設定ファイル外部化: パラメータをYAML/TOMLで管理
#   5. ログ機能強化: loggingモジュール導入
#
# ============================================================================
